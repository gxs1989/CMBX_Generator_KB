from __future__ import annotations

from pathlib import Path

from embedded_report_extractor import parse_report_sheet_objects, parse_report_sheets
from foq_contract_report import build_report_cell_value_map
from report_calculation_map import ReportCellCalculation
from report_formula_evaluator import FormulaEvaluation


def write_report_template_workbook(
    path: str | Path,
    report_name: str,
    report_xml: str,
    injection_name: str = "",
    formula_sheets: dict[str, list[FormulaEvaluation]] | None = None,
    calculation_map: list[ReportCellCalculation] | None = None,
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build report template workbooks.") from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    formula_sheets = formula_sheets or {}
    cell_values = build_report_cell_value_map(formula_sheets, calculation_map or [])
    sheets = parse_report_sheets(report_xml, report_name, injection_name)
    objects = parse_report_sheet_objects(report_xml, report_name)

    workbook = Workbook()
    workbook.remove(workbook.active)
    if not sheets:
        workbook.create_sheet("Report Template")

    title_fill = PatternFill("solid", fgColor="EAF2F8")
    formula_fill = PatternFill("solid", fgColor="FFF2CC")
    value_fill = PatternFill("solid", fgColor="D9EAD3")
    derived_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(bottom=Side(style="thin", color="D9D9D9"))
    sheet_by_name = {}
    for sheet in sheets:
        ws = workbook.create_sheet(_safe_sheet_name(sheet.sheet_name))
        sheet_by_name[sheet.sheet_name] = ws
        ws.sheet_view.showGridLines = True
        ws["A1"] = sheet.sheet_name
        ws["A1"].font = Font(name="Calibri", size=13, bold=True)
        ws["A1"].fill = title_fill
        ws["A2"] = "Report Template"
        ws["B2"] = report_name
        ws["A3"] = "Injection"
        ws["B3"] = injection_name or "(template only)"
        ws["A4"] = "Applies to Injection"
        ws["B4"] = sheet.applies_to_injection
        for row_cells in ws["A1:B4"]:
            for cell in row_cells:
                cell.border = border
        for column in range(1, 18):
            ws.column_dimensions[_column_letter(column)].width = 14
        ws.column_dimensions["A"].width = 22
        ws.freeze_panes = "A6"

    for obj in objects:
        ws = sheet_by_name.get(obj.sheet_name)
        if ws is None:
            continue
        start_cell = _top_left_cell(obj.excel_range)
        if not start_cell:
            continue
        cell = ws[start_cell]
        value = _direct_formula_value(formula_sheets.get(obj.sheet_name, []), start_cell)
        if value is None:
            value = _value_from_map(cell_values, obj.sheet_name, start_cell)
        if value is not None:
            cell.value = value
            cell.fill = value_fill
        elif obj.formula:
            cell.value = "[CM formula]"
            cell.fill = formula_fill
        elif obj.object_type:
            cell.value = f"[{obj.object_type}]"
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        comment_text = _object_comment(obj)
        if comment_text:
            cell.comment = Comment(comment_text, "CMBX Data Explorer")

    for (sheet_name, cell_ref), cell_value in cell_values.items():
        ws = _worksheet_by_normalized_name(sheet_by_name, sheet_name)
        if ws is None:
            continue
        cell = ws[cell_ref]
        if cell.value in (None, "", "[CM formula]"):
            cell.value = cell_value.value
            cell.fill = derived_fill
            cell.comment = Comment(f"Derived report value\nStatus: {cell_value.status}\n{cell_value.detail}", "CMBX Data Explorer")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if formula_sheets:
        _write_values_index(workbook.create_sheet("_Filled Values"), formula_sheets, cell_values, Font, PatternFill, Alignment, Border, Side)
    _write_object_index(workbook.create_sheet("_Template Objects"), objects, Font, PatternFill, Alignment, Border, Side)

    workbook.save(_windows_long_path(output_path))
    return output_path


def _direct_formula_value(rows: list[FormulaEvaluation], cell_ref: str) -> object | None:
    target = cell_ref.upper()
    for row in rows:
        if row.status == "ok" and row.excel_range.upper() == target:
            return _coerce_value(row.value)
    return None


def _value_from_map(cell_values, sheet_name: str, cell_ref: str) -> object | None:
    value = cell_values.get((sheet_name.strip().lower(), cell_ref.strip().upper()))
    if value and value.status == "ok":
        return value.value
    return None


def _worksheet_by_normalized_name(sheet_by_name, normalized_name: str):
    for sheet_name, ws in sheet_by_name.items():
        if sheet_name.strip().lower() == normalized_name:
            return ws
    return None


def _object_comment(obj) -> str:
    lines = [
        f"Object: {obj.object_id}",
        f"Type: {obj.object_type}",
        f"Range: {obj.excel_range}",
    ]
    if obj.fixed_channel:
        lines.append(f"FixedChannel: {obj.fixed_channel}")
    if obj.formula:
        lines.append(f"Formula: {obj.formula}")
    return "\n".join(line for line in lines if line)


def _write_values_index(ws, formula_sheets, cell_values, Font, PatternFill, Alignment, Border, Side) -> None:
    headers = ["Sheet", "Cell", "Value", "Status", "Source"]
    ws.append(headers)
    for sheet_name, rows in formula_sheets.items():
        for row in rows:
            ws.append([sheet_name, row.excel_range, _coerce_value(row.value), row.status, row.detail or row.formula])
    for (sheet_name, cell_ref), value in sorted(cell_values.items()):
        ws.append([sheet_name, cell_ref, value.value, value.status, value.detail])
    _format_index_sheet(ws, headers, Font, PatternFill, Alignment, Border, Side)


def _write_object_index(ws, objects, Font, PatternFill, Alignment, Border, Side) -> None:
    headers = ["Sheet", "Range", "ObjectType", "ObjectId", "FixedChannel", "Formula", "PlotType", "TableType"]
    ws.append(headers)
    for obj in objects:
        ws.append([obj.sheet_name, obj.excel_range, obj.object_type, obj.object_id, obj.fixed_channel, obj.formula, obj.plot_type, obj.table_type])
    _format_index_sheet(ws, headers, Font, PatternFill, Alignment, Border, Side)


def _format_index_sheet(ws, headers, Font, PatternFill, Alignment, Border, Side) -> None:
    fill = PatternFill("solid", fgColor="D6E4F0")
    border = Border(bottom=Side(style="thin", color="D9D9D9"))
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [24, 12, 20, 16, 22, 70, 18, 18]
    for idx, width in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[_column_letter(idx)].width = width
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def _top_left_cell(excel_range: str) -> str:
    return (excel_range or "").split(":", 1)[0].upper()


def _coerce_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    try:
        return float(text)
    except ValueError:
        return text


def _safe_sheet_name(value: str) -> str:
    cleaned = "".join("_" if ch in "[]:*?/\\" else ch for ch in value).strip()
    return (cleaned or "Sheet")[:31]


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _windows_long_path(path: Path) -> str:
    resolved = path.resolve()
    text = str(resolved)
    if not text.startswith("\\\\?\\") and len(text) >= 240:
        return "\\\\?\\" + text
    return text
