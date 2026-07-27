from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

FOQ_TD_TITLE = "FOQ Test Description (FOQ_TD) - VX-C10-A"
FOQ_TD_SOURCE_FILE = "FOQ_Testdescription_VX-C10-A.docm"
FOQ_TD_SCOPE = "TCC common FOQ TD; document manager lists VH-C10-A and VC-C10-A"


@dataclass(frozen=True)
class FoqKbRunRow:
    order: int
    td_item: str
    injection: str
    instrument_method: str
    processing_method: str
    report_sheets: tuple[str, ...]
    td_intent: str
    method_contract: str
    report_contract: str
    design_questions: tuple[str, ...]


def foq_td_title() -> str:
    return FOQ_TD_TITLE


def foq_td_safe_stem() -> str:
    return "FOQ_TD_VX-C10-A"


def foq_td_source_summary() -> str:
    return f"{FOQ_TD_SOURCE_FILE}; {FOQ_TD_SCOPE}"


_ROW_CN: dict[int, dict[str, object]] = {
    1: {
        "item": "Column ID 芯片/位置识别",
        "intent": "确认四个 Column ID 位置能读到预期的 A/B/C/D 描述，证明接口和位置映射正确。",
        "method": "方法脚本主要负责提示插卡、等待 CardState 正常、读取并记录 Column_A 到 Column_D 的 Description；异常时中止。",
        "report": "报告用 AUDIT.Column_A-D.Description 直接判断 A/B/C/D 是否匹配。",
        "questions": ("生成独立方法时是否保留 IRC processing method，需要按 TD/CM 行为确认。",),
    },
    2: {
        "item": "Preheater 连接与响应",
        "intent": "确认左右 preheater 是否存在、memory state 是否正常、加热反馈和噪声是否符合要求。",
        "method": "方法脚本需要设置 preheater 温度、等待 45 C/55 C 等触发点、写 RetTime，并在结束时恢复临时 PID/控制参数。",
        "report": "报告消费 RetTime1..4、precond ModulePresent/MemoryState、温度平均值、响应和 noise 窗口。",
        "questions": ("临时 PID tuning 和 restore command 是脚本合同的一部分，不能只复制温度设定。",),
    },
    3: {
        "item": "阀和按键行为",
        "intent": "确认上下阀可以切换到目标位置，并确认 Fast Cool/按键相关动作可用。",
        "method": "方法脚本需要按顺序切换上下阀位置，记录 CurrentPosition/Precision，并包含必要的人工按键提示。",
        "report": "报告读取阀位置、precision 和 FastCoolState 的 audit 值。",
        "questions": ("是否存在上下阀、是否需要人工 keypad 步骤，取决于实际仪器配置。",),
    },
    4: {
        "item": "Burn-In 热稳定预处理",
        "intent": "通过高低温循环让 TCC 在后续测量前进入更稳定的热状态。",
        "method": "方法脚本按 model 分支执行高/低/高温循环，并检查外部温度计行为；主要是条件化设备。",
        "report": "当前 DB mapping 中不是主要结果行，更多是 audit 和后续测试前置证据。",
        "questions": ("短流程生成时能否省略 burn-in，必须由 TD 或工艺规则确认。",),
    },
    5: {
        "item": "Temperature Calibration 温度校准",
        "intent": "用外部上下温度计和内部 CC sensor 的比较写入校准点/偏差。",
        "method": "方法脚本需要按 setpoint ladder 运行，写 RetTime1..8 和 CCCalib.CalPoint/CalDev U/L audit 值。",
        "report": "报告读取 CC_Temp、RetTime delta、外部 drift 窗口以及温度校准点/偏差 audit 值。",
        "questions": ("生成 calibration 前必须先解完整逐点脚本表和阈值。",),
    },
    6: {
        "item": "Temperature Accuracy 温度准确度",
        "intent": "在每个 nominal setpoint 稳定后，用外部上下温度计测得温度与 nominal 比较，取偏差更大的通道。",
        "method": "方法脚本需要先到达前置温度状态，再切换到目标 setpoint；外部温度计稳定后写 RetTimeN，RetTimeN 是报告计算的锚点。",
        "report": "报告在 RetTimeN-1.0 到 RetTimeN-0.2 分钟窗口平均 ExtTemp_LowerCC/UpperCC，取最大偏差并和 Accuracy criterion 比较。",
        "questions": ("单点 accuracy 的前置温度/接近路径不能猜，例如 40 C 是否从 20 C 上升，需要 TD 或用户确认。",),
    },
    7: {
        "item": "Temperature Precision 与 Fan",
        "intent": "确认外部温度计重复性，并验证 fan/mode 行为。",
        "method": "方法脚本围绕 45/50 C 等状态采集外部温度计和 fan 行为，并切换温度/模式状态。",
        "report": "Precision 用上下温度计各自三个窗口的 range，取较大者；fan 由后续固定窗口/audit 判断。",
        "questions": ("fan-only 或 precision-only 生成前，需要补齐 fan report formula map。",),
    },
    8: {
        "item": "Temperature Stability 与 PCC",
        "intent": "验证 70 C 下 CC 稳定性；VH 路径还验证 PCC accuracy、drift 和 cooldown。",
        "method": "方法脚本设置 CC 70 C 和 PCC 相关温度，采集外部/PCC 通道，并在 PCC heat/cool 过程中写 RetTime。",
        "report": "Stability 用 45..60 分钟窗口的一分钟平均；PCC 用固定窗口、drift 和 RetTime4-RetTime3 cooldown。",
        "questions": ("非 VH 设备需要分离 CC-only 与 PCC 分支。",),
    },
    9: {
        "item": "HeatUp/CoolDown 升降温时间",
        "intent": "测量 20->50 C 升温和 50->20 C 降温时间，并扣除脚本中包含的稳定保持时间。",
        "method": "方法脚本写 RetTime1/3/4/6 分别作为升温开始、升温完成、降温开始、降温完成锚点。",
        "report": "报告计算 HeatUp=RetTime3-RetTime1-2.0，CoolDown=RetTime6-RetTime4-2.0。",
        "questions": ("如果脚本触发结构改变，2 分钟扣除规则也必须同步检查。",),
    },
    10: {
        "item": "Liquid Leak 漏液传感器",
        "intent": "通过人工加水确认 leak sensor 和 mute/cleanup 流程。",
        "method": "方法脚本提示加水，等待 LiquidLeak=Leak，记录状态，再提示 mute/清理并关闭传感器。",
        "report": "报告读取 AUDIT.LiquidLeak 和 precond.LiquidLeakCalibrationValue。",
        "questions": ("人工提示是测试逻辑的一部分，不能随意删除。",),
    },
    11: {
        "item": "Qualification Service Done 完成记录",
        "intent": "记录服务/qualification 完成证据。",
        "method": "方法脚本记录 Wellness service/qualification last-date 等属性。",
        "report": "主要是 metadata/audit 驱动，不依赖复杂 raw signal 公式。",
        "questions": ("需要确认哪些 CM property 需要写入，哪些只是记录。",),
    },
    12: {
        "item": "Factory Default 恢复默认/身份记录",
        "intent": "恢复服务提醒、日志、设备状态，并记录设备身份。",
        "method": "方法脚本进入 service mode，禁用/清理相关 interval/log/revision 字段，并提示最终物理检查。",
        "report": "报告读取 ModelNo、SerialNo、FirmwareVersion、HardwareVersion、ModuleHardwareRevision 等 metadata。",
        "questions": ("这是会改变状态的清理步骤，只有 procedure 需要时才纳入生成流程。",),
    },
    13: {
        "item": "Error Log Check 错误日志检查",
        "intent": "检查最终错误日志状态，并让设备处在安全的非运行状态。",
        "method": "方法脚本关闭 preheater/CC 温控，并复位相关连接/活动状态。",
        "report": "主要输出 error log/audit table，不是 FormulaObject raw signal 计算为主。",
        "questions": ("最终 pass/fail 解释规则仍需从 TD 中精确补齐。",),
    },
}


def foq_kb_row_chinese(row: FoqKbRunRow) -> dict[str, object]:
    return _ROW_CN.get(
        row.order,
        {"item": "", "intent": "", "method": "", "report": "", "questions": ()},
    )


def tcc_vh_foq_kb_rows() -> tuple[FoqKbRunRow, ...]:
    return (
        FoqKbRunRow(
            1,
            "Column ID",
            "ColumnIDs",
            "ColumnID",
            "CORRECT_STABILITY_INJ_INSERTION",
            ("Column ID",),
            "Verify that four column-ID slots read chip card descriptions A, B, C, and D in the expected ports.",
            "Prompt operator to insert cards; wait for CardState=OK; log Column_A-D.Description; abort if descriptions do not match slots.",
            "AUDIT.Column_A-D.Description(0,\"forward\") feed pass/fail cells for A/B/C/D slot identity.",
            ("Confirm whether standalone generated Column ID should keep the production IRC processing method.",),
        ),
        FoqKbRunRow(
            2,
            "Preheater Connection Test",
            "Preheater Connection Test",
            "PREHEATER",
            "No_Integration",
            ("Preheater Ports_Noise",),
            "Verify left/right preheater presence, memory state, heater feedback, thermal response, and noise.",
            "Set both preheaters to 40 C ready state, heat each toward 60 C, write RetTime1/2 at 45 C and RetTime3/4 at 55 C, restore PID values.",
            "Report uses RetTime1..4, precond ModulePresent/MemoryState, preheater/heater averages, max response, and chm.noise windows.",
            ("Confirm generated methods must preserve temporary PID tuning and restore commands.",),
        ),
        FoqKbRunRow(
            3,
            "Valve / Keypad",
            "Valve",
            "VALVES",
            "No_Integration",
            ("Valve_Keypad",),
            "Verify upper/lower valve switching and keypad Fast Cool / valve action behavior.",
            "Switch upper/lower valves 6_1 -> 1_2 -> 6_1; log precision; prompt operator for keypad actions and reconnect.",
            "Report reads AUDIT.UpperValve/LowerValve.CurrentPosition and Precision at fixed audit times plus FastCoolState.",
            ("Generated method requires matching valve hardware configuration; decide whether keypad/manual step is required.",),
        ),
        FoqKbRunRow(
            4,
            "Burn-In",
            "VTCC_BurnIn",
            "BURNIN",
            "NO_INTEGRATION",
            ("Temp Stability_Noise", "Fan"),
            "Thermally condition the TCC before downstream measurements because sensors can shift after heating.",
            "Branch by model limits; cycle high/low/high temperatures; abort if model or external thermometer behavior is invalid.",
            "Mostly conditioning and audit evidence; not a primary DB formula row in current mapping.",
            ("Decide when generated short procedures can omit burn-in without violating TD.",),
        ),
        FoqKbRunRow(
            5,
            "Temperature Calibration",
            "Temperature Calibration",
            "TEMPERATURE_CALIBRATION",
            "CORRECT_ACCURACY_INJ_INSERTION",
            ("Temp_Calib_Internal",),
            "Write CC calibration point/deviation audit values from internal CC sensors versus external upper/lower thermometers.",
            "Run model-specific setpoint ladder; write RetTime1..8 and CCCalib.CalPoint/CalDev U/L audit values; abort on large deviation.",
            "Report reads CC_Temp signal values, RetTime deltas, external drift windows, and audit TempCalibrationPoint/Deviation values.",
            ("Need exact point-by-point script table before generating calibration methods.",),
        ),
        FoqKbRunRow(
            6,
            "Temperature Accuracy",
            "Temperature Accuracy_H",
            "TEMPERATURE_ACCURACY",
            "ACCURACY_IRC_STOP_H",
            ("Temp Accuracy",),
            "At each nominal setpoint, wait until CC and external upper/lower thermometers are stable, then compare observed external temperature to nominal.",
            "Use stability state machine on external thermometers; write RetTimeN after the stable window for each setpoint.",
            "For each RetTimeN, report averages ExtTemp_LowerCC/UpperCC over RetTimeN-1.0..RetTimeN-0.2, chooses larger deviation, and compares to Temperature Accuracy criterion.",
            ("For custom single-point tests, baseline/approach temperature is a TD/design decision, not inferred automatically.",),
        ),
        FoqKbRunRow(
            7,
            "Temperature Precision and Fan",
            "Temperature Precision_and_Fan",
            "TEMPERATURE_PRECISION_AND_FAN",
            "CORRECT_STABILITY_INJ_INSERTION",
            ("Temp Precision", "Fan"),
            "Measure repeatability at repeated external thermometer windows and verify fan/mode behavior.",
            "Precondition around 45/50 C, acquire external thermometers and fan, toggle temperature and mode states.",
            "Precision uses three fixed windows per external thermometer; raw precision=max(range(lower), range(upper)); fan report uses later fixed windows.",
            ("Need full fan formula map before generating fan-only or precision-only variants.",),
        ),
        FoqKbRunRow(
            8,
            "Temperature Stability and PCC",
            "Temperature Stability_and_PCC_H",
            "TEMPERATURE_STABILITY_AND_PCC_70_H",
            "NO_INTEGRATION",
            ("Temp Stability_Noise", "PCC"),
            "Measure CC stability at 70 C and VH PCC behavior including accuracy, drift, and cooldown.",
            "Set CC 70 C and PCC 40 C, acquire external/PCC channels, emit PCC RetTimes during heat/cool sequence.",
            "Stability uses 45..60 min external one-minute averages; PCC uses fixed windows, drift 19..24, and RetTime4-RetTime3 cooldown.",
            ("Separate CC-only and PCC branches for non-VH generation.",),
        ),
        FoqKbRunRow(
            9,
            "HeatUp and CoolDown",
            "HeatUp and CoolDownTime",
            "TEMP_HEAT_UP_DOWN_20_50_20",
            "No_Integration",
            ("HeatUp&CoolDown",),
            "Measure heat-up 20->50 C and cooldown 50->20 C using internal and external readiness guards.",
            "Emit RetTime1 heat start, RetTime3 heat end, RetTime4 cooldown start, RetTime6 cooldown end after hold windows.",
            "Report calculates HeatUp=RetTime3-RetTime1-2.0 and CoolDown=RetTime6-RetTime4-2.0.",
            ("Preserve the 2 minute hold subtraction if trigger structure is preserved.",),
        ),
        FoqKbRunRow(
            10,
            "Liquid Leak",
            "LiquidLeaktest",
            "LIQUID LEAK",
            "No_Integration",
            ("Liquid Leak Test",),
            "Verify leak sensor and alarm mute workflow with manual water addition.",
            "Prompt operator to add water, wait LiquidLeak=Leak, log LiquidLeak, prompt mute/cleanup, switch sensor off.",
            "Report reads AUDIT.LiquidLeak(100,\"backward\") and precond.LiquidLeakCalibrationValue.",
            ("Manual operator prompts are part of the test, not optional decorations.",),
        ),
        FoqKbRunRow(
            11,
            "Qualification Service Done",
            "Qualification_Service_Done",
            "Qualification_Service_Done",
            "No_Integration",
            ("Internal Use",),
            "Record service/qualification completion evidence after tests.",
            "Log ColumnComp_Wellness Service/Qualification last-date fields.",
            "Metadata/audit driven; no heavy raw signal report calculation.",
            ("Confirm exact CM properties that must be written versus only logged.",),
        ),
        FoqKbRunRow(
            12,
            "Factory Default",
            "Factory Default",
            "FACTORYDEFAULT",
            "No_Integration",
            ("Definitions", "Internal Use"),
            "Return service/reminder/log/device metadata to expected factory/default state and record identity.",
            "Enter service mode, disable service/qualification intervals, clear logs/revision fields, prompt final physical checks.",
            "Report uses seq/precond/audit metadata such as ModelNo, SerialNo, FirmwareVersion, HardwareVersion, ModuleHardwareRevision.",
            ("Treat as state-mutating cleanup; include only when procedure requires it.",),
        ),
        FoqKbRunRow(
            13,
            "Error Log Check",
            "Error Log Check",
            "CHECKERRORLOG",
            "No_Integration",
            ("Error Log",),
            "Check final error-log state and leave device in a safe non-running condition.",
            "Turn preheater and CC temp controls off; reset relevant connection/activity state in full procedure.",
            "Mostly error-log/audit table output, not formula-object driven.",
            ("Need final TD rule for pass/fail interpretation of error-log table.",),
        ),
    )


def write_foq_kb_to_run_alignment_workbook(rows: tuple[FoqKbRunRow, ...], output_path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to export FOQ KB alignment workbooks.") from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "FOQ KB to Run"
    headers = (
        "Order",
        "TD Title",
        "TD Item",
        "中文测试项",
        "Injection",
        "Instrument Method",
        "Processing Method",
        "Report Sheets",
        "TD Intent",
        "中文测试逻辑",
        "Method Contract",
        "方法脚本实现关系",
        "Report Contract",
        "报告计算关系",
        "Design Questions",
        "需人工确认",
    )
    ws.append(headers)
    for row in rows:
        cn = foq_kb_row_chinese(row)
        ws.append(
            (
                row.order,
                FOQ_TD_TITLE,
                row.td_item,
                cn["item"],
                row.injection,
                row.instrument_method,
                row.processing_method,
                ", ".join(row.report_sheets),
                row.td_intent,
                cn["intent"],
                row.method_contract,
                cn["method"],
                row.report_contract,
                cn["report"],
                "\n".join(row.design_questions),
                "\n".join(cn["questions"]),
            )
        )
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    ws.freeze_panes = "A2"
    widths = {
        "A": 8,
        "B": 38,
        "C": 28,
        "D": 34,
        "E": 32,
        "F": 36,
        "G": 32,
        "H": 34,
        "I": 58,
        "J": 58,
        "K": 78,
        "L": 78,
        "M": 78,
        "N": 78,
        "O": 58,
        "P": 58,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    wb.save(output_path)
    return output_path
