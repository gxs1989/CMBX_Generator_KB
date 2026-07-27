from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from pathlib import Path

from foq_result_locations import FoqResultLocation
from report_calculation_map import ReportCellCalculation
from report_formula_evaluator import FormulaEvaluation
from report_workbook_builder import build_accuracy_rows, build_heatup_cooldown_report


@dataclass(frozen=True)
class ReportCellValue:
    sheet_name: str
    cell: str
    value: object
    status: str
    detail: str


@dataclass(frozen=True)
class FoqContractValue:
    location: FoqResultLocation
    injection_name: str
    value: object
    status: str
    detail: str


@dataclass(frozen=True)
class FoqDependencyTrace:
    db_field: str
    final_cell: str
    status: str
    calculation_tier: str
    definition_source: str
    intermediate_cells: str
    cm_formula_cells: str
    raw_data_sources: str
    trace_summary: str
    next_parser_work: str


def build_report_cell_value_map(
    formula_sheets: dict[str, list[FormulaEvaluation]],
    calculation_map: list[ReportCellCalculation] | None = None,
) -> dict[tuple[str, str], ReportCellValue]:
    values: dict[tuple[str, str], ReportCellValue] = {}
    for sheet_name, rows in formula_sheets.items():
        for row in rows:
            if not row.excel_range or ":" in row.excel_range:
                continue
            values[_key(sheet_name, row.excel_range)] = ReportCellValue(
                sheet_name=sheet_name,
                cell=row.excel_range.upper(),
                value=_coerce_value(row.value),
                status=row.status,
                detail=_formula_detail(row),
            )
        normalized = sheet_name.lower()
        if normalized == "temp accuracy":
            values.update(_temp_accuracy_cell_values(sheet_name, rows))
        elif "heatup" in normalized and "cooldown" in normalized:
            values.update(_heatup_cooldown_cell_values(sheet_name, rows, calculation_map or []))
        elif normalized == "temp precision":
            values.update(_temp_precision_cell_values(sheet_name, rows, calculation_map or []))
        elif normalized == "temp stability_noise":
            values.update(_temp_stability_noise_cell_values(sheet_name, rows, calculation_map or []))
        elif normalized == "pcc":
            values.update(_pcc_cell_values(sheet_name, rows, calculation_map or []))
        elif normalized == "preheater ports_noise":
            values.update(_preheater_ports_noise_cell_values(sheet_name, rows))
        elif normalized == "temp_calib_internal":
            values.update(_temp_calib_internal_cell_values(sheet_name, rows))
        elif normalized == "column id":
            values.update(_column_id_cell_values(sheet_name, rows))
    values.update(_cross_sheet_cell_values(values))
    return _normalize_cell_values(values)


def resolve_contract_values(
    locations: list[FoqResultLocation],
    cell_values_by_injection: dict[str, dict[tuple[str, str], ReportCellValue]],
    injection_by_report_file: dict[str, str],
) -> list[FoqContractValue]:
    rows: list[FoqContractValue] = []
    for location in locations:
        injection_name = injection_by_report_file.get(_normalize_report_file(location.report_file), "")
        if not injection_name:
            rows.append(FoqContractValue(location, "", "", "missing_injection", "No injection matched xlsReportFile"))
            continue
        cell_values = cell_values_by_injection.get(injection_name, {})
        cell_value = cell_values.get(_key(location.report_sheet, location.report_cell))
        if not cell_value:
            rows.append(FoqContractValue(location, injection_name, "", "missing_cell", "Cell is not evaluated yet"))
            continue
        rows.append(FoqContractValue(location, injection_name, cell_value.value, cell_value.status, cell_value.detail))
    return rows


def write_foq_contract_workbook(
    path: str | Path,
    device_type: str,
    mapping_sheet: str,
    values: list[FoqContractValue],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build FOQ contract workbooks.") from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    db_ws = workbook.active
    db_ws.title = "DB Data"
    coverage_ws = workbook.create_sheet("Mapping Coverage")
    values_ws = workbook.create_sheet("Cell Values")
    trace_ws = workbook.create_sheet("Dependency Trace")

    header_fill = PatternFill("solid", fgColor="D6E4F0")
    ok_fill = PatternFill("solid", fgColor="D9EAD3")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Calibri", size=10, bold=True)
    text_font = Font(name="Calibri", size=10)

    display_values = [_display_value_for_db_field(row.location.db_field, row.value) for row in values]
    fields = ["DeviceType", "MappingSheet"] + [row.location.db_field for row in values]
    data = [device_type, mapping_sheet] + display_values
    db_ws.append(fields)
    db_ws.append(data)
    for cell in db_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in db_ws[2]:
        cell.font = text_font
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        if _is_date_field(str(db_ws.cell(row=1, column=cell.column).value or "")):
            cell.number_format = "yyyy-mm-dd"
    db_ws.freeze_panes = "A2"

    coverage_headers = [
        "dbField",
        "Value",
        "Status",
        "Injection",
        "ReportFile",
        "ReportSheet",
        "ReportCell",
        "Unit",
        "Description",
        "Detail",
    ]
    values_ws.append(coverage_headers)
    trace_headers = [
        "dbField",
        "Final Report Cell",
        "Status",
        "Calculation Tier",
        "Definition / Criterion Source",
        "Intermediate Cells",
        "CM Formula Cells",
        "Raw Data Sources",
        "Reverse Trace Summary",
        "Next Parser Work",
    ]
    trace_ws.append(trace_headers)
    coverage_ws.append(["Metric", "Value"])
    status_counts: dict[str, int] = {}
    for row in values:
        status_counts[row.status] = status_counts.get(row.status, 0) + 1
        values_ws.append(
            [
                row.location.db_field,
                _display_value_for_db_field(row.location.db_field, row.value),
                row.status,
                row.injection_name,
                row.location.report_file,
                row.location.report_sheet,
                row.location.report_cell,
                row.location.unit,
                row.location.description,
                row.detail,
            ]
        )
        trace = dependency_trace_for_contract_value(row)
        trace_ws.append(
            [
                trace.db_field,
                trace.final_cell,
                trace.status,
                trace.calculation_tier,
                trace.definition_source,
                trace.intermediate_cells,
                trace.cm_formula_cells,
                trace.raw_data_sources,
                trace.trace_summary,
                trace.next_parser_work,
            ]
        )
    for row_cells in values_ws.iter_rows(min_row=2):
        if _is_date_field(str(row_cells[0].value or "")):
            row_cells[1].number_format = "yyyy-mm-dd"
    coverage_rows = [
        ("DeviceType", device_type),
        ("MappingSheet", mapping_sheet),
        ("Mapped Fields", len(values)),
        ("OK Fields", status_counts.get("ok", 0)),
        ("Missing Injection", status_counts.get("missing_injection", 0)),
        ("Missing Cell", status_counts.get("missing_cell", 0)),
        ("Unsupported/Error", sum(count for status, count in status_counts.items() if status not in {"ok", "missing_injection", "missing_cell"})),
    ]
    for row in coverage_rows:
        coverage_ws.append(row)

    for ws in (coverage_ws, values_ws, trace_ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows(min_row=2):
            status = str(row_cells[2].value if ws.title in {"Cell Values", "Dependency Trace"} and len(row_cells) > 2 else "").lower()
            for cell in row_cells:
                cell.font = text_font
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ws.title in {"Cell Values", "Dependency Trace"}:
                fill = ok_fill if status == "ok" else warn_fill if status.startswith("missing") else fail_fill
                for cell in row_cells:
                    cell.fill = fill
        ws.freeze_panes = "A2"

    _fit_columns(db_ws, 28)
    _fit_columns(coverage_ws, 32)
    _fit_columns(values_ws, 48)
    _fit_columns(trace_ws, 64)
    workbook.save(_windows_long_path(output_path))
    return output_path


def dependency_trace_for_contract_value(row: FoqContractValue) -> FoqDependencyTrace:
    location = row.location
    sheet = location.report_sheet
    cell = location.report_cell.upper()
    final_cell = f"{location.report_file} | {sheet}!{cell}"
    status = row.status
    if status == "missing_injection":
        return FoqDependencyTrace(
            location.db_field,
            final_cell,
            status,
            "0 - report not found",
            "",
            "",
            "",
            "",
            "The DB contract points to a report file/injection that was not found in the CMBX package.",
            "Improve xlsReportFile to injection-name matching or verify this test exists in the CMBX.",
        )
    lowered_sheet = sheet.lower()
    if lowered_sheet == "temp accuracy":
        return _temp_accuracy_trace(row, final_cell)
    if "heatup" in lowered_sheet and "cooldown" in lowered_sheet:
        return _heatup_cooldown_trace(row, final_cell)
    if lowered_sheet == "temp precision":
        return _derived_sheet_trace(
            row,
            final_cell,
            "2 - derived from precision replicate cells" if row.status == "ok" else "3 - workbook-derived cell not mapped yet",
            "Definitions!Temperature Precision = 0.1",
            "Temp Precision!K65:L67",
            "Temp Precision!K65:L67 chm.sig_value average windows",
            "ExtTemp_LowerCC and ExtTemp_UpperCC raw channels",
            "Temperature precision is calculated as the larger of the lower-sensor repeatability range and upper-sensor repeatability range, then displayed to the report cell precision and compared with the criterion using the raw range.",
        )
    if lowered_sheet == "temp stability_noise":
        return _derived_sheet_trace(
            row,
            final_cell,
            "2 - derived from stability segment cells" if row.status == "ok" else "3 - workbook-derived cell not mapped yet",
            "Definitions!Temperature Stability = 0.05",
            "Temp Stability_Noise!K61:L75",
            "Temp Stability_Noise!K61:L75 chm.sig_value one-minute average windows",
            "ExtTemp_LowerCC and ExtTemp_UpperCC raw channels; CC_Temp/PCC_Temp raw channels for noise cells",
            "Temperature stability is calculated as the larger of the lower-sensor segment range and upper-sensor segment range, then displayed to the report cell precision and compared with the criterion using the raw range.",
        )
    if lowered_sheet == "pcc":
        return _derived_sheet_trace(
            row,
            final_cell,
            "2 - derived from PCC RetTime cells" if row.status == "ok" else "3 - workbook-derived cell not mapped yet",
            "Definitions!PCC CoolDownTime = 2 min",
            "PCC!K105/L105",
            "PCC!K105 RetTime3; PCC!L105 RetTime4",
            "Audit RetTimes from injection audit trail; PCC_Temp raw channel for PCC accuracy/drift fields",
            "PCC performance is calculated from the RetTime interval and compared with the PCC cooldown criterion.",
        )
    if lowered_sheet == "preheater ports_noise" and row.location.report_cell.upper() in {"C26", "C27"}:
        return _derived_sheet_trace(
            row,
            final_cell,
            "3 - workbook-derived preheater result cell not mapped yet",
            "",
            "Preheater Ports_Noise!C26:C27",
            "Preheater Ports_Noise!J72:J73/K72:K73 RetTime formulas; J117:K118 module/memory-state formulas",
            "Audit RetTimes and precondition/module state metadata",
            "Preheater port result cells are workbook pass/fail cells. The likely source cells are the RetTime pair and module state cells, but the exact pass/fail condition still needs confirmation.",
        )
    if lowered_sheet == "preheater ports_noise":
        return _derived_sheet_trace(
            row,
            final_cell,
            "2 - derived from preheater temperature cells" if row.status == "ok" else "3 - workbook-derived cell not mapped yet",
            "",
            "Preheater Ports_Noise!J82:K83",
            "Preheater temperature and heater-temperature chm.sig_value average windows",
            "PrehtLeft_Temp, PrehtRight_Temp, PREH_L_HeaterTemp_Actual, PREH_R_HeaterTemp_Actual raw channels",
            "Preheater difference cells are calculated as heater temperature average minus external preheater temperature average.",
        )
    if lowered_sheet == "temp_calib_internal":
        return _derived_sheet_trace(
            row,
            final_cell,
            "2 - compatibility alias to 5 deg C calibration cells" if row.status == "ok" else "3 - workbook-derived cell not mapped yet",
            "",
            "Temp_Calib_Internal!J15:J17",
            "RetTime8 duration and external upper/lower drift formulas",
            "Audit RetTimes and ExtTemp_UpperCC / ExtTemp_LowerCC raw channels",
            "The DB contract expects the 5 deg C calibration values at C22/D22/E22. In this report template the corresponding CM formula cells are placed at J15/J16/J17.",
        )
    if row.detail and ("chm." in row.detail or "AUDIT." in row.detail):
        return FoqDependencyTrace(
            location.db_field,
            final_cell,
            status,
            "1 - direct CM formula cell",
            "",
            "",
            f"{sheet}!{cell}: {row.detail}",
            _raw_source_hint(row.detail),
            "The final DB field is backed directly by a report SheetObject formula that can be evaluated from audit/raw signal data.",
            "",
        )
    if lowered_sheet == "column id":
        return _derived_sheet_trace(
            row,
            final_cell,
            "3 - workbook-derived result cell not mapped yet",
            "",
            "Column ID!C26:C29",
            "Column ID!L46:L49 audit Column_A-D.Description formulas",
            "Column description values from injection audit trail",
            "Column ID result cells are workbook pass/fail cells. The likely source cells are L46:L49, but the pass/fail condition still needs to be mapped from the embedded workbook.",
        )
    if lowered_sheet == "definitions" and cell == "J8":
        return _derived_sheet_trace(
            row,
            final_cell,
            "2 - derived from model and module revision cells" if row.status == "ok" else "3 - workbook-derived cell not mapped yet",
            "",
            "Definitions!C15; Internal Use!E12",
            "Definitions!C15 AUDIT.ColumnComp.ModelNo; Internal Use!E12 precond.ColumnComp.ModuleHardwareRevision",
            "ColumnComp model number and module hardware revision from injection audit precondition",
            "ModelVariant is reconstructed from the module hardware revision, matching the Definitions acceptance-criteria variant labels.",
        )
    if lowered_sheet == "internal use" and cell == "F10":
        return _derived_sheet_trace(
            row,
            final_cell,
            "2 - derived from sequence name and device serial number" if row.status == "ok" else "3 - workbook-derived cell not mapped yet",
            "",
            "Definitions!D15; Internal Use!B20",
            "Definitions!D15 precond.ColumnComp.SerialNo; Internal Use!B20 seq.name",
            "ColumnComp serial number from audit precondition and CMBX sequence name",
            "Serial-number check passes when the device serial number matches the sequence name used for the factory run.",
        )
    if status == "missing_cell":
        return FoqDependencyTrace(
            location.db_field,
            final_cell,
            status,
            "3 - workbook-derived cell not mapped yet",
            _definition_hint(sheet, cell),
            "",
            "",
            "",
            "The target cell is listed by the DB contract, but no direct SheetObject formula or known derived-cell rule currently produces it.",
            "Add a sheet-specific reverse rule or parse the embedded FormulaOne workbook formula for this cell.",
        )
    return FoqDependencyTrace(
        location.db_field,
        final_cell,
        status,
        "2 - calculated/metadata cell",
        _definition_hint(sheet, cell),
        "",
        row.detail,
        _raw_source_hint(row.detail),
        "The field is populated by a known calculation rule or metadata formula.",
        "" if status == "ok" else "Extend evaluator support for this formula or metadata source.",
    )


def _temp_accuracy_trace(row: FoqContractValue, final_cell: str) -> FoqDependencyTrace:
    cell = row.location.report_cell.upper()
    sheet = row.location.report_sheet
    if cell.startswith("D") and cell[1:].isdigit() and 66 <= int(cell[1:]) <= 70:
        excel_row = int(cell[1:])
        ret_index = excel_row - 65
        return FoqDependencyTrace(
            row.location.db_field,
            final_cell,
            row.status,
            "2 - derived from CM formula cells",
            "Definitions!Temperature Accuracy criterion is used for related pass/fail cells.",
            f"{sheet}!B{excel_row}/C{excel_row}/N{excel_row}",
            f"{sheet}!I{excel_row} nominal setpoint; {sheet}!K{excel_row} RetTime{ret_index}; {sheet}!L{excel_row}/{sheet}!M{excel_row} chm.sig_value average windows",
            "Audit RetTimeN and ColumnComp.CC.Temperature.Nominal; ExtTemp_LowerCC and ExtTemp_UpperCC raw channels",
            "DB field reads the deviation cell. The deviation is observed max-deviation temperature minus adjusted temperature; observed temperature is selected from lower/upper external thermometer averages.",
            "",
        )
    if cell == "J56":
        return FoqDependencyTrace(
            row.location.db_field,
            final_cell,
            row.status,
            "2 - compatibility alias",
            "Definitions!Temperature Accuracy criterion is used for related pass/fail cells.",
            f"{sheet}!D66",
            f"{sheet}!I66/K66/L66/M66",
            "Audit RetTime1 and external thermometer raw channels",
            "This legacy mapping points to J56, but the generated contract value aliases the first accuracy deviation to the 10 deg C row calculation.",
            "Confirm the native FormulaOne J56 link if exact legacy cell placement is required.",
        )
    if cell == "E26":
        return FoqDependencyTrace(
            row.location.db_field,
            final_cell,
            row.status,
            "2 - pass/fail derived from summary",
            "Definitions!Temperature Accuracy = 0.5",
            f"{sheet}!D26",
            f"{sheet}!D66:D70 derived deviations",
            "Audit RetTimes; nominal temperature audit values; ExtTemp_LowerCC and ExtTemp_UpperCC raw channels",
            "Pass/fail is derived by comparing the observed maximum absolute deviation with the temperature accuracy criterion.",
            "",
        )
    return FoqDependencyTrace(
        row.location.db_field,
        final_cell,
        row.status,
        "1 - direct or known temperature accuracy cell",
        "Definitions!Temperature Accuracy criterion where applicable",
        "",
        row.detail,
        _raw_source_hint(row.detail),
        "Temperature accuracy field resolved by the current accuracy reverse rule.",
        "",
    )


def _heatup_cooldown_trace(row: FoqContractValue, final_cell: str) -> FoqDependencyTrace:
    cell = row.location.report_cell.upper()
    sheet = row.location.report_sheet
    if cell in {"D26", "D65", "E26"}:
        return FoqDependencyTrace(
            row.location.db_field,
            final_cell,
            row.status,
            "2 - derived from audit RetTime cells",
            "Definitions!HeatUp & Cool Down = 15 min",
            f"{sheet}!D65 -> {sheet}!D26; {sheet}!E26 pass/fail",
            f"{sheet}!J65 RetTime1 start; {sheet}!K65 RetTime3 end",
            "Audit RetTimes from injection audit trail",
            "Heat-up observed time is end RetTime minus start RetTime minus the 2 minute stable-hold subtraction, then compared with the criterion.",
            "",
        )
    if cell in {"D27", "D66", "E27"}:
        return FoqDependencyTrace(
            row.location.db_field,
            final_cell,
            row.status,
            "2 - derived from audit RetTime cells",
            "Definitions!HeatUp & Cool Down = 15 min",
            f"{sheet}!D66 -> {sheet}!D27; {sheet}!E27 pass/fail",
            f"{sheet}!L65 RetTime4 start; {sheet}!M65 RetTime6 end",
            "Audit RetTimes from injection audit trail",
            "Cool-down observed time is end RetTime minus start RetTime minus the 2 minute stable-hold subtraction, then compared with the criterion.",
            "",
        )
    return FoqDependencyTrace(
        row.location.db_field,
        final_cell,
        row.status,
        "1 - direct heatup/cooldown formula cell",
        "Definitions!HeatUp & Cool Down where applicable",
        "",
        row.detail,
        "Audit RetTimes",
        "HeatUp/CoolDown field resolved by direct RetTime formula or known derived rule.",
        "",
    )


def _derived_sheet_trace(
    row: FoqContractValue,
    final_cell: str,
    calculation_tier: str,
    definition_source: str,
    intermediate_cells: str,
    cm_formula_cells: str,
    raw_data_sources: str,
    summary: str,
) -> FoqDependencyTrace:
    return FoqDependencyTrace(
        row.location.db_field,
        final_cell,
        row.status,
        calculation_tier,
        definition_source,
        intermediate_cells,
        cm_formula_cells,
        raw_data_sources,
        summary if row.status == "ok" else "The target cell is listed by the DB contract, but the sheet-specific reverse rule did not produce a value.",
        "" if row.status == "ok" else "Add the missing source-cell rule or parse the embedded FormulaOne workbook formula for this cell.",
    )


def _definition_hint(sheet: str, cell: str) -> str:
    lowered = sheet.lower()
    if lowered == "temp accuracy":
        return "Definitions!Temperature Accuracy"
    if "heatup" in lowered and "cooldown" in lowered:
        return "Definitions!HeatUp & Cool Down"
    if "stability" in lowered:
        return "Definitions!Temperature Stability"
    if "precision" in lowered:
        return "Definitions!Temperature Precision"
    if lowered == "pcc":
        return "Definitions!PCC T Accuracy / PCC T Drift / PCC CoolDownTime"
    return ""


def _raw_source_hint(detail: str) -> str:
    text = detail or ""
    sources = []
    if "chm." in text:
        sources.append("raw signal channel")
    if "AUDIT.RetTime" in text:
        sources.append("audit RetTime")
    if "AUDIT." in text and "AUDIT.RetTime" not in text:
        sources.append("audit property history")
    return "; ".join(sources)


def _temp_accuracy_cell_values(sheet_name: str, rows: list[FormulaEvaluation]) -> dict[tuple[str, str], ReportCellValue]:
    values: dict[tuple[str, str], ReportCellValue] = {}
    accuracy_rows = build_accuracy_rows(rows)
    max_dev = max((_summary_abs_deviation(row) for row in accuracy_rows if row.observed_max_deviation_temp_c is not None), default=None)
    overall = "Test passed" if max_dev is not None and max_dev <= 0.5 else "Test failed"
    values[_key(sheet_name, "C26")] = ReportCellValue(sheet_name, "C26", 0.5, "ok", "Temperature accuracy criterion")
    values[_key(sheet_name, "D26")] = ReportCellValue(sheet_name, "D26", max_dev, "ok", "max(abs(D66:D70))")
    values[_key(sheet_name, "E26")] = ReportCellValue(sheet_name, "E26", overall, "ok", "D26 <= C26")
    for index, row in enumerate(accuracy_rows, 66):
        observed = row.observed_max_deviation_temp_c
        deviation = (observed - row.adjusted_temperature_c) if observed is not None else None
        row_result = "ok" if deviation is not None and abs(deviation) <= 0.5 else "failed"
        values[_key(sheet_name, f"B{index}")] = ReportCellValue(sheet_name, f"B{index}", row.adjusted_temperature_c, "ok", "Adjusted temperature")
        values[_key(sheet_name, f"C{index}")] = ReportCellValue(sheet_name, f"C{index}", observed, "ok", "Observed max-deviation temperature")
        values[_key(sheet_name, f"D{index}")] = ReportCellValue(sheet_name, f"D{index}", deviation, "ok", "Observed - adjusted")
        values[_key(sheet_name, f"E{index}")] = ReportCellValue(sheet_name, f"E{index}", row_result, "ok", "abs(report-cell deviation) <= criterion")
        values[_key(sheet_name, f"N{index}")] = ReportCellValue(sheet_name, f"N{index}", observed, "ok", "Observed temperature with maximum deviation")
    if accuracy_rows:
        first = accuracy_rows[0]
        observed = first.observed_max_deviation_temp_c
        deviation = (observed - first.adjusted_temperature_c) if observed is not None else None
        values[_key(sheet_name, "J56")] = ReportCellValue(sheet_name, "J56", deviation, "ok", "Compatibility alias for TempAcc10 raw deviation")
    return values


def _heatup_cooldown_cell_values(
    sheet_name: str,
    rows: list[FormulaEvaluation],
    calculation_map: list[ReportCellCalculation],
) -> dict[tuple[str, str], ReportCellValue]:
    report = build_heatup_cooldown_report(rows, calculation_map)
    values = _single_cell_values(rows)
    heatup_raw = _delta_minus_two(values, "J66", "K66")
    if heatup_raw is None:
        heatup_raw = report.heatup_observed_min
    cooldown_raw = _delta_minus_two(values, "L66", "M66")
    if cooldown_raw is None:
        cooldown_raw = report.cooldown_observed_min
    heatup_observed = _cm_display_number(heatup_raw, 1)
    cooldown_observed = _cm_display_number(cooldown_raw, 1)
    heatup_result = "Test passed" if heatup_raw is not None and heatup_raw <= report.criterion_min else "Test failed"
    cooldown_result = "Test passed" if cooldown_raw is not None and cooldown_raw <= report.criterion_min else "Test failed"
    return {
        _key(sheet_name, "C26"): ReportCellValue(sheet_name, "C26", report.criterion_min, "ok", "Heat-up criterion"),
        _key(sheet_name, "D26"): ReportCellValue(sheet_name, "D26", heatup_observed, "ok" if heatup_raw is not None else "missing_cell", "K66 - J66 - 2 min; displayed to 1 decimal"),
        _key(sheet_name, "E26"): ReportCellValue(sheet_name, "E26", heatup_result, "ok", "D26 <= C26"),
        _key(sheet_name, "C27"): ReportCellValue(sheet_name, "C27", report.criterion_min, "ok", "Cool-down criterion"),
        _key(sheet_name, "D27"): ReportCellValue(sheet_name, "D27", cooldown_observed, "ok" if cooldown_raw is not None else "missing_cell", "M66 - L66 - 2 min; displayed to 1 decimal"),
        _key(sheet_name, "E27"): ReportCellValue(sheet_name, "E27", cooldown_result, "ok", "D27 <= C27"),
        _key(sheet_name, "D65"): ReportCellValue(sheet_name, "D65", heatup_observed, "ok" if heatup_raw is not None else "missing_cell", "K66 - J66 - 2 min; displayed to 1 decimal"),
        _key(sheet_name, "D66"): ReportCellValue(sheet_name, "D66", cooldown_observed, "ok" if cooldown_raw is not None else "missing_cell", "M66 - L66 - 2 min; displayed to 1 decimal"),
    }


def _temp_precision_cell_values(
    sheet_name: str,
    rows: list[FormulaEvaluation],
    calculation_map: list[ReportCellCalculation],
) -> dict[tuple[str, str], ReportCellValue]:
    values = _single_cell_values(rows)
    lower_range = _cell_range(values, ["K65", "K66", "K67"])
    upper_range = _cell_range(values, ["L65", "L66", "L67"])
    precision_raw = max([value for value in [lower_range, upper_range] if value is not None], default=None)
    precision = _cm_display_number(precision_raw, 2)
    criterion = _criterion_from_calculation_map(calculation_map, "Temperature Precision", 0.1)
    result = _pass_fail(precision_raw, criterion)
    return {
        _key(sheet_name, "C26"): ReportCellValue(sheet_name, "C26", criterion, "ok", "Temperature precision criterion"),
        _key(sheet_name, "D26"): ReportCellValue(sheet_name, "D26", precision, "ok" if precision_raw is not None else "missing_cell", "max(range(K65:K67), range(L65:L67)); displayed to 2 decimals"),
        _key(sheet_name, "E26"): ReportCellValue(sheet_name, "E26", result, "ok" if precision_raw is not None else "missing_cell", "raw D26 <= C26"),
    }


def _temp_stability_noise_cell_values(
    sheet_name: str,
    rows: list[FormulaEvaluation],
    calculation_map: list[ReportCellCalculation],
) -> dict[tuple[str, str], ReportCellValue]:
    values = _single_cell_values(rows)
    lower_range = _cell_range(values, [f"K{row}" for row in range(61, 76)])
    upper_range = _cell_range(values, [f"L{row}" for row in range(61, 76)])
    stability_raw = max([value for value in [lower_range, upper_range] if value is not None], default=None)
    stability = _cm_display_number(stability_raw, 2)
    criterion = _criterion_from_calculation_map(calculation_map, "Temperature Stability", 0.05)
    result = _pass_fail(stability_raw, criterion)
    return {
        _key(sheet_name, "C26"): ReportCellValue(sheet_name, "C26", criterion, "ok", "Temperature stability criterion"),
        _key(sheet_name, "D26"): ReportCellValue(sheet_name, "D26", stability, "ok" if stability_raw is not None else "missing_cell", "max(range(K61:K75), range(L61:L75)); displayed to 2 decimals"),
        _key(sheet_name, "E26"): ReportCellValue(sheet_name, "E26", result, "ok" if stability_raw is not None else "missing_cell", "raw D26 <= C26"),
    }


def _pcc_cell_values(
    sheet_name: str,
    rows: list[FormulaEvaluation],
    calculation_map: list[ReportCellCalculation],
) -> dict[tuple[str, str], ReportCellValue]:
    values = _single_cell_values(rows)
    start = _to_float(values.get("K105").value if values.get("K105") else None)
    end = _to_float(values.get("L105").value if values.get("L105") else None)
    cooldown_raw = (end - start) if start is not None and end is not None else None
    cooldown = _cm_display_number(cooldown_raw, 2)
    criterion = _criterion_from_calculation_map(calculation_map, "PCC CoolDownTime", 2.0)
    result = _pass_fail(cooldown_raw, criterion)
    return {
        _key(sheet_name, "C26"): ReportCellValue(sheet_name, "C26", criterion, "ok", "PCC cooldown criterion"),
        _key(sheet_name, "D26"): ReportCellValue(sheet_name, "D26", cooldown, "ok" if cooldown_raw is not None else "missing_cell", "L105 - K105; displayed to 2 decimals"),
        _key(sheet_name, "E26"): ReportCellValue(sheet_name, "E26", result, "ok" if cooldown_raw is not None else "missing_cell", "raw D26 <= C26"),
    }


def _preheater_ports_noise_cell_values(sheet_name: str, rows: list[FormulaEvaluation]) -> dict[tuple[str, str], ReportCellValue]:
    values = _single_cell_values(rows)
    left_temp = _to_float(values.get("J82").value if values.get("J82") else None)
    left_heater = _to_float(values.get("K82").value if values.get("K82") else None)
    right_temp = _to_float(values.get("J83").value if values.get("J83") else None)
    right_heater = _to_float(values.get("K83").value if values.get("K83") else None)
    left_diff_raw = (left_heater - left_temp) if left_temp is not None and left_heater is not None else None
    right_diff_raw = (right_heater - right_temp) if right_temp is not None and right_heater is not None else None
    left_diff = _cm_display_number(left_diff_raw, 1)
    right_diff = _cm_display_number(right_diff_raw, 1)
    left_result = _preheater_port_result(values, "J72", "K72", "J117", "K117")
    right_result = _preheater_port_result(values, "J73", "K73", "J118", "K118")
    return {
        _key(sheet_name, "C26"): ReportCellValue(sheet_name, "C26", left_result, "ok" if left_result else "missing_cell", "RetTime present and preheater left ModulePresent/MemoryState check"),
        _key(sheet_name, "C27"): ReportCellValue(sheet_name, "C27", right_result, "ok" if right_result else "missing_cell", "RetTime present and preheater right ModulePresent/MemoryState check"),
        _key(sheet_name, "L82"): ReportCellValue(sheet_name, "L82", left_diff, "ok" if left_diff_raw is not None else "missing_cell", "K82 - J82; displayed to 1 decimal"),
        _key(sheet_name, "L83"): ReportCellValue(sheet_name, "L83", right_diff, "ok" if right_diff_raw is not None else "missing_cell", "K83 - J83; displayed to 1 decimal"),
    }


def _column_id_cell_values(sheet_name: str, rows: list[FormulaEvaluation]) -> dict[tuple[str, str], ReportCellValue]:
    values = _single_cell_values(rows)
    mapping = {
        "C26": ("L46", "A"),
        "C27": ("L47", "B"),
        "C28": ("L48", "C"),
        "C29": ("L49", "D"),
    }
    result: dict[tuple[str, str], ReportCellValue] = {}
    for target_cell, (source_cell, expected) in mapping.items():
        source_value = values.get(source_cell)
        if not source_value or source_value.status != "ok":
            result[_key(sheet_name, target_cell)] = ReportCellValue(
                sheet_name,
                target_cell,
                "",
                "missing_cell",
                f"Column ID source {source_cell} is not evaluated",
            )
            continue
        actual = str(source_value.value).strip()
        passed = actual == expected
        result[_key(sheet_name, target_cell)] = ReportCellValue(
            sheet_name,
            target_cell,
            "Test passed" if passed else "Test failed",
            "ok",
            f"{source_cell} Description expected {expected}, observed {actual}",
        )
    return result


def _cross_sheet_cell_values(values: dict[tuple[str, str], ReportCellValue]) -> dict[tuple[str, str], ReportCellValue]:
    result: dict[tuple[str, str], ReportCellValue] = {}
    model = values.get(_key("Definitions", "C15"))
    module_revision = values.get(_key("Internal Use", "E12"))
    if model and module_revision and model.status == "ok" and module_revision.status == "ok":
        variant = _variant_revision_text(module_revision.value)
        result[_key("Definitions", "J8")] = ReportCellValue(
            "Definitions",
            "J8",
            variant,
            "ok",
            "Model variant derived from Internal Use!E12",
        )

    serial = values.get(_key("Definitions", "D15"))
    sequence_name = values.get(_key("Internal Use", "B20"))
    if serial and sequence_name and serial.status == "ok" and sequence_name.status == "ok":
        serial_text = _identifier_text(serial.value)
        sequence_text = _identifier_text(sequence_name.value)
        passed = bool(serial_text) and (serial_text == sequence_text or serial_text in sequence_text)
        result[_key("Internal Use", "F10")] = ReportCellValue(
            "Internal Use",
            "F10",
            "ok" if passed else "failed",
            "ok",
            f"Serial number check: Definitions!D15={serial_text}, Internal Use!B20={sequence_text}",
        )
    return result


def _variant_revision_text(value: object) -> str:
    number = _to_float(value)
    if number is not None and number.is_integer():
        return f"{int(number):02d}"
    return str(value).strip()


def _identifier_text(value: object) -> str:
    number = _to_float(value)
    if number is not None and number.is_integer():
        return str(int(number))
    return str(value).strip()


def _temp_calib_internal_cell_values(sheet_name: str, rows: list[FormulaEvaluation]) -> dict[tuple[str, str], ReportCellValue]:
    values = _single_cell_values(rows)
    aliases = {
        "C22": "J15",
        "D22": "J16",
        "E22": "J17",
    }
    result: dict[tuple[str, str], ReportCellValue] = {}
    for target, source in aliases.items():
        source_value = values.get(source)
        if not source_value:
            result[_key(sheet_name, target)] = ReportCellValue(sheet_name, target, "", "missing_cell", f"Alias source {source} is not evaluated")
            continue
        result[_key(sheet_name, target)] = ReportCellValue(
            sheet_name,
            target,
            source_value.value,
            source_value.status,
            f"Alias to {source}: {source_value.detail}",
        )
    return result


def _summary_abs_deviation(row) -> float:
    observed = round(row.observed_max_deviation_temp_c, 2) if row.observed_max_deviation_temp_c is not None else row.adjusted_temperature_c
    return abs(observed - row.adjusted_temperature_c)


def _single_cell_values(rows: list[FormulaEvaluation]) -> dict[str, ReportCellValue]:
    values: dict[str, ReportCellValue] = {}
    for row in rows:
        if not row.excel_range or ":" in row.excel_range:
            continue
        values[row.excel_range.upper()] = ReportCellValue(
            sheet_name=row.sheet_name,
            cell=row.excel_range.upper(),
            value=_coerce_value(row.value),
            status=row.status,
            detail=_formula_detail(row),
        )
    return values


def _normalize_cell_values(values: dict[tuple[str, str], ReportCellValue]) -> dict[tuple[str, str], ReportCellValue]:
    return {key: _normalize_report_cell_value(value) for key, value in values.items()}


def _normalize_report_cell_value(value: ReportCellValue) -> ReportCellValue:
    return ReportCellValue(
        value.sheet_name,
        value.cell,
        _normalize_number(value.value),
        value.status,
        value.detail,
    )


def _normalize_number(value: object) -> object:
    if not isinstance(value, float):
        return value
    if not value.is_integer():
        return float(f"{value:.15g}")
    return int(value)


def _display_value_for_db_field(field_name: str, value: object) -> object:
    if not isinstance(value, float):
        return value
    decimals = _display_decimals_for_db_field(field_name)
    if decimals is None:
        return _normalize_number(value)
    displayed = _cm_display_number(value, decimals)
    if displayed is None:
        return value
    return displayed


def _display_decimals_for_db_field(field_name: str) -> int | None:
    field = str(field_name or "")
    compact = field.replace(" ", "").replace("-", "_")
    lowered = compact.lower()
    if any(token in lowered for token in ("heatup_time", "cooldown_time", "diff_phleft", "diff_phright")):
        return 1
    if (
        lowered.startswith("tempacc")
        or lowered.startswith("tempcal")
        or lowered.startswith("pcc_acc")
        or "temperatureaccuracy" in lowered
        or "tempprecision" in lowered
        or "tempstability" in lowered
        or "pcc_cooldown" in lowered
        or lowered in {"performance_pcc"}
    ):
        return 2
    if lowered.startswith("noise_") or lowered.startswith("slope_") or "drift" in lowered:
        return 3
    return None


def _float_cells(values: dict[str, ReportCellValue], cells: list[str]) -> list[float]:
    numbers: list[float] = []
    for cell in cells:
        value = values.get(cell.upper())
        if not value or value.status != "ok":
            continue
        number = _to_float(value.value)
        if number is not None:
            numbers.append(number)
    return numbers


def _cell_range(values: dict[str, ReportCellValue], cells: list[str]) -> float | None:
    numbers = _float_cells(values, cells)
    if not numbers:
        return None
    return max(numbers) - min(numbers)


def _delta_minus_two(values: dict[str, ReportCellValue], start_cell: str, end_cell: str) -> float | None:
    start = _to_float(values.get(start_cell).value if values.get(start_cell) else None)
    end = _to_float(values.get(end_cell).value if values.get(end_cell) else None)
    if start is None or end is None:
        return None
    return end - start - 2


def _cm_display_number(value: float | None, decimals: int) -> float | None:
    if value is None:
        return None
    quant = Decimal("1").scaleb(-decimals)
    try:
        return float(Decimal(f"{value:.12g}").quantize(quant, rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return value


def _criterion_from_calculation_map(rows: list[ReportCellCalculation], label: str, fallback: float) -> float:
    for row in rows:
        if row.sheet_name == "Definitions" and row.label == label:
            value = _to_float(row.expression)
            if value is not None:
                return value
    for row in rows:
        if row.criterion:
            value = _to_float(row.criterion)
            if value is not None and label.lower() in (row.label + " " + row.dependencies + " " + row.rule).lower():
                return value
    return fallback


def _preheater_port_result(values: dict[str, ReportCellValue], start_cell: str, end_cell: str, present_cell: str, memory_cell: str) -> str:
    required = [values.get(cell) for cell in (start_cell, end_cell, present_cell, memory_cell)]
    if any(value is None or value.status != "ok" for value in required):
        return ""
    start = _to_float(required[0].value)
    end = _to_float(required[1].value)
    module_present = str(required[2].value).strip().lower()
    memory_state = str(required[3].value).strip().lower()
    passed = start is not None and end is not None and module_present == "yes" and memory_state == "ok"
    return "Test passed" if passed else "Test failed"


def _pass_fail(value: float | None, criterion: float) -> str:
    if value is not None and value <= criterion:
        return "Test passed"
    return "Test failed"


def _to_float(value: object) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _coerce_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    try:
        return float(text)
    except ValueError:
        return text


def _formula_detail(row: FormulaEvaluation) -> str:
    if row.status == "ok":
        return row.detail or row.formula
    if row.formula and row.detail:
        return f"{row.formula} | {row.detail}"
    return row.detail or row.formula


def _key(sheet_name: str, cell: str) -> tuple[str, str]:
    return (sheet_name.strip().lower(), cell.strip().upper())


def _normalize_report_file(value: str) -> str:
    text = str(value or "").strip().lower()
    for suffix in (".xlsx", ".xls"):
        if text.endswith(suffix):
            return text[: -len(suffix)]
    return text


def _is_date_field(field_name: str) -> bool:
    return "date" in field_name.replace("_", "").lower()


def _fit_columns(ws, max_width: int) -> None:
    for column_cells in ws.columns:
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), max_width)


def _windows_long_path(path: Path) -> str:
    resolved = path.resolve()
    text = str(resolved)
    if not text.startswith("\\\\?\\") and len(text) >= 240:
        return "\\\\?\\" + text
    return text
