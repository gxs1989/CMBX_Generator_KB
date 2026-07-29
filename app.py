from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
import ctypes
import json
from datetime import datetime, timedelta
import os
import queue
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import tkinter as tk
import traceback
import urllib.error
import urllib.request
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk

from chromeleon_method_decoder import decode_cpxm_method_xml
from chromeleon_bridge import export_audit_raw, export_signal_raw
from cmbx_container import CmbxElement, CmbxPackage, extract_cmbx_entry, injection_for_element, load_cmbx_package, rename_cmbx_header_element, report_templates_for_sequence, split_cmbx_sequences, summarize_package
from chromeleon_runtime import runtime_status_text
from db_upload_service import DatabaseUploadConfig, discover_foq_db_workbooks, test_database_connection, upload_foq_db_workbooks
from embedded_method_extractor import extract_embedded_instrument_method
from embedded_report_extractor import ReportSheet, decode_report_template_xml, parse_report_sheet_objects, parse_report_sheets
from export_service import export_all_report_workbooks, export_elements, export_filled_report_template_workbook, export_foq_contract_report, export_report_workbook, export_sequence_report_sheets_workbook
from foq_result_locations import load_foq_workbook, read_device_type_mappings, read_result_locations
from instrument_method_parser import InstrumentMethodText, discover_external_instrument_methods
from method_xml_flow import build_method_flow_rows
from method_role_map import MethodRoleMapAudit, classify_method_role_map
from method_semantic_analyzer import analyze_cm_method_rows, cm_method_variable_name, cm_numeric_value
from method_md_linter import lint_error_rows, lint_method_rows
from method_script_kb import MethodScriptKbEntry, find_method_script_kb_entry, flow_tsv_to_cm_preview_rows, load_method_script_rows_from_kb
from report_formula_evaluator import build_report_formula_context, evaluate_report_formulas
from sequence_cmd_parser import InjectionMethodLink, build_embedded_object_summary, build_injection_method_links, get_injection_method_link
from skills_catalog import SkillCatalogEntry, discover_skill_catalog_entries, skill_catalog_entry_markdown, skill_catalog_overview_markdown
from foq_alignment_catalog import (
    FoqAlignmentRecord,
    build_test_plan_modification_steps,
    build_foq_alignment_records,
    device_options,
    family_options,
    filter_alignment_records,
    intent_tool_options,
    record_detail_sections,
    record_intent_gate,
    record_intent_preview,
    record_modifiability_summary,
    render_intent_conflict_matrix_markdown,
    render_test_plan_assistant_markdown,
    test_intent_options,
    tcc_black_box_coverage_for_record,
    open_verification_topics_for_record,
    write_foq_alignment_workbook,
    write_intent_draft_asset_packet,
    write_intent_action_plan_markdown,
    write_intent_review_markdown,
)
from kb_index import (
    KbIndexEntry,
    discover_kb_index_entries,
    filter_kb_index_entries,
    kb_index_category_options,
    kb_index_entries_for_scope,
    kb_index_entry_category,
    kb_index_entry_detail,
    kb_index_entry_full_markdown,
    kb_index_entry_group,
    kb_index_group_options,
    parse_kb_index_entries,
    read_kb_index_text,
    resolve_kb_entry_files,
    resolve_kb_index_path,
)
from tcc_project_generator import (
    build_single_point_temperature_accuracy_project,
    instrument_method_script_text,
    report_calculation_spec_text,
    single_point_temperature_accuracy_project_to_dict,
    write_single_point_temperature_accuracy_excel_workbooks,
    write_single_point_temperature_accuracy_project,
)
from tools.render_cm_method_md import parse_md_to_rows, write_cm_workbook
from tools.compile_method_md_to_standalone_cmbx import compile_method_md_to_cmbx
from report_template_md_compiler import (
    ReportTemplateCompileResult,
    compile_report_template_md_to_cmbx,
    parse_report_template_md,
)
from external_report_window import ExternalReportWindow


APP_NAME = "CMBX Data Explorer"
APP_VERSION = "V1.4"
AUTO_SCAN_CMBX_LIMIT = 80
EXCEL_PREVIEW_LETTERS = tuple(chr(ord("A") + index) for index in range(16))
DEFAULT_SAMPLE = Path(__file__).resolve().parents[1] / "tcc_temperature_control_analyzer_staging" / "DATA" / "20260701_New" / "20260701_New.cmbx"
DEFAULT_APP_WORKSPACE = Path(os.environ.get("CMBX_DATA_EXPLORER_WORKSPACE", "") or Path(os.environ.get("PROGRAMDATA", r"C:\ProgramData")) / "CMBX Data Explorer Workspace")
DEFAULT_CMBX_SOURCE_FOLDER = DEFAULT_APP_WORKSPACE / "packages"
DEFAULT_CACHE_FOLDER = DEFAULT_APP_WORKSPACE / "cache"
DEFAULT_EXPORT_FOLDER = DEFAULT_APP_WORKSPACE / "exports"
DEFAULT_DB_FOLDER = DEFAULT_APP_WORKSPACE / "db"
DEFAULT_REPORT_FOLDER = DEFAULT_APP_WORKSPACE / "reports"
DEFAULT_LOG_FOLDER = DEFAULT_APP_WORKSPACE / "logs"
DEFAULT_FOQ_MAPPING_FOLDER = DEFAULT_APP_WORKSPACE / "DB MAPPING"
DEFAULT_DATABASE_CONFIG_FILE = DEFAULT_APP_WORKSPACE / "database_config.json"
DEFAULT_AI_CONFIG_FILE = DEFAULT_APP_WORKSPACE / "ai_config.json"
DEFAULT_METHOD_CMBX_TEMPLATE_CANDIDATES = (
    DEFAULT_APP_WORKSPACE / "KB" / "FOQ Template" / "TEMPERATURE_CALIBRATION_720.cmbx",
    DEFAULT_APP_WORKSPACE / "KB" / "FOQ Template" / "TEMP_HEAT_UP_DOWN_20_50_20.cmbx",
)
DEFAULT_METHOD_CMBX_TEMPLATE = next(
    (path for path in DEFAULT_METHOD_CMBX_TEMPLATE_CANDIDATES if path.exists()),
    DEFAULT_METHOD_CMBX_TEMPLATE_CANDIDATES[-1],
)
DEFAULT_REPORT_TEMPLATE_CARRIER_FOLDER = DEFAULT_APP_WORKSPACE / "KB" / "Method Script Generator" / "TCC" / "report_template_cmbx"
METHOD_GENERATOR_KB_DOCS = (
    (
        "MD to CMBX Packaging",
        "MD_TO_STANDALONE_METHOD_CMBX_PACKAGING.md",
        Path(__file__).resolve().parent / "docs" / "MD_TO_STANDALONE_METHOD_CMBX_PACKAGING.md",
    ),
    (
        "MD Format Spec",
        "CM_METHOD_SCRIPT_MD_FORMAT_SPEC.md",
        Path(__file__).resolve().parent / "docs" / "CM_METHOD_SCRIPT_MD_FORMAT_SPEC.md",
    ),
    (
        "CM Compiler Rules",
        "CM Compiler Rules.MD",
        Path(__file__).resolve().parent / "docs" / "CM Compiler Rules.MD",
    ),
)
UI_FONT_FAMILY = "Calibri"
PARSING_NOTES = Path(__file__).with_name("CMBX_PARSING_NOTES.md")
REPO_FOQ_MAPPING = Path(__file__).resolve().parents[1] / "foq" / "FOQResultLocations_V2.83.xls"
LOCAL_STARTUP_LOG = Path(__file__).resolve().parent / "logs" / "app_startup_last.log"


def _write_startup_log(message: str) -> None:
    try:
        LOCAL_STARTUP_LOG.parent.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        with LOCAL_STARTUP_LOG.open("a", encoding="utf-8") as handle:
            handle.write(f"[{timestamp}] {message}\n")
    except Exception:
        pass


def discover_default_cmbx_source_folder() -> Path:
    """Return the fixed application workspace folder for source CMBX packages."""
    env_value = os.environ.get("CMBX_RAW_DATA_FOLDER", "").strip()
    candidates = [
        Path(env_value) if env_value else None,
        DEFAULT_CMBX_SOURCE_FOLDER,
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return DEFAULT_CMBX_SOURCE_FOLDER


def discover_default_foq_mapping() -> Path | None:
    env_value = os.environ.get("CMBX_FOQ_MAPPING_FILE", "").strip()
    candidates: list[Path] = []
    if env_value:
        candidates.append(Path(env_value))
    candidates.append(DEFAULT_FOQ_MAPPING_FOLDER / "FOQResultLocations_V2.83.xls")
    if DEFAULT_FOQ_MAPPING_FOLDER.exists():
        candidates.extend(sorted(DEFAULT_FOQ_MAPPING_FOLDER.glob("*.xls")))
        candidates.extend(sorted(DEFAULT_FOQ_MAPPING_FOLDER.glob("*.xlsx")))
    candidates.append(REPO_FOQ_MAPPING)
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate).lower()
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _generator_report_anchor(spec: dict[str, object]) -> str:
    try:
        method_contract = spec["method_contract"]  # type: ignore[index]
        required_output = method_contract["required_report_output"]  # type: ignore[index]
        return f"{required_output['db_field']} @ {required_output['report_cell']}"  # type: ignore[index]
    except Exception:
        return "report contract"


def _baseline_label_for_app(project) -> str:
    value = getattr(project, "baseline_c", None)
    if value is None:
        return "designer-confirmed"
    return f"{value:g} C"


class CmbxExplorerApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self._main_thread_id = threading.get_ident()
        self._ui_queue: queue.Queue = queue.Queue()
        self.scale = 1.0
        self.colors = {
            "bg": "#F4F7FA",
            "card": "#FFFFFF",
            "card_alt": "#FBFCFE",
            "card_border": "#E5E7EB",
            "text": "#0F172A",
            "text_secondary": "#65738A",
            "primary": "#15803D",
            "primary_hover": "#166534",
            "primary_pressed": "#14532D",
            "secondary_btn": "#2563EB",
            "secondary_btn_hover": "#1D4ED8",
            "secondary_pressed": "#1E40AF",
            "neutral": "#FFFFFF",
            "neutral_hover": "#F3F6FA",
            "neutral_pressed": "#E9EEF5",
            "input_bg": "#F8FAFC",
            "input_border": "#B9C6D4",
            "list_select": "#DBEAFE",
            "match": "#DCFCE7",
            "sequence": "#E0F2FE",
            "injection": "#ECFDF5",
            "method": "#F5F3FF",
            "warning": "#B45309",
        }
        self.package: CmbxPackage | None = None
        self.loaded_packages: list[CmbxPackage] = []
        self.tree_item_context: dict[str, tuple[CmbxPackage | None, CmbxElement | None]] = {}
        self.tree_fs_context: dict[str, Path] = {}
        self.starred_signal_keys: set[str] = set()
        self.current_injection: CmbxElement | None = None
        self.current_sequence: CmbxElement | None = None
        self.injection_method_links: dict[str, InjectionMethodLink] = {}
        self.external_instrument_methods: dict[str, InstrumentMethodText] = {}
        self.report_xml_cache: dict[str, str] = {}
        self.report_sheet_cache: dict[tuple[str, str], list[ReportSheet]] = {}
        self.report_preview_cache: dict[tuple[str, str, str, str], list[tuple[int, dict[int, str], str]]] = {}
        self.foq_device_cache: dict[tuple[str, str], tuple[str, str]] = {}
        self.signal_points_cache: dict[tuple[str, str, str], list[tuple[float, float]]] = {}
        self.raw_plot_series: list[dict[str, object]] = []
        self.raw_plot_benchmark_key: str = ""
        self.foq_candidate_sequences: list[tuple[CmbxPackage, CmbxElement, str, str, str]] = []
        self.method_benchmark_name = ""
        self.method_benchmark_package_path = ""
        self.method_compare_items: list[tuple[CmbxPackage, CmbxElement]] = []
        self._method_drag_iid = ""
        self._tree_drag_iid = ""
        self._tree_drag_started = False
        self._tree_drag_start_xy = (0, 0)
        self._tree_drag_label: tk.Toplevel | None = None
        self._tree_drag_label_text: tk.Label | None = None
        self._tree_drop_target_iid = ""
        self._tree_drop_target_tags: tuple[str, ...] = ()
        self._tree_restore_open_keys: set[tuple[str, str]] = set()
        self._foq_report_template_combo: ttk.Combobox | None = None
        self._foq_filter_refreshing = False
        self._last_thread_status_time = 0.0
        self._last_thread_progress_percent = -1.0
        self.foq_filter_device_options: list[str] = []
        self.foq_filter_db_field_options: list[str] = []
        self.foq_selected_devices: set[str] = set()
        self.foq_selected_db_fields: set[str] = set()
        self.db_upload_files: list[Path] = []
        self.raw_filter_sequence_var = tk.StringVar(value="")
        self.raw_filter_injection_var = tk.StringVar(value="")
        self.raw_filter_channel_var = tk.StringVar(value="")
        self.report_template_sheet_var = tk.StringVar(value="")
        self.cmbx_path_var = tk.StringVar(value=str(discover_default_cmbx_source_folder()))
        self.output_folder_var = tk.StringVar(value=str(DEFAULT_EXPORT_FOLDER))
        default_foq_mapping = discover_default_foq_mapping()
        self.foq_mapping_path_var = tk.StringVar(value=str(default_foq_mapping) if default_foq_mapping else "")
        self.cmbx_display_var = tk.StringVar(value=self._compact_path(self.cmbx_path_var.get()))
        self.foq_mapping_display_var = tk.StringVar(value=self._compact_path(self.foq_mapping_path_var.get()))
        self.foq_device_type_var = tk.StringVar(value="")
        self.foq_device_search_var = tk.StringVar(value="")
        self.foq_db_field_search_var = tk.StringVar(value="")
        self.foq_device_count_var = tk.StringVar(value="")
        self.foq_db_field_count_var = tk.StringVar(value="")
        database_defaults = self._load_database_config_defaults()
        self.db_server_var = tk.StringVar(value=database_defaults.get("server", "10.68.178.52"))
        self.db_database_var = tk.StringVar(value=database_defaults.get("database", "QCLab"))
        self.db_username_var = tk.StringVar(value=database_defaults.get("username", "QCUser"))
        self.db_password_var = tk.StringVar(value=os.environ.get("CMBX_DB_PASSWORD", database_defaults.get("password", "")))
        self.db_schema_var = tk.StringVar(value=database_defaults.get("schema", "dbo"))
        self.db_table_var = tk.StringVar(value=database_defaults.get("table", "AUTO"))
        self.db_driver_var = tk.StringVar(value=database_defaults.get("driver", "ODBC Driver 17 for SQL Server"))
        self.db_trust_cert_var = tk.BooleanVar(value=bool(database_defaults.get("trust_server_certificate", True)))
        ai_defaults = self._load_ai_config_defaults()
        self.ai_base_url_var = tk.StringVar(value=str(ai_defaults.get("base_url", "https://api.openai.com/v1")))
        self.ai_model_var = tk.StringVar(value=str(ai_defaults.get("model", "gpt-5.5")))
        self.ai_api_key_var = tk.StringVar(value=str(ai_defaults.get("api_key", "")))
        self.method_generator_family_var = tk.StringVar(value="TCC")
        self.method_generator_intent_var = tk.StringVar(value="")
        self.generator_device_var = tk.StringVar(value="VH-C10-A")
        self.generator_test_var = tk.StringVar(value="Temperature Accuracy")
        self.generator_baseline_var = tk.StringVar(value="")
        self.generator_setpoint_var = tk.StringVar(value="")
        self.foq_alignment_family_var = tk.StringVar(value="TCC")
        self.foq_alignment_test_var = tk.StringVar(value="")
        self.foq_alignment_intent_var = tk.StringVar(value="Search / Recommend")
        self.foq_alignment_intent_parameter_var = tk.StringVar(value="")
        self.foq_alignment_records: tuple[FoqAlignmentRecord, ...] = ()
        self.foq_alignment_filtered_records: tuple[FoqAlignmentRecord, ...] = ()
        self.foq_alignment_selected_devices: set[str] = set()
        self.test_plan_family_var = tk.StringVar(value="TCC")
        self.test_plan_test_var = tk.StringVar(value="")
        self.test_plan_device_var = tk.StringVar(value="VH-C10-A")
        self.test_plan_intent_var = tk.StringVar(value="Crop / Modify")
        self.test_plan_free_intent_var = tk.StringVar(value="")
        self.test_plan_parameter_var = tk.StringVar(value="")
        self.test_plan_selected_record: FoqAlignmentRecord | None = None
        self.test_plan_accepted_review_text = ""
        self.test_plan_source_status_var = tk.StringVar(value="No method intent analyzed")
        self.method_editor_md_path_var = tk.StringVar(value="")
        self.method_editor_xlsx_path_var = tk.StringVar(value="")
        self.method_editor_example_path_var = tk.StringVar(value=str(Path(os.environ.get("USERPROFILE", "")) / "OneDrive - Thermo Fisher Scientific" / "Desktop" / "Book2.xlsx"))
        self.method_editor_status_var = tk.StringVar(value="Select an MD method script, preview it, then compile a standalone instrument-method CMBX.")
        self.method_editor_kb_doc_var = tk.StringVar(value=METHOD_GENERATOR_KB_DOCS[0][1])
        self.method_editor_rows: list[dict[str, str]] = []
        self.report_generator_md_path_var = tk.StringVar(value="")
        self.report_generator_output_path_var = tk.StringVar(value="")
        self.report_generator_sheet_var = tk.StringVar(value="")
        self.report_generator_status_var = tk.StringVar(value="Select a report-template Markdown file, preview the staged CMBX, then generate the final standalone report CMBX.")
        self.report_generator_result: ReportTemplateCompileResult | None = None
        self.report_generator_preview_cmbx: Path | None = None
        self.report_generator_preview_package: CmbxPackage | None = None
        self.report_generator_preview_report: CmbxElement | None = None
        self.kb_index_entries: tuple[KbIndexEntry, ...] = ()
        self.kb_index_filtered_entries: tuple[KbIndexEntry, ...] = ()
        self.kb_index_row_context: dict[str, KbIndexEntry | str] = {}
        self.kb_index_category_context: dict[str, str] = {}
        self.kb_index_category_var = tk.StringVar(value="All")
        self.kb_index_search_var = tk.StringVar(value="")
        self.kb_index_scope = "All"
        self._kb_index_populating = False
        self.skills_catalog_entries: tuple[SkillCatalogEntry, ...] = ()
        self.skills_catalog_row_context: dict[str, SkillCatalogEntry] = {}
        self.external_report_window: ExternalReportWindow | None = None
        self.status_var = tk.StringVar(value="Ready")
        self.progress_var = tk.DoubleVar(value=0.0)
        self._path_display_editing: set[str] = set()
        self._setup_window()
        self._setup_styles()
        self._build_ui()
        self.root.after(50, self._drain_ui_queue)

    def _setup_window(self) -> None:
        self.root.title(f"{APP_NAME} - {APP_VERSION}")
        self.root.geometry("1500x940")
        self.root.minsize(1220, 760)
        self.root.configure(bg=self.colors["bg"])
        try:
            dpi = self.root.winfo_fpixels("1i")
            self.scale = max(0.88, min(1.25, dpi / 96))
            self.root.tk.call("tk", "scaling", dpi / 72)
        except tk.TclError:
            self.scale = 1.0

    def _setup_styles(self) -> None:
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Explorer.Treeview", font=self._font(9), rowheight=25, background=self.colors["card_alt"], fieldbackground=self.colors["card_alt"], foreground=self.colors["text"])
        style.configure("Explorer.Treeview.Heading", font=self._font(9, "bold"), background="#E5E7EB", foreground=self.colors["text"], relief="flat")
        style.map("Explorer.Treeview", background=[("selected", self.colors["list_select"])], foreground=[("selected", self.colors["text"])])
        style.map("Explorer.Treeview.Heading", background=[("active", "#E5E7EB")])
        style.configure("TNotebook", background=self.colors["card"], borderwidth=0)
        style.configure("TNotebook.Tab", font=self._font(9, "bold"), padding=(14, 7), background="#EEF2F7")
        style.map("TNotebook.Tab", background=[("selected", self.colors["card"])])

    def _font(self, size: int, weight: str = "normal", family: str = UI_FONT_FAMILY) -> tuple[str, int, str]:
        return (family, max(8, round(size * self.scale)), weight)

    def _build_ui(self) -> None:
        main = tk.Frame(self.root, bg=self.colors["bg"])
        main.pack(fill="both", expand=True, padx=26, pady=22)
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=0)
        main.rowconfigure(2, weight=1)

        header = tk.Frame(main, bg=self.colors["bg"])
        header.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 16))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text=APP_NAME, font=self._font(20, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w")
        tk.Label(header, text=f"Chromeleon package browser | {APP_VERSION}", font=self._font(9), bg=self.colors["bg"], fg=self.colors["text_secondary"]).grid(row=1, column=0, sticky="w", pady=(3, 0))
        self._make_button(header, "External Report Engine", self.show_external_report_engine, kind="secondary", width=20).grid(row=0, column=1, rowspan=2, sticky="e", padx=(8, 0))
        self._make_button(header, "AI Settings", self.show_ai_settings, kind="neutral", width=10).grid(row=0, column=2, rowspan=2, sticky="e", padx=(8, 0))
        self._make_button(header, "CMBX Notes", self.show_parsing_notes, kind="neutral", width=12).grid(row=0, column=3, rowspan=2, sticky="e", padx=(8, 0))

        source_card = self._make_card(main)
        source_card.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 14))
        source_card.columnconfigure(1, weight=1)
        tk.Label(source_card, text="Raw Data Folder", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=(18, 10), pady=(16, 8))
        self.cmbx_entry = self._make_path_entry(source_card, self.cmbx_path_var, self.cmbx_display_var, "cmbx")
        self.cmbx_entry.grid(row=0, column=1, sticky="ew", pady=(16, 8), ipady=5)
        self._make_button(source_card, "Select Folder", self.browse_cmbx_folder, kind="neutral", width=13).grid(row=0, column=2, padx=10, pady=(16, 8))
        self._make_button(source_card, "Add CMBX Files", self.browse_cmbx, kind="neutral", width=13).grid(row=0, column=3, padx=(0, 10), pady=(16, 8))
        self._make_button(source_card, "Scan CMBX", self.load_package, kind="secondary", width=22).grid(row=0, column=4, padx=(0, 18), pady=(16, 8), sticky="e")

        self.main_paned = tk.PanedWindow(main, orient=tk.HORIZONTAL, sashwidth=8, sashrelief="flat", bg=self.colors["bg"], bd=0)
        self.main_paned.grid(row=2, column=0, columnspan=2, sticky="nsew")
        tree_card = self._make_card(main)
        tree_card.columnconfigure(0, weight=1)
        tree_card.rowconfigure(1, weight=1)
        tk.Label(tree_card, text="Sequence Data", font=self._font(12, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 10))
        tree_frame = tk.Frame(tree_card, bg=self.colors["card"])
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        self.package_tree = ttk.Treeview(tree_frame, columns=("type",), show="tree headings", style="Explorer.Treeview", selectmode="extended")
        self.package_tree.heading("#0", text="Name")
        self.package_tree.heading("type", text="Type")
        self.package_tree.column("#0", width=360)
        self.package_tree.column("type", width=150, anchor="center")
        self.package_tree.tag_configure("package", background="#F8FAFC", foreground=self.colors["text"])
        self.package_tree.tag_configure("folder", background="#F8FAFC", foreground="#334155")
        self.package_tree.tag_configure("file", foreground="#475569")
        self.package_tree.tag_configure("drop_target", background="#93C5FD", foreground="#0F172A")
        self.package_tree.tag_configure("sequence", background=self.colors["sequence"])
        self.package_tree.tag_configure("injection", background=self.colors["injection"])
        self.package_tree.tag_configure("signal", foreground="#0F766E")
        self.package_tree.tag_configure("audit", foreground="#7C2D12")
        self.package_tree.grid(row=0, column=0, sticky="nsew")
        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.package_tree.yview)
        scroll_y.grid(row=0, column=1, sticky="ns")
        self.package_tree.configure(yscrollcommand=scroll_y.set)
        self.package_tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.package_tree.bind("<Button-3>", self._show_sequence_tree_menu)
        self.package_tree.bind("<ButtonPress-1>", self._begin_tree_drag, add="+")
        self.package_tree.bind("<B1-Motion>", self._update_tree_drag, add="+")
        self.package_tree.bind("<ButtonRelease-1>", self._finish_tree_drag, add="+")

        detail_card = self._make_card(main)
        detail_card.columnconfigure(0, weight=1)
        detail_card.rowconfigure(2, weight=1)
        tk.Label(detail_card, text="Selection Details", font=self._font(12, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 8))
        self.summary_label = tk.Label(detail_card, text="No CMBX loaded.", anchor="w", justify="left", font=self._font(9), bg=self.colors["card"], fg=self.colors["text_secondary"])
        self.summary_label.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 10))
        self.notebook = ttk.Notebook(detail_card)
        self.notebook.grid(row=2, column=0, sticky="nsew", padx=18, pady=(0, 12))
        self._build_raw_plot_tab()
        self._build_foq_db_tab()
        self._build_database_upload_tab()
        self._build_channels_tab()
        self._build_audit_tab()
        self._build_report_sheets_tab()
        self._build_method_context_tab()
        self._build_method_editor_tab()
        self._build_processing_methods_tab()
        self._build_report_templates_tab()
        self._build_report_template_generator_tab()
        self._build_info_tab()
        self._build_kb_index_tab()
        self._build_skills_tab()
        self._build_foq_kb_to_run_tab()
        self.main_paned.add(tree_card, minsize=360, width=560)
        self.main_paned.add(detail_card, minsize=560, width=900)

        export_card = self._make_card(main)
        export_card.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(14, 0))
        export_card.columnconfigure(1, weight=1)
        tk.Label(export_card, text="Output Folder", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=(18, 10), pady=14)
        self._make_entry(export_card, self.output_folder_var).grid(row=0, column=1, sticky="ew", pady=14, ipady=5)
        self._make_button(export_card, "Browse", self.browse_output_folder, kind="neutral", width=10).grid(row=0, column=2, padx=10, pady=14)
        self._make_button(export_card, "Open", self.open_output_folder, kind="secondary", width=10).grid(row=0, column=3, padx=(0, 18), pady=14)
        status_frame = tk.Frame(main, bg=self.colors["bg"])
        status_frame.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        status_frame.columnconfigure(0, weight=1)
        tk.Label(status_frame, textvariable=self.status_var, anchor="w", font=self._font(9), bg=self.colors["bg"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="ew")
        self.progress_bar = ttk.Progressbar(status_frame, variable=self.progress_var, maximum=100, mode="determinate", length=210)
        self.progress_bar.grid(row=0, column=1, sticky="e", padx=(12, 0))
        self._populate_tree()
        self._show_workspace_summary()
        self._enable_cmbx_drop_target()
        # Let the window paint before the default workspace scan starts; otherwise
        # a large package folder can make startup look like a failed launch.
        self.root.after(1200, self._auto_scan_default_folder)

    def _build_channels_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Advanced raw channel filter. Use Sequence Data right-click for direct exports, and Raw Plot for previews.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Apply Filter", self.apply_raw_channel_filter, kind="secondary", width=14).grid(row=0, column=1, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Export Filtered", self.export_filtered_channels, kind="primary", width=16).grid(row=0, column=2, sticky="e", padx=(8, 0))
        filter_frame = tk.Frame(tab, bg=self.colors["card"])
        filter_frame.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        filter_frame.columnconfigure(1, weight=1)
        filter_frame.columnconfigure(3, weight=1)
        filter_frame.columnconfigure(5, weight=1)
        tk.Label(filter_frame, text="Sequence", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6))
        self._make_entry(filter_frame, self.raw_filter_sequence_var).grid(row=0, column=1, sticky="ew", padx=(0, 12), ipady=4)
        tk.Label(filter_frame, text="Injection", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="w", padx=(0, 6))
        self._make_entry(filter_frame, self.raw_filter_injection_var).grid(row=0, column=3, sticky="ew", padx=(0, 12), ipady=4)
        tk.Label(filter_frame, text="Channel", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=4, sticky="w", padx=(0, 6))
        self._make_entry(filter_frame, self.raw_filter_channel_var).grid(row=0, column=5, sticky="ew", ipady=4)
        self.channel_table = self._make_table(
            tab,
            ("cmbx", "sequence", "injection", "channel", "size", "raw"),
            {"cmbx": "CMBX", "sequence": "Sequence", "injection": "Injection", "channel": "Channel", "size": "Size", "raw": "Raw File"},
        )
        self.channel_table.grid(row=2, column=0, sticky="nsew")
        channel_widget = self._table_widget(self.channel_table)
        self._enable_drag_row_selection(channel_widget)
        channel_widget.bind("<Double-1>", self.open_selected_channels)
        self.channel_plot_points: list[tuple[float, float]] = []
        self.channel_plot_title = ""
        self.notebook.add(tab, text="Raw Filter Export")

    def _build_raw_plot_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Compare raw channels. Add channels from Sequence Data, then select rows in Selected Channels to draw them.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Set Benchmark", self.set_raw_plot_benchmark, kind="neutral", width=14).grid(row=0, column=1, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Clear Plot", self.clear_raw_plot, kind="neutral", width=12).grid(row=0, column=2, sticky="e", padx=(8, 0))
        raw_plot_paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        raw_plot_paned.grid(row=1, column=0, sticky="nsew")
        selection_frame = tk.Frame(raw_plot_paned, bg=self.colors["card"])
        selection_frame.columnconfigure(0, weight=1)
        selection_frame.rowconfigure(1, weight=1)
        tk.Label(selection_frame, text="Selected Channels", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.raw_plot_selection_table = self._make_table(
            selection_frame,
            ("benchmark", "cmbx", "sequence", "injection", "channel"),
            {"benchmark": "Benchmark", "cmbx": "CMBX", "sequence": "Sequence", "injection": "Injection", "channel": "Channel"},
        )
        self.raw_plot_selection_table.grid(row=1, column=0, sticky="nsew")
        raw_plot_table = self._table_widget(self.raw_plot_selection_table)
        raw_plot_table.bind("<<TreeviewSelect>>", lambda _event: self._draw_raw_plot_series())
        raw_plot_table.bind("<Button-3>", self._show_raw_plot_selection_menu)
        plot_frame = tk.Frame(raw_plot_paned, bg=self.colors["card"])
        plot_frame.columnconfigure(0, weight=1)
        plot_frame.rowconfigure(0, weight=1)
        self.raw_plot_canvas = tk.Canvas(plot_frame, bg=self.colors["card_alt"], highlightbackground=self.colors["card_border"], highlightthickness=1)
        self.raw_plot_canvas.grid(row=0, column=0, sticky="nsew")
        self.raw_plot_canvas.bind("<Configure>", lambda _e: self._draw_raw_plot_series())
        raw_plot_paned.add(selection_frame, minsize=105, height=165)
        raw_plot_paned.add(plot_frame, minsize=260, height=460)
        self.notebook.add(tab, text="Raw Plot")

    def _build_foq_db_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)

        mapping_frame = tk.Frame(tab, bg=self.colors["card"])
        mapping_frame.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        mapping_frame.columnconfigure(1, weight=1)
        tk.Label(mapping_frame, text="FOQ Mapping", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self._make_path_entry(mapping_frame, self.foq_mapping_path_var, self.foq_mapping_display_var, "foq").grid(row=0, column=1, sticky="ew", ipady=4)
        self._make_button(mapping_frame, "Browse", self.browse_foq_mapping, kind="neutral", width=10).grid(row=0, column=2, sticky="e", padx=(8, 0))
        tk.Label(mapping_frame, text="Detected From CMBX", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=3, sticky="e", padx=(14, 6))
        device_entry = self._make_entry(mapping_frame, self.foq_device_type_var)
        device_entry.configure(width=18)
        device_entry.configure(state="readonly")
        device_entry.grid(row=0, column=4, sticky="e", ipady=4)

        filter_frame = tk.Frame(tab, bg=self.colors["card"])
        filter_frame.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        for col in (1, 3):
            filter_frame.columnconfigure(col, weight=1)
        tk.Label(filter_frame, text="Devices", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="nw", padx=(0, 6))
        device_frame = tk.Frame(filter_frame, bg=self.colors["card"])
        device_frame.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        device_frame.columnconfigure(0, weight=1)
        device_search = self._make_entry(device_frame, self.foq_device_search_var)
        device_search.grid(row=0, column=0, columnspan=2, sticky="ew", ipady=3, pady=(0, 4))
        self.foq_filter_device_listbox = tk.Listbox(
            device_frame,
            height=6,
            selectmode="extended",
            exportselection=False,
            font=self._font(9),
            bg=self.colors["input_bg"],
            fg=self.colors["text"],
            relief="solid",
            bd=1,
        )
        self.foq_filter_device_listbox.grid(row=1, column=0, sticky="ew")
        self.foq_filter_device_listbox.bind("<<ListboxSelect>>", self._on_foq_device_selection_changed)
        device_scroll = ttk.Scrollbar(device_frame, orient="vertical", command=self.foq_filter_device_listbox.yview)
        device_scroll.grid(row=1, column=1, sticky="ns")
        self.foq_filter_device_listbox.configure(yscrollcommand=device_scroll.set)
        device_buttons = tk.Frame(device_frame, bg=self.colors["card"])
        device_buttons.grid(row=1, column=2, sticky="n", padx=(8, 0))
        self._make_button(device_buttons, "All", self.select_all_foq_devices, kind="neutral", width=8).grid(row=0, column=0, sticky="ew")
        self._make_button(device_buttons, "Clear", self.clear_foq_devices, kind="neutral", width=8).grid(row=1, column=0, sticky="ew", pady=(4, 0))
        tk.Label(device_frame, textvariable=self.foq_device_count_var, font=self._font(8), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=2, column=0, columnspan=2, sticky="w", pady=(3, 0))
        tk.Label(filter_frame, text="DB Fields", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="nw", padx=(0, 6))
        db_field_frame = tk.Frame(filter_frame, bg=self.colors["card"])
        db_field_frame.grid(row=0, column=3, sticky="ew")
        db_field_frame.columnconfigure(0, weight=1)
        db_field_search = self._make_entry(db_field_frame, self.foq_db_field_search_var)
        db_field_search.grid(row=0, column=0, columnspan=2, sticky="ew", ipady=3, pady=(0, 4))
        self.foq_filter_db_field_listbox = tk.Listbox(
            db_field_frame,
            height=6,
            selectmode="extended",
            exportselection=False,
            font=self._font(9),
            bg=self.colors["input_bg"],
            fg=self.colors["text"],
            relief="solid",
            bd=1,
        )
        self.foq_filter_db_field_listbox.grid(row=1, column=0, sticky="ew")
        self.foq_filter_db_field_listbox.bind("<<ListboxSelect>>", self._on_foq_db_field_selection_changed)
        db_field_scroll = ttk.Scrollbar(db_field_frame, orient="vertical", command=self.foq_filter_db_field_listbox.yview)
        db_field_scroll.grid(row=1, column=1, sticky="ns")
        self.foq_filter_db_field_listbox.configure(yscrollcommand=db_field_scroll.set)
        field_buttons = tk.Frame(db_field_frame, bg=self.colors["card"])
        field_buttons.grid(row=1, column=2, sticky="n", padx=(8, 0))
        self._make_button(field_buttons, "All", self.select_all_foq_db_fields, kind="neutral", width=8).grid(row=0, column=0, sticky="ew")
        self._make_button(field_buttons, "Clear", self.clear_foq_db_fields, kind="neutral", width=8).grid(row=1, column=0, sticky="ew", pady=(4, 0))
        self._make_button(field_buttons, "Preview", self.preview_foq_db_fields, kind="secondary", width=10).grid(row=2, column=0, sticky="ew", pady=(4, 0))
        tk.Label(db_field_frame, textvariable=self.foq_db_field_count_var, font=self._font(8), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=2, column=0, columnspan=2, sticky="w", pady=(3, 0))
        self.foq_device_search_var.trace_add("write", lambda *_args: self._populate_foq_device_filter_listbox())
        self.foq_db_field_search_var.trace_add("write", lambda *_args: self._populate_foq_db_field_filter_listbox())

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=2, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Build one FOQ DB workbook per candidate sequence. Add from Sequence Data or use filters.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Add Selected Sequence", self.add_selected_sequences_to_foq_candidates, kind="secondary", width=20).grid(row=0, column=1, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Add Filter Matches", self.add_filtered_sequences_to_foq_candidates, kind="secondary", width=18).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Remove", self.remove_selected_foq_candidates, kind="neutral", width=10).grid(row=0, column=3, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Clear", self.clear_foq_candidates, kind="neutral", width=10).grid(row=0, column=4, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Export Candidate CMBX", self.export_foq_candidate_sequences, kind="neutral", width=19).grid(row=0, column=5, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Export Candidate DB", self.export_foq_candidate_db, kind="primary", width=20).grid(row=0, column=6, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Upload Candidate DB", self.upload_foq_candidate_db, kind="primary", width=20).grid(row=0, column=7, sticky="e", padx=(8, 0))

        foq_paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        foq_paned.grid(row=3, column=0, sticky="nsew")
        candidate_frame = tk.Frame(foq_paned, bg=self.colors["card"])
        candidate_frame.columnconfigure(0, weight=1)
        candidate_frame.rowconfigure(0, weight=1)
        preview_frame = tk.Frame(foq_paned, bg=self.colors["card"])
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(1, weight=1)
        preview_frame.rowconfigure(3, weight=0)

        self.foq_candidate_table = self._make_table(
            candidate_frame,
            ("device", "device_source", "report_template", "cmbx", "sequence", "injections", "channels", "path"),
            {
                "device": "Device",
                "device_source": "Device Source",
                "report_template": "Report Template",
                "cmbx": "CMBX",
                "sequence": "Sequence",
                "injections": "Injections",
                "channels": "Channels",
                "path": "Source",
            },
        )
        self.foq_candidate_table.grid(row=0, column=0, sticky="nsew")
        foq_table = self._table_widget(self.foq_candidate_table)
        foq_table.bind("<Button-1>", self._maybe_edit_foq_report_template_cell, add="+")
        tk.Label(preview_frame, text="Selected DB Field Preview", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6))
        self.foq_preview_table = self._make_table(
            preview_frame,
            ("sequence", "device", "status"),
            {"sequence": "Sequence", "device": "Device", "status": "Status"},
        )
        self.foq_preview_table.grid(row=1, column=0, sticky="nsew")
        tk.Label(preview_frame, text="Preview Log", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=2, column=0, sticky="w", pady=(8, 4))
        preview_log_frame = tk.Frame(preview_frame, bg=self.colors["card"])
        preview_log_frame.grid(row=3, column=0, sticky="ew")
        preview_log_frame.columnconfigure(0, weight=1)
        self.foq_preview_log = tk.Text(
            preview_log_frame,
            height=6,
            wrap="word",
            font=self._font(8),
            bg=self.colors["input_bg"],
            fg=self.colors["text"],
            relief="solid",
            bd=1,
        )
        self.foq_preview_log.grid(row=0, column=0, sticky="ew")
        self.foq_preview_log.configure(state="disabled")
        preview_log_scroll = ttk.Scrollbar(preview_log_frame, orient="vertical", command=self.foq_preview_log.yview)
        preview_log_scroll.grid(row=0, column=1, sticky="ns")
        self.foq_preview_log.configure(yscrollcommand=preview_log_scroll.set)
        foq_paned.add(candidate_frame, minsize=180, height=330)
        foq_paned.add(preview_frame, minsize=180, height=260)
        self._refresh_foq_mapping_filter_options()
        self.notebook.add(tab, text="FOQ DB")

    def _build_database_upload_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        tab.rowconfigure(4, weight=0)

        config = tk.Frame(tab, bg=self.colors["card"])
        config.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        for column in (1, 3, 5):
            config.columnconfigure(column, weight=1)
        tk.Label(config, text="Server", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6), pady=(0, 6))
        self._make_entry(config, self.db_server_var).grid(row=0, column=1, sticky="ew", padx=(0, 12), pady=(0, 6), ipady=4)
        tk.Label(config, text="Database", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="w", padx=(0, 6), pady=(0, 6))
        self._make_entry(config, self.db_database_var).grid(row=0, column=3, sticky="ew", padx=(0, 12), pady=(0, 6), ipady=4)
        tk.Label(config, text="Driver", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=4, sticky="w", padx=(0, 6), pady=(0, 6))
        self._make_entry(config, self.db_driver_var).grid(row=0, column=5, sticky="ew", pady=(0, 6), ipady=4)

        tk.Label(config, text="User", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(0, 6))
        self._make_entry(config, self.db_username_var).grid(row=1, column=1, sticky="ew", padx=(0, 12), pady=(0, 6), ipady=4)
        tk.Label(config, text="Password", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=1, column=2, sticky="w", padx=(0, 6), pady=(0, 6))
        password_entry = tk.Entry(config, textvariable=self.db_password_var, show="*", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        password_entry.grid(row=1, column=3, sticky="ew", padx=(0, 12), pady=(0, 6), ipady=4)
        tk.Label(config, text="Schema", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=1, column=4, sticky="w", padx=(0, 6), pady=(0, 6))
        self._make_entry(config, self.db_schema_var).grid(row=1, column=5, sticky="ew", pady=(0, 6), ipady=4)

        tk.Label(config, text="Table", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=2, column=0, sticky="w", padx=(0, 6))
        self._make_entry(config, self.db_table_var).grid(row=2, column=1, sticky="ew", padx=(0, 12), ipady=4)
        tk.Checkbutton(config, text="Trust server certificate", variable=self.db_trust_cert_var, bg=self.colors["card"], fg=self.colors["text"], activebackground=self.colors["card"], font=self._font(9)).grid(row=2, column=2, columnspan=2, sticky="w")
        self._make_button(config, "Save Config", self.save_database_config, kind="neutral", width=13).grid(row=2, column=4, sticky="e", padx=(0, 8))
        self._make_button(config, "Test Connection", self.test_database_upload_connection, kind="secondary", width=16).grid(row=2, column=5, sticky="e")

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=1, column=0, sticky="ew", pady=(4, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Upload exported FOQ DB workbooks. One DB workbook row is inserted per sequence.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Add DB Files", self.add_database_upload_files, kind="neutral", width=13).grid(row=0, column=1, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Add Folder", self.add_database_upload_folder, kind="neutral", width=12).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Clear", self.clear_database_upload_files, kind="neutral", width=10).grid(row=0, column=3, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Upload DB Files", self.upload_database_files, kind="primary", width=16).grid(row=0, column=4, sticky="e", padx=(8, 0))

        self.db_upload_table = self._make_table(
            tab,
            ("sequence", "fields", "path"),
            {"sequence": "Sequence", "fields": "DB Fields", "path": "Workbook"},
        )
        self.db_upload_table.grid(row=2, column=0, sticky="nsew")

        tk.Label(tab, text="Database Upload Log", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=3, column=0, sticky="w", pady=(8, 4))
        log_frame = tk.Frame(tab, bg=self.colors["card"])
        log_frame.grid(row=4, column=0, sticky="ew")
        log_frame.columnconfigure(0, weight=1)
        self.db_upload_log = tk.Text(log_frame, height=7, wrap="word", font=self._font(8), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        self.db_upload_log.grid(row=0, column=0, sticky="ew")
        self.db_upload_log.configure(state="disabled")
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.db_upload_log.yview)
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.db_upload_log.configure(yscrollcommand=log_scroll.set)
        self.notebook.add(tab, text="Database Upload")

    def _build_foq_kb_to_run_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)

        controls = tk.Frame(tab, bg=self.colors["card"])
        controls.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        controls.columnconfigure(3, weight=1)
        tk.Label(controls, text="FOQ TD", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6))
        td_combo = ttk.Combobox(controls, textvariable=self.foq_kb_td_var, values=(foq_td_title(),), state="readonly", width=38)
        td_combo.grid(row=0, column=1, sticky="w", padx=(0, 12))
        tk.Label(controls, text="FOQ Source", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="w", padx=(0, 6))
        tk.Label(controls, text=foq_td_source_summary(), font=self._font(9), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=3, sticky="w")
        self._make_button(controls, "Export Alignment Excel", self.export_foq_kb_to_run_alignment, kind="primary", width=22).grid(row=0, column=4, sticky="e", padx=(8, 0))

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(
            toolbar,
            "Select and align FOQ TD items to injections, method scripts, processing bindings, and report calculations before creating runnable methods.",
        ).grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Open Output", self.open_output_folder, kind="neutral", width=12).grid(row=0, column=1, sticky="e", padx=(8, 0))

        paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        paned.grid(row=2, column=0, sticky="nsew")
        top = tk.Frame(paned, bg=self.colors["card"])
        bottom = tk.Frame(paned, bg=self.colors["card"])
        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        bottom.columnconfigure(0, weight=1)
        bottom.rowconfigure(1, weight=1)

        self.foq_kb_to_run_table = self._make_table(
            top,
            ("order", "td_item", "injection", "instrument_method", "processing_method", "report_sheets"),
            {
                "order": "#",
                "td_item": "FOQ TD Item / 测试项",
                "injection": "Injection",
                "instrument_method": "Instrument Method",
                "processing_method": "Processing Method",
                "report_sheets": "Report",
            },
        )
        self.foq_kb_to_run_table.grid(row=0, column=0, sticky="nsew")
        kb_table = self._table_widget(self.foq_kb_to_run_table)
        kb_table.column("order", width=44, anchor="e")
        kb_table.bind("<<TreeviewSelect>>", self._preview_selected_foq_kb_to_run_row)

        tk.Label(bottom, text="TD -> Method -> Report Alignment", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 4))
        detail_frame = tk.Frame(bottom, bg=self.colors["card"])
        detail_frame.grid(row=1, column=0, sticky="nsew")
        detail_frame.columnconfigure(0, weight=1)
        detail_frame.rowconfigure(0, weight=1)
        self.foq_kb_to_run_detail = tk.Text(detail_frame, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        self.foq_kb_to_run_detail.grid(row=0, column=0, sticky="nsew")
        detail_scroll = ttk.Scrollbar(detail_frame, orient="vertical", command=self.foq_kb_to_run_detail.yview)
        detail_scroll.grid(row=0, column=1, sticky="ns")
        self.foq_kb_to_run_detail.configure(yscrollcommand=detail_scroll.set)

        paned.add(top, minsize=220, height=320)
        paned.add(bottom, minsize=220, height=360)
        self.notebook.add(tab, text="FOQ KB to Run")
        self._populate_foq_kb_to_run_table()

    def _build_foq_kb_to_run_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)

        controls = tk.Frame(tab, bg=self.colors["card"])
        controls.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        controls.columnconfigure(6, weight=1)
        tk.Label(controls, text="Family", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.foq_alignment_family_combo = ttk.Combobox(controls, textvariable=self.foq_alignment_family_var, values=("TCC", "VDAD"), state="readonly", width=12)
        self.foq_alignment_family_combo.grid(row=0, column=1, sticky="w", padx=(0, 12))
        self.foq_alignment_family_combo.bind("<<ComboboxSelected>>", self._refresh_foq_alignment_filters)

        tk.Label(controls, text="TestIntent / TD", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="w", padx=(0, 6))
        self.foq_alignment_test_combo = ttk.Combobox(controls, textvariable=self.foq_alignment_test_var, values=(), width=30)
        self.foq_alignment_test_combo.grid(row=0, column=3, sticky="w", padx=(0, 12))
        self.foq_alignment_test_combo.bind("<<ComboboxSelected>>", self._populate_foq_kb_to_run_table)
        self.foq_alignment_test_combo.bind("<KeyRelease>", self._populate_foq_kb_to_run_table)

        self._make_button(controls, "Refresh KB", self._refresh_foq_alignment_catalog, kind="neutral", width=12).grid(row=0, column=4, sticky="w")
        self._make_button(controls, "Export Progress Excel", self.export_foq_kb_to_run_alignment, kind="primary", width=22).grid(row=0, column=6, sticky="e", padx=(8, 0))

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(2, weight=1)
        toolbar.columnconfigure(8, weight=1)
        tk.Label(toolbar, text="Devices", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="nw", padx=(0, 8))
        self.foq_alignment_device_listbox = tk.Listbox(
            toolbar,
            selectmode="extended",
            height=4,
            exportselection=False,
            font=self._font(9),
            bg=self.colors["input_bg"],
            fg=self.colors["text"],
            relief="solid",
            bd=1,
        )
        self.foq_alignment_device_listbox.grid(row=0, column=1, sticky="w", padx=(0, 12))
        self.foq_alignment_device_listbox.bind("<<ListboxSelect>>", self._populate_foq_kb_to_run_table)
        self._make_toolbar_label(
            toolbar,
            "Knowledge progress only. This page reads existing KB/MD alignment records and does not scan or load CMBX packages.",
        ).grid(row=0, column=2, sticky="ew")
        self._make_button(toolbar, "Open Output", self.open_output_folder, kind="neutral", width=12).grid(row=0, column=3, sticky="e", padx=(8, 0))

        paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        paned.grid(row=2, column=0, sticky="nsew")
        top = tk.Frame(paned, bg=self.colors["card"])
        bottom = tk.Frame(paned, bg=self.colors["card"])
        top.columnconfigure(0, weight=1)
        top.rowconfigure(2, weight=1)
        bottom.columnconfigure(0, weight=1)
        bottom.rowconfigure(1, weight=1)

        summary = tk.Frame(top, bg=self.colors["card_alt"], highlightbackground=self.colors["card_border"], highlightthickness=1)
        summary.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        summary.columnconfigure(0, weight=1)
        summary.columnconfigure(1, weight=0)
        self.foq_alignment_summary_title = tk.Label(
            summary,
            text="当前验证: 请选择一个测试",
            font=self._font(12, "bold"),
            bg=self.colors["card_alt"],
            fg=self.colors["text"],
            anchor="w",
        )
        self.foq_alignment_summary_title.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 2))
        self.foq_alignment_summary_status = tk.Label(
            summary,
            text="拆解状态: -    Open Verification: -    最后更新: -",
            font=self._font(9),
            bg=self.colors["card_alt"],
            fg=self.colors["text_secondary"],
            anchor="w",
        )
        self.foq_alignment_summary_status.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))
        self._make_button(summary, "Export Verification Report", self.export_foq_kb_to_run_alignment, kind="primary", width=24).grid(row=0, column=1, rowspan=2, sticky="e", padx=12, pady=10)

        cards = tk.Frame(top, bg=self.colors["card"])
        cards.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        self.foq_alignment_contract_cards: dict[str, dict[str, tk.Label | tk.Frame]] = {}
        for index, title in enumerate(("Config", "Method", "Report", "DB", "Processing", "Open")):
            cards.columnconfigure(index, weight=1, uniform="foq_contract_cards")
            card = tk.Frame(cards, bg="#F8FAFC", highlightbackground=self.colors["card_border"], highlightthickness=1)
            card.grid(row=0, column=index, sticky="nsew", padx=(0 if index == 0 else 6, 0))
            title_label = tk.Label(card, text=title, font=self._font(8, "bold"), bg="#F8FAFC", fg=self.colors["text_secondary"], anchor="w")
            title_label.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 0))
            status_label = tk.Label(card, text="-", font=self._font(10, "bold"), bg="#F8FAFC", fg=self.colors["text"], anchor="w")
            status_label.grid(row=1, column=0, sticky="ew", padx=8, pady=(2, 0))
            detail_label = tk.Label(card, text="-", font=self._font(8), bg="#F8FAFC", fg=self.colors["text_secondary"], anchor="w", justify="left", wraplength=170)
            detail_label.grid(row=2, column=0, sticky="ew", padx=8, pady=(2, 6))
            target_tab = {
                "Config": "Config / Open",
                "Method": "Method",
                "Processing": "Processing",
                "Report": "Report / DB",
                "DB": "Report / DB",
                "Open": "Config / Open",
            }[title]
            for widget in (card, title_label, status_label, detail_label):
                widget.bind("<Button-1>", lambda _event, tab_name=target_tab: self._select_foq_alignment_detail_tab(tab_name))
            self.foq_alignment_contract_cards[title] = {"frame": card, "title": title_label, "status": status_label, "detail": detail_label}

        self.foq_kb_to_run_table = self._make_table(
            top,
            (
                "family",
                "td_item",
                "device",
                "coverage",
                "open_gaps_count",
                "action",
            ),
            {
                "family": "Family",
                "td_item": "Test",
                "device": "Device",
                "coverage": "Coverage",
                "open_gaps_count": "Open",
                "action": "Next / Blocker",
            },
        )
        self.foq_kb_to_run_table.grid(row=2, column=0, sticky="nsew")
        kb_table = self._table_widget(self.foq_kb_to_run_table)
        kb_table.column("family", width=90)
        kb_table.column("td_item", width=280)
        kb_table.column("device", width=190)
        kb_table.column("coverage", width=150)
        kb_table.column("open_gaps_count", width=70, anchor="center")
        kb_table.column("action", width=520)
        kb_table.tag_configure("alignment_ready", background="#DCFCE7")
        kb_table.tag_configure("alignment_review", background="#FEF3C7")
        kb_table.tag_configure("alignment_blocked", background="#FEE2E2")
        kb_table.bind("<<TreeviewSelect>>", self._preview_selected_foq_kb_to_run_row)

        tk.Label(bottom, text="Selected Test Verification", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 4))
        self.foq_alignment_detail_notebook = ttk.Notebook(bottom)
        self.foq_alignment_detail_notebook.grid(row=1, column=0, sticky="nsew")
        self.foq_alignment_detail_texts: dict[str, tk.Text] = {}
        for title in ("Overview", "Config / Open", "Method", "Report / DB", "Processing"):
            detail_tab = tk.Frame(self.foq_alignment_detail_notebook, bg=self.colors["card"])
            detail_tab.columnconfigure(0, weight=1)
            detail_tab.rowconfigure(0, weight=1)
            text_widget = tk.Text(detail_tab, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
            text_widget.grid(row=0, column=0, sticky="nsew")
            scroll = ttk.Scrollbar(detail_tab, orient="vertical", command=text_widget.yview)
            scroll.grid(row=0, column=1, sticky="ns")
            text_widget.configure(yscrollcommand=scroll.set, state="disabled")
            self.foq_alignment_detail_texts[title] = text_widget
            self.foq_alignment_detail_notebook.add(detail_tab, text=title)

        paned.add(top, minsize=220, height=320)
        paned.add(bottom, minsize=220, height=360)
        self.notebook.add(tab, text="FOQ Knowledge Alignment")
        self._refresh_foq_alignment_catalog()

    def _build_test_plan_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)

        controls = tk.Frame(tab, bg=self.colors["card"])
        controls.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        controls.columnconfigure(11, weight=1)

        tk.Label(controls, text="Family", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.test_plan_family_combo = ttk.Combobox(controls, textvariable=self.test_plan_family_var, values=("TCC", "VDAD"), state="readonly", width=10)
        self.test_plan_family_combo.grid(row=0, column=1, sticky="w", padx=(0, 10))
        self.test_plan_family_combo.bind("<<ComboboxSelected>>", self._refresh_test_plan_options)

        tk.Label(controls, text="Test / 原稿", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="w", padx=(0, 6))
        self.test_plan_test_combo = ttk.Combobox(controls, textvariable=self.test_plan_test_var, values=(), width=30)
        self.test_plan_test_combo.grid(row=0, column=3, sticky="w", padx=(0, 10))
        self.test_plan_test_combo.bind("<<ComboboxSelected>>", self._analyze_test_plan_intent)

        tk.Label(controls, text="Device", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=4, sticky="w", padx=(0, 6))
        self.test_plan_device_combo = ttk.Combobox(controls, textvariable=self.test_plan_device_var, values=(), width=14)
        self.test_plan_device_combo.grid(row=0, column=5, sticky="w", padx=(0, 10))
        self.test_plan_device_combo.bind("<<ComboboxSelected>>", self._analyze_test_plan_intent)

        tk.Label(controls, text="Action", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=6, sticky="w", padx=(0, 6))
        self.test_plan_intent_combo = ttk.Combobox(controls, textvariable=self.test_plan_intent_var, values=intent_tool_options(), state="readonly", width=18)
        self.test_plan_intent_combo.grid(row=0, column=7, sticky="w", padx=(0, 10))
        self.test_plan_intent_combo.bind("<<ComboboxSelected>>", self._analyze_test_plan_intent)

        tk.Label(controls, text="Parameter", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=8, sticky="w", padx=(0, 6))
        parameter_entry = tk.Entry(controls, textvariable=self.test_plan_parameter_var, width=20, font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        parameter_entry.grid(row=0, column=9, sticky="w", padx=(0, 10))
        parameter_entry.bind("<Return>", self._analyze_test_plan_intent)

        self._make_button(controls, "Analyze", self._analyze_test_plan_intent, kind="primary", width=12).grid(row=0, column=10, sticky="w")
        self._make_button(controls, "Export Draft", self.export_test_plan_draft_packet, kind="neutral", width=14).grid(row=0, column=12, sticky="e", padx=(8, 0))
        self._make_button(controls, "Open Output", self.open_output_folder, kind="neutral", width=12).grid(row=0, column=13, sticky="e", padx=(8, 0))

        intent_row = tk.Frame(tab, bg=self.colors["card"])
        intent_row.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        intent_row.columnconfigure(1, weight=1)
        tk.Label(intent_row, text="Your intent / 输入意图", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        free_entry = tk.Entry(intent_row, textvariable=self.test_plan_free_intent_var, font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        free_entry.grid(row=0, column=1, sticky="ew")
        free_entry.bind("<Return>", self._analyze_test_plan_intent)
        self._make_toolbar_label(intent_row, "Examples: accuracy 40 C; HeatUp 20->50->20; 合并 Accuracy Stability").grid(row=0, column=2, sticky="e", padx=(10, 0))

        paned = tk.PanedWindow(tab, orient=tk.HORIZONTAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        paned.grid(row=2, column=0, sticky="nsew")
        left = tk.Frame(paned, bg=self.colors["card"])
        right = tk.Frame(paned, bg=self.colors["card"])
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        tk.Label(left, text="Concrete Modification Targets", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.test_plan_template_table = self._make_table(
            left,
            ("asset", "template", "location", "change", "output"),
            {
                "asset": "Asset",
                "template": "Use Template",
                "location": "Modify Exact Location",
                "change": "Change To / Check",
                "output": "Output",
            },
        )
        self.test_plan_template_table.grid(row=1, column=0, sticky="nsew")
        template_widget = self._table_widget(self.test_plan_template_table)
        template_widget.column("asset", width=110)
        template_widget.column("template", width=200)
        template_widget.column("location", width=280)
        template_widget.column("change", width=360)
        template_widget.column("output", width=240)

        tk.Label(right, text="Modification Advice / Review Plan", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        text_frame = tk.Frame(right, bg=self.colors["card"])
        text_frame.grid(row=1, column=0, sticky="nsew")
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        self.test_plan_preview_text = tk.Text(text_frame, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        self.test_plan_preview_text.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.test_plan_preview_text.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.test_plan_preview_text.configure(yscrollcommand=scroll.set, state="disabled")

        paned.add(left, minsize=420, width=620)
        paned.add(right, minsize=520, width=880)
        self.notebook.add(tab, text="Test Plan")
        self._refresh_test_plan_options()

    def _build_test_plan_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)

        controls = tk.Frame(tab, bg=self.colors["card"])
        controls.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        controls.columnconfigure(11, weight=1)

        tk.Label(controls, text="Family", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.test_plan_family_combo = ttk.Combobox(controls, textvariable=self.test_plan_family_var, values=("TCC", "VDAD"), state="readonly", width=10)
        self.test_plan_family_combo.grid(row=0, column=1, sticky="w", padx=(0, 10))
        self.test_plan_family_combo.bind("<<ComboboxSelected>>", self._refresh_test_plan_options)

        tk.Label(controls, text="Test / Template", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="w", padx=(0, 6))
        self.test_plan_test_combo = ttk.Combobox(controls, textvariable=self.test_plan_test_var, values=(), width=30)
        self.test_plan_test_combo.grid(row=0, column=3, sticky="w", padx=(0, 10))
        self.test_plan_test_combo.bind("<<ComboboxSelected>>", self._analyze_test_plan_intent)

        tk.Label(controls, text="Device", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=4, sticky="w", padx=(0, 6))
        self.test_plan_device_combo = ttk.Combobox(controls, textvariable=self.test_plan_device_var, values=(), width=14)
        self.test_plan_device_combo.grid(row=0, column=5, sticky="w", padx=(0, 10))
        self.test_plan_device_combo.bind("<<ComboboxSelected>>", self._analyze_test_plan_intent)

        self._make_button(controls, "AI 分析", self.ai_analyze_test_plan_intent, kind="primary", width=12).grid(row=0, column=6, sticky="w", padx=(8, 0))
        self._make_button(controls, "规则分析", self._analyze_test_plan_intent, kind="neutral", width=12).grid(row=0, column=7, sticky="w", padx=(8, 0))
        self._make_toolbar_label(controls, "AI identifies intent; method/report rendering remains evidence-based.").grid(row=0, column=11, sticky="w", padx=(12, 0))
        self._make_button(controls, "Export Assets", self.export_test_plan_draft_packet, kind="secondary", width=14).grid(row=0, column=12, sticky="e", padx=(8, 0))
        self._make_button(controls, "Open Output", self.open_output_folder, kind="neutral", width=12).grid(row=0, column=13, sticky="e", padx=(8, 0))

        intent_row = tk.Frame(tab, bg=self.colors["card"])
        intent_row.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        intent_row.columnconfigure(1, weight=1)
        tk.Label(intent_row, text="Intent", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        free_entry = tk.Entry(intent_row, textvariable=self.test_plan_free_intent_var, font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        free_entry.grid(row=0, column=1, sticky="ew")
        free_entry.bind("<Return>", self._analyze_test_plan_intent)
        self._make_toolbar_label(intent_row, "Examples: accuracy 40 C; HeatUp 25->45->25; merge Accuracy Stability").grid(row=0, column=2, sticky="e", padx=(10, 0))

        candidates_panel = tk.Frame(tab, bg=self.colors["card"])
        candidates_panel.grid(row=2, column=0, sticky="nsew", pady=(0, 8))
        candidates_panel.columnconfigure(0, weight=1)
        candidates_panel.rowconfigure(1, weight=1)
        tk.Label(candidates_panel, text="1. Template Candidates", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.test_plan_candidate_table = self._make_table(
            candidates_panel,
            ("layer", "template", "source", "confidence", "output"),
            {
                "layer": "Layer",
                "template": "Selected Template / Source",
                "source": "Evidence Location",
                "confidence": "Confidence",
                "output": "Primary Output",
            },
        )
        self.test_plan_candidate_table.grid(row=1, column=0, sticky="nsew")
        candidate_widget = self._table_widget(self.test_plan_candidate_table)
        candidate_widget.column("layer", width=130)
        candidate_widget.column("template", width=260)
        candidate_widget.column("source", width=520)
        candidate_widget.column("confidence", width=160)
        candidate_widget.column("output", width=280)

        detail_notebook = ttk.Notebook(tab, style="Explorer.TNotebook")
        detail_notebook.grid(row=3, column=0, sticky="nsew")

        method_tab = tk.Frame(detail_notebook, bg=self.colors["card"])
        method_tab.columnconfigure(0, weight=1)
        method_tab.columnconfigure(1, weight=1)
        method_tab.rowconfigure(1, weight=1)
        tk.Label(method_tab, text="2. Instrument Method - Original", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6), padx=(0, 8))
        tk.Label(method_tab, text="Modified Preview", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=1, sticky="w", pady=(8, 6), padx=(8, 0))
        self.test_plan_method_source_table = self._make_method_preview_table(method_tab)
        self.test_plan_method_source_table.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
        self.test_plan_method_modified_table = self._make_method_preview_table(method_tab)
        self.test_plan_method_modified_table.grid(row=1, column=1, sticky="nsew", padx=(8, 0))
        self.test_plan_method_table = self.test_plan_method_modified_table

        report_tab = tk.Frame(detail_notebook, bg=self.colors["card"])
        report_tab.columnconfigure(0, weight=1)
        report_tab.rowconfigure(1, weight=1)
        tk.Label(report_tab, text="3. Report Template / Formula Diff", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6))
        self.test_plan_report_table = self._make_test_plan_diff_table(report_tab)
        self.test_plan_report_table.grid(row=1, column=0, sticky="nsew")

        config_tab = tk.Frame(detail_notebook, bg=self.colors["card"])
        config_tab.columnconfigure(0, weight=1)
        config_tab.rowconfigure(1, weight=1)
        tk.Label(config_tab, text="4. Config Checklist / Open Verification", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6))
        self.test_plan_config_table = self._make_table(
            config_tab,
            ("item", "source", "status", "action"),
            {
                "item": "Item",
                "source": "Source",
                "status": "Status",
                "action": "Required Action",
            },
        )
        self.test_plan_config_table.grid(row=1, column=0, sticky="nsew")
        config_widget = self._table_widget(self.test_plan_config_table)
        config_widget.column("item", width=270)
        config_widget.column("source", width=420)
        config_widget.column("status", width=180)
        config_widget.column("action", width=620)

        review_tab = tk.Frame(detail_notebook, bg=self.colors["card"])
        review_tab.columnconfigure(0, weight=1)
        review_tab.rowconfigure(1, weight=1)
        review_toolbar = tk.Frame(review_tab, bg=self.colors["card"])
        review_toolbar.grid(row=0, column=0, sticky="ew", pady=(8, 6))
        review_toolbar.columnconfigure(1, weight=1)
        tk.Label(review_toolbar, text="Review Plan - edit this text first, then generate method diff", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w")
        self._make_button(review_toolbar, "Generate Method Diff", self.apply_test_plan_review_to_method_diff, kind="primary", width=20).grid(row=0, column=2, sticky="e")
        text_frame = tk.Frame(review_tab, bg=self.colors["card"])
        text_frame.grid(row=1, column=0, sticky="nsew")
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        self.test_plan_preview_text = tk.Text(text_frame, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        self.test_plan_preview_text.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.test_plan_preview_text.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.test_plan_preview_text.configure(yscrollcommand=scroll.set)

        detail_notebook.add(method_tab, text="Method Script Diff")
        detail_notebook.add(report_tab, text="Report Template Diff")
        detail_notebook.add(config_tab, text="Config / Open Verification")
        detail_notebook.add(review_tab, text="Review Plan")
        self.notebook.add(tab, text="Test Plan")
        self._refresh_test_plan_options()

    def _make_test_plan_diff_table(self, parent: tk.Widget) -> tk.Frame:
        table_frame = self._make_table(
            parent,
            ("asset", "template", "location", "change", "output"),
            {
                "asset": "Asset",
                "template": "Use Template",
                "location": "Modify Exact Location",
                "change": "Change To / Check",
                "output": "Output",
            },
        )
        table_widget = self._table_widget(table_frame)
        table_widget.column("asset", width=140)
        table_widget.column("template", width=220)
        table_widget.column("location", width=360)
        table_widget.column("change", width=560)
        table_widget.column("output", width=260)
        return table_frame

    def _make_method_preview_table(self, parent: tk.Widget) -> tk.Frame:
        table_frame = self._make_table(
            parent,
            ("row", "kind", "time", "command", "value", "comment"),
            {
                "row": "#",
                "kind": "Kind",
                "time": "Time",
                "command": "Command",
                "value": "Value",
                "comment": "Comment",
            },
        )
        table_widget = self._table_widget(table_frame)
        table_widget.column("row", width=54, anchor="e")
        table_widget.column("kind", width=90)
        table_widget.column("time", width=115)
        table_widget.column("command", width=360)
        table_widget.column("value", width=300)
        table_widget.column("comment", width=460)
        table_widget.tag_configure("modified", background="#FEE2E2", foreground="#7F1D1D")
        table_widget.tag_configure("removed", background="#F3F4F6", foreground="#6B7280")
        table_widget.tag_configure("source", background="#F8FAFC", foreground=self.colors["text"])
        table_widget.tag_configure("cm_initial", background="#FED7AA")
        table_widget.tag_configure("cm_condition", background="#BEF29A")
        table_widget.tag_configure("cm_comment", foreground="#15803D")
        table_widget.tag_configure("cm_header", foreground="#15803D")
        return table_frame

    def _build_test_plan_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)

        controls = tk.Frame(tab, bg=self.colors["card"])
        controls.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        controls.columnconfigure(4, weight=1)

        tk.Label(controls, text="Family", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.test_plan_family_combo = ttk.Combobox(controls, textvariable=self.test_plan_family_var, values=("TCC", "VDAD"), state="readonly", width=10)
        self.test_plan_family_combo.grid(row=0, column=1, sticky="w", padx=(0, 10))
        self.test_plan_family_combo.bind("<<ComboboxSelected>>", self._refresh_test_plan_options)

        self._make_button(controls, "Analyze Method Roles", self._analyze_test_plan_intent, kind="primary", width=20).grid(row=0, column=2, sticky="w", padx=(8, 0))
        self._make_button(controls, "Generate Method Script", self.apply_test_plan_review_to_method_diff, kind="secondary", width=22).grid(row=0, column=3, sticky="w", padx=(8, 0))
        self._make_toolbar_label(controls, "Method basis comes from KB method scripts. CMBX is only used offline to refresh the KB.").grid(row=0, column=4, sticky="ew", padx=(14, 0))

        intent_row = tk.Frame(tab, bg=self.colors["card"])
        intent_row.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        intent_row.columnconfigure(1, weight=1)
        tk.Label(intent_row, text="Intent", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        free_entry = tk.Entry(intent_row, textvariable=self.test_plan_free_intent_var, font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        free_entry.grid(row=0, column=1, sticky="ew")
        free_entry.bind("<Return>", self._analyze_test_plan_intent)
        self._make_toolbar_label(intent_row, "Example: accuracy from stable 20 C to target 40 C").grid(row=0, column=2, sticky="e", padx=(10, 0))

        basis_panel = tk.Frame(tab, bg=self.colors["card"])
        basis_panel.grid(row=2, column=0, sticky="ew", pady=(0, 8))
        basis_panel.columnconfigure(0, weight=1)
        tk.Label(basis_panel, text="Intent Pipeline", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.test_plan_pipeline_table = self._make_table(
            basis_panel,
            ("step", "result", "evidence", "decision"),
            {
                "step": "Step",
                "result": "Current Result",
                "evidence": "Evidence Source",
                "decision": "Decision / Constraint",
            },
        )
        self.test_plan_pipeline_table.grid(row=1, column=0, sticky="ew")
        pipeline_widget = self._table_widget(self.test_plan_pipeline_table)
        pipeline_widget.column("step", width=180)
        pipeline_widget.column("result", width=440)
        pipeline_widget.column("evidence", width=420)
        pipeline_widget.column("decision", width=480)

        tk.Label(basis_panel, text="Method Basis", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=2, column=0, sticky="w", pady=(8, 6))
        self.test_plan_basis_table = self._make_table(
            basis_panel,
            ("item", "value", "status"),
            {
                "item": "Item",
                "value": "Basis / Detected Value",
                "status": "Status",
            },
        )
        self.test_plan_basis_table.grid(row=3, column=0, sticky="ew")
        basis_widget = self._table_widget(self.test_plan_basis_table)
        basis_widget.column("item", width=180)
        basis_widget.column("value", width=820)
        basis_widget.column("status", width=340)
        tk.Label(basis_panel, text="Learned Method Roles", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=4, column=0, sticky="w", pady=(8, 6))
        self.test_plan_role_table = self._make_table(
            basis_panel,
            ("role", "rows", "meaning", "handling"),
            {
                "role": "Role",
                "rows": "Row(s)",
                "meaning": "Learned Meaning",
                "handling": "How It Serves Changes",
            },
        )
        self.test_plan_role_table.grid(row=5, column=0, sticky="ew")
        role_widget = self._table_widget(self.test_plan_role_table)
        role_widget.column("role", width=220)
        role_widget.column("rows", width=160)
        role_widget.column("meaning", width=620)
        role_widget.column("handling", width=520)

        main_paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        main_paned.grid(row=3, column=0, sticky="nsew")

        review_frame = tk.Frame(main_paned, bg=self.colors["card"])
        review_frame.columnconfigure(0, weight=1)
        review_frame.rowconfigure(1, weight=1)
        tk.Label(review_frame, text="1. Editable Method Change Plan", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        text_frame = tk.Frame(review_frame, bg=self.colors["card"])
        text_frame.grid(row=1, column=0, sticky="nsew")
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        self.test_plan_preview_text = tk.Text(text_frame, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        self.test_plan_preview_text.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.test_plan_preview_text.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.test_plan_preview_text.configure(yscrollcommand=scroll.set)

        method_frame = tk.Frame(main_paned, bg=self.colors["card"])
        method_frame.columnconfigure(0, weight=1)
        method_frame.columnconfigure(1, weight=1)
        method_frame.rowconfigure(1, weight=1)
        tk.Label(method_frame, text="2. Original Method Script", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6), padx=(0, 8))
        tk.Label(method_frame, text="Modified Method Script Preview", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=1, sticky="w", pady=(8, 6), padx=(8, 0))
        self.test_plan_method_source_table = self._make_method_preview_table(method_frame)
        self.test_plan_method_source_table.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
        self.test_plan_method_modified_table = self._make_method_preview_table(method_frame)
        self.test_plan_method_modified_table.grid(row=1, column=1, sticky="nsew", padx=(8, 0))
        self.test_plan_method_table = self.test_plan_method_modified_table

        main_paned.add(review_frame, minsize=150, height=240)
        main_paned.add(method_frame, minsize=260, height=520)
        self.notebook.add(tab, text="Test Plan")
        self._refresh_test_plan_options()

    def _build_audit_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Audit trails for the selected injection or sequence.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Export Selected Audits", self.export_selected_audits, kind="primary", width=20).grid(row=0, column=1, sticky="e", padx=(8, 0))
        audit_paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        audit_paned.grid(row=1, column=0, sticky="nsew")
        audit_top = tk.Frame(audit_paned, bg=self.colors["card"])
        audit_top.columnconfigure(0, weight=1)
        audit_top.rowconfigure(0, weight=1)
        audit_bottom = tk.Frame(audit_paned, bg=self.colors["card"])
        audit_bottom.columnconfigure(0, weight=1)
        audit_bottom.rowconfigure(1, weight=1)
        self.audit_table = self._make_table(audit_top, ("name", "size", "raw", "url"), {"name": "Audit", "size": "Size", "raw": "Raw File", "url": "URL"})
        self.audit_table.grid(row=0, column=0, sticky="nsew")
        audit_widget = self._table_widget(self.audit_table)
        audit_widget.bind("<Double-1>", self.open_selected_audits)
        audit_widget.bind("<<TreeviewSelect>>", self.preview_selected_audit)
        tk.Label(audit_bottom, text="CM Audit Trail Preview", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6))
        self.audit_preview_table = self._make_table(
            audit_bottom,
            ("row", "day_time", "ret_time", "message"),
            {"row": "", "day_time": "Day Time", "ret_time": "Ret. Time", "message": "Command/Message"},
        )
        self.audit_preview_table.grid(row=1, column=0, sticky="nsew")
        preview_widget = self._table_widget(self.audit_preview_table)
        preview_widget.tag_configure("audit_title", background="#FFFFFF", foreground="#000000")
        preview_widget.tag_configure("audit_meta", background="#FFFFFF", foreground="#000000")
        preview_widget.tag_configure("audit_header", background="#E5E7EB", foreground="#000000")
        preview_widget.tag_configure("audit_stage", background="#FED7AA")
        preview_widget.tag_configure("audit_condition", background="#BEF29A")
        audit_paned.add(audit_top, minsize=110, height=175)
        audit_paned.add(audit_bottom, minsize=240, height=430)
        self.notebook.add(tab, text="Audit Trails")

    def _build_report_sheets_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Report sheets for the selected injection or sequence. Export filled Chromeleon-style .xls reports here.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Export Filled CM Report", self.export_report_sheets_workbook, kind="primary", width=23).grid(row=0, column=1, sticky="e", padx=(8, 0))
        report_paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        report_paned.grid(row=1, column=0, sticky="nsew")
        report_top = tk.Frame(report_paned, bg=self.colors["card"])
        report_top.columnconfigure(0, weight=1)
        report_top.rowconfigure(0, weight=1)
        report_bottom = tk.Frame(report_paned, bg=self.colors["card"])
        report_bottom.columnconfigure(0, weight=1)
        report_bottom.rowconfigure(1, weight=1)
        self.report_sheet_table = self._make_table(
            report_top,
            ("injection", "sheet", "applies", "report", "objects", "formulas", "active", "each", "condition", "id"),
            {
                "injection": "Injection",
                "sheet": "Sheet",
                "applies": "For Injection",
                "report": "Report Template",
                "objects": "Objects",
                "formulas": "Formulas",
                "active": "Active",
                "each": "Each Inj.",
                "condition": "Condition / Reason",
                "id": "Sheet Id",
            },
        )
        self.report_sheet_table.grid(row=0, column=0, sticky="nsew")
        report_sheet_widget = self._table_widget(self.report_sheet_table)
        report_sheet_widget.tag_configure("name_match", background=self.colors["match"])
        report_sheet_widget.bind("<Double-1>", self.open_report_sheet_workbook)
        report_sheet_widget.bind("<<TreeviewSelect>>", self.preview_selected_report_sheet)
        tk.Label(report_bottom, text="Report Sheet Preview", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6))
        self.report_sheet_preview_table = self._make_excel_preview_table(report_bottom)
        self.report_sheet_preview_table.grid(row=1, column=0, sticky="nsew")
        self._configure_excel_preview_tags(self._table_widget(self.report_sheet_preview_table))
        report_paned.add(report_top, minsize=150, height=260)
        report_paned.add(report_bottom, minsize=240, height=420)
        self.notebook.add(tab, text="Report Sheets")

    def _build_method_context_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Instrument methods only. Select a method to preview; drag methods into the compare box below.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Add To Compare", self.add_selected_method_to_compare, kind="secondary", width=16).grid(row=0, column=1, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Export Selected Methods", self.export_selected_methods, kind="primary", width=21).grid(row=0, column=2, sticky="e", padx=(8, 0))
        method_paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        method_paned.grid(row=1, column=0, sticky="nsew")
        method_top = tk.Frame(method_paned, bg=self.colors["card"])
        method_top.columnconfigure(0, weight=1)
        method_top.rowconfigure(0, weight=1)
        method_top.rowconfigure(2, weight=0)
        method_top.rowconfigure(3, weight=1)
        method_bottom = tk.Frame(method_paned, bg=self.colors["card"])
        method_bottom.columnconfigure(0, weight=1)
        method_bottom.rowconfigure(1, weight=1)
        self.method_context_table = self._make_table(
            method_top,
            ("role", "name", "type", "source"),
            {"role": "Role", "name": "Name", "type": "Object Type", "source": "CMBX Source / Reference"},
        )
        self.method_context_table.grid(row=0, column=0, sticky="nsew")
        method_context_widget = self._table_widget(self.method_context_table)
        method_context_widget.bind("<Double-1>", self.open_selected_methods)
        method_context_widget.bind("<<TreeviewSelect>>", self.preview_selected_method)
        method_context_widget.bind("<ButtonPress-1>", self._begin_method_drag, add="+")
        method_context_widget.bind("<ButtonRelease-1>", self._finish_method_drag, add="+")
        tk.Label(method_top, text="Selected Methods For Compare", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=2, column=0, sticky="w", pady=(8, 6))
        self.method_compare_selection_table = self._make_table(
            method_top,
            ("benchmark", "cmbx", "sequence", "method", "source"),
            {"benchmark": "Benchmark", "cmbx": "CMBX", "sequence": "Sequence", "method": "Method", "source": "Source"},
        )
        self.method_compare_selection_table.grid(row=3, column=0, sticky="nsew")
        compare_widget = self._table_widget(self.method_compare_selection_table)
        compare_widget.bind("<Double-1>", self._remove_method_compare_item)
        tk.Label(method_bottom, text="CM Method Table Preview", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6))
        self.method_flow_table = self._make_table(
            method_bottom,
            ("row", "kind", "time", "command", "value", "comment"),
            {
                "row": "#",
                "kind": "Kind",
                "time": "Time",
                "command": "Command",
                "value": "Value",
                "comment": "Comment",
            },
        )
        self.method_flow_table.grid(row=1, column=0, sticky="nsew")
        flow_widget = self._table_widget(self.method_flow_table)
        flow_widget.column("row", width=54, anchor="e")
        flow_widget.tag_configure("cm_initial", background="#FED7AA")
        flow_widget.tag_configure("cm_condition", background="#BEF29A")
        flow_widget.tag_configure("cm_comment", foreground="#15803D")
        flow_widget.tag_configure("cm_header", foreground="#15803D")
        method_paned.add(method_top, minsize=150, height=220)
        method_paned.add(method_bottom, minsize=180, height=360)
        self.notebook.add(tab, text="Instrument Methods")

    def _build_method_editor_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(1, weight=1)
        tk.Label(toolbar, text="MD Script", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self._make_entry(toolbar, self.method_editor_md_path_var).grid(row=0, column=1, sticky="ew", ipady=4)
        self._make_button(toolbar, "Browse MD", self.browse_method_editor_md, kind="neutral", width=12).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Preview", self.render_method_editor_md, kind="secondary", width=12).grid(row=0, column=3, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Generate CMBX", self.export_method_editor_xlsx, kind="primary", width=16).grid(row=0, column=4, sticky="e", padx=(8, 0))

        out_row = tk.Frame(tab, bg=self.colors["card"])
        out_row.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        out_row.columnconfigure(1, weight=1)
        tk.Label(out_row, text="Output CMBX", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self._make_entry(out_row, self.method_editor_xlsx_path_var).grid(row=0, column=1, sticky="ew", ipady=4)
        self._make_button(out_row, "Save As", self.browse_method_editor_output, kind="neutral", width=10).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(out_row, "Open Output", self.open_method_editor_xlsx, kind="neutral", width=13).grid(row=0, column=3, sticky="e", padx=(8, 0))

        status = tk.Label(tab, textvariable=self.method_editor_status_var, anchor="w", justify="left", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"])
        status.grid(row=2, column=0, sticky="ew", pady=(0, 8))

        content_paned = ttk.PanedWindow(tab, orient=tk.VERTICAL)
        content_paned.grid(row=3, column=0, sticky="nsew")

        table_frame = tk.Frame(content_paned, bg=self.colors["card"])
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)
        tk.Label(table_frame, text="CM Method Preview", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.method_editor_table = self._make_table(
            table_frame,
            ("row", "kind", "time", "command", "value", "comment"),
            {
                "row": "#",
                "kind": "Kind",
                "time": "Time",
                "command": "Command",
                "value": "Value",
                "comment": "Comment",
            },
        )
        self.method_editor_table.grid(row=1, column=0, sticky="nsew")
        editor_widget = self._table_widget(self.method_editor_table)
        editor_widget.column("row", width=54, anchor="e")
        editor_widget.column("kind", width=95)
        editor_widget.column("time", width=130)
        editor_widget.column("command", width=430)
        editor_widget.column("value", width=460)
        editor_widget.column("comment", width=560)
        editor_widget.tag_configure("cm_stage", background="#F4B183", foreground=self.colors["text"])
        editor_widget.tag_configure("cm_branch", background="#C6EFCE", foreground="#14532D")
        editor_widget.tag_configure("cm_comment", foreground="#15803D")
        editor_widget.tag_configure("cm_command", background="#F8FAFC", foreground=self.colors["text"])
        editor_widget.tag_configure("cm_end", background="#F4B183", foreground="#7C2D12")
        editor_widget.tag_configure("cm_invalid", background="#FEE2E2", foreground="#991B1B")
        content_paned.add(table_frame, weight=3)

        kb_frame = tk.Frame(content_paned, bg=self.colors["card"])
        kb_frame.columnconfigure(0, weight=1)
        kb_frame.rowconfigure(1, weight=1)
        kb_toolbar = tk.Frame(kb_frame, bg=self.colors["card"])
        kb_toolbar.grid(row=0, column=0, sticky="ew", pady=(8, 6))
        kb_toolbar.columnconfigure(5, weight=1)
        tk.Label(kb_toolbar, text="KB Reference", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=(0, 10))
        self._make_button(kb_toolbar, "Packaging KB", lambda: self.render_method_editor_kb_doc("MD_TO_STANDALONE_METHOD_CMBX_PACKAGING.md"), kind="neutral", width=16).grid(row=0, column=1, sticky="w", padx=(0, 8))
        self._make_button(kb_toolbar, "Format Spec", lambda: self.render_method_editor_kb_doc("CM_METHOD_SCRIPT_MD_FORMAT_SPEC.md"), kind="neutral", width=14).grid(row=0, column=2, sticky="w", padx=(0, 8))
        self._make_button(kb_toolbar, "Compiler Rules", lambda: self.render_method_editor_kb_doc("CM Compiler Rules.MD"), kind="neutral", width=15).grid(row=0, column=3, sticky="w", padx=(0, 8))
        self._make_button(kb_toolbar, "Open KB File", self.open_method_editor_kb_doc, kind="secondary", width=13).grid(row=0, column=4, sticky="w")
        tk.Label(kb_toolbar, textvariable=self.method_editor_kb_doc_var, font=self._font(8), bg=self.colors["card"], fg=self.colors["text_secondary"], anchor="e").grid(row=0, column=5, sticky="e")
        kb_text_frame = tk.Frame(kb_frame, bg=self.colors["card"])
        kb_text_frame.grid(row=1, column=0, sticky="nsew")
        kb_text_frame.columnconfigure(0, weight=1)
        kb_text_frame.rowconfigure(0, weight=1)
        self.method_editor_kb_text = tk.Text(kb_text_frame, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1, padx=8, pady=6)
        self.method_editor_kb_text.grid(row=0, column=0, sticky="nsew")
        kb_scroll = ttk.Scrollbar(kb_text_frame, orient="vertical", command=self.method_editor_kb_text.yview)
        kb_scroll.grid(row=0, column=1, sticky="ns")
        self.method_editor_kb_text.configure(yscrollcommand=kb_scroll.set, state="disabled")
        content_paned.add(kb_frame, weight=1)
        self.render_method_editor_kb_doc(self.method_editor_kb_doc_var.get())
        self.notebook.add(tab, text="Method Script Generator")

    def _build_processing_methods_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Processing methods for the selected sequence or injection context. Kept separate from instrument method logic.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Export Selected Processing", self.export_selected_processing, kind="primary", width=24).grid(row=0, column=1, sticky="e", padx=(8, 0))
        self.processing_method_table = self._make_table(
            tab,
            ("role", "name", "type", "source"),
            {"role": "Role", "name": "Name", "type": "Object Type", "source": "CMBX Source"},
        )
        self.processing_method_table.grid(row=1, column=0, sticky="nsew")
        self._table_widget(self.processing_method_table).bind("<Double-1>", self.open_selected_processing)
        self.notebook.add(tab, text="Processing Methods")

    def _build_report_templates_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(0, weight=1)
        self._make_toolbar_label(toolbar, "Report templates only. Export or double-click to open the blank Chromeleon .xls template.").grid(row=0, column=0, sticky="ew")
        self._make_button(toolbar, "Export Blank Template", self.export_selected_templates, kind="primary", width=22).grid(row=0, column=1, sticky="e", padx=(8, 0))
        template_paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        template_paned.grid(row=1, column=0, sticky="nsew")
        template_top = tk.Frame(template_paned, bg=self.colors["card"])
        template_top.columnconfigure(0, weight=1)
        template_top.rowconfigure(0, weight=1)
        template_bottom = tk.Frame(template_paned, bg=self.colors["card"])
        template_bottom.columnconfigure(0, weight=1)
        template_bottom.rowconfigure(2, weight=1)
        self.report_template_table = self._make_table(
            template_top,
            ("role", "name", "type", "source"),
            {"role": "Role", "name": "Name", "type": "Object Type", "source": "CMBX Source"},
        )
        self.report_template_table.grid(row=0, column=0, sticky="nsew")
        template_widget = self._table_widget(self.report_template_table)
        template_widget.bind("<Double-1>", self.open_selected_templates)
        template_widget.bind("<<TreeviewSelect>>", self.preview_selected_report_template)
        template_header = tk.Frame(template_bottom, bg=self.colors["card"])
        template_header.grid(row=0, column=0, sticky="ew", pady=(8, 6))
        template_header.columnconfigure(1, weight=1)
        tk.Label(template_header, text="Report Template Preview", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=(0, 12))
        self.report_template_sheet_combo = ttk.Combobox(template_header, textvariable=self.report_template_sheet_var, state="readonly", width=32, font=self._font(9))
        self.report_template_sheet_combo.grid(row=0, column=1, sticky="w")
        self.report_template_sheet_combo.bind("<<ComboboxSelected>>", self.preview_selected_report_template)
        self.report_template_preview_table = self._make_excel_preview_table(template_bottom)
        self.report_template_preview_table.grid(row=2, column=0, sticky="nsew")
        self._configure_excel_preview_tags(self._table_widget(self.report_template_preview_table))
        template_paned.add(template_top, minsize=140, height=220)
        template_paned.add(template_bottom, minsize=240, height=460)
        self.notebook.add(tab, text="Report Templates")

    def _build_report_template_generator_tab(self) -> None:
        """Build the standalone Report MD -> CMBX clone-and-patch workbench."""
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(1, weight=1)
        tk.Label(toolbar, text="Report MD", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self._make_entry(toolbar, self.report_generator_md_path_var).grid(row=0, column=1, sticky="ew", ipady=4)
        self._make_button(toolbar, "Browse MD", self.browse_report_generator_md, kind="neutral", width=12).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Preview", self.render_report_generator_md, kind="secondary", width=12).grid(row=0, column=3, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Generate CMBX", self.export_report_generator_cmbx, kind="primary", width=16).grid(row=0, column=4, sticky="e", padx=(8, 0))

        output_row = tk.Frame(tab, bg=self.colors["card"])
        output_row.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        output_row.columnconfigure(1, weight=1)
        tk.Label(output_row, text="Output CMBX", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self._make_entry(output_row, self.report_generator_output_path_var).grid(row=0, column=1, sticky="ew", ipady=4)
        self._make_button(output_row, "Save As", self.browse_report_generator_output, kind="neutral", width=10).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(output_row, "Open Output", self.open_report_generator_output, kind="neutral", width=13).grid(row=0, column=3, sticky="e", padx=(8, 0))
        self._make_button(output_row, "Report Spec", lambda: self._open_report_generator_kb("CM_REPORT_TEMPLATE_MD_TO_CMBX_SPEC.md"), kind="neutral", width=12).grid(row=0, column=4, sticky="e", padx=(8, 0))
        self._make_button(output_row, "Formula Reference", lambda: self._open_report_generator_kb("CM_REPORT_FORMULA_LANGUAGE_REFERENCE.md"), kind="neutral", width=16).grid(row=0, column=5, sticky="e", padx=(8, 0))

        tk.Label(tab, textvariable=self.report_generator_status_var, anchor="w", justify="left", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=2, column=0, sticky="ew", pady=(0, 8))

        body = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        body.grid(row=3, column=0, sticky="nsew")
        checks_frame = tk.Frame(body, bg=self.colors["card"])
        checks_frame.columnconfigure(0, weight=1)
        checks_frame.rowconfigure(1, weight=1)
        tk.Label(checks_frame, text="Preflight / Input Contract", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.report_generator_check_table = self._make_table(
            checks_frame,
            ("check", "result", "status"),
            {"check": "Check", "result": "Result", "status": "Status"},
        )
        self.report_generator_check_table.grid(row=1, column=0, sticky="nsew")
        check_widget = self._table_widget(self.report_generator_check_table)
        check_widget.column("check", width=210)
        check_widget.column("result", width=900)
        check_widget.column("status", width=115)

        preview_frame = tk.Frame(body, bg=self.colors["card"])
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(1, weight=1)
        preview_header = tk.Frame(preview_frame, bg=self.colors["card"])
        preview_header.grid(row=0, column=0, sticky="ew", pady=(8, 6))
        preview_header.columnconfigure(2, weight=1)
        tk.Label(preview_header, text="CMBX-equivalent Report Preview", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=(0, 12))
        tk.Label(preview_header, text="Sheet", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=1, sticky="w", padx=(0, 6))
        self.report_generator_sheet_combo = ttk.Combobox(preview_header, textvariable=self.report_generator_sheet_var, state="readonly", width=32, font=self._font(9))
        self.report_generator_sheet_combo.grid(row=0, column=2, sticky="w")
        self.report_generator_sheet_combo.bind("<<ComboboxSelected>>", self.preview_report_generator_sheet)
        self.report_generator_preview_table = self._make_excel_preview_table(preview_frame)
        self.report_generator_preview_table.grid(row=1, column=0, sticky="nsew")
        self._configure_excel_preview_tags(self._table_widget(self.report_generator_preview_table))
        body.add(checks_frame, minsize=160, height=220)
        body.add(preview_frame, minsize=260, height=480)
        self.notebook.add(tab, text="Report Template Generator")

    def _build_info_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(0, weight=1)
        self.info_text = tk.Text(tab, height=12, wrap="word", font=self._font(9), bg=self.colors["card_alt"], fg=self.colors["text"], relief="solid", bd=1, padx=10, pady=8)
        self.info_text.grid(row=0, column=0, sticky="nsew")
        self.info_text.configure(state="disabled")
        self.notebook.add(tab, text="CMBX Info")

    def _build_kb_index_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        toolbar.columnconfigure(1, weight=1)
        tk.Label(toolbar, text="KB Index", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.kb_index_path_var = tk.StringVar(value="")
        tk.Label(toolbar, textvariable=self.kb_index_path_var, anchor="w", font=self._font(8), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=1, sticky="ew")
        self._make_button(toolbar, "Open Source", self._open_selected_kb_index_source, kind="neutral", width=12).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Refresh", self._refresh_kb_index_tab, kind="neutral", width=10).grid(row=0, column=3, sticky="e", padx=(8, 0))

        filters = tk.Frame(tab, bg=self.colors["card"])
        filters.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        filters.columnconfigure(5, weight=1)
        tk.Label(filters, text="Category", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.kb_index_category_combo = ttk.Combobox(filters, textvariable=self.kb_index_category_var, values=("All",), state="readonly", width=22)
        self.kb_index_category_combo.grid(row=0, column=1, sticky="w", padx=(0, 12))
        self.kb_index_category_combo.bind("<<ComboboxSelected>>", self._select_kb_index_combo_category)
        tk.Label(filters, text="Search", font=self._font(8, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"]).grid(row=0, column=2, sticky="w", padx=(0, 6))
        search_entry = tk.Entry(filters, textvariable=self.kb_index_search_var, font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        search_entry.grid(row=0, column=3, sticky="ew", padx=(0, 12), ipadx=4, ipady=2)
        search_entry.bind("<KeyRelease>", self._populate_kb_index_table)
        self._make_button(filters, "Clear", self._clear_kb_index_filters, kind="neutral", width=8).grid(row=0, column=4, sticky="w")

        paned = tk.PanedWindow(tab, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["card"], bd=0)
        paned.grid(row=2, column=0, sticky="nsew")
        top = tk.Frame(paned, bg=self.colors["card"])
        bottom = tk.Frame(paned, bg=self.colors["card"])
        top.columnconfigure(0, weight=1)
        top.rowconfigure(0, weight=1)
        bottom.columnconfigure(0, weight=1)
        bottom.rowconfigure(0, weight=1)

        top.columnconfigure(0, weight=0)
        top.columnconfigure(1, weight=1)
        top.rowconfigure(0, weight=1)
        category_panel = tk.Frame(top, bg=self.colors["card"])
        category_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        category_panel.columnconfigure(0, weight=1)
        category_panel.rowconfigure(1, weight=1)
        tk.Label(category_panel, text="Categories", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 4))
        self.kb_index_category_tree = ttk.Treeview(category_panel, columns=("count",), show="tree headings", style="Explorer.Treeview", selectmode="browse")
        self.kb_index_category_tree.heading("#0", text="Knowledge Layer / Category")
        self.kb_index_category_tree.heading("count", text="KBs")
        self.kb_index_category_tree.column("#0", width=260, anchor="w")
        self.kb_index_category_tree.column("count", width=58, anchor="e")
        self.kb_index_category_tree.tag_configure("kb_scope_all", background="#E0F2FE", foreground=self.colors["text"])
        self.kb_index_category_tree.tag_configure("kb_scope_group", background="#F1F5F9", foreground=self.colors["text"])
        self.kb_index_category_tree.tag_configure("kb_scope_category", background=self.colors["card_alt"], foreground=self.colors["text"])
        self.kb_index_category_tree.grid(row=1, column=0, sticky="nsew")
        category_scroll = ttk.Scrollbar(category_panel, orient="vertical", command=self.kb_index_category_tree.yview)
        category_scroll.grid(row=1, column=1, sticky="ns")
        self.kb_index_category_tree.configure(yscrollcommand=category_scroll.set)
        self.kb_index_category_tree.bind("<<TreeviewSelect>>", self._select_kb_index_category)

        entry_panel = tk.Frame(top, bg=self.colors["card"])
        entry_panel.grid(row=0, column=1, sticky="nsew")
        entry_panel.columnconfigure(0, weight=1)
        entry_panel.rowconfigure(1, weight=1)
        tk.Label(entry_panel, text="Knowledge Bases", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 4))
        self.kb_index_table = self._make_table(
            entry_panel,
            ("kb_name", "category", "version", "update_date", "coverage", "status", "local_files"),
            {
                "kb_name": "KB",
                "category": "Category",
                "version": "Version",
                "update_date": "Update Date",
                "coverage": "Coverage",
                "status": "Status",
                "local_files": "Local File(s)",
            },
        )
        self.kb_index_table.grid(row=1, column=0, sticky="nsew")
        table = self._table_widget(self.kb_index_table)
        table.column("kb_name", width=220)
        table.column("category", width=150)
        table.column("version", width=90)
        table.column("update_date", width=130)
        table.column("coverage", width=260)
        table.column("status", width=190)
        table.column("local_files", width=620)
        table.tag_configure("kb_category", background="#E0F2FE", foreground=self.colors["text"])
        table.tag_configure("kb_entry", background=self.colors["card_alt"], foreground=self.colors["text"])
        table.bind("<<TreeviewSelect>>", self._preview_selected_kb_index_entry)
        table.bind("<Double-1>", self._open_selected_kb_index_source)

        preview_panel = tk.Frame(bottom, bg=self.colors["card"])
        preview_panel.grid(row=0, column=0, sticky="nsew", pady=(8, 0))
        preview_panel.columnconfigure(0, weight=1)
        preview_panel.rowconfigure(1, weight=1)
        tk.Label(preview_panel, text="Selected KB Preview", font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 4))
        self.kb_index_preview_text = tk.Text(preview_panel, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1, padx=8, pady=6)
        self.kb_index_preview_text.grid(row=1, column=0, sticky="nsew")
        preview_scroll = ttk.Scrollbar(preview_panel, orient="vertical", command=self.kb_index_preview_text.yview)
        preview_scroll.grid(row=1, column=1, sticky="ns")
        self.kb_index_preview_text.configure(yscrollcommand=preview_scroll.set, state="disabled")

        paned.add(top, minsize=180, height=300)
        paned.add(bottom, minsize=260, height=520)
        self.notebook.add(tab, text="KB Index")
        self._refresh_kb_index_tab()

    def _refresh_kb_index_tab(self) -> None:
        if not hasattr(self, "kb_index_table"):
            return
        table = self._table_widget(self.kb_index_table)
        table.delete(*table.get_children())
        try:
            path = resolve_kb_index_path()
            markdown = read_kb_index_text(path)
            entries = parse_kb_index_entries(markdown)
            entries = (*entries, *discover_kb_index_entries(path, existing_entries=entries))
        except Exception as exc:
            path = resolve_kb_index_path()
            markdown = f"Failed to load KB index:\n\n{exc}"
            entries = ()
        self.kb_index_entries = entries
        if hasattr(self, "kb_index_path_var"):
            self.kb_index_path_var.set(str(path))
        if hasattr(self, "kb_index_category_combo"):
            categories = ("All", *kb_index_category_options(entries))
            self.kb_index_category_combo.configure(values=categories)
            if self.kb_index_category_var.get() not in categories:
                self.kb_index_category_var.set("All")
                self.kb_index_scope = "All"
        self._populate_kb_index_table()
        self.status_var.set(f"Loaded KB index: {len(entries)} entr{'y' if len(entries) == 1 else 'ies'}")

    def _populate_kb_index_table(self, _event=None) -> None:
        if not hasattr(self, "kb_index_table"):
            return
        category_tree = self.kb_index_category_tree
        table = self._table_widget(self.kb_index_table)
        category_tree.delete(*category_tree.get_children())
        table.delete(*table.get_children())
        self.kb_index_category_context = {}
        self.kb_index_row_context = {}
        search_matched_entries = filter_kb_index_entries(
            self.kb_index_entries,
            category="All",
            search_text=self.kb_index_search_var.get(),
        )
        entries = kb_index_entries_for_scope(search_matched_entries, self.kb_index_scope)
        if self.kb_index_scope != "All" and not entries:
            self.kb_index_scope = "All"
            self.kb_index_category_var.set("All")
            entries = search_matched_entries
        self.kb_index_filtered_entries = entries
        by_category: dict[str, list[KbIndexEntry]] = {}
        for entry in search_matched_entries:
            by_category.setdefault(kb_index_entry_category(entry), []).append(entry)
        scope_to_iid: dict[str, str] = {}
        all_iid = "scope:all"
        self.kb_index_category_context[all_iid] = "All"
        scope_to_iid["All"] = all_iid
        category_tree.insert("", "end", iid=all_iid, text="All Knowledge Bases", values=(len(search_matched_entries),), tags=("kb_scope_all",), open=True)
        for group in kb_index_group_options(search_matched_entries):
            group_scope = f"group:{group}"
            group_iid = f"scope:group:{group}"
            group_entries = [entry for entry in search_matched_entries if kb_index_entry_group(entry) == group]
            self.kb_index_category_context[group_iid] = group_scope
            scope_to_iid[group_scope] = group_iid
            category_tree.insert(all_iid, "end", iid=group_iid, text=group, values=(len(group_entries),), tags=("kb_scope_group",), open=True)
            group_categories = [
                category
                for category in kb_index_category_options(search_matched_entries)
                if kb_index_entry_group(category) == group
            ]
            for category in group_categories:
                category_scope = f"category:{category}"
                category_iid = f"scope:category:{category}"
                self.kb_index_category_context[category_iid] = category_scope
                scope_to_iid[category_scope] = category_iid
                category_tree.insert(group_iid, "end", iid=category_iid, text=category, values=(len(by_category[category]),), tags=("kb_scope_category",))
        selected_iid = scope_to_iid.get(self.kb_index_scope, all_iid)
        self._kb_index_populating = True
        try:
            if selected_iid in category_tree.get_children("") or category_tree.exists(selected_iid):
                category_tree.selection_set(selected_iid)
                category_tree.focus(selected_iid)
                parent = category_tree.parent(selected_iid)
                if parent:
                    category_tree.item(parent, open=True)
        finally:
            self._kb_index_populating = False
        row_index = 0
        for entry in entries:
            row_index += 1
            iid = f"entry:{row_index}"
            self.kb_index_row_context[iid] = entry
            table.insert(
                "",
                "end",
                iid=iid,
                values=(
                    entry.kb_name,
                    kb_index_entry_category(entry),
                    entry.version,
                    entry.update_date,
                    entry.coverage,
                    entry.status,
                    "; ".join(entry.local_files),
                ),
                tags=("kb_entry",),
            )
        self._preview_selected_kb_index_category()
        if not table.get_children() and not category_tree.get_children():
            self._set_kb_index_preview("No KB index entries found.")

    def _clear_kb_index_filters(self) -> None:
        self.kb_index_category_var.set("All")
        self.kb_index_search_var.set("")
        self.kb_index_scope = "All"
        self._populate_kb_index_table()

    def _select_kb_index_combo_category(self, _event=None) -> None:
        category = self.kb_index_category_var.get() or "All"
        self.kb_index_scope = "All" if category == "All" else f"category:{category}"
        self._populate_kb_index_table()
        self._preview_selected_kb_index_category()

    def _select_kb_index_category(self, _event=None) -> None:
        if self._kb_index_populating:
            return
        if not hasattr(self, "kb_index_category_tree"):
            return
        category_tree = self.kb_index_category_tree
        selected = category_tree.selection()
        if not selected:
            return
        scope = self.kb_index_category_context.get(selected[0], "All")
        if scope == self.kb_index_scope:
            if hasattr(self, "kb_index_table"):
                table = self._table_widget(self.kb_index_table)
                table.selection_remove(table.selection())
            self._preview_selected_kb_index_category()
            return
        self.kb_index_scope = scope
        if scope == "All":
            self.kb_index_category_var.set("All")
        elif scope.startswith("category:"):
            self.kb_index_category_var.set(scope.partition(":")[2])
        else:
            self.kb_index_category_var.set("All")
        self._populate_kb_index_table()
        if hasattr(self, "kb_index_table"):
            table = self._table_widget(self.kb_index_table)
            table.selection_remove(table.selection())
        self._preview_selected_kb_index_category()

    def _preview_selected_kb_index_category(self) -> None:
        scope = self.kb_index_scope or "All"
        entries = self.kb_index_filtered_entries
        summary = [
            f"Scope: {scope}",
            f"KB Count: {len(entries)}",
            "",
            "Select one KB from the table above to preview its resolved Markdown here.",
            "",
            "Visible entries:",
        ]
        if entries:
            summary.extend(f"- {kb_index_entry_category(entry)} :: {entry.kb_name} ({entry.version}, {entry.status})" for entry in entries)
        else:
            summary.append("- none")
        self._set_kb_index_preview("\n".join(summary))

    def _preview_selected_kb_index_entry(self, _event=None) -> None:
        if not hasattr(self, "kb_index_table"):
            return
        table = self._table_widget(self.kb_index_table)
        selected = table.selection()
        if not selected:
            self._preview_selected_kb_index_category()
            return
        context = self.kb_index_row_context.get(selected[0])
        if isinstance(context, KbIndexEntry):
            preview = "\n\n".join(
                (
                    kb_index_entry_detail(context),
                    "---",
                    kb_index_entry_full_markdown(context, index_path=self.kb_index_path_var.get()),
                )
            )
            self._set_kb_index_preview(preview)
            return
        self._preview_selected_kb_index_category()

    def _set_kb_index_preview(self, text: str) -> None:
        if hasattr(self, "kb_index_preview_text"):
            self._set_text_widget(self.kb_index_preview_text, text)

    def _selected_kb_index_context(self) -> KbIndexEntry | str | None:
        if not hasattr(self, "kb_index_table"):
            return None
        table = self._table_widget(self.kb_index_table)
        selected = table.selection()
        if selected:
            context = self.kb_index_row_context.get(selected[0])
            if context is not None:
                return context
        if hasattr(self, "kb_index_category_tree"):
            category_tree = self.kb_index_category_tree
            category_selected = category_tree.selection()
            if category_selected:
                return self.kb_index_category_context.get(category_selected[0], self.kb_index_scope or "All")
        return self.kb_index_scope or "All"

    def _open_selected_kb_index_source(self) -> None:
        context = self._selected_kb_index_context()
        if not isinstance(context, KbIndexEntry):
            messagebox.showinfo(APP_NAME, "Select a specific KB entry to open its source file.")
            return
        paths = [resolved.path for resolved in resolve_kb_entry_files(context, index_path=self.kb_index_path_var.get()) if resolved.exists and resolved.path and resolved.path.is_file()]
        if not paths:
            messagebox.showinfo(APP_NAME, "No readable source file was resolved for this KB entry.")
            return
        self._open_preferred_path(paths)

    def _build_skills_tab(self) -> None:
        tab = tk.Frame(self.notebook, bg=self.colors["card"])
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)

        toolbar = tk.Frame(tab, bg=self.colors["card"])
        toolbar.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 8))
        toolbar.columnconfigure(1, weight=1)
        tk.Label(toolbar, text="Skills", font=self._font(10, "bold"), bg=self.colors["card"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w")
        tk.Label(
            toolbar,
            text="Local Codex skills and the KB/evidence files each skill references.",
            font=self._font(8),
            bg=self.colors["card"],
            fg=self.colors["text_secondary"],
        ).grid(row=0, column=1, sticky="w", padx=14)
        self._make_button(toolbar, "Refresh", self._refresh_skills_catalog_tab, kind="secondary", width=10).grid(row=0, column=2, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Open Skill", self._open_selected_skill_folder, kind="neutral", width=11).grid(row=0, column=3, sticky="e", padx=(8, 0))
        self._make_button(toolbar, "Open KB Ref", self._open_selected_skill_reference, kind="neutral", width=12).grid(row=0, column=4, sticky="e", padx=(8, 0))

        paned = ttk.Panedwindow(tab, orient=tk.VERTICAL)
        paned.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))

        columns = ("name", "source", "references", "kb_refs", "description", "path")
        headings = {
            "name": "Skill",
            "source": "Source",
            "references": "Reference Files",
            "kb_refs": "KB Refs",
            "description": "Description",
            "path": "Path",
        }
        table_frame = self._make_table(paned, columns, headings)
        self.skills_catalog_table = table_frame
        table = self._table_widget(table_frame)
        table.configure(selectmode="browse")
        table.column("name", width=260)
        table.column("source", width=95)
        table.column("references", width=105, anchor="e")
        table.column("kb_refs", width=85, anchor="e")
        table.column("description", width=760)
        table.column("path", width=520)
        table.bind("<<TreeviewSelect>>", self._preview_selected_skill_catalog_entry)
        table.bind("<Double-1>", lambda _event: self._show_selected_skill_rendered())
        paned.add(table_frame, weight=2)

        detail_notebook = ttk.Notebook(paned)
        self.skills_catalog_detail_texts: dict[str, tk.Text] = {}
        for title in ("Selected Skill", "KB References", "Catalog Overview"):
            detail_tab = tk.Frame(detail_notebook, bg=self.colors["card"])
            detail_tab.columnconfigure(0, weight=1)
            detail_tab.rowconfigure(0, weight=1)
            text_widget = tk.Text(detail_tab, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1, padx=8, pady=6)
            text_widget.grid(row=0, column=0, sticky="nsew")
            scroll = ttk.Scrollbar(detail_tab, orient="vertical", command=text_widget.yview)
            scroll.grid(row=0, column=1, sticky="ns")
            text_widget.configure(yscrollcommand=scroll.set, state="disabled")
            self.skills_catalog_detail_texts[title] = text_widget
            detail_notebook.add(detail_tab, text=title)
        paned.add(detail_notebook, weight=3)

        self.notebook.add(tab, text="Skills")
        self._refresh_skills_catalog_tab()

    def _refresh_skills_catalog_tab(self) -> None:
        if not hasattr(self, "skills_catalog_table"):
            return
        self.skills_catalog_entries = discover_skill_catalog_entries()
        self._populate_skills_catalog_table()
        if hasattr(self, "skills_catalog_detail_texts"):
            self._set_text_widget(self.skills_catalog_detail_texts["Catalog Overview"], skill_catalog_overview_markdown(self.skills_catalog_entries))
            self._set_text_widget(self.skills_catalog_detail_texts["Selected Skill"], "Select a skill to inspect its instructions, references, and KB links.")
            self._set_text_widget(self.skills_catalog_detail_texts["KB References"], "Select a skill to inspect its KB/evidence references.")
        self.status_var.set(f"Loaded {len(self.skills_catalog_entries)} skills.")

    def _populate_skills_catalog_table(self) -> None:
        table = self._table_widget(self.skills_catalog_table)
        table.delete(*table.get_children())
        self.skills_catalog_row_context = {}
        for entry in self.skills_catalog_entries:
            row_id = table.insert(
                "",
                "end",
                values=(
                    entry.name,
                    entry.source,
                    len(entry.reference_files),
                    len(entry.kb_references),
                    entry.description,
                    str(entry.path),
                ),
            )
            self.skills_catalog_row_context[row_id] = entry

    def _selected_skill_catalog_entry(self) -> SkillCatalogEntry | None:
        if not hasattr(self, "skills_catalog_table"):
            return None
        table = self._table_widget(self.skills_catalog_table)
        selected = table.selection()
        if not selected:
            return None
        return self.skills_catalog_row_context.get(selected[0])

    def _preview_selected_skill_catalog_entry(self, _event=None) -> None:
        entry = self._selected_skill_catalog_entry()
        if entry is None or not hasattr(self, "skills_catalog_detail_texts"):
            return
        self._set_text_widget(self.skills_catalog_detail_texts["Selected Skill"], skill_catalog_entry_markdown(entry))
        refs = ["# KB / Evidence References", ""]
        if entry.kb_references:
            refs.extend(f"- `{ref}`" for ref in entry.kb_references)
        else:
            refs.append("- No explicit KB reference paths found in skill text.")
        refs.extend(["", "## Skill Reference Files", ""])
        if entry.reference_files:
            refs.extend(f"- `{path}`" for path in entry.reference_files)
        else:
            refs.append("- None")
        self._set_text_widget(self.skills_catalog_detail_texts["KB References"], "\n".join(refs))

    def _show_selected_skill_rendered(self) -> None:
        entry = self._selected_skill_catalog_entry()
        if entry is None:
            return
        self._show_markdown_window(f"Skill: {entry.name}", skill_catalog_entry_markdown(entry))

    def _open_selected_skill_folder(self) -> None:
        entry = self._selected_skill_catalog_entry()
        if entry is None:
            messagebox.showinfo(APP_NAME, "Select a skill first.")
            return
        self._open_preferred_path([entry.path])

    def _open_selected_skill_reference(self) -> None:
        entry = self._selected_skill_catalog_entry()
        if entry is None:
            messagebox.showinfo(APP_NAME, "Select a skill first.")
            return
        paths = list(entry.reference_files) or [entry.skill_file]
        self._open_preferred_path(paths)

    def _make_table(self, parent: tk.Widget, columns: tuple[str, ...], headings: dict[str, str]) -> ttk.Treeview:
        frame = tk.Frame(parent, bg=self.colors["card"])
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        table = ttk.Treeview(frame, columns=columns, show="headings", style="Explorer.Treeview", selectmode="extended")
        widths = {
            "injection": 220,
            "device": 150,
            "device_source": 260,
            "report_template": 220,
            "cmbx": 220,
            "sequence": 220,
            "injections": 95,
            "channels": 95,
            "path": 420,
            "sheet": 280,
            "report": 230,
            "condition": 420,
            "source": 360,
            "url": 340,
            "name": 220,
            "raw": 220,
            "id": 180,
            "row": 54,
            "kind": 92,
            "time": 120,
            "day_time": 140,
            "ret_time": 130,
            "command": 360,
            "value": 420,
            "comment": 520,
            "message": 760,
        }
        for col in columns:
            table.heading(col, text=headings[col])
            table.column(col, width=widths.get(col, 140), anchor="w")
        table.grid(row=0, column=0, sticky="nsew")
        y = ttk.Scrollbar(frame, orient="vertical", command=table.yview)
        y.grid(row=0, column=1, sticky="ns")
        x = ttk.Scrollbar(frame, orient="horizontal", command=table.xview)
        x.grid(row=1, column=0, sticky="ew")
        table.configure(yscrollcommand=y.set, xscrollcommand=x.set)
        table.container = frame  # type: ignore[attr-defined]
        return frame  # type: ignore[return-value]

    def _make_excel_preview_table(self, parent: tk.Widget) -> tk.Frame:
        columns = ("row", *EXCEL_PREVIEW_LETTERS)
        headings = {column: column for column in columns}
        headings["row"] = ""
        frame = self._make_table(parent, columns, headings)
        table = self._table_widget(frame)
        table.column("row", width=54, anchor="e")
        for column in columns[1:]:
            table.column(column, width=150, anchor="w")
        return frame

    def _configure_excel_preview_tags(self, table: ttk.Treeview) -> None:
        table.tag_configure("preview_header", background="#E5E7EB", foreground=self.colors["text"])
        table.tag_configure("preview_formula", background="#FEF3C7")
        table.tag_configure("preview_value", background="#DCFCE7")
        table.tag_configure("preview_object", background="#E0F2FE")
        table.tag_configure("preview_title", background="#F8FAFC", foreground=self.colors["text"])

    def _table_widget(self, frame: tk.Frame) -> ttk.Treeview:
        return next(child for child in frame.winfo_children() if isinstance(child, ttk.Treeview))

    def _set_text_widget(self, widget: tk.Text, text: str) -> None:
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert(tk.END, text)
        widget.configure(state="disabled")

    def _load_database_config_defaults(self) -> dict[str, object]:
        defaults: dict[str, object] = {
            "server": "10.68.178.52",
            "database": "QCLab",
            "username": "QCUser",
            "password": "",
            "schema": "dbo",
            "table": "AUTO",
            "driver": "ODBC Driver 17 for SQL Server",
            "trust_server_certificate": True,
        }
        if not DEFAULT_DATABASE_CONFIG_FILE.exists():
            return defaults
        try:
            data = json.loads(DEFAULT_DATABASE_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            return defaults
        defaults.update({key: value for key, value in data.items() if key != "password"})
        return defaults

    def _load_ai_config_defaults(self) -> dict[str, object]:
        defaults: dict[str, object] = {
            "base_url": "https://api.openai.com/v1",
            "model": "gpt-5.5",
            "api_key": "",
        }
        if not DEFAULT_AI_CONFIG_FILE.exists():
            return defaults
        try:
            data = json.loads(DEFAULT_AI_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            return defaults
        if isinstance(data, dict):
            defaults.update({key: value for key, value in data.items() if key in defaults})
        return defaults

    def _save_ai_config(self) -> None:
        data = {
            "base_url": self.ai_base_url_var.get().strip() or "https://api.openai.com/v1",
            "model": self.ai_model_var.get().strip() or "gpt-5.5",
            "api_key": self.ai_api_key_var.get(),
        }
        DEFAULT_AI_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        DEFAULT_AI_CONFIG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def save_database_config(self) -> None:
        data = {
            "server": self.db_server_var.get().strip(),
            "database": self.db_database_var.get().strip(),
            "username": self.db_username_var.get().strip(),
            "schema": self.db_schema_var.get().strip(),
            "table": self.db_table_var.get().strip(),
            "driver": self.db_driver_var.get().strip(),
            "trust_server_certificate": bool(self.db_trust_cert_var.get()),
        }
        try:
            DEFAULT_DATABASE_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
            DEFAULT_DATABASE_CONFIG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))
            return
        self._append_db_upload_log(f"Saved database config to {DEFAULT_DATABASE_CONFIG_FILE} (password is not saved).")

    def _database_upload_config(self) -> DatabaseUploadConfig:
        return DatabaseUploadConfig(
            server=self.db_server_var.get().strip(),
            database=self.db_database_var.get().strip(),
            username=self.db_username_var.get().strip(),
            password=self.db_password_var.get(),
            schema=self.db_schema_var.get().strip() or "dbo",
            table=self.db_table_var.get().strip() or "AUTO",
            driver=self.db_driver_var.get().strip() or "ODBC Driver 17 for SQL Server",
            trust_server_certificate=bool(self.db_trust_cert_var.get()),
        )

    def _append_db_upload_log(self, message: str) -> None:
        if not hasattr(self, "db_upload_log"):
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.db_upload_log.configure(state="normal")
        self.db_upload_log.insert(tk.END, f"[{timestamp}] {message}\n")
        line_count = int(self.db_upload_log.index("end-1c").split(".", 1)[0])
        if line_count > 300:
            self.db_upload_log.delete("1.0", "50.0")
        self.db_upload_log.see(tk.END)
        self.db_upload_log.configure(state="disabled")

    def _thread_db_upload_log(self, message: str) -> None:
        self._call_ui(lambda message=message: self._append_db_upload_log(message))

    def add_database_upload_files(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Select FOQ DB workbooks",
            filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")],
            initialdir=Path(self.output_folder_var.get().strip()) if self.output_folder_var.get().strip() else DEFAULT_EXPORT_FOLDER,
        )
        if paths:
            self._add_database_upload_paths([Path(path) for path in paths])

    def add_database_upload_folder(self) -> None:
        folder = filedialog.askdirectory(
            title="Select folder containing FOQ DB workbooks",
            initialdir=Path(self.output_folder_var.get().strip()) if self.output_folder_var.get().strip() else DEFAULT_EXPORT_FOLDER,
        )
        if not folder:
            return
        paths = discover_foq_db_workbooks(folder)
        if not paths:
            messagebox.showinfo(APP_NAME, "No FOQ DB workbooks were found in that folder.")
            return
        self._add_database_upload_paths(paths)

    def _add_database_upload_paths(self, paths: list[Path]) -> None:
        known = {str(path).lower() for path in self.db_upload_files}
        for path in paths:
            if str(path).lower() not in known:
                self.db_upload_files.append(path)
                known.add(str(path).lower())
        self._refresh_database_upload_table()
        self._append_db_upload_log(f"Added {len(paths)} workbook path(s). Upload list has {len(self.db_upload_files)} item(s).")

    def clear_database_upload_files(self) -> None:
        self.db_upload_files.clear()
        self._refresh_database_upload_table()
        self._append_db_upload_log("Cleared upload workbook list.")

    def _refresh_database_upload_table(self) -> None:
        if not hasattr(self, "db_upload_table"):
            return
        table = self._table_widget(self.db_upload_table)
        table.delete(*table.get_children())
        for path in self.db_upload_files:
            sequence_name = re.sub(r"(?i)_foq_contract_db$", "", path.stem)
            field_count = ""
            try:
                from db_upload_service import read_foq_db_workbook

                field_count = str(len(read_foq_db_workbook(path).values))
            except Exception:
                field_count = "?"
            table.insert("", "end", values=(sequence_name, field_count, str(path)))

    def test_database_upload_connection(self) -> None:
        config = self._database_upload_config()
        self.status_var.set("Testing database connection...")
        self._set_buttons_state("disabled")
        self._append_db_upload_log(f"Testing SQL connection: {config.server} / {config.database}")

        def worker() -> None:
            try:
                message = test_database_connection(config)
                self._call_ui(lambda: self._database_connection_ok(message))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._database_upload_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _database_connection_ok(self, message: str) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(100.0)
        self.status_var.set("Database connection OK")
        self._append_db_upload_log(message)
        messagebox.showinfo(APP_NAME, message)

    def upload_database_files(self) -> None:
        selected_paths = self._selected_database_upload_paths()
        if not selected_paths:
            selected_paths = list(self.db_upload_files)
        if not selected_paths:
            messagebox.showinfo(APP_NAME, "Add one or more exported FOQ DB workbooks first.")
            return
        self._upload_database_workbooks(selected_paths)

    def _selected_database_upload_paths(self) -> list[Path]:
        if not hasattr(self, "db_upload_table"):
            return []
        table = self._table_widget(self.db_upload_table)
        selected: list[Path] = []
        for iid in table.selection():
            values = table.item(iid, "values")
            if len(values) >= 3:
                selected.append(Path(str(values[2])))
        return selected

    def _upload_database_workbooks(self, paths: list[Path]) -> None:
        config = self._database_upload_config()
        self.progress_var.set(0.0)
        self.status_var.set(f"Uploading {len(paths)} FOQ DB workbook(s) to {config.schema}.{config.table}...")
        self._append_db_upload_log(f"Starting upload: {len(paths)} workbook(s) -> {config.database}.{config.schema}.{config.table}")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                results = upload_foq_db_workbooks(paths, config, log=self._thread_db_upload_log)
                self._call_ui(lambda: self._database_upload_done(results))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._database_upload_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _database_upload_done(self, results) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(100.0 if results else 0.0)
        self.status_var.set(f"Uploaded {len(results)} FOQ DB workbook(s)")
        for result in results:
            self._append_db_upload_log(f"Uploaded {result.sequence_name} -> {self.db_schema_var.get().strip()}.{result.table_name}: {result.field_count} field(s) from {result.path.name}")
        messagebox.showinfo(APP_NAME, f"Uploaded {len(results)} FOQ DB workbook(s). See Database Upload Log for target table details.")

    def _database_upload_failed(self, exc: Exception) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(0.0)
        self.status_var.set("Database upload failed")
        self._append_db_upload_log(f"ERROR: {exc}")
        messagebox.showerror(APP_NAME, str(exc))

    def _enable_drag_row_selection(self, table: ttk.Treeview) -> None:
        state = {"anchor": ""}

        def on_press(event) -> None:
            row = table.identify_row(event.y)
            if row:
                state["anchor"] = row
                table.selection_set(row)

        def on_drag(event) -> None:
            anchor = state.get("anchor")
            row = table.identify_row(event.y)
            if not anchor or not row:
                return
            rows = list(table.get_children(""))
            if anchor not in rows or row not in rows:
                return
            start = rows.index(anchor)
            end = rows.index(row)
            low, high = sorted((start, end))
            table.selection_set(rows[low : high + 1])

        table.bind("<ButtonPress-1>", on_press, add="+")
        table.bind("<B1-Motion>", on_drag, add="+")

    def browse_cmbx(self) -> None:
        initialdir = discover_default_cmbx_source_folder()
        paths = filedialog.askopenfilenames(title="Select CMBX package(s)", initialdir=str(initialdir), filetypes=[("Chromeleon CMBX", "*.cmbx"), ("All files", "*.*")])
        if paths:
            self.cmbx_path_var.set("; ".join(paths))
            self._sync_path_display(self.cmbx_path_var, self.cmbx_display_var, "cmbx")

    def browse_cmbx_folder(self) -> None:
        path = filedialog.askdirectory(title="Select raw data folder", initialdir=str(discover_default_cmbx_source_folder()))
        if path:
            self.cmbx_path_var.set(path)
            self._sync_path_display(self.cmbx_path_var, self.cmbx_display_var, "cmbx")

    def browse_output_folder(self) -> None:
        path = filedialog.askdirectory(title="Select export folder")
        if path:
            self.output_folder_var.set(path)

    def open_output_folder(self) -> None:
        output = Path(self.output_folder_var.get().strip())
        try:
            output.mkdir(parents=True, exist_ok=True)
            os.startfile(str(output))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Cannot open output folder:\n{exc}")

    def _default_method_editor_output_path(self, md_path: Path | None = None) -> Path:
        source = md_path or Path(self.method_editor_md_path_var.get().strip() or "method_script.md")
        stem = self._safe_temp_stem(source.stem)
        return Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER) / "method_script_generator" / f"{stem}.cmbx"

    def _method_editor_kb_doc_path(self, filename: str) -> Path:
        filename = Path(filename).name
        workspace_path = DEFAULT_APP_WORKSPACE / "KB" / "FOQ Template" / filename
        if workspace_path.exists():
            return workspace_path
        for _, doc_filename, repo_path in METHOD_GENERATOR_KB_DOCS:
            if doc_filename == filename:
                return repo_path
        return workspace_path

    def render_method_editor_kb_doc(self, filename: str) -> None:
        if not hasattr(self, "method_editor_kb_text"):
            return
        path = self._method_editor_kb_doc_path(filename)
        self.method_editor_kb_doc_var.set(path.name)
        if not path.exists():
            self._set_text_widget(self.method_editor_kb_text, f"KB file not found:\n{path}")
            return
        try:
            text = path.read_text(encoding="utf-8-sig", errors="replace")
        except Exception as exc:
            self._set_text_widget(self.method_editor_kb_text, f"Could not read KB file:\n{path}\n\n{exc}")
            return
        title = f"# {path.name}\n\nSource: {path}\n\n"
        self._set_text_widget(self.method_editor_kb_text, title + text)
        self.status_var.set(f"Rendered method generator KB: {path.name}")

    def open_method_editor_kb_doc(self) -> None:
        path = self._method_editor_kb_doc_path(self.method_editor_kb_doc_var.get())
        if not path.exists():
            messagebox.showinfo(APP_NAME, f"KB file not found:\n{path}")
            return
        try:
            os.startfile(str(path))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Cannot open KB file:\n{exc}")

    def browse_method_editor_md(self) -> None:
        initial = Path(self.method_editor_md_path_var.get().strip() or Path.home())
        initialdir = initial.parent if initial.exists() else Path.home()
        path = filedialog.askopenfilename(
            title="Select method script Markdown",
            initialdir=str(initialdir),
            filetypes=[("Markdown", "*.md"), ("Text", "*.txt"), ("All files", "*.*")],
        )
        if not path:
            return
        self.method_editor_md_path_var.set(path)
        self.method_editor_xlsx_path_var.set(str(self._default_method_editor_output_path(Path(path))))
        self.render_method_editor_md()

    def browse_method_editor_output(self) -> None:
        current = Path(self.method_editor_xlsx_path_var.get().strip() or str(self._default_method_editor_output_path()))
        path = filedialog.asksaveasfilename(
            title="Save CM method workbook",
            initialdir=str(current.parent),
            initialfile=current.name,
            defaultextension=".cmbx",
            filetypes=[("Chromeleon CMBX", "*.cmbx"), ("All files", "*.*")],
        )
        if path:
            self.method_editor_xlsx_path_var.set(path)

    def browse_method_editor_example(self) -> None:
        current = Path(self.method_editor_example_path_var.get().strip() or Path.home())
        initialdir = current.parent if current.exists() else Path.home()
        path = filedialog.askopenfilename(
            title="Select example CM XLSX",
            initialdir=str(initialdir),
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.method_editor_example_path_var.set(path)

    def render_method_editor_md(self) -> None:
        md_path = Path(self.method_editor_md_path_var.get().strip())
        if not md_path.exists() or not md_path.is_file():
            messagebox.showinfo(APP_NAME, "Select a valid Markdown method script first.")
            return
        try:
            rows = parse_md_to_rows(md_path)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not render Markdown method script:\n{exc}")
            return
        self.method_editor_rows = rows
        if not self.method_editor_xlsx_path_var.get().strip():
            self.method_editor_xlsx_path_var.set(str(self._default_method_editor_output_path(md_path)))
        self._populate_method_editor_table(rows)
        warnings = self._method_editor_row_warnings(rows)
        warning_text = f" Warnings: {len(warnings)}. " + "; ".join(warnings[:3]) if warnings else ""
        self.method_editor_status_var.set(
            f"Previewed {len(rows)} row(s). Generate CMBX uses template carrier: {DEFAULT_METHOD_CMBX_TEMPLATE}.{warning_text}"
        )
        self.status_var.set(f"Rendered method MD: {md_path.name}")

    def _method_editor_row_warnings(self, rows: list[dict[str, str]]) -> list[str]:
        return [issue.display() for issue in lint_method_rows(rows)]

    def _populate_method_editor_table(self, rows: list[dict[str, str]]) -> None:
        if not hasattr(self, "method_editor_table"):
            return
        table = self._table_widget(self.method_editor_table)
        table.delete(*table.get_children())
        invalid_rows = lint_error_rows(rows)
        for index, row in enumerate(rows):
            kind = row.get("Kind", "")
            row_id = row.get("#", str(index))
            if row_id in invalid_rows:
                tag = "cm_invalid"
            elif kind == "Stage":
                tag = "cm_stage"
            elif kind == "Branch":
                tag = "cm_branch"
            elif kind == "Comment":
                tag = "cm_comment"
            elif kind == "End":
                tag = "cm_end"
            else:
                tag = "cm_command"
            table.insert(
                "",
                "end",
                iid=f"method_editor:{index}",
                values=(
                    row_id,
                    kind,
                    row.get("Time", ""),
                    row.get("Command", ""),
                    row.get("Value", ""),
                    row.get("Comment", ""),
                ),
                tags=(tag,),
            )

    def export_method_editor_xlsx(self) -> None:
        if not self.method_editor_rows:
            self.render_method_editor_md()
            if not self.method_editor_rows:
                return
        md_path = Path(self.method_editor_md_path_var.get().strip())
        if not md_path.exists() or not md_path.is_file():
            messagebox.showinfo(APP_NAME, "Select a valid Markdown method script first.")
            return
        if not DEFAULT_METHOD_CMBX_TEMPLATE.exists():
            messagebox.showerror(APP_NAME, f"Default standalone method CMBX template was not found:\n{DEFAULT_METHOD_CMBX_TEMPLATE}")
            return
        issues = lint_method_rows(self.method_editor_rows)
        blocking_issues = [issue.display() for issue in issues if issue.severity == "error"]
        if blocking_issues:
            messagebox.showerror(
                APP_NAME,
                "This MD failed method-script preflight.\n\n"
                + "\n".join(blocking_issues[:8])
                + "\n\nUpdate the MD according to CM Compiler Rules.MD / CM_METHOD_SCRIPT_MD_FORMAT_SPEC.md before generating CMBX.",
            )
            self.method_editor_status_var.set(f"Generation blocked: {len(blocking_issues)} preflight error(s). See Format Spec.")
            return
        output = Path(self.method_editor_xlsx_path_var.get().strip() or str(self._default_method_editor_output_path()))
        try:
            method_name = md_path.stem.replace("_", " ").strip() or "Generated method"
            stats = compile_method_md_to_cmbx(DEFAULT_METHOD_CMBX_TEMPLATE, md_path, output, method_name=method_name)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not generate method CMBX:\n{exc}")
            return
        self.method_editor_xlsx_path_var.set(str(output))
        self.method_editor_status_var.set(f"Generated CMBX: {output} | {stats}")
        self.status_var.set(f"Generated method CMBX: {output.name}")
        messagebox.showinfo(APP_NAME, f"Generated standalone instrument-method CMBX:\n\n{output}")

    def open_method_editor_xlsx(self) -> None:
        path = Path(self.method_editor_xlsx_path_var.get().strip())
        if not path.exists():
            messagebox.showinfo(APP_NAME, "Generate the CMBX first.")
            return
        try:
            os.startfile(str(path))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Cannot open CM XLSX:\n{exc}")

    def _default_report_generator_output_path(self, md_path: Path | None = None) -> Path:
        source = md_path or Path(self.report_generator_md_path_var.get().strip() or "report_template.md")
        stem = self._safe_temp_stem(source.stem)
        return Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER) / "report_template_generator" / f"{stem}.cmbx"

    def _report_generator_search_roots(self) -> tuple[Path, ...]:
        roots = [DEFAULT_REPORT_TEMPLATE_CARRIER_FOLDER, DEFAULT_APP_WORKSPACE / "KB" / "FOQ Template"]
        md_path = Path(self.report_generator_md_path_var.get().strip())
        if md_path.exists():
            roots.insert(0, md_path.parent)
        return tuple(root for root in roots if root.exists())

    def _report_generator_kb_path(self, filename: str) -> Path:
        filename = Path(filename).name
        candidates = (
            DEFAULT_REPORT_TEMPLATE_CARRIER_FOLDER / filename,
            DEFAULT_APP_WORKSPACE / "KB" / "FOQ Template" / filename,
            Path(__file__).resolve().parent / "docs" / filename,
        )
        return next((candidate for candidate in candidates if candidate.exists()), candidates[-1])

    def _open_report_generator_kb(self, filename: str) -> None:
        path = self._report_generator_kb_path(filename)
        if not path.exists():
            messagebox.showinfo(APP_NAME, f"KB file not found:\n{path}")
            return
        try:
            os.startfile(str(path))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Cannot open KB file:\n{exc}")

    def browse_report_generator_md(self) -> None:
        current = Path(self.report_generator_md_path_var.get().strip() or DEFAULT_REPORT_TEMPLATE_CARRIER_FOLDER)
        initialdir = current.parent if current.exists() and current.is_file() else current
        path = filedialog.askopenfilename(
            title="Select report-template Markdown",
            initialdir=str(initialdir if initialdir.exists() else Path.home()),
            filetypes=[("Markdown", "*.md *.MD"), ("All files", "*.*")],
        )
        if not path:
            return
        self.report_generator_md_path_var.set(path)
        self.report_generator_output_path_var.set(str(self._default_report_generator_output_path(Path(path))))
        self.render_report_generator_md()

    def browse_report_generator_output(self) -> None:
        current = Path(self.report_generator_output_path_var.get().strip() or str(self._default_report_generator_output_path()))
        path = filedialog.asksaveasfilename(
            title="Save standalone report-template CMBX",
            initialdir=str(current.parent),
            initialfile=current.name,
            defaultextension=".cmbx",
            filetypes=[("Chromeleon CMBX", "*.cmbx"), ("All files", "*.*")],
        )
        if path:
            self.report_generator_output_path_var.set(path)

    def render_report_generator_md(self) -> None:
        md_path = Path(self.report_generator_md_path_var.get().strip())
        if not md_path.exists() or not md_path.is_file():
            messagebox.showinfo(APP_NAME, "Select a valid report-template Markdown file first.")
            return
        try:
            spec = parse_report_template_md(md_path)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not parse report-template Markdown:\n{exc}")
            return
        if not self.report_generator_output_path_var.get().strip():
            self.report_generator_output_path_var.set(str(self._default_report_generator_output_path(md_path)))
        preview_dir = Path(tempfile.gettempdir()) / "CmbxDataExplorer" / "report_template_generator_preview"
        preview_dir.mkdir(parents=True, exist_ok=True)
        preview_path = preview_dir / f"{self._safe_temp_stem(spec.template_name or md_path.stem)}_preview.cmbx"
        staged = compile_report_template_md_to_cmbx(spec, preview_path, self._report_generator_search_roots())
        self.report_generator_result = staged
        self._populate_report_generator_checks(staged)
        if not staged.ready:
            self.report_generator_status_var.set(f"Preview packaging blocked: {len(staged.errors)} error(s).")
            self.status_var.set(f"Report MD preview blocked: {md_path.name}")
            return
        try:
            package = load_cmbx_package(preview_path)
            report = next(element for element in package.methods_and_reports if element.kind == "report_template")
            _embedded, xml_text = decode_report_template_xml(package, report)
            sheets = parse_report_sheets(xml_text, report.name)
        except Exception as exc:
            self.report_generator_status_var.set(f"Preview decode failed: {exc}")
            self.status_var.set("Report CMBX preview decode failed")
            return
        self.report_generator_preview_cmbx = preview_path
        self.report_generator_preview_package = package
        self.report_generator_preview_report = report
        sheet_names = [sheet.sheet_name for sheet in sheets]
        self.report_generator_sheet_combo.configure(values=sheet_names)
        selected = self.report_generator_sheet_var.get().strip()
        if selected not in sheet_names:
            selected = self._first_preview_sheet_name(sheets)
            self.report_generator_sheet_var.set(selected)
        self.preview_report_generator_sheet()
        self.report_generator_status_var.set(
            f"Previewed staged CMBX '{report.name}': {len(sheet_names)} sheet(s), {len(staged.applied_patches)} direct CM formula object(s), "
            f"{len(staged.applied_workbook_patches)} FormulaOne cell write(s), {len(staged.applied_dynamic_tables)} native dynamic table(s)."
        )
        self.status_var.set(f"Rendered report MD and staged CMBX preview: {md_path.name}")

    def _populate_report_generator_checks(self, result: ReportTemplateCompileResult) -> None:
        table = self._table_widget(self.report_generator_check_table)
        table.delete(*table.get_children())
        table.tag_configure("check_ok", background="#DCFCE7", foreground=self.colors["text"])
        table.tag_configure("check_warning", background="#FEF3C7", foreground=self.colors["text"])
        table.tag_configure("check_error", background="#FEE2E2", foreground="#991B1B")
        rows: list[tuple[str, str, str]] = []
        spec = result.spec
        rows.append(("Template identity", spec.template_name or "Missing", "OK" if spec.template_name else "ERROR"))
        carrier_label = "Internal neutral carrier" if spec.generation_mode == "create_from_blank" else "Reference carrier"
        rows.append((carrier_label, str(result.source_cmbx or spec.reference_cmbx or "Missing"), "OK" if result.source_cmbx else "ERROR"))
        rows.append(("Carrier serialization", result.source_report_name or "Not decoded", "OK" if result.source_report_name else "ERROR"))
        workbook_ok = spec.workbook_policy == "create_static" or spec.workbook_policy == "preserve" or (
            spec.workbook_policy == "existing_cells_only" and not result.errors
        )
        workbook_detail = (
            f"{spec.workbook_policy}; {len(result.sheets)} new logical sheet(s), "
            f"{len(result.applied_workbook_patches)} FormulaOne cell write(s)."
            if spec.generation_mode == "create_from_blank"
            else f"{spec.workbook_policy or 'Missing'}; {len(result.applied_workbook_patches)} existing-cell FormulaOne patch(es) applied; all layout/table structure preserved."
        )
        rows.append((
            "Workbook policy",
            workbook_detail,
            "OK" if workbook_ok else "ERROR",
        ))
        rows.append(("Sheets", f"{len(result.sheets)} detected: {', '.join(result.sheets[:6])}" + (" ..." if len(result.sheets) > 6 else ""), "OK" if result.sheets else "ERROR"))
        rows.append(("Direct CM formula patches", f"{len(result.applied_patches)} applied / {len(spec.patches)} declared", "OK" if len(result.applied_patches) == len(spec.patches) else "ERROR"))
        if result.preserved_report_tables:
            table_labels = ", ".join(
                f"{sheet_name} / {excel_range} ({table_type})"
                for sheet_name, excel_range, table_type in result.preserved_report_tables
            )
            rows.append((
                "Dynamic report table(s)",
                f"{len(result.preserved_report_tables)} preserved unchanged: {table_labels}",
                "REVIEW",
            ))
        if result.applied_dynamic_tables:
            table_labels = ", ".join(
                f"{item.sheet_name} / {item.excel_range} ({item.table_type})"
                for item in result.applied_dynamic_tables
            )
            rows.append((
                "Created dynamic report table(s)",
                f"{len(result.applied_dynamic_tables)} created from MD: {table_labels}",
                "OK",
            ))
        for patch in result.applied_patches[:12]:
            rows.append((f"Patch {patch.sheet_name} / {patch.excel_range}", f"{patch.object_type}: {patch.formula}", "OK"))
        for patch in result.applied_workbook_patches[:12]:
            rows.append((f"Workbook {patch.sheet_name} / {patch.excel_range}", f"{patch.value_type}: {patch.value}", "OK"))
        for warning in result.warnings:
            rows.append(("Open verification", warning, "REVIEW"))
        for error in result.errors:
            rows.append(("Preflight error", error, "ERROR"))
        for index, (check, detail, status) in enumerate(rows):
            tag = "check_ok" if status == "OK" else "check_warning" if status == "REVIEW" else "check_error"
            table.insert("", "end", iid=f"report-generator-check:{index}", values=(check, detail, status), tags=(tag,))

    def preview_report_generator_sheet(self, _event=None) -> None:
        package = self.report_generator_preview_package
        report = self.report_generator_preview_report
        if package is None or report is None:
            return
        selected_sheet = self.report_generator_sheet_var.get().strip()
        export_error = ""
        try:
            temp_root = Path(tempfile.gettempdir()) / "CmbxDataExplorer" / "report_template_generator_xls"
            temp_root.mkdir(parents=True, exist_ok=True)
            paths = export_elements(package, [report], temp_root)
            workbook = next((path for path in paths if path.suffix.lower() in {".xls", ".xlsx"}), None)
            rows = self._workbook_preview_rows(workbook, selected_sheet) if workbook else []
            dynamic_tables = (
                [item for item in self.report_generator_result.spec.dynamic_tables if item.sheet_name == selected_sheet]
                if self.report_generator_result is not None
                else []
            )
            if dynamic_tables:
                rows = self._merge_dynamic_table_preview_rows(rows, dynamic_tables)
        except Exception as exc:
            rows = []
            export_error = str(exc)
        if not rows:
            rows = self._report_generator_spec_preview_rows(selected_sheet)
        if not rows:
            try:
                _embedded, xml_text = decode_report_template_xml(package, report)
                objects = parse_report_sheet_objects(xml_text, report.name, selected_sheet)
                rows = self._build_report_preview_rows(objects, {}, selected_sheet)
            except Exception as exc:
                export_error = export_error or str(exc)
        self._set_excel_preview_rows(self.report_generator_preview_table, rows)
        if export_error and rows:
            self.report_generator_status_var.set(
                "Preview uses the embedded/MD fallback because native XLS export was unavailable. "
                f"CMBX generation remains valid; export detail: {export_error}"
            )
        elif export_error:
            self.report_generator_status_var.set(f"Preview could not be built: {export_error}")

    def _report_generator_spec_preview_rows(self, selected_sheet: str):
        """Build a deterministic preview without invoking CM's .NET XLS exporter."""
        result = self.report_generator_result
        if result is None:
            return []
        row_cells: dict[int, dict[int, str]] = {1: {1: selected_sheet or "Report Sheet"}}
        row_tags: dict[int, str] = {1: "preview_title"}
        for patch in result.spec.workbook_patches:
            if patch.sheet_name != selected_sheet:
                continue
            parsed = self._cell_ref_to_row_col(self._top_left_cell(patch.excel_range))
            if not parsed:
                continue
            row, column = parsed
            text = str(patch.value)
            if patch.value_type == "formula" and not text.startswith("="):
                text = f"={text}"
            row_cells.setdefault(row + 1, {})[column] = text
            row_tags[row + 1] = "preview_formula" if patch.value_type == "formula" else "preview_value"
        for patch in result.spec.patches:
            if patch.sheet_name != selected_sheet:
                continue
            parsed = self._cell_ref_to_row_col(self._top_left_cell(patch.excel_range))
            if not parsed:
                continue
            row, column = parsed
            row_cells.setdefault(row + 1, {})[column] = self._short_preview_text(patch.formula)
            if row_tags.get(row + 1) != "preview_value":
                row_tags[row + 1] = "preview_formula"
        rows = [
            (row_number, row_cells[row_number], row_tags.get(row_number, ""))
            for row_number in sorted(row_cells)
            if row_number <= 160
        ]
        dynamic_tables = [item for item in result.spec.dynamic_tables if item.sheet_name == selected_sheet]
        return self._merge_dynamic_table_preview_rows(rows, dynamic_tables) if dynamic_tables else rows

    def _merge_dynamic_table_preview_rows(self, rows, dynamic_tables):
        row_cells = {row_number: dict(cells) for row_number, cells, _tag in rows}
        row_tags = {row_number: tag for row_number, _cells, tag in rows}
        for table in dynamic_tables:
            top_left = self._cell_ref_to_row_col(self._top_left_cell(table.excel_range))
            if not top_left:
                continue
            excel_row, first_column = top_left
            preview_row = excel_row + 1  # Row 1 in the widget is the sheet-name banner.
            if table.table_type == "audittrail":
                headers = []
                if table.show_day_time:
                    headers.append("Day Time")
                headers.append("Retention Time")
                if table.show_device:
                    headers.append("Device")
                headers.append("Command / Message")
                marker = "[runtime audit event]"
            else:
                headers = [column.header for column in table.columns]
                marker = "[runtime injection row]"
            row_cells.setdefault(preview_row, {}).update(
                {first_column + offset: header for offset, header in enumerate(headers)}
            )
            row_tags[preview_row] = "preview_formula"
            for offset in range(1, min(table.body_rows, 5) + 1):
                current = preview_row + offset
                row_cells.setdefault(current, {})[first_column] = marker
                row_tags[current] = "preview_object"
        return [
            (row_number, row_cells[row_number], row_tags.get(row_number, ""))
            for row_number in sorted(row_cells)
            if row_number <= 160
        ]

    def export_report_generator_cmbx(self) -> None:
        md_path = Path(self.report_generator_md_path_var.get().strip())
        if not md_path.exists() or not md_path.is_file():
            messagebox.showinfo(APP_NAME, "Select a valid report-template Markdown file first.")
            return
        try:
            spec = parse_report_template_md(md_path)
            output = Path(self.report_generator_output_path_var.get().strip() or str(self._default_report_generator_output_path(md_path)))
            result = compile_report_template_md_to_cmbx(spec, output, self._report_generator_search_roots())
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not generate report-template CMBX:\n{exc}")
            return
        self.report_generator_result = result
        self._populate_report_generator_checks(result)
        if not result.ready:
            self.report_generator_status_var.set(f"Generation blocked: {len(result.errors)} preflight error(s).")
            messagebox.showerror(APP_NAME, "Report-template generation is blocked:\n\n" + "\n".join(result.errors[:10]))
            return
        self.report_generator_output_path_var.set(str(output))
        self.report_generator_status_var.set(f"Generated standalone report-template CMBX: {output}. Open/import it in Chromeleon for final validation.")
        self.status_var.set(f"Generated report-template CMBX: {output.name}")
        messagebox.showinfo(
            APP_NAME,
            f"Generated standalone report-template CMBX:\n\n{output}\n\n"
            f"Created {len(result.sheets)} report sheet(s), {len(result.applied_patches)} CM formula object(s), "
            f"{len(result.applied_workbook_patches)} FormulaOne cell write(s), and "
            f"{len(result.applied_dynamic_tables)} native dynamic report table(s).",
        )

    def open_report_generator_output(self) -> None:
        path = Path(self.report_generator_output_path_var.get().strip())
        if not path.exists():
            messagebox.showinfo(APP_NAME, "Generate the report-template CMBX first.")
            return
        try:
            os.startfile(str(path))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Cannot open report CMBX:\n{exc}")

    def _generated_project_output_root(self) -> Path:
        return Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER) / "generated_projects"

    def _foq_kb_to_run_rows(self) -> tuple[FoqKbRunRow, ...]:
        # Current high-confidence alignment is the VX-C10-A FOQ TD path. Device
        # variants reuse many rows, but every concrete module still needs its
        # own TD/method/report evidence before generation.
        return tcc_vh_foq_kb_rows()

    def _populate_foq_kb_to_run_table(self) -> None:
        if not hasattr(self, "foq_kb_to_run_table"):
            return
        table = self._table_widget(self.foq_kb_to_run_table)
        table.delete(*table.get_children())
        for row in self._foq_kb_to_run_rows():
            table.insert(
                "",
                "end",
                iid=str(row.order),
                values=(
                    row.order,
                    row.td_item,
                    row.injection,
                    row.instrument_method,
                    row.processing_method,
                    ", ".join(row.report_sheets),
                ),
            )
        children = table.get_children()
        if children:
            table.selection_set(children[0])
            table.focus(children[0])
            self._preview_selected_foq_kb_to_run_row()

    def _selected_foq_kb_to_run_row(self) -> FoqKbRunRow | None:
        if not hasattr(self, "foq_kb_to_run_table"):
            return None
        table = self._table_widget(self.foq_kb_to_run_table)
        selected = table.selection()
        if not selected:
            return None
        try:
            order = int(selected[0])
        except ValueError:
            return None
        return next((row for row in self._foq_kb_to_run_rows() if row.order == order), None)

    def _preview_selected_foq_kb_to_run_row(self, _event=None) -> None:
        row = self._selected_foq_kb_to_run_row()
        if row is None or not hasattr(self, "foq_kb_to_run_detail"):
            return
        cn = foq_kb_row_chinese(row)
        text = "\n".join(
            [
                f"FOQ TD: {foq_td_title()}",
                f"FOQ TD Item: {row.td_item}",
                f"中文测试项: {cn['item']}",
                f"Injection: {row.injection}",
                f"Instrument Method: {row.instrument_method}",
                f"Processing Method: {row.processing_method}",
                f"Report Sheet(s): {', '.join(row.report_sheets)}",
                "",
                "TD Intent",
                row.td_intent,
                "",
                "TD 测试逻辑",
                str(cn["intent"]),
                "",
                "Method Script Contract",
                row.method_contract,
                "",
                "方法脚本实现关系",
                str(cn["method"]),
                "",
                "Report Calculation Contract",
                row.report_contract,
                "",
                "报告计算关系",
                str(cn["report"]),
                "",
                "Design Questions / Human Decisions",
                "\n".join(f"- {item}" for item in row.design_questions) or "- none",
                "",
                "需人工确认",
                "\n".join(f"- {item}" for item in cn["questions"]) or "- 暂无",
                "",
                "Next Useful Action",
                "Use this row as the unit of work: confirm the TD logic, then inspect the decoded method script and report formulas. The generation target is not a copied method fragment; it is a script/report pair that emits and consumes the same evidence.",
            ]
        )
        self.foq_kb_to_run_detail.configure(state="normal")
        self.foq_kb_to_run_detail.delete("1.0", tk.END)
        self.foq_kb_to_run_detail.insert(tk.END, text)
        self.foq_kb_to_run_detail.configure(state="disabled")

    def export_foq_kb_to_run_alignment(self) -> None:
        try:
            rows = self._foq_kb_to_run_rows()
            output = Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER) / "foq_kb_to_run" / f"{foq_td_safe_stem()}_FOQ_KB_to_Run.xlsx"
            path = write_foq_kb_to_run_alignment_workbook(rows, output)
            self.status_var.set(f"Exported FOQ KB alignment: {path}")
            messagebox.showinfo(APP_NAME, f"Exported FOQ KB alignment:\n\n{path}")
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def _refresh_foq_alignment_catalog(self, _event=None) -> None:
        self.foq_alignment_records = build_foq_alignment_records()
        if hasattr(self, "foq_alignment_family_combo"):
            families = family_options(self.foq_alignment_records)
            values = ("All", *families)
            self.foq_alignment_family_combo.configure(values=values)
            if self.foq_alignment_family_var.get() not in values:
                self.foq_alignment_family_var.set(families[0] if families else "All")
        self._refresh_foq_alignment_filters()
        if hasattr(self, "test_plan_family_combo"):
            self._refresh_test_plan_options()

    def _test_plan_records(self) -> tuple[FoqAlignmentRecord, ...]:
        if not self.foq_alignment_records:
            self.foq_alignment_records = build_foq_alignment_records()
        return self.foq_alignment_records

    def _refresh_test_plan_options(self, _event=None) -> None:
        records = self._test_plan_records()
        family = self.test_plan_family_var.get().strip() or "TCC"
        family_records = filter_alignment_records(records, family=family)
        if hasattr(self, "test_plan_family_combo"):
            families = family_options(records)
            self.test_plan_family_combo.configure(values=families or ("TCC", "VDAD"))
        if hasattr(self, "test_plan_test_combo"):
            self.test_plan_test_combo.configure(values=("", *test_intent_options(family_records)))
        if hasattr(self, "test_plan_device_combo"):
            devices = device_options(family_records)
            self.test_plan_device_combo.configure(values=devices)
            if devices and self.test_plan_device_var.get() not in devices:
                self.test_plan_device_var.set(devices[0])
        self._analyze_test_plan_intent()

    def _resolve_test_plan_intent_token(self, requested: str, family: str) -> str:
        requested_text = (requested or "").strip()
        records = filter_alignment_records(self._test_plan_records(), family=family or "TCC")
        options = test_intent_options(records)
        if not requested_text:
            return ""
        if requested_text in options:
            return requested_text
        normalized = re.sub(r"[^a-z0-9]+", "_", requested_text.lower()).strip("_")
        aliases = {
            "heatup": ("heatup", "cooldown"),
            "heat_up": ("heatup", "cooldown"),
            "cooldown": ("heatup", "cooldown"),
            "cool_down": ("heatup", "cooldown"),
            "heatup_cooldown": ("heatup", "cooldown"),
            "accuracy": ("temperature", "accuracy"),
            "temperature_accuracy": ("temperature", "accuracy"),
            "calibration": ("calibration",),
            "temperature_calibration": ("calibration",),
            "stability": ("stability",),
            "temperature_stability": ("stability",),
            "precision": ("precision",),
            "temperature_precision": ("precision",),
        }
        terms = aliases.get(normalized, (normalized,))
        for option in options:
            option_norm = re.sub(r"[^a-z0-9]+", "_", option.lower()).strip("_")
            if normalized == option_norm:
                return option
        for option in options:
            option_norm = re.sub(r"[^a-z0-9]+", "_", option.lower()).strip("_")
            if all(term in option_norm for term in terms):
                return option
        for option in options:
            option_norm = re.sub(r"[^a-z0-9]+", "_", option.lower()).strip("_")
            if normalized in option_norm or option_norm in normalized:
                return option
        return requested_text

    def _test_plan_numbers_from_text(self, text: str) -> list[float]:
        numbers: list[float] = []
        for token in re.findall(r"-?\d+(?:\.\d+)?", text or ""):
            try:
                numbers.append(float(token))
            except ValueError:
                continue
        return numbers

    def _format_test_plan_number(self, value: float) -> str:
        return f"{value:g}"

    def _test_plan_parameter_from_text(self, text: str, test_intent: str = "") -> str:
        normalized = text or ""
        transition = self._test_plan_temperature_transition_from_text(normalized)
        if transition is not None:
            return "->".join(self._format_test_plan_number(value) for value in transition)
        arrow_match = re.search(
            r"(-?\d+(?:\.\d+)?)\s*(?:->|→|to|到)\s*(-?\d+(?:\.\d+)?)(?:\s*(?:->|→|to|到)\s*(-?\d+(?:\.\d+)?))?",
            normalized,
            flags=re.I,
        )
        if arrow_match:
            values = [float(part) for part in arrow_match.groups() if part is not None]
            return "->".join(self._format_test_plan_number(value) for value in values)
        numbers = self._test_plan_numbers_from_text(normalized)
        if "heatup" in test_intent or "cooldown" in test_intent:
            if len(numbers) >= 3:
                return "->".join(self._format_test_plan_number(value) for value in numbers[-3:])
            if len(numbers) >= 2:
                return "->".join(self._format_test_plan_number(value) for value in numbers[-2:])
        if numbers:
            return f"{self._format_test_plan_number(numbers[-1])} C"
        return ""

    def _infer_test_plan_inputs_from_free_text(self) -> None:
        text = self.test_plan_free_intent_var.get().strip()
        if not text:
            return
        lowered = text.lower()
        self.test_plan_test_var.set("")
        self.test_plan_parameter_var.set("")
        if any(token in lowered for token in ("merge", "combine", "合并")):
            self.test_plan_intent_var.set("Merge")
        elif any(token in lowered for token in ("compare", "对比", "比较")):
            self.test_plan_intent_var.set("Compare")
        elif any(token in lowered for token in ("search", "recommend", "检索", "推荐")):
            self.test_plan_intent_var.set("Search / Recommend")
        else:
            self.test_plan_intent_var.set("Crop / Modify")

        if True:
            if any(token in lowered for token in ("heatup", "heat up", "cooldown", "cool down", "升温", "降温")):
                self.test_plan_test_var.set(self._resolve_test_plan_intent_token("heatup_cooldown", self.test_plan_family_var.get().strip() or "TCC"))
            elif any(token in lowered for token in ("accuracy", "准确", "精度")):
                self.test_plan_test_var.set(self._resolve_test_plan_intent_token("temperature_accuracy", self.test_plan_family_var.get().strip() or "TCC"))
            elif any(token in lowered for token in ("calibration", "校准")):
                self.test_plan_test_var.set(self._resolve_test_plan_intent_token("temperature_calibration", self.test_plan_family_var.get().strip() or "TCC"))
            elif any(token in lowered for token in ("stability", "稳定")):
                self.test_plan_test_var.set(self._resolve_test_plan_intent_token("temperature_stability", self.test_plan_family_var.get().strip() or "TCC"))
            elif any(token in lowered for token in ("precision", "重复", "precision")):
                self.test_plan_test_var.set(self._resolve_test_plan_intent_token("temperature_precision", self.test_plan_family_var.get().strip() or "TCC"))

        if True:
            parameter = self._test_plan_parameter_from_text(lowered, self.test_plan_test_var.get().strip())
            if parameter:
                self.test_plan_parameter_var.set(parameter)
                return
            range_match = re.search(r"(-?\d+(?:\.\d+)?)\s*(?:->|→|to|到)\s*(-?\d+(?:\.\d+)?)(?:\s*(?:->|→|to|到)\s*(-?\d+(?:\.\d+)?))?", lowered)
            if range_match:
                parts = [part for part in range_match.groups() if part is not None]
                self.test_plan_parameter_var.set("->".join(parts))
                return
            temp_match = re.search(r"(-?\d+(?:\.\d+)?)\s*(?:°?\s*c|degc|摄氏|度)", lowered)
            if temp_match:
                self.test_plan_parameter_var.set(f"{temp_match.group(1)} C")

    def _selected_test_plan_record(self) -> FoqAlignmentRecord | None:
        self._infer_test_plan_inputs_from_free_text()
        records = self._test_plan_records()
        family = self.test_plan_family_var.get().strip() or "TCC"
        test_text = self.test_plan_test_var.get().strip()
        family_records = filter_alignment_records(records, family=family)
        matches = filter_alignment_records(family_records, test_text=test_text) if test_text else family_records
        if not matches:
            return None
        device = self.test_plan_device_var.get().strip()
        if device:
            device_matches = filter_alignment_records(matches, devices=(device,))
            if device_matches:
                matches = device_matches
        return matches[0]

    def _analyze_test_plan_intent(self, _event=None) -> None:
        record = self._selected_test_plan_record()
        self.test_plan_selected_record = record
        self.test_plan_accepted_review_text = ""
        if record is None:
            self._populate_test_plan_template_table(None, generate_method_diff=False)
            self._set_test_plan_preview("No matching test plan source found. Try selecting a test or entering an intent such as `accuracy 40 C`.")
            return
        device = self.test_plan_device_var.get().strip()
        if device and device not in record.device_models and record.device_models:
            device = record.device_models[0]
            self.test_plan_device_var.set(device)
        intent = self.test_plan_intent_var.get().strip() or "Search / Recommend"
        parameter = self.test_plan_parameter_var.get().strip()
        selected_records = (record,)
        candidates = filter_alignment_records(self._test_plan_records(), family=self.test_plan_family_var.get().strip() or "TCC")
        self._populate_test_plan_template_table(record, generate_method_diff=False)
        fallback = render_test_plan_assistant_markdown(
            record,
            intent,
            parameter,
            device_model=device,
            selected_records=selected_records,
            candidate_records=candidates,
        )
        content = self._build_test_plan_review_plan(record, device, fallback=fallback)
        self._set_test_plan_preview(content)

    def ai_analyze_test_plan_intent(self, _event=None) -> None:
        self.status_var.set("AI analyzing test intent...")
        self.progress_var.set(8.0)

        def worker() -> None:
            try:
                result = self._ai_or_local_test_plan_intent()
                self._call_ui(lambda result=result: self._apply_ai_test_plan_intent(result))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._ai_test_plan_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _ai_or_local_test_plan_intent(self) -> dict[str, object]:
        local = self._local_test_plan_intent_result()
        key = self.ai_api_key_var.get().strip()
        if not key:
            return local | {"source": "local fallback"}
        try:
            remote = self._call_ai_test_plan_intent_parser(local)
        except Exception as exc:
            return local | {"source": f"local fallback after AI error: {exc}"}
        return local | remote | {"source": "AI"}

    def _local_test_plan_intent_result(self) -> dict[str, object]:
        text = self.test_plan_free_intent_var.get().strip()
        lowered = text.lower()
        family = self.test_plan_family_var.get().strip() or "TCC"
        device = self.test_plan_device_var.get().strip()
        test_intent = self.test_plan_test_var.get().strip()
        if any(token in lowered for token in ("accuracy", "准确")):
            test_intent = "temperature_accuracy"
        elif any(token in lowered for token in ("heatup", "cooldown", "heat up", "cool down", "升温", "降温")):
            test_intent = "heatup_cooldown"
        elif any(token in lowered for token in ("calibration", "校准")):
            test_intent = "temperature_calibration"
        elif any(token in lowered for token in ("stability", "稳定")):
            test_intent = "temperature_stability"
        elif any(token in lowered for token in ("precision", "精密", "重复")):
            test_intent = "temperature_precision"

        setpoint_c: float | None = None
        test_intent = self._resolve_test_plan_intent_token(test_intent, family)
        numbers = self._test_plan_numbers_from_text(lowered)
        if numbers:
            setpoint_c = numbers[-1]
        parameter = self._test_plan_parameter_from_text(lowered, test_intent)
        action = "Crop / Modify" if any(token in lowered for token in ("only", "single", "crop", "modify", "change", "edit")) else self.test_plan_intent_var.get()
        return {
            "family": family,
            "device_model": device,
            "test_intent": test_intent,
            "action": action or "Crop / Modify",
            "parameter": parameter,
            "setpoint_c": setpoint_c,
            "explanation": "Local parser matched test/template and extracted temperature parameters; AI can refine this using the local KB context when configured.",
        }
        temp_match = re.search(r"(-?\d+(?:\.\d+)?)\s*(?:c|°c|℃|度|摄氏度)\b?", lowered)
        if temp_match:
            setpoint_c = float(temp_match.group(1))
        range_match = re.search(r"(-?\d+(?:\.\d+)?)\s*(?:->|→|to)\s*(-?\d+(?:\.\d+)?)", lowered)
        parameter = f"{setpoint_c:g} C" if setpoint_c is not None else ""
        if range_match:
            parameter = f"{float(range_match.group(1)):g}->{float(range_match.group(2)):g}"

        action = "Crop / Modify" if any(token in lowered for token in ("only", "single", "crop", "modify", "只", "单点", "改")) else self.test_plan_intent_var.get()
        return {
            "family": family,
            "device_model": device,
            "test_intent": test_intent,
            "action": action or "Crop / Modify",
            "parameter": parameter,
            "setpoint_c": setpoint_c,
            "explanation": "Local parser matched test keywords and temperature/range parameters.",
        }

    def _build_test_plan_ai_context(self, local_hint: dict[str, object]) -> dict[str, object]:
        family = str(local_hint.get("family") or self.test_plan_family_var.get() or "TCC")
        requested_test = str(local_hint.get("test_intent") or self.test_plan_test_var.get() or "")
        test_intent = self._resolve_test_plan_intent_token(requested_test, family)
        records = self._test_plan_records()
        family_records = filter_alignment_records(records, family=family)
        matches = filter_alignment_records(family_records, test_text=test_intent) if test_intent else family_records
        device = str(local_hint.get("device_model") or self.test_plan_device_var.get() or "")
        if device:
            device_matches = filter_alignment_records(matches, devices=(device,))
            if device_matches:
                matches = device_matches
        record = matches[0] if matches else None
        candidates = [
            {
                "test_intent": candidate.test_intent,
                "td_test": self._clip_ai_text(candidate.td_test, 80),
                "devices": list(candidate.device_models[:6]),
            }
            for candidate in family_records[:30]
        ]
        if record is None:
            return {"available_candidates": candidates, "selected": None, "source_method_rows": []}
        source_rows = self._test_plan_source_method_rows(record)
        compact_rows = self._compact_test_plan_method_rows(record, source_rows, local_hint)
        method_rows = [
            {
                "row": row[0],
                "kind": row[1],
                "time": row[2],
                "command": self._clip_ai_text(row[3], 140),
                "value": self._clip_ai_text(row[4], 100),
                "comment": self._clip_ai_text(row[5], 140),
            }
            for row in compact_rows
        ]
        plan_rows = build_test_plan_modification_steps(
            record,
            str(local_hint.get("action") or self.test_plan_intent_var.get() or "Crop / Modify"),
            str(local_hint.get("parameter") or self.test_plan_parameter_var.get() or ""),
            device_model=device,
        )
        return {
            "available_candidates": candidates,
            "selected": {
                "family": record.family,
                "test_intent": record.test_intent,
                "td_test": self._clip_ai_text(record.td_test, 120),
                "device_models": list(record.device_models[:8]),
                "td_meaning": self._clip_ai_text(record.td_meaning, 360),
                "key_conditions": self._clip_ai_list(record.key_conditions, 8, 160),
                "injection": record.injection,
                "instrument_method": record.instrument_method,
                "expected_ret_times": list(record.expected_ret_times[:12]),
                "expected_channels": list(record.expected_channels[:12]),
                "expected_audit_properties": list(record.expected_audit_properties[:12]),
                "required_config": self._clip_ai_list(record.required_config, 10, 150),
                "method_evidence": self._clip_ai_list(record.method_evidence, 10, 180),
                "coverage_status": record.coverage_status,
                "open_gaps": self._clip_ai_list(record.open_gaps, 8, 180),
            },
            "source_method_row_count": len(source_rows),
            "source_method_rows_sent": len(method_rows),
            "source_method_rows_truncated": len(method_rows) < len(source_rows),
            "source_method_rows": method_rows,
            "existing_rule_plan_rows": [
                {
                    "asset": row[0],
                    "location": self._clip_ai_text(row[2], 160),
                    "change_to": self._clip_ai_text(row[3], 180),
                }
                for row in plan_rows[:12]
            ],
        }

    def _clip_ai_text(self, value: object, limit: int = 180) -> str:
        text = re.sub(r"\s+", " ", str(value or "")).strip()
        if len(text) <= limit:
            return text
        return text[: max(0, limit - 1)].rstrip() + "…"

    def _clip_ai_list(self, values: tuple[str, ...] | list[str], limit: int = 8, text_limit: int = 180) -> list[str]:
        output = [self._clip_ai_text(value, text_limit) for value in list(values or ())[:limit]]
        remaining = max(0, len(values or ()) - limit)
        if remaining:
            output.append(f"+{remaining} more")
        return output

    def _compact_test_plan_method_rows(
        self,
        record: FoqAlignmentRecord,
        rows: list[tuple[object, ...]],
        local_hint: dict[str, object],
        max_rows: int = 45,
    ) -> list[tuple[object, ...]]:
        if len(rows) <= max_rows:
            return rows
        intent_text = " ".join(
            str(value or "")
            for value in (
                self.test_plan_free_intent_var.get(),
                self.test_plan_parameter_var.get(),
                local_hint.get("parameter"),
                local_hint.get("test_intent"),
                record.test_intent,
                record.td_test,
            )
        ).lower()
        number_tokens = {match.group(0) for match in re.finditer(r"-?\d+(?:\.\d+)?", intent_text)}
        base_tokens = {
            "temperature",
            "nominal",
            "ready",
            "delay",
            "rettime",
            "ret time",
            "retimes",
            "modelno",
            "generic",
            "audit",
            "log",
        }
        if "heat" in intent_text or "cool" in intent_text:
            base_tokens.update({"heat", "cool", "50", "20"})
        if "accuracy" in intent_text:
            base_tokens.update({"accuracy", "10", "20", "40", "60", "80", "85", "120"})
        if "stability" in intent_text or "precision" in intent_text:
            base_tokens.update({"stability", "precision", "range"})
        selected: set[int] = set()
        for index, row in enumerate(rows):
            row_text = " ".join(str(cell) for cell in row[:6]).lower()
            kind = str(row[1]).lower() if len(row) > 1 else ""
            command = str(row[3]).lower() if len(row) > 3 else ""
            score = 0
            if kind in {"stage", "branch", "end"}:
                score += 1
            if any(token in row_text for token in base_tokens):
                score += 3
            if number_tokens and any(token in row_text for token in number_tokens):
                score += 2
            if "data_collection_rate" in command and "temperature.nominal" not in command:
                score -= 2
            if score > 0:
                for neighbor in range(max(0, index - 1), min(len(rows), index + 2)):
                    selected.add(neighbor)
        if not selected:
            selected.update(range(min(len(rows), max_rows)))
        ordered = sorted(selected)
        if len(ordered) > max_rows:
            ordered = ordered[:max_rows]
        return [rows[index] for index in ordered]

    def _call_ai_test_plan_intent_parser(self, local_hint: dict[str, object]) -> dict[str, object]:
        base_url = self.ai_base_url_var.get().strip().rstrip("/") or "https://api.openai.com/v1"
        endpoint = f"{base_url}/chat/completions"
        model = self.ai_model_var.get().strip() or "gpt-5.5"
        local_context = self._build_test_plan_ai_context(local_hint)
        prompt = (
            "You analyze a CMBX Test Plan intent using the provided local KB/CMBX context. "
            "Return only compact JSON with keys: family, device_model, test_intent, action, parameter, setpoint_c, "
            "explanation, method_edit_plan, report_suggestions, config_notes, open_verification. "
            "Use available options exactly when possible. method_edit_plan must be an array of objects with keys "
            "step, location, original, proposed, rationale, confidence. For now, focus on instrument method changes; "
            "report_suggestions should be empty unless the user explicitly asks for report edits. Work from local_context as source evidence; do not "
            "invent verified CM commands. If exact commands are missing, say REVIEW REQUIRED in confidence/rationale.\n\n"
            f"Local KB/CMBX context: {json.dumps(local_context, ensure_ascii=False)}\n"
            f"Current UI hint: {json.dumps(local_hint, ensure_ascii=False)}\n"
            f"User intent: {self.test_plan_free_intent_var.get().strip()}"
        )
        body = {
            "model": model,
            "messages": [
                {"role": "system", "content": "Return JSON only. You are a CMBX method/report modification planner grounded in the supplied local KB context."},
                {"role": "user", "content": prompt},
            ],
            "max_completion_tokens": 1200,
        }
        request = urllib.request.Request(
            endpoint,
            data=json.dumps(body).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {self.ai_api_key_var.get().strip()}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=15) as response:
            payload = json.loads(response.read().decode("utf-8"))
        content = payload["choices"][0]["message"]["content"]
        match = re.search(r"\{.*\}", content, flags=re.S)
        if not match:
            raise ValueError("AI response did not contain JSON.")
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else {}

    def _apply_ai_test_plan_intent(self, result: dict[str, object]) -> None:
        family = str(result.get("family") or self.test_plan_family_var.get() or "TCC")
        test_intent = self._resolve_test_plan_intent_token(str(result.get("test_intent") or self.test_plan_test_var.get() or ""), family)
        device = str(result.get("device_model") or self.test_plan_device_var.get() or "")
        action = str(result.get("action") or "Crop / Modify")
        parameter = str(result.get("parameter") or "")
        if not parameter and result.get("setpoint_c") is not None:
            try:
                parameter = f"{float(result['setpoint_c']):g} C"
            except (TypeError, ValueError):
                parameter = ""
        self.test_plan_family_var.set(family)
        self._refresh_test_plan_options()
        if test_intent:
            self.test_plan_test_var.set(test_intent)
        if device:
            self.test_plan_device_var.set(device)
        self.test_plan_intent_var.set(action)
        self.test_plan_parameter_var.set(parameter)
        result["user_intent_text"] = self.test_plan_free_intent_var.get().strip()
        self.test_plan_ai_result = result
        self.test_plan_accepted_review_text = ""
        record = self._selected_test_plan_record()
        self.test_plan_selected_record = record
        if record is not None:
            self._populate_test_plan_template_table(record, generate_method_diff=False)
        source = str(result.get("source") or "AI")
        explanation = str(result.get("explanation") or "")
        self.status_var.set(f"{source} intent: {test_intent or '(unknown)'} {parameter}".strip())
        if record is not None:
            ai_plan = self._test_plan_ai_result_markdown(result)
            self._set_test_plan_preview(self._build_test_plan_review_plan(record, device, ai_result=result, ai_summary=ai_plan, explanation=explanation))
        self.progress_var.set(100.0)

    def _test_plan_ai_result_markdown(self, result: dict[str, object]) -> str:
        lines = ["## AI Method Plan", ""]
        method_plan = result.get("method_edit_plan")
        if isinstance(method_plan, list) and method_plan:
            lines.extend(["### Method edit plan", ""])
            for item in method_plan[:20]:
                if not isinstance(item, dict):
                    continue
                lines.append(
                    "- "
                    + str(item.get("location") or item.get("step") or "method")
                    + ": "
                    + str(item.get("proposed") or "")
                    + (" [" + str(item.get("confidence")) + "]" if item.get("confidence") else "")
                )
        config_notes = result.get("config_notes")
        if isinstance(config_notes, list) and config_notes:
            lines.extend(["", "### Config notes", ""])
            lines.extend(f"- {note}" for note in config_notes[:12])
        open_items = result.get("open_verification")
        if isinstance(open_items, list) and open_items:
            lines.extend(["", "### Open verification", ""])
            lines.extend(f"- {note}" for note in open_items[:12])
        return "\n".join(lines).strip()

    def _build_test_plan_review_plan(
        self,
        record: FoqAlignmentRecord,
        device: str,
        ai_result: dict[str, object] | None = None,
        ai_summary: str = "",
        explanation: str = "",
        fallback: str = "",
    ) -> str:
        intent_text = self.test_plan_free_intent_var.get().strip()
        parameter_text = self.test_plan_parameter_var.get().strip()
        if ai_result:
            parameter_text = str(ai_result.get("parameter") or parameter_text)
        if record.family == "TCC" and record.test_intent == "temperature_accuracy":
            setpoint = self._test_plan_accuracy_setpoint_from_text(" ".join([intent_text, parameter_text, explanation]))
            if setpoint is None:
                setpoint = self._test_plan_setpoint_from_parameter(parameter_text) or 40.0
            transition = self._test_plan_temperature_transition_from_text(" ".join([intent_text, parameter_text, explanation]))
            baseline = transition[0] if transition else None
            source_rows = self._test_plan_source_method_rows(record)
            semantic = analyze_cm_method_rows(source_rows)
            has_stability_word = any(token in intent_text.lower() for token in ("stability", "稳定", "稳定性"))
            related_methods = [record.instrument_method or "<missing>"]
            if has_stability_word:
                related_methods.append("TEMPERATURE_STABILITY_70_C / TEMPERATURE_STABILITY_AND_PCC_70_H (reference for stability report meaning, not the current script basis)")
            lines = [
                "# Method Modification Review",
                "",
                f"Intent interpreted by method roles: baseline {baseline:g} C -> target {setpoint:g} C" if baseline is not None else f"Intent interpreted by method roles: target {setpoint:g} C",
                f"Device: {device or record.device_label}",
                f"Template method: {record.instrument_method or '<missing>'}",
                "",
                "## Method scripts implicated",
                "",
                "| Role | Method script | Why it is involved |",
                "|---|---|---|",
                f"| Primary editable basis | {related_methods[0]} | Contains the variable-driven temperature ladder, stability gate, and RetTime anchors for Accuracy. |",
            ]
            if len(related_methods) > 1:
                lines.append(f"| Reference only | {related_methods[1]} | User intent mentions stability; confirm whether final report metric is Accuracy deviation or Stability range. |")
            lines.extend([
                "",
                "## Natural-language modification contract",
                "",
                f"- [CHANGE] Treat {setpoint:g} C as the target measurement point.",
                f"- [CHANGE] Treat {baseline:g} C as pre-equilibration/baseline, not as the target report point." if baseline is not None else "- [OPEN] Baseline/pre-equilibration temperature is not specified.",
                "- [LOCKED] Preserve CC.TempReady and external upper/lower thermometer stability gates before writing RetTime.",
                "- [LOCKED] Preserve RetTime semantics unless report cells are explicitly remapped.",
                "- [LOCKED] Do not modify the final direct 20 C reset command unless cleanup behavior changes.",
                "- [OPEN] Confirm whether the final report should calculate Accuracy deviation or Stability range; the method script alone cannot decide this.",
                "",
                "## Editable assumptions",
                "",
                "| Item | Current proposal | User editable note |",
                "|---|---|---|",
                f"| Target accuracy point | {setpoint:g} C | Change this if the target test point is different. |",
                f"| Pre-equilibration | {baseline:g} C | This is a setup state before target measurement. |" if baseline is not None else "| Pre-equilibration | keep source method behavior | Add a sentence such as `first equilibrate at 20 C, then test 40 C` if required. |",
                "| Non-target accuracy points | skip/remove their measurement blocks | Confirm whether RetTimes/report cells should be remapped or left sparse. |",
                "| Safety reset | keep source reset/final stabilization commands | Only edit if CM run requirement changes. |",
                "",
                "## High-impact editable script roles",
                "",
                "| Status | Script role | Exact object | Meaning |",
                "|---|---|---|---|",
            ])
            if baseline is not None:
                lines.append(f"| [CHANGE] | Baseline source variable | GenericDouble* = {baseline:g} C | Use as first stable state before target. |")
            lines.extend([
                f"| [CHANGE] | Target source variable | GenericDouble* = {setpoint:g} C | Retain this measurement block and its RetTime anchor. |",
                "| [LOCKED] | Stability gate | TempReady + external probe stability counters | Required before valid report window. |",
                "| [LOCKED] | Final reset | Direct literal 20 C after final RetTime | Cleanup/reset, not a measurement point. |",
                "",
                "## Method semantic anchors detected",
                "",
                "| Role | Row(s) | Meaning | Proposed handling |",
                "|---|---:|---|---|",
            ])
            target_blocks = semantic.blocks_for_setpoint(setpoint)
            if target_blocks:
                for block in target_blocks:
                    ret_times = ", ".join(item.ret_time for item in block.ret_times) or "(no RetTime)"
                    lines.append(
                        f"| Target measurement block | {block.setpoint.row_number}..{source_rows[block.end_row_index][0]} | "
                        f"{block.setpoint.target} = {block.setpoint.value}; RetTimes: {ret_times} | keep |"
                    )
            else:
                lines.append(f"| Target measurement block | open | No exact {setpoint:g} C block detected in current method rows. | review before diff |")
            for block in semantic.measurement_blocks:
                if block.setpoint.numeric_value is None or abs(block.setpoint.numeric_value - setpoint) <= 1e-9:
                    continue
                if block.role == "safety_reset":
                    continue
                lines.append(
                    f"| Non-target measurement block | {block.setpoint.row_number}..{source_rows[block.end_row_index][0]} | "
                    f"{block.setpoint.numeric_value:g} C block | skip/remove preview |"
                )
            for row_index in semantic.safety_reset_rows:
                row = source_rows[row_index]
                lines.append(f"| [LOCKED] Safety/reset command | {row[0]} | {row[3]} = {row[4]} | keep |")
            if ai_summary:
                lines.extend(["", "## AI notes", "", ai_summary])
            lines.extend(
                [
                    "",
                    "## Human confirmation",
                    "",
                    "Edit this section until it matches the intended test. Then click `Generate Method Diff`.",
                    "",
                    "- Confirmed target point:",
                    f"- Confirmed baseline/equilibration requirement: {'not specified' if '平衡' not in intent_text and 'equilibr' not in intent_text.lower() else intent_text}",
                    "- Confirmed RetTime/report remap requirement:",
                ]
            )
            return "\n".join(lines)
        return fallback or self._test_plan_ai_result_markdown(ai_result or {})

    def apply_test_plan_review_to_method_diff(self) -> None:
        record = self.test_plan_selected_record or self._selected_test_plan_record()
        if record is None:
            messagebox.showwarning(APP_NAME, "Select or enter an intent before generating the method diff.")
            return
        self.test_plan_selected_record = record
        self.test_plan_accepted_review_text = self._get_test_plan_review_text()
        target = self._test_plan_accuracy_setpoint_from_text(self.test_plan_accepted_review_text)
        if target is not None and record.family == "TCC" and record.test_intent == "temperature_accuracy":
            self.test_plan_parameter_var.set(f"{target:g} C")
        device = self.test_plan_device_var.get().strip() or record.device_label
        self._populate_test_plan_template_table(record, generate_method_diff=True)
        self.status_var.set("Generated method diff from accepted Review Plan")

    def _get_test_plan_review_text(self) -> str:
        text_widget = getattr(self, "test_plan_preview_text", None)
        if text_widget is None:
            return ""
        return text_widget.get("1.0", tk.END).strip()

    def _test_plan_accuracy_setpoint_from_text(self, text: str) -> float | None:
        normalized = text or ""
        transition = self._test_plan_temperature_transition_from_text(normalized)
        if transition is not None:
            return transition[1]
        preferred_patterns = (
            r"(?:target|目标|目标点|测试点|测量点|上升到|升到|到)[^\d\-]{0,18}(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度|度)?",
            r"(?:accuracy|准确|准确性|测试|只测|only|target|目标)[^\d\-]{0,24}(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度)?",
            r"(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度)[^\n\r]{0,18}(?:accuracy|准确|准确性|测试|only|目标)",
        )
        for pattern in preferred_patterns:
            match = re.search(pattern, normalized, flags=re.I)
            if match:
                try:
                    return float(match.group(1))
                except ValueError:
                    pass
        candidates: list[tuple[float, str]] = []
        for match in re.finditer(r"(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度)?", normalized, flags=re.I):
            try:
                value = float(match.group(1))
            except ValueError:
                continue
            window = normalized[max(0, match.start() - 18) : min(len(normalized), match.end() + 18)].lower()
            candidates.append((value, window))
        non_baseline = [
            value
            for value, window in candidates
            if not any(token in window for token in ("平衡", "稳定态", "equilibr", "baseline", "预平衡", "起始", "start", "from", "从"))
        ]
        if non_baseline:
            return non_baseline[-1] if any(token in normalized.lower() for token in ("到", "上升", "升到", "to", "target")) else non_baseline[0]
        return candidates[0][0] if candidates else None

    def _test_plan_temperature_transition_from_text(self, text: str) -> tuple[float, float] | None:
        normalized = text or ""
        patterns = (
            r"(?:从|from)\s*(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度|度)?[^\d\-]{0,30}(?:到|至|上升到|升到|to|->|→)\s*(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度|度)?",
            r"(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度|度)?\s*(?:->|→|to|到|至)\s*(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|摄氏度|度)?",
        )
        for pattern in patterns:
            match = re.search(pattern, normalized, flags=re.I)
            if not match:
                continue
            try:
                return float(match.group(1)), float(match.group(2))
            except ValueError:
                continue
        return None

    def _ai_test_plan_failed(self, exc: Exception) -> None:
        self.progress_var.set(0.0)
        self.status_var.set("AI analysis failed")
        messagebox.showerror(APP_NAME, f"AI analysis failed:\n{exc}")

    def _populate_test_plan_template_table(self, record: FoqAlignmentRecord | None, generate_method_diff: bool = False) -> None:
        for table_name in ("test_plan_candidate_table", "test_plan_method_table", "test_plan_report_table", "test_plan_config_table", "test_plan_pipeline_table", "test_plan_basis_table", "test_plan_role_table"):
            table_frame = getattr(self, table_name, None)
            if table_frame is None:
                continue
            table = self._table_widget(table_frame)
            table.delete(*table.get_children())
        if record is None:
            self.test_plan_source_status_var.set("No matching method basis")
            basis_frame = getattr(self, "test_plan_basis_table", None)
            if basis_frame is not None:
                basis_table = self._table_widget(basis_frame)
                basis_table.insert("", "end", iid="basis:none", values=("Method basis", "No matching test/method source found", "enter a clearer intent or refresh Method Script KB"))
            return
        device = self.test_plan_device_var.get().strip() or record.device_label
        source_rows = self._test_plan_source_method_rows(record)
        self._populate_test_plan_pipeline_table(record, device, source_rows)
        self._populate_test_plan_basis_table(record, device, source_rows)
        self._populate_test_plan_role_table(record, source_rows)
        self._populate_test_plan_method_preview(record, device, generate_modified=generate_method_diff)

    def _populate_test_plan_pipeline_table(self, record: FoqAlignmentRecord, device: str, source_rows: list[tuple[object, ...]]) -> None:
        table_frame = getattr(self, "test_plan_pipeline_table", None)
        if table_frame is None:
            return
        table = self._table_widget(table_frame)
        table.delete(*table.get_children())
        intent_text = self.test_plan_free_intent_var.get().strip() or "<empty intent>"
        ai_result = getattr(self, "test_plan_ai_result", None)
        ai_source = "AI parser" if isinstance(ai_result, dict) and ai_result.get("user_intent_text") == self.test_plan_free_intent_var.get().strip() else "local rules"
        method_loaded = self._has_real_cm_method_rows(source_rows)
        semantic = analyze_cm_method_rows(source_rows) if method_loaded else None
        report_bits: list[str] = []
        if record.report_template:
            report_bits.append(record.report_template)
        if record.report_sheet_label:
            report_bits.append(record.report_sheet_label)
        if record.db_fields:
            field_text = ", ".join(record.db_fields[:5])
            if len(record.db_fields) > 5:
                field_text += f", +{len(record.db_fields) - 5}"
            report_bits.append(field_text)
        report_decision = "Report is a calculation constraint; this method-only step does not rewrite report Excel."
        if not report_bits:
            report_decision = "BLOCKED: report/formula binding is missing."
        rows = (
            (
                "1. Natural Semantic",
                intent_text,
                ai_source,
                "Confirm this sentence first; do not edit method rows until the semantic target is right.",
            ),
            (
                "2. Test Intent",
                f"{record.family} / {record.test_intent or record.td_test} / {device or record.device_label}",
                record.td_source or "FOQ KB + TKN alignment",
                self._test_plan_intent_decision_text(record),
            ),
            (
                "3. CM Execution Mechanism",
                self._test_plan_cm_mechanism_summary(semantic, record),
                "Method Role Map from full KB method script + CM command KB",
                "Variables, RetTimes, waits, triggers, and logs must be moved as coupled roles.",
            ),
            (
                "4. Method Script",
                f"{record.instrument_method or '<missing>'}; {len(source_rows)} rendered row(s)",
                r"C:\ProgramData\CMBX Data Explorer Workspace\KB\CMBX Method Scripts",
                "KB script is the only generation basis." if method_loaded else "BLOCKED: full method script is not in KB.",
            ),
            (
                "5. Report Calculation Constraint",
                " | ".join(report_bits) if report_bits else "<missing report/formula binding>",
                self._test_plan_report_evidence_summary(record),
                report_decision,
            ),
        )
        for index, row in enumerate(rows, start=1):
            table.insert("", "end", iid=f"pipeline:{index}", values=row)

    def _test_plan_intent_decision_text(self, record: FoqAlignmentRecord) -> str:
        gaps = len(record.open_gaps or ())
        if record.coverage_status == "complete" and not gaps:
            return "Intent is covered by current KB; method-level change can be planned."
        if gaps:
            return f"{record.coverage_status or 'partial'}; {gaps} open verification item(s) remain."
        return record.coverage_status or "coverage not classified"

    def _test_plan_cm_mechanism_summary(self, semantic, record: FoqAlignmentRecord) -> str:
        if semantic is None:
            evidence = "; ".join(record.method_evidence[:3])
            return evidence or "<no method role map; full script missing from KB>"
        parts: list[str] = []
        if semantic.temperature_variables:
            parts.append(f"{len(semantic.temperature_variables)} temperature variable(s)")
        if semantic.measurement_blocks:
            parts.append(f"{len(semantic.measurement_blocks)} measurement block(s)")
        if semantic.ret_times:
            parts.append(f"{len([item for item in semantic.ret_times if item.emission])} RetTime anchor(s)")
        if semantic.waits:
            parts.append(f"{len(semantic.waits)} ready/wait gate(s)")
        if semantic.triggers:
            parts.append(f"{len(semantic.triggers)} trigger(s)")
        if not parts:
            parts.append("script loaded; no special roles detected")
        return "; ".join(parts)

    def _test_plan_report_evidence_summary(self, record: FoqAlignmentRecord) -> str:
        evidence: list[str] = []
        if record.report_evidence:
            evidence.extend(record.report_evidence[:2])
        if record.db_evidence:
            evidence.extend(record.db_evidence[:2])
        if record.expected_ret_times:
            evidence.append("RetTimes: " + ", ".join(record.expected_ret_times[:5]))
        return "; ".join(evidence) if evidence else "Report KB / DB mapping evidence not recorded"

    def _populate_test_plan_basis_table(self, record: FoqAlignmentRecord, device: str, source_rows: list[tuple[object, ...]]) -> None:
        table_frame = getattr(self, "test_plan_basis_table", None)
        if table_frame is None:
            return
        table = self._table_widget(table_frame)
        table.delete(*table.get_children())
        method_basis = "KB method script rows" if self._has_real_cm_method_rows(source_rows) else "KB/alignment evidence only"
        if self._has_real_cm_method_rows(source_rows):
            method_basis = "KB method script rows"
            status = "exact original method script available from KB"
        else:
            status = "refresh Method Script KB from known-good CMBX evidence"
        rows = (
            ("Method Script KB", method_basis, status),
            ("Detected test", record.td_test or record.test_intent, record.test_intent),
            ("Device", device or record.device_label or ", ".join(record.device_models), "inferred from intent/KB"),
            ("Instrument method", record.instrument_method or "<missing>", status),
            ("CMBX dependency", "not required at generation time", "CMBX is only an offline KB refresh source"),
        )
        for index, row in enumerate(rows, start=1):
            table.insert("", "end", iid=f"basis:{index}", values=row)
        self.test_plan_source_status_var.set(f"{record.test_intent or record.td_test}: {method_basis}")

    def _populate_test_plan_role_table(self, record: FoqAlignmentRecord, source_rows: list[tuple[object, ...]]) -> None:
        table_frame = getattr(self, "test_plan_role_table", None)
        if table_frame is None:
            return
        table = self._table_widget(table_frame)
        table.delete(*table.get_children())
        if not self._has_real_cm_method_rows(source_rows):
            table.insert(
                "",
                "end",
                iid="role:none",
                values=(
                    "Open",
                    "",
                    "Exact method script is not loaded from KB.",
                    "Refresh Method Script KB from known-good CMBX evidence before row-level generation.",
                ),
            )
            return
        semantic = analyze_cm_method_rows(source_rows)
        rows: list[tuple[str, str, str, str]] = []
        temp_vars = []
        for variable in semantic.temperature_variables:
            assignments = semantic.assignments_for_variable(variable)
            values = ", ".join(f"{item.row_number}={item.value}" for item in assignments[:3])
            numeric_values = [item.numeric_value for item in assignments if item.numeric_value is not None]
            label = f"{variable}"
            if numeric_values:
                label += f" ({numeric_values[0]:g} C)"
            temp_vars.append(label)
            rows.append((
                "Temperature ladder variable",
                values or ", ".join(item.row_number for item in assignments),
                f"{variable} is a source temperature point consumed by Temperature.Nominal.",
                "Editable only through the ladder role; do not confuse with final reset literals.",
            ))
        if temp_vars:
            rows.insert(0, ("Temperature ladder", ", ".join(temp_vars), "Model-specific setpoint ladder learned from GenericDouble variables.", "Choose target/baseline from this ladder before editing."))
        for block in semantic.measurement_blocks[:12]:
            ret_times = ", ".join(item.ret_time for item in block.ret_times) or "no RetTime"
            setpoint = block.setpoint.numeric_value
            setpoint_text = f"{setpoint:g} C" if setpoint is not None else block.setpoint.value
            rows.append((
                "Measurement block" if block.role != "safety_reset" else "Safety/reset block",
                f"{block.setpoint.row_number}..{source_rows[block.end_row_index][0]}",
                f"{setpoint_text}; {ret_times}; waits={len(block.waits)}",
                "Keep target block, skip/remap non-target blocks, preserve RetTime semantics." if block.role != "safety_reset" else "Keep unless the run cleanup requirement changes.",
            ))
        if semantic.waits:
            waits = ", ".join(item.row_number for item in semantic.waits[:8])
            rows.append(("Ready/stability gates", waits, "Wait/ready conditions gate valid measurement windows.", "Locked for accuracy/stability unless a new validation rule is written."))
        for trigger in semantic.triggers[:12]:
            if trigger.time_window_start is not None and trigger.time_window_end is not None:
                window = f"{trigger.time_window_start:g}-{trigger.time_window_end:g} min"
            else:
                window = "open retention window"
            interval = f"{trigger.rearm_minutes * 60:g} s" if trigger.rearm_minutes is not None else "unknown interval"
            valves = "; ".join(trigger.valve_positions) or "no valve position command"
            logs = ", ".join(trigger.logged_properties[:5]) or "no explicit Log command"
            rows.append((
                "Periodic trigger",
                trigger.row_number,
                f"{trigger.name}: {window}; re-arm={interval}; {valves}; logs={logs}",
                "Edit trigger window, re-arm interval, valve positions, and bool handoff as one coupled role.",
            ))
        if semantic.safety_reset_rows:
            reset_rows = ", ".join(str(source_rows[index][0]) for index in semantic.safety_reset_rows)
            rows.append(("Final reset", reset_rows, "Direct nominal temperature after final RetTime, usually cleanup/reset.", "Do not treat as a test point."))
        for index, row in enumerate(rows or [("Open", "", "No semantic roles detected.", "Review method manually.")], start=1):
            table.insert("", "end", iid=f"role:{index}", values=row)

    def _populate_test_plan_method_preview(self, record: FoqAlignmentRecord, device: str, generate_modified: bool = True) -> None:
        source_frame = getattr(self, "test_plan_method_source_table", None)
        modified_frame = getattr(self, "test_plan_method_modified_table", None)
        if source_frame is None or modified_frame is None:
            return
        source_table = self._table_widget(source_frame)
        modified_table = self._table_widget(modified_frame)
        source_table.delete(*source_table.get_children())
        modified_table.delete(*modified_table.get_children())
        for index, row in enumerate(self._test_plan_source_method_rows(record), start=1):
            tag = str(row[6]) if len(row) > 6 else ""
            source_table.insert("", "end", iid=f"source:{index}", values=row[:6], tags=(tag or "source",))
        if not generate_modified:
            modified_table.insert(
                "",
                "end",
                iid="modified:pending",
                values=("", "Pending", "", "Edit Review Plan, then click Generate Method Diff", "", "Modified script is generated only after the language contract is accepted."),
                tags=("source",),
            )
            return
        modified_rows = self._test_plan_modified_method_rows(record, device)
        for index, (row, changed) in enumerate(modified_rows, start=1):
            tag = str(row[6]) if len(row) > 6 else ""
            tags = ("modified",) if changed else ((tag,) if tag else ())
            modified_table.insert("", "end", iid=f"modified:{index}", values=row[:6], tags=tags)

    def _populate_test_plan_ai_report_rows(self, record: FoqAlignmentRecord, device: str) -> None:
        result = getattr(self, "test_plan_ai_result", None)
        if not isinstance(result, dict):
            return
        if str(result.get("user_intent_text") or "") != self.test_plan_free_intent_var.get().strip():
            return
        family = str(result.get("family") or "")
        test_intent = self._resolve_test_plan_intent_token(str(result.get("test_intent") or ""), family or record.family)
        result_device = str(result.get("device_model") or "")
        if family and family != record.family:
            return
        if test_intent and test_intent != record.test_intent:
            return
        if result_device and device and result_device != device:
            return
        suggestions = result.get("report_suggestions")
        if not isinstance(suggestions, list) or not suggestions:
            return
        frame = getattr(self, "test_plan_report_table", None)
        if frame is None:
            return
        table = self._table_widget(frame)
        base = len(table.get_children()) + 1
        for offset, item in enumerate(suggestions[:30], start=0):
            if not isinstance(item, dict):
                continue
            table.insert(
                "",
                "end",
                iid=f"ai_report:{base + offset}",
                values=(
                    "Report Template",
                    record.report_template or "<report>",
                    item.get("location") or "AI report suggestion",
                    item.get("proposed") or "",
                    item.get("rationale") or item.get("confidence") or "AI grounded in local KB context.",
                ),
            )

    def _test_plan_source_method_rows(self, record: FoqAlignmentRecord) -> list[tuple[object, ...]]:
        method_name = record.instrument_method
        kb_rows = load_method_script_rows_from_kb(
            method_name,
            family=record.family or self.test_plan_family_var.get().strip() or "TCC",
            device_model=self.test_plan_device_var.get().strip() or record.device_label,
            workspace_root=Path.cwd(),
        )
        if kb_rows:
            return [tuple(row) for row in kb_rows]
        rows: list[tuple[object, ...]] = []
        rows.append(("", "Source", "", method_name or "<not bound>", "", "No full method script was found in KB. Refresh KB from known-good CMBX evidence before generating.", "source"))
        for index, evidence in enumerate(record.method_evidence[:12], start=1):
            rows.append((str(index), "Comment", "", evidence, "", "Method evidence from KB/alignment.", "cm_comment"))
        if len(rows) == 1:
            rows.append(("", "Open", "", "MISSING", "", "No method evidence found in KB.", "modified"))
        return rows

    def _cm_method_rows_for_package_method(self, package: CmbxPackage, method: CmbxElement) -> list[tuple[str, str, str, str, str, str, str]]:
        external = discover_external_instrument_methods(package.path, [method.name]).get(method.name)
        if external:
            return self._external_method_rows(external)
        try:
            embedded = extract_embedded_instrument_method(package, method)
            if not embedded:
                return []
            temp_dir = Path(tempfile.gettempdir()) / "cmbx_data_explorer_test_plan_method"
            temp_dir.mkdir(parents=True, exist_ok=True)
            stem = self._safe_temp_stem(f"{package.path.name}_{method.id or method.name}")
            cpxm_path = temp_dir / f"{stem}.cpxm.bin"
            xml_path = temp_dir / f"{stem}.xml"
            cpxm_path.write_bytes(embedded.cpxm_payload)
            decode_result = decode_cpxm_method_xml(cpxm_path, xml_path)
            if not decode_result.ok:
                return []
            rows, error = build_method_flow_rows(xml_path.read_text(encoding="utf-8"))
            if error:
                return []
            return self._cm_method_rows_from_flow_rows(rows)
        except Exception:
            return []

    def _method_compare_line_to_preview_row(self, index: int, line: str) -> tuple[object, ...]:
        parts = line.split("\t")
        if len(parts) >= 4:
            return (index, parts[0], parts[1], parts[2], "", parts[3])
        if len(parts) == 3:
            return (index, parts[0], parts[1], parts[2], "", "")
        if len(parts) == 2:
            return (index, "", parts[0], parts[1], "", "")
        return (index, "", "", line, "", "")

    def _test_plan_modified_method_rows(self, record: FoqAlignmentRecord, device: str) -> list[tuple[tuple[object, ...], bool]]:
        source_rows = self._test_plan_source_method_rows(record)
        if record.family == "TCC" and record.test_intent == "heatup_cooldown_20_50_20":
            return self._apply_heatup_cooldown_to_cm_rows(source_rows, record, device)
        if record.family == "TCC" and record.test_intent == "temperature_accuracy":
            try:
                accepted_text = self.test_plan_accepted_review_text or self._get_test_plan_review_text()
                setpoint = (
                    self._test_plan_accuracy_setpoint_from_text(accepted_text)
                    or self._test_plan_setpoint_from_parameter(self.test_plan_parameter_var.get())
                    or 40.0
                )
                if self._has_real_cm_method_rows(source_rows):
                    return self._apply_accuracy_setpoint_to_cm_rows(source_rows, setpoint)
                project = build_single_point_temperature_accuracy_project(device or "VH-C10-A", setpoint)
                rows = []
                for row in self._single_point_accuracy_preview_rows(project):
                    changed = self._method_preview_row_is_changed(row, setpoint)
                    rows.append((row, changed))
                return rows
            except Exception as exc:
                return [(("", "Error", "", "OPEN", "", str(exc), "modified"), True)]
        ai_rows = self._test_plan_ai_method_rows(record, device, source_rows)
        if ai_rows:
            return ai_rows
        if source_rows:
            return [(row, False) for row in source_rows]
        return [(("", "Open", "", "NOT GENERATED", record.instrument_method, "No source method rows available.", "modified"), True)]

    def _test_plan_ai_method_rows(self, record: FoqAlignmentRecord, device: str, source_rows: list[tuple[object, ...]] | None = None) -> list[tuple[tuple[object, ...], bool]]:
        result = getattr(self, "test_plan_ai_result", None)
        if not isinstance(result, dict):
            return []
        if str(result.get("user_intent_text") or "") != self.test_plan_free_intent_var.get().strip():
            return []
        family = str(result.get("family") or "")
        test_intent = self._resolve_test_plan_intent_token(str(result.get("test_intent") or ""), family or record.family)
        result_device = str(result.get("device_model") or "")
        if family and family != record.family:
            return []
        if test_intent and test_intent != record.test_intent:
            return []
        if result_device and device and result_device != device:
            return []
        plan = result.get("method_edit_plan")
        if not isinstance(plan, list) or not plan:
            return []
        if source_rows:
            marked = self._mark_ai_referenced_method_rows(source_rows, plan)
            if marked:
                return marked
        rows: list[tuple[tuple[object, ...], bool]] = []
        for index, item in enumerate(plan[:80], start=1):
            if not isinstance(item, dict):
                continue
            rows.append(
                (
                    (
                        str(item.get("step") or index),
                        "Command",
                        "",
                        item.get("location") or "AI plan",
                        item.get("proposed") or "",
                        item.get("rationale") or item.get("confidence") or "",
                        "modified",
                    ),
                    True,
                )
            )
        return rows

    def _mark_ai_referenced_method_rows(self, source_rows: list[tuple[object, ...]], plan: list[object]) -> list[tuple[tuple[object, ...], bool]]:
        references: list[str] = []
        for item in plan[:80]:
            if not isinstance(item, dict):
                continue
            for key in ("location", "original", "proposed"):
                value = str(item.get(key) or "").strip().lower()
                if value:
                    references.append(value)
        output: list[tuple[tuple[object, ...], bool]] = []
        for row in source_rows:
            row_text = " ".join(str(cell) for cell in row[:6]).lower()
            changed = any(token and token in row_text for token in references)
            output.append((row, changed))
        return output

    def _has_real_cm_method_rows(self, rows: list[tuple[object, ...]]) -> bool:
        return any(str(row[1]) in {"Stage", "Command", "Branch", "Comment", "End", "Time"} and str(row[0]).strip().isdigit() for row in rows)

    def _apply_heatup_cooldown_to_cm_rows(self, source_rows: list[tuple[object, ...]], record: FoqAlignmentRecord, device: str) -> list[tuple[tuple[object, ...], bool]]:
        if not self._has_real_cm_method_rows(source_rows):
            return self._heatup_cooldown_preview_rows(record, device)
        values = self._test_plan_numbers_from_text(self.test_plan_parameter_var.get())
        if len(values) >= 3:
            start, target, return_temp = values[-3], values[-2], values[-1]
        else:
            start, target, return_temp = 20.0, 50.0, 20.0
        replacements = {
            20.0: self._format_test_plan_number(start),
            50.0: self._format_test_plan_number(target),
        }
        if abs(return_temp - start) > 1e-9:
            replacements[20.0] = self._format_test_plan_number(return_temp)
        output: list[tuple[tuple[object, ...], bool]] = []
        changed_count = 0
        for row in source_rows:
            row_list = list(row)
            changed = False
            if len(row_list) >= 6 and str(row_list[1]) == "Command" and "temperature" in str(row_list[3]).lower():
                new_value, changed = self._replace_known_temperature_value(str(row_list[4]), replacements)
                if changed:
                    row_list[4] = new_value
                    changed_count += 1
            output.append((tuple(row_list), changed))
        if changed_count == 0:
            hint = (
                "",
                "Comment",
                "",
                f"Intent review: change HeatUp/CoolDown range to {start:g}->{target:g}->{return_temp:g}",
                "",
                "No direct 20/50 temperature setpoint row was found in the rendered method. Review the source script manually.",
                "modified",
            )
            output.insert(0, (hint, True))
        return output

    def _apply_accuracy_setpoint_to_cm_rows(
        self,
        source_rows: list[tuple[object, ...]],
        setpoint: float,
        *,
        spec: dict[str, object] | None = None,
        device: str = "",
    ) -> list[tuple[tuple[object, ...], bool]]:
        semantic = analyze_cm_method_rows(source_rows)
        if semantic.temperature_variables:
            return self._apply_variable_driven_accuracy_setpoint_to_cm_rows(source_rows, setpoint, semantic, spec=spec or {}, device=device)
        output: list[tuple[tuple[object, ...], bool]] = []
        changed_any = False
        for row in source_rows:
            row_list = list(row)
            changed = False
            if len(row_list) >= 6:
                command_text = str(row_list[3]).lower()
                value_text = str(row_list[4])
                if "temperature" in command_text or "nominal" in command_text:
                    new_value, changed = self._replace_known_temperature_value(value_text, {10.0: f"{setpoint:g}", 20.0: f"{setpoint:g}", 40.0: f"{setpoint:g}", 60.0: f"{setpoint:g}", 80.0: f"{setpoint:g}", 85.0: f"{setpoint:g}", 120.0: f"{setpoint:g}"})
            if changed:
                row_list[4] = new_value
                changed_any = True
            output.append((tuple(row_list), changed))
        if not changed_any:
            hint = (
                "",
                "Comment",
                "",
                f"Intent review: reduce Temperature Accuracy to {setpoint:g} C only",
                "",
                "No directly replaceable setpoint row was found. Full row deletion/RetTime renumbering requires manual review.",
                "modified",
            )
            output.insert(0, (hint, True))
        return output

    def _apply_variable_driven_accuracy_setpoint_to_cm_rows(
        self,
        source_rows: list[tuple[object, ...]],
        setpoint: float,
        semantic,
        *,
        spec: dict[str, object] | None = None,
        device: str = "",
    ) -> list[tuple[tuple[object, ...], bool]]:
        spec = spec or {}
        baseline = self._method_generator_baseline_temperature(spec)
        duration = self._method_generator_float_or_none(spec.get("duration_minutes"))
        device_ladder = self._accuracy_device_ladder_assignments(semantic, device)
        temperature_variables = tuple(item.variable for item in device_ladder) or semantic.temperature_variables
        variable_values = {item.variable: item.numeric_value for item in device_ladder if item.numeric_value is not None}
        assignment_rows_by_var = {item.variable: [item.row_index] for item in device_ladder}
        if not device_ladder:
            for variable in temperature_variables:
                assignments_for_var = semantic.assignments_for_variable(variable)
                assignment_rows_by_var[variable] = [item.row_index for item in assignments_for_var]
                variable_values[variable] = next((item.numeric_value for item in assignments_for_var if item.numeric_value is not None), None)
        temperature_uses: dict[str, list[int]] = {}
        for event in semantic.temperature_setpoints:
            if event.variable:
                temperature_uses.setdefault(event.variable, []).append(event.row_index)
        target_variables = {variable for variable, value in variable_values.items() if value is not None and abs(value - setpoint) <= 1e-9}
        changed_assignment_rows: dict[int, str] = {}
        if not target_variables and temperature_variables:
            chosen = self._choose_accuracy_target_variable(device_ladder, setpoint, baseline) or temperature_variables[0]
            target_variables.add(chosen)
            for row_index in assignment_rows_by_var.get(chosen, []):
                changed_assignment_rows[row_index] = f"{setpoint:g}"
        skipped_variables = set(temperature_variables) - target_variables
        skipped_indices = self._cm_accuracy_measurement_block_indices(semantic, skipped_variables)
        target_use_indices = {index for variable in target_variables for index in temperature_uses.get(variable, [])}
        target_block_start = min(target_use_indices) if target_use_indices else -1
        target_block_indices = self._cm_accuracy_measurement_block_indices(semantic, set(target_variables))
        direct_temp_indices = set(semantic.safety_reset_rows)
        assignment_variable_by_row = {
            row_index: variable
            for variable, row_indices in assignment_rows_by_var.items()
            for row_index in row_indices
        }
        output: list[tuple[tuple[object, ...], bool]] = []
        for index, row in enumerate(source_rows):
            row_list = list(row)
            if len(row_list) < 7:
                row_list.extend([""] * (7 - len(row_list)))
            changed = False
            variable = assignment_variable_by_row.get(index, "")
            if baseline is not None and index == target_block_start:
                row_id = str(row_list[0])
                output.extend(self._accuracy_baseline_preview_rows(row_id, baseline, setpoint, duration))
            if index in changed_assignment_rows:
                row_list[4] = changed_assignment_rows[index]
                row_list[5] = self._append_preview_note(row_list[5], f"AI/local contract: set device-specific source variable for single-point accuracy {setpoint:g} C.")
                row_list[6] = "modified"
                changed = True
            elif variable in target_variables:
                original_value = variable_values.get(variable)
                note = f"AI/local contract: keep target accuracy source variable ({variable} = {original_value:g} C)." if original_value is not None else f"AI/local contract: keep target accuracy source variable {variable}."
                row_list[5] = self._append_preview_note(row_list[5], note)
                row_list[6] = "modified"
                changed = True
            elif variable in skipped_variables:
                row_list[5] = self._append_preview_note(row_list[5], f"AI/local contract: skip device-specific source variable {variable} for single-point {setpoint:g} C preview.")
                row_list[6] = "removed"
            elif index in target_use_indices:
                used_variable = cm_method_variable_name(str(row_list[4]))
                row_list[5] = self._append_preview_note(row_list[5], f"AI/local contract: target measurement block uses {used_variable}, not a direct temperature literal.")
                row_list[6] = "modified"
                changed = True
            elif index in target_block_indices:
                row_list[5] = self._append_preview_note(row_list[5], f"AI/local contract: keep target accuracy block for {setpoint:g} C.")
                row_list[6] = "modified"
                changed = True
            elif index in skipped_indices:
                row_list[5] = self._append_preview_note(row_list[5], f"AI/local contract: remove/skip this non-target temperature measurement block for {setpoint:g} C only.")
                row_list[6] = "removed"
            elif index in direct_temp_indices:
                row_list[5] = self._append_preview_note(row_list[5], "AI/local contract: keep direct temperature command; likely safety/reset/final stabilization, not the accuracy ladder variable source.")
            output.append((tuple(row_list), changed))
        return output

    def _cm_variable_measurement_block_indices(
        self,
        rows: list[tuple[object, ...]],
        temperature_uses: dict[str, list[int]],
        skipped_variables: set[str],
    ) -> set[int]:
        all_use_indices = sorted(index for indices in temperature_uses.values() for index in indices)
        skipped: set[int] = set()
        for variable in skipped_variables:
            for start_index in temperature_uses.get(variable, ()):
                next_start = next((index for index in all_use_indices if index > start_index), len(rows))
                for row_index in range(start_index, min(next_start, len(rows))):
                    skipped.add(row_index)
        return skipped

    def _accuracy_device_ladder_assignments(self, semantic, device: str):
        assignments = []
        for variable in semantic.temperature_variables:
            assignments.extend(
                item
                for item in semantic.assignments_for_variable(variable)
                if item.numeric_value is not None
            )
        assignments = sorted(assignments, key=lambda item: item.row_index)
        groups: list[list[object]] = []
        current: list[object] = []
        previous_index = -99
        for item in assignments:
            if current and item.row_index != previous_index + 1:
                groups.append(current)
                current = []
            current.append(item)
            previous_index = item.row_index
        if current:
            groups.append(current)
        groups = [group for group in groups if len(group) >= 3]
        if not groups:
            return []
        device_prefix = (device or "").upper().split("-", 1)[0]
        if device_prefix == "VH":
            candidates = [group for group in groups if any((item.numeric_value or 0) >= 100 for item in group)]
            if candidates:
                return candidates[0]
        if device_prefix in {"VC", "VA"}:
            candidates = [group for group in groups if any(abs((item.numeric_value or 0) - 85.0) <= 1e-9 for item in group)]
            if candidates:
                return candidates[0]
        return groups[0]

    def _choose_accuracy_target_variable(self, ladder_assignments, target: float, baseline: float | None) -> str:
        clean = [
            item
            for item in ladder_assignments
            if getattr(item, "numeric_value", None) is not None and getattr(item, "variable", "")
        ]
        if not clean:
            return ""
        exact = [item for item in clean if abs(item.numeric_value - target) <= 1e-9]
        if exact:
            return exact[0].variable
        if baseline is not None and target > baseline:
            after_baseline = [item for item in clean if item.numeric_value > baseline]
            if after_baseline:
                return sorted(after_baseline, key=lambda item: (abs(item.numeric_value - target), item.numeric_value))[0].variable
        if baseline is not None and target < baseline:
            before_baseline = [item for item in clean if item.numeric_value < baseline]
            if before_baseline:
                return sorted(before_baseline, key=lambda item: (abs(item.numeric_value - target), -item.numeric_value))[0].variable
        return sorted(clean, key=lambda item: abs(item.numeric_value - target))[0].variable

    def _cm_accuracy_measurement_block_indices(self, semantic, variables: set[str]) -> set[int]:
        indices: set[int] = set()
        if not variables:
            return indices
        for block in semantic.measurement_blocks:
            if block.setpoint.variable not in variables:
                continue
            if not block.ret_times:
                continue
            for row_index in range(block.start_row_index, block.end_row_index + 1):
                indices.add(row_index)
        return indices

    def _accuracy_baseline_preview_rows(
        self,
        source_row_number: str,
        baseline: float,
        target: float,
        duration: float | None,
    ) -> list[tuple[tuple[object, ...], bool]]:
        duration_value = f"{duration:g}" if duration is not None else "30"
        return [
            ((f"{source_row_number}.b0", "Comment", "", f"AI/local contract: baseline pre-equilibration before accuracy target {target:g} C.", "", f"Set baseline {baseline:g} C, wait ready, hold {duration_value} min before target measurement.", "modified"), True),
            ((f"{source_row_number}.b1", "Command", "", "ColumnComp.CC.Temperature.Nominal", f"{baseline:g}", "Inserted baseline setpoint from natural-language intent.", "modified"), True),
            ((f"{source_row_number}.b2", "Command", "", "Wait", "ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady, Run=Continue", "Inserted baseline readiness gate; preserves external thermometer readiness logic.", "modified"), True),
            ((f"{source_row_number}.b3", "Command", "", "Delay", duration_value, "Inserted baseline hold duration; confirm CM Delay unit in live method editor.", "modified"), True),
            ((f"{source_row_number}.b4", "Comment", "", f"AI/local contract: ramp from {baseline:g} C to {target:g} C and keep the following RetTime/report anchor for the target block.", "", "", "modified"), True),
        ]

    def _append_preview_note(self, existing: object, note: str) -> str:
        text = str(existing or "").strip()
        if not text:
            return note
        if note in text:
            return text
        return f"{text} | {note}"

    def _replace_known_temperature_value(self, value_text: str, replacements: dict[float, str]) -> tuple[str, bool]:
        text = value_text or ""
        numeric_match = re.search(r"(-?\d+(?:\.\d+)?)", text)
        current = cm_numeric_value(text)
        if not numeric_match or current is None:
            return value_text, False
        for source_value, replacement in replacements.items():
            if abs(current - source_value) <= 1e-9:
                new_text = text[: numeric_match.start(1)] + replacement + text[numeric_match.end(1) :]
                return new_text, new_text != text
        return value_text, False

    def _heatup_cooldown_preview_rows(self, record: FoqAlignmentRecord, device: str) -> list[tuple[tuple[object, ...], bool]]:
        values = self._test_plan_numbers_from_text(self.test_plan_parameter_var.get())
        if len(values) >= 3:
            start, target, return_temp = values[-3], values[-2], values[-1]
        else:
            start, target, return_temp = 20.0, 50.0, 20.0
        changed = (abs(start - 20.0) > 1e-9) or (abs(target - 50.0) > 1e-9) or (abs(return_temp - 20.0) > 1e-9)
        return [
            (("", "Stage", "{Initial Time}", record.instrument_method or "TEMP_HEAT_UP_DOWN_20_50_20", "", f"Device: {device or record.device_label}", "cm_initial"), False),
            (("1", "Command", "", "ColumnComp.CC.Temperature.Nominal", f"{start:g} C", "Replace source precondition/start setpoint.", ""), changed),
            (("2", "Command", "", "ColumnComp.CC.Ready/Stability", "True", "Keep readiness/stability gate from source method.", ""), False),
            (("3", "Command", "", "RetTimes.RetTime1", "", "Heat-up start anchor before upward transition.", ""), False),
            (("4", "Command", "", "ColumnComp.CC.Temperature.Nominal", f"{target:g} C", "Replace source heat-up target setpoint.", ""), changed),
            (("5", "Command", "", "ColumnComp.CC.Ready/Stability", "True", "Keep source timeout and stable-window logic unless reviewed.", ""), False),
            (("6", "Command", "", "RetTimes.RetTime3", "", "Heat-up end/stable anchor used by report.", ""), False),
            (("7", "Command", "", "RetTimes.RetTime4", "", "Cool-down start anchor.", ""), False),
            (("8", "Command", "", "ColumnComp.CC.Temperature.Nominal", f"{return_temp:g} C", "Replace source cool-down return setpoint.", ""), changed),
            (("9", "Command", "", "ColumnComp.CC.Ready/Stability", "True", "Keep source timeout and stable-window logic unless reviewed.", ""), False),
            (("10", "Command", "", "RetTimes.RetTime6", "", "Cool-down end/stable anchor used by report.", ""), False),
            (("", "Comment", "", "HeatUp&CoolDown row 66 / D26-D27", f"{start:g}->{target:g}->{return_temp:g}", "Report labels/formulas must be reviewed later; report template is not modified in this step.", "modified"), changed),
        ]

    def _test_plan_setpoint_from_parameter(self, parameter: str) -> float | None:
        match = re.search(r"(-?\d+(?:\.\d+)?)", parameter or "")
        if not match:
            return None
        try:
            return float(match.group(1))
        except ValueError:
            return None

    def _single_point_accuracy_preview_rows(self, project) -> list[tuple[object, ...]]:
        script = instrument_method_script_text(project)
        rows: list[tuple[object, ...]] = []
        current_stage = ""
        step = 1
        for raw in script.splitlines():
            line = raw.strip()
            if not line:
                continue
            if line.startswith("[") and line.endswith("]"):
                current_stage = line.strip("[]")
                rows.append(("", "Stage", "{Initial Time}" if not rows else "", current_stage, "", "", "cm_initial"))
                continue
            if line.startswith("#"):
                rows.append(("", "Comment", "", line[1:].strip(), "", "", "cm_comment"))
                continue
            command, target, value = self._parse_method_preview_command(line)
            rows.append((str(step), "Command", "", target or command, value, "Generated modified preview.", ""))
            step += 1
        return rows

    def _parse_method_preview_command(self, line: str) -> tuple[str, str, str]:
        if line.startswith("SET "):
            body = line[4:]
            if " = " in body:
                target, value = body.split(" = ", 1)
                return "SET", target, value
            return "SET", body, ""
        if line.startswith("RUN "):
            return "RUN", line[4:], ""
        if line.startswith("IF "):
            return "IF", line[3:], ""
        if line in {"ELSE", "END IF"}:
            return line, "", ""
        return "CMD", line, ""

    def _method_preview_row_is_changed(self, row: tuple[object, ...], setpoint: float) -> bool:
        text = " ".join(str(item) for item in row).lower()
        token = f"{setpoint:g}".lower()
        return (
            token in text
            or "single" in text
            or "removed" in text
            or "rettimes." in text
            or "baseline" in text
            or "nominal" in text
        )

    def _populate_test_plan_candidate_table(self, record: FoqAlignmentRecord, device: str) -> None:
        table_frame = getattr(self, "test_plan_candidate_table", None)
        if table_frame is None:
            return
        table = self._table_widget(table_frame)
        sheets = record.report_sheet_label or "<not mapped>"
        db_fields = ", ".join(record.db_fields[:6]) + (f", +{len(record.db_fields) - 6}" if len(record.db_fields) > 6 else "")
        rows = (
            ("Device / Intent", device or record.device_label, record.td_test, "from selected intent", "config checklist"),
            ("Instrument Method", record.instrument_method or "<missing>", record.injection or "CMBX injection binding", self._test_plan_confidence(record.instrument_method), "method script Excel"),
            ("Report Template", record.report_template or "<missing>", sheets, self._test_plan_confidence(record.report_template), "report template/calculation Excel"),
            ("DB / Formula Contract", db_fields or "<not mapped>", "; ".join(record.db_evidence[:2]) or "FOQ mapping", self._test_plan_confidence(db_fields), "report formula map / DB trace"),
        )
        for index, row in enumerate(rows, start=1):
            table.insert("", "end", iid=f"candidate:{index}", values=row)

    def _populate_test_plan_config_table(self, record: FoqAlignmentRecord, rows: tuple[tuple[str, str, str, str, str], ...], device: str) -> None:
        table_frame = getattr(self, "test_plan_config_table", None)
        if table_frame is None:
            return
        table = self._table_widget(table_frame)
        index = 1
        for item in record.required_config:
            table.insert("", "end", iid=f"config:{index}", values=(item, "KB required_config", "required", "Confirm in CM Instrument Setup before running."))
            index += 1
        for row in rows:
            asset = str(row[0]).lower()
            if "config" in asset or "validation" in asset:
                table.insert("", "end", iid=f"config:{index}", values=(row[2], row[1], "review", row[3]))
                index += 1
        for gap in record.open_gaps:
            table.insert("", "end", iid=f"gap:{index}", values=(gap, "Open Verification", "open", "Close with CMBX evidence or manual CM confirmation."))
            index += 1
        if index == 1:
            table.insert("", "end", iid="config:none", values=(f"{device or record.device_label} has no explicit open config gap in KB.", "KB/CMBX evidence", "ok", "Still review generated method/report before running."))

    def _test_plan_table_for_asset(self, asset: str) -> ttk.Treeview:
        lowered = asset.lower()
        if "report" in lowered or "formula" in lowered or "db" in lowered:
            frame = getattr(self, "test_plan_report_table", None)
        elif "config" in lowered or "validation" in lowered:
            frame = getattr(self, "test_plan_config_table", None)
        else:
            frame = getattr(self, "test_plan_method_table", None)
        if frame is None:
            frame = getattr(self, "test_plan_method_table")
        return self._table_widget(frame)

    def _test_plan_confidence(self, value: str) -> str:
        text = (value or "").strip().lower()
        if not text or "missing" in text or "open verification" in text or "not mapped" in text:
            return "needs evidence"
        if "device-specific" in text:
            return "device-specific"
        return "high"

    def _set_test_plan_preview(self, content: str) -> None:
        text_widget = getattr(self, "test_plan_preview_text", None)
        if text_widget is None:
            return
        text_widget.configure(state="normal")
        text_widget.delete("1.0", tk.END)
        text_widget.insert(tk.END, content)
        text_widget.tag_configure("tp_change", background="#FEF3C7", foreground="#92400E")
        text_widget.tag_configure("tp_locked", background="#E0F2FE", foreground="#075985")
        text_widget.tag_configure("tp_open", background="#FEE2E2", foreground="#991B1B")
        text_widget.tag_configure("tp_heading", foreground="#111827", font=self._font(9, "bold"))
        for line_no, line in enumerate(content.splitlines(), start=1):
            lower = line.lower()
            tag = ""
            if "[change]" in lower:
                tag = "tp_change"
            elif "[locked]" in lower:
                tag = "tp_locked"
            elif "[open]" in lower or "review before diff" in lower:
                tag = "tp_open"
            elif line.startswith("#"):
                tag = "tp_heading"
            if tag:
                text_widget.tag_add(tag, f"{line_no}.0", f"{line_no}.end")

    def export_test_plan_draft_packet(self) -> None:
        record = self.test_plan_selected_record or self._selected_test_plan_record()
        if record is None:
            messagebox.showwarning(APP_NAME, "Select or enter an intent before exporting a test plan draft.")
            return
        try:
            device = self.test_plan_device_var.get().strip()
            if not device and len(record.device_models) == 1:
                device = record.device_models[0]
            if not device or "," in device:
                raise ValueError("Select exactly one device model before exporting a test plan draft.")
            intent = self.test_plan_intent_var.get().strip() or "Search / Recommend"
            parameter = self.test_plan_parameter_var.get().strip()
            candidates = filter_alignment_records(self._test_plan_records(), family=self.test_plan_family_var.get().strip() or "TCC")
            output_root = (
                Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER)
                / "test_plan"
                / "draft_packets"
            )
            paths = write_intent_draft_asset_packet(
                output_root,
                record,
                intent,
                parameter,
                device_model=device,
                selected_records=(record,),
                candidate_records=candidates,
            )
            self.status_var.set(f"Exported test plan draft: {paths['project_dir']}")
            output_lines = self._test_plan_output_summary_lines(paths)
            messagebox.showinfo(
                APP_NAME,
                "Exported test plan draft:\n\n"
                f"{paths['project_dir']}\n\n"
                + "\n".join(output_lines)
                + "\n\nReview Config -> Method -> Report before treating any output as runnable.",
            )
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def _test_plan_output_summary_lines(self, paths: dict[str, Path]) -> list[str]:
        labels = (
            ("Method script Excel", "method_excel"),
            ("Report calculation Excel", "report_excel"),
            ("Config checklist", "config_method_report_review"),
            ("Sequence template", "sequence_template"),
            ("Config contract", "config_contract"),
            ("Report/DB contract", "report_db_contract"),
            ("Manifest", "manifest"),
        )
        lines: list[str] = []
        for label, key in labels:
            path = paths.get(key)
            if path and Path(path).exists():
                lines.append(f"{label}: {Path(path).name}")
        project_dir = paths.get("project_dir")
        if project_dir:
            for script_name in ("method_script_40C_only.txt", "method_script_40C.txt", "instrument_method_draft.txt"):
                script = Path(project_dir) / script_name
                if script.exists():
                    lines.insert(0, f"CM method script text: {script.name}")
                    break
        return lines or ["Review files were exported."]

    def _refresh_foq_alignment_filters(self, _event=None) -> None:
        family = self.foq_alignment_family_var.get().strip() or "All"
        family_records = filter_alignment_records(self.foq_alignment_records, family=family)
        if hasattr(self, "foq_alignment_test_combo"):
            self.foq_alignment_test_combo.configure(values=("", *test_intent_options(family_records)))
        if hasattr(self, "foq_alignment_device_listbox"):
            listbox = self.foq_alignment_device_listbox
            previous = set(self.foq_alignment_selected_devices)
            listbox.delete(0, tk.END)
            for index, device in enumerate(device_options(family_records)):
                listbox.insert(tk.END, device)
                if device in previous:
                    listbox.selection_set(index)
        self._populate_foq_kb_to_run_table()

    def _selected_foq_alignment_devices(self) -> tuple[str, ...]:
        if not hasattr(self, "foq_alignment_device_listbox"):
            return ()
        listbox = self.foq_alignment_device_listbox
        devices = tuple(listbox.get(index) for index in listbox.curselection())
        self.foq_alignment_selected_devices = set(devices)
        return devices

    def _foq_kb_to_run_rows(self) -> tuple[FoqAlignmentRecord, ...]:
        records = filter_alignment_records(
            self.foq_alignment_records,
            family=self.foq_alignment_family_var.get().strip() or "All",
            devices=self._selected_foq_alignment_devices(),
            test_text=self.foq_alignment_test_var.get(),
        )
        self.foq_alignment_filtered_records = records
        return records

    def _populate_foq_kb_to_run_table(self, _event=None) -> None:
        if not hasattr(self, "foq_kb_to_run_table"):
            return
        table = self._table_widget(self.foq_kb_to_run_table)
        table.delete(*table.get_children())
        for index, row in enumerate(self._foq_kb_to_run_rows(), start=1):
            gate = record_intent_gate(
                row,
                self.foq_alignment_intent_var.get(),
                self.foq_alignment_intent_parameter_var.get(),
                selected_records=(row,),
            )
            open_count = self._foq_alignment_open_count(row)
            action = self._foq_alignment_action_summary(row, gate)
            tag = self._foq_alignment_row_tag(row, gate)
            table.insert(
                "",
                "end",
                iid=str(index),
                values=(
                    row.family,
                    row.td_test,
                    row.device_label,
                    row.coverage_status,
                    open_count,
                    action,
                ),
                tags=(tag,),
            )
        children = table.get_children()
        if children:
            table.selection_set(children[0])
            table.focus(children[0])
            self._preview_selected_foq_kb_to_run_row()
        else:
            self._clear_foq_alignment_detail()

    def _selected_foq_kb_to_run_row(self) -> FoqAlignmentRecord | None:
        rows = self._selected_foq_kb_to_run_rows()
        return rows[0] if rows else None

    def _selected_foq_kb_to_run_rows(self) -> tuple[FoqAlignmentRecord, ...]:
        if not hasattr(self, "foq_kb_to_run_table"):
            return ()
        table = self._table_widget(self.foq_kb_to_run_table)
        selected = table.selection()
        if not selected:
            return ()
        rows = self.foq_alignment_filtered_records or self._foq_kb_to_run_rows()
        result: list[FoqAlignmentRecord] = []
        for item in selected:
            try:
                index = int(item) - 1
            except ValueError:
                continue
            if 0 <= index < len(rows):
                result.append(rows[index])
        return tuple(result)

    def _foq_alignment_open_count(self, row: FoqAlignmentRecord) -> int:
        topics = open_verification_topics_for_record(row) if row.family == "TCC" else ()
        return len(row.open_gaps) + len(topics)

    def _foq_alignment_action_summary(self, row: FoqAlignmentRecord, gate) -> str:
        if gate.blockers:
            first = gate.blockers[0]
            return f"Blocked: {first[:120]}"
        if row.coverage_status == "complete":
            return "Review contract details; candidate after CM validation"
        return "Review open evidence and relationship rules"

    def _foq_alignment_row_tag(self, row: FoqAlignmentRecord, gate) -> str:
        if gate.blockers:
            return "alignment_blocked"
        if row.coverage_status == "complete" and not row.open_gaps:
            return "alignment_ready"
        return "alignment_review"

    def _select_foq_alignment_detail_tab(self, title: str) -> None:
        if not hasattr(self, "foq_alignment_detail_notebook"):
            return
        text_widget = getattr(self, "foq_alignment_detail_texts", {}).get(title)
        if text_widget is not None:
            self.foq_alignment_detail_notebook.select(text_widget.master)

    def _set_foq_alignment_text(self, title: str, content: str) -> None:
        text_widget = getattr(self, "foq_alignment_detail_texts", {}).get(title)
        if text_widget is None:
            return
        text_widget.configure(state="normal")
        text_widget.delete("1.0", tk.END)
        text_widget.insert(tk.END, content)
        text_widget.configure(state="disabled")

    def _render_foq_alignment_summary(self, row: FoqAlignmentRecord) -> None:
        selected_devices = self._selected_foq_alignment_devices()
        device_label = ", ".join(selected_devices) if selected_devices else row.device_label
        open_count = self._foq_alignment_open_count(row)
        coverage = tcc_black_box_coverage_for_record(row) if row.family == "TCC" else None
        six_contract_status = "六契约完整" if coverage and coverage.status.startswith("documented") else "契约待补全"
        title = f"当前验证: {row.td_test}    设备: {device_label}"
        status = (
            f"拆解状态: {self._status_icon(open_count == 0 and row.coverage_status == 'complete')} {six_contract_status}    "
            f"Coverage: {row.coverage_status}    "
            f"Open Verification: {open_count} 未确认    "
            f"最后更新: {datetime.now().strftime('%m-%d')}"
        )
        if hasattr(self, "foq_alignment_summary_title"):
            self.foq_alignment_summary_title.configure(text=title)
        if hasattr(self, "foq_alignment_summary_status"):
            self.foq_alignment_summary_status.configure(text=status)

    def _render_foq_alignment_contract_cards(self, row: FoqAlignmentRecord) -> None:
        coverage = tcc_black_box_coverage_for_record(row) if row.family == "TCC" else None
        topics = open_verification_topics_for_record(row) if row.family == "TCC" else ()
        topic_categories = {topic.category for topic in topics}
        cards = {
            "Method": self._contract_card_status(
                bool(row.instrument_method) and (coverage.contract_1_method if coverage else True),
                f"{row.instrument_method or '(not bound)'} | RetTimes {len(row.expected_ret_times)}",
                issue="method evidence missing",
            ),
            "Processing": self._contract_card_status(
                bool(row.processing_method) and (coverage.contract_2_processing if coverage else True),
                f"{row.processing_method or '(not bound)'}",
                issue="processing not decoded",
                warn=("Processing Method" in topic_categories or row.processing_method.upper().startswith(("CORRECT", "ACCURACY_IRC"))),
            ),
            "Report": self._contract_card_status(
                bool(row.report_sheets) and (coverage.contract_3_report if coverage else True),
                f"{row.report_sheet_label or '(not mapped)'}",
                issue="report formula missing",
                warn=("Report Formula" in topic_categories),
            ),
            "DB": self._contract_card_status(
                bool(row.db_fields) and (coverage.contract_4_db if coverage else True),
                f"{len(row.db_fields)} fields",
                issue="DB contract missing",
            ),
            "Config": self._contract_card_status(
                bool(row.required_config) and (coverage.contract_5_config if coverage else True),
                f"{len(row.required_config)} dependencies",
                issue="config evidence missing",
                warn=("Config Requirement" in topic_categories),
            ),
            "Open": self._contract_card_status(
                self._foq_alignment_open_count(row) == 0,
                f"{self._foq_alignment_open_count(row)} item(s)",
                issue="needs evidence",
            ),
        }
        for title, (status, detail, color) in cards.items():
            card = getattr(self, "foq_alignment_contract_cards", {}).get(title)
            if not card:
                continue
            frame = card["frame"]
            status_label = card["status"]
            detail_label = card["detail"]
            title_label = card["title"]
            if isinstance(frame, tk.Frame):
                frame.configure(bg=color, highlightbackground=self.colors["card_border"])
            for widget in (title_label, status_label, detail_label):
                if isinstance(widget, tk.Label):
                    widget.configure(bg=color)
            if isinstance(status_label, tk.Label):
                status_label.configure(text=status, fg=self.colors["text"])
            if isinstance(detail_label, tk.Label):
                detail_label.configure(text=detail)

    def _contract_card_status(self, ok: bool, detail: str, issue: str, warn: bool = False) -> tuple[str, str, str]:
        if ok and not warn:
            return "✅ 完整", detail, "#DCFCE7"
        if ok and warn:
            return "🟡 需评审", detail, "#FEF3C7"
        return "🔴 缺口", issue, "#FEE2E2"

    def _status_icon(self, ok: bool) -> str:
        return "✅" if ok else "🟡"

    def _render_foq_alignment_workbench_details(self, row: FoqAlignmentRecord) -> None:
        sections = record_detail_sections(row)
        overview = "\n\n".join(
            (
                "## Test Overview",
                sections.get("TKN Node", ""),
                sections.get("Cross-KB Mapping", ""),
                sections.get("TD Meaning", ""),
                sections.get("Generation Readiness", ""),
            )
        )
        method = "\n\n".join(
            (
                "## Method Command Contract",
                sections.get("Method Evidence", ""),
                "## Dependency Impact",
                sections.get("Dependency Impact", ""),
            )
        )
        processing = "\n\n".join(
            (
                "## Processing Method Contract",
                sections.get("M2 Processing Targets", ""),
                "## Processing / Closure Queue",
                sections.get("M2 Closure Tasks", ""),
                "## Open Topics",
                sections.get("Open Verification Topics", ""),
            )
        )
        report_db = "\n\n".join(
            (
                "## Report Formula Contract",
                sections.get("Report Evidence", ""),
                sections.get("M2 Report Targets", ""),
                "## DB Contract",
                sections.get("DB Evidence", ""),
            )
        )
        config_open = "\n\n".join(
            (
                "## Design Actions",
                sections.get("Design Actions", ""),
                "## Resolution Choices",
                sections.get("Resolution Choices", ""),
                "## Open Verification",
                sections.get("Open Verification", ""),
                "## Relationship Audit",
                sections.get("Relationship Audit", ""),
            )
        )
        self._set_foq_alignment_text("Overview", overview)
        self._set_foq_alignment_text("Method", method)
        self._set_foq_alignment_text("Processing", processing)
        self._set_foq_alignment_text("Report / DB", report_db)
        self._set_foq_alignment_text("Config / Open", config_open)

    def _preview_selected_foq_kb_to_run_row(self, _event=None) -> None:
        row = self._selected_foq_kb_to_run_row()
        if row is None:
            self._clear_foq_alignment_detail()
            return
        if not hasattr(self, "foq_alignment_detail_texts"):
            return
        self._render_foq_alignment_summary(row)
        self._render_foq_alignment_contract_cards(row)
        self._render_foq_alignment_workbench_details(row)
        self._render_foq_alignment_test_plan(select_tab=False)
        self._render_foq_alignment_intent_preview(select_tab=False)
        self._render_foq_alignment_intent_conflict(select_tab=False)

    def preview_foq_alignment_intent(self, _event=None) -> None:
        self._render_foq_alignment_test_plan(select_tab=True)
        self._render_foq_alignment_intent_preview(select_tab=False)
        self._render_foq_alignment_intent_conflict(select_tab=False)

    def _refresh_foq_alignment_intent_view(self, _event=None) -> None:
        self._populate_foq_kb_to_run_table()
        self._render_foq_alignment_test_plan(select_tab=True)
        self._render_foq_alignment_intent_preview(select_tab=False)
        self._render_foq_alignment_intent_conflict(select_tab=False)

    def _render_foq_alignment_test_plan(self, select_tab: bool = False) -> None:
        row = self._selected_foq_kb_to_run_row()
        text_widget = getattr(self, "foq_alignment_detail_texts", {}).get("Test Plan")
        if row is None or text_widget is None:
            return
        selected_rows = self._selected_foq_kb_to_run_rows()
        candidates = self.foq_alignment_filtered_records or self._foq_kb_to_run_rows()
        selected_devices = self._selected_foq_alignment_devices()
        device_model = selected_devices[0] if len(selected_devices) == 1 else ""
        content = render_test_plan_assistant_markdown(
            row,
            self.foq_alignment_intent_var.get(),
            self.foq_alignment_intent_parameter_var.get(),
            device_model=device_model,
            selected_records=selected_rows,
            candidate_records=candidates,
        )
        text_widget.configure(state="normal")
        text_widget.delete("1.0", tk.END)
        text_widget.insert(tk.END, content)
        text_widget.configure(state="disabled")
        if select_tab and hasattr(self, "foq_alignment_detail_notebook"):
            self.foq_alignment_detail_notebook.select(text_widget.master)

    def _render_foq_alignment_intent_preview(self, select_tab: bool = False) -> None:
        row = self._selected_foq_kb_to_run_row()
        text_widget = getattr(self, "foq_alignment_detail_texts", {}).get("Intent")
        if row is None or text_widget is None:
            return
        selected_rows = self._selected_foq_kb_to_run_rows()
        candidates = self.foq_alignment_filtered_records or self._foq_kb_to_run_rows()
        preview = record_intent_preview(
            row,
            self.foq_alignment_intent_var.get(),
            self.foq_alignment_intent_parameter_var.get(),
            selected_records=selected_rows,
            candidate_records=candidates,
        )
        conflict = render_intent_conflict_matrix_markdown(
            selected_rows or (row,),
            device_model=(self._selected_foq_alignment_devices()[0] if len(self._selected_foq_alignment_devices()) == 1 else ""),
            intent=self.foq_alignment_intent_var.get(),
        )
        content = "\n\n".join(("Intent Preview", preview, "Intent Conflict / Impact Matrix", conflict))
        text_widget.configure(state="normal")
        text_widget.delete("1.0", tk.END)
        text_widget.insert(tk.END, content)
        text_widget.configure(state="disabled")
        if select_tab and hasattr(self, "foq_alignment_detail_notebook"):
            self.foq_alignment_detail_notebook.select(text_widget.master)

    def _render_foq_alignment_intent_conflict(self, select_tab: bool = False) -> None:
        row = self._selected_foq_kb_to_run_row()
        if "Intent Conflict" not in getattr(self, "foq_alignment_detail_texts", {}):
            return
        text_widget = getattr(self, "foq_alignment_detail_texts", {}).get("Intent Conflict")
        if row is None or text_widget is None:
            return
        selected_rows = self._selected_foq_kb_to_run_rows() or (row,)
        selected_devices = self._selected_foq_alignment_devices()
        device_model = selected_devices[0] if len(selected_devices) == 1 else ""
        content = render_intent_conflict_matrix_markdown(
            selected_rows,
            device_model=device_model,
            intent=self.foq_alignment_intent_var.get(),
        )
        text_widget.configure(state="normal")
        text_widget.delete("1.0", tk.END)
        text_widget.insert(tk.END, content)
        text_widget.configure(state="disabled")
        if select_tab and hasattr(self, "foq_alignment_detail_notebook"):
            self.foq_alignment_detail_notebook.select(text_widget.master)

    def export_foq_alignment_intent_review(self) -> None:
        row = self._selected_foq_kb_to_run_row()
        if row is None:
            messagebox.showwarning(APP_NAME, "Select an alignment row before exporting an intent review.")
            return
        try:
            selected_rows = self._selected_foq_kb_to_run_rows()
            candidates = self.foq_alignment_filtered_records or self._foq_kb_to_run_rows()
            intent = self.foq_alignment_intent_var.get().strip() or "Search / Recommend"
            parameter = self.foq_alignment_intent_parameter_var.get().strip()
            stem_parts = [row.family, row.test_intent, intent.replace("/", "_").replace(" ", "_")]
            if parameter:
                stem_parts.append(parameter)
            stem = safe_filename("_".join(stem_parts)) or "FOQ_Intent_Review"
            output = (
                Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER)
                / "foq_knowledge_alignment"
                / "intent_reviews"
                / f"{stem}.md"
            )
            path = write_intent_review_markdown(
                output,
                row,
                intent,
                parameter,
                selected_records=selected_rows,
                candidate_records=candidates,
            )
            self.status_var.set(f"Exported FOQ intent review: {path}")
            messagebox.showinfo(APP_NAME, f"Exported FOQ intent review:\n\n{path}")
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def export_foq_alignment_intent_action_plan(self) -> None:
        row = self._selected_foq_kb_to_run_row()
        if row is None:
            messagebox.showwarning(APP_NAME, "Select an alignment row before exporting an action plan.")
            return
        try:
            selected_rows = self._selected_foq_kb_to_run_rows()
            candidates = self.foq_alignment_filtered_records or self._foq_kb_to_run_rows()
            intent = self.foq_alignment_intent_var.get().strip() or "Search / Recommend"
            parameter = self.foq_alignment_intent_parameter_var.get().strip()
            stem_parts = [row.family, row.test_intent, intent.replace("/", "_").replace(" ", "_"), "Action_Plan"]
            if parameter:
                stem_parts.append(parameter)
            stem = safe_filename("_".join(stem_parts)) or "FOQ_Intent_Action_Plan"
            output = (
                Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER)
                / "foq_knowledge_alignment"
                / "intent_action_plans"
                / f"{stem}.md"
            )
            path = write_intent_action_plan_markdown(
                output,
                row,
                intent,
                parameter,
                selected_records=selected_rows,
                candidate_records=candidates,
            )
            self.status_var.set(f"Exported FOQ intent action plan: {path}")
            messagebox.showinfo(APP_NAME, f"Exported FOQ intent action plan:\n\n{path}")
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def export_foq_alignment_draft_asset_packet(self) -> None:
        row = self._selected_foq_kb_to_run_row()
        if row is None:
            messagebox.showwarning(APP_NAME, "Select an alignment row before exporting a draft asset packet.")
            return
        try:
            selected_devices = self._selected_foq_alignment_devices()
            if len(selected_devices) > 1:
                raise ValueError("Select exactly one device model before exporting a draft asset packet.")
            device_model = selected_devices[0] if selected_devices else ""
            selected_rows = self._selected_foq_kb_to_run_rows()
            candidates = self.foq_alignment_filtered_records or self._foq_kb_to_run_rows()
            intent = self.foq_alignment_intent_var.get().strip() or "Search / Recommend"
            parameter = self.foq_alignment_intent_parameter_var.get().strip()
            output_root = (
                Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER)
                / "foq_knowledge_alignment"
                / "draft_asset_packets"
            )
            paths = write_intent_draft_asset_packet(
                output_root,
                row,
                intent,
                parameter,
                device_model=device_model,
                selected_records=selected_rows,
                candidate_records=candidates,
            )
            self.status_var.set(f"Exported FOQ draft asset packet: {paths['project_dir']}")
            messagebox.showinfo(
                APP_NAME,
                "Exported FOQ draft asset packet:\n\n"
                f"{paths['project_dir']}\n\n"
                "This is a reviewable asset packet, not a runnable CMBX.",
            )
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def _clear_foq_alignment_detail(self) -> None:
        for text_widget in getattr(self, "foq_alignment_detail_texts", {}).values():
            text_widget.configure(state="normal")
            text_widget.delete("1.0", tk.END)
            text_widget.insert(tk.END, "No alignment record selected.")
            text_widget.configure(state="disabled")
        if hasattr(self, "foq_alignment_summary_title"):
            self.foq_alignment_summary_title.configure(text="当前验证: 请选择一个测试")
        if hasattr(self, "foq_alignment_summary_status"):
            self.foq_alignment_summary_status.configure(text="拆解状态: -    Open Verification: -    最后更新: -")
        for card in getattr(self, "foq_alignment_contract_cards", {}).values():
            frame = card["frame"]
            status = card["status"]
            detail = card["detail"]
            if isinstance(frame, tk.Frame):
                frame.configure(bg="#F8FAFC", highlightbackground=self.colors["card_border"])
            if isinstance(status, tk.Label):
                status.configure(text="-", bg="#F8FAFC", fg=self.colors["text"])
            if isinstance(detail, tk.Label):
                detail.configure(text="-", bg="#F8FAFC")

    def export_foq_kb_to_run_alignment(self) -> None:
        try:
            rows = self._foq_kb_to_run_rows()
            output = Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER) / "foq_knowledge_alignment" / "FOQ_Knowledge_Alignment.xlsx"
            path = write_foq_alignment_workbook(
                rows,
                output,
                intent=self.foq_alignment_intent_var.get(),
                parameter=self.foq_alignment_intent_parameter_var.get(),
            )
            self.status_var.set(f"Exported FOQ knowledge alignment: {path}")
            messagebox.showinfo(APP_NAME, f"Exported FOQ knowledge alignment:\n\n{path}")
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def _build_generator_project(self):
        if self.generator_test_var.get().strip() != "Temperature Accuracy":
            raise ValueError("Only Temperature Accuracy generation is implemented in this first generator version.")
        device = self.generator_device_var.get().strip() or "VH-C10-A"
        if not self.generator_setpoint_var.get().strip():
            raise ValueError("Enter the requested accuracy setpoint C. Example: 40 or 120.")
        try:
            setpoint = float(self.generator_setpoint_var.get().strip())
        except ValueError as exc:
            raise ValueError("Setpoint C must be a number, for example 40.") from exc
        baseline_text = self.generator_baseline_var.get().strip()
        baseline = None
        if baseline_text:
            try:
                baseline = float(baseline_text)
            except ValueError as exc:
                raise ValueError("Baseline C must be a number or blank if still unknown.") from exc
        return build_single_point_temperature_accuracy_project(device, setpoint, baseline_c=baseline)

    def _show_generator_placeholder(self) -> None:
        method_text = (
            "FOQ KB to Run is now the primary design workbench.\n\n"
            "Enter:\n"
            "- Device, for example VH-C10-A\n"
            "- Setpoint C, for example 40 or 120\n"
            "- Baseline C, for example 20 or 80, only when TD/method design confirms it\n\n"
            "Then click Preview. The exported Excel is a candidate design package, not a signed Chromeleon binary method."
        )
        report_text = (
            "Report calculation preview will show:\n\n"
            "- RetTime selected for the requested setpoint\n"
            "- chm.sig_value windows\n"
            "- observed/deviation workbook rules\n"
            "- DB field and report cell mapping\n\n"
            "The generator will not guess whether 120 C should start from 80 C or 20 C."
        )
        self.generator_method_preview.configure(state="normal")
        self.generator_method_preview.delete("1.0", tk.END)
        self.generator_method_preview.insert(tk.END, method_text)
        self.generator_method_preview.configure(state="disabled")
        self.generator_report_preview.configure(state="normal")
        self.generator_report_preview.delete("1.0", tk.END)
        self.generator_report_preview.insert(tk.END, report_text)
        self.generator_report_preview.configure(state="disabled")

    def preview_method_report_generation(self) -> None:
        try:
            project = self._build_generator_project()
            method_text = instrument_method_script_text(project)
            report_text = report_calculation_spec_text(project)
            spec = single_point_temperature_accuracy_project_to_dict(project)
            header = [
                f"Generated object: {project.instrument_method}",
                f"Device: {project.device_model}",
                f"Intent: {project.test_intent}",
                f"Setpoint: {_baseline_label_for_app(project)} baseline -> {project.setpoint_c:g} C report point",
                f"Report anchor: {project.report_sheet} / {_generator_report_anchor(spec)}",
                "",
            ]
            self.generator_method_preview.configure(state="normal")
            self.generator_method_preview.delete("1.0", tk.END)
            self.generator_method_preview.insert(tk.END, "\n".join(header) + method_text)
            self.generator_method_preview.configure(state="disabled")

            self.generator_report_preview.configure(state="normal")
            self.generator_report_preview.delete("1.0", tk.END)
            self.generator_report_preview.insert(tk.END, report_text)
            self.generator_report_preview.configure(state="disabled")
            self.status_var.set(f"Generated preview for {project.device_model} {project.db_field}")
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def export_method_report_generation_excel(self) -> None:
        try:
            project = self._build_generator_project()
            output_root = self._generated_project_output_root()
            paths = write_single_point_temperature_accuracy_excel_workbooks(project, output_root)
            self.status_var.set(f"Generated method/report Excel: {paths['method_excel'].parent}")
            messagebox.showinfo(
                APP_NAME,
                "Generated Excel workbooks:\n\n"
                + "\n".join(str(path) for path in paths.values()),
            )
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def export_method_report_generation_project(self) -> None:
        try:
            project = self._build_generator_project()
            output_root = self._generated_project_output_root()
            project_dir = write_single_point_temperature_accuracy_project(project, output_root)
            paths = write_single_point_temperature_accuracy_excel_workbooks(project, output_root)
            self.status_var.set(f"Generated full method/report project: {project_dir}")
            messagebox.showinfo(
                APP_NAME,
                "Generated project folder:\n\n"
                f"{project_dir}\n\n"
                "Excel workbooks:\n"
                + "\n".join(str(path) for path in paths.values()),
            )
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))

    def browse_foq_mapping(self) -> None:
        initialdir = DEFAULT_FOQ_MAPPING_FOLDER if DEFAULT_FOQ_MAPPING_FOLDER.exists() else (Path(self.foq_mapping_path_var.get()).parent if self.foq_mapping_path_var.get().strip() else Path.cwd())
        path = filedialog.askopenfilename(
            title="Select FOQResultLocations mapping",
            initialdir=str(initialdir),
            filetypes=[("Excel mapping", "*.xls *.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.foq_mapping_path_var.set(path)
            self._sync_path_display(self.foq_mapping_path_var, self.foq_mapping_display_var, "foq")
            self._refresh_foq_mapping_filter_options()

    def load_package(self) -> None:
        self._commit_path_display(self.cmbx_path_var, self.cmbx_display_var, "cmbx")
        self._load_packages_from_source_text(self.cmbx_path_var.get(), empty_status="No CMBX packages found; showing folder tree")

    def _load_packages_from_source_text(self, source_text: str, *, empty_status: str) -> None:
        self.status_var.set("Scanning CMBX packages...")
        self.progress_var.set(0.0)
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths = self._cmbx_paths_from_text(source_text)
                missing = [path for path in paths if not path.exists()]
                if missing:
                    self._call_ui(lambda missing=missing: self._load_packages_failed(ValueError("CMBX path not found:\n" + "\n".join(str(path) for path in missing[:8]))))
                    return
                packages_by_index: dict[int, CmbxPackage] = {}
                total = max(len(paths), 1)
                if not paths:
                    self._call_ui(lambda: self._load_packages_done([]))
                    return
                worker_count = min(8, max(1, os.cpu_count() or 1), len(paths))
                self._thread_status(f"__PROGRESS__=2.0|Loading {len(paths)} CMBX package(s) with {worker_count} worker(s)")
                load_failures: list[tuple[Path, str]] = []

                def load_one(load_index: int, load_path: Path) -> CmbxPackage:
                    self._thread_status(f"__PROGRESS__=2.0|Reading CMBX {load_index + 1}/{len(paths)}: {load_path.name}")
                    _write_startup_log(f"scan reading: {load_path}")
                    return load_cmbx_package(load_path)

                with ThreadPoolExecutor(max_workers=worker_count) as executor:
                    futures = {executor.submit(load_one, index, path): (index, path) for index, path in enumerate(paths)}
                    for done_count, future in enumerate(as_completed(futures), 1):
                        index, path = futures[future]
                        try:
                            packages_by_index[index] = future.result()
                        except Exception as exc:
                            load_failures.append((path, str(exc)))
                            _write_startup_log(f"scan failed: {path}: {exc}")
                            self._thread_status(f"__PROGRESS__={done_count / total * 95:.1f}|Skipped CMBX {done_count}/{len(paths)}: {path.name}")
                            continue
                        self._thread_status(f"__PROGRESS__={done_count / total * 95:.1f}|Loaded CMBX {done_count}/{len(paths)}: {path.name}")
                self._thread_status("__PROGRESS__=98.0|Refreshing package tree")
                packages = [packages_by_index[index] for index in range(len(paths)) if index in packages_by_index]
                self._call_ui(lambda packages=packages, failures=load_failures: self._load_packages_done(packages, failures, empty_status))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._load_packages_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _load_packages_done(self, packages: list[CmbxPackage], failures: list[tuple[Path, str]] | None = None, empty_status: str | None = None) -> None:
        self._set_buttons_state("normal")
        failures = failures or []
        self.loaded_packages = packages
        self.starred_signal_keys.clear()
        self.foq_device_cache.clear()
        self._set_workspace_packages(packages)
        if packages:
            self.progress_var.set(100.0)
            suffix = f"; skipped {len(failures)} failed package(s)" if failures else ""
            self.status_var.set(f"Loaded {len(packages)} package(s){suffix}")
            if failures:
                preview = "\n".join(f"{path.name}: {message}" for path, message in failures[:8])
                if len(failures) > 8:
                    preview += f"\n... and {len(failures) - 8} more. See logs."
                messagebox.showwarning(APP_NAME, f"Some CMBX packages could not be loaded and were skipped:\n\n{preview}")
        else:
            self.progress_var.set(0.0)
            self.status_var.set(empty_status or "No CMBX packages found; showing folder tree")

    def _load_packages_failed(self, exc: Exception) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(0.0)
        self.status_var.set("Load failed")
        messagebox.showerror(APP_NAME, str(exc))

    def _auto_scan_default_folder(self) -> None:
        root = discover_default_cmbx_source_folder()
        current = Path(self.cmbx_path_var.get().strip().strip('"')) if self.cmbx_path_var.get().strip() else root
        if current != root and self.loaded_packages:
            return
        self.cmbx_path_var.set(str(root))
        self._sync_path_display(self.cmbx_path_var, self.cmbx_display_var, "cmbx")
        self.status_var.set("Checking default package folder...")

        def worker() -> None:
            try:
                package_count = self._cmbx_path_count_from_text(str(root), limit=AUTO_SCAN_CMBX_LIMIT + 1)
                if package_count > AUTO_SCAN_CMBX_LIMIT:
                    _write_startup_log(f"auto scan skipped: more than {AUTO_SCAN_CMBX_LIMIT} CMBX package(s)")

                    def skipped() -> None:
                        self.progress_var.set(0.0)
                        self.status_var.set(
                            f"Default folder has more than {AUTO_SCAN_CMBX_LIMIT} CMBX package(s); "
                            "showing folder tree only. Select a folder/package or click Scan CMBX to load them."
                        )

                    self._call_ui(skipped)
                    return
                _write_startup_log(f"auto scan starting: {package_count} CMBX package(s)")
                self._call_ui(self.load_package)
            except Exception as exc:
                _write_startup_log(f"auto scan check failed: {exc}")
                self._call_ui(lambda exc=exc: self.status_var.set(f"Default folder auto-scan check failed: {exc}"))

        threading.Thread(target=worker, daemon=True).start()

    def _cmbx_path_count_from_text(self, text: str, limit: int | None = None) -> int:
        cleaned = text.strip()
        if not cleaned:
            return 0
        pieces = [piece.strip().strip('"') for piece in cleaned.split(";") if piece.strip()]
        count = 0
        seen: set[str] = set()
        for piece in pieces:
            path = Path(piece)
            candidates = path.rglob("*.cmbx") if path.is_dir() else (candidate for candidate in [path])
            for candidate in candidates:
                if not candidate.is_file() or self._path_is_in_deleted_folder(candidate):
                    continue
                key = str(candidate.resolve()) if candidate.exists() else str(candidate)
                if key in seen:
                    continue
                seen.add(key)
                count += 1
                if limit is not None and count >= limit:
                    return count
        return count

    def _cmbx_paths_from_text(self, text: str) -> list[Path]:
        cleaned = text.strip()
        if not cleaned:
            return []
        pieces = [piece.strip().strip('"') for piece in cleaned.split(";") if piece.strip()]
        paths: list[Path] = []
        seen: set[str] = set()
        for piece in pieces:
            path = Path(piece)
            candidates = sorted(candidate for candidate in path.rglob("*.cmbx") if candidate.is_file()) if path.is_dir() else [path]
            for candidate in candidates:
                if self._path_is_in_deleted_folder(candidate):
                    continue
                key = str(candidate.resolve()) if candidate.exists() else str(candidate)
                if key not in seen:
                    seen.add(key)
                    paths.append(candidate)
        return paths

    def _path_is_in_deleted_folder(self, path: Path) -> bool:
        return any(part.lower() == "deleted" for part in path.parts)

    def _package_tree_root(self) -> Path:
        cleaned = self.cmbx_path_var.get().strip()
        pieces = [piece.strip().strip('"') for piece in cleaned.split(";") if piece.strip()]
        if len(pieces) == 1:
            path = Path(pieces[0])
            if path.is_dir():
                return path
            if path.is_file():
                return path.parent
        return DEFAULT_CMBX_SOURCE_FOLDER

    def _set_workspace_packages(self, packages: list[CmbxPackage]) -> None:
        self.package = packages[0] if packages else None
        self.current_injection = None
        self.current_sequence = None
        self._load_context_for_package(self.package)
        self._prune_foq_candidates()
        self._populate_tree()
        self._clear_context_tables()
        self._show_workspace_summary()

    def _load_context_for_package(self, package: CmbxPackage | None) -> None:
        if not package:
            self.injection_method_links = {}
            self.external_instrument_methods = {}
            return
        self.package = package
        self.injection_method_links = build_injection_method_links(package)
        method_names = [method.name for method in package.methods_and_reports if method.kind == "instrument_method"]
        self.external_instrument_methods = discover_external_instrument_methods(package.path, method_names)
        self.report_xml_cache.clear()
        self.report_sheet_cache.clear()
        self.report_preview_cache.clear()

    def _enable_cmbx_drop_target(self) -> None:
        try:
            self.root.tk.call("package", "require", "tkdnd")
            self.root.tk.call("tkdnd::drop_target", "register", self.root, "DND_Files")
            self.root.bind("<<Drop>>", self._handle_cmbx_drop)
        except tk.TclError:
            return

    def _handle_cmbx_drop(self, event) -> None:
        paths = [path for path in self._split_dropped_paths(getattr(event, "data", "")) if path.lower().endswith(".cmbx") or Path(path).is_dir()]
        if not paths:
            return
        self.cmbx_path_var.set("; ".join(paths))
        self._sync_path_display(self.cmbx_path_var, self.cmbx_display_var, "cmbx")
        self.load_package()

    def _split_dropped_paths(self, data: str) -> list[str]:
        if not data:
            return []
        try:
            return list(self.root.tk.splitlist(data))
        except tk.TclError:
            return [part.strip("{} ") for part in data.split() if part.strip()]

    def _populate_tree(self) -> None:
        restore_open_keys = set(self._tree_restore_open_keys)
        self._tree_restore_open_keys = set()
        self.package_tree.delete(*self.package_tree.get_children())
        self.tree_item_context.clear()
        self.tree_fs_context.clear()
        root = self._package_tree_root()
        root_iid = "fs:root"
        self.tree_fs_context[root_iid] = root
        root_open = True
        self.package_tree.insert("", "end", iid=root_iid, text=str(root), values=("Folder",), tags=("folder",), open=root_open)
        packages_by_path: dict[str, tuple[int, CmbxPackage]] = {}
        for index, package in enumerate(self.loaded_packages):
            packages_by_path[self._path_key(package.path)] = (index, package)
        if root.exists() and root.is_dir():
            self._insert_filesystem_children(root_iid, root, packages_by_path, restore_open_keys)
            return
        for package_index, package in enumerate(self.loaded_packages):
            self._insert_loaded_package(root_iid, package_index, package, restore_open_keys, open_node=len(self.loaded_packages) == 1)

    def _insert_filesystem_children(
        self,
        parent_iid: str,
        folder: Path,
        packages_by_path: dict[str, tuple[int, CmbxPackage]],
        restore_open_keys: set[tuple[str, str]],
    ) -> None:
        try:
            children = sorted(folder.iterdir(), key=self._filesystem_tree_sort_key)
        except OSError:
            return
        for child in children:
            if child.name.startswith("."):
                continue
            if child.is_dir():
                child_iid = f"fs:{len(self.tree_fs_context)}"
                self.tree_fs_context[child_iid] = child
                open_node = self._tree_open_key_for_fs_path(child) in restore_open_keys
                self.package_tree.insert(parent_iid, "end", iid=child_iid, text=child.name, values=("Folder",), tags=("folder",), open=open_node)
                self._insert_filesystem_children(child_iid, child, packages_by_path, restore_open_keys)
                continue
            suffix = child.suffix.lower()
            if suffix == ".cmbx":
                child_iid = f"fs:{len(self.tree_fs_context)}"
                self.tree_fs_context[child_iid] = child
                open_node = self._tree_open_key_for_fs_path(child) in restore_open_keys
                self.package_tree.insert(parent_iid, "end", iid=child_iid, text=child.name, values=("CMBX",), tags=("package",), open=open_node)
                package_info = packages_by_path.get(self._path_key(child))
                if package_info:
                    package_index, package = package_info
                    self.tree_item_context[child_iid] = (package, None)
                    for element in package.root_elements:
                        if element.kind in {"instrument_method", "processing_method", "report_template"}:
                            continue
                        self._insert_element(child_iid, element, package, package_index, restore_open_keys, open_node=False)
                continue

    def _filesystem_tree_sort_key(self, item: Path) -> tuple[int, int, str]:
        is_deleted = item.is_dir() and item.name.lower() == "deleted"
        return (1 if is_deleted else 0, 0 if item.is_dir() else 1, item.name.lower())

    def _insert_loaded_package(
        self,
        parent_iid: str,
        package_index: int,
        package: CmbxPackage,
        restore_open_keys: set[tuple[str, str]],
        open_node: bool = False,
    ) -> str:
        package_iid = f"pkg:{package_index}"
        self.tree_item_context[package_iid] = (package, None)
        self.tree_fs_context[package_iid] = package.path
        package_open = open_node or self._tree_open_key_for_fs_path(package.path) in restore_open_keys
        self.package_tree.insert(parent_iid, "end", iid=package_iid, text=package.path.name, values=("CMBX",), tags=("package",), open=package_open)
        for element in package.root_elements:
            if element.kind in {"instrument_method", "processing_method", "report_template"}:
                continue
            self._insert_element(package_iid, element, package, package_index, restore_open_keys, open_node=open_node)
        return package_iid

    def _insert_element(
        self,
        parent_iid: str,
        element: CmbxElement,
        package: CmbxPackage,
        package_index: int,
        restore_open_keys: set[tuple[str, str]],
        open_node: bool = False,
    ) -> str:
        iid = f"pkg:{package_index}:element:{len(self.tree_item_context)}"
        self.tree_item_context[iid] = (package, element)
        tag = element.kind if element.kind in {"folder", "sequence", "injection", "signal", "audit"} else ""
        element_open = open_node or self._tree_open_key_for_element(package, element) in restore_open_keys
        self.package_tree.insert(parent_iid, "end", iid=iid, text=element.name, values=(element.kind,), tags=(tag,), open=element_open)
        for child in element.children:
            if child.kind in {"instrument_method", "processing_method", "report_template"}:
                continue
            self._insert_element(iid, child, package, package_index, restore_open_keys)
        return iid

    def on_tree_select(self, _event=None) -> None:
        selected = self.package_tree.selection()
        if not selected:
            return
        package, element = self.tree_item_context.get(selected[0], (None, None))
        if not package:
            fs_path = self.tree_fs_context.get(selected[0])
            if fs_path:
                self.current_injection = None
                self.current_sequence = None
                self._clear_context_tables()
                self.summary_label.config(text=f"{fs_path.name or fs_path}   {fs_path}")
                self._set_info(f"Path: {fs_path}\nType: {'Folder' if fs_path.is_dir() else 'File'}")
            return
        if package is not self.package:
            self._load_context_for_package(package)
        if not element:
            self.current_injection = None
            self.current_sequence = None
            self._clear_context_tables()
            self._show_package_summary()
            return
        if element.kind == "sequence":
            self.current_injection = None
            self.current_sequence = element
            self._populate_sequence(element)
            return
        injection = element if element.kind == "injection" else injection_for_element(self.package, element)
        if injection:
            self.current_injection = injection
            self.current_sequence = self.package.elements_by_id.get(injection.parent_id or "") if self.package else None
            self._populate_injection(injection)
        else:
            self.current_injection = None
            self.current_sequence = None
            self._clear_context_tables()
            self._show_element_info(element)

    def _on_tree_click(self, event) -> None:
        region = self.package_tree.identify("region", event.x, event.y)
        column = self.package_tree.identify_column(event.x)
        if region != "cell" or column != "#2":
            return
        iid = self.package_tree.identify_row(event.y)
        package, element = self.tree_item_context.get(iid, (None, None))
        if not package or not element or element.kind != "signal":
            return
        key = self._signal_cache_key_for_package(package, element)
        joined = "|".join(key)
        if joined in self.starred_signal_keys:
            self.starred_signal_keys.remove(joined)
        else:
            self.starred_signal_keys.add(joined)
        values = list(self.package_tree.item(iid, "values"))
        if len(values) >= 2:
            values[1] = self._signal_star_for_package(package, element)
            self.package_tree.item(iid, values=values)
        self.plot_starred_raw_channels()
        return "break"

    def _show_sequence_tree_menu(self, event) -> None:
        iid = self.package_tree.identify_row(event.y)
        if not iid:
            return
        if iid not in self.package_tree.selection():
            self.package_tree.selection_set(iid)
        package, element = self.tree_item_context.get(iid, (None, None))
        menu = tk.Menu(self.root, tearoff=0)
        fs_paths = self._selected_filesystem_paths()
        scannable_paths = self._scannable_tree_paths(fs_paths)
        if scannable_paths:
            count = len(scannable_paths)
            label = "Scan CMBX In This Folder" if count == 1 and scannable_paths[0].is_dir() else f"Scan CMBX In Selected Scope ({count})"
            menu.add_command(label=label, command=lambda paths=scannable_paths: self.scan_tree_paths(paths))
        renamable_path = self._single_renamable_tree_path(fs_paths)
        if renamable_path:
            menu.add_command(label="Rename", command=lambda path=renamable_path: self.rename_tree_path(path))
        deletable_paths = [path for path in fs_paths if self._is_deletable_tree_path(path)]
        if deletable_paths:
            count = len(deletable_paths)
            menu.add_command(label=f"Delete To packages\\deleted ({count})", command=lambda paths=deletable_paths: self.delete_tree_paths(paths))
        if package and element:
            if menu.index("end") is not None:
                menu.add_separator()
            if element.kind in {"sequence", "injection", "folder"}:
                menu.add_command(label="Rename", command=lambda p=package, e=element: self.rename_cmbx_tree_element(p, e))
            elements = self._raw_channel_export_elements_for_tree_element(element)
            if elements:
                if element.kind == "signal":
                    label = "Export This Channel"
                elif element.kind == "injection":
                    label = "Export Injection Channels"
                elif element.kind == "sequence":
                    label = "Export Sequence Channels"
                else:
                    label = "Export Raw Channels"
                menu.add_command(label=label, command=lambda p=package, items=elements: self._export_elements_by_package([(p, items)]))
            selected_signals = self._selected_tree_signal_items()
            if selected_signals:
                count = len(selected_signals)
                menu.add_command(label=f"Add To Raw Plot Selected Channels ({count})", command=lambda items=selected_signals: self.add_tree_channels_to_raw_plot(items))
        sequence_groups = self._selected_sequences_by_package()
        if sequence_groups:
            if menu.index("end") is not None:
                menu.add_separator()
            count = sum(len(items) for _pkg, items in sequence_groups)
            menu.add_command(label=f"Add Selected Sequence(s) To FOQ DB Candidates ({count})", command=lambda groups=sequence_groups: self.add_sequence_groups_to_foq_candidates(groups))
            menu.add_command(label=f"Unpack Selected Sequence(s) To Folder ({count})", command=lambda groups=sequence_groups: self.unpack_selected_sequences(groups))
        if menu.index("end") is None:
            return
        menu.tk_popup(event.x_root, event.y_root)

    def _scannable_tree_paths(self, paths: list[Path]) -> list[Path]:
        scannable: list[Path] = []
        seen: set[str] = set()
        for path in paths:
            if self._path_is_in_deleted_folder(path):
                continue
            is_cmbx = path.is_file() and path.suffix.lower() == ".cmbx"
            if not path.is_dir() and not is_cmbx:
                continue
            key = self._path_key(path)
            if key in seen:
                continue
            seen.add(key)
            scannable.append(path)
        return self._collapse_nested_paths(scannable)

    def scan_tree_paths(self, paths: list[Path]) -> None:
        scan_paths = self._scannable_tree_paths(paths)
        if not scan_paths:
            messagebox.showinfo(APP_NAME, "Select a Sequence Data folder or CMBX file to scan.")
            return
        source_text = ";".join(f'"{path}"' for path in scan_paths)
        self.cmbx_path_var.set(source_text)
        self._sync_path_display(self.cmbx_path_var, self.cmbx_display_var, "cmbx")
        self._tree_restore_open_keys = set()
        self._load_packages_from_source_text(
            source_text,
            empty_status="No CMBX packages found under the selected Sequence Data scope.",
        )

    def _raw_channel_export_elements_for_tree_element(self, element: CmbxElement) -> list[CmbxElement]:
        if element.kind == "signal":
            return [element]
        if element.kind == "injection":
            return [child for child in element.children if child.kind == "signal"]
        if element.kind == "sequence":
            return [signal for injection in element.children if injection.kind == "injection" for signal in injection.children if signal.kind == "signal"]
        if element.kind == "folder":
            return [
                signal
                for sequence in element.children
                if sequence.kind == "sequence"
                for injection in sequence.children
                if injection.kind == "injection"
                for signal in injection.children
                if signal.kind == "signal"
            ]
        return []

    def _selected_tree_signal_items(self) -> list[tuple[CmbxPackage, CmbxElement]]:
        signals: list[tuple[CmbxPackage, CmbxElement]] = []
        seen: set[str] = set()
        for selected_iid in self.package_tree.selection():
            package, element = self.tree_item_context.get(selected_iid, (None, None))
            if package and element and element.kind == "signal":
                key = "|".join(self._signal_cache_key_for_package(package, element))
                if key not in seen:
                    seen.add(key)
                    signals.append((package, element))
        return signals

    def _selected_sequences_by_package(self) -> list[tuple[CmbxPackage, list[CmbxElement]]]:
        grouped: dict[str, tuple[CmbxPackage, list[CmbxElement], set[str]]] = {}
        for iid in self.package_tree.selection():
            package, element = self.tree_item_context.get(iid, (None, None))
            if not package:
                continue
            if element is None:
                sequences = package.sequences
            else:
                sequences = self._sequence_descendants(element)
            if not sequences:
                continue
            key = str(package.path)
            if key not in grouped:
                grouped[key] = (package, [], set())
            _package, items, seen = grouped[key]
            for sequence in sequences:
                sequence_key = sequence.id or sequence.url or sequence.name
                if sequence_key in seen:
                    continue
                seen.add(sequence_key)
                items.append(sequence)
        return [(package, items) for package, items, _seen in grouped.values() if items]

    def _foq_sequence_search_items(self) -> list[tuple[CmbxPackage, CmbxElement]]:
        selected_items: list[tuple[CmbxPackage, CmbxElement]] = []
        seen: set[tuple[str, str]] = set()

        def add_sequence(package: CmbxPackage, sequence: CmbxElement) -> None:
            if self._foq_package_is_deleted(package):
                return
            key = self._foq_candidate_key(package, sequence)
            if key in seen:
                return
            seen.add(key)
            selected_items.append((package, sequence))

        for package, sequences in self._selected_sequences_by_package():
            for sequence in sequences:
                add_sequence(package, sequence)

        selected_paths = [self.tree_fs_context[iid] for iid in self.package_tree.selection() if iid in self.tree_fs_context]
        for fs_path in selected_paths:
            try:
                resolved = fs_path.resolve()
            except OSError:
                continue
            for package in self.loaded_packages:
                try:
                    package_path = package.path.resolve()
                except OSError:
                    continue
                in_scope = package_path == resolved if resolved.is_file() else package_path == resolved or resolved in package_path.parents
                if not in_scope:
                    continue
                for sequence in package.sequences:
                    add_sequence(package, sequence)

        if selected_items:
            return selected_items
        return [
            (package, sequence)
            for package in self.loaded_packages
            if not self._foq_package_is_deleted(package)
            for sequence in package.sequences
        ]

    def _foq_sequence_search_scope_text(self, item_count: int) -> str:
        selected_paths = [self.tree_fs_context[iid] for iid in self.package_tree.selection() if iid in self.tree_fs_context]
        if selected_paths:
            names = ", ".join(path.name or str(path) for path in selected_paths[:2])
            if len(selected_paths) > 2:
                names += f", +{len(selected_paths) - 2}"
            return f"{item_count} sequence(s) under selected Sequence Data scope: {names}"
        if self.package_tree.selection():
            return f"{item_count} selected sequence(s)"
        return f"{item_count} sequence(s) in workspace"

    def _sequence_descendants(self, element: CmbxElement) -> list[CmbxElement]:
        if element.kind == "sequence":
            return [element]
        sequences: list[CmbxElement] = []
        for child in element.children:
            sequences.extend(self._sequence_descendants(child))
        return sequences

    def add_selected_sequences_to_foq_candidates(self) -> None:
        self.add_sequence_groups_to_foq_candidates(self._selected_sequences_by_package())

    def add_sequence_groups_to_foq_candidates(self, grouped: list[tuple[CmbxPackage, list[CmbxElement]]]) -> None:
        items = [(package, sequence) for package, sequences in grouped for sequence in sequences]
        self._add_foq_candidate_sequences_async(items)

    def add_filtered_sequences_to_foq_candidates(self) -> None:
        self._commit_path_display(self.foq_mapping_path_var, self.foq_mapping_display_var, "foq")
        self._refresh_foq_mapping_filter_options()
        device_filters = {device.lower() for device in self._selected_foq_devices()}
        sequences = self._foq_sequence_search_items()
        self.status_var.set(f"Finding FOQ candidate sequences: {self._foq_sequence_search_scope_text(len(sequences))}")
        self.progress_var.set(0.0)
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                records: list[tuple[CmbxPackage, CmbxElement, str, str, str]] = []
                total = max(len(sequences), 1)
                for index, (package, sequence) in enumerate(sequences, 1):
                    self._thread_status(f"__PROGRESS__={(index - 1) / total * 95:.1f}|Checking sequence {index}/{len(sequences)}: {sequence.name}")
                    device, source = self._resolve_foq_device_type(package, sequence)
                    if device_filters and device.lower() not in device_filters:
                        continue
                    records.append((package, sequence, device, source, self._default_foq_report_template_name(sequence)))
                self._call_ui(lambda records=records: self._add_foq_candidate_records(records))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _add_foq_candidate_sequences(self, items: list[tuple[CmbxPackage, CmbxElement]]) -> None:
        if not items:
            messagebox.showinfo(APP_NAME, "No matching sequence was found.")
            return
        seen = {self._foq_candidate_key(package, sequence) for package, sequence, _device, _source, _report in self.foq_candidate_sequences}
        added = 0
        for package, sequence in items:
            key = self._foq_candidate_key(package, sequence)
            if key in seen:
                continue
            seen.add(key)
            device, source = self._resolve_foq_device_type(package, sequence)
            report_template = self._default_foq_report_template_name(sequence)
            self.foq_candidate_sequences.append((package, sequence, device, source, report_template))
            added += 1
            if device != "unresolved":
                self.foq_device_type_var.set(device)
        self._refresh_foq_candidate_table()
        self.status_var.set(f"Added {added} FOQ candidate sequence(s)")

    def _add_foq_candidate_sequences_async(self, items: list[tuple[CmbxPackage, CmbxElement]]) -> None:
        if not items:
            messagebox.showinfo(APP_NAME, "No matching sequence was found.")
            return
        self.status_var.set(f"Adding {len(items)} FOQ candidate sequence(s)...")
        self.progress_var.set(0.0)
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                records: list[tuple[CmbxPackage, CmbxElement, str, str, str]] = []
                total = max(len(items), 1)
                for index, (package, sequence) in enumerate(items, 1):
                    self._thread_status(f"__PROGRESS__={(index - 1) / total * 95:.1f}|Resolving device {index}/{len(items)}: {sequence.name}")
                    device, source = self._resolve_foq_device_type(package, sequence)
                    records.append((package, sequence, device, source, self._default_foq_report_template_name(sequence)))
                self._call_ui(lambda records=records: self._add_foq_candidate_records(records))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _add_foq_candidate_records(self, records: list[tuple[CmbxPackage, CmbxElement, str, str, str]]) -> None:
        self._set_buttons_state("normal")
        if not records:
            self.progress_var.set(0.0)
            self.status_var.set("No matching sequence was found")
            messagebox.showinfo(APP_NAME, "No matching sequence was found.")
            return
        seen = {self._foq_candidate_key(package, sequence) for package, sequence, _device, _source, _report in self.foq_candidate_sequences}
        added = 0
        for package, sequence, device, source, report_template in records:
            key = self._foq_candidate_key(package, sequence)
            if key in seen:
                continue
            seen.add(key)
            self.foq_candidate_sequences.append((package, sequence, device, source, report_template))
            added += 1
            if device != "unresolved":
                self.foq_device_type_var.set(device)
        self.progress_var.set(100.0)
        self._refresh_foq_candidate_table()
        self.status_var.set(f"Added {added} FOQ candidate sequence(s)")

    def _foq_candidate_key(self, package: CmbxPackage, sequence: CmbxElement) -> tuple[str, str]:
        return (self._path_key(package.path), sequence.id or sequence.url or sequence.name)

    def _foq_package_is_deleted(self, package: CmbxPackage) -> bool:
        try:
            root = self._package_tree_root().resolve()
            deleted = (root / "deleted").resolve()
            package_path = package.path.resolve()
        except OSError:
            return False
        return package_path == deleted or deleted in package_path.parents

    def _foq_report_templates_for_sequence(self, sequence: CmbxElement) -> list[CmbxElement]:
        reports = [child for child in sequence.children if child.kind == "report_template" and "ReportDefinition" in child.item_type]
        if reports:
            return reports
        return [child for child in sequence.children if child.kind == "report_template"]

    def _default_foq_report_template_name(self, sequence: CmbxElement) -> str:
        reports = self._foq_report_templates_for_sequence(sequence)
        return reports[0].name if reports else ""

    def _refresh_foq_mapping_filter_options(self) -> None:
        if not hasattr(self, "foq_filter_device_listbox"):
            return
        mapping_path = Path(self.foq_mapping_path_var.get().strip())
        devices: list[str] = []
        if mapping_path.exists():
            try:
                workbook = load_foq_workbook(mapping_path)
                devices = sorted({mapping.device_type for mapping in read_device_type_mappings(workbook).values()})
            except Exception:
                devices = []
        available = set(devices)
        self.foq_selected_devices.intersection_update(available)
        self.foq_filter_device_options = devices
        self._populate_foq_device_filter_listbox()
        self._refresh_foq_db_field_filter_options()

    def _filter_foq_options(self, options: list[str], query: str) -> list[str]:
        tokens = [token for token in query.lower().split() if token]
        if not tokens:
            return options
        return [option for option in options if all(token in option.lower() for token in tokens)]

    def _visible_listbox_values(self, listbox: tk.Listbox) -> list[str]:
        return [str(listbox.get(index)) for index in range(listbox.size())]

    def _populate_foq_device_filter_listbox(self) -> None:
        if not hasattr(self, "foq_filter_device_listbox"):
            return
        self._foq_filter_refreshing = True
        try:
            visible = self._filter_foq_options(self.foq_filter_device_options, self.foq_device_search_var.get())
            self.foq_filter_device_listbox.delete(0, tk.END)
            for device in visible:
                self.foq_filter_device_listbox.insert(tk.END, device)
                if device in self.foq_selected_devices:
                    self.foq_filter_device_listbox.selection_set(tk.END)
        finally:
            self._foq_filter_refreshing = False
        self._update_foq_filter_counts()

    def _sync_foq_device_selection_from_listbox(self) -> None:
        if not hasattr(self, "foq_filter_device_listbox"):
            return
        visible = set(self._visible_listbox_values(self.foq_filter_device_listbox))
        selected = {str(self.foq_filter_device_listbox.get(index)) for index in self.foq_filter_device_listbox.curselection()}
        self.foq_selected_devices.difference_update(visible)
        self.foq_selected_devices.update(selected)

    def _on_foq_device_selection_changed(self, _event: tk.Event | None = None) -> None:
        if self._foq_filter_refreshing:
            return
        self._sync_foq_device_selection_from_listbox()
        self._refresh_foq_db_field_filter_options()

    def _selected_foq_devices(self) -> list[str]:
        if hasattr(self, "foq_filter_device_listbox") and not self._foq_filter_refreshing:
            self._sync_foq_device_selection_from_listbox()
        return [device for device in self.foq_filter_device_options if device in self.foq_selected_devices]

    def select_all_foq_devices(self) -> None:
        if hasattr(self, "foq_filter_device_listbox"):
            self.foq_selected_devices.update(self._visible_listbox_values(self.foq_filter_device_listbox))
            self._populate_foq_device_filter_listbox()
            self._refresh_foq_db_field_filter_options()

    def clear_foq_devices(self) -> None:
        if hasattr(self, "foq_filter_device_listbox"):
            self.foq_selected_devices.clear()
            self._populate_foq_device_filter_listbox()
            self._refresh_foq_db_field_filter_options()

    def _refresh_foq_db_field_filter_options(self) -> None:
        if not hasattr(self, "foq_filter_db_field_listbox"):
            return
        fields = self._foq_db_fields_for_filter_device()
        self.foq_selected_db_fields.intersection_update(set(fields))
        self.foq_filter_db_field_options = fields
        self._populate_foq_db_field_filter_listbox()

    def _populate_foq_db_field_filter_listbox(self) -> None:
        if not hasattr(self, "foq_filter_db_field_listbox"):
            return
        self._foq_filter_refreshing = True
        try:
            visible = self._filter_foq_options(self.foq_filter_db_field_options, self.foq_db_field_search_var.get())
            self.foq_filter_db_field_listbox.delete(0, tk.END)
            for field in visible:
                self.foq_filter_db_field_listbox.insert(tk.END, field)
                if field in self.foq_selected_db_fields:
                    self.foq_filter_db_field_listbox.selection_set(tk.END)
        finally:
            self._foq_filter_refreshing = False
        self._update_foq_filter_counts()

    def _sync_foq_db_field_selection_from_listbox(self) -> None:
        if not hasattr(self, "foq_filter_db_field_listbox"):
            return
        visible = set(self._visible_listbox_values(self.foq_filter_db_field_listbox))
        selected = {str(self.foq_filter_db_field_listbox.get(index)) for index in self.foq_filter_db_field_listbox.curselection()}
        self.foq_selected_db_fields.difference_update(visible)
        self.foq_selected_db_fields.update(selected)

    def _on_foq_db_field_selection_changed(self, _event: tk.Event | None = None) -> None:
        if self._foq_filter_refreshing:
            return
        self._sync_foq_db_field_selection_from_listbox()
        self._update_foq_filter_counts()

    def _update_foq_filter_counts(self) -> None:
        if hasattr(self, "foq_device_count_var"):
            visible_devices = len(self._filter_foq_options(self.foq_filter_device_options, self.foq_device_search_var.get()))
            self.foq_device_count_var.set(f"{len(self.foq_selected_devices)} selected / {len(self.foq_filter_device_options)} devices ({visible_devices} shown)")
        if hasattr(self, "foq_db_field_count_var"):
            visible_fields = len(self._filter_foq_options(self.foq_filter_db_field_options, self.foq_db_field_search_var.get()))
            self.foq_db_field_count_var.set(f"{len(self.foq_selected_db_fields)} selected / {len(self.foq_filter_db_field_options)} fields ({visible_fields} shown)")

    def _selected_foq_db_fields(self) -> list[str]:
        if hasattr(self, "foq_filter_db_field_listbox") and not self._foq_filter_refreshing:
            self._sync_foq_db_field_selection_from_listbox()
        return [field for field in self.foq_filter_db_field_options if field in self.foq_selected_db_fields]

    def select_all_foq_db_fields(self) -> None:
        if hasattr(self, "foq_filter_db_field_listbox"):
            self.foq_selected_db_fields.update(self._visible_listbox_values(self.foq_filter_db_field_listbox))
            self._populate_foq_db_field_filter_listbox()

    def clear_foq_db_fields(self) -> None:
        if hasattr(self, "foq_filter_db_field_listbox"):
            self.foq_selected_db_fields.clear()
            self._populate_foq_db_field_filter_listbox()

    def _clear_foq_log(self) -> None:
        if not hasattr(self, "foq_preview_log"):
            return
        self.foq_preview_log.configure(state="normal")
        self.foq_preview_log.delete("1.0", tk.END)
        self.foq_preview_log.configure(state="disabled")

    def _foq_log(self, message: str) -> None:
        self._call_ui(lambda message=message: self._append_foq_log(message))

    def _append_foq_log(self, message: str) -> None:
        if not hasattr(self, "foq_preview_log"):
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.foq_preview_log.configure(state="normal")
        self.foq_preview_log.insert(tk.END, f"[{timestamp}] {message}\n")
        line_count = int(self.foq_preview_log.index("end-1c").split(".", 1)[0])
        if line_count > 300:
            self.foq_preview_log.delete("1.0", "50.0")
        self.foq_preview_log.see(tk.END)
        self.foq_preview_log.configure(state="disabled")

    def preview_foq_db_fields(self) -> None:
        self._commit_path_display(self.foq_mapping_path_var, self.foq_mapping_display_var, "foq")
        self._refresh_foq_mapping_filter_options()
        mapping_path = Path(self.foq_mapping_path_var.get().strip())
        if not mapping_path.exists():
            messagebox.showinfo(APP_NAME, "Select a valid FOQResultLocations mapping file first.")
            return
        device_filters = self._selected_foq_devices()
        if not device_filters:
            messagebox.showinfo(APP_NAME, "Select one or more Devices first.")
            return
        fields = self._selected_foq_db_fields()
        if not fields:
            messagebox.showinfo(APP_NAME, "Select one or more DB fields to preview.")
            return
        self._configure_foq_preview_table(fields)
        self._clear_foq_log()
        self.progress_var.set(0.0)
        scope_items = self._foq_sequence_search_items() if not self.foq_candidate_sequences else []
        scope_text = self._foq_sequence_search_scope_text(len(scope_items)) if scope_items else f"{len(self.foq_candidate_sequences)} candidate sequence(s)"
        self.status_var.set(f"Finding sequence candidates for {len(fields)} DB field(s): {scope_text}")
        self._append_foq_log(f"Preview requested: {len(device_filters)} device(s), {len(fields)} DB field(s).")
        self._append_foq_log(f"Search scope: {scope_text}.")
        self._append_foq_log(f"Mapping file: {mapping_path}")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                candidates = self._foq_preview_candidates(device_filters)
                if not candidates:
                    self._call_ui(lambda: self._foq_preview_no_candidates())
                    return
                package_count = len({str(package.path) for package, *_rest in candidates})
                self._foq_log(f"Found {len(candidates)} matching sequence(s) in {package_count} CMBX package(s).")
                self._thread_status(f"Previewing {len(fields)} DB field(s) for {len(candidates)} sequence(s)...")
                payload_rows = self._run_foq_preview_batch_worker(candidates, mapping_path, fields)
                rows: list[tuple[str, ...]] = []
                for payload in payload_rows:
                    values = payload.get("values", {})
                    rows.append(
                        (
                            str(payload.get("sequence", "")),
                            str(payload.get("device", "")),
                            str(payload.get("report_template", "")),
                            *[self._format_workbook_preview_value(values.get(field, "")) for field in fields],
                        )
                    )
                self._call_ui(lambda: self._set_foq_preview_rows(fields, rows))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _run_foq_preview_batch_worker(
        self,
        candidates: list[tuple[CmbxPackage, CmbxElement, str, str, str]],
        mapping_path: Path,
        fields: list[str],
    ) -> list[dict[str, object]]:
        worker_path = Path(__file__).with_name("foq_preview_worker.py")
        jobs = [
            {
                "package": str(package.path),
                "sequence_key": sequence.id or sequence.url or sequence.name,
                "sequence_name": sequence.name,
                "device": device,
                "report_template": report_template,
            }
            for package, sequence, device, _source, report_template in candidates
        ]
        jobs_file = tempfile.NamedTemporaryFile("w", suffix=".json", encoding="utf-8", delete=False)
        jobs_path = Path(jobs_file.name)
        try:
            json.dump({"jobs": jobs}, jobs_file, ensure_ascii=False)
            jobs_file.close()
        except Exception:
            jobs_file.close()
            jobs_path.unlink(missing_ok=True)
            raise
        command = [
            sys.executable,
            "-B",
            str(worker_path),
            "--mapping",
            str(mapping_path),
            "--fields-json",
            json.dumps(fields),
            "--jobs-file",
            str(jobs_path),
        ]
        creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" and hasattr(subprocess, "CREATE_NO_WINDOW") else 0
        self._foq_log("Starting one batched preview worker. CMBX packages will be loaded once per package.")
        try:
            process = subprocess.Popen(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=creationflags,
            )
            result_rows: list[dict[str, object]] = []
            assert process.stdout is not None
            for raw_line in process.stdout:
                line = raw_line.strip()
                if not line:
                    continue
                try:
                    event = json.loads(line)
                except json.JSONDecodeError:
                    self._foq_log(line)
                    continue
                event_type = event.get("type")
                if event_type == "log":
                    self._foq_log(str(event.get("message", "")))
                elif event_type == "result":
                    result_rows = list(event.get("rows", []))
            return_code = process.wait()
            if return_code != 0:
                raise RuntimeError(f"preview worker exited with code {return_code}")
            self._foq_log(f"Preview worker finished with {len(result_rows)} result row(s).")
            return result_rows
        finally:
            jobs_path.unlink(missing_ok=True)

    def _foq_preview_no_candidates(self) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(0.0)
        self.status_var.set("FOQ DB preview found no matching sequence")
        self._append_foq_log("No matching non-deleted sequence was found.")
        messagebox.showinfo(APP_NAME, "No non-deleted sequence matched the selected Device(s).")

    def _foq_preview_candidates(self, device_filters: list[str]) -> list[tuple[CmbxPackage, CmbxElement, str, str, str]]:
        device_filter_keys = {device.lower() for device in device_filters}
        source = self.foq_candidate_sequences
        candidates: list[tuple[CmbxPackage, CmbxElement, str, str, str]] = []
        if source:
            for package, sequence, device, source_text, report_template in source:
                if self._foq_package_is_deleted(package):
                    continue
                if not device_filter_keys or device.lower() in device_filter_keys:
                    candidates.append((package, sequence, device, source_text, report_template))
            return candidates
        for package, sequence in self._foq_sequence_search_items():
            device, source_text = self._resolve_foq_device_type(package, sequence)
            if device_filter_keys and device.lower() not in device_filter_keys:
                continue
            candidates.append((package, sequence, device, source_text, self._default_foq_report_template_name(sequence)))
        return candidates

    def _configure_foq_preview_table(self, fields: list[str]) -> None:
        table = self._table_widget(self.foq_preview_table)
        columns = ("sequence", "device", "report_template", *fields)
        table.configure(columns=columns)
        headings = {"sequence": "Sequence", "device": "Device", "report_template": "Report Template", **{field: field for field in fields}}
        widths = {"sequence": 220, "device": 130, "report_template": 220}
        for column in columns:
            table.heading(column, text=headings[column])
            table.column(column, width=widths.get(column, 150), anchor="w")
        table.delete(*table.get_children())

    def _set_foq_preview_rows(self, fields: list[str], rows: list[tuple[str, ...]]) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(100.0 if rows else 0.0)
        self._configure_foq_preview_table(fields)
        table = self._table_widget(self.foq_preview_table)
        if not rows:
            table.insert("", "end", values=("No preview data was found.", "", "", *["" for _field in fields]))
            self.status_var.set("FOQ DB field preview found no data")
            return
        for row in rows:
            table.insert("", "end", values=row)
        self.status_var.set(f"Previewed {len(rows)} sequence(s)")
        self._append_foq_log(f"Preview table updated: {len(rows)} sequence row(s).")

    def _read_foq_preview_db_values(self, workbook_path: Path) -> dict[str, object]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            sheet = workbook["DB Data"]
            rows = list(sheet.iter_rows(max_row=2, values_only=True))
            if len(rows) < 2:
                return {}
            return {str(header): value for header, value in zip(rows[0], rows[1]) if header}
        finally:
            workbook.close()

    def _foq_db_fields_for_filter_device(self) -> list[str]:
        mapping_path = Path(self.foq_mapping_path_var.get().strip())
        if not mapping_path.exists():
            return []
        selected_devices = self._selected_foq_devices()
        try:
            workbook = load_foq_workbook(mapping_path)
            mappings = read_device_type_mappings(workbook)
        except Exception:
            return []
        field_sets: list[set[str]] = []
        for mapping in mappings.values():
            if selected_devices and mapping.device_type not in selected_devices:
                continue
            try:
                field_sets.append({location.db_field for location in read_result_locations(workbook, mapping.sheet_name) if location.db_field})
            except Exception:
                continue
        if not field_sets:
            return []
        fields = set.intersection(*field_sets) if selected_devices else set.union(*field_sets)
        return sorted(fields)

    def _maybe_edit_foq_report_template_cell(self, event) -> None:
        table = self._table_widget(self.foq_candidate_table)
        region = table.identify_region(event.x, event.y)
        if region != "cell":
            self._destroy_foq_report_template_combo()
            return
        row_iid = table.identify_row(event.y)
        column_id = table.identify_column(event.x)
        columns = list(table["columns"])
        try:
            column_name = columns[int(column_id.removeprefix("#")) - 1]
        except (ValueError, IndexError):
            self._destroy_foq_report_template_combo()
            return
        if column_name != "report_template" or not row_iid.startswith("foqcandidate:"):
            self._destroy_foq_report_template_combo()
            return
        index = int(row_iid.split(":", 1)[1])
        if not (0 <= index < len(self.foq_candidate_sequences)):
            return
        _package, sequence, _device, _source, current_report = self.foq_candidate_sequences[index]
        names = [report.name for report in self._foq_report_templates_for_sequence(sequence)]
        if not names:
            return
        bbox = table.bbox(row_iid, column_id)
        if not bbox:
            return
        self._destroy_foq_report_template_combo()
        x, y, width, height = bbox
        combo = ttk.Combobox(table, values=names, state="readonly", font=self._font(9))
        combo.set(current_report if current_report in names else names[0])
        combo.place(x=x, y=y, width=width, height=height)
        combo.focus_set()

        def commit(_event=None) -> None:
            selected = combo.get().strip()
            if selected:
                self._set_foq_candidate_report_template(index, selected)
            self._destroy_foq_report_template_combo()

        combo.bind("<<ComboboxSelected>>", commit)
        combo.bind("<FocusOut>", commit)
        combo.bind("<Escape>", lambda _event=None: self._destroy_foq_report_template_combo())
        self._foq_report_template_combo = combo

    def _set_foq_candidate_report_template(self, index: int, report_template: str) -> None:
        if not (0 <= index < len(self.foq_candidate_sequences)):
            return
        package, sequence, device, source, _current_report = self.foq_candidate_sequences[index]
        valid = {report.name.lower(): report.name for report in self._foq_report_templates_for_sequence(sequence)}
        matched = valid.get(report_template.lower())
        if not matched:
            return
        self.foq_candidate_sequences[index] = (package, sequence, device, source, matched)
        self._refresh_foq_candidate_table()
        self.status_var.set(f"Set report template for {sequence.name}: {matched}")

    def _destroy_foq_report_template_combo(self) -> None:
        if self._foq_report_template_combo is not None:
            try:
                self._foq_report_template_combo.destroy()
            except tk.TclError:
                pass
            self._foq_report_template_combo = None

    def _foq_sequence_matches_filters(self, package: CmbxPackage, sequence: CmbxElement) -> bool:
        device_filters = {device.lower() for device in self._selected_foq_devices()}
        if device_filters:
            device, _source = self._resolve_foq_device_type(package, sequence)
            if device.lower() not in device_filters:
                return False
        return True

    def _contains_filter(self, value: str, filter_text: str) -> bool:
        tokens = [token for token in re.split(r"[\s,;]+", filter_text.strip().lower()) if token]
        haystack = str(value or "").lower()
        return all(token in haystack for token in tokens)

    def _resolve_foq_device_type(self, package: CmbxPackage, sequence: CmbxElement) -> tuple[str, str]:
        cache_key = (self._path_key(package.path), sequence.id or sequence.url or sequence.name)
        cached = self.foq_device_cache.get(cache_key)
        if cached:
            return cached
        audit_device = self._foq_device_from_audit(package, sequence)
        if audit_device:
            self.foq_device_cache[cache_key] = audit_device
            return audit_device
        report_device = self._foq_device_from_report(package, sequence)
        if report_device:
            self.foq_device_cache[cache_key] = report_device
            return report_device
        unresolved = ("unresolved", "No ColumnComp.ModelNo value found in report formulas or audit/precondition records")
        self.foq_device_cache[cache_key] = unresolved
        return unresolved

    def _foq_device_from_report(self, package: CmbxPackage, sequence: CmbxElement) -> tuple[str, str] | None:
        injections = [child for child in sequence.children if child.kind == "injection"]
        reports = [child for child in sequence.children if child.kind == "report_template" and "ReportDefinition" in child.item_type]
        if not injections or not reports:
            return None
        for report in reports:
            try:
                xml_text = self.report_xml_cache.get(report.id)
                if xml_text is None:
                    _embedded, xml_text = decode_report_template_xml(package, report)
                    self.report_xml_cache[report.id] = xml_text
            except Exception:
                continue
            model_sheets = ["Definitions", "Internal Use", ""]
            for injection in injections:
                try:
                    context = build_report_formula_context(package, injection)
                except Exception:
                    continue
                for sheet_name in model_sheets:
                    try:
                        evaluations = evaluate_report_formulas(package, injection, report.name, xml_text, sheet_name, context=context)
                    except Exception:
                        continue
                    for row in evaluations:
                        if row.status != "ok":
                            continue
                        if not self._is_foq_device_model_formula(row.formula, row.detail):
                            continue
                        device = self._normalize_foq_device_candidate(row.value)
                        if device:
                            location = f"{row.sheet_name or sheet_name}!{row.excel_range}".strip("!")
                            return device, f"Report formula {location}: {row.formula}"
        return None

    def _is_foq_device_model_formula(self, formula: str, detail: str = "") -> bool:
        # FOQ DB device selection must come from Chromeleon's ModelNo field, not from names.
        normalized = re.sub(r"\s+", "", str(formula or "")).lower()
        if normalized == "audit.columncomp.modelno":
            return True
        return "audit.columncomp.modelno" in re.sub(r"\s+", "", str(detail or "")).lower()

    def _foq_device_from_audit(self, package: CmbxPackage, sequence: CmbxElement) -> tuple[str, str] | None:
        injections = [child for child in sequence.children if child.kind == "injection"]
        for injection in injections:
            try:
                context = build_report_formula_context(package, injection)
            except Exception:
                continue
            for record in context.audit_records:
                record_path = f"{record.device}.{record.property_name}"
                if "modelno" in record_path.lower():
                    device = self._normalize_foq_device_candidate(record.property_value)
                    if device:
                        source = "audit precondition" if record.retention_time_min is None else f"audit {record.retention_time_min:.6g} min"
                        return device, f"{source}: {record_path}"
                for text in (record.property_value, record.message):
                    if not text:
                        continue
                    device = self._normalize_foq_device_candidate(text)
                    if device and "modelno" in f"{record_path} {record.message}".lower():
                        return device, f"audit message: {record_path}"
        return None

    def _normalize_foq_device_candidate(self, value: str) -> str:
        raw = str(value or "").strip().strip("'\"")
        if not raw:
            return ""
        lookup = self._foq_device_type_lookup()
        normalized = raw.upper().replace("_", "-")
        compact = re.sub(r"[^A-Z0-9-]+", " ", normalized)
        if normalized in lookup:
            return lookup[normalized]
        compact_tokens = set(compact.split())
        for key, device in sorted(lookup.items(), key=lambda item: len(item[0]), reverse=True):
            if key in compact_tokens or key in normalized or key in compact:
                return device
        match = re.search(r"\bV[A-Z]-[A-Z0-9]{2,4}-[A-Z0-9]{1,4}\b", compact)
        return match.group(0) if match else ""

    def _foq_device_type_lookup(self) -> dict[str, str]:
        fallback = {device: device for device in ("VH-C10-A", "VC-C10-A", "VN-C10-A", "VA-C10-A", "VF-C10-A")}
        mapping_path = Path(self.foq_mapping_path_var.get().strip())
        if not mapping_path.exists():
            return fallback
        try:
            workbook = load_foq_workbook(mapping_path)
            mappings = read_device_type_mappings(workbook)
        except Exception:
            return fallback
        lookup = {key.upper().replace("_", "-"): mapping.device_type for key, mapping in mappings.items()}
        return lookup or fallback

    def _refresh_foq_candidate_table(self) -> None:
        if not hasattr(self, "foq_candidate_table"):
            return
        table = self._table_widget(self.foq_candidate_table)
        table.delete(*table.get_children())
        for index, (package, sequence, device, source, report_template) in enumerate(self.foq_candidate_sequences):
            injections = [child for child in sequence.children if child.kind == "injection"]
            channels = [signal for injection in injections for signal in injection.children if signal.kind == "signal"]
            table.insert(
                "",
                "end",
                iid=f"foqcandidate:{index}",
                values=(device, source, report_template, package.path.name, sequence.name, len(injections), len(channels), str(package.path)),
            )

    def remove_selected_foq_candidates(self) -> None:
        if not hasattr(self, "foq_candidate_table"):
            return
        table = self._table_widget(self.foq_candidate_table)
        indexes = sorted((int(iid.split(":", 1)[1]) for iid in table.selection() if iid.startswith("foqcandidate:")), reverse=True)
        for index in indexes:
            if 0 <= index < len(self.foq_candidate_sequences):
                self.foq_candidate_sequences.pop(index)
        self._refresh_foq_candidate_table()

    def clear_foq_candidates(self) -> None:
        self.foq_candidate_sequences = []
        self._refresh_foq_candidate_table()

    def _selected_foq_candidates(self) -> list[tuple[CmbxPackage, CmbxElement, str, str, str]]:
        if not hasattr(self, "foq_candidate_table"):
            return self.foq_candidate_sequences
        table = self._table_widget(self.foq_candidate_table)
        selected = [int(iid.split(":", 1)[1]) for iid in table.selection() if iid.startswith("foqcandidate:")]
        if not selected:
            return list(self.foq_candidate_sequences)
        return [self.foq_candidate_sequences[index] for index in selected if 0 <= index < len(self.foq_candidate_sequences)]

    def _prune_foq_candidates(self) -> None:
        if not self.foq_candidate_sequences:
            return
        valid = {self._foq_candidate_key(package, sequence) for package in self.loaded_packages for sequence in package.sequences}
        self.foq_candidate_sequences = [
            (package, sequence, device, source, report_template)
            for package, sequence, device, source, report_template in self.foq_candidate_sequences
            if self._foq_candidate_key(package, sequence) in valid
        ]
        self._refresh_foq_candidate_table()

    def _selected_filesystem_paths(self) -> list[Path]:
        paths: list[Path] = []
        seen: set[str] = set()
        for iid in self.package_tree.selection():
            path = self.tree_fs_context.get(iid)
            if not path:
                continue
            key = self._path_key(path)
            if key in seen:
                continue
            seen.add(key)
            paths.append(path)
        return self._collapse_nested_paths(paths)

    def _path_key(self, path: Path) -> str:
        try:
            return str(path.resolve()).lower()
        except OSError:
            return str(path).lower()

    def _tree_open_key_for_fs_path(self, path: Path) -> tuple[str, str]:
        return ("fs", self._path_key(path))

    def _tree_open_key_for_element(self, package: CmbxPackage, element: CmbxElement) -> tuple[str, str]:
        element_key = element.id or element.url or element.raw_filename or element.filename or element.name
        return ("element", f"{self._path_key(package.path)}::{element_key}")

    def _capture_tree_open_keys(self) -> set[tuple[str, str]]:
        open_keys: set[tuple[str, str]] = set()

        def walk(iid: str) -> None:
            if self.package_tree.item(iid, "open"):
                fs_path = self.tree_fs_context.get(iid)
                package, element = self.tree_item_context.get(iid, (None, None))
                if fs_path:
                    open_keys.add(self._tree_open_key_for_fs_path(fs_path))
                elif package and element:
                    open_keys.add(self._tree_open_key_for_element(package, element))
            for child_iid in self.package_tree.get_children(iid):
                walk(child_iid)

        for root_iid in self.package_tree.get_children(""):
            walk(root_iid)
        return open_keys

    def _collapse_nested_paths(self, paths: list[Path]) -> list[Path]:
        resolved = [(path, self._path_key(path)) for path in paths]
        result: list[Path] = []
        for path, key in resolved:
            parent_selected = False
            for other_path, other_key in resolved:
                if key == other_key:
                    continue
                try:
                    path.resolve().relative_to(other_path.resolve())
                    parent_selected = True
                    break
                except (OSError, ValueError):
                    continue
            if not parent_selected:
                result.append(path)
        return result

    def _is_deletable_tree_path(self, path: Path) -> bool:
        root = self._package_tree_root()
        deleted = root / "deleted"
        try:
            resolved = path.resolve()
            if resolved == root.resolve() or resolved == deleted.resolve():
                return False
            resolved.relative_to(root.resolve())
        except (OSError, ValueError):
            return False
        return path.exists()

    def _single_renamable_tree_path(self, paths: list[Path]) -> Path | None:
        if len(paths) != 1:
            return None
        path = paths[0]
        return path if self._is_renamable_tree_path(path) else None

    def _is_renamable_tree_path(self, path: Path) -> bool:
        root = self._package_tree_root()
        deleted = root / "deleted"
        try:
            resolved = path.resolve()
            if resolved == root.resolve() or resolved == deleted.resolve():
                return False
            resolved.relative_to(root.resolve())
        except (OSError, ValueError):
            return False
        return path.exists()

    def rename_tree_path(self, path: Path) -> None:
        if not self._is_renamable_tree_path(path):
            return
        new_name = simpledialog.askstring(APP_NAME, "Rename to:", initialvalue=path.name, parent=self.root)
        if new_name is None:
            return
        new_name = new_name.strip()
        if not new_name or new_name == path.name:
            return
        if any(separator in new_name for separator in ("\\", "/")):
            messagebox.showerror(APP_NAME, "Name cannot contain path separators.")
            return
        destination = path.with_name(new_name)
        if destination.exists():
            messagebox.showerror(APP_NAME, f"An item named '{new_name}' already exists.")
            return
        self._tree_restore_open_keys = self._capture_tree_open_keys()
        try:
            path.rename(destination)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not rename item:\n{exc}")
            return
        self.status_var.set(f"Renamed {path.name} to {destination.name}")
        self._reload_packages_from_source()

    def rename_cmbx_tree_element(self, package: CmbxPackage, element: CmbxElement) -> None:
        if element.kind not in {"sequence", "injection", "folder"}:
            return
        new_name = simpledialog.askstring(APP_NAME, "Rename to:", initialvalue=element.name, parent=self.root)
        if new_name is None:
            return
        new_name = new_name.strip()
        if not new_name or new_name == element.name:
            return
        if any(separator in new_name for separator in ("\\", "/")):
            messagebox.showerror(APP_NAME, "Name cannot contain path separators.")
            return
        self._tree_restore_open_keys = self._capture_tree_open_keys()
        try:
            rename_cmbx_header_element(package, element, new_name)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not rename CMBX item:\n{exc}")
            return
        self.status_var.set(f"Renamed {element.kind}: {element.name} to {new_name}")
        self._reload_packages_from_source()

    def delete_tree_paths(self, paths: list[Path]) -> None:
        paths = [path for path in self._collapse_nested_paths(paths) if self._is_deletable_tree_path(path)]
        if not paths:
            return
        preview = "\n".join(str(path) for path in paths[:8])
        if len(paths) > 8:
            preview += f"\n... {len(paths) - 8} more"
        if not messagebox.askyesno(APP_NAME, f"Move {len(paths)} item(s) to packages\\deleted?\n\n{preview}"):
            return
        deleted_folder = self._package_tree_root() / "deleted"
        try:
            deleted_folder.mkdir(parents=True, exist_ok=True)
            for path in paths:
                destination = self._unique_destination(deleted_folder / path.name)
                shutil.move(str(path), str(destination))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not move item(s) to deleted:\n{exc}")
            return
        self.status_var.set(f"Moved {len(paths)} item(s) to {deleted_folder}")
        self._reload_packages_from_source()

    def _unique_destination(self, destination: Path) -> Path:
        if not destination.exists():
            return destination
        stem = destination.stem
        suffix = destination.suffix
        parent = destination.parent
        counter = 2
        while True:
            candidate = parent / f"{stem} ({counter}){suffix}"
            if not candidate.exists():
                return candidate
            counter += 1

    def _reload_packages_from_source(self, preserve_tree_state: bool = True) -> None:
        if preserve_tree_state and not self._tree_restore_open_keys:
            self._tree_restore_open_keys = self._capture_tree_open_keys()
        source_text = self.cmbx_path_var.get()
        self.status_var.set("Refreshing CMBX package tree...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths = self._cmbx_paths_from_text(source_text)
                packages: list[CmbxPackage] = []
                total = max(len(paths), 1)
                for index, path in enumerate(paths, 1):
                    self._thread_status(f"__PROGRESS__={(index - 1) / total * 95:.1f}|Refreshing CMBX {index}/{len(paths)}: {path.name}")
                    packages.append(load_cmbx_package(path))
                self._call_ui(lambda packages=packages: self._reload_packages_done(packages))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._reload_packages_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _reload_packages_done(self, packages: list[CmbxPackage]) -> None:
        self._set_buttons_state("normal")
        self.loaded_packages = packages
        self.foq_device_cache.clear()
        self._set_workspace_packages(packages)
        self.status_var.set(f"Refreshed {len(packages)} package(s)")

    def _reload_packages_failed(self, exc: Exception) -> None:
        self._set_buttons_state("normal")
        self._tree_restore_open_keys = set()
        self.status_var.set("Refresh failed")
        messagebox.showerror(APP_NAME, f"Could not rescan CMBX packages:\n{exc}")

    def _begin_tree_drag(self, event) -> None:
        iid = self.package_tree.identify_row(event.y)
        self._tree_drag_iid = iid or ""
        self._tree_drag_started = False
        self._tree_drag_start_xy = (event.x_root, event.y_root)
        self._destroy_tree_drag_label()

    def _update_tree_drag(self, event) -> None:
        if not self._tree_drag_iid:
            return
        start_x, start_y = self._tree_drag_start_xy
        if not self._tree_drag_started:
            if abs(event.x_root - start_x) < 5 and abs(event.y_root - start_y) < 5:
                return
            self._tree_drag_started = True
            self._show_tree_drag_label(event)
        self._update_tree_drop_target(event.y)
        self._update_tree_drag_label_operation(event)
        self._move_tree_drag_label(event)

    def _finish_tree_drag(self, event) -> None:
        source_iid = self._tree_drag_iid
        was_dragging = self._tree_drag_started
        self._tree_drag_iid = ""
        self._tree_drag_started = False
        self._destroy_tree_drag_label()
        self._clear_tree_drop_target()
        if not source_iid or not was_dragging:
            return
        target_iid = self.package_tree.identify_row(event.y)
        if not target_iid or target_iid == source_iid:
            return
        target_folder = self._target_folder_for_tree_drop(target_iid)
        if not target_folder:
            return
        copy_mode = self._tree_drag_copy_mode(event)
        if source_iid not in self.package_tree.selection():
            self.package_tree.selection_set(source_iid)
        fs_paths = self._selected_filesystem_paths()
        if fs_paths and self.tree_fs_context.get(source_iid):
            self._transfer_filesystem_paths_to_folder(fs_paths, target_folder, copy_mode=copy_mode)
            return
        sequence_groups = self._selected_sequences_by_package()
        if sequence_groups:
            self._unpack_sequence_groups_to_folder(sequence_groups, target_folder)

    def _show_tree_drag_label(self, event) -> None:
        items = self.package_tree.selection()
        count = len(items) if items else 1
        source_name = self.package_tree.item(self._tree_drag_iid, "text") or "Item"
        text = source_name if count <= 1 else f"{source_name} + {count - 1}"
        label_window = tk.Toplevel(self.root)
        label_window.overrideredirect(True)
        label_window.attributes("-topmost", True)
        label = tk.Label(
            label_window,
            text=text,
            font=self._font(9),
            bg="#FFFFFF",
            fg=self.colors["text"],
            bd=1,
            relief="solid",
            padx=10,
            pady=5,
        )
        label.pack()
        self._tree_drag_label = label_window
        self._tree_drag_label_text = label
        self._update_tree_drag_label_operation(event)
        self._move_tree_drag_label(event)

    def _tree_drag_copy_mode(self, event) -> bool:
        return bool(getattr(event, "state", 0) & 0x0004)

    def _update_tree_drag_label_operation(self, event) -> None:
        if not self._tree_drag_label_text:
            return
        items = self.package_tree.selection()
        count = len(items) if items else 1
        source_name = self.package_tree.item(self._tree_drag_iid, "text") or "Item"
        base = source_name if count <= 1 else f"{source_name} + {count - 1}"
        prefix = "Copy: " if self._tree_drag_copy_mode(event) else "Move: "
        self._tree_drag_label_text.config(text=f"{prefix}{base}")

    def _move_tree_drag_label(self, event) -> None:
        if self._tree_drag_label:
            self._tree_drag_label.geometry(f"+{event.x_root + 14}+{event.y_root + 14}")

    def _destroy_tree_drag_label(self) -> None:
        if self._tree_drag_label:
            try:
                self._tree_drag_label.destroy()
            except tk.TclError:
                pass
        self._tree_drag_label = None
        self._tree_drag_label_text = None

    def _update_tree_drop_target(self, y: int) -> None:
        target_iid = self.package_tree.identify_row(y)
        target_iid = self._target_folder_iid_for_tree_drop(target_iid) if target_iid else ""
        if target_iid == self._tree_drop_target_iid:
            return
        self._clear_tree_drop_target()
        if target_iid:
            tags = tuple(self.package_tree.item(target_iid, "tags"))
            self._tree_drop_target_tags = tags
            self.package_tree.item(target_iid, tags=("drop_target",))
            self._tree_drop_target_iid = target_iid

    def _clear_tree_drop_target(self) -> None:
        if not self._tree_drop_target_iid:
            return
        try:
            self.package_tree.item(self._tree_drop_target_iid, tags=self._tree_drop_target_tags)
        except tk.TclError:
            pass
        self._tree_drop_target_iid = ""
        self._tree_drop_target_tags = ()

    def _target_folder_iid_for_tree_drop(self, iid: str) -> str:
        path = self.tree_fs_context.get(iid)
        if path and path.is_dir():
            return iid
        parent = self.package_tree.parent(iid)
        while parent:
            path = self.tree_fs_context.get(parent)
            if path:
                return parent if path.is_dir() else self.package_tree.parent(parent)
            parent = self.package_tree.parent(parent)
        return ""

    def _target_folder_for_tree_drop(self, iid: str) -> Path | None:
        folder_iid = self._target_folder_iid_for_tree_drop(iid)
        path = self.tree_fs_context.get(folder_iid) if folder_iid else None
        if path:
            return path if path.is_dir() else None
        return None

    def _transfer_filesystem_paths_to_folder(self, paths: list[Path], target_folder: Path, copy_mode: bool = False) -> None:
        paths = [path for path in self._collapse_nested_paths(paths) if path.exists()]
        if not paths:
            return
        target_folder.mkdir(parents=True, exist_ok=True)
        try:
            for path in paths:
                if self._path_key(path) == self._path_key(target_folder):
                    continue
                if not copy_mode and self._path_key(path.parent) == self._path_key(target_folder):
                    continue
                if path.is_dir():
                    try:
                        target_folder.resolve().relative_to(path.resolve())
                        continue
                    except (OSError, ValueError):
                        pass
                destination = self._unique_destination(target_folder / path.name)
                if copy_mode:
                    if path.is_dir():
                        shutil.copytree(str(path), str(destination))
                    else:
                        shutil.copy2(str(path), str(destination))
                else:
                    shutil.move(str(path), str(destination))
        except Exception as exc:
            action = "copy" if copy_mode else "move"
            messagebox.showerror(APP_NAME, f"Could not {action} item(s):\n{exc}")
            return
        self.status_var.set(f"{'Copied' if copy_mode else 'Moved'} {len(paths)} item(s) to {target_folder}")
        self._reload_packages_from_source()

    def _unpack_sequence_groups_to_folder(self, grouped: list[tuple[CmbxPackage, list[CmbxElement]]], target_folder: Path) -> None:
        if not grouped:
            return
        target_folder.mkdir(parents=True, exist_ok=True)
        try:
            paths: list[Path] = []
            for package, sequences in grouped:
                paths.extend(split_cmbx_sequences(package, sequences, target_folder))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Could not unpack dragged sequence(s):\n{exc}")
            return
        self.status_var.set(f"Unpacked {len(paths)} sequence CMBX file(s) to {target_folder}")
        self._reload_packages_from_source()

    def unpack_selected_sequences(self, grouped: list[tuple[CmbxPackage, list[CmbxElement]]] | None = None) -> None:
        grouped = grouped or self._selected_sequences_by_package()
        if not grouped:
            messagebox.showinfo(APP_NAME, "Select one or more sequence nodes to unpack.")
            return
        total = sum(len(items) for _package, items in grouped)
        if not messagebox.askyesno(
            APP_NAME,
            f"Unpack {total} selected sequence(s) into standalone CMBX files next to the source package?\n\n"
            "The original packed CMBX will not be modified.",
        ):
            return
        self.status_var.set(f"Unpacking {total} sequence(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                for package, sequences in grouped:
                    paths.extend(split_cmbx_sequences(package, sequences, package.path.parent))
                scan_paths = self._cmbx_paths_from_text(self.cmbx_path_var.get())
                seen_paths = {str(path.resolve()) for path in scan_paths if path.exists()}
                for path in paths:
                    resolved = str(path.resolve())
                    if resolved not in seen_paths:
                        scan_paths.append(path)
                        seen_paths.add(resolved)
                packages = [load_cmbx_package(path) for path in scan_paths]
                self._call_ui(lambda: self._unpack_sequences_done(paths, packages))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._unpack_sequences_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _unpack_sequences_done(self, paths: list[Path], packages: list[CmbxPackage]) -> None:
        self._set_buttons_state("normal")
        self.loaded_packages = packages
        self._set_workspace_packages(packages)
        self.status_var.set(f"Unpacked {len(paths)} sequence CMBX file(s)")
        preview = "\n".join(str(path) for path in paths[:12])
        if len(paths) > 12:
            preview += f"\n... {len(paths) - 12} more"
        messagebox.showinfo(APP_NAME, f"Created {len(paths)} CMBX file(s):\n\n{preview}")

    def _unpack_sequences_failed(self, exc: Exception) -> None:
        self._set_buttons_state("normal")
        self.status_var.set("Unpack failed")
        messagebox.showerror(APP_NAME, f"Could not unpack selected sequence(s):\n{exc}")

    def _populate_sequence(self, sequence: CmbxElement) -> None:
        injections = [child for child in sequence.children if child.kind == "injection"]
        channels = [grandchild for injection in injections for grandchild in injection.children if grandchild.kind == "signal"]
        audits = [grandchild for injection in injections for grandchild in injection.children if grandchild.kind == "audit"]
        self._fill_table(
            self._table_widget(self.channel_table),
            channels,
            lambda e: (self.package.path.name if self.package else "", sequence.name, self._parent_injection_name(e), e.name, e.size or "", e.raw_filename),
        )
        self.channel_plot_points = []
        self.channel_plot_title = ""
        self._redraw_channel_plot()
        self._fill_table(self._table_widget(self.audit_table), audits, lambda e: (f"{self._parent_injection_name(e)} / {e.name}", e.size or "", e.raw_filename, e.url))
        self._clear_audit_preview()
        self._populate_sequence_methods(sequence)
        self._populate_processing_methods(sequence)
        self._populate_report_templates(sequence)
        self._populate_report_sheets_for_sequence(sequence)
        self._clear_report_template_preview()
        self.summary_label.config(text=f"Sequence: {sequence.name}   Injections: {len(injections)}   Channels: {len(channels)}   Audit trails: {len(audits)}")
        self._show_element_info(sequence)

    def _populate_injection(self, injection: CmbxElement) -> None:
        channels = [child for child in injection.children if child.kind == "signal"]
        audits = [child for child in injection.children if child.kind == "audit"]
        sequence_name = self.current_sequence.name if self.current_sequence else ""
        self._fill_table(
            self._table_widget(self.channel_table),
            channels,
            lambda e: (self.package.path.name if self.package else "", sequence_name, injection.name, e.name, e.size or "", e.raw_filename),
        )
        self.channel_plot_points = []
        self.channel_plot_title = ""
        self._redraw_channel_plot()
        self._fill_table(self._table_widget(self.audit_table), audits, lambda e: (e.name, e.size or "", e.raw_filename, e.url))
        self._clear_audit_preview()
        link = get_injection_method_link(self.injection_method_links, injection)
        method_text = f"   IM: {link.instrument_method}   PM: {link.processing_method}" if link else ""
        self.summary_label.config(text=f"Injection: {injection.name}   Channels: {len(channels)}   Audit trails: {len(audits)}{method_text}")
        self._populate_method_context(injection)
        self._populate_processing_methods_for_injection(injection)
        self._populate_report_templates_for_injection(injection)
        self._populate_report_sheets(injection)
        self._clear_report_template_preview()
        self._show_element_info(injection)

    def _populate_method_context(self, injection: CmbxElement) -> None:
        table = self._table_widget(self.method_context_table)
        table.delete(*table.get_children())
        if not self.package:
            return
        link = get_injection_method_link(self.injection_method_links, injection)
        if not link:
            table.insert("", "end", iid=f"context:none:{injection.id}", values=("Method link", "Not found", "", "sequence .cmd"))
            return
        instrument = self._method_by_name(link.instrument_method, "instrument_method")
        rows: list[tuple[str, str, str, str, str]] = []
        rows.append(((instrument.id if instrument else "context:im:" + injection.id), "Linked Instrument Method", link.instrument_method, "instrument_method", instrument.url if instrument else "sequence .cmd"))
        external = self.external_instrument_methods.get(link.instrument_method)
        if external:
            rows.append(("context:external:" + injection.id, "Reference TXT (view only)", external.path.name, "instrument_method_txt", str(external.path)))
        for iid, role, name, kind, source in rows:
            table.insert("", "end", iid=iid, values=(role, name, kind, source))

    def _populate_sequence_methods(self, sequence: CmbxElement) -> None:
        table = self._table_widget(self.method_context_table)
        table.delete(*table.get_children())
        for method in [child for child in sequence.children if child.kind == "instrument_method"]:
            table.insert("", "end", iid=method.id, values=("Sequence Instrument Method", method.name, method.kind, method.url or method.filename or method.raw_filename))
            external = self.external_instrument_methods.get(method.name)
            if external:
                table.insert("", "end", iid=f"context:external:{method.id}", values=("Reference TXT (view only)", external.path.name, "instrument_method_txt", str(external.path)))

    def _populate_processing_methods(self, sequence: CmbxElement) -> None:
        table = self._table_widget(self.processing_method_table)
        table.delete(*table.get_children())
        for method in [child for child in sequence.children if child.kind == "processing_method"]:
            table.insert("", "end", iid=method.id, values=("Sequence Processing Method", method.name, method.kind, method.url or method.filename or method.raw_filename))

    def _populate_processing_methods_for_injection(self, injection: CmbxElement) -> None:
        table = self._table_widget(self.processing_method_table)
        table.delete(*table.get_children())
        link = get_injection_method_link(self.injection_method_links, injection)
        if not link:
            table.insert("", "end", iid=f"context:none:pm:{injection.id}", values=("Processing Method", "Not found", "", "sequence .cmd"))
            return
        processing = self._method_by_name(link.processing_method, "processing_method")
        table.insert(
            "",
            "end",
            iid=processing.id if processing else f"context:pm:{injection.id}",
            values=("Linked Processing Method", link.processing_method, "processing_method", processing.url if processing else "sequence .cmd"),
        )

    def _populate_report_templates(self, sequence: CmbxElement) -> None:
        table = self._table_widget(self.report_template_table)
        table.delete(*table.get_children())
        for report in report_templates_for_sequence(self.package, sequence) if self.package else ():
            table.insert("", "end", iid=report.id, values=("Sequence Report Template", report.name, report.kind, report.url or report.filename or report.raw_filename))

    def _populate_report_templates_for_injection(self, injection: CmbxElement) -> None:
        table = self._table_widget(self.report_template_table)
        table.delete(*table.get_children())
        for report in self._reports_for_injection(injection):
            table.insert("", "end", iid=report.id, values=("Applicable Report Template", report.name, report.kind, report.url or report.filename or report.raw_filename))

    def _clear_method_context(self) -> None:
        self._table_widget(self.method_context_table).delete(*self._table_widget(self.method_context_table).get_children())
        self._clear_method_flow()

    def _clear_report_sheets(self) -> None:
        self._table_widget(self.report_sheet_table).delete(*self._table_widget(self.report_sheet_table).get_children())
        self._clear_report_sheet_preview()

    def _clear_context_tables(self) -> None:
        self._table_widget(self.channel_table).delete(*self._table_widget(self.channel_table).get_children())
        self.channel_plot_points = []
        self.channel_plot_title = ""
        self._redraw_channel_plot()
        self._table_widget(self.audit_table).delete(*self._table_widget(self.audit_table).get_children())
        self._clear_audit_preview()
        self._clear_method_context()
        self._table_widget(self.processing_method_table).delete(*self._table_widget(self.processing_method_table).get_children())
        self._table_widget(self.report_template_table).delete(*self._table_widget(self.report_template_table).get_children())
        self._clear_report_template_preview()
        self._clear_report_sheets()

    def _populate_report_sheets(self, injection: CmbxElement) -> None:
        table = self._table_widget(self.report_sheet_table)
        table.delete(*table.get_children())
        self._insert_report_sheet_rows(table, injection)

    def _populate_report_sheets_for_sequence(self, sequence: CmbxElement) -> None:
        table = self._table_widget(self.report_sheet_table)
        table.delete(*table.get_children())
        for injection in [child for child in sequence.children if child.kind == "injection"]:
            self._insert_report_sheet_rows(table, injection)

    def _insert_report_sheet_rows(self, table: ttk.Treeview, injection: CmbxElement) -> None:
        for report in self._reports_for_injection(injection):
            cache_key = (report.id, injection.name)
            try:
                sheets = self.report_sheet_cache.get(cache_key)
                if sheets is None:
                    xml_text = self.report_xml_cache.get(report.id)
                    if xml_text is None:
                        _embedded, xml_text = decode_report_template_xml(self.package, report)  # type: ignore[arg-type]
                        self.report_xml_cache[report.id] = xml_text
                    sheets = parse_report_sheets(xml_text, report.name, injection.name)
                    self.report_sheet_cache[cache_key] = sheets
            except Exception as exc:
                table.insert("", "end", values=(injection.name, "Decode failed", "", report.name, "", "", "", "", str(exc), ""))
                continue
            for index, sheet in enumerate(sheets):
                if sheet.applies_to_injection == "No" and not self._is_common_report_sheet(sheet.sheet_name):
                    continue
                tags = ("name_match",) if sheet.applies_to_injection == "Yes" or self._names_match(injection.name, sheet.sheet_name) else ()
                table.insert(
                    "",
                    "end",
                    iid=f"report-sheet:{report.id}:{injection.id}:{index}",
                    values=(
                        injection.name,
                        sheet.sheet_name,
                        sheet.applies_to_injection,
                        sheet.report_name,
                        sheet.object_count,
                        sheet.formula_count,
                        sheet.is_active,
                        sheet.each_injection,
                        sheet.reason,
                        sheet.sheet_id,
                    ),
                    tags=tags,
                )

    def preview_selected_report_sheet(self, _event=None) -> None:
        if not self.package or not hasattr(self, "report_sheet_preview_table"):
            return
        selected_context, message = self._selected_report_sheet_context()
        if not selected_context:
            self._set_report_sheet_preview_message(message)
            return
        report, injection, sheet_name = selected_context
        self.status_var.set(f"Building report sheet preview: {sheet_name}...")
        try:
            preview_rows = self._filled_report_preview_rows(injection, report, sheet_name)
            if preview_rows:
                self._set_excel_preview_rows(self.report_sheet_preview_table, preview_rows)
                self.status_var.set(f"Previewed filled report sheet: {sheet_name}")
                return
            xml_text = self._report_xml(report)
            objects = parse_report_sheet_objects(xml_text, report.name, sheet_name)
            formula_values = {}
            try:
                context = build_report_formula_context(self.package, injection)
                evaluations = evaluate_report_formulas(self.package, injection, report.name, xml_text, sheet_name, context=context)
                formula_values = {self._top_left_cell(row.excel_range): row.value for row in evaluations if row.status == "ok"}
            except Exception:
                formula_values = {}
            self._set_excel_preview_rows(self.report_sheet_preview_table, self._build_report_preview_rows(objects, formula_values, sheet_name))
            self.status_var.set(f"Previewed report sheet: {sheet_name}")
        except Exception as exc:
            self._set_report_sheet_preview_message(str(exc))
            self.status_var.set("Report sheet preview failed")

    def _selected_report_sheet_context(self) -> tuple[tuple[CmbxElement, CmbxElement, str] | None, str]:
        if not self.package or not hasattr(self, "report_sheet_table"):
            return None, "Load a CMBX package first."
        table = self._table_widget(self.report_sheet_table)
        selected = table.selection()
        if not selected:
            return None, "Select a report sheet row to preview."
        iid = selected[0]
        if not iid.startswith("report-sheet:"):
            return None, "The selected row is not a report sheet."
        parts = iid.split(":")
        if len(parts) < 4:
            return None, "Cannot resolve selected report sheet."
        report = self.package.elements_by_id.get(parts[1])
        injection = self.package.elements_by_id.get(parts[2])
        values = table.item(iid, "values")
        sheet_name = str(values[1]) if len(values) > 1 else ""
        if not report or not injection:
            return None, "Cannot resolve report template or injection for this sheet."
        return (report, injection, sheet_name), ""

    def preview_selected_report_template(self, _event=None) -> None:
        if not self.package or not hasattr(self, "report_template_preview_table"):
            return
        table = self._table_widget(self.report_template_table)
        selected = table.selection()
        if not selected:
            self._set_report_template_preview_message("Select a report template row to preview.")
            return
        report = self.package.elements_by_id.get(selected[0])
        if not report or report.kind != "report_template":
            self._set_report_template_preview_message("The selected row is not a report template.")
            return
        self.status_var.set(f"Building report template preview: {report.name}...")
        try:
            xml_text = self._report_xml(report)
            sheets = parse_report_sheets(xml_text, report.name, self.current_injection.name if self.current_injection else "")
            sheet_names = self._report_template_preview_sheet_names(report, sheets)
            if hasattr(self, "report_template_sheet_combo"):
                self.report_template_sheet_combo.configure(values=sheet_names)
            current_sheet = self.report_template_sheet_var.get()
            if sheet_names and current_sheet not in sheet_names:
                current_sheet = sheet_names[0]
                self.report_template_sheet_var.set(current_sheet)
            template_rows = self._blank_report_template_preview_rows(report, current_sheet)
            if template_rows:
                self._set_excel_preview_rows(self.report_template_preview_table, template_rows)
                self.status_var.set(f"Previewed blank report template: {report.name} / {current_sheet}")
                return
            sheet_name = current_sheet or self._first_preview_sheet_name(sheets)
            objects = parse_report_sheet_objects(xml_text, report.name, sheet_name)
            self._set_excel_preview_rows(self.report_template_preview_table, self._build_report_preview_rows(objects, {}, sheet_name))
            self.status_var.set(f"Previewed report template: {report.name} / {sheet_name}")
        except Exception as exc:
            self._set_report_template_preview_message(str(exc))
            self.status_var.set("Report template preview failed")

    def _report_xml(self, report: CmbxElement) -> str:
        xml_text = self.report_xml_cache.get(report.id)
        if xml_text is None:
            _embedded, xml_text = decode_report_template_xml(self.package, report)  # type: ignore[arg-type]
            self.report_xml_cache[report.id] = xml_text
        return xml_text

    def _filled_report_preview_rows(self, injection: CmbxElement, report: CmbxElement, sheet_name: str) -> list[tuple[int, dict[int, str], str]]:
        if not self.package:
            return []
        cache_key = (str(self.package.path), report.id, injection.id, sheet_name)
        cached = self.report_preview_cache.get(cache_key)
        if cached is not None:
            return cached
        try:
            temp_root = Path(tempfile.gettempdir()) / "cmbx_data_explorer_report_preview"
            path = export_filled_report_template_workbook(self.package, injection, report, temp_root, sheet_names=[sheet_name] if sheet_name else None)
            rows = self._workbook_preview_rows(path, sheet_name)
            self.report_preview_cache[cache_key] = rows
            return rows
        except Exception:
            return []

    def _report_template_preview_sheet_names(self, report: CmbxElement, sheets: list[ReportSheet]) -> list[str]:
        workbook_names = self._blank_report_workbook_sheet_names(report)
        if workbook_names:
            return workbook_names
        names = [sheet.sheet_name for sheet in sheets if sheet.sheet_name]
        return names or ["Sheet1"]

    def _blank_report_workbook_sheet_names(self, report: CmbxElement) -> list[str]:
        if not self.package:
            return []
        try:
            import xlrd
        except ImportError:
            return []
        try:
            workbook = self._export_blank_report_template_workbook(report)
            return list(xlrd.open_workbook(str(workbook), on_demand=True).sheet_names())
        except Exception:
            return []

    def _export_blank_report_template_workbook(self, report: CmbxElement) -> Path | None:
        if not self.package:
            return None
        temp_root = Path(tempfile.gettempdir()) / "cmbx_data_explorer_template_preview"
        paths = export_elements(self.package, [report], temp_root)
        return next((path for path in paths if path.suffix.lower() in {".xls", ".xlsx"}), None)

    def _blank_report_template_preview_rows(self, report: CmbxElement, sheet_name: str = "") -> list[tuple[int, dict[int, str], str]]:
        if not self.package:
            return []
        cache_key = (str(self.package.path), report.id, sheet_name, "blank")
        cached = self.report_preview_cache.get(cache_key)
        if cached is not None:
            return cached
        try:
            workbook = self._export_blank_report_template_workbook(report)
            if not workbook:
                return []
            rows = self._workbook_preview_rows(workbook, sheet_name)
            self.report_preview_cache[cache_key] = rows
            return rows
        except Exception:
            return []

    def _workbook_preview_rows(self, path: Path, preferred_sheet: str = "") -> list[tuple[int, dict[int, str], str]]:
        try:
            import xlrd
        except ImportError:
            return []
        workbook = xlrd.open_workbook(str(path), on_demand=True)
        sheet = workbook.sheet_by_name(preferred_sheet) if preferred_sheet and preferred_sheet in workbook.sheet_names() else workbook.sheet_by_index(0)
        rows: list[tuple[int, dict[int, str], str]] = []
        max_rows = min(sheet.nrows, 160)
        max_cols = min(sheet.ncols, len(EXCEL_PREVIEW_LETTERS))
        rows.append((1, {1: sheet.name}, "preview_title"))
        for row_index in range(max_rows):
            cells: dict[int, str] = {}
            for col_index in range(max_cols):
                value = sheet.cell_value(row_index, col_index)
                if value not in ("", None):
                    cells[col_index + 1] = self._format_workbook_preview_value(value)
            if cells:
                rows.append((row_index + 2, cells, "preview_value"))
        return rows

    def _format_workbook_preview_value(self, value) -> str:
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return f"{value:.6g}"
        return str(value)

    def _first_preview_sheet_name(self, sheets: list[ReportSheet]) -> str:
        for sheet in sheets:
            if sheet.object_count and sheet.sheet_name not in {"Title", "Test Procedures", "COC", "FOQ VTCC History"}:
                return sheet.sheet_name
        return sheets[0].sheet_name if sheets else ""

    def _build_report_preview_rows(self, objects, formula_values: dict[str, str], sheet_name: str) -> list[tuple[int, dict[int, str], str]]:
        row_cells: dict[int, dict[int, str]] = {}
        row_tags: dict[int, str] = {}
        row_cells[1] = {1: sheet_name or "Report Sheet"}
        row_tags[1] = "preview_title"
        for obj in objects:
            parsed = self._cell_ref_to_row_col(self._top_left_cell(obj.excel_range))
            if not parsed:
                continue
            row, column = parsed
            if row > 160 or column > len(EXCEL_PREVIEW_LETTERS):
                continue
            value = formula_values.get(self._top_left_cell(obj.excel_range))
            if value not in (None, ""):
                text = str(value)
                tag = "preview_value"
            elif obj.formula:
                text = self._short_preview_text(obj.formula)
                tag = "preview_formula"
            else:
                label = obj.table_type or obj.plot_type or obj.object_type
                text = f"[{label}]" if label else ""
                tag = "preview_object" if text else ""
            if not text:
                continue
            row_cells.setdefault(row, {})[column] = text
            if tag and row_tags.get(row) != "preview_value":
                row_tags[row] = tag
        rows = []
        for row in range(1, min(max(row_cells.keys(), default=1), 160) + 1):
            cells = row_cells.get(row, {})
            if cells or row <= 3:
                rows.append((row, cells, row_tags.get(row, "")))
        return rows

    def _set_excel_preview_rows(self, table_frame: tk.Frame, rows: list[tuple[int, dict[int, str], str]]) -> None:
        table = self._table_widget(table_frame)
        table.delete(*table.get_children())
        if not rows:
            table.insert("", "end", values=self._excel_preview_message_values("No preview data was found."))
            return
        for row_number, cells, tag in rows:
            values = [str(row_number)]
            values.extend(cells.get(index, "") for index in range(1, len(EXCEL_PREVIEW_LETTERS) + 1))
            table.insert("", "end", values=values, tags=(tag,) if tag else ())

    def _excel_preview_message_values(self, message: str) -> tuple[str, ...]:
        return ("", message, *("" for _ in range(max(len(EXCEL_PREVIEW_LETTERS) - 1, 0))))

    def _clear_report_sheet_preview(self) -> None:
        if hasattr(self, "report_sheet_preview_table"):
            self._table_widget(self.report_sheet_preview_table).delete(*self._table_widget(self.report_sheet_preview_table).get_children())

    def _clear_report_template_preview(self) -> None:
        if hasattr(self, "report_template_preview_table"):
            self._table_widget(self.report_template_preview_table).delete(*self._table_widget(self.report_template_preview_table).get_children())

    def _set_report_sheet_preview_message(self, message: str) -> None:
        if hasattr(self, "report_sheet_preview_table"):
            table = self._table_widget(self.report_sheet_preview_table)
            table.delete(*table.get_children())
            table.insert("", "end", values=self._excel_preview_message_values(message))

    def _set_report_template_preview_message(self, message: str) -> None:
        if hasattr(self, "report_template_preview_table"):
            table = self._table_widget(self.report_template_preview_table)
            table.delete(*table.get_children())
            table.insert("", "end", values=self._excel_preview_message_values(message))

    def _top_left_cell(self, excel_range: str) -> str:
        return (excel_range or "").split(":", 1)[0].strip().upper()

    def _cell_ref_to_row_col(self, cell_ref: str) -> tuple[int, int] | None:
        import re

        match = re.fullmatch(r"([A-Z]+)([0-9]+)", cell_ref)
        if not match:
            return None
        column = 0
        for char in match.group(1):
            column = column * 26 + ord(char) - 64
        return int(match.group(2)), column

    def _short_preview_text(self, value: str, limit: int = 80) -> str:
        text = str(value).strip()
        return text if len(text) <= limit else text[: limit - 1] + "..."

    def _is_common_report_sheet(self, sheet_name: str) -> bool:
        return sheet_name in {"Definitions", "Title", "Test Procedures", "COC", "FOQ VTCC History"}

    def _method_by_name(self, name: str, kind: str) -> CmbxElement | None:
        if not self.package:
            return None
        return next((method for method in self.package.methods_and_reports if method.name == name and method.kind == kind), None)

    def _reports_for_injection(self, injection: CmbxElement) -> list[CmbxElement]:
        if not self.package:
            return []
        sequence = self.package.elements_by_id.get(injection.parent_id or "")
        if not sequence:
            return []
        return [
            report for report in report_templates_for_sequence(self.package, sequence)
            if "ReportDefinition" in report.item_type
        ]

    def _parent_injection_name(self, element: CmbxElement) -> str:
        if not self.package or not element.parent_id:
            return ""
        parent = self.package.elements_by_id.get(element.parent_id)
        return parent.name if parent else ""

    def _parent_injection_name_for_package(self, package: CmbxPackage, element: CmbxElement) -> str:
        if not element.parent_id:
            return ""
        parent = package.elements_by_id.get(element.parent_id)
        return parent.name if parent else ""

    def _names_match(self, left: str, right: str) -> bool:
        def normalize(value: str) -> str:
            text = "".join(ch.lower() for ch in value if ch.isalnum())
            text = text.replace("temperature", "temp")
            if len(text) > 1 and text[-1] in {"h", "c"}:
                text = text[:-1]
            return text

        left_key = normalize(left)
        right_key = normalize(right)
        return left_key == right_key or left_key in right_key or right_key in left_key

    def _fill_table(self, table: ttk.Treeview, elements: list[CmbxElement], row_factory) -> None:
        table.delete(*table.get_children())
        for element in elements:
            table.insert("", "end", iid=element.id, values=row_factory(element))

    def _show_package_summary(self) -> None:
        if not self.package:
            return
        counts = summarize_package(self.package)
        self.summary_label.config(
            text=(
                f"{self.package.path.name}   "
                f"Sequences: {counts['sequences']}   Injections: {counts['injections']}   "
                f"Channels: {counts['channels']}   Audits: {counts['audits']}   Entries: {counts['entries']}"
            )
        )
        lines = [f"Package: {self.package.path}", ""]
        lines.extend(f"{key.replace('_', ' ').title()}: {value}" for key, value in counts.items())
        self._set_info("\n".join(lines))

    def _show_workspace_summary(self) -> None:
        if not self.loaded_packages:
            self.summary_label.config(text="No CMBX loaded. Choose the workspace package folder or add CMBX files.")
            lines = [
                "CMBX Data Explorer Workspace",
                "----------------------------",
                f"Workspace: {DEFAULT_APP_WORKSPACE}",
                "",
                f"packages: {DEFAULT_CMBX_SOURCE_FOLDER}",
                f"cache:    {DEFAULT_CACHE_FOLDER}",
                f"exports:  {DEFAULT_EXPORT_FOLDER}",
                f"db:       {DEFAULT_DB_FOLDER}",
                f"reports:  {DEFAULT_REPORT_FOLDER}",
                f"logs:     {DEFAULT_LOG_FOLDER}",
                "",
                "Put CMBX packages under packages, then click Scan CMBX.",
            ]
            self._set_info("\n".join(lines))
            return
        totals = {"sequences": 0, "injections": 0, "channels": 0, "audits": 0, "entries": 0}
        lines = [f"Loaded CMBX packages: {len(self.loaded_packages)}", ""]
        for package in self.loaded_packages:
            counts = summarize_package(package)
            for key in totals:
                totals[key] += int(counts.get(key, 0))
            lines.append(
                f"{package.path.name}: "
                f"Sequences {counts['sequences']}, Injections {counts['injections']}, "
                f"Channels {counts['channels']}, Audits {counts['audits']}"
            )
        self.summary_label.config(
            text=(
                f"Workspace: {len(self.loaded_packages)} CMBX   "
                f"Sequences: {totals['sequences']}   Injections: {totals['injections']}   "
                f"Channels: {totals['channels']}   Audits: {totals['audits']}"
            )
        )
        self._set_info("\n".join(lines))

    def _show_element_info(self, element: CmbxElement) -> None:
        lines = [
            f"Name: {element.name}",
            f"Type: {element.kind}",
            f"ItemType: {element.item_type}",
            f"URL: {element.url}",
            f"RawDataFilename: {element.raw_filename}",
            f"Filename: {element.filename}",
            f"Size: {element.size if element.size is not None else ''}",
            f"RawDataFileId: {element.raw_data_file_id}",
        ]
        if element.kind == "injection":
            link = get_injection_method_link(self.injection_method_links, element)
            lines.append("")
            lines.append("Instrument Method Link")
            lines.append("---------------------")
            if link:
                lines.append(f"Processing Method: {link.processing_method}")
                lines.append(f"Instrument Method: {link.instrument_method}")
                lines.append(f"Sequence Cmd Offset: {link.occurrence}")
                lines.append("")
                lines.append("Export")
                lines.append("------")
                lines.append("Select the Embedded Instrument Method row in Instrument Methods, then click Export Selected.")
                lines.append("When Chromeleon DLLs are available, export includes decoded XML plus TXT/TSV flow views.")
                external = self.external_instrument_methods.get(link.instrument_method)
                if external:
                    lines.extend(["", external.summary_text(max_lines=80)])
            else:
                lines.append("No method link was found in the sequence command object.")
        if element.kind in {"instrument_method", "processing_method", "report_template"}:
            try:
                summary = build_embedded_object_summary(self.package, element) if self.package else None
            except Exception as exc:
                lines.extend(["", f"Embedded Summary Error: {exc}"])
            else:
                if summary:
                    lines.extend(["", summary.to_text(max_section_chars=4500)])
            if element.kind == "instrument_method" and element.name in self.external_instrument_methods:
                lines.extend(["", self.external_instrument_methods[element.name].summary_text(max_lines=100)])
        self._set_info("\n".join(lines))

    def _set_info(self, text: str) -> None:
        self.info_text.configure(state="normal")
        self.info_text.delete("1.0", "end")
        self.info_text.insert("end", text)
        self.info_text.configure(state="disabled")

    def export_selected(self) -> None:
        if not self.package:
            messagebox.showinfo(APP_NAME, "Load a CMBX package first.")
            return
        selected_elements = self._selected_export_elements()
        if not selected_elements:
            messagebox.showinfo(APP_NAME, "Select raw channel rows, audit rows, CMBX method/report rows, or a tree node to export.")
            return
        output = Path(self.output_folder_var.get().strip())
        self.status_var.set(f"Exporting {len(selected_elements)} item(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths = export_elements(self.package, selected_elements, output)
                self._call_ui(lambda: self._export_done(paths))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def export_selected_channels(self) -> None:
        elements_by_package = self._selected_channel_table_elements_by_package()
        if elements_by_package:
            self._export_elements_by_package(elements_by_package)
            return
        self._export_selected_table(self.channel_table, "Select raw channel rows to export.")

    def apply_raw_channel_filter(self) -> None:
        rows = self._filtered_raw_channel_items()
        table = self._table_widget(self.channel_table)
        table.delete(*table.get_children())
        for package, sequence, injection, signal in rows:
            key = "|".join(self._signal_cache_key_for_package(package, signal))
            table.insert(
                "",
                "end",
                iid=f"rawfilter:{key}",
                values=(package.path.name, sequence.name, injection.name, signal.name, signal.size or "", signal.raw_filename),
            )
        self.status_var.set(f"Raw channel filter matched {len(rows)} channel(s)")

    def export_filtered_channels(self) -> None:
        rows = self._filtered_raw_channel_items()
        if not rows:
            messagebox.showinfo(APP_NAME, "No raw channels matched the current filter.")
            return
        grouped_map: dict[str, tuple[CmbxPackage, list[CmbxElement]]] = {}
        for package, _sequence, _injection, signal in rows:
            key = str(package.path)
            if key not in grouped_map:
                grouped_map[key] = (package, [])
            grouped_map[key][1].append(signal)
        self._export_elements_by_package(list(grouped_map.values()))

    def _filtered_raw_channel_items(self) -> list[tuple[CmbxPackage, CmbxElement, CmbxElement, CmbxElement]]:
        seq_filter = self.raw_filter_sequence_var.get().strip().lower()
        inj_filter = self.raw_filter_injection_var.get().strip().lower()
        ch_filter = self.raw_filter_channel_var.get().strip().lower()
        rows: list[tuple[CmbxPackage, CmbxElement, CmbxElement, CmbxElement]] = []
        packages = self.loaded_packages or ([self.package] if self.package else [])
        for package in [item for item in packages if item is not None]:
            for sequence in package.sequences:
                if seq_filter and seq_filter not in sequence.name.lower():
                    continue
                for injection in [child for child in sequence.children if child.kind == "injection"]:
                    if inj_filter and inj_filter not in injection.name.lower():
                        continue
                    for signal in [child for child in injection.children if child.kind == "signal"]:
                        if ch_filter and ch_filter not in signal.name.lower():
                            continue
                        rows.append((package, sequence, injection, signal))
        return rows

    def _selected_channel_table_elements_by_package(self) -> list[tuple[CmbxPackage, list[CmbxElement]]]:
        if not hasattr(self, "channel_table"):
            return []
        table = self._table_widget(self.channel_table)
        grouped_map: dict[str, tuple[CmbxPackage, list[CmbxElement]]] = {}
        for iid in table.selection():
            if not iid.startswith("rawfilter:"):
                continue
            key = iid.removeprefix("rawfilter:")
            for package in self.loaded_packages:
                for element in package.elements_by_id.values():
                    if element.kind == "signal" and "|".join(self._signal_cache_key_for_package(package, element)) == key:
                        package_key = str(package.path)
                        if package_key not in grouped_map:
                            grouped_map[package_key] = (package, [])
                        grouped_map[package_key][1].append(element)
                        break
        return list(grouped_map.values())

    def _export_elements_by_package(self, grouped: list[tuple[CmbxPackage, list[CmbxElement]]], open_after: bool = False, show_message: bool = True) -> None:
        if not grouped:
            messagebox.showinfo(APP_NAME, "No exportable items were selected.")
            return
        output = Path(self.output_folder_var.get().strip())
        total = sum(len(items) for _package, items in grouped)
        self.status_var.set(f"Exporting {total} item(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                for package, elements in grouped:
                    paths.extend(export_elements(package, elements, output))
                self._call_ui(lambda: self._export_done(paths, open_after=open_after, show_message=show_message))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def export_selected_audits(self) -> None:
        if self._export_selected_audit_previews(open_after=False):
            return
        self._export_selected_table(self.audit_table, "Select audit trail rows to export.")

    def export_selected_methods(self) -> None:
        if not self.package:
            messagebox.showinfo(APP_NAME, "Load a CMBX package first.")
            return
        table = self._table_widget(self.method_context_table)
        selected = list(table.selection())
        if not selected:
            messagebox.showinfo(APP_NAME, "Select instrument method rows to export.")
            return
        selected_elements: list[CmbxElement] = []
        preview_jobs: list[tuple[str, list[tuple[str, str, str, str, str, str, str]]]] = []
        for iid in selected:
            if not iid.startswith("context:external:"):
                element = self.package.elements_by_id.get(iid)
                if element and element.kind == "instrument_method":
                    selected_elements.append(element)
            result = self._method_preview_rows_for_iid(iid)
            if result:
                preview_jobs.append(result)
        if not selected_elements and not preview_jobs:
            messagebox.showinfo(APP_NAME, "Select instrument method rows to export.")
            return
        export_format = self._ask_method_export_format()
        if not export_format:
            return
        if export_format == "bundle" and not selected_elements:
            messagebox.showinfo(APP_NAME, "Decode Bundle is only available for embedded CMBX instrument methods.")
            return
        if export_format in {"md", "xlsx"} and not preview_jobs:
            messagebox.showinfo(APP_NAME, "No previewable method script rows were found for the selected method(s).")
            return
        package = self.package
        output = Path(self.output_folder_var.get().strip())
        if export_format == "md":
            preview_output = self._preview_export_folder(package, "instrument_method_md")
        elif export_format == "xlsx":
            preview_output = self._preview_export_folder(package, "instrument_method_previews")
        else:
            preview_output = output
        self.status_var.set(f"Exporting {len(selected)} instrument method selection(s) as {export_format.upper()}...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                if export_format == "bundle" and selected_elements:
                    paths.extend(export_elements(package, selected_elements, output))
                elif export_format == "xlsx":
                    for title, rows in preview_jobs:
                        path = preview_output / f"{self._safe_temp_stem(title)}.xlsx"
                        self._write_method_preview_workbook(path, title, rows)
                        paths.append(path)
                elif export_format == "md":
                    for title, rows in preview_jobs:
                        path = preview_output / f"{self._safe_temp_stem(title)}.md"
                        self._write_method_preview_markdown(path, title, rows)
                        paths.append(path)
                self._call_ui(lambda: self._export_done(paths))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _ask_method_export_format(self) -> str | None:
        dialog = tk.Toplevel(self.root)
        dialog.title("Export Instrument Method")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=self.colors["card"])
        dialog.resizable(False, False)
        choice = tk.StringVar(value="")
        tk.Label(
            dialog,
            text="Choose one export format",
            font=self._font(11, "bold"),
            bg=self.colors["card"],
            fg=self.colors["text"],
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=14, pady=(12, 8))
        tk.Label(
            dialog,
            text="MD follows CM_METHOD_SCRIPT_MD_FORMAT_SPEC. Decode Bundle exports raw/bin/xml/flow evidence.",
            font=self._font(9),
            bg=self.colors["card"],
            fg=self.colors["text_secondary"],
            wraplength=460,
            justify="left",
        ).grid(row=1, column=0, columnspan=3, sticky="w", padx=14, pady=(0, 12))

        def set_choice(value: str) -> None:
            choice.set(value)
            dialog.destroy()

        self._make_button(dialog, "MD Script", lambda: set_choice("md"), kind="primary", width=16).grid(row=2, column=0, padx=(14, 8), pady=(0, 14))
        self._make_button(dialog, "XLSX Preview", lambda: set_choice("xlsx"), kind="secondary", width=16).grid(row=2, column=1, padx=8, pady=(0, 14))
        self._make_button(dialog, "Decode Bundle", lambda: set_choice("bundle"), kind="neutral", width=16).grid(row=2, column=2, padx=(8, 14), pady=(0, 14))
        dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)
        dialog.update_idletasks()
        x = self.root.winfo_rootx() + max(0, (self.root.winfo_width() - dialog.winfo_width()) // 2)
        y = self.root.winfo_rooty() + max(0, (self.root.winfo_height() - dialog.winfo_height()) // 2)
        dialog.geometry(f"+{x}+{y}")
        self.root.wait_window(dialog)
        return choice.get() or None

    def export_selected_processing(self) -> None:
        self._export_selected_table(self.processing_method_table, "Select processing method rows to export.")

    def export_selected_templates(self) -> None:
        self._export_selected_table(self.report_template_table, "Select report template rows to export.")

    def open_selected_channels(self, _event=None) -> None:
        grouped = self._selected_channel_table_elements_by_package()
        if grouped:
            self._export_elements_by_package(grouped, open_after=True, show_message=False)
            return
        self._export_selected_table(self.channel_table, "Select a raw channel row to open.", open_after=True, show_message=False)

    def preview_selected_channel(self, _event=None) -> None:
        if not self.package or not hasattr(self, "channel_plot_canvas"):
            return
        table = self._table_widget(self.channel_table)
        selected = table.selection()
        if not selected:
            self._set_channel_plot_message("Select a raw channel row to preview.")
            return
        element = self.package.elements_by_id.get(selected[0])
        if not element or element.kind != "signal":
            self._set_channel_plot_message("The selected row is not a raw signal channel.")
            return
        self.status_var.set(f"Decoding raw channel preview: {element.name}...")
        try:
            points = self._decode_signal_preview_points(element)
            self.channel_plot_points = points
            self.channel_plot_title = element.name
            self._redraw_channel_plot()
            self.status_var.set(f"Previewed raw channel: {element.name} ({len(points)} points)")
        except Exception as exc:
            self.channel_plot_points = []
            self.channel_plot_title = ""
            self._set_channel_plot_message(str(exc))
            self.status_var.set("Raw channel preview failed")

    def _decode_signal_preview_points(self, signal: CmbxElement) -> list[tuple[float, float]]:
        if not self.package:
            return []
        return self._decode_signal_points(self.package, signal)

    def _decode_signal_points(self, package: CmbxPackage, signal: CmbxElement) -> list[tuple[float, float]]:
        if not signal.raw_filename:
            raise ValueError(f"Signal element has no raw file: {signal.name}")
        cache_key = self._signal_cache_key_for_package(package, signal)
        cached = self.signal_points_cache.get(cache_key)
        if cached is not None:
            return cached
        temp_dir = Path(tempfile.gettempdir()) / "cmbx_data_explorer_signal_preview"
        temp_dir.mkdir(parents=True, exist_ok=True)
        stem = self._safe_temp_stem("|".join(cache_key))
        raw_path = temp_dir / f"{stem}.raw"
        tsv_path = temp_dir / f"{stem}.tsv"
        raw_path.write_bytes(extract_cmbx_entry(package.path, signal.raw_filename))
        export_signal_raw(raw_path, tsv_path, signal.name)
        points: list[tuple[float, float]] = []
        for line in tsv_path.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
            parts = line.split("\t")
            if len(parts) < 3 or parts[0] in {"Channel", "Time (min)"}:
                continue
            try:
                points.append((float(parts[0]), float(parts[2])))
            except ValueError:
                continue
        self.signal_points_cache[cache_key] = points
        return points

    def _signal_cache_key(self, signal: CmbxElement) -> tuple[str, str, str]:
        if not self.package:
            return ("", signal.id or signal.name, signal.raw_filename or "")
        return self._signal_cache_key_for_package(self.package, signal)

    def _signal_cache_key_for_package(self, package: CmbxPackage, signal: CmbxElement) -> tuple[str, str, str]:
        return (str(package.path), signal.id or signal.name, signal.raw_filename or "")

    def _signal_star_for_package(self, package: CmbxPackage, signal: CmbxElement) -> str:
        return "*" if "|".join(self._signal_cache_key_for_package(package, signal)) in self.starred_signal_keys else ""

    def add_tree_channel_to_raw_plot(self, package: CmbxPackage, signal: CmbxElement) -> None:
        self.add_tree_channels_to_raw_plot([(package, signal)])

    def add_tree_channels_to_raw_plot(self, items: list[tuple[CmbxPackage, CmbxElement]]) -> None:
        signals = [(package, signal) for package, signal in items if signal.kind == "signal"]
        if not signals:
            return
        existing = {str(item["key"]) for item in self.raw_plot_series}
        added_keys: list[str] = []
        skipped = 0
        errors: list[str] = []
        for package, signal in signals:
            key = "|".join(self._signal_cache_key_for_package(package, signal))
            if key in existing:
                skipped += 1
                continue
            try:
                points = self._decode_signal_points(package, signal)
            except Exception as exc:
                errors.append(f"{signal.name}: {exc}")
                continue
            if not points:
                skipped += 1
                continue
            self.raw_plot_series.append(
                {
                    "key": key,
                    "name": f"{package.path.stem} / {self._parent_injection_name_for_package(package, signal)} / {signal.name}",
                    "points": points,
                }
            )
            existing.add(key)
            added_keys.append(key)
        if not self.raw_plot_benchmark_key and self.raw_plot_series:
            self.raw_plot_benchmark_key = str(self.raw_plot_series[0]["key"])
        self._set_raw_plot_selection_rows(self._current_raw_plot_items(), selected_keys=set(added_keys) if added_keys else None)
        self._draw_raw_plot_series()
        self.notebook.select(self.raw_plot_canvas.master)
        if errors:
            messagebox.showwarning(APP_NAME, "Some raw channels could not be added:\n" + "\n".join(errors[:8]))
        self.status_var.set(f"Added {len(added_keys)} raw channel(s) to plot selection" + (f"; skipped {skipped}" if skipped else ""))

    def plot_selected_raw_channels(self) -> None:
        if not self.loaded_packages:
            messagebox.showinfo(APP_NAME, "Load a CMBX package first.")
            return
        items = self._selected_signal_items_for_plot()
        if not items:
            messagebox.showinfo(APP_NAME, "Select one or more raw channel rows or signal nodes first.")
            return
        self.status_var.set(f"Preparing {len(items)} raw channel plot(s)...")
        series: list[dict[str, object]] = []
        for package, signal in items:
            try:
                points = self._decode_signal_points(package, signal)
            except Exception as exc:
                self.status_var.set(f"Raw plot skipped {signal.name}: {exc}")
                continue
            if points:
                key = "|".join(self._signal_cache_key_for_package(package, signal))
                series.append(
                    {
                        "key": key,
                        "name": f"{package.path.stem} / {self._parent_injection_name_for_package(package, signal)} / {signal.name}",
                        "points": points,
                    }
                )
        self.raw_plot_series = series
        if not self.raw_plot_benchmark_key and series:
            self.raw_plot_benchmark_key = str(series[0]["key"])
        self._set_raw_plot_selection_rows(items, selected_keys={str(item["key"]) for item in series})
        self._draw_raw_plot_series()
        self.notebook.select(self.raw_plot_canvas.master)
        self.status_var.set(f"Plotted {len(series)} raw channel(s)")

    def plot_selected_raw_plot_channels(self) -> None:
        if not self.raw_plot_series:
            messagebox.showinfo(APP_NAME, "Add one or more raw channels to Selected Channels first.")
            return
        if not self._selected_raw_plot_keys():
            messagebox.showinfo(APP_NAME, "Select one or more rows in Selected Channels first.")
            return
        self._draw_raw_plot_series()
        self.status_var.set(f"Plotted {len(self._selected_raw_plot_series())} selected raw channel(s)")

    def plot_starred_raw_channels(self) -> None:
        starred = self._starred_signal_elements()
        series: list[dict[str, object]] = []
        for package, signal in starred:
            try:
                points = self._decode_signal_points(package, signal)
            except Exception as exc:
                self.status_var.set(f"Raw plot skipped {signal.name}: {exc}")
                continue
            if points:
                key = "|".join(self._signal_cache_key_for_package(package, signal))
                series.append(
                    {
                        "key": key,
                        "name": f"{package.path.stem} / {self._parent_injection_name_for_package(package, signal)} / {signal.name}",
                        "points": points,
                    }
                )
        self.raw_plot_series = series
        if self.raw_plot_benchmark_key not in {str(item["key"]) for item in series}:
            self.raw_plot_benchmark_key = str(series[0]["key"]) if series else ""
        self._set_raw_plot_selection_rows(starred, selected_keys={str(item["key"]) for item in series})
        self._draw_raw_plot_series()
        if series:
            self.notebook.select(self.raw_plot_canvas.master)
        self.status_var.set(f"Plotted {len(series)} starred raw channel(s)")

    def set_raw_plot_benchmark(self) -> None:
        selected_keys = self._selected_raw_plot_keys()
        if selected_keys:
            self.raw_plot_benchmark_key = selected_keys[0]
            label = self._raw_plot_label_for_key(selected_keys[0])
            self._set_raw_plot_selection_rows(self._current_raw_plot_items(), selected_keys=set(selected_keys))
            self._draw_raw_plot_series()
            self.status_var.set(f"Benchmark channel: {label}")
            return
        items = self._selected_signal_items_for_plot()
        if not items:
            messagebox.showinfo(APP_NAME, "Select a raw channel to use as benchmark.")
            return
        package, signal = items[0]
        self.raw_plot_benchmark_key = "|".join(self._signal_cache_key_for_package(package, signal))
        if not self.raw_plot_series:
            self.add_tree_channels_to_raw_plot(items)
            return
        self._set_raw_plot_selection_rows(self._current_raw_plot_items())
        self._draw_raw_plot_series()
        self.status_var.set(f"Benchmark channel: {package.path.stem} / {self._parent_injection_name_for_package(package, signal)} / {signal.name}")

    def clear_raw_plot(self) -> None:
        self.raw_plot_series = []
        self.raw_plot_benchmark_key = ""
        self._set_raw_plot_selection_rows([])
        self._draw_raw_plot_series()

    def _selected_signal_items_for_plot(self) -> list[tuple[CmbxPackage, CmbxElement]]:
        if not self.loaded_packages:
            return []
        signals: list[tuple[CmbxPackage, CmbxElement]] = []
        seen: set[str] = set()
        for iid in self.package_tree.selection():
            package, element = self.tree_item_context.get(iid, (None, None))
            if package and element and element.kind == "signal":
                key = "|".join(self._signal_cache_key_for_package(package, element))
                if key not in seen:
                    seen.add(key)
                    signals.append((package, element))
        channel_table = self._table_widget(self.channel_table)
        for iid in channel_table.selection():
            if iid.startswith("rawfilter:"):
                key = iid.removeprefix("rawfilter:")
                for package in self.loaded_packages:
                    element = next(
                        (
                            item
                            for item in package.elements_by_id.values()
                            if item.kind == "signal" and "|".join(self._signal_cache_key_for_package(package, item)) == key
                        ),
                        None,
                    )
                    if element and key not in seen:
                        seen.add(key)
                        signals.append((package, element))
                        break
                continue
            if not self.package:
                continue
            element = self.package.elements_by_id.get(iid)
            if element and element.kind == "signal":
                key = "|".join(self._signal_cache_key_for_package(self.package, element))
                if key not in seen:
                    seen.add(key)
                    signals.append((self.package, element))
        return signals

    def _selected_signal_elements_for_plot(self) -> list[CmbxElement]:
        return [signal for _package, signal in self._selected_signal_items_for_plot()]

    def _starred_signal_elements(self) -> list[tuple[CmbxPackage, CmbxElement]]:
        selected: list[tuple[CmbxPackage, CmbxElement]] = []
        for package in self.loaded_packages:
            for element in package.elements_by_id.values():
                if element.kind != "signal":
                    continue
                if "|".join(self._signal_cache_key_for_package(package, element)) in self.starred_signal_keys:
                    selected.append((package, element))
        return selected

    def _current_raw_plot_items(self) -> list[tuple[CmbxPackage, CmbxElement]]:
        keys = {str(item["key"]) for item in self.raw_plot_series}
        items: list[tuple[CmbxPackage, CmbxElement]] = []
        for package in self.loaded_packages:
            for element in package.elements_by_id.values():
                if element.kind != "signal":
                    continue
                if "|".join(self._signal_cache_key_for_package(package, element)) in keys:
                    items.append((package, element))
        return items

    def _set_raw_plot_selection_rows(self, items: list[tuple[CmbxPackage, CmbxElement]], selected_keys: set[str] | None = None) -> None:
        if not hasattr(self, "raw_plot_selection_table"):
            return
        table = self._table_widget(self.raw_plot_selection_table)
        previous_selection = selected_keys if selected_keys is not None else set(self._selected_raw_plot_keys())
        table.delete(*table.get_children())
        iids_to_select: list[str] = []
        for package, signal in items:
            key = "|".join(self._signal_cache_key_for_package(package, signal))
            sequence_name, injection_name = self._signal_location_for_package(package, signal)
            iid = f"rawplot:{key}"
            table.insert(
                "",
                "end",
                iid=iid,
                values=("Yes" if key == self.raw_plot_benchmark_key else "", package.path.name, sequence_name, injection_name, signal.name),
            )
            if key in previous_selection:
                iids_to_select.append(iid)
        if iids_to_select:
            table.selection_set(*iids_to_select)

    def _selected_raw_plot_keys(self) -> list[str]:
        if not hasattr(self, "raw_plot_selection_table"):
            return []
        table = self._table_widget(self.raw_plot_selection_table)
        keys: list[str] = []
        for iid in table.selection():
            if iid.startswith("rawplot:"):
                keys.append(iid.removeprefix("rawplot:"))
        return keys

    def _selected_raw_plot_series(self) -> list[dict[str, object]]:
        selected = set(self._selected_raw_plot_keys())
        if not selected:
            return []
        return [item for item in self.raw_plot_series if str(item["key"]) in selected]

    def _raw_plot_label_for_key(self, key: str) -> str:
        return next((str(item["name"]) for item in self.raw_plot_series if str(item["key"]) == key), key)

    def remove_selected_raw_plot_channels(self) -> None:
        selected = set(self._selected_raw_plot_keys())
        if not selected:
            messagebox.showinfo(APP_NAME, "Select one or more rows in Selected Channels first.")
            return
        self.raw_plot_series = [item for item in self.raw_plot_series if str(item["key"]) not in selected]
        if self.raw_plot_benchmark_key in selected:
            self.raw_plot_benchmark_key = str(self.raw_plot_series[0]["key"]) if self.raw_plot_series else ""
        self._set_raw_plot_selection_rows(self._current_raw_plot_items())
        self._draw_raw_plot_series()
        self.status_var.set(f"Removed {len(selected)} raw channel(s) from plot selection")

    def _show_raw_plot_selection_menu(self, event) -> None:
        if not hasattr(self, "raw_plot_selection_table"):
            return
        table = self._table_widget(self.raw_plot_selection_table)
        row_iid = table.identify_row(event.y)
        if row_iid and row_iid not in table.selection():
            table.selection_set(row_iid)
        menu = tk.Menu(self.root, tearoff=0)
        selected_count = len(self._selected_raw_plot_keys())
        menu.add_command(label=f"Plot Selected Channels ({selected_count})", command=self.plot_selected_raw_plot_channels, state="normal" if selected_count else "disabled")
        menu.add_command(label="Set Benchmark From Selection", command=self.set_raw_plot_benchmark, state="normal" if selected_count else "disabled")
        menu.add_separator()
        menu.add_command(label=f"Remove Selected ({selected_count})", command=self.remove_selected_raw_plot_channels, state="normal" if selected_count else "disabled")
        menu.add_command(label="Clear Plot", command=self.clear_raw_plot, state="normal" if self.raw_plot_series else "disabled")
        menu.tk_popup(event.x_root, event.y_root)

    def _signal_location_for_package(self, package: CmbxPackage, signal: CmbxElement) -> tuple[str, str]:
        injection = package.elements_by_id.get(signal.parent_id or "")
        sequence = package.elements_by_id.get(injection.parent_id or "") if injection else None
        return (sequence.name if sequence else "", injection.name if injection else "")

    def _draw_raw_plot_series(self) -> None:
        if not hasattr(self, "raw_plot_canvas"):
            return
        canvas = self.raw_plot_canvas
        canvas.delete("all")
        if not self.raw_plot_series:
            canvas.create_text(24, 24, anchor="nw", text="Add raw channels from Sequence Data, then select rows in Selected Channels.", fill=self.colors["text_secondary"], font=self._font(10))
            return
        visible_series = self._selected_raw_plot_series()
        if not visible_series:
            canvas.create_text(24, 24, anchor="nw", text="Select one or more rows in Selected Channels to draw them.", fill=self.colors["text_secondary"], font=self._font(10))
            return
        width = max(260, canvas.winfo_width())
        height = max(180, canvas.winfo_height())
        margin_left, margin_right, margin_top, margin_bottom = 62, 150, 36, 42
        plot_w = max(1, width - margin_left - margin_right)
        plot_h = max(1, height - margin_top - margin_bottom)
        sampled_series = []
        for item in visible_series:
            points = self._sample_plot_points(item["points"], 1400)  # type: ignore[arg-type]
            if points:
                sampled_series.append((item, points))
        xs = [point[0] for _item, points in sampled_series for point in points]
        ys = [point[1] for _item, points in sampled_series for point in points]
        if not xs or not ys:
            canvas.create_text(24, 24, anchor="nw", text="No plottable raw data points.", fill=self.colors["text_secondary"], font=self._font(10))
            return
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)
        if max_x == min_x:
            max_x = min_x + 1
        if max_y == min_y:
            max_y = min_y + 1
        y_pad = (max_y - min_y) * 0.06
        min_y -= y_pad
        max_y += y_pad
        canvas.create_rectangle(margin_left, margin_top, margin_left + plot_w, margin_top + plot_h, outline="#CBD5E1", fill="#FFFFFF")
        canvas.create_text(margin_left, 14, anchor="w", text="Raw Channel Comparison", fill=self.colors["text"], font=self._font(10, "bold"))
        for index in range(5):
            y = margin_top + plot_h * index / 4
            value = max_y - (max_y - min_y) * index / 4
            canvas.create_line(margin_left, y, margin_left + plot_w, y, fill="#E5E7EB")
            canvas.create_text(margin_left - 8, y, anchor="e", text=f"{value:.3g}", fill=self.colors["text_secondary"], font=self._font(8))
        for index in range(5):
            x = margin_left + plot_w * index / 4
            value = min_x + (max_x - min_x) * index / 4
            canvas.create_line(x, margin_top, x, margin_top + plot_h, fill="#F1F5F9")
            canvas.create_text(x, margin_top + plot_h + 16, anchor="n", text=f"{value:.3g}", fill=self.colors["text_secondary"], font=self._font(8))
        palette = ["#2563EB", "#16A34A", "#DC2626", "#9333EA", "#EA580C", "#0891B2", "#4F46E5", "#BE123C"]
        for index, (item, points) in enumerate(sampled_series):
            key = str(item["key"])
            color = "#111827" if key == self.raw_plot_benchmark_key else palette[index % len(palette)]
            width_px = 3 if key == self.raw_plot_benchmark_key else 2
            coords: list[float] = []
            for x_value, y_value in points:
                x = margin_left + (x_value - min_x) / (max_x - min_x) * plot_w
                y = margin_top + (max_y - y_value) / (max_y - min_y) * plot_h
                coords.extend([x, y])
            if len(coords) >= 4:
                canvas.create_line(*coords, fill=color, width=width_px, smooth=False)
            legend_y = margin_top + 20 * index
            canvas.create_line(width - margin_right + 16, legend_y, width - margin_right + 42, legend_y, fill=color, width=width_px)
            label = str(item["name"])
            if key == self.raw_plot_benchmark_key:
                label = "BENCHMARK  " + label
            canvas.create_text(width - margin_right + 48, legend_y, anchor="w", text=label[:42], fill=self.colors["text"], font=self._font(8, "bold" if key == self.raw_plot_benchmark_key else "normal"))

    def _redraw_channel_plot(self) -> None:
        if not hasattr(self, "channel_plot_canvas"):
            return
        canvas = self.channel_plot_canvas
        canvas.delete("all")
        if not self.channel_plot_points:
            self._set_channel_plot_message("Select a raw channel row to preview.")
            return
        width = max(200, canvas.winfo_width())
        height = max(160, canvas.winfo_height())
        margin_left, margin_right, margin_top, margin_bottom = 54, 18, 34, 38
        plot_w = max(1, width - margin_left - margin_right)
        plot_h = max(1, height - margin_top - margin_bottom)
        points = self._sample_plot_points(self.channel_plot_points, 1800)
        xs = [item[0] for item in points]
        ys = [item[1] for item in points]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)
        if max_x == min_x:
            max_x = min_x + 1
        if max_y == min_y:
            max_y = min_y + 1
        canvas.create_rectangle(margin_left, margin_top, margin_left + plot_w, margin_top + plot_h, outline="#CBD5E1", fill="#FFFFFF")
        canvas.create_text(margin_left, 12, anchor="w", text=self.channel_plot_title, fill=self.colors["text"], font=self._font(9, "bold"))
        for index in range(5):
            y = margin_top + plot_h * index / 4
            value = max_y - (max_y - min_y) * index / 4
            canvas.create_line(margin_left, y, margin_left + plot_w, y, fill="#E5E7EB")
            canvas.create_text(margin_left - 8, y, anchor="e", text=f"{value:.3g}", fill=self.colors["text_secondary"], font=self._font(8))
        for index in range(5):
            x = margin_left + plot_w * index / 4
            value = min_x + (max_x - min_x) * index / 4
            canvas.create_line(x, margin_top + plot_h, x, margin_top + plot_h + 4, fill="#94A3B8")
            canvas.create_text(x, margin_top + plot_h + 18, anchor="n", text=f"{value:.3g}", fill=self.colors["text_secondary"], font=self._font(8))
        coords: list[float] = []
        for x_value, y_value in points:
            x = margin_left + (x_value - min_x) / (max_x - min_x) * plot_w
            y = margin_top + plot_h - (y_value - min_y) / (max_y - min_y) * plot_h
            coords.extend([x, y])
        if len(coords) >= 4:
            canvas.create_line(*coords, fill="#2563EB", width=1.5)
        canvas.create_text(margin_left + plot_w / 2, height - 7, anchor="s", text="Time (min)", fill=self.colors["text_secondary"], font=self._font(8, "bold"))

    def _sample_plot_points(self, points: list[tuple[float, float]], limit: int) -> list[tuple[float, float]]:
        if len(points) <= limit:
            return points
        step = max(1, len(points) // limit)
        sampled = points[::step]
        return sampled if sampled[-1] == points[-1] else sampled + [points[-1]]

    def _set_channel_plot_message(self, message: str) -> None:
        if not hasattr(self, "channel_plot_canvas"):
            return
        canvas = self.channel_plot_canvas
        canvas.delete("all")
        canvas.create_text(max(120, canvas.winfo_width() // 2), max(80, canvas.winfo_height() // 2), text=message, fill=self.colors["text_secondary"], font=self._font(9), width=max(240, canvas.winfo_width() - 40))

    def open_selected_audits(self, _event=None) -> None:
        if self._export_selected_audit_previews(open_after=True, show_message=False):
            return
        self._export_selected_table(self.audit_table, "Select an audit trail row to open.", open_after=True, show_message=False)

    def preview_selected_audit(self, _event=None) -> None:
        if not self.package or not hasattr(self, "audit_preview_table"):
            return
        table = self._table_widget(self.audit_table)
        selected = table.selection()
        if not selected:
            self._set_audit_preview_message("Select an audit trail row to preview.")
            return
        audit = self.package.elements_by_id.get(selected[0])
        if not audit or audit.kind != "audit":
            self._set_audit_preview_message("The selected row is not an audit trail.")
            return
        self.status_var.set(f"Decoding audit preview: {audit.name}...")
        try:
            rows = self._decode_audit_preview_rows(audit)
            self._set_audit_preview_rows(rows)
            self.status_var.set(f"Previewed audit trail: {self._parent_injection_name(audit) or audit.name}")
        except Exception as exc:
            self._set_audit_preview_message(str(exc))
            self.status_var.set("Audit preview failed")

    def _decode_audit_preview_rows(self, audit: CmbxElement) -> list[tuple[str, str, str, str, str]]:
        if not self.package:
            return []
        if not audit.raw_filename:
            raise ValueError(f"Audit element has no raw file: {audit.name}")
        temp_dir = Path(tempfile.gettempdir()) / "cmbx_data_explorer_audit_preview"
        temp_dir.mkdir(parents=True, exist_ok=True)
        stem = self._safe_temp_stem(audit.id or audit.name)
        raw_path = temp_dir / f"{stem}.raw"
        tsv_path = temp_dir / f"{stem}_audit.tsv"
        raw_path.write_bytes(extract_cmbx_entry(self.package.path, audit.raw_filename))
        export_audit_raw(raw_path, tsv_path)
        records = self._read_audit_preview_tsv(tsv_path)
        start_time = self._infer_audit_start_time(records)
        if start_time:
            for record in records:
                if not record["day_time"]:
                    record["day_time"] = self._audit_day_time_from_retention(start_time, record["ret_time"])
        injection = self.package.elements_by_id.get(audit.parent_id or "")
        injection_name = injection.name if injection else self._parent_injection_name(audit)
        link = get_injection_method_link(self.injection_method_links, injection or injection_name)
        run_time = self._audit_run_time(records)
        injection_time = self._format_audit_injection_time(start_time)
        metadata = [
            ("4", "Injection Name:", "", injection_name, "audit_meta"),
            ("5", "Injection Type:", "", "Unknown", "audit_meta"),
            ("6", "Instrument Method:", "", link.instrument_method if link else "", "audit_meta"),
            ("7", "Processing Method:", "", link.processing_method if link else "", "audit_meta"),
            ("8", "Injection Time:", "", injection_time, "audit_meta"),
            ("9", "Run Time (min):", "", run_time, "audit_meta"),
            ("10", "Channel:", "", audit.name, "audit_meta"),
        ]
        rows: list[tuple[str, str, str, str, str]] = [
            ("1", "", "", "", ""),
            ("2", "", "Audit Trail", "", "audit_title"),
            ("3", "", "", "", ""),
            *metadata,
            ("11", "", "", "", ""),
            ("12", "", "", "", ""),
            ("13", "", "", "", ""),
            ("14", "Day Time", "Ret. Time", "Command/Message", "audit_header"),
        ]
        for index, record in enumerate(records, 15):
            rows.append((str(index), record["day_time"], record["ret_time"], record["message"], self._audit_preview_tag(record["message"])))
        return rows

    def _read_audit_preview_tsv(self, path: Path) -> list[dict[str, str]]:
        text = path.read_text(encoding="utf-8-sig", errors="ignore")
        rows: list[dict[str, str]] = []
        for line in text.splitlines():
            parts = line.split("\t")
            if len(parts) < 7 or parts[0] == "Index":
                continue
            message = self._audit_command_message(parts)
            rows.append({"day_time": self._format_audit_day_time(parts[7] if len(parts) > 7 else ""), "ret_time": self._format_ret_time(parts[1]), "message": message})
        return rows

    def _audit_command_message(self, parts: list[str]) -> str:
        device = parts[2].strip()
        message = parts[3].strip()
        trigger = parts[4].strip()
        property_name = parts[5].strip()
        property_value = parts[6].strip()
        if message:
            return message
        if property_name:
            left = f"{device}.{property_name}" if device else property_name
            return f"{left} = {property_value}" if property_value else left
        if trigger:
            return f"Trigger {trigger}"
        return ""

    def _format_ret_time(self, value: str) -> str:
        try:
            return f"{float(value):.3f}"
        except ValueError:
            return ""

    def _infer_audit_start_time(self, records: list[dict[str, str]]) -> datetime | None:
        for record in records:
            parsed = self._parse_audit_day_time(record["day_time"])
            if not parsed:
                continue
            try:
                ret_time = float(record["ret_time"])
            except ValueError:
                continue
            return parsed - timedelta(minutes=ret_time)
        return None

    def _audit_day_time_from_retention(self, start_time: datetime, ret_time_text: str) -> str:
        try:
            value = float(ret_time_text)
        except ValueError:
            return ""
        return (start_time + timedelta(minutes=value)).strftime("%H:%M:%S")

    def _format_audit_day_time(self, value: str) -> str:
        parsed = self._parse_audit_day_time(value)
        return parsed.strftime("%H:%M:%S") if parsed else ""

    def _format_audit_injection_time(self, start_time: datetime | None) -> str:
        if not start_time:
            return ""
        if start_time.year == 1900:
            return start_time.strftime("%H:%M:%S")
        return start_time.strftime("%m/%d/%Y %H:%M:%S")

    def _parse_audit_day_time(self, value: str) -> datetime | None:
        text = value.strip()
        if not text:
            return None
        normalized = text.replace("T", " ").replace("Z", "")
        if "." in normalized:
            normalized = normalized.split(".", 1)[0]
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%H:%M:%S",
            "%H:%M",
        ]
        for fmt in formats:
            try:
                parsed = datetime.strptime(normalized, fmt)
                return parsed
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None

    def _audit_run_time(self, records: list[dict[str, str]]) -> str:
        values = []
        for record in records:
            try:
                values.append(float(record["ret_time"]))
            except ValueError:
                continue
        return f"{max(values):.2f}" if values else ""

    def _audit_preview_tag(self, message: str) -> str:
        lower = message.lower()
        if "entered stage" in lower:
            return "audit_stage"
        if lower.startswith("if ") or lower.startswith("else") or lower.startswith("end if") or ": yes" in lower or ": no" in lower:
            return "audit_condition"
        return ""

    def _clear_audit_preview(self) -> None:
        if not hasattr(self, "audit_preview_table"):
            return
        table = self._table_widget(self.audit_preview_table)
        table.delete(*table.get_children())

    def _set_audit_preview_message(self, message: str) -> None:
        self._clear_audit_preview()
        table = self._table_widget(self.audit_preview_table)
        table.insert("", "end", values=("", "", "", message))

    def _set_audit_preview_rows(self, rows: list[tuple[str, str, str, str, str]]) -> None:
        self._clear_audit_preview()
        table = self._table_widget(self.audit_preview_table)
        if not rows:
            self._set_audit_preview_message("No audit records were found.")
            return
        for row, day_time, ret_time, message, tag in rows:
            table.insert("", "end", values=(row, day_time, ret_time, message), tags=(tag,) if tag else ())

    def _export_selected_audit_previews(self, open_after: bool = False, show_message: bool = True) -> bool:
        if not self.package:
            return False
        table = self._table_widget(self.audit_table)
        audits = [self.package.elements_by_id.get(iid) for iid in table.selection()]
        audits = [audit for audit in audits if audit and audit.kind == "audit"]
        if not audits:
            return False
        package = self.package
        output = self._preview_export_folder(package, "audit_previews")
        self.status_var.set(f"Exporting {len(audits)} styled audit preview workbook(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                for audit in audits:
                    rows = self._decode_audit_preview_rows(audit)
                    injection = package.elements_by_id.get(audit.parent_id or "")
                    stem = self._safe_temp_stem(f"{injection.name if injection else audit.name}_audit")
                    path = output / f"{stem}.xlsx"
                    self._write_audit_preview_workbook(path, rows)
                    paths.append(path)
                self._call_ui(lambda: self._export_done(paths, open_after=open_after, show_message=show_message))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()
        return True

    def _write_audit_preview_workbook(self, path: Path, rows: list[tuple[str, str, str, str, str]]) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Audit Trail"
        thin = Side(style="thin", color="C7C7C7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fills = {
            "audit_header": PatternFill("solid", fgColor="E7E7EC"),
            "audit_title": PatternFill("solid", fgColor="FFFFFF"),
            "audit_stage": PatternFill("solid", fgColor="F9C18B"),
            "audit_condition": PatternFill("solid", fgColor="BFF29A"),
        }
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 120
        sheet.freeze_panes = "A15"
        for row_text, day_time, ret_time, message, tag in rows:
            try:
                row_index = max(1, int(row_text))
            except ValueError:
                row_index = sheet.max_row + 1
            if tag == "audit_meta":
                sheet.cell(row=row_index, column=1, value=day_time)
                sheet.cell(row=row_index, column=3, value=message)
                sheet.cell(row=row_index, column=1).font = Font(name=UI_FONT_FAMILY, size=11, italic=True)
                sheet.cell(row=row_index, column=3).font = Font(name=UI_FONT_FAMILY, size=11, bold=True)
            elif tag == "audit_title":
                sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
                sheet.cell(row=row_index, column=1, value=ret_time or "Audit Trail")
                sheet.cell(row=row_index, column=1).font = Font(name=UI_FONT_FAMILY, size=16, bold=True)
                sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center")
            else:
                sheet.cell(row=row_index, column=1, value=day_time)
                sheet.cell(row=row_index, column=2, value=ret_time)
                sheet.cell(row=row_index, column=3, value=message)
            if tag in fills:
                for col in range(1, 4):
                    sheet.cell(row=row_index, column=col).fill = fills[tag]
            for col in range(1, 4):
                cell = sheet.cell(row=row_index, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col == 3))
                if tag == "audit_header":
                    cell.font = Font(name=UI_FONT_FAMILY, size=11, bold=True)
                elif tag == "audit_condition":
                    cell.font = Font(name=UI_FONT_FAMILY, size=10)
                elif tag == "":
                    cell.font = Font(name=UI_FONT_FAMILY, size=10)
        workbook.save(path)

    def open_selected_methods(self, _event=None) -> None:
        if self._export_selected_method_previews(open_after=True, show_message=False):
            return
        self._export_selected_table(self.method_context_table, "Select an instrument method row to open.", open_after=True, show_message=False)

    def add_selected_method_to_compare(self) -> None:
        table = self._table_widget(self.method_context_table)
        selected = table.selection()
        if not selected:
            messagebox.showinfo(APP_NAME, "Select an instrument method row first.")
            return
        self._add_method_to_compare(selected[0])

    def _begin_method_drag(self, event) -> None:
        table = self._table_widget(self.method_context_table)
        self._method_drag_iid = table.identify_row(event.y) or ""

    def _finish_method_drag(self, event) -> None:
        if not self._method_drag_iid or not self.package:
            return
        target = self._table_widget(self.method_compare_selection_table)
        x = target.winfo_pointerx() - target.winfo_rootx()
        y = target.winfo_pointery() - target.winfo_rooty()
        if 0 <= x <= target.winfo_width() and 0 <= y <= target.winfo_height():
            self._add_method_to_compare(self._method_drag_iid)
        self._method_drag_iid = ""

    def _add_method_to_compare(self, iid: str) -> None:
        if not self.package:
            return
        if iid.startswith("context:external:"):
            messagebox.showinfo(APP_NAME, "Drag the embedded instrument method row, not the external reference row.")
            return
        method = self.package.elements_by_id.get(iid)
        if not method or method.kind != "instrument_method":
            return
        item = (self.package, method)
        self.method_compare_items = [existing for existing in self.method_compare_items if not (existing[0] is self.package and existing[1].id == method.id)]
        if len(self.method_compare_items) >= 2:
            self.method_compare_items.pop(0)
        self.method_compare_items.append(item)
        self._refresh_method_compare_selection_table()
        if len(self.method_compare_items) == 2:
            self.compare_loaded_instrument_methods()

    def _remove_method_compare_item(self, _event=None) -> None:
        table = self._table_widget(self.method_compare_selection_table)
        selected = table.selection()
        if not selected:
            return
        try:
            index = int(selected[0].split(":", 1)[1])
        except (IndexError, ValueError):
            return
        if 0 <= index < len(self.method_compare_items):
            self.method_compare_items.pop(index)
            self._refresh_method_compare_selection_table()

    def _refresh_method_compare_selection_table(self) -> None:
        if not hasattr(self, "method_compare_selection_table"):
            return
        table = self._table_widget(self.method_compare_selection_table)
        table.delete(*table.get_children())
        for index, (package, method) in enumerate(self.method_compare_items[:2]):
            parent = package.elements_by_id.get(method.parent_id or "")
            table.insert(
                "",
                "end",
                iid=f"methodcompare:{index}",
                values=("Yes" if index == 0 else "", package.path.name, parent.name if parent else "", method.name, method.url or method.filename or method.raw_filename),
            )

    def set_method_benchmark(self) -> None:
        method_name = self._selected_method_name()
        if not method_name:
            messagebox.showinfo(APP_NAME, "Select an instrument method row first.")
            return
        self.method_benchmark_name = method_name
        self.method_benchmark_package_path = str(self.package.path) if self.package else ""
        self.status_var.set(f"Benchmark instrument method: {method_name}")

    def compare_loaded_instrument_methods(self) -> None:
        if not self.loaded_packages:
            messagebox.showinfo(APP_NAME, "Load one or more CMBX packages first.")
            return
        if len(self.method_compare_items) == 2:
            comparisons = [
                (package, method, self._method_compare_lines(package, method))
                for package, method in self.method_compare_items
            ]
            benchmark_package, _benchmark_method, benchmark_lines = comparisons[0]
            self._set_method_compare_selection_rows(comparisons, benchmark_package)
            self._show_method_compare_window(comparisons[0][1].name, benchmark_package, benchmark_lines, comparisons)
            return
        packages = self._selected_packages_for_method_compare()
        if len(packages) > 2:
            messagebox.showinfo(APP_NAME, "Instrument method comparison supports up to two CMBX packages at a time.")
            return
        if len(packages) < 2:
            packages = self.loaded_packages[:2]
        if len(packages) < 2:
            messagebox.showinfo(APP_NAME, "Load or select two CMBX packages for method comparison.")
            return
        method_name = self.method_benchmark_name or self._selected_method_name() or self._first_instrument_method_name(packages[0])
        if not method_name:
            messagebox.showinfo(APP_NAME, "No instrument method was found for comparison.")
            return
        comparisons: list[tuple[CmbxPackage, CmbxElement | None, list[str]]] = []
        for package in packages:
            method = next((item for item in package.methods_and_reports if item.kind == "instrument_method" and item.name == method_name), None)
            if method is None:
                method = next((item for item in package.methods_and_reports if item.kind == "instrument_method"), None)
            lines = self._method_compare_lines(package, method) if method else ["<instrument method not found>"]
            comparisons.append((package, method, lines))
        if not comparisons:
            return
        benchmark_index = next((index for index, (package, _method, _lines) in enumerate(comparisons) if str(package.path) == self.method_benchmark_package_path), 0)
        if benchmark_index:
            comparisons.insert(0, comparisons.pop(benchmark_index))
        benchmark_package, _benchmark_method, benchmark_lines = comparisons[0]
        self._set_method_compare_selection_rows(comparisons, benchmark_package)
        self._show_method_compare_window(method_name, benchmark_package, benchmark_lines, comparisons)

    def _selected_packages_for_method_compare(self) -> list[CmbxPackage]:
        packages: list[CmbxPackage] = []
        for iid in self.package_tree.selection():
            package, _element = self.tree_item_context.get(iid, (None, None))
            if package and package not in packages:
                packages.append(package)
        return packages

    def _first_instrument_method_name(self, package: CmbxPackage) -> str:
        method = next((item for item in package.methods_and_reports if item.kind == "instrument_method"), None)
        return method.name if method else ""

    def _set_method_compare_selection_rows(self, comparisons: list[tuple[CmbxPackage, CmbxElement | None, list[str]]], benchmark_package: CmbxPackage) -> None:
        if not hasattr(self, "method_compare_selection_table"):
            return
        table = self._table_widget(self.method_compare_selection_table)
        table.delete(*table.get_children())
        for index, (package, method, _lines) in enumerate(comparisons[:2]):
            sequence_name = ""
            source = ""
            if method:
                parent = package.elements_by_id.get(method.parent_id or "")
                sequence_name = parent.name if parent else ""
                source = method.url or method.filename or method.raw_filename
            table.insert(
                "",
                "end",
                iid=f"methodcompare:{index}",
                values=("Yes" if package is benchmark_package else "", package.path.name, sequence_name, method.name if method else "", source),
            )

    def _selected_method_name(self) -> str:
        if not hasattr(self, "method_context_table"):
            return ""
        table = self._table_widget(self.method_context_table)
        selected = table.selection()
        if not selected:
            return ""
        values = table.item(selected[0], "values")
        return str(values[1]) if len(values) > 1 else ""

    def _method_compare_lines(self, package: CmbxPackage, method: CmbxElement | None) -> list[str]:
        if method is None:
            return ["<instrument method not found>"]
        external = discover_external_instrument_methods(package.path, [method.name]).get(method.name)
        if external:
            return ["\t".join([line.time, line.command, line.value, line.comment]).rstrip() for line in external.lines]
        try:
            embedded = extract_embedded_instrument_method(package, method)
            if not embedded:
                return ["<embedded instrument method payload not found>"]
            temp_dir = Path(tempfile.gettempdir()) / "cmbx_data_explorer_method_compare"
            temp_dir.mkdir(parents=True, exist_ok=True)
            stem = self._safe_temp_stem(f"{package.path.name}_{method.id or method.name}")
            cpxm_path = temp_dir / f"{stem}.cpxm.bin"
            xml_path = temp_dir / f"{stem}.xml"
            cpxm_path.write_bytes(embedded.cpxm_payload)
            decode_result = decode_cpxm_method_xml(cpxm_path, xml_path)
            if not decode_result.ok:
                return [decode_result.message]
            rows, error = build_method_flow_rows(xml_path.read_text(encoding="utf-8"))
            if error:
                return [error]
            return [
                "\t".join(cm_row[1:5]).rstrip()
                for cm_row in self._cm_method_rows_from_flow_rows(rows)
            ]
        except Exception as exc:
            return [str(exc)]

    def _show_method_compare_window(
        self,
        method_name: str,
        benchmark_package: CmbxPackage,
        benchmark_lines: list[str],
        comparisons: list[tuple[CmbxPackage, CmbxElement | None, list[str]]],
    ) -> None:
        window = tk.Toplevel(self.root)
        window.title(f"Instrument Method Compare - {method_name}")
        window.geometry("1500x780")
        window.configure(bg=self.colors["bg"])
        header = tk.Label(
            window,
            text=f"Benchmark: {benchmark_package.path.name} | Method: {method_name} | Double-click a cell to edit the preview text",
            font=self._font(11, "bold"),
            bg=self.colors["bg"],
            fg=self.colors["text"],
        )
        header.pack(anchor="w", padx=14, pady=(12, 8))
        body = tk.Frame(window, bg=self.colors["bg"])
        body.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)
        left = comparisons[0]
        right = comparisons[1] if len(comparisons) > 1 else (benchmark_package, None, [])
        left_rows = self._method_compare_table_rows(left[2])
        right_rows = self._method_compare_table_rows(right[2])
        max_lines = max(len(left_rows), len(right_rows))
        left_table = self._build_method_side_table(body, left[0], left[1], benchmark_package, 0)
        right_table = self._build_method_side_table(body, right[0], right[1], benchmark_package, 1)
        for index in range(max_lines):
            left_values = left_rows[index] if index < len(left_rows) else ("", "<missing row>", "", "")
            right_values = right_rows[index] if index < len(right_rows) else ("", "<missing row>", "", "")
            left_text = "\t".join(left_values)
            right_text = "\t".join(right_values)
            tag = "different" if left_text != right_text else ""
            left_table.insert("", "end", values=(index, *left_values), tags=(tag,) if tag else ())
            right_table.insert("", "end", values=(index, *right_values), tags=(tag,) if tag else ())

    def _build_method_side_table(self, parent: tk.Widget, package: CmbxPackage, method: CmbxElement | None, benchmark_package: CmbxPackage, column: int) -> ttk.Treeview:
        frame = tk.Frame(parent, bg=self.colors["card"])
        frame.grid(row=0, column=column, sticky="nsew", padx=(0, 7) if column == 0 else (7, 0))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        sequence_name = ""
        if method:
            parent_element = package.elements_by_id.get(method.parent_id or "")
            sequence_name = parent_element.name if parent_element else ""
        title = f"{'BENCHMARK | ' if package is benchmark_package else ''}{package.path.name} | {sequence_name} | {method.name if method else ''}"
        tk.Label(frame, text=title, font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text"], anchor="w").grid(row=0, column=0, sticky="ew", pady=(0, 6))
        table = ttk.Treeview(frame, columns=("row", "time", "command", "value", "comment"), show="headings", style="Explorer.Treeview")
        for key, title_text, width in (("row", "", 52), ("time", "Time", 115), ("command", "Command", 260), ("value", "Value", 300), ("comment", "Comment", 360)):
            table.heading(key, text=title_text)
            table.column(key, width=width, anchor="w")
        table.tag_configure("different", background="#FEE2E2")
        table.grid(row=1, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=table.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        table.configure(yscrollcommand=scroll.set)
        table.bind("<Double-1>", self._edit_treeview_cell)
        return table

    def _method_compare_table_rows(self, lines: list[str]) -> list[tuple[str, str, str, str]]:
        rows: list[tuple[str, str, str, str]] = []
        for line in lines:
            parts = line.split("\t")
            padded = (parts + ["", "", "", ""])[:4]
            rows.append((padded[0], padded[1], padded[2], padded[3]))
        return rows

    def _edit_treeview_cell(self, event) -> None:
        table = event.widget
        row_id = table.identify_row(event.y)
        column_id = table.identify_column(event.x)
        if not row_id or column_id == "#1":
            return
        column_index = int(column_id.replace("#", "")) - 1
        bbox = table.bbox(row_id, column_id)
        if not bbox:
            return
        x, y, width, height = bbox
        values = list(table.item(row_id, "values"))
        entry = tk.Entry(table, font=self._font(9))
        entry.insert(0, str(values[column_index]) if column_index < len(values) else "")
        entry.select_range(0, "end")
        entry.place(x=x, y=y, width=width, height=height)

        def commit(_event=None) -> None:
            if column_index < len(values):
                values[column_index] = entry.get()
                table.item(row_id, values=values)
            entry.destroy()

        entry.bind("<Return>", commit)
        entry.bind("<FocusOut>", commit)
        entry.focus_set()

    def preview_selected_method(self, _event=None) -> None:
        if not self.package or not hasattr(self, "method_flow_table"):
            return
        table = self._table_widget(self.method_context_table)
        selected = table.selection()
        if not selected:
            self._set_method_flow_message("Select an instrument method row to preview decoded flow.")
            return
        iid = selected[0]
        if iid.startswith("context:external:"):
            self._preview_external_method(iid)
            return
        element = self.package.elements_by_id.get(iid)
        if not element or element.kind != "instrument_method":
            self._set_method_flow_message("The selected row is not an embedded instrument method.")
            return
        external = self.external_instrument_methods.get(element.name)
        if external:
            self._set_cm_method_rows(self._external_method_rows(external))
            self.status_var.set(f"Previewed CM-like method table from reference TXT: {external.path.name}")
            return
        self.status_var.set(f"Decoding method preview: {element.name}...")
        try:
            embedded = extract_embedded_instrument_method(self.package, element)
            if not embedded:
                self._set_method_flow_message("Embedded instrument method payload was not found in the sequence command object.")
                return
            temp_dir = Path(tempfile.gettempdir()) / "cmbx_data_explorer_method_preview"
            temp_dir.mkdir(parents=True, exist_ok=True)
            stem = self._safe_temp_stem(element.id or element.name)
            cpxm_path = temp_dir / f"{stem}.cpxm.bin"
            xml_path = temp_dir / f"{stem}.xml"
            cpxm_path.write_bytes(embedded.cpxm_payload)
            decode_result = decode_cpxm_method_xml(cpxm_path, xml_path)
            if not decode_result.ok:
                self._set_method_flow_message(decode_result.message)
                self.status_var.set("Method preview decode failed")
                return
            rows, error = build_method_flow_rows(xml_path.read_text(encoding="utf-8"))
            if error:
                self._set_method_flow_message(error)
                self.status_var.set("Method preview parse failed")
                return
            self._set_cm_method_rows(self._cm_method_rows_from_flow_rows(rows))
            self.status_var.set(f"Previewed method flow: {element.name}")
        except Exception as exc:
            self._set_method_flow_message(str(exc))
            self.status_var.set("Method preview failed")

    def _preview_external_method(self, iid: str) -> None:
        table = self._table_widget(self.method_context_table)
        values = table.item(iid, "values")
        source = str(values[3]) if len(values) >= 4 else ""
        external = next((item for item in self.external_instrument_methods.values() if str(item.path) == source), None)
        if not external:
            self._set_method_flow_message("External reference TXT was not found.")
            return
        self._set_cm_method_rows(self._external_method_rows(external))
        self.status_var.set(f"Previewed reference TXT: {external.path.name}")

    def _external_method_rows(self, method: InstrumentMethodText) -> list[tuple[str, str, str, str, str, str, str]]:
        rows: list[tuple[str, str, str, str, str, str, str]] = []
        condition_depth = 0
        for index, line in enumerate(method.lines):
            time_text = line.time
            command = line.command
            value = line.value
            comment = line.comment
            tag = self._cm_method_tag(time_text, command, value, comment, condition_depth)
            kind = self._cm_method_kind(time_text, command, value, comment, tag)
            rows.append((str(index), kind, time_text, command, value, comment, tag))
            lower_time = time_text.lower()
            if lower_time in {"if", "trigger"}:
                condition_depth += 1
            elif lower_time in {"else if", "else"} and condition_depth == 0:
                condition_depth = 1
            elif lower_time in {"end if", "end trigger"}:
                condition_depth = max(0, condition_depth - 1)
        return rows

    def _cm_method_tag(self, time_text: str, command: str, value: str, comment: str, condition_depth: int) -> str:
        lower_time = time_text.lower()
        lower_command = command.lower()
        lower_value = value.lower()
        if "{initial time}" in lower_time or lower_command == "equilibration":
            return "cm_initial"
        if lower_time in {"if", "else if", "else", "end if", "trigger", "end trigger"} or condition_depth > 0:
            return "cm_condition"
        if command.startswith("=") or (command.startswith("-") and not value):
            return "cm_header"
        if comment and not command and not value:
            return "cm_comment"
        if command and not value and not comment and any(token in lower_command for token in ("====", "hplc-system", "column compartment", "parameters")):
            return "cm_header"
        if lower_value.startswith("columncomp.modelno") or "trigger" in lower_time:
            return "cm_condition"
        return ""

    def _cm_method_rows_from_flow_rows(self, rows) -> list[tuple[str, str, str, str, str, str, str]]:
        cm_rows: list[tuple[str, str, str, str, str, str, str]] = []
        last_time = ""
        for index, row in enumerate(rows):
            display_time = ""
            if row.action == "STAGE":
                stage_name = self._cm_stage_display_name(row.value or row.stage)
                display_time = row.time or ("{Initial Time}" if stage_name == "Instrument Setup" else "")
                last_time = display_time
            elif row.action not in {"IF", "ELSE IF", "ELSE", "END IF"} and row.time and row.time != last_time:
                display_time = row.time
                last_time = row.time
            cm_rows.append(self._cm_method_row_from_flow_row(index, row, display_time))
            if row.action == "TRIGGER":
                trigger_lines = [line.strip() for line in row.value.splitlines() if line.strip()]
                for param_index, param in enumerate(trigger_lines[1:], 1):
                    cm_rows.append((f"{index}.{param_index}", "Command", "", param, "", "", "cm_condition"))
        return cm_rows

    def _cm_method_row_from_flow_row(self, index: int, row, display_time: str = "") -> tuple[str, str, str, str, str, str, str]:
        indent = "    " * max(0, getattr(row, "level", 0))
        action = row.action
        if action == "STAGE":
            stage_name = self._cm_stage_display_name(row.value or row.stage)
            return (
                str(index),
                "Stage",
                display_time or row.time or ("{Initial Time}" if stage_name == "Instrument Setup" else ""),
                stage_name,
                row.comment,
                "",
                "cm_initial",
            )
        if action == "COMMENT":
            return (str(index), "Comment", display_time, f"{indent}{row.comment}", "", "", "cm_comment")
        if action == "TRIGGER":
            trigger_lines = [line.strip() for line in row.value.splitlines() if line.strip()]
            trigger_name = trigger_lines[0] if trigger_lines else ""
            return (str(index), "Stage", f"{indent}Trigger", trigger_name, "", row.comment, "cm_condition")
        if action == "END TRIGGER":
            return (str(index), "Stage", f"{indent}End Trigger", "", "", "", "cm_condition")
        if action == "IF":
            return (str(index), "Branch", f"{indent}If", row.condition, "", "", "cm_condition")
        if action == "ELSE IF":
            return (str(index), "Branch", f"{indent}Else If", row.condition, "", "", "cm_condition")
        if action == "ELSE":
            return (str(index), "Branch", f"{indent}Else", "", "", "", "cm_condition")
        if action == "END IF":
            return (str(index), "Branch", f"{indent}End If", "", "", "", "cm_condition")
        if action == "END":
            return (str(index), "End", "", "End", "", "", "")
        if action in {"SET", "RUN"}:
            return (str(index), "Command", display_time, f"{indent}{row.target}", row.value, row.comment, "")
        if action == "TIME":
            return (str(index), "Time", row.time, "", "", row.comment, "")
        tag = self._xml_flow_tag(action)
        return (str(index), self._cm_method_kind(row.time, row.target or row.action, row.value, row.comment, tag), row.time, f"{indent}{row.target or row.action}", row.value, row.comment, tag)

    def _cm_method_kind(self, time_text: str, command: str, value: str, comment: str, tag: str) -> str:
        lower_time = (time_text or "").strip().lower()
        lower_command = (command or "").strip().lower()
        if tag in {"cm_comment", "cm_header"}:
            return "Comment"
        if tag == "cm_initial":
            return "Stage"
        if tag == "cm_condition" or lower_time in {"if", "else if", "else", "end if", "trigger", "end trigger"}:
            return "Branch"
        if lower_command == "end":
            return "End"
        if command or value:
            return "Command"
        return ""

    def _cm_stage_display_name(self, value: str) -> str:
        mapping = {
            "InstrumentSetup": "Instrument Setup",
            "StartRun": "Start Run",
            "StopRun": "Stop Run",
            "PostRun": "Post Run",
            "InjectPreparation": "Inject Preparation",
            "MassSpectrometerSync": "Mass Spectrometer Sync",
        }
        return mapping.get(value, value)

    def _xml_flow_command(self, action: str, target: str, value: str) -> str:
        if action == "STAGE":
            return value
        if action in {"IF", "ELSE", "ELSE IF", "END IF"}:
            return action.title()
        if action == "TIME":
            return "Time Step"
        return target

    def _xml_flow_tag(self, action: str) -> str:
        if action == "STAGE":
            return "cm_initial"
        if action in {"IF", "ELSE", "ELSE IF", "END IF"}:
            return "cm_condition"
        if action == "COMMENT":
            return "cm_comment"
        return ""

    def _clear_method_flow(self) -> None:
        if not hasattr(self, "method_flow_table"):
            return
        table = self._table_widget(self.method_flow_table)
        table.delete(*table.get_children())

    def _set_method_flow_message(self, message: str) -> None:
        self._clear_method_flow()
        table = self._table_widget(self.method_flow_table)
        table.insert("", "end", values=("", "", "", "", "", message))

    def _set_cm_method_rows(self, rows: list[tuple[str, str, str, str, str, str, str]]) -> None:
        self._clear_method_flow()
        table = self._table_widget(self.method_flow_table)
        if not rows:
            self._set_method_flow_message("No method flow rows were found.")
            return
        for row, kind, time_text, command, value, comment, tag in rows:
            table.insert("", "end", values=(row, kind, time_text, command, value, comment), tags=(tag,) if tag else ())

    def _export_selected_method_previews(self, open_after: bool = False, show_message: bool = True) -> bool:
        if not self.package:
            return False
        table = self._table_widget(self.method_context_table)
        selected = list(table.selection())
        if not selected:
            return False
        jobs: list[tuple[str, list[tuple[str, str, str, str, str, str, str]]]] = []
        for iid in selected:
            result = self._method_preview_rows_for_iid(iid)
            if result:
                jobs.append(result)
        if not jobs:
            return False
        package = self.package
        output = self._preview_export_folder(package, "instrument_method_previews")
        self.status_var.set(f"Exporting {len(jobs)} styled instrument method preview workbook(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                for title, rows in jobs:
                    path = output / f"{self._safe_temp_stem(title)}.xlsx"
                    self._write_method_preview_workbook(path, title, rows)
                    paths.append(path)
                self._call_ui(lambda: self._export_done(paths, open_after=open_after, show_message=show_message))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()
        return True

    def _method_preview_rows_for_iid(self, iid: str) -> tuple[str, list[tuple[str, str, str, str, str, str, str]]] | None:
        if not self.package:
            return None
        if iid.startswith("context:external:"):
            table = self._table_widget(self.method_context_table)
            values = table.item(iid, "values")
            source = str(values[3]) if len(values) >= 4 else ""
            external = next((item for item in self.external_instrument_methods.values() if str(item.path) == source), None)
            if not external:
                return None
            return (external.path.stem, self._external_method_rows(external))
        element = self.package.elements_by_id.get(iid)
        if not element or element.kind != "instrument_method":
            return None
        external = self.external_instrument_methods.get(element.name)
        if external:
            return (element.name, self._external_method_rows(external))
        embedded = extract_embedded_instrument_method(self.package, element)
        if not embedded:
            return None
        temp_dir = Path(tempfile.gettempdir()) / "cmbx_data_explorer_method_preview_export"
        temp_dir.mkdir(parents=True, exist_ok=True)
        stem = self._safe_temp_stem(element.id or element.name)
        cpxm_path = temp_dir / f"{stem}.cpxm.bin"
        xml_path = temp_dir / f"{stem}.xml"
        cpxm_path.write_bytes(embedded.cpxm_payload)
        decode_result = decode_cpxm_method_xml(cpxm_path, xml_path)
        if not decode_result.ok:
            return None
        rows, error = build_method_flow_rows(xml_path.read_text(encoding="utf-8"))
        if error:
            return None
        return (
            element.name,
            self._cm_method_rows_from_flow_rows(rows),
        )

    def _write_method_preview_workbook(self, path: Path, title: str, rows: list[tuple[str, str, str, str, str, str, str]]) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self._excel_sheet_title(title)
        thin = Side(style="thin", color="C7C7C7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fills = {
            "cm_initial": PatternFill("solid", fgColor="F9C18B"),
            "cm_condition": PatternFill("solid", fgColor="BFF29A"),
        }
        headers = ("#", "Kind", "Time", "Command", "Value", "Comment")
        widths = (8, 14, 16, 48, 72, 90)
        for col, (header, width) in enumerate(zip(headers, widths), 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = PatternFill("solid", fgColor="D8D8DE")
            cell.font = Font(name=UI_FONT_FAMILY, size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            sheet.column_dimensions[chr(64 + col)].width = width
        sheet.freeze_panes = "A2"
        for out_row, (row_number, kind, time_text, command, value, comment, tag) in enumerate(rows, 2):
            values = (row_number, kind, time_text, command, value, comment)
            for col, value_text in enumerate(values, 1):
                cell = sheet.cell(row=out_row, column=col, value=value_text)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col in {4, 5, 6}), horizontal="right" if col == 1 else "left")
                cell.font = Font(name=UI_FONT_FAMILY, size=10)
                if tag in fills:
                    cell.fill = fills[tag]
                if tag in {"cm_comment", "cm_header"}:
                    cell.font = Font(name=UI_FONT_FAMILY, size=10, italic=True, color="178A24")
        workbook.save(path)

    def _write_method_preview_markdown(self, path: Path, title: str, rows: list[tuple[str, str, str, str, str, str, str]]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        lines = [
            f"# {title} Instrument Method Script",
            "",
            "Source: decoded CMBX instrument method preview",
            "Format: `CM_METHOD_SCRIPT_MD_FORMAT_SPEC.md` / `CM Compiler Rules.MD`",
            "",
            "```tsv",
            "Time\tCommand\tValue\tComment",
        ]
        for _row_number, kind, time_text, command, value, comment, _tag in rows:
            time_out, command_out, value_out, comment_out = self._method_preview_row_to_md_cells(kind, time_text, command, value, comment)
            lines.append("\t".join(self._clean_method_md_cell(cell) for cell in (time_out, command_out, value_out, comment_out)))
        lines.extend(["```", ""])
        path.write_text("\r\n".join(lines), encoding="utf-8")

    def _method_preview_row_to_md_cells(self, kind: str, time_text: str, command: str, value: str, comment: str) -> tuple[str, str, str, str]:
        time_out = (time_text or "").strip()
        command_out = (command or "").lstrip()
        value_out = value or ""
        comment_out = comment or ""
        if kind == "Branch" and time_out in {"If", "Else If"} and not command_out and value_out:
            command_out, value_out = value_out, ""
        if kind == "Branch" and time_out in {"Else", "End If"}:
            command_out, value_out = "", ""
        if command_out == "End Trigger" and not time_out:
            time_out, command_out, value_out = "End Trigger", "", ""
        if time_out == "End Trigger":
            command_out, value_out = "", ""
        if kind == "End" and command_out == "End":
            time_out, value_out = "", ""
        return time_out, command_out, value_out, comment_out

    @staticmethod
    def _clean_method_md_cell(value: str) -> str:
        return str(value or "").replace("\t", " ").replace("\r", " ").replace("\n", "\\n")

    def _preview_export_folder(self, package: CmbxPackage, category: str) -> Path:
        output = Path(self.output_folder_var.get().strip())
        folder = output / package.path.stem / "_preview" / category
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def _excel_sheet_title(self, value: str) -> str:
        text = re.sub(r"[\[\]\:\*\?\/\\]", "_", value).strip()
        return (text or "Sheet1")[:31]

    def open_selected_processing(self, _event=None) -> None:
        self._export_selected_table(self.processing_method_table, "Select a processing method row to open.", open_after=True, show_message=False)

    def open_selected_templates(self, _event=None) -> None:
        self._export_selected_table(self.report_template_table, "Select a report template row to open.", open_after=True, show_message=False)

    def _export_selected_table(self, table_frame: tk.Frame, empty_message: str, open_after: bool = False, show_message: bool = True) -> None:
        if not self.package:
            messagebox.showinfo(APP_NAME, "Load a CMBX package first.")
            return
        selected_iids = list(self._table_widget(table_frame).selection())
        selected_elements = self._selected_elements_from_table(table_frame)
        if not selected_elements:
            if selected_iids and all(iid.startswith("context:external:") for iid in selected_iids):
                return
            messagebox.showinfo(APP_NAME, empty_message)
            return
        output = Path(self.output_folder_var.get().strip())
        self.status_var.set(f"Exporting {len(selected_elements)} item(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths = export_elements(self.package, selected_elements, output)
                self._call_ui(lambda: self._export_done(paths, open_after=open_after, show_message=show_message))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def export_report_workbook(self, open_after: bool = False, show_message: bool = True) -> None:
        if not self.package or not self.current_injection:
            messagebox.showinfo(APP_NAME, "Select an injection first.")
            return
        reports = self._reports_for_injection(self.current_injection)
        if not reports:
            messagebox.showinfo(APP_NAME, "No report definition was found for the selected injection.")
            return
        report = reports[0]
        output = Path(self.output_folder_var.get().strip())
        injection = self.current_injection
        self.status_var.set(f"Exporting report workbook for {injection.name}...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                path = export_report_workbook(self.package, injection, report, output, progress=self._thread_status)  # type: ignore[arg-type]
                self._call_ui(lambda: self._export_done([path], open_after=open_after, show_message=show_message))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def open_report_sheet_workbook(self, _event=None) -> None:
        selected_context, _message = self._selected_report_sheet_context()
        if selected_context and self.package:
            report, injection, sheet_name = selected_context
            output = Path(self.output_folder_var.get().strip())
            self.status_var.set(f"Opening filled CM report sheet: {injection.name} / {sheet_name}...")
            self._set_buttons_state("disabled")

            def worker_selected_sheet() -> None:
                try:
                    path = export_filled_report_template_workbook(
                        self.package,
                        injection,
                        report,
                        output,
                        progress=self._thread_status,  # type: ignore[arg-type]
                        sheet_names=[sheet_name] if sheet_name else None,
                    )
                    self._call_ui(lambda: self._export_done([path], open_after=True, show_message=False))
                except Exception as exc:
                    self._call_ui(lambda exc=exc: self._export_failed(exc))

            threading.Thread(target=worker_selected_sheet, daemon=True).start()
            return
        self.export_report_sheets_workbook(open_after=True, show_message=False)

    def export_report_sheets_workbook(self, open_after: bool = True, show_message: bool = True) -> None:
        if not self.package:
            messagebox.showinfo(APP_NAME, "Load a CMBX package first.")
            return
        output = Path(self.output_folder_var.get().strip())
        if self.current_injection:
            reports = self._reports_for_injection(self.current_injection)
            if not reports:
                messagebox.showinfo(APP_NAME, "No report definition was found for the selected injection.")
                return
            injection = self.current_injection
            report = reports[0]
            self.status_var.set(f"Exporting filled CM report for {injection.name}...")
            self._set_buttons_state("disabled")

            def worker_injection() -> None:
                try:
                    path = export_filled_report_template_workbook(self.package, injection, report, output, progress=self._thread_status)  # type: ignore[arg-type]
                    self._call_ui(lambda: self._export_done([path], open_after=open_after, show_message=show_message))
                except Exception as exc:
                    self._call_ui(lambda exc=exc: self._export_failed(exc))

            threading.Thread(target=worker_injection, daemon=True).start()
            return
        if self.current_sequence:
            sequence = self.current_sequence
            self.status_var.set(f"Exporting sequence report sheets for {sequence.name}...")
            self._set_buttons_state("disabled")

            def worker_sequence() -> None:
                try:
                    path = export_sequence_report_sheets_workbook(self.package, sequence, output, progress=self._thread_status)  # type: ignore[arg-type]
                    self._call_ui(lambda: self._export_done([path], open_after=open_after, show_message=show_message))
                except Exception as exc:
                    self._call_ui(lambda exc=exc: self._export_failed(exc))

            threading.Thread(target=worker_sequence, daemon=True).start()
            return
        messagebox.showinfo(APP_NAME, "Select a sequence or injection first.")

    def export_all_reports(self) -> None:
        if not self.package:
            messagebox.showinfo(APP_NAME, "Load a CMBX package first.")
            return
        output = Path(self.output_folder_var.get().strip())
        self.status_var.set("Exporting report workbooks for all injections...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths = export_all_report_workbooks(self.package, output, progress=self._thread_status)  # type: ignore[arg-type]
                self._call_ui(lambda: self._export_done(paths))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def export_foq_candidate_sequences(self) -> None:
        candidates = self._selected_foq_candidates()
        if not candidates:
            messagebox.showinfo(APP_NAME, "Add one or more FOQ candidate sequences first.")
            return
        output = Path(self.output_folder_var.get().strip()) / "foq_candidate_sequences"
        self.status_var.set(f"Exporting {len(candidates)} candidate sequence CMBX file(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                for package, sequence, _device, _source, _report_template in candidates:
                    self._thread_status(f"Exporting candidate sequence: {sequence.name}")
                    paths.extend(split_cmbx_sequences(package, [sequence], output))
                self._call_ui(lambda: self._export_done(paths))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def export_foq_candidate_db(self) -> None:
        candidates = self._selected_foq_candidates()
        if not candidates:
            messagebox.showinfo(APP_NAME, "Add one or more FOQ candidate sequences first.")
            return
        self._commit_path_display(self.foq_mapping_path_var, self.foq_mapping_display_var, "foq")
        self._refresh_foq_mapping_filter_options()
        mapping_path = Path(self.foq_mapping_path_var.get().strip())
        db_field_filter = self._selected_foq_db_fields()
        if not mapping_path.exists():
            messagebox.showinfo(APP_NAME, "Select a valid FOQResultLocations mapping file first.")
            return
        unresolved = [sequence.name for _package, sequence, device, _source, _report_template in candidates if not device or device == "unresolved"]
        if unresolved:
            messagebox.showinfo(
                APP_NAME,
                "Device type could not be resolved from CMBX report/audit for:\n"
                + "\n".join(unresolved[:12])
                + ("\n..." if len(unresolved) > 12 else ""),
            )
            return
        output = Path(self.output_folder_var.get().strip())
        self.progress_var.set(0.0)
        self.status_var.set(f"Exporting {len(candidates)} candidate FOQ DB workbook(s)...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                with tempfile.TemporaryDirectory(prefix="cmbx_foq_candidates_") as temp_dir:
                    temp_root = Path(temp_dir)
                    total = max(len(candidates), 1)
                    span = 100.0 / total
                    for index, (package, sequence, device, _source, report_template) in enumerate(candidates):
                        device_type = device
                        report_suffix = f", {report_template}" if report_template else ""
                        progress = self._scaled_thread_progress(index * span, span)
                        progress(f"Exporting FOQ DB: {sequence.name} ({device_type}{report_suffix})")
                        job_root = temp_root / f"candidate_{index}"
                        job_root.mkdir(parents=True, exist_ok=True)
                        sequence_paths = split_cmbx_sequences(package, [sequence], job_root)
                        if not sequence_paths:
                            continue
                        sequence_package = load_cmbx_package(sequence_paths[0])
                        paths.append(export_foq_contract_report(sequence_package, mapping_path, device_type, output, progress=progress, report_template_name=report_template, db_field_filter=db_field_filter))
                self._call_ui(lambda: self._export_done(paths))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def upload_foq_candidate_db(self) -> None:
        candidates = self._selected_foq_candidates()
        if not candidates:
            messagebox.showinfo(APP_NAME, "Add one or more FOQ candidate sequences first.")
            return
        self._commit_path_display(self.foq_mapping_path_var, self.foq_mapping_display_var, "foq")
        self._refresh_foq_mapping_filter_options()
        mapping_path = Path(self.foq_mapping_path_var.get().strip())
        db_field_filter = self._selected_foq_db_fields()
        if not mapping_path.exists():
            messagebox.showinfo(APP_NAME, "Select a valid FOQResultLocations mapping file first.")
            return
        unresolved = [sequence.name for _package, sequence, device, _source, _report_template in candidates if not device or device == "unresolved"]
        if unresolved:
            messagebox.showinfo(
                APP_NAME,
                "Device type could not be resolved from CMBX report/audit for:\n"
                + "\n".join(unresolved[:12])
                + ("\n..." if len(unresolved) > 12 else ""),
            )
            return
        output = Path(self.output_folder_var.get().strip())
        config = self._database_upload_config()
        self.progress_var.set(0.0)
        self.status_var.set(f"Exporting and uploading {len(candidates)} candidate FOQ DB workbook(s)...")
        self._append_db_upload_log(f"Candidate upload requested: {len(candidates)} sequence(s) -> {config.database}.{config.schema}.{config.table}")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                with tempfile.TemporaryDirectory(prefix="cmbx_foq_candidates_") as temp_dir:
                    temp_root = Path(temp_dir)
                    total = max(len(candidates), 1)
                    span = 70.0 / total
                    for index, (package, sequence, device, _source, report_template) in enumerate(candidates):
                        device_type = device
                        report_suffix = f", {report_template}" if report_template else ""
                        progress = self._scaled_thread_progress(index * span, span)
                        progress(f"Exporting FOQ DB: {sequence.name} ({device_type}{report_suffix})")
                        self._thread_db_upload_log(f"Exporting DB workbook for {sequence.name} ({device_type}{report_suffix})")
                        job_root = temp_root / f"candidate_{index}"
                        job_root.mkdir(parents=True, exist_ok=True)
                        sequence_paths = split_cmbx_sequences(package, [sequence], job_root)
                        if not sequence_paths:
                            continue
                        sequence_package = load_cmbx_package(sequence_paths[0])
                        paths.append(export_foq_contract_report(sequence_package, mapping_path, device_type, output, progress=progress, report_template_name=report_template, db_field_filter=db_field_filter))
                self._thread_status("Uploading candidate FOQ DB workbook(s) to SQL Server...")
                results = upload_foq_db_workbooks(paths, config, log=self._thread_db_upload_log)
                self._call_ui(lambda: self._database_upload_done(results))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._database_upload_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def export_foq_contract_db(self) -> None:
        packages = self.loaded_packages or ([self.package] if self.package else [])
        packages = [package for package in packages if package is not None]
        if not packages:
            messagebox.showinfo(APP_NAME, "Load a CMBX package first.")
            return
        self._commit_path_display(self.foq_mapping_path_var, self.foq_mapping_display_var, "foq")
        self._refresh_foq_mapping_filter_options()
        mapping_path = Path(self.foq_mapping_path_var.get().strip())
        device_type = self.foq_device_type_var.get().strip()
        db_field_filter = self._selected_foq_db_fields()
        if not mapping_path.exists():
            messagebox.showinfo(APP_NAME, "Select a valid FOQResultLocations mapping file first.")
            return
        if not device_type:
            messagebox.showinfo(APP_NAME, "Enter a device type, for example VH-C10-A.")
            return
        output = Path(self.output_folder_var.get().strip())
        self.progress_var.set(0.0)
        self.status_var.set(f"Exporting {len(packages)} FOQ DB workbook(s) for {device_type}...")
        self._set_buttons_state("disabled")

        def worker() -> None:
            try:
                paths: list[Path] = []
                total = max(len(packages), 1)
                span = 100.0 / total
                for index, package in enumerate(packages):
                    progress = self._scaled_thread_progress(index * span, span)
                    progress(f"Exporting FOQ DB workbook: {package.path.name}...")
                    paths.append(export_foq_contract_report(package, mapping_path, device_type, output, progress=progress, db_field_filter=db_field_filter))
                self._call_ui(lambda: self._export_done(paths))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._export_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _thread_status(self, message: str) -> None:
        now = datetime.now().timestamp()
        text = str(message or "")
        match = re.match(r"__PROGRESS__=([0-9.]+)\|(.*)", text)
        if match:
            percent = float(match.group(1))
            clean_message = match.group(2)
            if percent < 100 and percent - self._last_thread_progress_percent < 0.5 and now - self._last_thread_status_time < 0.25:
                return
            self._last_thread_progress_percent = percent
            self._last_thread_status_time = now
            self._call_ui(lambda: self._set_progress_status(clean_message, percent))
            return
        if now - self._last_thread_status_time < 0.25:
            return
        self._last_thread_status_time = now
        self._call_ui(lambda: self.status_var.set(text))

    def _call_ui(self, callback) -> None:
        if threading.get_ident() == self._main_thread_id:
            callback()
            return
        self._ui_queue.put(callback)

    def _drain_ui_queue(self) -> None:
        handled = 0
        while handled < 200:
            try:
                callback = self._ui_queue.get_nowait()
            except queue.Empty:
                break
            try:
                callback()
            finally:
                handled += 1
        self.root.after(50, self._drain_ui_queue)

    def _set_progress_status(self, message: str, percent: float) -> None:
        self.progress_var.set(max(0.0, min(100.0, percent)))
        self.status_var.set(f"{percent:.0f}% - {message}")

    def _scaled_thread_progress(self, base_percent: float, span_percent: float):
        def callback(message: str) -> None:
            text = str(message or "")
            match = re.match(r"__PROGRESS__=([0-9.]+)\|(.*)", text)
            if match:
                inner_percent = float(match.group(1))
                scaled = base_percent + span_percent * inner_percent / 100.0
                self._thread_status(f"__PROGRESS__={scaled:.1f}|{match.group(2)}")
                return
            self._thread_status(text)

        return callback

    def _selected_elements_from_table(self, table_frame: tk.Frame) -> list[CmbxElement]:
        if not self.package:
            return []
        table = self._table_widget(table_frame)
        elements: list[CmbxElement] = []
        opened_external = False
        for iid in table.selection():
            if iid.startswith("context:external:"):
                self._open_external_reference(iid)
                opened_external = True
                continue
            element = self.package.elements_by_id.get(iid)
            if element and element not in elements:
                elements.append(element)
        if opened_external and not elements:
            return []
        return elements

    def _selected_export_elements(self) -> list[CmbxElement]:
        if not self.package:
            return []
        elements: list[CmbxElement] = []
        for table_frame in (self.channel_table, self.audit_table):
            table = self._table_widget(table_frame)
            for iid in table.selection():
                element = self.package.elements_by_id.get(iid)
                if element and element not in elements:
                    elements.append(element)
        for table_frame in (self.method_context_table, self.processing_method_table, self.report_template_table):
            table = self._table_widget(table_frame)
            for iid in table.selection():
                if iid.startswith("context:external:"):
                    continue
                element = self.package.elements_by_id.get(iid)
                if element and element not in elements:
                    elements.append(element)
        if elements:
            return elements
        for iid in self.package_tree.selection():
            package, element = self.tree_item_context.get(iid, (None, None))
            if not package or package is not self.package or not element:
                continue
            if element.kind in {"signal", "audit", "sequence"}:
                elements.append(element)
            elif element.kind == "injection":
                elements.extend([child for child in element.children if child.kind in {"signal", "audit"}])
        return elements

    def _export_done(self, paths: list[Path], open_after: bool = False, show_message: bool = True) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(100.0 if paths else 0.0)
        self.status_var.set(f"Exported {len(paths)} item(s) to {self.output_folder_var.get()}")
        if open_after:
            self._open_preferred_path(paths)
        if not show_message:
            return
        preview = "\n".join(str(path) for path in paths[:12])
        if len(paths) > 12:
            preview += f"\n... {len(paths) - 12} more"
        messagebox.showinfo(APP_NAME, preview or "Export complete.")

    def _export_failed(self, exc: Exception) -> None:
        self._set_buttons_state("normal")
        self.progress_var.set(0.0)
        self.status_var.set("Export failed")
        messagebox.showerror(APP_NAME, str(exc))

    def _open_preferred_path(self, paths: list[Path]) -> None:
        if not paths:
            return
        patterns = [
            ".xls",
            "_report_template.xls",
            "_report.xlsx",
            "_embedded_method_flow.tsv",
            "_embedded_method_flow.txt",
            "_report_sheets.tsv",
            "_report_sheet_objects.tsv",
            "_embedded_report_metadata.txt",
            "_embedded_metadata.txt",
        ]
        suffixes = [".xlsx", ".xls", ".tsv", ".csv", ".txt", ".xml", ".bin"]
        selected = None
        for pattern in patterns:
            selected = next((path for path in paths if path.name.endswith(pattern)), None)
            if selected:
                break
        if selected is None:
            for suffix in suffixes:
                selected = next((path for path in paths if path.suffix.lower() == suffix), None)
                if selected:
                    break
        selected = selected or paths[0]
        self._open_path(selected)

    def _open_external_reference(self, iid: str) -> None:
        if not iid.startswith("context:external:"):
            return
        table = self._table_widget(self.method_context_table)
        values = table.item(iid, "values")
        if len(values) >= 4:
            self._open_path(Path(str(values[3])))

    def _open_path(self, path: Path) -> None:
        try:
            os.startfile(str(path))
            self.status_var.set(f"Opened {path}")
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Cannot open file:\n{path}\n\n{exc}")

    def show_parsing_notes(self) -> None:
        text = PARSING_NOTES.read_text(encoding="utf-8") if PARSING_NOTES.exists() else "CMBX parsing notes file was not found."
        text = text + "\n\n" + runtime_status_text()
        self._show_markdown_window("CMBX Parsing Notes", text)

    def show_external_report_engine(self) -> None:
        try:
            self.external_report_window = ExternalReportWindow(
                self.root,
                initial_paths=[package.path for package in self.loaded_packages],
                output_folder=Path(self.output_folder_var.get().strip() or DEFAULT_EXPORT_FOLDER),
            )
            self.external_report_window.top.lift()
            self.external_report_window.top.focus_force()
        except Exception as exc:
            _write_startup_log(f"External Report Engine launch failed: {exc}\n{traceback.format_exc()}")
            messagebox.showerror(APP_NAME, f"Cannot open External Report Engine:\n\n{exc}", parent=self.root)

    def show_method_script_generator(self) -> None:
        top = tk.Toplevel(self.root)
        top.title("Method Script Generator")
        top.configure(bg=self.colors["bg"])
        top.geometry("1380x820")
        top.minsize(1120, 700)
        top.resizable(True, True)
        try:
            top.state("zoomed")
        except tk.TclError:
            pass

        shell = tk.Frame(top, bg=self.colors["bg"])
        shell.pack(fill="both", expand=True, padx=20, pady=18)
        shell.columnconfigure(0, weight=1)
        shell.rowconfigure(3, weight=1)

        title_row = tk.Frame(shell, bg=self.colors["bg"])
        title_row.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        title_row.columnconfigure(4, weight=1)
        tk.Label(title_row, text="Method Script Generator", font=self._font(16, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=(0, 14))
        tk.Label(title_row, text="Family", font=self._font(9, "bold"), bg=self.colors["bg"], fg=self.colors["text_secondary"]).grid(row=0, column=1, sticky="w")
        family_combo = ttk.Combobox(title_row, textvariable=self.method_generator_family_var, values=("TCC", "VDAD"), state="readonly", width=10)
        family_combo.grid(row=0, column=2, sticky="w", padx=(6, 12))
        self._make_button(title_row, "AI Settings", self.show_ai_settings, kind="neutral", width=11).grid(row=0, column=3, sticky="w")
        self._make_toolbar_label(title_row, "Flow: Natural language -> AI structured spec -> local KB route -> CM mechanism plan -> generated script -> report contract check").grid(row=0, column=4, sticky="ew", padx=(14, 0))

        intent_row = tk.Frame(shell, bg=self.colors["bg"])
        intent_row.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        intent_row.columnconfigure(1, weight=1)
        tk.Label(intent_row, text="Natural Language", font=self._font(9, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", padx=(0, 8))
        intent_entry = tk.Entry(intent_row, textvariable=self.method_generator_intent_var, font=self._font(10), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        intent_entry.grid(row=0, column=1, sticky="ew", ipady=5)
        self._make_button(intent_row, "Run Flow", lambda: self._run_method_script_generator_flow(), kind="primary", width=14).grid(row=0, column=2, sticky="e", padx=(10, 0))
        self._make_button(intent_row, "Close", top.destroy, kind="neutral", width=10).grid(row=0, column=3, sticky="e", padx=(8, 0))

        flow_frame = tk.Frame(shell, bg=self.colors["bg"])
        flow_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        flow_frame.columnconfigure(0, weight=1)
        self.method_generator_flow_table = self._make_table(
            flow_frame,
            ("stage", "result", "evidence", "status"),
            {
                "stage": "Stage",
                "result": "Result",
                "evidence": "Evidence",
                "status": "Status",
            },
        )
        self.method_generator_flow_table.grid(row=0, column=0, sticky="ew")
        flow_table = self._table_widget(self.method_generator_flow_table)
        flow_table.column("stage", width=210)
        flow_table.column("result", width=430)
        flow_table.column("evidence", width=430)
        flow_table.column("status", width=360)

        body = tk.PanedWindow(shell, orient=tk.VERTICAL, sashwidth=7, sashrelief="flat", bg=self.colors["bg"], bd=0)
        body.grid(row=3, column=0, sticky="nsew")

        upper = tk.Frame(body, bg=self.colors["bg"])
        upper.columnconfigure(0, weight=1)
        upper.columnconfigure(1, weight=1)
        upper.columnconfigure(2, weight=1)
        upper.rowconfigure(1, weight=1)
        tk.Label(upper, text="1. Structured Spec", font=self._font(10, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w")
        tk.Label(upper, text="2. Local KB Route", font=self._font(10, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=1, sticky="w", padx=(8, 0))
        tk.Label(upper, text="3. Report Contract Check", font=self._font(10, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=2, sticky="w", padx=(8, 0))
        self.method_generator_spec_text = tk.Text(upper, wrap="word", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        self.method_generator_spec_text.grid(row=1, column=0, sticky="nsew", pady=(6, 0), padx=(0, 8))
        self.method_generator_route_table = self._make_table(
            upper,
            ("role", "intent", "method", "status"),
            {"role": "Role", "intent": "Intent", "method": "Method Script", "status": "Status"},
        )
        self.method_generator_route_table.grid(row=1, column=1, sticky="nsew", pady=(6, 0), padx=(8, 8))
        route_table = self._table_widget(self.method_generator_route_table)
        route_table.column("role", width=120)
        route_table.column("intent", width=180)
        route_table.column("method", width=240)
        route_table.column("status", width=220)
        self.method_generator_report_table = self._make_table(
            upper,
            ("intent", "report", "fields", "constraint"),
            {"intent": "Intent", "report": "Report", "fields": "DB / Formula", "constraint": "Constraint"},
        )
        self.method_generator_report_table.grid(row=1, column=2, sticky="nsew", pady=(6, 0), padx=(8, 0))
        report_table = self._table_widget(self.method_generator_report_table)
        report_table.column("intent", width=150)
        report_table.column("report", width=220)
        report_table.column("fields", width=260)
        report_table.column("constraint", width=320)

        lower = tk.Frame(body, bg=self.colors["bg"])
        lower.columnconfigure(0, weight=1)
        lower.columnconfigure(1, weight=1)
        lower.rowconfigure(1, weight=1)
        tk.Label(lower, text="4. Original KB Method Script", font=self._font(10, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(8, 6), padx=(0, 8))
        tk.Label(lower, text="5. Generated / Planned Method Script", font=self._font(10, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=1, sticky="w", pady=(8, 6), padx=(8, 0))
        self.method_generator_original_table = self._make_method_preview_table(lower)
        self.method_generator_original_table.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
        self.method_generator_generated_table = self._make_method_preview_table(lower)
        self.method_generator_generated_table.grid(row=1, column=1, sticky="nsew", padx=(8, 0))

        body.add(upper, minsize=220, height=310)
        body.add(lower, minsize=280, height=470)

        self._clear_method_generator_window()
        intent_entry.focus_set()

    def _clear_method_generator_window(self) -> None:
        for frame_name in (
            "method_generator_flow_table",
            "method_generator_route_table",
            "method_generator_report_table",
            "method_generator_original_table",
            "method_generator_generated_table",
        ):
            frame = getattr(self, frame_name, None)
            if frame is not None:
                table = self._table_widget(frame)
                table.delete(*table.get_children())
        text = getattr(self, "method_generator_spec_text", None)
        if text is not None:
            text.configure(state="normal")
            text.delete("1.0", tk.END)
            text.insert("1.0", "Enter an intent and click Run Flow.")
            text.configure(state="disabled")

    def _run_method_script_generator_flow(self) -> None:
        intent = self.method_generator_intent_var.get().strip()
        if not intent:
            messagebox.showinfo(APP_NAME, "Enter a natural-language method intent first.")
            return
        self.status_var.set("Method Script Generator running...")
        self.progress_var.set(10.0)
        self._render_method_generator_pending_flow(intent)

        def worker() -> None:
            try:
                result = self._build_method_generator_result(intent, self.method_generator_family_var.get().strip() or "TCC")
                self._call_ui(lambda result=result: self._render_method_generator_result(result))
            except Exception as exc:
                self._call_ui(lambda exc=exc: self._method_generator_failed(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _build_method_generator_result(self, intent: str, family: str) -> dict[str, object]:
        self._method_generator_worker_status("Method Script Generator: loading AI settings...", 12.0)
        self._refresh_method_generator_ai_config_from_disk()
        self._method_generator_worker_status("Method Script Generator: parsing local semantic spec...", 18.0)
        local_spec = self._method_generator_local_spec(intent, family)
        spec = local_spec
        ai_status = "disabled: no API key configured"
        if self.ai_api_key_var.get().strip():
            try:
                self._method_generator_worker_status("Method Script Generator: AI structured spec request...", 28.0)
                ai_spec = self._call_method_generator_ai_spec(local_spec)
                spec = self._normalize_method_generator_spec(local_spec | ai_spec | {"source": "AI + local KB route"}, local_spec)
                ai_status = f"success: {self.ai_model_var.get().strip() or 'gpt-5.5'}"
            except Exception as exc:
                spec = self._normalize_method_generator_spec(local_spec | {"source": f"local fallback after AI error: {exc}"}, local_spec)
                ai_status = f"fallback: {exc}"
        else:
            spec = self._normalize_method_generator_spec(local_spec | {"source": "local semantic parser"}, local_spec)
        self._method_generator_worker_status("Method Script Generator: routing local KB...", 45.0)
        route_records = self._method_generator_route_records(spec)
        primary = next((target for _role, target in route_records if isinstance(target, FoqAlignmentRecord)), None)
        if primary is None and route_records:
            primary = route_records[0][1]
        source_rows: list[tuple[object, ...]] = []
        generated_rows: list[tuple[tuple[object, ...], bool]] = []
        cm_summary = "<no routed method>"
        role_audit: MethodRoleMapAudit | None = None
        operation_contract_draft: dict[str, object] = {}
        operation_contract_ai_status = "not run"
        operation_validation: dict[str, object] = {"status": "not run", "summary": "Operation contract draft was not requested."}
        row_transformer_status = "not run"
        if isinstance(primary, FoqAlignmentRecord):
            device = str(spec.get("device_model") or primary.device_label)
            self._method_generator_worker_status(f"Method Script Generator: loading method script {primary.instrument_method}...", 58.0)
            source_rows = self._method_generator_source_rows(primary, device)
            if source_rows:
                self._method_generator_worker_status("Method Script Generator: analyzing CM mechanism and Method Role Map...", 68.0)
                semantic = analyze_cm_method_rows(source_rows)
                cm_summary = self._test_plan_cm_mechanism_summary(semantic, primary)
                role_audit = classify_method_role_map(
                    source_rows,
                    family=primary.family,
                    method_name=primary.instrument_method,
                    test_intent=str(spec.get("primary_intent") or primary.test_intent),
                    device_model=device,
                )
                operation_contract_draft, operation_contract_ai_status = self._method_generator_build_operation_contract_draft(spec, role_audit, route_records, primary)
                operation_validation = self._method_generator_validate_operation_contract(operation_contract_draft, spec, role_audit, route_records)
                if self._method_generator_can_generate_method_preview(primary, operation_contract_draft, operation_validation, role_audit):
                    self._method_generator_worker_status("Method Script Generator: generating role-map-gated script preview...", 82.0)
                    generated_rows = self._method_generator_generate_rows(primary, source_rows, spec, device)
                    row_transformer_status = "generated"
                else:
                    generated_rows = self._method_generator_blocked_rows(
                        source_rows,
                        self._method_generator_generation_block_reason(role_audit, operation_validation),
                    )
                    row_transformer_status = "blocked"
            else:
                cm_summary = "BLOCKED: full method script missing from Method Script KB"
        elif isinstance(primary, MethodScriptKbEntry):
            self._method_generator_worker_status(f"Method Script Generator: loading method script {primary.method_name}...", 58.0)
            source_rows = [tuple(row) for row in flow_tsv_to_cm_preview_rows(primary.path)]
            if source_rows:
                self._method_generator_worker_status("Method Script Generator: analyzing CM mechanism and Method Role Map...", 68.0)
                cm_summary = self._method_generator_method_entry_summary(primary, source_rows)
                role_audit = classify_method_role_map(
                    source_rows,
                    family=family,
                    method_name=primary.method_name,
                    test_intent=str(spec.get("primary_intent") or ""),
                    device_model=str(spec.get("device_model") or ""),
                )
                operation_contract_draft, operation_contract_ai_status = self._method_generator_build_operation_contract_draft(spec, role_audit, route_records, primary)
                operation_validation = self._method_generator_validate_operation_contract(operation_contract_draft, spec, role_audit, route_records)
                if self._method_generator_can_generate_method_preview(primary, operation_contract_draft, operation_validation, role_audit):
                    self._method_generator_worker_status("Method Script Generator: generating role-map-gated script preview...", 82.0)
                    generated_rows = [(row, False) for row in source_rows]
                    row_transformer_status = "generated"
                else:
                    generated_rows = self._method_generator_blocked_rows(
                        source_rows,
                        self._method_generator_generation_block_reason(role_audit, operation_validation),
                    )
                    row_transformer_status = "blocked"
        related_summaries: list[str] = []
        for _role, target in route_records:
            if target is primary:
                continue
            if isinstance(target, MethodScriptKbEntry):
                rows = [tuple(row) for row in flow_tsv_to_cm_preview_rows(target.path)]
                related_summaries.append(self._method_generator_method_entry_summary(target, rows))
            elif isinstance(target, FoqAlignmentRecord):
                related_summaries.append(f"{target.test_intent}; method {target.instrument_method}; report {target.report_template}")
        if related_summaries:
            cm_summary = f"{cm_summary}; related mechanism KB routed, merge not yet generated: " + " | ".join(related_summaries[:3])
            if generated_rows:
                note = (
                    "",
                    "Comment",
                    "",
                    "BLOCKED: related trigger/stress method routed from KB but not merged into the primary script.",
                    "",
                    "A role-level merge is required before this can be treated as a complete generated method. Evidence: " + " ; ".join(related_summaries[:3]),
                    "modified",
                )
                generated_rows.insert(0, (note, True))
                if row_transformer_status == "generated":
                    row_transformer_status = "partial_merge_blocked"
        operation_contract = self._method_generator_operation_contract_summary(spec, role_audit, route_records)
        self._method_generator_worker_status("Method Script Generator: checking report contract and rendering result...", 92.0)
        return {
            "spec": spec,
            "routes": route_records,
            "source_rows": source_rows,
            "generated_rows": generated_rows,
            "cm_summary": cm_summary,
            "ai_status": ai_status,
            "role_audit": role_audit,
            "operation_contract": operation_contract,
            "operation_contract_draft": operation_contract_draft,
            "operation_contract_ai_status": operation_contract_ai_status,
            "operation_validation": operation_validation,
            "row_transformer_status": row_transformer_status,
        }

    def _method_generator_worker_status(self, message: str, progress: float | None = None) -> None:
        def update() -> None:
            self.status_var.set(message)
            if progress is not None:
                self.progress_var.set(progress)

        self._call_ui(update)

    def _method_generator_blocked_rows(self, source_rows: list[tuple[object, ...]], reason: str) -> list[tuple[tuple[object, ...], bool]]:
        note = (
            "",
            "Comment",
            "",
            "BLOCKED: role-map-gated script generation is disabled.",
            "",
            reason,
            "modified",
        )
        return [(note, True), *[(row, False) for row in source_rows]]

    def _method_generator_generation_block_reason(self, role_audit: MethodRoleMapAudit, operation_validation: dict[str, object] | None = None) -> str:
        if role_audit.status != "complete":
            return role_audit.summary
        validation_status = str((operation_validation or {}).get("status") or "")
        validation_summary = str((operation_validation or {}).get("summary") or "")
        if validation_status and validation_status not in {"ok", "not run", "method_ok_report_blocked"}:
            return f"{role_audit.summary}. Operation contract validation is {validation_status}: {validation_summary}"
        return (
            f"{role_audit.summary}. Role Map is complete, but no generic local operation generator is certified for "
            f"`{role_audit.method_name}` / `{role_audit.test_intent}` yet. Add a role-map operation contract before script generation."
        )

    def _method_generator_can_generate_method_preview(
        self,
        primary: object,
        operation_contract: dict[str, object],
        operation_validation: dict[str, object],
        role_audit: MethodRoleMapAudit | None,
    ) -> bool:
        if not isinstance(role_audit, MethodRoleMapAudit) or role_audit.status != "complete":
            return False
        if str(operation_validation.get("status") or "") not in {"ok", "method_ok_report_blocked"}:
            return False
        operation_type = str(operation_contract.get("operation_type") or "")
        if isinstance(primary, FoqAlignmentRecord):
            if primary.family == "TCC" and primary.test_intent == "temperature_accuracy" and operation_type.startswith("temperature_accuracy"):
                return True
            if primary.family == "TCC" and "stability" in primary.test_intent and operation_type.startswith("temperature_stability"):
                return True
            if primary.family == "TCC" and primary.test_intent == "heatup_cooldown_20_50_20" and operation_type.startswith("heatup_cooldown"):
                return True
        if role_audit.generation_mode == "local_operation_supported":
            return True
        return False

    def _method_generator_build_operation_contract_draft(
        self,
        spec: dict[str, object],
        role_audit: MethodRoleMapAudit | None,
        routes: list[tuple[str, object]],
        primary: object,
    ) -> tuple[dict[str, object], str]:
        if not isinstance(role_audit, MethodRoleMapAudit) or role_audit.status != "complete":
            return {}, "skipped: Method Role Map is not complete"
        local_draft = self._method_generator_local_operation_contract_draft(spec, role_audit, routes, primary)
        if not self.ai_api_key_var.get().strip():
            return local_draft, "disabled: no API key configured; local draft only"
        try:
            self._method_generator_worker_status("Method Script Generator: AI operation contract draft...", 76.0)
            ai_draft = self._call_method_generator_ai_operation_contract(spec, role_audit, routes, local_draft)
            merged = local_draft | ai_draft | {"source": "AI operation contract + local guardrails"}
            return merged, f"success: {self.ai_model_var.get().strip() or 'gpt-5.5'}"
        except Exception as exc:
            return local_draft | {"source": f"local draft after AI operation contract error: {exc}"}, f"fallback: {exc}"

    def _method_generator_local_operation_contract_draft(
        self,
        spec: dict[str, object],
        role_audit: MethodRoleMapAudit,
        routes: list[tuple[str, object]],
        primary: object,
    ) -> dict[str, object]:
        editable_roles = self._method_generator_roles_by_edit_status(role_audit, allowed_prefixes=("editable", "required_when_requested"))
        locked_roles = self._method_generator_roles_by_edit_status(role_audit, blocked_prefixes=("locked",))
        target = self._method_generator_target_temperature(spec)
        baseline = self._method_generator_baseline_temperature(spec)
        report_status, report_note = self._method_generator_report_contract_for_routes(routes, spec)
        operation_type = "role_map_review"
        if isinstance(primary, FoqAlignmentRecord) and primary.family == "TCC" and primary.test_intent == "temperature_accuracy":
            accuracy_points = spec.get("accuracy_points_c")
            if isinstance(accuracy_points, list) and len(accuracy_points) > 1:
                operation_type = "temperature_accuracy_multi_point_review"
            else:
                operation_type = "temperature_accuracy_baseline_to_single_target" if baseline is not None else "temperature_accuracy_single_target"
        elif isinstance(primary, FoqAlignmentRecord) and "stability" in primary.test_intent:
            operation_type = "temperature_stability_target_hold"
        elif isinstance(primary, FoqAlignmentRecord) and primary.test_intent == "heatup_cooldown_20_50_20":
            operation_type = "heatup_cooldown_transition_edit"
        return {
            "source": "local role-map draft",
            "operation_type": operation_type,
            "intent_summary": str(spec.get("user_intent") or ""),
            "baseline_c": baseline,
            "target_c": target,
            "duration_minutes": self._method_generator_float_or_none(spec.get("duration_minutes")),
            "editable_roles": editable_roles,
            "locked_roles": locked_roles,
            "modify_roles": editable_roles,
            "preserve_roles": locked_roles,
            "composition_plan": [
                {
                    "module": operation_type,
                    "basis_method": role_audit.method_name,
                    "editable_roles": editable_roles,
                    "locked_roles": locked_roles,
                }
            ],
            "row_edit_plan": [
                {
                    "role_id": role,
                    "operation": "modify_or_insert_by_role",
                    "target_c": target,
                    "baseline_c": baseline,
                }
                for role in editable_roles
            ],
            "report_impact": {
                "status": report_status,
                "summary": report_note,
            },
            "open_questions": [],
        }

    def _method_generator_roles_by_edit_status(
        self,
        role_audit: MethodRoleMapAudit,
        *,
        allowed_prefixes: tuple[str, ...] = (),
        blocked_prefixes: tuple[str, ...] = (),
    ) -> list[str]:
        roles: list[str] = []
        for match in role_audit.matches:
            status = match.edit_status.lower()
            include = False
            if allowed_prefixes and any(status.startswith(prefix) for prefix in allowed_prefixes):
                include = True
            if blocked_prefixes and any(status.startswith(prefix) for prefix in blocked_prefixes):
                include = True
            if include and match.role_id not in roles:
                roles.append(match.role_id)
        return roles

    def _call_method_generator_ai_operation_contract(
        self,
        spec: dict[str, object],
        role_audit: MethodRoleMapAudit,
        routes: list[tuple[str, object]],
        local_draft: dict[str, object],
    ) -> dict[str, object]:
        route_context: list[dict[str, object]] = []
        for role, target in routes[:4]:
            if isinstance(target, FoqAlignmentRecord):
                route_context.append({
                    "route_role": role,
                    "test_intent": target.test_intent,
                    "method": target.instrument_method,
                    "report": target.report_template,
                    "db_fields": list(target.db_fields[:12]),
                })
            elif isinstance(target, MethodScriptKbEntry):
                route_context.append({
                    "route_role": role,
                    "test_intent": "method_script_kb_entry",
                    "method": target.method_name,
                    "source": target.source,
                })
        role_context = {
            "method_name": role_audit.method_name,
            "test_intent": role_audit.test_intent,
            "role_counts": role_audit.role_counts(),
            "matched_roles": [
                {
                    "role_id": match.role_id,
                    "edit_status": match.edit_status,
                    "reason": match.reason,
                }
                for match in role_audit.matches[:40]
            ],
        }
        source_excerpt = self._method_generator_ai_source_excerpt(role_audit, routes)
        prompt = (
            "Return JSON only. You are the primary CMBX method designer. Use the structured spec, local KB route, "
            "Method Role Map, and source script excerpt to draft a method composition plan and row edit plan. "
            "The local program will only validate guardrails and render the preview. Do not invent roles or CM commands. "
            "Use only role_id values present in matched_roles. "
            "Required keys: operation_type, intent_summary, baseline_c, target_c, duration_minutes, modify_roles, "
            "preserve_roles, composition_plan, row_edit_plan, insert_blocks, report_impact, assumptions, open_questions. "
            "modify_roles and preserve_roles must be arrays of string role_id values, not objects. "
            "row_edit_plan must be an array of objects with role_id, operation, old_semantics, new_semantics, row_selection, rationale. "
            "If report_impact is blocked, say so; do not claim script generation is safe.\n\n"
            f"spec={json.dumps(spec, ensure_ascii=False)}\n"
            f"role_context={json.dumps(role_context, ensure_ascii=False)}\n"
            f"route_context={json.dumps(route_context, ensure_ascii=False)}\n"
            f"source_script_excerpt={json.dumps(source_excerpt, ensure_ascii=False)}\n"
            f"local_draft={json.dumps(local_draft, ensure_ascii=False)}"
        )
        body = {
            "model": self.ai_model_var.get().strip() or "gpt-5.5",
            "messages": [
                {"role": "system", "content": "Return JSON only. You draft role-map constrained CMBX method operation contracts, not method scripts."},
                {"role": "user", "content": prompt},
            ],
            "max_completion_tokens": 1200,
        }
        request = urllib.request.Request(
            self._ai_chat_completions_endpoint(),
            data=json.dumps(body).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {self.ai_api_key_var.get().strip()}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=18) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ValueError(f"HTTP {exc.code} from {self._ai_chat_completions_endpoint()} using model {body['model']}: {detail[:500]}") from exc
        content = payload["choices"][0]["message"]["content"]
        match = re.search(r"\{.*\}", content, flags=re.S)
        if not match:
            raise ValueError("AI operation contract response did not contain JSON.")
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else {}

    def _method_generator_ai_source_excerpt(
        self,
        role_audit: MethodRoleMapAudit,
        routes: list[tuple[str, object]],
    ) -> list[dict[str, object]]:
        source_rows: list[tuple[object, ...]] = []
        primary = next((target for _role, target in routes if isinstance(target, FoqAlignmentRecord)), None)
        if isinstance(primary, FoqAlignmentRecord):
            source_rows = self._method_generator_source_rows(primary, primary.device_label or "VH-C10-A")
        excerpts: list[dict[str, object]] = []
        matched_indices = {match.row_index for match in role_audit.matches[:80]}
        for index, row in enumerate(source_rows[:260]):
            if index not in matched_indices and len(excerpts) > 80:
                continue
            if index not in matched_indices and str(row[1]) not in {"Stage", "Branch"}:
                continue
            excerpts.append({
                "row_index": index,
                "row_no": row[0] if len(row) > 0 else "",
                "kind": row[1] if len(row) > 1 else "",
                "time": row[2] if len(row) > 2 else "",
                "command": row[3] if len(row) > 3 else "",
                "value": row[4] if len(row) > 4 else "",
                "comment": row[5] if len(row) > 5 else "",
            })
        return excerpts[:120]

    def _method_generator_validate_operation_contract(
        self,
        contract: dict[str, object],
        spec: dict[str, object],
        role_audit: MethodRoleMapAudit | None,
        routes: list[tuple[str, object]],
    ) -> dict[str, object]:
        if not isinstance(role_audit, MethodRoleMapAudit) or role_audit.status != "complete":
            return {"status": "blocked", "summary": "Method Role Map is not complete."}
        if not contract:
            return {"status": "blocked", "summary": "AI/local operation contract draft is missing."}
        available_roles = {match.role_id for match in role_audit.matches}
        locked_roles = {
            match.role_id
            for match in role_audit.matches
            if match.edit_status.lower().startswith("locked")
        }
        modify_roles = set(self._method_generator_contract_role_list(contract, "modify_roles"))
        edit_plan_roles = set(self._method_generator_contract_plan_role_list(contract, "row_edit_plan"))
        modify_roles.update(edit_plan_roles)
        unknown_roles = sorted(role for role in modify_roles if role not in available_roles)
        locked_modified = sorted(role for role in modify_roles if role in locked_roles)
        report_status, report_note = self._method_generator_report_contract_for_routes(routes, spec)
        blockers: list[str] = []
        if not str(contract.get("operation_type") or "").strip():
            blockers.append("operation_type is missing")
        if str(contract.get("operation_type") or "") == "temperature_accuracy_multi_point_review":
            blockers.append("multi-point Temperature Accuracy requires ladder/RetTime/report-cell selection before script generation")
        if unknown_roles:
            blockers.append("unknown modify_roles: " + ", ".join(unknown_roles))
        if locked_modified:
            blockers.append("locked roles requested for modification: " + ", ".join(locked_modified))
        if blockers:
            return {
                "status": "blocked",
                "summary": " ; ".join(blockers),
                "report_status": report_status,
                "report_note": report_note,
            }
        if report_status.startswith("blocked"):
            return {
                "status": "method_ok_report_blocked",
                "summary": "Operation contract draft is role-map valid for method preview; report/DB contract is blocked separately.",
                "report_status": report_status,
                "report_note": report_note,
            }
        return {
            "status": "ok",
            "summary": "Operation contract draft is role-map valid. Script generation still requires a certified row transformer.",
            "report_status": report_status,
            "report_note": report_note,
        }

    def _method_generator_contract_role_list(self, contract: dict[str, object], key: str) -> list[str]:
        value = contract.get(key)
        if isinstance(value, list):
            roles: list[str] = []
            for item in value:
                if isinstance(item, dict):
                    role = str(item.get("role_id") or item.get("id") or item.get("role") or "").strip()
                else:
                    role = str(item).strip()
                if role:
                    roles.append(role)
            return roles
        if isinstance(value, str) and value.strip():
            return [item.strip() for item in re.split(r"[,;]", value) if item.strip()]
        return []

    def _method_generator_contract_plan_role_list(self, contract: dict[str, object], key: str) -> list[str]:
        value = contract.get(key)
        roles: list[str] = []
        if not isinstance(value, list):
            return roles
        for item in value:
            if isinstance(item, dict):
                role = str(item.get("role_id") or item.get("role") or "").strip()
                if role:
                    roles.append(role)
        return roles

    def _method_generator_report_contract_for_routes(self, routes: list[tuple[str, object]], spec: dict[str, object]) -> tuple[str, str]:
        target = self._method_generator_target_temperature(spec)
        notes: list[str] = []
        statuses: list[str] = []
        for _role, record in routes:
            if isinstance(record, FoqAlignmentRecord) and record.family == "TCC" and record.test_intent == "temperature_accuracy" and target is not None:
                status, note = self._method_generator_accuracy_report_contract_for_target(record, target)
                statuses.append(status)
                notes.append(note)
        if any(status == "blocked" for status in statuses):
            return ("blocked", " ; ".join(notes))
        if any(status == "open" for status in statuses):
            return ("open", " ; ".join(notes))
        if notes:
            return ("ok", " ; ".join(notes))
        return ("not_checked", "No target-specific report contract rule was applied.")

    def _method_generator_operation_contract_summary(
        self,
        spec: dict[str, object],
        role_audit: MethodRoleMapAudit | None,
        routes: list[tuple[str, object]],
    ) -> str:
        if not isinstance(role_audit, MethodRoleMapAudit):
            return "No Method Role Map contract was resolved."
        target = self._method_generator_target_temperature(spec)
        baseline = self._method_generator_baseline_temperature(spec)
        primary = next((target_record for _role, target_record in routes if isinstance(target_record, FoqAlignmentRecord)), None)
        if (
            isinstance(primary, FoqAlignmentRecord)
            and primary.family == "TCC"
            and primary.test_intent == "temperature_accuracy"
            and target is not None
        ):
            report_status, report_note = self._method_generator_accuracy_report_contract_for_target(primary, target)
            baseline_text = f"baseline {baseline:g} C -> " if baseline is not None else ""
            return (
                f"Operation contract needed: TCC accuracy {baseline_text}target {target:g} C. "
                "Editable roles: temperature_ladder_assignment + accuracy_measurement_setpoint. "
                "Locked roles: device_branch, ret_time_anchor, external_stability_gate, final_reset. "
                f"Report contract: {report_status}: {report_note}"
            )
        return (
            "Operation contract needed before script mutation. Role Map is available, but this intent has no certified "
            "row-level operation policy yet."
        )

    def _method_generator_accuracy_report_contract_for_target(self, record: FoqAlignmentRecord, target: float) -> tuple[str, str]:
        points: list[float] = []
        for field in record.db_fields:
            match = re.search(r"TempAcc(-?\d+(?:\.\d+)?)", str(field), flags=re.I)
            if not match:
                continue
            try:
                points.append(float(match.group(1)))
            except ValueError:
                continue
        unique_points = sorted({point for point in points})
        if not unique_points:
            return ("open", "No TempAcc* DB/report fields were resolved for this record.")
        if any(abs(point - target) <= 1e-9 for point in unique_points):
            return ("ok", f"target {target:g} C exists in mapped TempAcc fields: {', '.join(f'{point:g}' for point in unique_points)}.")
        return (
            "blocked",
            f"target {target:g} C is not present in mapped TempAcc fields ({', '.join(f'{point:g}' for point in unique_points)}); report/DB contract redesign is required.",
        )

    def _render_method_generator_pending_flow(self, intent: str) -> None:
        frame = getattr(self, "method_generator_flow_table", None)
        if frame is None:
            return
        table = self._table_widget(frame)
        table.delete(*table.get_children())
        rows = (
            ("1. Natural language", intent, "UI input", "parsed"),
            ("2. AI structured spec", "<running>", "AI call is capped; local fallback will be used if it times out", "running"),
            ("3. Local KB route", "<queued>", "FOQ alignment + Method Script KB", "queued"),
            ("4. Method Role Map", "<queued>", "TCC_METHOD_ROLE_MAP.json + Method Script KB", "queued"),
            ("5. Local CM mechanism plan", "<queued>", "method_semantic_analyzer + role contract", "queued"),
            ("6. AI composition / row-edit plan", "<queued>", "AI is primary planner; local fallback if unavailable", "queued"),
            ("7. Local guardrail validation", "<queued>", "Role Map + locked-role guardrails + report contract", "queued"),
            ("8. Script rendering", "<queued>", "AI/local plan rendered with role-map checks", "queued"),
            ("9. Report contract", "<queued>", "report template / DB mapping / RetTime evidence", "queued"),
        )
        for index, row in enumerate(rows, start=1):
            table.insert("", "end", iid=f"flow:pending:{index}", values=row)

    def _refresh_method_generator_ai_config_from_disk(self) -> None:
        try:
            defaults = self._load_ai_config_defaults()
        except Exception:
            return
        if not self.ai_api_key_var.get().strip() and defaults.get("api_key"):
            self.ai_api_key_var.set(str(defaults.get("api_key") or ""))
        current_model = self.ai_model_var.get().strip()
        if (not current_model or current_model == "chatgpt-5.4") and defaults.get("model"):
            self.ai_model_var.set(str(defaults.get("model") or "gpt-5.5"))
        if not self.ai_base_url_var.get().strip() and defaults.get("base_url"):
            self.ai_base_url_var.set(str(defaults.get("base_url") or "https://api.openai.com/v1"))

    def _normalize_method_generator_spec(self, spec: dict[str, object], local_spec: dict[str, object]) -> dict[str, object]:
        normalized = dict(spec)
        intent = str(local_spec.get("user_intent") or normalized.get("user_intent") or "")
        if self._method_generator_is_setup_stability_for_accuracy(intent):
            normalized["primary_intent"] = "temperature_accuracy"
            related = normalized.get("related_intents")
            if isinstance(related, list):
                normalized["related_intents"] = [item for item in related if str(item) != "temperature_stability"]
        for key in ("temperatures_c", "transition_c", "target_c", "baseline_c", "duration_minutes", "accuracy_points_c"):
            local_value = local_spec.get(key)
            if local_value not in (None, "", []):
                normalized[key] = local_value
        return normalized

    def _method_generator_local_spec(self, intent: str, family: str) -> dict[str, object]:
        lowered = intent.lower()
        if "外部" in intent:
            lowered += " external"
        if "内部" in intent:
            lowered += " internal"
        device = "VH-C10-A"
        for candidate in ("VH-C10-A", "VC-C10-A", "VA-C10-A"):
            if candidate.lower() in lowered or candidate.split("-")[0].lower() in lowered:
                device = candidate
                break
        intents: list[str] = []
        if any(token in lowered for token in ("stability", "稳定")):
            intents.append("temperature_stability")
        if any(token in lowered for token in ("accuracy", "准确")):
            intents.append("temperature_accuracy")
        if any(token in lowered for token in ("precision", "精密", "重复")):
            intents.append("temperature_precision")
        if any(token in lowered for token in ("heatup", "cooldown", "heat up", "cool down", "升温", "降温", "爬坡")):
            if "heatup_cooldown_20_50_20" not in intents:
                intents.append("heatup_cooldown_20_50_20")
        if any(token in lowered for token in ("valve", "阀", "trigger", "切换", "转动")):
            intents.append("valve_stress")
        if "稳定" in intent and "temperature_stability" not in intents:
            intents.insert(0, "temperature_stability")
        if "准确" in intent and "temperature_accuracy" not in intents:
            intents.append("temperature_accuracy")
        if any(token in intent for token in ("精密", "重复")) and "temperature_precision" not in intents:
            intents.append("temperature_precision")
        if any(token in intent for token in ("升温", "降温", "爬坡")) and "heatup_cooldown_20_50_20" not in intents:
            intents.append("heatup_cooldown_20_50_20")
        if any(token in intent for token in ("阀", "切换", "转动")) and "valve_stress" not in intents:
            intents.append("valve_stress")
        intents = self._method_generator_order_intents_by_text(intent, intents)
        if self._method_generator_is_setup_stability_for_accuracy(intent):
            intents = [item for item in intents if item != "temperature_stability"]
            if "temperature_accuracy" in intents:
                intents = ["temperature_accuracy", *[item for item in intents if item != "temperature_accuracy"]]
            else:
                intents.insert(0, "temperature_accuracy")
        if not intents:
            intents.append("temperature_stability")
        numbers = self._test_plan_numbers_from_text(intent)
        temperature_numbers = self._method_generator_temperatures_from_text(intent)
        transition = self._test_plan_temperature_transition_from_text(intent)
        if transition is None and len(temperature_numbers) >= 2 and self._method_generator_text_has_ramp(intent):
            transition = (temperature_numbers[0], temperature_numbers[-1])
        duration_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:min|minute|minutes|分钟)", intent, flags=re.I)
        duration_minutes = float(duration_match.group(1)) if duration_match else None
        trigger_seconds = None
        trigger_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:s|sec|second|seconds|秒)", intent, flags=re.I)
        if trigger_match and any(item == "valve_stress" for item in intents):
            trigger_seconds = float(trigger_match.group(1))
        target = numbers[-1] if numbers else None
        trigger_matches = [float(match) for match in re.findall(r"(\d+(?:\.\d+)?)\s*(?:s|sec|second|seconds|秒)", intent, flags=re.I)]
        if trigger_matches and any(item == "valve_stress" for item in intents):
            trigger_seconds = trigger_matches[0]
        if temperature_numbers:
            target = temperature_numbers[-1]
        if transition:
            target = transition[-1]
        intent_targets: dict[str, float] = {}
        if "temperature_stability" in intents and temperature_numbers:
            intent_targets["temperature_stability"] = temperature_numbers[0]
            stability_target = self._method_generator_explicit_stability_target(intent)
            if stability_target is not None:
                intent_targets["temperature_stability"] = stability_target
        if "temperature_accuracy" in intents and temperature_numbers:
            intent_targets["temperature_accuracy"] = temperature_numbers[-1]
        if "temperature_precision" in intents and temperature_numbers:
            intent_targets["temperature_precision"] = temperature_numbers[0]
        primary_target = intent_targets.get(intents[0], target)
        baseline_c = None
        if temperature_numbers and len(temperature_numbers) >= 2:
            primary_target_float = self._method_generator_float_or_none(primary_target)
            if primary_target_float is not None:
                for value in temperature_numbers:
                    if abs(value - primary_target_float) > 1e-9:
                        baseline_c = value
                        break
        if transition and len(transition) >= 2:
            baseline_c = transition[0]
        unique_temperatures: list[float] = []
        for value in temperature_numbers:
            if not any(abs(value - existing) <= 1e-9 for existing in unique_temperatures):
                unique_temperatures.append(value)
        stable_baseline_for_accuracy = (
            "temperature_accuracy" in intents
            and self._method_generator_is_setup_stability_for_accuracy(intent)
        )
        explicit_multi_point_accuracy = (
            "temperature_accuracy" in intents
            and len(unique_temperatures) > 1
            and any(token in intent.lower() for token in ("和", "以及", "and", ",", "，", "/", "+"))
        )
        accuracy_points = unique_temperatures if explicit_multi_point_accuracy and not stable_baseline_for_accuracy else []
        return {
            "user_intent": intent,
            "family": family,
            "device_model": device,
            "primary_intent": intents[0],
            "related_intents": intents[1:],
            "temperatures_c": temperature_numbers or numbers,
            "transition_c": list(transition) if transition else [],
            "target_c": primary_target,
            "baseline_c": baseline_c,
            "accuracy_points_c": accuracy_points,
            "final_target_c": target,
            "intent_targets_c": intent_targets,
            "duration_minutes": duration_minutes,
            "trigger_intervals_seconds": trigger_matches,
            "trigger_interval_seconds": trigger_seconds,
            "sensor_scope": "external only" if any(token in lowered for token in ("external", "外部")) else ("internal only" if any(token in lowered for token in ("internal", "内部")) else "source method default"),
        }

    def _method_generator_is_setup_stability_for_accuracy(self, text: str) -> bool:
        lowered = (text or "").lower()
        has_accuracy = any(token in lowered for token in ("accuracy", "准确"))
        if not has_accuracy:
            return False
        explicit_stability_test = any(token in lowered for token in ("stability", "稳定性", "测试稳定性", "测稳定性"))
        if explicit_stability_test:
            return False
        has_setup_stability = any(token in lowered for token in ("稳定态", "稳定后", "稳定", "平衡", "equilibrated", "after stable", "baseline"))
        has_after_marker = any(token in lowered for token in ("后", "after", "then", "之后"))
        return has_setup_stability and has_after_marker

    def _method_generator_text_has_ramp(self, text: str) -> bool:
        lowered = (text or "").lower()
        return any(token in lowered for token in ("从", "上升", "升到", "升至", "到", "至", "爬坡", "from", "to", "ramp"))

    def _method_generator_order_intents_by_text(self, text: str, intents: list[str]) -> list[str]:
        seen: list[str] = []
        for item in intents:
            if item not in seen:
                seen.append(item)
        positions = {item: self._method_generator_intent_position(text, item) for item in seen}
        return sorted(seen, key=lambda item: (positions.get(item, 10**9), seen.index(item)))

    def _method_generator_intent_position(self, text: str, intent: str) -> int:
        lowered = (text or "").lower()
        keyword_map = {
            "temperature_stability": ("stability", "稳定性", "测试稳定", "稳定"),
            "temperature_accuracy": ("accuracy", "准确性", "准确"),
            "temperature_precision": ("precision", "精密", "重复"),
            "heatup_cooldown_20_50_20": ("heatup", "cooldown", "heat up", "cool down", "升温", "降温", "爬坡"),
            "valve_stress": ("valve", "trigger", "阀", "切换", "转动"),
        }
        positions: list[int] = []
        for keyword in keyword_map.get(intent, (intent,)):
            index = lowered.find(keyword.lower())
            if index >= 0:
                positions.append(index)
        return min(positions) if positions else 10**9

    def _method_generator_temperatures_from_text(self, text: str) -> list[float]:
        values: list[float] = []
        for match in re.finditer(r"(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|℃|摄氏度|度)", text or "", flags=re.I):
            try:
                values.append(float(match.group(1)))
            except ValueError:
                continue
        return values

    def _method_generator_explicit_stability_target(self, text: str) -> float | None:
        patterns = (
            r"(?:升到|上升到|到|至)\s*(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|℃|摄氏度|度)[^。；;,.]{0,20}(?:测试)?稳定性",
            r"stability\s+(?:at|to)\s*(-?\d+(?:\.\d+)?)\s*(?:°?\s*C|℃)?",
        )
        for pattern in patterns:
            match = re.search(pattern, text or "", flags=re.I)
            if not match:
                continue
            try:
                return float(match.group(1))
            except ValueError:
                continue
        return None

    def _call_method_generator_ai_spec(self, local_spec: dict[str, object]) -> dict[str, object]:
        family = str(local_spec.get("family") or "TCC")
        candidates = [
            {
                "test_intent": item.test_intent,
                "td_test": item.td_test,
                "devices": list(item.device_models),
                "method": item.instrument_method,
            }
            for item in filter_alignment_records(self._test_plan_records(), family=family)[:40]
        ]
        prompt = (
            "Return JSON only. Convert the user natural language into a structured CMBX method generation spec. "
            "Do not invent CM commands. Choose primary_intent and related_intents from available_candidates. "
            "Keys: family, device_model, primary_intent, related_intents, temperatures_c, transition_c, target_c, "
            "duration_minutes, trigger_interval_seconds, sensor_scope, assumptions, open_questions. "
            "Keep the JSON compact; do not include method scripts.\n\n"
            f"local_spec={json.dumps(local_spec, ensure_ascii=False)}\n"
            f"available_candidates={json.dumps(candidates, ensure_ascii=False)}"
        )
        body = {
            "model": self.ai_model_var.get().strip() or "gpt-5.5",
            "messages": [
                {"role": "system", "content": "Return JSON only. You are a structured intent parser for local KB routing."},
                {"role": "user", "content": prompt},
            ],
            "max_completion_tokens": 900,
        }
        endpoint = self._ai_chat_completions_endpoint()
        request = urllib.request.Request(
            endpoint,
            data=json.dumps(body).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {self.ai_api_key_var.get().strip()}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=15) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ValueError(f"HTTP {exc.code} from {endpoint} using model {body['model']}: {detail[:500]}") from exc
        content = payload["choices"][0]["message"]["content"]
        match = re.search(r"\{.*\}", content, flags=re.S)
        if not match:
            raise ValueError("AI response did not contain JSON.")
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else {}

    def _ai_chat_completions_endpoint(self) -> str:
        base = self.ai_base_url_var.get().strip() or "https://api.openai.com/v1"
        base = base.rstrip("/")
        if base.endswith("/chat/completions"):
            return base
        if re.fullmatch(r"https?://api\.openai\.com", base):
            return f"{base}/v1/chat/completions"
        return f"{base}/chat/completions"

    def _method_generator_route_records(self, spec: dict[str, object]) -> list[tuple[str, object]]:
        family = str(spec.get("family") or "TCC")
        device = str(spec.get("device_model") or "")
        routed: list[tuple[str, object]] = []
        intents = [str(spec.get("primary_intent") or "")]
        related = spec.get("related_intents")
        if isinstance(related, list):
            intents.extend(str(item) for item in related)
        for index, intent in enumerate([item for item in intents if item]):
            if intent == "valve_stress":
                entry = self._method_generator_valve_stress_entry(spec)
                if entry is not None:
                    routed.append(("Primary" if index == 0 else "Related", entry))
                continue
            token = self._resolve_test_plan_intent_token(intent, family)
            records = filter_alignment_records(self._test_plan_records(), family=family, test_text=token)
            if device:
                device_records = filter_alignment_records(records, devices=(device,))
                if device_records:
                    records = device_records
            if records:
                routed.append(("Primary" if index == 0 else "Related", records[0]))
        return routed

    def _method_generator_valve_stress_entry(self, spec: dict[str, object]) -> MethodScriptKbEntry | None:
        family = str(spec.get("family") or "TCC")
        device = str(spec.get("device_model") or "VH-C10-A")
        target = self._method_generator_target_temperature(spec)
        candidates: list[str] = []
        if target is not None:
            if abs(target - 70.0) <= 1e-9:
                candidates.extend((
                    "Stress test_5s_Preheater Box_pcc70c",
                    "Stress test_5s_RealPreheater_pcc_ON_10min equibration",
                    "Stress test_5s_RealPreheater_pcc_on",
                ))
            elif abs(target - 60.0) <= 1e-9:
                candidates.append("Stress test_5s_Preheater Box_pcc60c")
            elif abs(target - 40.0) <= 1e-9:
                candidates.append("Stress test_5s_Preheater Box_pcc40c")
        candidates.extend((
            "Stress test_5s_RealPreheater_pcc_ON_10min equibration",
            "Stress test_5s_RealPreheater_pcc_on",
            "stress test_5s",
        ))
        for method_name in candidates:
            entry = find_method_script_kb_entry(method_name, family=family, device_model=device, workspace_root=Path.cwd())
            if entry is not None:
                return entry
        return None

    def _method_generator_method_entry_summary(self, entry: MethodScriptKbEntry, rows: list[tuple[object, ...]]) -> str:
        semantic = analyze_cm_method_rows(rows)
        parts = [entry.method_name]
        if semantic.triggers:
            intervals = []
            for trigger in semantic.triggers[:4]:
                if trigger.rearm_minutes is not None:
                    intervals.append(f"{trigger.name}={trigger.rearm_minutes * 60:g}s")
                else:
                    intervals.append(trigger.name)
            parts.append(f"{len(semantic.triggers)} trigger(s): " + ", ".join(intervals))
        if semantic.temperature_setpoints:
            temps = [item.numeric_value for item in semantic.temperature_setpoints if item.numeric_value is not None]
            if temps:
                parts.append("setpoints " + ", ".join(f"{value:g}C" for value in temps[:5]))
        return "; ".join(parts)

    def _method_generator_source_rows(self, record: FoqAlignmentRecord, device: str) -> list[tuple[object, ...]]:
        rows = load_method_script_rows_from_kb(
            record.instrument_method,
            family=record.family or "TCC",
            device_model=device or record.device_label,
            workspace_root=Path.cwd(),
        )
        return [tuple(row) for row in rows]

    def _method_generator_generate_rows(
        self,
        record: FoqAlignmentRecord,
        source_rows: list[tuple[object, ...]],
        spec: dict[str, object],
        device: str,
    ) -> list[tuple[tuple[object, ...], bool]]:
        target = self._method_generator_target_temperature(spec)
        if record.family == "TCC" and record.test_intent == "temperature_accuracy" and target is not None:
            accuracy_points = spec.get("accuracy_points_c")
            if isinstance(accuracy_points, list) and len(accuracy_points) > 1:
                points = [self._method_generator_float_or_none(item) for item in accuracy_points]
                clean_points = [item for item in points if item is not None]
                if len(clean_points) > 1:
                    return self._apply_accuracy_multi_point_review_to_cm_rows(source_rows, clean_points)
            return self._apply_accuracy_setpoint_to_cm_rows(source_rows, target, spec=spec, device=device)
        if record.family == "TCC" and "stability" in record.test_intent and target is not None:
            return self._apply_stability_setpoint_to_cm_rows(source_rows, target, spec)
        if record.family == "TCC" and record.test_intent == "heatup_cooldown_20_50_20":
            previous = self.test_plan_parameter_var.get()
            transition = spec.get("transition_c")
            points = transition if isinstance(transition, list) and len(transition) >= 2 else None
            if points is None:
                temperatures = spec.get("temperatures_c")
                if isinstance(temperatures, list) and len(temperatures) >= 3:
                    points = temperatures[-3:]
            if isinstance(points, list) and len(points) >= 2:
                self.test_plan_parameter_var.set("->".join(self._format_test_plan_number(float(item)) for item in points))
            try:
                return self._apply_heatup_cooldown_to_cm_rows(source_rows, record, device)
            finally:
                self.test_plan_parameter_var.set(previous)
        return [(row, False) for row in source_rows]

    def _apply_accuracy_multi_point_review_to_cm_rows(
        self,
        source_rows: list[tuple[object, ...]],
        points: list[float],
    ) -> list[tuple[tuple[object, ...], bool]]:
        point_text = ", ".join(f"{point:g} C" for point in points)
        warning = (
            "",
            "Comment",
            "",
            f"BLOCKED: multi-point Temperature Accuracy requested ({point_text}).",
            "",
            "Do not collapse this to a single target. Requires ladder/RetTime/report-cell selection and possible row deletion/renumbering review.",
            "modified",
        )
        output: list[tuple[tuple[object, ...], bool]] = [(warning, True)]
        semantic = analyze_cm_method_rows(source_rows)
        requested = set(points)
        assignment_by_row: dict[int, float] = {}
        for variable in semantic.temperature_variables:
            for assignment in semantic.assignments_for_variable(variable):
                if assignment.numeric_value is not None:
                    assignment_by_row[assignment.row_index] = assignment.numeric_value
        for index, row in enumerate(source_rows):
            row_list = list(row)
            if len(row_list) < 7:
                row_list.extend([""] * (7 - len(row_list)))
            changed = False
            value = assignment_by_row.get(index)
            if value is not None:
                if any(abs(value - point) <= 1e-9 for point in requested):
                    row_list[5] = self._append_preview_note(row_list[5], f"Generator review: requested accuracy point {value:g} C.")
                    row_list[6] = "modified"
                    changed = True
                else:
                    row_list[5] = self._append_preview_note(row_list[5], f"Generator review: non-requested source accuracy point {value:g} C; remove only after RetTime/report review.")
                    row_list[6] = "removed"
                    changed = True
            output.append((tuple(row_list), changed))
        return output

    def _method_generator_target_temperature(self, spec: dict[str, object]) -> float | None:
        for key in ("target_c", "setpoint_c"):
            value = spec.get(key)
            try:
                if value is not None and str(value) != "":
                    return float(value)
            except (TypeError, ValueError):
                pass
        values = spec.get("temperatures_c")
        if isinstance(values, list) and values:
            try:
                return float(values[-1])
            except (TypeError, ValueError):
                return None
        return None

    def _method_generator_baseline_temperature(self, spec: dict[str, object]) -> float | None:
        baseline = self._method_generator_float_or_none(spec.get("baseline_c"))
        target = self._method_generator_target_temperature(spec)
        if baseline is not None and (target is None or abs(baseline - target) > 1e-9):
            return baseline
        values = spec.get("transition_c")
        if isinstance(values, list) and len(values) >= 2:
            first = self._method_generator_float_or_none(values[0])
            last = self._method_generator_float_or_none(values[-1])
            if first is not None and (last is None or abs(first - last) > 1e-9):
                return first
        return None

    def _method_generator_float_or_none(self, value: object) -> float | None:
        try:
            if value is None or str(value).strip() == "":
                return None
            return float(value)
        except (TypeError, ValueError):
            return None

    def _apply_stability_setpoint_to_cm_rows(
        self,
        source_rows: list[tuple[object, ...]],
        target: float,
        spec: dict[str, object],
    ) -> list[tuple[tuple[object, ...], bool]]:
        output: list[tuple[tuple[object, ...], bool]] = []
        changed_any = False
        duration = self._method_generator_float_or_none(spec.get("duration_minutes"))
        baseline = self._method_generator_baseline_temperature(spec)
        sensor_scope = str(spec.get("sensor_scope") or "")
        inserted_baseline = False
        for row in source_rows:
            row_list = list(row)
            changed = False
            command_text = str(row_list[3])
            command_lower = command_text.lower()
            is_cc_stability_setpoint = (
                str(row_list[1]) == "Command"
                and command_lower == "columncomp.cc.temperature.nominal"
            )
            if baseline is not None and is_cc_stability_setpoint and not inserted_baseline:
                row_id = str(row_list[0])
                output.extend((
                    ((f"{row_id}.1", "Comment", "", f"Generator plan: pre-equilibrate CC at {baseline:g} C before the target stability run.", "", "Inserted from natural-language baseline request.", "modified"), True),
                    ((f"{row_id}.2", "Command", "", "ColumnComp.CC.Temperature.Nominal", f"{baseline:g}", "Generator: set requested baseline/pre-equilibration temperature.", "modified"), True),
                    ((f"{row_id}.3", "Command", "", "Wait", "CC.TempReady AND PCC.TempReady", "Generator: wait for requested baseline to become ready; verify PCC dependency for non-VH methods.", "modified"), True),
                ))
                if duration:
                    output.append(((f"{row_id}.4", "Command", "", "Delay", f"{duration:g}", "Generator: requested baseline hold duration in minutes; report window still requires review.", "modified"), True))
                output.append(((f"{row_id}.5", "Comment", "", f"Generator plan: ramp from {baseline:g} C to {target:g} C for stability measurement.", "", "The following source setpoint is the target measurement setpoint.", "modified"), True))
                inserted_baseline = True
                changed_any = True
            if len(row_list) >= 6 and is_cc_stability_setpoint:
                numeric = cm_numeric_value(str(row_list[4]))
                if numeric is not None and abs(numeric - target) > 1e-9:
                    row_list[4] = re.sub(r"-?\d+(?:\.\d+)?", f"{target:g}", str(row_list[4]), count=1)
                    row_list[5] = self._append_preview_note(row_list[5], f"Generator: set stability temperature to {target:g} C.")
                    row_list[6] = "modified"
                    changed = True
                    changed_any = True
                elif baseline is not None and numeric is not None:
                    row_list[5] = self._append_preview_note(row_list[5], f"Generator: target stability setpoint remains {target:g} C after baseline.")
                    row_list[6] = "modified"
                    changed = True
            elif (
                str(row_list[1]) == "Command"
                and "columncomp.pcc.temperature.nominal" in command_lower
                and baseline is not None
            ):
                row_list[5] = self._append_preview_note(row_list[5], "Generator note: PCC branch setpoint preserved; it is not the CC stability target.")
            if sensor_scope in {"external only", "internal only"} and any(token in str(row_list[3]).lower() for token in ("data_collection_rate", "log")):
                row_list[5] = self._append_preview_note(row_list[5], f"Generator note: requested sensor scope = {sensor_scope}; channel/report contract must be checked.")
            output.append((tuple(row_list), changed))
        if not changed_any:
            output.insert(0, (("", "Comment", "", f"Generator could not find a direct stability setpoint row for {target:g} C.", "", "Review Method Role Map / KB script.", "modified"), True))
        return output

    def _render_method_generator_result(self, result: dict[str, object]) -> None:
        spec = result.get("spec")
        routes = result.get("routes") if isinstance(result.get("routes"), list) else []
        source_rows = result.get("source_rows") if isinstance(result.get("source_rows"), list) else []
        generated_rows = result.get("generated_rows") if isinstance(result.get("generated_rows"), list) else []
        self._set_method_generator_spec_text(
            json.dumps(
                {
                    "structured_spec": spec,
                    "ai_composition_plan": result.get("operation_contract_draft") or {},
                    "local_guardrail_validation": result.get("operation_validation") or {},
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        self._render_method_generator_flow(result)
        self._render_method_generator_routes(routes)
        self._render_method_generator_report_contract(routes, spec if isinstance(spec, dict) else {})
        self._render_method_generator_method_tables(source_rows, generated_rows)
        self.status_var.set("Method Script Generator finished")
        self.progress_var.set(100.0)

    def _set_method_generator_spec_text(self, text: str) -> None:
        widget = getattr(self, "method_generator_spec_text", None)
        if widget is None:
            return
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert("1.0", text)
        widget.configure(state="disabled")

    def _render_method_generator_flow(self, result: dict[str, object]) -> None:
        frame = getattr(self, "method_generator_flow_table", None)
        if frame is None:
            return
        table = self._table_widget(frame)
        table.delete(*table.get_children())
        spec = result.get("spec") if isinstance(result.get("spec"), dict) else {}
        routes = result.get("routes") if isinstance(result.get("routes"), list) else []
        source_rows = result.get("source_rows") if isinstance(result.get("source_rows"), list) else []
        generated_rows = result.get("generated_rows") if isinstance(result.get("generated_rows"), list) else []
        role_audit = result.get("role_audit")
        role_status = role_audit.status if isinstance(role_audit, MethodRoleMapAudit) else "missing"
        role_summary = role_audit.summary if isinstance(role_audit, MethodRoleMapAudit) else "No Method Role Map audit."
        operation_mode = role_audit.generation_mode if isinstance(role_audit, MethodRoleMapAudit) else "blocked"
        row_transformer_status = str(result.get("row_transformer_status") or "")
        script_status = "preview only" if row_transformer_status == "generated" and generated_rows else "blocked"
        operation_contract = str(result.get("operation_contract") or "")
        operation_draft = result.get("operation_contract_draft") if isinstance(result.get("operation_contract_draft"), dict) else {}
        operation_validation = result.get("operation_validation") if isinstance(result.get("operation_validation"), dict) else {}
        operation_type = str(operation_draft.get("operation_type") or "<missing>")
        validation_status = str(operation_validation.get("status") or "not run")
        validation_summary = str(operation_validation.get("summary") or "")
        report_status, report_evidence = self._method_generator_overall_report_contract_status(routes, spec)
        script_result = f"{len(generated_rows)} row(s)"
        if script_status == "blocked" and operation_contract:
            script_result = operation_contract
        rows = (
            ("1. Natural language", str(spec.get("user_intent") or self.method_generator_intent_var.get()), str(spec.get("source") or ""), "parsed"),
            ("2. AI structured spec", str(spec.get("primary_intent") or ""), json.dumps({k: spec.get(k) for k in ("device_model", "baseline_c", "target_c", "duration_minutes", "sensor_scope")}, ensure_ascii=False), str(result.get("ai_status") or "unknown")),
            ("3. Local KB route", f"{len(routes)} routed method(s)", "FOQ alignment + Method Script KB", "ok" if routes else "blocked"),
            ("4. Method Role Map", role_summary, "TCC_METHOD_ROLE_MAP.json + Method Script KB", role_status),
            ("5. Local CM mechanism plan", str(result.get("cm_summary") or ""), "method_semantic_analyzer + role contract", "ok" if source_rows and role_status == "complete" else "blocked"),
            ("6. AI composition / row-edit plan", operation_type, str(result.get("operation_contract_ai_status") or "not run"), "ok" if operation_draft else "blocked"),
            ("7. Local guardrail validation", validation_summary, "Role Map + locked-role guardrails + report contract", validation_status),
            ("8. Script rendering", script_result, f"AI/local plan rendered with role-map checks; mode={operation_mode}; transformer={row_transformer_status or 'not run'}", script_status),
            ("9. Report contract", report_evidence, "report template / DB mapping / RetTime evidence", report_status),
        )
        for index, row in enumerate(rows, start=1):
            table.insert("", "end", iid=f"flow:{index}", values=row)

    def _render_method_generator_routes(self, routes: list[object]) -> None:
        frame = getattr(self, "method_generator_route_table", None)
        if frame is None:
            return
        table = self._table_widget(frame)
        table.delete(*table.get_children())
        if not routes:
            table.insert("", "end", iid="route:none", values=("Open", "<none>", "<missing>", "No KB route found"))
            return
        for index, item in enumerate(routes, start=1):
            role, target = item
            if isinstance(target, FoqAlignmentRecord):
                values = (role, target.test_intent, target.instrument_method, target.coverage_status or "mapped")
            elif isinstance(target, MethodScriptKbEntry):
                values = (role, "valve_stress / trigger KB", target.method_name, f"method script KB: {target.source}")
            else:
                values = (role, "<unknown>", "<missing>", "Open")
            table.insert("", "end", iid=f"route:{index}", values=values)

    def _render_method_generator_report_contract(self, routes: list[object], spec: dict[str, object]) -> None:
        frame = getattr(self, "method_generator_report_table", None)
        if frame is None:
            return
        table = self._table_widget(frame)
        table.delete(*table.get_children())
        if not routes:
            table.insert("", "end", iid="report:none", values=("<none>", "<missing>", "<missing>", "No report contract can be checked."))
            return
        duration = self._method_generator_float_or_none(spec.get("duration_minutes"))
        for index, item in enumerate(routes, start=1):
            _role, record = item
            if isinstance(record, MethodScriptKbEntry):
                table.insert(
                    "",
                    "end",
                    iid=f"report:{index}",
                    values=("valve_stress / trigger KB", record.method_name, "Log valve position / trigger timing", "Mechanism evidence only; report formula must be designed/verified separately."),
                )
                continue
            if not isinstance(record, FoqAlignmentRecord):
                continue
            fields = ", ".join(record.db_fields[:6])
            if len(record.db_fields) > 6:
                fields += f", +{len(record.db_fields) - 6}"
            constraint = "OK: report binding exists" if record.report_template or fields else "Open: missing report/formula binding"
            target = self._method_generator_target_temperature(spec)
            if record.family == "TCC" and record.test_intent == "temperature_accuracy" and target is not None:
                status, note = self._method_generator_accuracy_report_contract_for_target(record, target)
                constraint = f"{status.upper()}: {note}"
            if duration and "stability" in record.test_intent:
                constraint = f"Review: requested {duration:g} min may conflict with fixed stability report windows."
            table.insert("", "end", iid=f"report:{index}", values=(record.test_intent, record.report_template or "<missing>", fields or "<missing>", constraint))

    def _method_generator_overall_report_contract_status(self, routes: list[object], spec: dict[str, object]) -> tuple[str, str]:
        target = self._method_generator_target_temperature(spec)
        statuses: list[str] = []
        notes: list[str] = []
        for item in routes:
            _role, record = item
            if isinstance(record, FoqAlignmentRecord) and record.family == "TCC" and record.test_intent == "temperature_accuracy" and target is not None:
                status, note = self._method_generator_accuracy_report_contract_for_target(record, target)
                statuses.append(status)
                notes.append(note)
        if any(status == "blocked" for status in statuses):
            return ("blocked: report redesign required", " ; ".join(notes))
        if any(status == "open" for status in statuses):
            return ("open verification", " ; ".join(notes))
        if notes:
            return ("ok", " ; ".join(notes))
        return ("method changes must not violate report windows", "checked below")

    def _render_method_generator_method_tables(self, source_rows: list[object], generated_rows: list[object]) -> None:
        source_frame = getattr(self, "method_generator_original_table", None)
        generated_frame = getattr(self, "method_generator_generated_table", None)
        if source_frame is None or generated_frame is None:
            return
        source_table = self._table_widget(source_frame)
        generated_table = self._table_widget(generated_frame)
        source_table.delete(*source_table.get_children())
        generated_table.delete(*generated_table.get_children())
        for index, row in enumerate(source_rows, start=1):
            tag = str(row[6]) if len(row) > 6 else ""
            source_table.insert("", "end", iid=f"mgsrc:{index}", values=tuple(row)[:6], tags=(tag,) if tag else ())
        for index, item in enumerate(generated_rows, start=1):
            row, changed = item
            tag = "modified" if changed else (str(row[6]) if len(row) > 6 else "")
            generated_table.insert("", "end", iid=f"mgg:{index}", values=tuple(row)[:6], tags=(tag,) if tag else ())

    def _method_generator_failed(self, exc: Exception) -> None:
        self.progress_var.set(0.0)
        self.status_var.set("Method Script Generator failed")
        messagebox.showerror(APP_NAME, f"Method Script Generator failed:\n{exc}")

    def show_ai_settings(self) -> None:
        top = tk.Toplevel(self.root)
        top.title("AI 设置")
        top.configure(bg=self.colors["bg"])
        top.transient(self.root)
        top.geometry("640x300")
        top.minsize(560, 260)

        shell = tk.Frame(top, bg=self.colors["bg"])
        shell.pack(fill="both", expand=True, padx=24, pady=22)
        shell.columnconfigure(1, weight=1)
        tk.Label(shell, text="AI Analysis Settings", font=self._font(14, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 14))

        tk.Label(shell, text="Base URL", font=self._font(9, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        self._make_entry(shell, self.ai_base_url_var).grid(row=1, column=1, sticky="ew", pady=6, ipady=4)

        tk.Label(shell, text="Model ID", font=self._font(9, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        self._make_entry(shell, self.ai_model_var).grid(row=2, column=1, sticky="ew", pady=6, ipady=4)

        tk.Label(shell, text="API Key", font=self._font(9, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=3, column=0, sticky="w", padx=(0, 10), pady=6)
        key_entry = tk.Entry(shell, textvariable=self.ai_api_key_var, show="*", font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1)
        key_entry.grid(row=3, column=1, sticky="ew", pady=6, ipady=4)

        note = "OpenAI-compatible /chat/completions endpoint is expected. Base URL may be https://api.openai.com/v1 or a full /chat/completions endpoint. If a 404 occurs, check both Base URL and Model ID. If key is empty or request fails, local KB routing is used."
        tk.Label(shell, text=note, wraplength=560, justify="left", font=self._font(8), bg=self.colors["bg"], fg=self.colors["text_secondary"]).grid(row=4, column=0, columnspan=2, sticky="w", pady=(10, 6))

        buttons = tk.Frame(shell, bg=self.colors["bg"])
        buttons.grid(row=5, column=0, columnspan=2, sticky="e", pady=(12, 0))

        def save_and_close() -> None:
            try:
                self._save_ai_config()
            except Exception as exc:
                messagebox.showerror(APP_NAME, str(exc))
                return
            self.status_var.set(f"Saved AI settings: {self.ai_model_var.get().strip() or 'gpt-5.5'}")
            top.destroy()

        self._make_button(buttons, "Save", save_and_close, kind="primary", width=10).grid(row=0, column=0, padx=(0, 8))
        self._make_button(buttons, "Close", top.destroy, kind="neutral", width=10).grid(row=0, column=1)

    def _show_markdown_window(self, title: str, text: str) -> None:
        top = tk.Toplevel(self.root)
        top.title(title)
        top.configure(bg=self.colors["bg"])
        top.transient(self.root)
        top.geometry("980x700")
        top.minsize(760, 520)

        shell = tk.Frame(top, bg=self.colors["bg"])
        shell.pack(fill="both", expand=True, padx=24, pady=22)
        shell.columnconfigure(0, weight=1)
        shell.rowconfigure(1, weight=1)
        tk.Label(shell, text=title, font=self._font(14, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).grid(row=0, column=0, sticky="w", pady=(0, 12))
        body = tk.Text(shell, wrap="word", font=self._font(9), bg=self.colors["card_alt"], fg=self.colors["text"], relief="solid", bd=1, padx=12, pady=10)
        body.grid(row=1, column=0, sticky="nsew")
        body.insert("1.0", text)
        body.configure(state="disabled")
        scroll = ttk.Scrollbar(shell, orient="vertical", command=body.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        body.configure(yscrollcommand=scroll.set)
        self._make_button(shell, "Close", top.destroy, kind="neutral", width=10).grid(row=2, column=0, columnspan=2, sticky="e", pady=(12, 0))

    def _set_buttons_state(self, state: str) -> None:
        for widget in self.root.winfo_children():
            self._set_button_state_recursive(widget, state)

    def _set_button_state_recursive(self, widget: tk.Widget, state: str) -> None:
        if isinstance(widget, tk.Button):
            widget.configure(state=state)
        for child in widget.winfo_children():
            self._set_button_state_recursive(child, state)

    def _make_card(self, parent: tk.Widget) -> tk.Frame:
        return tk.Frame(parent, bg=self.colors["card"], highlightbackground=self.colors["card_border"], highlightthickness=1)

    def _make_entry(self, parent: tk.Widget, variable: tk.StringVar) -> tk.Entry:
        return tk.Entry(parent, textvariable=variable, font=self._font(9), bg=self.colors["input_bg"], fg=self.colors["text"], relief="solid", bd=1, highlightbackground=self.colors["input_border"], highlightthickness=1)

    def _make_path_entry(self, parent: tk.Widget, source_var: tk.StringVar, display_var: tk.StringVar, key: str) -> tk.Entry:
        entry = self._make_entry(parent, display_var)

        def focus_in(_event=None) -> None:
            self._path_display_editing.add(key)
            display_var.set(source_var.get())
            self.root.after_idle(lambda: (entry.selection_range(0, "end"), entry.icursor("end")))

        def commit(_event=None) -> None:
            self._commit_path_display(source_var, display_var, key)

        entry.bind("<FocusIn>", focus_in)
        entry.bind("<Return>", commit)
        entry.bind("<FocusOut>", commit)
        return entry

    def _commit_path_display(self, source_var: tk.StringVar, display_var: tk.StringVar, key: str) -> None:
        if key in self._path_display_editing:
            value = display_var.get().strip()
            if value:
                source_var.set(value)
        self._sync_path_display(source_var, display_var, key)

    def _sync_path_display(self, source_var: tk.StringVar, display_var: tk.StringVar, key: str) -> None:
        self._path_display_editing.discard(key)
        display_var.set(self._compact_path(source_var.get()))

    def _compact_path(self, value: str, max_chars: int = 92) -> str:
        text = str(value).strip()
        if len(text) <= max_chars:
            return text
        path = Path(text)
        parts = path.parts
        name = path.name or text[-32:]
        if len(parts) >= 4:
            candidate = str(Path(parts[0], "...", parts[-2], name))
        elif len(parts) >= 2:
            candidate = str(Path(parts[0], "...", name))
        else:
            candidate = "..." + text[-(max_chars - 3) :]
        if len(candidate) <= max_chars:
            return candidate
        return "..." + candidate[-(max_chars - 3) :]

    def _make_toolbar_label(self, parent: tk.Widget, text: str) -> tk.Label:
        return tk.Label(parent, text=text, font=self._font(9, "bold"), bg=self.colors["card"], fg=self.colors["text_secondary"], anchor="w", justify="left", wraplength=max(460, round(700 * self.scale)))

    def _safe_temp_stem(self, value: str) -> str:
        stem = "".join(ch if ch.isalnum() else "_" for ch in value)
        return stem[:90] or "method_preview"

    def _make_button(self, parent: tk.Widget, text: str, command, kind: str = "neutral", width: int | None = None) -> tk.Button:
        palette = {
            "primary": (self.colors["primary"], self.colors["primary_hover"], self.colors["primary_pressed"], "white", 0),
            "secondary": (self.colors["secondary_btn"], self.colors["secondary_btn_hover"], self.colors["secondary_pressed"], "white", 0),
            "neutral": (self.colors["neutral"], self.colors["neutral_hover"], self.colors["neutral_pressed"], self.colors["text"], 1),
        }
        bg, hover, pressed, fg, border = palette[kind]
        btn = tk.Button(parent, text=text, command=command, font=self._font(9, "bold"), bg=bg, fg=fg, activebackground=hover, activeforeground=fg, relief="flat", bd=0, cursor="hand2", padx=16, pady=5, width=width)
        if border:
            btn.configure(highlightbackground="#C9D3DF", highlightthickness=1)
        btn.bind("<Enter>", lambda _e, b=btn, c=hover: b.config(bg=c) if str(b["state"]) == "normal" else None)
        btn.bind("<Leave>", lambda _e, b=btn, c=bg: b.config(bg=c) if str(b["state"]) == "normal" else None)
        btn.bind("<ButtonPress-1>", lambda _e, b=btn, c=pressed: b.config(bg=c) if str(b["state"]) == "normal" else None)
        btn.bind("<ButtonRelease-1>", lambda _e, b=btn, c=hover: b.config(bg=c) if str(b["state"]) == "normal" else None)
        return btn


def main() -> None:
    try:
        try:
            LOCAL_STARTUP_LOG.parent.mkdir(parents=True, exist_ok=True)
            LOCAL_STARTUP_LOG.write_text("", encoding="utf-8")
        except Exception:
            pass
        _write_startup_log("main start")
        _enable_dpi_awareness()
        _write_startup_log("dpi awareness complete")
        root = tk.Tk()
        _write_startup_log("tk root created")
        CmbxExplorerApp(root)
        _write_startup_log("app initialized")
        try:
            root.update_idletasks()
            root.deiconify()
            root.lift()
            root.focus_force()
            _write_startup_log("root shown")
        except Exception as exc:
            _write_startup_log(f"root show warning: {exc}")
        _write_startup_log("entering mainloop")
        root.mainloop()
        _write_startup_log("mainloop exited")
    except Exception:
        _write_startup_log("fatal exception:\n" + traceback.format_exc())
        raise


def _enable_dpi_awareness() -> None:
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def _looks_float(value: str) -> bool:
    try:
        float(str(value).strip())
        return True
    except (TypeError, ValueError):
        return False


if __name__ == "__main__":
    main()
