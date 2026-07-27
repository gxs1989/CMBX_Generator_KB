from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from semantic_generation import load_tcc_semantic_catalog


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_GENERATED_PROJECT_ROOT = PROJECT_ROOT / "cmbx_data_explorer" / "outputs" / "generated_projects"


@dataclass(frozen=True)
class SinglePointTemperatureAccuracyProject:
    family: str
    device_model: str
    test_intent: str
    baseline_c: float | None
    setpoint_c: float
    db_field: str
    report_template: str
    report_sheet: str
    injection_name: str
    instrument_method: str
    processing_method: str
    ret_time: str
    source_method: str
    source_report_template: str
    required_channels: tuple[str, ...]
    required_config: tuple[str, ...]


def build_single_point_temperature_accuracy_project(
    device_model: str,
    setpoint_c: float,
    baseline_c: float | None = 20.0,
    catalog: dict[str, Any] | None = None,
) -> SinglePointTemperatureAccuracyProject:
    catalog = catalog or load_tcc_semantic_catalog()
    device = catalog["device_models"][device_model]
    test = catalog["test_intents"]["temperature_accuracy"]
    binding = test["device_bindings"][device_model]
    db_field = _temperature_accuracy_db_field(setpoint_c)
    ret_time = _temperature_accuracy_ret_time(device_model, setpoint_c)

    return SinglePointTemperatureAccuracyProject(
        family=str(catalog.get("family", "TCC")),
        device_model=device_model,
        test_intent="temperature_accuracy",
        baseline_c=float(baseline_c) if baseline_c is not None else None,
        setpoint_c=float(setpoint_c),
        db_field=db_field,
        report_template=f"{device['report_template']}__single_{_setpoint_token(setpoint_c)}",
        report_sheet="Temp Accuracy",
        injection_name=str(binding["injection_name"]),
        instrument_method=f"{test['instrument_method']}__single_{_setpoint_token(setpoint_c)}",
        processing_method=str(binding["processing_method"]),
        ret_time=ret_time,
        source_method=str(test["instrument_method"]),
        source_report_template=str(device["report_template"]),
        required_channels=tuple(test.get("channels", ())),
        required_config=(
            "ColumnComp.ModelNo = AUDIT.ColumnComp.ModelNo",
            "ColumnComp.CC.TempCtrl = On",
            "ColumnComp.CC.Mode = StillAir",
            "ColumnComp.CC.TempReady available",
            "Thermometer1.ExtTemp_UpperCC configured and acquired",
            "Thermometer1.ExtTemp_LowerCC configured and acquired",
        ),
    )


def single_point_temperature_accuracy_project_to_dict(
    project: SinglePointTemperatureAccuracyProject,
) -> dict[str, Any]:
    return {
        "schema": "cmbx-data-explorer.generated-tcc-project.v1",
        "family": project.family,
        "device_model": project.device_model,
        "test_intent": project.test_intent,
        "baseline_c": project.baseline_c,
        "setpoint_c": project.setpoint_c,
        "db_field": project.db_field,
        "sequence_row": {
            "injection_name": project.injection_name,
            "instrument_method": project.instrument_method,
            "processing_method": project.processing_method,
            "sample_type_default": "Unknown",
            "status_default": "Idle",
        },
        "source_objects": {
            "instrument_method": project.source_method,
            "report_template": project.source_report_template,
        },
        "generated_objects": {
            "instrument_method": project.instrument_method,
            "report_template": project.report_template,
            "report_sheet": project.report_sheet,
        },
        "method_contract": {
            "baseline_temperature_c": project.baseline_c,
            "nominal_temperature_c": project.setpoint_c,
            "evaluation_ret_time": project.ret_time,
            "required_channels": list(project.required_channels),
            "required_config": list(project.required_config),
            "required_report_output": {
                "db_field": project.db_field,
                "report_file": f"{project.injection_name}.XLS",
                "report_sheet": project.report_sheet,
                "report_cell": _temperature_accuracy_report_cell(project.device_model, project.setpoint_c),
                "value_semantics": "deviation at nominal temperature",
                "unit": "K",
            },
        },
        "report_contract": _report_contract(project),
    }


def write_single_point_temperature_accuracy_project(
    project: SinglePointTemperatureAccuracyProject,
    output_root: str | Path | None = None,
) -> Path:
    output_root = Path(output_root or DEFAULT_GENERATED_PROJECT_ROOT)
    project_dir = output_root / _project_dir_name(project)
    project_dir.mkdir(parents=True, exist_ok=True)

    spec = single_point_temperature_accuracy_project_to_dict(project)
    (project_dir / "project_spec.json").write_text(
        json.dumps(spec, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    (project_dir / "instrument_method_draft.txt").write_text(
        instrument_method_draft_text(project),
        encoding="utf-8",
    )
    token = _setpoint_token(project.setpoint_c)
    method_script = instrument_method_script_text(project)
    (project_dir / f"method_script_{token}.txt").write_text(
        method_script,
        encoding="utf-8",
    )
    if token == "40C":
        (project_dir / "method_script_40C_only.txt").write_text(
            method_script,
            encoding="utf-8",
        )
    (project_dir / "required_configuration.md").write_text(
        required_configuration_text(project),
        encoding="utf-8",
    )
    (project_dir / "sequence_template.tsv").write_text(
        sequence_template_tsv(project),
        encoding="utf-8",
    )
    (project_dir / "processing_method_binding.md").write_text(
        processing_method_binding_text(project),
        encoding="utf-8",
    )
    (project_dir / "report_calculation_spec.md").write_text(
        report_calculation_spec_text(project),
        encoding="utf-8",
    )
    (project_dir / "config_method_report_review.md").write_text(
        config_method_report_review_text(project),
        encoding="utf-8",
    )
    (project_dir / f"report_formula_map_{token}.tsv").write_text(
        report_formula_map_tsv(project),
        encoding="utf-8",
    )
    (project_dir / "generation_notes.md").write_text(
        generation_notes_text(project),
        encoding="utf-8",
    )
    return project_dir


def write_single_point_temperature_accuracy_excel_workbooks(
    project: SinglePointTemperatureAccuracyProject,
    output_root: str | Path | None = None,
) -> dict[str, Path]:
    """Write Excel workbooks for the generated method script and report contract.

    These workbooks are intentionally semantic exports: they are designed to be
    reviewed and copied into Chromeleon method/report editors. They are not a
    signed Chromeleon binary payload.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build generated method/report Excel workbooks.") from exc

    output_root = Path(output_root or DEFAULT_GENERATED_PROJECT_ROOT)
    project_dir = output_root / _project_dir_name(project)
    project_dir.mkdir(parents=True, exist_ok=True)

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="E2F0D9")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    mono = Font(name="Consolas", size=10)
    bold = Font(bold=True)

    def style_table(ws, header_row: int = 1) -> None:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row == header_row:
                    cell.font = bold
                    cell.fill = header_fill
        ws.freeze_panes = f"A{header_row + 1}"

    method_wb = Workbook()
    ws = method_wb.active
    ws.title = "Method Script"
    ws.append(("Step", "Stage", "Command Type", "Target / Condition", "Value / Action", "Why It Exists"))
    for row in _single_point_accuracy_method_rows(project):
        ws.append(row)
    for cell in ws["D"] + ws["E"]:
        cell.font = mono
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 3).value in {"SECTION", "COMMENT"}:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = section_fill if ws.cell(row, 3).value == "SECTION" else warning_fill
    style_table(ws)
    widths = {"A": 8, "B": 18, "C": 18, "D": 52, "E": 42, "F": 68}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws = method_wb.create_sheet("RetTime Contract")
    ws.append(("RetTime", "Emitted At", "Used By Report", "Meaning"))
    for row in _single_point_accuracy_rettime_rows(project):
        ws.append(row)
    style_table(ws)
    for column, width in {"A": 16, "B": 44, "C": 24, "D": 70}.items():
        ws.column_dimensions[column].width = width

    ws = method_wb.create_sheet("Configuration")
    ws.append(("Type", "Required Symbol / Setting", "Reason"))
    for item in project.required_config:
        ws.append(("Required", item, "Method cannot run or report cannot calculate without this CM configuration."))
    for channel in project.required_channels:
        ws.append(("Channel", channel, "Acquired raw signal used by report formula."))
    ws.append(("Variable", "StabVars.*", "Upper/lower external thermometer stability state machine."))
    ws.append(("Variable", f"{_baseline_ret_time(project)} and RetTimes.{project.ret_time}", "Baseline marker and requested report anchor."))
    style_table(ws)
    for column, width in {"A": 18, "B": 54, "C": 80}.items():
        ws.column_dimensions[column].width = width

    method_path = project_dir / f"{project.instrument_method}_method_script.xlsx"
    method_wb.save(method_path)

    report_wb = Workbook()
    ws = report_wb.active
    ws.title = "Report Formula Map"
    for line in report_formula_map_tsv(project).splitlines():
        ws.append(line.split("\t"))
    style_table(ws)
    for column, width in {"A": 24, "B": 26, "C": 20, "D": 10, "E": 58, "F": 24, "G": 78, "H": 18, "I": 10}.items():
        ws.column_dimensions[column].width = width

    ws = report_wb.create_sheet("Calculation Contract")
    ws.append(("Item", "Formula / Rule", "Precision", "Source Contract"))
    lower_formula = f'chm.sig_value("average", AUDIT.{project.ret_time}(1,"forward")-1, AUDIT.{project.ret_time}(1,"forward")-0.2)'
    upper_formula = lower_formula
    ws.append(("Lower average", lower_formula, "raw", "FixedChannel = ExtTemp_LowerCC"))
    ws.append(("Upper average", upper_formula, "raw", "FixedChannel = ExtTemp_UpperCC"))
    ws.append(("Observed", f"lower if abs(lower - {project.setpoint_c:g}) >= abs(upper - {project.setpoint_c:g}) else upper", "raw", "Workbook-derived rule"))
    ws.append(("Deviation", f"observed - {project.setpoint_c:g}", "display 2 decimals", project.db_field))
    ws.append(("Pass/Fail", "abs(raw deviation) <= Definitions!Temperature Accuracy", "raw comparison", "Definitions sheet"))
    style_table(ws)
    for column, width in {"A": 22, "B": 88, "C": 22, "D": 44}.items():
        ws.column_dimensions[column].width = width

    ws = report_wb.create_sheet("Single Point Layout")
    ws.append(("Report Sheet", "Cell", "Value / Formula", "Meaning"))
    row = _report_row(project)
    ws.append((project.report_sheet, f"J{row}", project.setpoint_c, f"Nominal {project.setpoint_c:g} C row for accuracy"))
    ws.append((project.report_sheet, f"K{row}", f"AUDIT.{project.ret_time}(1,\"forward\")", f"{project.setpoint_c:g} C stable RetTime emitted by method"))
    ws.append((project.report_sheet, f"L{row}", lower_formula, "Lower external thermometer average"))
    ws.append((project.report_sheet, f"M{row}", upper_formula, "Upper external thermometer average"))
    ws.append((project.report_sheet, f"C{row}", "Observed value", "External sensor with larger absolute deviation"))
    ws.append((project.report_sheet, f"D{row}", f"Observed - {project.setpoint_c:g} C", f"DB field {project.db_field}"))
    ws.append((project.report_sheet, f"E{row}", f"Test passed if abs(D{row} raw) <= threshold", "Pass/fail"))
    style_table(ws)
    for column, width in {"A": 22, "B": 12, "C": 86, "D": 58}.items():
        ws.column_dimensions[column].width = width

    report_path = project_dir / f"{project.report_template}_report_calculation.xlsx"
    report_wb.save(report_path)
    return {"method_excel": method_path, "report_excel": report_path}


def instrument_method_draft_text(project: SinglePointTemperatureAccuracyProject) -> str:
    lines = [
        f"# {project.instrument_method}",
        "",
        "Purpose: single-point VH TCC temperature accuracy test.",
        f"Source method: {project.source_method}",
        f"Device model: {project.device_model}",
        f"Baseline temperature: {_baseline_label(project)}",
        f"Nominal temperature: {project.setpoint_c:g} degC",
        f"Report/DB anchor: {project.db_field} via {project.ret_time}",
        "",
        "Required configuration:",
        *[f"- {item}" for item in project.required_config],
        "",
        "Required acquisition channels:",
        *[f"- {channel}" for channel in project.required_channels],
        "",
        "Instrument method draft:",
        "1. Enter Equilibration stage.",
        "2. Set ColumnComp.CC.ReadyTempDelta = 1.0 degC.",
        "3. Set ColumnComp.CC.EquilibrationTime = 0.5 min.",
        "4. Set ColumnComp.CC.TempCtrl = On.",
        "5. For VH-C10-A, run ColumnComp.CmdString Cmd=\"PCC.TempCtrl=0\".",
        "6. Set ColumnComp.CC.Mode = StillAir.",
        f"7. Set baseline variable = {_baseline_label(project)}.",
        f"8. Set Variables.GenericDouble3 = {project.setpoint_c:g} for the report point.",
        "9. Enter InjectPreparation stage.",
        "10. Initialize RetTimes.RetTime1 through RetTimes.RetTime5 to 0.",
        "11. Set ColumnComp.CC.ReadyTempDelta = 0.2 degC.",
        "12. Set ColumnComp.CC.EquilibrationTime = 3 min.",
        "13. Start stability triggers for upper/lower external thermometers.",
        "14. Enter StartRun stage.",
        "15. Acquire ColumnComp.CC_Temp.",
        "16. Acquire Thermometer1.ExtTemp_UpperCC.",
        "17. Acquire Thermometer1.ExtTemp_LowerCC.",
        "18. Enter Run stage.",
        f"19. Set ColumnComp.CC.Temperature.Nominal = {_baseline_label(project)}.",
        f"20. Wait for CC and upper/lower external thermometer stability at {_baseline_label(project)}.",
        f"21. Set {_baseline_ret_time(project)} = System.Retention as the baseline marker if this design needs it.",
        f"22. Set ColumnComp.CC.Temperature.Nominal = {project.setpoint_c:g} degC.",
        f"23. Wait for CC and upper/lower external thermometer stability at {project.setpoint_c:g} C.",
        f"24. Set RetTimes.{project.ret_time} = System.Retention after the stable averaging window exists.",
        f"25. Return ColumnComp.CC.Temperature.Nominal = {_safe_return_temperature(project)} degC.",
        "26. Turn acquisitions off and stop run.",
        "",
        "Binary status:",
        "- This is a generation contract and method draft, not a Chromeleon-signed binary method payload.",
        "- The source binary payload is TEMPERATURE_ACCURACY; a binary encoder/patcher must preserve CpXm structure before direct CMBX creation.",
    ]
    return "\n".join(lines) + "\n"


def instrument_method_script_text(project: SinglePointTemperatureAccuracyProject) -> str:
    lines = [
        f"# {project.instrument_method}",
        "# Intended Chromeleon method script logic, reduced from TEMPERATURE_ACCURACY.",
        "# This is a single-setpoint design candidate. It is not the unchanged multi-point source method.",
        "# Baseline/approach temperature is a design input, not something the generator should guess.",
        "",
        "[InstrumentSetup]",
        "COMMENT HPLC-System / TCC standalone temperature accuracy.",
        "COMMENT Required external thermometer channels must already exist in the CM instrument configuration.",
        "",
        "[Equilibration]",
        "SET ColumnComp.CC.ReadyTempDelta = 1.0",
        "SET ColumnComp.CC.EquilibrationTime = 0.5",
        "SET ColumnComp.CC.TempCtrl = On",
        "RUN ColumnComp.CmdString Cmd=\"PCC.TempCtrl=0\"  # VH only: disable PCC temp control",
        "SET ColumnComp.CC.Mode = StillAir",
        "SET ColumnComp.LeakSensorMode = Off",
        f"SET Variables.GenericDouble2 = {_baseline_value_text(project)}",
        f"SET Variables.GenericDouble3 = {project.setpoint_c:g}",
        f"SET ColumnComp.CC.Temperature.Nominal = {_baseline_value_text(project)}",
        "RUN Delay 5",
        "",
        "[InjectPreparation]",
        "RUN Wait CC.TempReady",
        "SET RetTimes.RetTime1 = 0",
        "SET RetTimes.RetTime2 = 0",
        "SET RetTimes.RetTime3 = 0",
        "SET RetTimes.RetTime4 = 0",
        "SET RetTimes.RetTime5 = 0",
        "SET StabVars.TriggerStab1 = 0",
        "SET StabVars.TriggerStab2 = 0",
        "SET StabVars.TempUpperHigh = 0",
        "SET StabVars.TempUpperLow = 0",
        "SET StabVars.TempLowerHigh = 0",
        "SET StabVars.TempLowerLow = 0",
        "SET StabVars.CounterUpper = 0",
        "SET StabVars.CounterLower = 0",
        "SET StabVars.UpperReady = 0",
        "SET StabVars.LowerReady = 0",
        "SET ColumnComp.CC.ReadyTempDelta = 0.2",
        "SET ColumnComp.CC.EquilibrationTime = 3",
        "RUN System.Trigger \"Gradient_1\", (StabVars.TriggerStab1=1) AND CC.TempReady, TrueTime=30, Delay=0",
        "SET StabVars.TriggerStab1 = 0",
        "SET StabVars.TriggerStab2 = 1",
        "IF (StabVars.TempUpperHigh<>0) AND ((Thermometer1.ExtTemp_UpperCC<=StabVars.TempUpperHigh) AND (Thermometer1.ExtTemp_UpperCC>=StabVars.TempUpperLow))",
        "  SET StabVars.CounterUpper = StabVars.CounterUpper+1",
        "ELSE",
        "  SET StabVars.CounterUpper = 0",
        "  SET StabVars.TempUpperHigh = Thermometer1.ExtTemp_UpperCC+0.05",
        "  SET StabVars.TempUpperLow = Thermometer1.ExtTemp_UpperCC-0.05",
        "END IF",
        "IF (StabVars.TempLowerHigh<>0) AND ((Thermometer1.ExtTemp_LowerCC<=StabVars.TempLowerHigh) AND (Thermometer1.ExtTemp_LowerCC>=StabVars.TempLowerLow))",
        "  SET StabVars.CounterLower = StabVars.CounterLower+1",
        "ELSE",
        "  SET StabVars.CounterLower = 0",
        "  SET StabVars.TempLowerHigh = Thermometer1.ExtTemp_LowerCC+0.05",
        "  SET StabVars.TempLowerLow = Thermometer1.ExtTemp_LowerCC-0.05",
        "END IF",
        "IF StabVars.CounterUpper>=4",
        "  SET StabVars.UpperReady = 1",
        "ELSE",
        "  SET StabVars.UpperReady = 0",
        "END IF",
        "IF StabVars.CounterLower>=4",
        "  SET StabVars.LowerReady = 1",
        "ELSE",
        "  SET StabVars.LowerReady = 0",
        "END IF",
        "RUN System.Trigger \"Gradient_2\", StabVars.TriggerStab2=1, TrueTime=30, Delay=0",
        "SET StabVars.TriggerStab1 = 1",
        "SET StabVars.TriggerStab2 = 0",
        "RUN System.Trigger \"ExitRange_Upper\", (StabVars.TempUpperHigh<>0) AND ((ColumnComp.CC.TempReady=0) OR (Thermometer1.ExtTemp_UpperCC>StabVars.TempUpperHigh) OR (Thermometer1.ExtTemp_UpperCC<StabVars.TempUpperLow)), TrueTime=5",
        "SET StabVars.TempUpperHigh = 0",
        "SET StabVars.TempUpperLow = 0",
        "SET StabVars.UpperReady = 0",
        "RUN System.Trigger \"ExitRange_Lower\", (StabVars.TempLowerHigh<>0) AND ((ColumnComp.CC.TempReady=0) OR (Thermometer1.ExtTemp_LowerCC>StabVars.TempLowerHigh) OR (Thermometer1.ExtTemp_LowerCC<StabVars.TempLowerLow)), TrueTime=5",
        "SET StabVars.TempLowerHigh = 0",
        "SET StabVars.TempLowerLow = 0",
        "SET StabVars.LowerReady = 0",
        f"RUN System.Trigger \"Abort\", (System.Retention>40 AND RetTimes.{project.ret_time}=0), TrueTime=0, AllowImmediateExecution=Yes",
        "RUN ColumnComp.CmdString Cmd=\"LedBar.ForceColor=1\"",
        "RUN Message \"QUEUE WAS ABORTED! Please check setup and repeat temperature accuracy test.\"",
        "RUN ColumnComp.CmdString Cmd=\"LedBar.ForceColor=0\"",
        "RUN System.AbortQueue",
        "",
        "[StartRun]",
        "RUN ColumnComp.CC_Temp.AcqOn",
        "RUN Thermometer1.ExtTemp_UpperCC.AcqOn",
        "RUN Thermometer1.ExtTemp_LowerCC.AcqOn",
        "",
        "[Run]",
        "RUN Delay 5",
        f"SET ColumnComp.CC.Temperature.Nominal = {_baseline_value_text(project)}",
        "SET StabVars.TriggerStab1 = 1",
        "RUN Delay 3",
        f"# First stabilize at {_baseline_label(project)}. This baseline must be confirmed from TD/method design.",
        "RUN Wait ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady, Run=Continue",
        "RUN Delay 60",
        "RUN Wait ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady, Run=Continue",
        f"SET {_baseline_ret_time(project)} = System.Retention",
        "RUN Delay 2",
        f"SET ColumnComp.CC.Temperature.Nominal = {project.setpoint_c:g}",
        "SET StabVars.UpperReady = 0",
        "SET StabVars.LowerReady = 0",
        "SET StabVars.TempUpperHigh = 0",
        "SET StabVars.TempUpperLow = 0",
        "SET StabVars.TempLowerHigh = 0",
        "SET StabVars.TempLowerLow = 0",
        "RUN Delay 3",
        "# Wait until the controller is ready and both external thermometer channels are stable.",
        "RUN Wait ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady, Run=Continue",
        f"# Collect a settled external-temperature window before writing {project.ret_time}.",
        f"# The report reads {project.ret_time} - 1.0 min to {project.ret_time} - 0.2 min.",
        "RUN Delay 60",
        "RUN Wait ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady, Run=Continue",
        f"SET RetTimes.{project.ret_time} = System.Retention",
        "RUN Delay 2",
        f"SET ColumnComp.CC.Temperature.Nominal = {_safe_return_temperature(project)}",
        "SET StabVars.TriggerStab1 = 0",
        "SET StabVars.TriggerStab2 = 0",
        "RUN Delay 2",
        "RUN ColumnComp.CC_Temp.AcqOff",
        "RUN Thermometer1.ExtTemp_UpperCC.AcqOff",
        "RUN Thermometer1.ExtTemp_LowerCC.AcqOff",
        "",
        "# Removed from the source multi-point method:",
        "# - Non-requested production setpoint transitions.",
        f"# - {_removed_accuracy_setpoints_label(project)} transitions for this {_setpoint_token(project.setpoint_c)}-only draft.",
        "# - Non-requested RetTime emissions.",
        f"# - {_baseline_ret_time(project)} is kept as an explicit baseline marker only if this design needs it.",
        "# - Ambient skip branch for the 10 degC point.",
        "# Preserved intentionally:",
        f"# - RetTimes.{project.ret_time} as the {project.setpoint_c:g} degC report/DB anchor.",
        "# - Upper/lower external thermometer stability trigger logic.",
    ]
    return "\n".join(lines) + "\n"


def required_configuration_text(project: SinglePointTemperatureAccuracyProject) -> str:
    return "\n".join(
        [
            f"# Required Configuration - {project.instrument_method}",
            "",
            f"This generated method assumes a {project.device_model} TCC CM instrument configuration with these symbols available.",
            "",
            "Required module identity:",
            "",
            f"- `AUDIT.ColumnComp.ModelNo` must resolve to `{project.device_model}`.",
            "- `ColumnComp.CC` must exist.",
            "- `ColumnComp.CC.TempReady` must exist.",
            "- `ColumnComp.CC.Temperature.Nominal` must be writable.",
            "",
            "Required external thermometer configuration:",
            "",
            "- `Thermometer1.ExtTemp_UpperCC` must exist and be acquirable.",
            "- `Thermometer1.ExtTemp_LowerCC` must exist and be acquirable.",
            "- These are CM instrument configuration channels, not values created by the report.",
            "",
            "Required method variables:",
            "",
            "- `RetTimes.RetTime1` through `RetTimes.RetTime5`",
            "- `StabVars.TriggerStab1`, `StabVars.TriggerStab2`",
            "- `StabVars.TempUpperHigh`, `StabVars.TempUpperLow`",
            "- `StabVars.TempLowerHigh`, `StabVars.TempLowerLow`",
            "- `StabVars.CounterUpper`, `StabVars.CounterLower`",
            "- `StabVars.UpperReady`, `StabVars.LowerReady`",
            "- `Variables.GenericDouble3`",
            "",
            "Optional diagnostic channels removed from the minimal script:",
            "",
            "- `ColumnComp.CC_U_Temp_Actual`",
            "- `ColumnComp.CC_L_Temp_Actual`",
            "- `ColumnComp.CC_UCTL_TempRear_Actual`",
            "- `ColumnComp.PWM_*`",
            "- `ColumnComp.Fan_Rear_ActualRPM`",
            "- leak-board diagnostic channels",
            "",
            "If CM requires these diagnostics for this instrument configuration, they can be re-added without changing the report calculation.",
        ]
    ) + "\n"


def sequence_template_tsv(project: SinglePointTemperatureAccuracyProject) -> str:
    rows = [
        (
            "Row",
            "Injection Name",
            "Type",
            "Status",
            "Instrument Method",
            "Processing Method",
            "Sample ID",
            "Replicate",
            "Report Template",
            "Report Sheet",
            "DB Field",
            "Review Note",
        ),
        (
            "1",
            project.injection_name,
            "Unknown",
            "Idle",
            project.instrument_method,
            project.processing_method,
            "Internal",
            "1",
            project.report_template,
            project.report_sheet,
            project.db_field,
            "Review-only sequence row; recreate in Chromeleon after method/report assets are validated.",
        ),
    ]
    return "\n".join("\t".join(row) for row in rows) + "\n"


def processing_method_binding_text(project: SinglePointTemperatureAccuracyProject) -> str:
    return "\n".join(
        [
            f"# Processing Method Binding - {project.processing_method}",
            "",
            "This file records the sequence-level processing binding for the generated draft asset packet.",
            "",
            "| Field | Value |",
            "|---|---|",
            f"| Device | `{project.device_model}` |",
            f"| Injection | `{project.injection_name}` |",
            f"| Generated instrument method | `{project.instrument_method}` |",
            f"| Processing method | `{project.processing_method}` |",
            f"| Source processing binding | `{project.processing_method}` |",
            f"| Report template draft | `{project.report_template}` |",
            f"| DB field | `{project.db_field}` |",
            "",
            "## Boundary",
            "",
            "The draft packet reuses the known processing method binding. It does not decode or rewrite the processing method payload.",
            "For `ACCURACY_IRC_STOP_H` / `ACCURACY_IRC_STOP_C`, the pass-action and stop behavior remain open-verification items until the processing method payload is fully decoded or confirmed in CM.",
            "",
            "## Review Checklist",
            "",
            "- Confirm whether the source processing method is appropriate for a single-point cropped accuracy injection.",
            "- Confirm whether IRC should be disabled, preserved, or redesigned for the generated sequence row.",
            "- Confirm any stop behavior in Chromeleon before using this draft in a live FOQ sequence.",
        ]
    ) + "\n"


def report_calculation_spec_text(project: SinglePointTemperatureAccuracyProject) -> str:
    lower = _channel_leaf(project.required_channels, "LowerCC")
    upper = _channel_leaf(project.required_channels, "UpperCC")
    return "\n".join(
        [
            f"# {project.report_template}",
            "",
            f"Source report template: {project.source_report_template}",
            f"Source sheet: {project.report_sheet}",
            f"Single-point DB field: {project.db_field}",
            f"Baseline: {_baseline_label(project)}",
            f"Nominal/report point: {project.setpoint_c:g} degC",
            f"Evaluation RetTime: AUDIT.{project.ret_time}(1,\"forward\")",
            "",
            "SheetObject formulas:",
            "",
            f"- Lower external thermometer ({lower}):",
            f"  `chm.sig_value(\"average\", AUDIT.{project.ret_time}(1,\"forward\")-1, AUDIT.{project.ret_time}(1,\"forward\")-0.2)`",
            f"- Upper external thermometer ({upper}):",
            f"  `chm.sig_value(\"average\", AUDIT.{project.ret_time}(1,\"forward\")-1, AUDIT.{project.ret_time}(1,\"forward\")-0.2)`",
            "",
            "Workbook-derived rule:",
            "",
            f"- `Observed = lower if abs(lower - {project.setpoint_c:g}) >= abs(upper - {project.setpoint_c:g}) else upper`",
            f"- `Deviation = Observed - {project.setpoint_c:g}`",
            "- Display observed/deviation to 2 decimals.",
            "- Pass/fail compares raw absolute deviation with Definitions!Temperature Accuracy.",
            f"- FOQ DB mapping: `{project.db_field}` -> `{project.injection_name}.XLS` / `{project.report_sheet}` / `{_temperature_accuracy_report_cell(project.device_model, project.setpoint_c)}`.",
            "",
            "Compatibility note:",
            f"- The requested setpoint maps to {project.ret_time} for {project.device_model}, so the report formulas preserve that row anchor.",
            "- A full multi-point report can leave other temperature rows blank; this single-point report should expose only the requested row/DB field.",
        ]
    ) + "\n"


def config_method_report_review_text(project: SinglePointTemperatureAccuracyProject) -> str:
    """Render the minimal black-box review for an intent-level method edit.

    This is deliberately narrower than the full draft packet. It answers the
    question the user actually needs before generation: does the selected edit
    preserve instrument configuration, method command semantics, and report
    formula semantics? Processing and DB are kept as downstream checks.
    """
    report_cell = _temperature_accuracy_report_cell(project.device_model, project.setpoint_c)
    lower_formula = (
        f'chm.sig_value("average", AUDIT.{project.ret_time}(1,"forward")-1, '
        f'AUDIT.{project.ret_time}(1,"forward")-0.2)'
    )
    removed_setpoints = _removed_accuracy_setpoints_label(project)
    lines = [
        f"# Config -> Method -> Report Review - {project.device_model} {project.setpoint_c:g} C Accuracy",
        "",
        "## Review Boundary",
        "",
        "This is an intent-level review for a single Temperature Accuracy setpoint crop. It is not a runnable CMBX claim.",
        "The review intentionally prioritizes instrument configuration, instrument method command semantics, and report formula semantics.",
        "Processing Method and DB upload remain downstream checks; they must not be used as the primary proof that the method is correct.",
        "",
        "## Intent",
        "",
        "| Item | Value |",
        "|---|---|",
        f"| Device | `{project.device_model}` |",
        "| Test | `Temperature Accuracy` |",
        f"| Requested setpoint | `{project.setpoint_c:g} C` |",
        f"| Source method | `{project.source_method}` |",
        f"| Draft method | `{project.instrument_method}` |",
        f"| Source report template | `{project.source_report_template}` |",
        f"| Draft report template | `{project.report_template}` |",
        f"| Report output | `{project.db_field}` / `{project.report_sheet}!{report_cell}` |",
        f"| Report RetTime anchor | `AUDIT.{project.ret_time}(1,\"forward\")` |",
        "",
        "## 1. Instrument Config Manifest",
        "",
        "The method and report require these CM-side symbols before any generated or manually copied method can run:",
        "",
        "| Type | Required evidence | Why it matters | Status for this draft |",
        "|---|---|---|---|",
    ]
    for item in project.required_config:
        lines.append(f"| Config | `{_md_escape(item)}` | Required by method command or model/report binding | Required, not created by the draft |")
    for channel in project.required_channels:
        lines.append(f"| Channel | `{_md_escape(channel)}` | Raw external thermometer data used by report formulas | Required, must exist in CM instrument config |")
    lines.extend(
        [
            "| Variable | `RetTimes.RetTime1..RetTime5` | Audit anchors consumed by report formulas | Required, emitted by method script |",
            "| Variable | `StabVars.*` | External thermometer stability state machine | Required, emitted/updated by method script |",
            "",
            "Open config question:",
            "",
            f"- The draft uses `{_baseline_label(project)}` as the approach/baseline. This is a design input and must be confirmed from FOQ TD or user intent; it is not proven just because the generator has a default.",
            "",
            "## 2. Instrument Method Script Delta",
            "",
            "The draft keeps only the command semantics needed to reach the selected stable setpoint and emit the retained report anchor.",
            "",
            "| Command group | Preserved / changed behavior | Evidence status |",
            "|---|---|---|",
            f"| Temperature approach | Stabilize at `{_baseline_label(project)}` before moving to `{project.setpoint_c:g} C` | Drafted; baseline requires human confirmation |",
            f"| Target setpoint | `SET ColumnComp.CC.Temperature.Nominal = {project.setpoint_c:g}` | Drafted from retained Accuracy setpoint |",
            "| Stability wait | Wait for `ColumnComp.CC.TempReady` and both external thermometer stability flags | Preserved from source method semantics |",
            f"| RetTime anchor | `SET RetTimes.{project.ret_time} = System.Retention` after the stable window exists | Preserved because report formulas use `{project.ret_time}` |",
            "| Acquisition | Acquire upper/lower external thermometer raw channels | Preserved because report formulas use those raw channels |",
            f"| Removed transitions | `{removed_setpoints}` | Review required; removed RetTimes must not be referenced by the single-point report |",
            "| Return state | Return CC nominal temperature to safe/default value | Drafted; CM operating convention should be confirmed |",
            "",
            "Runnable boundary:",
            "",
            "- The text script is a semantic draft. It is not a verified Chromeleon binary payload.",
            "- Direct CMBX generation still needs a safe encoder/patcher for the original method payload structure.",
            "",
            "## 3. Report Formula Chain",
            "",
            "The report remains valid only if the report template is narrowed to the retained RetTime/cell chain.",
            "",
            "| Output | Formula / rule | Source | Status |",
            "|---|---|---|---|",
            f"| Lower average | `{lower_formula}` | FixedChannel = `ExtTemp_LowerCC` | Preserved |",
            f"| Upper average | `{lower_formula}` | FixedChannel = `ExtTemp_UpperCC` | Preserved |",
            f"| Observed | sensor with larger `abs(sensor - {project.setpoint_c:g})` | workbook-derived rule | Must be retained/rebuilt |",
            f"| Deviation | `Observed - {project.setpoint_c:g}` | `{project.db_field}` / `{project.report_sheet}!{report_cell}` | Must display to 2 decimals |",
            "| Pass/fail | `abs(raw deviation) <= Definitions!Temperature Accuracy` | Definitions sheet | Must compare raw value, not rounded display |",
            "",
            "Report crop rule:",
            "",
            f"- Keep the `{project.setpoint_c:g} C` row anchored on `{project.ret_time}`.",
            "- Remove, hide, or mark not-applicable the other TempAcc fields.",
            "- Rebuild `RES_TempAccuracy` so it evaluates only retained setpoints.",
            "",
            "## 4. Minimal Verdict",
            "",
            "| Question | Current answer |",
            "|---|---|",
            "| Can the intent be reviewed? | Yes. The three core contracts are explicit enough to review. |",
            "| Can the method be manually recreated in CM? | Possibly, if the config manifest exists and the baseline/approach rule is confirmed. |",
            "| Can this packet be treated as runnable CMBX? | No. Binary method/report payload generation is not proven. |",
            "| What is the next best evidence to close? | Confirm baseline/approach rule and verify report FormulaOne narrowing for the single retained setpoint. |",
        ]
    )
    return "\n".join(lines) + "\n"


def report_formula_map_tsv(project: SinglePointTemperatureAccuracyProject) -> str:
    formula = (
        f'chm.sig_value("average", AUDIT.{project.ret_time}(1,"forward")-1,'
        f' AUDIT.{project.ret_time}(1,"forward")-0.2)'
    )
    rows = [
        (
            "DBField",
            "ReportFile",
            "Sheet",
            "Cell",
            "Meaning",
            "FixedChannel",
            "FormulaOrRule",
            "Display",
            "Unit",
        ),
        (
            project.db_field,
            f"{project.injection_name}.XLS",
            project.report_sheet,
            _temperature_accuracy_report_cell(project.device_model, project.setpoint_c),
            f"Deviation at {project.setpoint_c:g} degC",
            "",
            f"Observed - {project.setpoint_c:g}",
            "2 decimals",
            "K",
        ),
        (
            "_source_lower_average",
            f"{project.injection_name}.XLS",
            project.report_sheet,
            f"L{_report_row(project)}",
            f"Lower external thermometer average around {project.setpoint_c:g} degC stable RetTime",
            "ExtTemp_LowerCC",
            formula,
            "raw numeric",
            "degC",
        ),
        (
            "_source_upper_average",
            f"{project.injection_name}.XLS",
            project.report_sheet,
            f"M{_report_row(project)}",
            f"Upper external thermometer average around {project.setpoint_c:g} degC stable RetTime",
            "ExtTemp_UpperCC",
            formula,
            "raw numeric",
            "degC",
        ),
        (
            "_source_nominal",
            f"{project.injection_name}.XLS",
            project.report_sheet,
            f"J{_report_row(project)}",
            f"Nominal temperature for the {project.setpoint_c:g} degC row",
            "",
            f"{project.setpoint_c:g}",
            "0 decimals",
            "degC",
        ),
        (
            "_source_ret_time",
            f"{project.injection_name}.XLS",
            project.report_sheet,
            f"K{_report_row(project)}",
            "Stable-time marker emitted by the method",
            "",
            f"AUDIT.{project.ret_time}(1,\"forward\")",
            "raw numeric",
            "min",
        ),
    ]
    return "\n".join("\t".join(row) for row in rows) + "\n"


def generation_notes_text(project: SinglePointTemperatureAccuracyProject) -> str:
    return "\n".join(
        [
            f"# Generation Notes - {project.device_model} {project.db_field}",
            "",
            "This project is the first concrete TCC reverse-generation target.",
            "",
            f"Why {project.ret_time}:",
            f"- The requested setpoint {project.setpoint_c:g} degC maps to {project.ret_time} for {project.device_model}.",
            f"- The method writes RetTimes.{project.ret_time} after waiting for the requested stable condition.",
            f"- FOQ DB field {project.db_field} should therefore read the report row/formula anchored on {project.ret_time}.",
            "",
            "Next encoder work:",
            "- Clone the source TEMPERATURE_ACCURACY CpXm payload.",
            "- Remove or disable the non-requested temperature transitions while keeping required setup/acquisition symbols.",
            f"- Clone the source report template into a single-point report template or preserve the original template and only export {project.db_field}.",
            "- Build a sequence with one row using the generated instrument method and existing processing method.",
        ]
    ) + "\n"


def _report_contract(project: SinglePointTemperatureAccuracyProject) -> dict[str, Any]:
    formula = (
        f'chm.sig_value("average", AUDIT.{project.ret_time}(1,"forward")-1, '
        f'AUDIT.{project.ret_time}(1,"forward")-0.2)'
    )
    return {
        "sheet": project.report_sheet,
        "db_field": project.db_field,
        "nominal_temperature_c": project.setpoint_c,
        "evaluation_ret_time": project.ret_time,
        "source_formulas": {
            "lower_external_temperature": formula,
            "upper_external_temperature": formula,
        },
        "derived_values": {
            "observed": "external sensor value with larger absolute deviation from nominal",
            "deviation": "observed - nominal",
            "display_decimals": 2,
            "pass_fail_threshold": "Definitions!Temperature Accuracy",
        },
    }


def _md_escape(text: str) -> str:
    return str(text).replace("|", "\\|")


def _baseline_value_text(project: SinglePointTemperatureAccuracyProject) -> str:
    if project.baseline_c is None:
        return "<designer-confirmed baseline C>"
    return f"{project.baseline_c:g}"


def _baseline_label(project: SinglePointTemperatureAccuracyProject) -> str:
    if project.baseline_c is None:
        return "designer-confirmed baseline"
    return f"{project.baseline_c:g} C"


def _baseline_ret_time(project: SinglePointTemperatureAccuracyProject) -> str:
    if project.baseline_c is None:
        return "RetTimes.<BaselineRetTime>"
    try:
        return f"RetTimes.{_temperature_accuracy_ret_time(project.device_model, project.baseline_c)}"
    except Exception:
        return "RetTimes.<BaselineRetTime>"


def _safe_return_temperature(project: SinglePointTemperatureAccuracyProject) -> str:
    if project.baseline_c is None:
        return "20.0"
    return f"{project.baseline_c:g}"


def _report_row(project: SinglePointTemperatureAccuracyProject) -> int:
    return 65 + int(project.ret_time.replace("RetTime", ""))


def _single_point_accuracy_method_rows(project: SinglePointTemperatureAccuracyProject) -> list[tuple[object, ...]]:
    return [
        (1, "Overview", "COMMENT", "", f"{project.device_model} temperature accuracy, {_baseline_label(project)} baseline -> {project.setpoint_c:g} C report point", "Generated from semantic test intent, not copied as the full production method."),
        (2, "Setup", "SET", "ColumnComp.CC.TempCtrl", "On", "Enable CC control."),
        (3, "Setup", "SET", "ColumnComp.CC.Mode", "StillAir", "Match production accuracy method measurement mode."),
        (4, "Setup", "RUN", 'ColumnComp.CmdString Cmd="PCC.TempCtrl=0"', "VH only", "Disable PCC temperature control so it does not perturb accuracy measurement."),
        (5, "Setup", "SET", "ColumnComp.CC.ReadyTempDelta", "0.2 C", "Tighter readiness before external thermometer stability check."),
        (6, "Setup", "SET", "ColumnComp.CC.EquilibrationTime", "3 min", "Preserve accuracy-method stabilization contract."),
        (7, "Setup", "SET", "RetTimes.RetTime1..RetTime5", "0", "Initialize report/audit markers."),
        (8, "Acquisition", "RUN", "ColumnComp.CC_Temp.AcqOn", "", "Provide CC raw signal context."),
        (9, "Acquisition", "RUN", "Thermometer1.ExtTemp_UpperCC.AcqOn", "", "Report upper external thermometer source."),
        (10, "Acquisition", "RUN", "Thermometer1.ExtTemp_LowerCC.AcqOn", "", "Report lower external thermometer source."),
        (11, "Baseline", "SET", "ColumnComp.CC.Temperature.Nominal", _baseline_label(project), "Start from the designer-confirmed baseline before moving to the requested setpoint."),
        (12, "Baseline", "WAIT", "ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady", "Continue", "Internal CC and both external thermometers stable."),
        (13, "Baseline", "DELAY", "System.Retention", "60 s", "Ensure the report-style averaging window is meaningful if baseline is audited."),
        (14, "Baseline", "SET", _baseline_ret_time(project), "System.Retention", "Optional baseline marker; not the requested DB field."),
        (15, "Accuracy Point", "SET", "ColumnComp.CC.Temperature.Nominal", f"{project.setpoint_c:g} C", "Move from baseline to requested accuracy setpoint."),
        (16, "Accuracy Point", "RESET", "StabVars upper/lower ready/range fields", "0", "Force a new stability decision after the setpoint change."),
        (17, "Accuracy Point", "WAIT", "ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady", "Continue", f"Stable {project.setpoint_c:g} C condition."),
        (18, "Accuracy Point", "DELAY", "System.Retention", "60 s", f"The report reads {project.ret_time} -1.0 to -0.2 min."),
        (19, "Accuracy Point", "WAIT", "ColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady", "Continue", "Confirm stability immediately before writing RetTime."),
        (20, "Accuracy Point", "SET", f"RetTimes.{project.ret_time}", "System.Retention", f"Report anchor for {project.db_field} row {_temperature_accuracy_report_cell(project.device_model, project.setpoint_c)}."),
        (21, "Cleanup", "SET", "ColumnComp.CC.Temperature.Nominal", f"{_safe_return_temperature(project)} C", "Return to selected safe/default temperature."),
        (22, "Cleanup", "RUN", "ExtTemp/CC acquisitions AcqOff", "", "Stop raw data collection."),
    ]


def _single_point_accuracy_rettime_rows(project: SinglePointTemperatureAccuracyProject) -> list[tuple[str, str, str, str]]:
    ret_time_notes = _accuracy_ladder_ret_time_notes(project)
    return [
        ("RetTime1", "Not emitted unless this setpoint/baseline design needs it", "No", ret_time_notes.get("RetTime1", "Production multi-point method used it for one of the ladder points.")),
        (_baseline_ret_time(project), f"After baseline {_baseline_label(project)} is stable", "Optional audit context", "Confirms the method started from the chosen controlled baseline."),
        (project.ret_time, f"After {project.setpoint_c:g} C is stable and averaging window exists", "Yes", f"{project.db_field} report formulas read {project.ret_time}-1.0 to {project.ret_time}-0.2 min."),
        ("RetTime4", "Not emitted in reduced method", "No", ret_time_notes.get("RetTime4", "Production multi-point method used it for one of the ladder points.")),
        ("RetTime5", "Not emitted in reduced method", "No", ret_time_notes.get("RetTime5", "Production multi-point method used it for one of the ladder points.")),
    ]


def _temperature_accuracy_db_field(setpoint_c: float) -> str:
    value = int(round(setpoint_c))
    return f"TempAcc{value:02d}"


def _temperature_accuracy_ret_time(device_model: str, setpoint_c: float) -> str:
    setpoints = _temperature_accuracy_ladder(device_model)
    if not setpoints:
        raise KeyError(f"Unknown TCC device model: {device_model}")
    normalized = float(setpoint_c)
    for index, candidate in enumerate(setpoints, start=1):
        if abs(candidate - normalized) < 1e-9:
            return f"RetTime{index}"
    raise ValueError(f"{device_model} temperature accuracy setpoint is not in the known FOQ sequence: {setpoint_c:g}")


def _temperature_accuracy_ladder(device_model: str) -> tuple[float, ...]:
    setpoints_by_device = {
        "VA-C10-A": (10.0, 20.0, 40.0, 60.0, 85.0),
        "VC-C10-A": (10.0, 20.0, 40.0, 60.0, 85.0),
        "VH-C10-A": (10.0, 20.0, 40.0, 80.0, 120.0),
    }
    return setpoints_by_device.get(device_model, ())


def _removed_accuracy_setpoints_label(project: SinglePointTemperatureAccuracyProject) -> str:
    removed = [
        f"{setpoint:g}"
        for setpoint in _temperature_accuracy_ladder(project.device_model)
        if abs(setpoint - project.setpoint_c) >= 1e-9
    ]
    return "/".join(removed) + " degC" if removed else "no production setpoint"


def _accuracy_ladder_ret_time_notes(project: SinglePointTemperatureAccuracyProject) -> dict[str, str]:
    return {
        f"RetTime{index}": f"Production {project.device_model} method used it for the {setpoint:g} C point."
        for index, setpoint in enumerate(_temperature_accuracy_ladder(project.device_model), start=1)
    }


def _temperature_accuracy_report_cell(device_model: str, setpoint_c: float) -> str:
    ret_time = _temperature_accuracy_ret_time(device_model, setpoint_c)
    row = 65 + int(ret_time.replace("RetTime", ""))
    return f"D{row}"


def _project_dir_name(project: SinglePointTemperatureAccuracyProject) -> str:
    return f"{project.device_model}_{project.test_intent}_{_setpoint_token(project.setpoint_c)}"


def _setpoint_token(setpoint_c: float) -> str:
    if abs(setpoint_c - int(setpoint_c)) < 1e-9:
        return f"{int(setpoint_c)}C"
    return f"{str(setpoint_c).replace('.', 'p')}C"


def _channel_leaf(channels: tuple[str, ...], marker: str) -> str:
    for channel in channels:
        if marker in channel:
            return channel.split(".")[-1]
    return marker
