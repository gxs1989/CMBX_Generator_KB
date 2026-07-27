from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path


STAGE_COMMANDS = {
    "Instrument Setup",
    "Equilibration",
    "Inject Preparation",
    "Inject",
    "Start Run",
    "Run",
    "Stop Run",
    "Post Run",
}

STAGE_ALIASES = {
    "InstrumentSetup": "Instrument Setup",
    "InjectPreparation": "Inject Preparation",
    "StartRun": "Start Run",
    "StopRun": "Stop Run",
    "PostRun": "Post Run",
}

BRANCH_WORDS = {"If", "Else If", "Else", "End If"}
TRIGGER_PARAM_PREFIXES = ("TrueTime=", "Limit=", "Hysteresis=", "AllowImmediateExecution=")
TRIGGER_PARAM_NAMES = {"TrueTime", "Limit", "Hysteresis", "AllowImmediateExecution"}

KNOWN_COMMAND_PREFIXES = (
    "ColumnComp.",
    "PumpModule.",
    "SamplerModule.",
    "Thermometer",
    "Thermometer1.",
    "Variables.",
    "RetTimes.",
    "StabVars.",
    "TempVars.",
    "System.",
    "CC.",
    "VirtualChannel",
    "Protocol",
    "Delay",
    "Wait",
    "Log",
    "Message",
    "Condition",
    "TrueTime",
    "Limit",
    "Hysteresis",
    "AllowImmediateExecution",
    "End Trigger",
    "Trigger",
    "End",
)


def _extract_first_code_block(text: str) -> list[str]:
    match = re.search(r"```(?:text|tsv|csv)?\s*(.*?)```", text, flags=re.S | re.I)
    if not match:
        return [line.rstrip("\n") for line in text.splitlines()]
    return [line.rstrip("\n") for line in match.group(1).splitlines()]


def _split_line(line: str) -> list[str]:
    raw = line.rstrip("\r\n")
    if "\t" in raw:
        return [part.strip() for part in raw.split("\t")]
    stripped = raw.strip()
    if not stripped:
        return ["", "", "", ""]
    if stripped.startswith(";"):
        return ["", stripped.lstrip(";").strip(), "", ""]
    parts = [part.strip() for part in re.split(r"\s{2,}", stripped, maxsplit=3)]
    if len(parts) > 1:
        return parts
    branch_match = re.match(r"^(If|Else If)\s+(.+)$", stripped)
    if branch_match:
        return [branch_match.group(1), branch_match.group(2), "", ""]
    if stripped == "Else" or stripped == "End If":
        return [stripped, "", "", ""]
    trigger_match = re.match(r"^Trigger\s+(.+)$", stripped)
    if trigger_match and trigger_match.group(1).strip().startswith('"'):
        return ["Trigger", trigger_match.group(1), "", ""]
    if stripped.startswith(("Trigger ", "Triggers ", "Trigger:")):
        return ["", stripped, "", ""]
    end_trigger_match = re.match(r"^End\s+Trigger$", stripped)
    if end_trigger_match:
        return ["End Trigger", "", "", ""]
    command_match = re.match(r"^(Message|Protocol|Log|Delay|Wait)\s+(.+)$", stripped)
    if command_match:
        return ["", command_match.group(1), command_match.group(2), ""]
    assignment_match = re.match(r"^([A-Za-z][A-Za-z0-9_.]*|End)\s+(.+)$", stripped)
    if assignment_match and (
        assignment_match.group(1) == "End"
        or assignment_match.group(1).startswith(KNOWN_COMMAND_PREFIXES)
        or assignment_match.group(1) in {"ColumnComp", "Thermometer", "Thermometer1"}
    ):
        return ["", assignment_match.group(1), assignment_match.group(2), ""]
    # Keep free text comments in Command column.
    return ["", stripped, "", ""]


def _is_comment_text(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    if stripped.startswith(("Trigger ", "Triggers ", "Trigger:")):
        return True
    if stripped.startswith(KNOWN_COMMAND_PREFIXES):
        return False
    if _looks_like_cm_symbol(stripped):
        return False
    if len(stripped.split()) == 1 and re.fullmatch(r"[A-Za-z0-9_./\"=\-+(),:]+", stripped):
        return False
    if stripped.startswith("=") or stripped.startswith("-"):
        return True
    if stripped.endswith(":"):
        return True
    if stripped.startswith(("IM ", "HPLC-System", "Parameters ", "Variable ", "Initialize ", "Settings ", "Column compartment", "External ", "Pre-Run", "Trigger ", "Triggers ")):
        return True
    if stripped[0].islower():
        return False
    # Most generated prose lines have spaces and no executable value.
    return " " in stripped


def _looks_like_cm_symbol(value: str) -> bool:
    stripped = value.strip()
    if not stripped or " " in stripped:
        return False
    if stripped in {"End", "VirtualChannel"}:
        return True
    if "." not in stripped:
        return False
    return bool(re.match(r"^[A-Za-z_][A-Za-z0-9_.%]*$", stripped))


def _looks_trigger_param_fragment(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    if stripped.startswith(TRIGGER_PARAM_PREFIXES):
        return True
    if stripped.startswith("("):
        return True
    if stripped.endswith(",") and any(token in stripped for token in (" AND ", " OR ", "<=", ">=", "=", "System.Retention")):
        return True
    return False


def _normalize_row(fields: list[str]) -> dict[str, str]:
    fields = fields[:]
    while len(fields) < 4:
        fields.append("")

    f0, f1, f2, f3 = fields[:4]
    # Web models often copy the compact XML stage names. Accept them as input,
    # but normalize to the exact CM Script Editor display names immediately.
    f0 = STAGE_ALIASES.get(f0, f0)
    f1 = STAGE_ALIASES.get(f1, f1)
    time = ""
    command = ""
    value = ""
    comment = ""
    kind = "Command"

    if f0 and not f1 and not f2 and _looks_numeric_text(f0):
        time, command, value, comment = f0, "", "", f3
        kind = "Command"
    elif f0 and _looks_numeric_text(f0) and f1:
        time, command, value, comment = f0, f1, f2, f3
        if command in STAGE_COMMANDS:
            kind = "Stage"
        elif not value and _is_comment_text(command):
            kind = "Comment"
        else:
            kind = "Command"
    # Stage rows: Time + Command + optional Value/Comment.
    elif f0 in STAGE_COMMANDS and not f1:
        time, command, value, comment = "", f0, f1, f2 or f3
        kind = "Stage"
    elif f1 in STAGE_COMMANDS:
        time, command, value, comment = f0, f1, f2, f3
        if command == "Instrument Setup" and not time:
            time = "{Initial Time}"
        kind = "Stage" if command in STAGE_COMMANDS else "Command"
    elif not f0 and f1 == "End Trigger":
        time, command, value, comment = "End Trigger", "", "", f2 or f3
        kind = "Stage"
    elif not f0 and f1 == "End":
        time, command, value, comment = "", f1, f2, f3
        kind = "End" if f1 == "End" else "Stage"
    # Trigger header in generated drafts.
    elif f0 == "Trigger":
        time, command, value, comment = "Trigger", f2 or f1, "", f3
        kind = "Stage"
    elif f0 == "End Trigger":
        time, command, value, comment = "End Trigger", "", "", f1 or f2 or f3
        kind = "Stage"
    elif f0 == "End":
        time, command, value, comment = "", f0, f1 or f2, f3
        kind = "End" if f0 == "End" else "Stage"
    elif f0 in {"If", "Else If"}:
        time, command, value, comment = f0, "", f1 or f2, f3
        kind = "Branch"
    elif f0 in {"Else", "End If"}:
        time, command, value, comment = f0, "", "", f3 or f1 or f2
        kind = "Branch"
    elif f0.startswith(TRIGGER_PARAM_PREFIXES):
        time, command, value, comment = "", f0, f1 or f2, f3
        kind = "Command"
    elif not f0 and _looks_trigger_param_fragment(f1):
        time, command, value, comment = "", f1, f2, f3
        kind = "Command"
    elif not f0 and f1 in TRIGGER_PARAM_NAMES and f2:
        time, command, value, comment = "", f1, f2.rstrip(",").strip(), f3
        kind = "Command"
    # Free-text comments: blank time, text in first or second visible field.
    elif not f0 and _is_comment_text(f1):
        time, command, value, comment = "", f1, f2, f3
        kind = "Comment"
    elif f0 and not f1 and _is_comment_text(f0):
        time, command, value, comment = "", f0, "", ""
        kind = "Comment"
    # Four-column-ish command row.
    else:
        if f0 and not f1:
            command, value, comment = f0, f1, f2 or f3
        elif f0 and f1:
            command, value, comment = f0, f1, f2 or f3
        else:
            command, value, comment = f1, f2, f3
        kind = "Command"

    if kind == "Command" and not value and not comment and _is_comment_text(command) and not _looks_trigger_param_fragment(command):
        kind = "Comment"
    if kind == "Command" and not value and not comment and not command.startswith(KNOWN_COMMAND_PREFIXES) and not _looks_like_cm_symbol(command) and not _looks_trigger_param_fragment(command):
        kind = "Comment"

    if time == "End" and not command:
        command = time
        time = ""
        kind = "End"

    value = _normalize_value_for_command(command, value)

    return {
        "Kind": kind,
        "Time": time,
        "Command": command,
        "Value": value,
        "Comment": comment,
    }


def _normalize_value_for_command(command: str, value: str) -> str:
    if command.strip() == "Delay":
        match = re.fullmatch(r"\s*([-+]?\d+(?:\.\d+)?)\s*\[s\]\s*", value, flags=re.I)
        if match:
            return match.group(1)
    return value


def _looks_numeric_text(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def parse_md_to_rows(path: Path) -> list[dict[str, str]]:
    text = path.read_text(encoding="utf-8-sig")
    code_lines = _extract_first_code_block(text)
    rows: list[dict[str, str]] = []
    for line in code_lines:
        if not line.strip():
            continue
        fields = _split_line(line)
        normalized_headers = [field.strip().lower() for field in fields]
        if not rows and len(normalized_headers) >= 3 and normalized_headers[:3] == ["time", "command", "value"]:
            continue
        row = _normalize_row(fields)
        if not (row["Time"] or row["Command"] or row["Value"] or row["Comment"]):
            continue
        row["#"] = str(len(rows))
        rows.append(row)
    if rows and rows[-1]["Kind"] != "End":
        last_command = rows[-1].get("Command", "").strip()
        last_time = rows[-1].get("Time", "").strip()
        if last_command == "End" or last_time == "End":
            rows[-1]["Kind"] = "End"
            rows[-1]["Command"] = "End"
            rows[-1]["Time"] = ""
        else:
            rows.append({"#": str(len(rows)), "Kind": "End", "Time": "", "Command": "End", "Value": "", "Comment": "Inserted by renderer because source script did not end with End row"})
    return rows


def write_tsv(rows: list[dict[str, str]], path: Path, paste_only: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["Time", "Command", "Value", "Comment"] if paste_only else ["#", "Kind", "Time", "Command", "Value", "Comment"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(rows: list[dict[str, str]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# CM Rendered Method Script",
        "",
        "| # | Kind | Time | Command | Value | Comment |",
        "|---:|---|---|---|---|---|",
    ]
    for row in rows:
        vals = [row.get(k, "") for k in ["#", "Kind", "Time", "Command", "Value", "Comment"]]
        vals = [v.replace("|", "\\|").replace("\n", "<br>") for v in vals]
        lines.append("| " + " | ".join(vals) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(rows: list[dict[str, str]], path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "CM Rendered"

    headers = ["#", "Kind", "Time", "Command", "Value", "Comment"]
    ws.append(headers)
    fills = {
        "Stage": PatternFill("solid", fgColor="FFD7A3"),
        "Comment": PatternFill("solid", fgColor="FFFFFF"),
        "Branch": PatternFill("solid", fgColor="B7F29B"),
        "Command": PatternFill("solid", fgColor="F7F8FA"),
        "End": PatternFill("solid", fgColor="FFD7A3"),
    }
    header_fill = PatternFill("solid", fgColor="D9DEE7")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(k, "") for k in headers])
        excel_row = ws.max_row
        fill = fills.get(row.get("Kind", "Command"), fills["Command"])
        for cell in ws[excel_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row.get("Kind") == "Comment":
            for cell in ws[excel_row]:
                cell.font = Font(color="008000", italic=True)
        elif row.get("Kind") == "Branch":
            ws.cell(excel_row, 2).font = Font(color="006100", bold=True)

    widths = [8, 12, 16, 42, 56, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return True


def write_cm_workbook(rows: list[dict[str, str]], path: Path) -> bool:
    """Write a Chromeleon-style workbook matching the observed Book2.xlsx shape.

    The first sheet intentionally has only three columns:
    Time, Command, Value.  Row kind/comment metadata is preserved in a second
    review sheet so the main sheet remains close to the CM import format.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return False

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Match the sparse CM workbook shape: no header row, no Kind column.
    for row in rows:
        ws.append([row.get("Time", ""), row.get("Command", ""), row.get("Value", "")])
        excel_row = ws.max_row
        kind = row.get("Kind", "")
        for cell in ws[excel_row]:
            cell.alignment = Alignment(vertical="top")
        if kind == "Comment":
            ws.cell(excel_row, 2).font = Font(color="008000", italic=True)
        elif kind == "Branch":
            for cell in ws[excel_row]:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif kind in {"Stage", "End"}:
            for cell in ws[excel_row]:
                cell.fill = PatternFill("solid", fgColor="F4B183")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 92
    ws.column_dimensions["C"].width = 70

    review = wb.create_sheet("Rendered Rows")
    headers = ["#", "Kind", "Time", "Command", "Value", "Comment"]
    review.append(headers)
    for cell in review[1]:
        cell.fill = PatternFill("solid", fgColor="D9E2F3")
        cell.font = Font(bold=True)
    for row in rows:
        review.append([row.get(k, "") for k in headers])
    widths = [8, 14, 16, 52, 60, 72]
    for idx, width in enumerate(widths, start=1):
        review.column_dimensions[chr(ord("A") + idx - 1)].width = width
    review.freeze_panes = "A2"

    wb.save(path)
    return True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_md", type=Path)
    parser.add_argument("--out-dir", type=Path, default=Path("cmbx_data_explorer/outputs/cm_rendered"))
    args = parser.parse_args()

    rows = parse_md_to_rows(args.input_md)
    stem = args.input_md.stem.replace(" ", "_")
    out_dir = args.out_dir
    write_markdown(rows, out_dir / f"{stem}_CM_rendered.md")
    write_tsv(rows, out_dir / f"{stem}_CM_rendered.tsv")
    write_tsv(rows, out_dir / f"{stem}_CM_paste.tsv", paste_only=True)
    cm_ok = write_cm_workbook(rows, out_dir / f"{stem}_CM_method.xlsx")
    xlsx_ok = write_xlsx(rows, out_dir / f"{stem}_CM_rendered.xlsx")
    print(f"rows={len(rows)}")
    print(f"markdown={out_dir / f'{stem}_CM_rendered.md'}")
    print(f"tsv={out_dir / f'{stem}_CM_rendered.tsv'}")
    print(f"paste_tsv={out_dir / f'{stem}_CM_paste.tsv'}")
    print(f"cm_xlsx={'ok' if cm_ok else 'skipped'}:{out_dir / f'{stem}_CM_method.xlsx'}")
    print(f"xlsx={'ok' if xlsx_ok else 'skipped'}:{out_dir / f'{stem}_CM_rendered.xlsx'}")


if __name__ == "__main__":
    main()
