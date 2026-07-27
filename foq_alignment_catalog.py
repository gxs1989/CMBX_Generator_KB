from __future__ import annotations

from dataclasses import dataclass, replace
from pathlib import Path
import re
from typing import Any, Iterable

from sequence_cmd_parser import build_injection_method_links
from processing_method_inspector import inspect_processing_method
from tcc_project_generator import (
    build_single_point_temperature_accuracy_project,
    write_single_point_temperature_accuracy_excel_workbooks,
    write_single_point_temperature_accuracy_project,
)


DEFAULT_KB_ROOT = Path(r"C:\ProgramData\CMBX Data Explorer Workspace\KB")


@dataclass(frozen=True)
class FoqAlignmentRecord:
    order: int
    family: str
    test_intent: str
    td_test: str
    device_models: tuple[str, ...]
    td_source: str
    td_meaning: str
    key_conditions: tuple[str, ...]
    injection: str
    instrument_method: str
    processing_method: str
    report_template: str
    report_sheets: tuple[str, ...]
    report_evidence: tuple[str, ...]
    db_fields: tuple[str, ...]
    db_evidence: tuple[str, ...]
    expected_ret_times: tuple[str, ...]
    expected_channels: tuple[str, ...]
    expected_audit_properties: tuple[str, ...]
    required_config: tuple[str, ...]
    method_evidence: tuple[str, ...]
    cmbx_sources: tuple[str, ...]
    coverage_status: str
    open_gaps: tuple[str, ...]
    generation_readiness: str

    @property
    def device_label(self) -> str:
        return ", ".join(self.device_models)

    @property
    def report_sheet_label(self) -> str:
        return ", ".join(self.report_sheets)

    @property
    def db_field_label(self) -> str:
        return ", ".join(self.db_fields)


@dataclass(frozen=True)
class TccRelationshipRow:
    category: str
    item_id: str
    source: str
    target: str
    strength: str
    impact: str
    generation_rule: str
    evidence: str


@dataclass(frozen=True)
class IntentGate:
    status: str
    can_export_generic_packet: bool
    can_export_specialized_packet: bool
    runnable_generation_allowed: bool
    blockers: tuple[str, ...]
    next_actions: tuple[str, ...]


@dataclass(frozen=True)
class IntentConflictRow:
    category: str
    aspect: str
    values: str
    status: str
    impact: str
    required_action: str


@dataclass(frozen=True)
class IntentParameterImpact:
    normalized_parameter: str
    parameter_kind: str
    setpoint_c: float | None
    selected_db_fields: tuple[str, ...]
    retained_setpoints: tuple[str, ...]
    removed_db_fields: tuple[str, ...]
    affected_models: tuple[str, ...]
    supported: bool
    notes: tuple[str, ...]


@dataclass(frozen=True)
class TccBlackBoxCoverageRow:
    milestone: str
    test_id: str
    test_name: str
    document: str
    exists: bool
    contract_1_method: bool
    contract_2_processing: bool
    contract_3_report: bool
    contract_4_db: bool
    contract_5_config: bool
    contract_6_open_verification: bool
    open_verification_present: bool
    open_verification_count: int
    open_verification_topics: tuple[str, ...]
    evidence_sources_present: bool
    model_branches: str
    mermaid_present: bool
    word_count: int
    status: str


@dataclass(frozen=True)
class TccOpenVerificationTopicRow:
    milestone: str
    test_id: str
    test_name: str
    document: str
    category: str
    topic: str
    likely_evidence_source: str
    closure_action: str


@dataclass(frozen=True)
class TccMilestoneStatusRow:
    milestone: str
    objective: str
    evidence: str
    status: str
    open_work: str


@dataclass(frozen=True)
class TccTemperatureContractMatrixRow:
    test_id: str
    test_name: str
    document: str
    method_contract: str
    processing_contract: str
    report_contract: str
    db_contract: str
    config_contract: str
    open_verification_contract: str
    open_topic_categories: str
    next_closure_actions: str
    template_readiness: str


@dataclass(frozen=True)
class TccContractClosureTaskRow:
    milestone: str
    test_id: str
    test_name: str
    document: str
    contract: str
    priority: str
    evidence_group: str
    task_type: str
    topic: str
    likely_evidence_source: str
    closure_action: str
    generation_blocker: str


@dataclass(frozen=True)
class TccNextActionQueueRow:
    rank: int
    milestone: str
    priority: str
    evidence_group: str
    task_count: int
    tests: str
    contracts: str
    primary_blocker: str
    next_action: str
    unlocks: str
    generation_gate: str


@dataclass(frozen=True)
class TccEvidenceWorkstreamRow:
    milestone: str
    priority: str
    evidence_group: str
    task_count: int
    tests: str
    contracts: str
    likely_evidence_sources: str
    closure_actions: str
    unlocks: str
    next_action: str


@dataclass(frozen=True)
class TccP1EvidenceExtractionPlanRow:
    milestone: str
    test_id: str
    test_name: str
    document: str
    contract: str
    evidence_group: str
    topic: str
    evidence_source: str
    extraction_steps: str
    validation_outputs: str
    closure_update: str
    status: str


@dataclass(frozen=True)
class TccProcessingMethodTargetRow:
    milestone: str
    test_id: str
    test_name: str
    device_model: str
    injection: str
    instrument_method: str
    processing_method: str
    topic: str
    expected_behavior: str
    extraction_target: str
    source_document: str
    closure_action: str
    readiness: str


@dataclass(frozen=True)
class TccReportFormulaTargetRow:
    milestone: str
    test_id: str
    test_name: str
    device_model: str
    topic: str
    report_template: str
    report_sheets: str
    db_fields: str
    formula_id: str
    extraction_target: str
    source_document: str
    closure_action: str
    readiness: str


@dataclass(frozen=True)
class TccReportFormulaExtractionPlanRow:
    milestone: str
    test_id: str
    test_name: str
    device_model: str
    report_template: str
    report_sheets: str
    formula_id: str
    db_fields: str
    extraction_target: str
    extraction_steps: str
    validation_outputs: str
    closure_update: str
    status: str


@dataclass(frozen=True)
class RelationshipResolutionChoice:
    rule_id: str
    relationship: str
    decision_required: str
    options: tuple[str, ...]
    evidence_to_capture: str
    default_recommendation: str


def record_modifiability_summary(record: FoqAlignmentRecord) -> str:
    """Return a compact design-workbench label for whether this test can be changed."""
    if record.coverage_status == "not applicable":
        return "N/A not applicable"
    if record.coverage_status in {"missing", "open verification"}:
        return "🟡 needs black-box closure"
    if "open verification" in record.coverage_status.lower() or record.open_gaps:
        if record.test_intent in {"heatup_cooldown_20_50_20", "temperature_accuracy"}:
            return "🟡 editable after review"
        return "🟡 verify before editing"
    if record.test_intent == "temperature_calibration":
        return "🔴 locked foundation"
    if record.test_intent in {"heatup_cooldown_20_50_20", "temperature_accuracy"}:
        return "🟢 editable contract"
    if record.test_intent in {"temperature_stability_and_pcc", "temperature_stability_no_pcc", "temperature_precision_and_fan"}:
        return "🟡 merge/crop with dependency review"
    if record.test_intent in {"factory_default_metadata", "error_log_check", "qualification_service_done"}:
        return "🔴 preserve audit/state step"
    return "🟡 review required"


def record_cut_point_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    """List currently known editable/locked design points for a selected alignment row."""
    intent = record.test_intent
    if intent == "temperature_calibration":
        return (
            "🔴 Locked: calibration ladder writes device calibration variables and should not be trimmed independently.",
            "🟡 Review: low-temperature 10/5 deg C fallback can only be changed after report reach-check formula is decoded.",
            "🔴 Locked: external calibrated thermometers remain the reference source.",
            "🔴 Locked: deviation semantics are external thermometer minus internal CC actual temperature.",
        )
    if intent == "temperature_accuracy":
        return (
            "🟡 Editable after review: setpoint list can be shortened for a custom single-point accuracy method.",
            "🔴 Locked: approach/baseline rule must be explicit before removing neighboring temperature points.",
            "🔴 Locked: RetTime anchors the stable external thermometer averaging window.",
            "🔴 Locked: external upper/lower thermometers are the reference source.",
        )
    if intent == "heatup_cooldown_20_50_20":
        return (
            "🟢 Editable: temperature range, for example 20->50->20 to another explicit ramp.",
            "🟢 Editable with report update: hold-time subtraction constant, currently 2.0 min in the verified report rule.",
            "🔴 Locked: RetTime start/end anchors must remain paired with method trigger semantics.",
            "🟡 Review: modified range may require Calibration/Accuracy evidence for the selected temperature region.",
        )
    if intent in {"temperature_stability_and_pcc", "temperature_stability_no_pcc"}:
        return (
            "🟡 Editable after dependency model: stability window length and nominal setpoint may be parameterized.",
            "🔴 Locked: evaluate upper/lower external sensors separately and use the worse range.",
            "🔴 Locked for VH: PCC branch must remain VH-only.",
        )
    if intent == "temperature_precision_and_fan":
        return (
            "🟡 Editable after full black-box split: repeat count and setpoint may be parameterized.",
            "🔴 Locked: lower and upper sensor ranges must be calculated separately.",
            "🟡 Review: fan checks depend on available symbols and report branch.",
        )
    if intent in {"column_id", "factory_default_metadata", "error_log_check", "qualification_service_done"}:
        return (
            "🔴 Locked: audit/service identity checks are not numerical temperature contracts.",
            "🟡 Review: include/exclude decisions should be explicit sequence policy, not method trimming.",
        )
    if intent in {"valve_keypad", "liquid_leak_keypad", "preheater_connection"}:
        return (
            "🟡 Editable only by device configuration: availability depends on valves, leak sensor, or preheater hardware.",
            "🔴 Locked: operator/audit evidence cannot be replaced by raw temperature formulas.",
        )
    return ("🟡 No cut-point rules recorded yet; complete black-box decomposition first.",)


def record_dependency_graph_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    """Return a small text/mermaid dependency view for the selected test."""
    intent = record.test_intent
    if record.family == "TCC":
        edges = {
            "temperature_calibration": (
                "Temperature Calibration --> Temperature Accuracy",
                "Temperature Calibration --> Temperature Stability",
                "Temperature Calibration --> Temperature Precision",
                "Temperature Calibration --> HeatUp/CoolDown",
            ),
            "temperature_accuracy": (
                "Temperature Calibration --> Temperature Accuracy",
                "Temperature Accuracy --> Temperature Stability",
                "Temperature Accuracy --> Temperature Precision",
            ),
            "temperature_stability_and_pcc": (
                "Temperature Calibration --> Temperature Stability + PCC",
                "Temperature Accuracy --> Temperature Stability + PCC",
                "PCC configuration --> PCC performance",
            ),
            "temperature_stability_no_pcc": (
                "Temperature Calibration --> Temperature Stability",
                "Temperature Accuracy --> Temperature Stability",
            ),
            "temperature_precision_and_fan": (
                "Temperature Calibration --> Temperature Precision",
                "Temperature Accuracy --> Temperature Precision",
                "Fan symbol/configuration --> Fan evidence",
            ),
            "heatup_cooldown_20_50_20": (
                "Temperature Calibration --> HeatUp/CoolDown",
                "External upper thermometer --> RetTime trigger evidence",
                "RetTime pairs --> report duration rule",
            ),
            "preheater_connection": (
                "Preheater hardware --> Preheater Connection",
                "Precondition metadata --> Port pass/fail",
            ),
            "valve_keypad": (
                "Valve hardware --> Valve cycle",
                "Operator keypad action --> keypad audit evidence",
            ),
            "liquid_leak_keypad": (
                "Leak sensor configuration --> Liquid Leak",
                "Operator water/alarm action --> audit evidence",
            ),
        }.get(intent, ())
        if edges:
            return (
                "Mermaid sketch:",
                "```mermaid",
                "flowchart TD",
                *(f"    {edge}" for edge in edges),
                "```",
                "",
                "Impact notes:",
                *record_cut_point_lines(record),
                "",
                "Structured relationship rules:",
                *(record_relationship_rule_lines(record) or ("- no structured relationship rule matched",)),
            )
    if record.family == "VDAD":
        return (
            "Mermaid sketch:",
            "```mermaid",
            "flowchart TD",
            "    WarmUp --> DiagnosticCellTests",
            "    DiagnosticCellTests --> FluidicFlowCellTests",
            "    LampConfiguration --> TestApplicability",
            "```",
            "",
            "Impact notes:",
            "🟡 VDAD generation is alignment-only until method/report black-box decomposition is complete.",
        )
    return ("No dependency graph recorded yet.",)


def record_verification_action_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    gaps = record.open_gaps or ("No open verification item recorded for this alignment row.",)
    lines = [
        "Use this checklist as a review queue. Closing an item should update the source KB/decomposition document, not only the UI row.",
        "",
    ]
    for index, gap in enumerate(gaps, start=1):
        lines.extend(
            [
                f"{index}. [ ] {gap}",
                "   Evidence needed: CMBX decoded method/report object, CM UI screenshot, exported report workbook, or manually confirmed KB note.",
                "   Closure action: add the confirmation to the relevant black-box decomposition or module KB, then refresh the catalog.",
            ]
        )
    return tuple(lines)


INTENT_TOOL_OPTIONS = (
    "Search / Recommend",
    "Crop / Modify",
    "Merge",
    "Compare",
)


def intent_tool_options() -> tuple[str, ...]:
    return INTENT_TOOL_OPTIONS


def record_intent_gate(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
) -> IntentGate:
    """Return structured readiness for a selected intent.

    The gate separates review packet availability from runnable generation.
    Runnable generation remains closed unless all contracts are complete and CM
    validation has happened outside this workbench.
    """
    selected = _unique_records((*tuple(selected_records), record))
    normalized_intent = (intent or INTENT_TOOL_OPTIONS[0]).strip().lower()
    blockers: list[str] = []
    next_actions: list[str] = []

    if any(item.coverage_status == "not applicable" for item in selected):
        blockers.append("Selected row is not applicable for the chosen device/model branch.")
    if any(item.coverage_status in {"missing", "open verification"} for item in selected):
        blockers.append("Black-box coverage is missing or still open.")
    black_box_rows = tuple(tcc_black_box_coverage_for_record(item) for item in selected if item.family == "TCC")
    if any(row is None or not row.exists for row in black_box_rows):
        blockers.append("One or more selected TCC tests has no black-box decomposition document.")
    if any(row is not None and not row.status.startswith("documented") for row in black_box_rows):
        blockers.append("One or more selected TCC black-box documents is missing six-contract coverage.")
    if any(row is not None and not row.evidence_sources_present for row in black_box_rows):
        blockers.append("One or more selected TCC black-box documents has no Evidence Sources section.")
    topic_rows = _intent_gate_open_verification_topic_rows(selected)
    if topic_rows:
        categories = _unique(row.category for row in topic_rows)
        blockers.append("Black-box open verification topics remain by category: " + ", ".join(categories) + ".")
    if any(item.open_gaps for item in selected):
        blockers.append("Open verification remains in one or more selected rows.")
    if any("locked" in record_modifiability_summary(item).lower() or "foundation" in record_modifiability_summary(item).lower() for item in selected):
        blockers.append("One or more selected rows are locked/foundation/finalization steps.")
    if normalized_intent.startswith("merge") and len(selected) < 2:
        blockers.append("Merge preview needs at least two selected alignment rows.")
    relationship_rows = _intent_gate_relationship_rows(selected, normalized_intent)
    if relationship_rows:
        blockers.extend(_intent_gate_relationship_blockers(relationship_rows, normalized_intent))

    specialized = (
        len(selected) == 1
        and record.family == "TCC"
        and record.test_intent == "temperature_accuracy"
        and normalized_intent.startswith("crop")
        and bool(re.search(r"(-?\d+(?:\.\d+)?)", parameter or ""))
    )
    generic = not any(item.coverage_status == "not applicable" for item in selected)
    if specialized:
        next_actions.append("Specialized Temperature Accuracy single-setpoint draft packet can be exported for review.")
    elif generic:
        next_actions.append("Generic review-only draft packet can be exported after selecting exactly one device model.")
    if topic_rows:
        next_actions.extend(_intent_gate_topic_next_actions(topic_rows))
    if relationship_rows:
        next_actions.extend(_intent_gate_relationship_next_actions(relationship_rows))
    if blockers:
        next_actions.append("Close blockers in the relevant black-box decomposition or KB before runnable generation.")
    else:
        next_actions.append("Review method, processing, report, DB, and configuration contracts before CM validation.")

    if blockers:
        status = "blocked for runnable generation"
    elif specialized:
        status = "specialized draft available"
    elif normalized_intent.startswith(("crop", "merge", "compare", "search")):
        status = "review packet available"
    else:
        status = "review required"

    return IntentGate(
        status=status,
        can_export_generic_packet=generic,
        can_export_specialized_packet=specialized,
        runnable_generation_allowed=False,
        blockers=_unique(blockers),
        next_actions=_unique(next_actions),
    )


def _intent_gate_open_verification_topic_rows(
    selected_records: Iterable[FoqAlignmentRecord],
) -> tuple[TccOpenVerificationTopicRow, ...]:
    rows: list[TccOpenVerificationTopicRow] = []
    seen: set[tuple[str, str, str]] = set()
    for record in selected_records:
        for row in open_verification_topics_for_record(record):
            key = (row.test_id, row.category, row.topic)
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
    return tuple(rows)


def _intent_gate_topic_next_actions(
    topic_rows: Iterable[TccOpenVerificationTopicRow],
    limit: int = 3,
) -> tuple[str, ...]:
    rows = tuple(topic_rows)
    actions = [
        f"Close {row.category} topic for {row.test_id}: {row.topic.rstrip('.。 ')}. Evidence: {row.likely_evidence_source}."
        for row in rows[:limit]
    ]
    if len(rows) > limit:
        actions.append(f"Review {len(rows) - limit} additional open-verification topic(s) in the BlackBox Audit closure queue.")
    return tuple(actions)


def _intent_gate_relationship_rows(
    selected_records: Iterable[FoqAlignmentRecord],
    normalized_intent: str,
) -> tuple[TccRelationshipRow, ...]:
    rows: list[TccRelationshipRow] = []
    seen: set[str] = set()
    for record in selected_records:
        for row in record_relationship_rows(record, normalized_intent):
            if "hard" not in row.strength.lower():
                continue
            if row.item_id in seen:
                continue
            seen.add(row.item_id)
            rows.append(row)
    return tuple(rows)


def _intent_gate_relationship_blockers(
    rows: Iterable[TccRelationshipRow],
    normalized_intent: str,
    limit: int = 4,
) -> tuple[str, ...]:
    row_tuple = tuple(rows)
    if not row_tuple:
        return ()
    blockers = [
        f"Hard relationship rule {row.item_id} ({row.category}) must be resolved before {normalized_intent or 'intent'} can generate runnable output: {row.generation_rule}"
        for row in row_tuple[:limit]
    ]
    if len(row_tuple) > limit:
        blockers.append(f"{len(row_tuple) - limit} additional hard relationship rule(s) require review before runnable generation.")
    return tuple(blockers)


def _intent_gate_relationship_next_actions(
    rows: Iterable[TccRelationshipRow],
    limit: int = 4,
) -> tuple[str, ...]:
    row_tuple = tuple(rows)
    actions = [
        f"Resolve relationship {row.item_id}: {row.source} -> {row.target}. Rule: {row.generation_rule}"
        for row in row_tuple[:limit]
    ]
    if len(row_tuple) > limit:
        actions.append(f"Review {len(row_tuple) - limit} additional hard relationship rule(s) in Relationship Audit.")
    return tuple(actions)


def relationship_resolution_choices(
    record: FoqAlignmentRecord,
    intent: str = "Crop / Modify",
    selected_records: Iterable[FoqAlignmentRecord] = (),
) -> tuple[RelationshipResolutionChoice, ...]:
    selected = _unique_records((*tuple(selected_records), record))
    normalized_intent = (intent or "Crop / Modify").strip().lower()
    rows = _intent_gate_relationship_rows(selected, normalized_intent)
    choices: list[RelationshipResolutionChoice] = []
    seen: set[str] = set()
    for row in rows:
        if row.item_id in seen:
            continue
        seen.add(row.item_id)
        choices.append(_relationship_resolution_choice(row))
    return tuple(choices)


def _relationship_resolution_choice(row: TccRelationshipRow) -> RelationshipResolutionChoice:
    mapping: dict[str, RelationshipResolutionChoice] = {
        "ORDER_01": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "BurnIn / preconditioning scope decision",
            (
                "keep BurnIn before the cropped/full sequence",
                "replace BurnIn with an explicitly documented preconditioning method",
                "declare BurnIn out of scope for non-FOQ exploratory output",
            ),
            "method package note plus black-box/TD evidence explaining thermal preconditioning choice",
            "Keep BurnIn for full FOQ; require an explicit preconditioning note for cropped packages.",
        ),
        "ORDER_02": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "Calibration scope decision",
            (
                "reuse existing Temperature Calibration evidence",
                "rerun Temperature Calibration before the selected test",
                "declare Calibration out of scope for a non-FOQ exploratory run",
            ),
            "selected option, calibration evidence source, affected report/DB assumptions",
            "For generated FOQ-like output, rerun or explicitly bind Temperature Calibration evidence.",
        ),
        "DEP_01": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "Accuracy calibration dependency decision",
            (
                "reuse existing calibration ladder and cite source sequence",
                "include Temperature Calibration in the generated/cropped sequence",
                "block runnable generation until Calibration dependency is manually waived",
            ),
            "calibration source sequence/injection, report mapping impact, DB field subset decision",
            "Do not generate runnable Accuracy output until Calibration reuse/rerun is explicit.",
        ),
        "DEP_02": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "Precision/Stability calibration dependency decision",
            (
                "reuse calibration context and preserve separate lower/upper sensor rules",
                "include Calibration before the merged/cropped temperature window",
                "redesign report windows and mark the output as non-FOQ",
            ),
            "calibration evidence plus report formula proof for lower/upper sensor range semantics",
            "Preserve Calibration context and separate sensor-range formulas.",
        ),
        "DEP_03": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "VH PCC branch decision",
            (
                "keep VH Stability/PCC branch",
                "switch to no-PCC branch only when device is explicitly VC/VA",
                "block generation if PCC hardware availability is unknown",
            ),
            "AUDIT.ColumnComp.ModelNo, PCC channel/config evidence, report template branch",
            "Use PCC branch for VH; never infer from filename alone.",
        ),
        "ORDER_03": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "Finalization chain decision",
            (
                "keep Factory Default followed by Error Log Check for full FOQ",
                "omit both only for explicitly partial/non-final test packages",
                "split into a separate finalization package with documented state impact",
            ),
            "sequence scope note and audit/state evidence for final state behavior",
            "Keep the finalization chain together for full FOQ packages.",
        ),
        "RES_01": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "External thermometer configuration decision",
            (
                "confirm ExtTemp_UpperCC and ExtTemp_LowerCC channels in target CM configuration",
                "map equivalent Generic Device thermometer channels and update report formulas",
                "block runnable generation when thermometer channels are unavailable",
            ),
            "CM instrument configuration, required-symbol manifest, raw channel evidence",
            "Require explicit external thermometer channel evidence before method reuse.",
        ),
        "RES_02": RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "Device identity branch decision",
            (
                "use AUDIT.ColumnComp.ModelNo as source of truth",
                "manually select branch only when ModelNo evidence is absent and documented",
                "block DB/report branch selection when device identity is ambiguous",
            ),
            "AUDIT.ColumnComp.ModelNo trace, report template selection, DB mapping table",
            "Prefer AUDIT.ColumnComp.ModelNo; do not infer branch from filenames.",
        ),
    }
    return mapping.get(
        row.item_id,
        RelationshipResolutionChoice(
            row.item_id,
            f"{row.source} -> {row.target}",
            "Relationship rule decision",
            (row.generation_rule,),
            row.evidence,
            "Resolve this relationship rule before runnable generation.",
        ),
    )


def record_resolution_choice_lines(record: FoqAlignmentRecord, intent: str = "Crop / Modify") -> tuple[str, ...]:
    choices = relationship_resolution_choices(record, intent, selected_records=(record,))
    if not choices:
        return (
            "No hard relationship resolution choices are mapped for this row and intent.",
            "If generation is still blocked, review Relationship Audit and Open Verification Topics.",
        )
    lines = [
        "Relationship Resolution Choices",
        "",
        "These are design decisions that must be recorded before a generated package can be treated as runnable.",
        "",
        "| Rule | Relationship | Decision Required | Options | Evidence To Capture | Default Recommendation |",
        "|---|---|---|---|---|---|",
    ]
    for choice in choices:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    choice.rule_id,
                    choice.relationship,
                    choice.decision_required,
                    "; ".join(choice.options),
                    choice.evidence_to_capture,
                    choice.default_recommendation,
                )
            )
            + " |"
        )
    return tuple(lines)


def _resolution_choices_markdown_lines(choices: Iterable[RelationshipResolutionChoice]) -> tuple[str, ...]:
    choice_tuple = tuple(choices)
    lines = [
        "| Rule | Relationship | Decision Required | Options | Evidence To Capture | Default Recommendation |",
        "|---|---|---|---|---|---|",
    ]
    if not choice_tuple:
        lines.append("| (none) | (none) | (none) | (none) | (none) | (none) |")
        return tuple(lines)
    for choice in choice_tuple:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    choice.rule_id,
                    choice.relationship,
                    choice.decision_required,
                    "; ".join(choice.options),
                    choice.evidence_to_capture,
                    choice.default_recommendation,
                )
            )
            + " |"
        )
    return tuple(lines)


def relationship_decision_register_tsv(
    record: FoqAlignmentRecord,
    intent: str = "Crop / Modify",
    parameter: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
) -> str:
    """Render fillable decision-register rows for hard relationship choices."""
    selected = _unique_records((*tuple(selected_records), record))
    choices = relationship_resolution_choices(record, intent, selected_records=selected)
    headers = (
        "Rule ID",
        "Relationship",
        "Decision Required",
        "Available Options",
        "Selected Option",
        "Decision Status",
        "Evidence Path",
        "Owner",
        "Notes",
    )
    rows: list[tuple[object, ...]] = [headers]
    if not choices:
        rows.append(
            (
                "(none)",
                "(none)",
                "No hard relationship decision is mapped for this packet.",
                "",
                "",
                "Closed",
                "",
                "",
                f"Intent={intent or '(none)'}; Parameter={parameter or '(none)'}",
            )
        )
    for choice in choices:
        rows.append(
            (
                choice.rule_id,
                choice.relationship,
                choice.decision_required,
                " | ".join(choice.options),
                "",
                "Open",
                "",
                "",
                choice.default_recommendation,
            )
        )
    return "\n".join("\t".join(_tsv_escape(cell) for cell in row) for row in rows) + "\n"


def record_intent_preview(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> str:
    """Render a non-mutating intent impact preview for the alignment workbench."""
    selected = _unique_records((*tuple(selected_records), record))
    candidates = tuple(candidate_records)
    normalized_intent = (intent or INTENT_TOOL_OPTIONS[0]).strip().lower()
    gate = record_intent_gate(record, intent, parameter, selected_records=selected)
    if normalized_intent.startswith("crop"):
        preview = _crop_intent_preview(record, parameter)
    elif normalized_intent.startswith("merge"):
        preview = _merge_intent_preview(record, selected, candidates)
    elif normalized_intent.startswith("compare"):
        preview = _compare_intent_preview(record, selected)
    else:
        preview = _search_intent_preview(record, parameter, selected, candidates)
    return "\n".join((*_intent_gate_preview_lines(gate), "", preview))


def record_intent_parameter_impact(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    device_model: str = "",
) -> IntentParameterImpact:
    """Return structured impact for a user-entered intent parameter."""
    normalized_intent = (intent or "").strip().lower()
    normalized_parameter = (parameter or "").strip()
    if not normalized_intent.startswith("crop"):
        return IntentParameterImpact(
            normalized_parameter=normalized_parameter,
            parameter_kind="unstructured",
            setpoint_c=None,
            selected_db_fields=(),
            retained_setpoints=(),
            removed_db_fields=(),
            affected_models=(),
            supported=False,
            notes=("Structured parameter impact is currently defined for Crop / Modify intents.",),
        )
    if record.family != "TCC" or record.test_intent != "temperature_accuracy":
        return IntentParameterImpact(
            normalized_parameter=normalized_parameter,
            parameter_kind="unstructured",
            setpoint_c=None,
            selected_db_fields=(),
            retained_setpoints=(),
            removed_db_fields=(),
            affected_models=record.device_models,
            supported=False,
            notes=("No structured parameter model is defined for this test yet.",),
        )

    setpoint = _parse_numeric_setpoint(normalized_parameter)
    if setpoint is None:
        return IntentParameterImpact(
            normalized_parameter=normalized_parameter,
            parameter_kind="temperature_accuracy_setpoint",
            setpoint_c=None,
            selected_db_fields=(),
            retained_setpoints=(),
            removed_db_fields=(),
            affected_models=record.device_models,
            supported=False,
            notes=("Enter a numeric temperature setpoint, for example 40 C.",),
        )

    requested_devices = (device_model.strip(),) if device_model.strip() else record.device_models
    affected_models: list[str] = []
    selected_db_fields: list[str] = []
    removed_db_fields: list[str] = []
    retained_setpoints: list[str] = []
    notes: list[str] = []
    target_db_field = _temperature_accuracy_setpoint_db_field(setpoint)

    for device in requested_devices:
        if device not in record.device_models:
            notes.append(f"{device} is not applicable to {record.td_test}.")
            continue
        affected_models.append(device)
        fields = _device_db_fields(record, device)
        temp_fields = tuple(field for field in fields if field.startswith("TempAcc"))
        result_fields = tuple(field for field in fields if field.startswith("RES_"))
        if target_db_field not in temp_fields:
            notes.append(
                f"{device} does not expose {target_db_field}; known Accuracy fields are {', '.join(temp_fields) or 'none'}."
            )
            removed_db_fields.extend(temp_fields)
            continue
        selected_db_fields.append(target_db_field)
        selected_db_fields.extend(result_fields)
        retained_setpoints.append(f"{_format_setpoint(setpoint)} C")
        removed_db_fields.extend(field for field in temp_fields if field != target_db_field)

    supported = bool(affected_models) and any(field == target_db_field for field in selected_db_fields) and not any(
        "does not expose" in note or "not applicable" in note for note in notes
    )
    if supported:
        notes.append(
            "Report rows/cells for removed TempAcc fields must be removed, hidden, or explicitly marked not applicable."
        )
        notes.append("RES_TempAccuracy remains meaningful only if its workbook rule is narrowed to the retained setpoint.")

    return IntentParameterImpact(
        normalized_parameter=normalized_parameter,
        parameter_kind="temperature_accuracy_setpoint",
        setpoint_c=setpoint,
        selected_db_fields=_unique(selected_db_fields),
        retained_setpoints=_unique(retained_setpoints),
        removed_db_fields=_unique(removed_db_fields),
        affected_models=_unique(affected_models),
        supported=supported,
        notes=_unique(notes),
    )


def record_intent_parameter_impact_lines(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    device_model: str = "",
) -> tuple[str, ...]:
    impact = record_intent_parameter_impact(record, intent, parameter, device_model=device_model)
    if impact.parameter_kind == "unstructured" and not impact.normalized_parameter:
        return ()
    lines = [
        f"- Parameter kind: {impact.parameter_kind}",
        f"- Normalized parameter: {impact.normalized_parameter or '(none)'}",
        f"- Supported by selected scope: {'yes' if impact.supported else 'no'}",
    ]
    if impact.setpoint_c is not None:
        lines.append(f"- Parsed setpoint: {_format_setpoint(impact.setpoint_c)} C")
    lines.extend(
        (
            f"- Affected models: {', '.join(impact.affected_models) or '(none)'}",
            f"- Retained setpoints: {', '.join(impact.retained_setpoints) or '(none)'}",
            f"- Selected DB fields: {', '.join(impact.selected_db_fields) or '(none)'}",
            f"- Removed / unused DB fields: {', '.join(impact.removed_db_fields) or '(none)'}",
        )
    )
    if impact.notes:
        lines.append("- Notes:")
        lines.extend(f"  - {note}" for note in impact.notes)
    return tuple(lines)


def render_test_plan_assistant_markdown(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    device_model: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> str:
    """Render the user-facing test-plan assistant view.

    The view is intentionally simpler than the full alignment packet. It follows
    the user's desired workflow:

    1. User intent or existing intent selection.
    2. Editable templates found from the current framework.
    3. Modification/change plan and review boundary.
    """
    selected = _unique_records((*tuple(selected_records), record))
    device = device_model.strip() or (record.device_models[0] if len(record.device_models) == 1 else record.device_label)
    normalized_intent = intent.strip() or "Search / Recommend"
    normalized_parameter = parameter.strip()
    gate = record_intent_gate(record, normalized_intent, normalized_parameter, selected_records=selected)
    impact_lines = record_intent_parameter_impact_lines(record, normalized_intent, normalized_parameter, device_model=device if "," not in device else "")
    candidate_lines = _test_plan_candidate_template_lines(record, selected, device, gate)
    change_lines = _test_plan_change_plan_lines(record, normalized_intent, normalized_parameter, selected, gate)
    next_actions = gate.next_actions or ("Review Config -> Method -> Report evidence before changing the package.",)
    lines = [
        f"# Test Plan Assistant - {record.family} - {record.td_test}",
        "",
        "## 1. Intent Source",
        "",
        "| Item | Value |",
        "|---|---|",
        f"| Selected intent | {_md_cell(normalized_intent)} |",
        f"| Parameter | {_md_cell(normalized_parameter or '(none)')} |",
        f"| Anchor test | `{_md_cell(record.test_intent)}` |",
        f"| Device scope | `{_md_cell(device or '(not selected)')}` |",
        f"| Modifiability | {_md_cell(record_modifiability_summary(record))} |",
        f"| Intent gate | {_md_cell(gate.status)} |",
        f"| Runnable generation | {'allowed' if gate.runnable_generation_allowed else 'not allowed'} |",
        "",
        "Structured parameter interpretation:",
        "",
        *(impact_lines or ("- No structured parameter model for this intent/test yet.",)),
        "",
        "## 2. Editable Templates From Current Framework",
        "",
        "| Template | Existing framework source | Editable surface | Locked / must preserve |",
        "|---|---|---|---|",
        *candidate_lines,
        "",
        "## 3. Modification / Change Plan",
        "",
        *change_lines,
        "",
        "## Review Boundary",
        "",
        "- Primary proof order: Instrument Config -> Instrument Method Script -> Report Formula.",
        "- Processing Method is a downstream CM automation check.",
        "- DB mapping is a downstream export/upload check.",
        "- Do not claim runnable CMBX until binary method/report payload behavior is verified in CM.",
        "",
        "## Next Actions",
        "",
        *(f"- {action}" for action in next_actions),
    ]
    blockers = gate.blockers
    if blockers:
        lines.extend(("", "## Current Blockers", "", *(f"- {blocker}" for blocker in blockers)))
    return "\n".join(lines).rstrip() + "\n"


def build_test_plan_modification_steps(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    device_model: str = "",
) -> tuple[tuple[str, str, str, str, str], ...]:
    """Return concrete UI rows for what to modify and what file to expect.

    Columns are:
    Asset, Use Template, Modify Exact Location, Change To, Output.
    """
    device = device_model.strip() or (record.device_models[0] if record.device_models else "")
    normalized_intent = (intent or "").strip().lower()
    normalized_parameter = (parameter or "").strip()

    if record.family == "TCC" and record.test_intent == "temperature_accuracy" and normalized_intent.startswith("crop"):
        impact = record_intent_parameter_impact(record, intent, normalized_parameter, device_model=device)
        setpoint = impact.setpoint_c
        if setpoint is not None:
            token = f"{_format_setpoint(setpoint)}C".replace(".", "p")
            db_field = _temperature_accuracy_setpoint_db_field(setpoint)
            ret_time = _temperature_accuracy_ret_time_for_device(device, setpoint)
            row = _temperature_accuracy_report_row_for_device(device, setpoint)
            injection = _device_injection(record, device) or record.injection or "Temperature Accuracy_H/C"
            source_report = _device_report_template(record, device) or record.report_template
            return (
                (
                    "Config",
                    "required_configuration.md / CM Instrument Setup",
                    "ColumnComp + Thermometer1 + RetTimes/StabVars symbols",
                    f"Confirm `{device}` model, upper/lower external thermometer channels, and writable RetTimes before copying method.",
                    "required_configuration.md; method workbook Configuration sheet",
                ),
                (
                    "Instrument Method",
                    record.instrument_method or "TEMPERATURE_ACCURACY",
                    "[Equilibration] and [Run] setpoint ladder",
                    f"Keep baseline as explicit design input; set target nominal to `{_format_setpoint(setpoint)} C`; remove non-requested Accuracy transitions.",
                    f"method_script_{token}.txt; {record.instrument_method}__single_{token}_method_script.xlsx",
                ),
                (
                    "Instrument Method",
                    record.instrument_method or "TEMPERATURE_ACCURACY",
                    f"[Run] RetTime anchor `{ret_time}`",
                    f"Emit `RetTimes.{ret_time} = System.Retention` only after CC ready + upper/lower external thermometer stable window exists.",
                    f"method_script_{token}.txt",
                ),
                (
                    "Report Template",
                    source_report or "(report template open verification)",
                    f"`Temp Accuracy` row {row}: J/K/L/M/C/D/E",
                    f"Keep nominal `{_format_setpoint(setpoint)}`, K{row}=`AUDIT.{ret_time}`, L/M average window, D{row}=`{db_field}`, E{row}=pass/fail.",
                    f"{source_report}__single_{token}_report_calculation.xlsx; report_formula_map_{token}.tsv",
                ),
                (
                    "Sequence",
                    injection,
                    "Sequence row Instrument Method cell",
                    f"Use generated method name `{record.instrument_method}__single_{token}`; keep processing binding only as downstream CM behavior review.",
                    "sequence_template.tsv",
                ),
            )

    if record.family == "TCC" and record.test_intent == "heatup_cooldown_20_50_20":
        same_range = _looks_like_heatup_default_range(normalized_parameter)
        change = (
            "No method range change needed for 20->50->20; use source method as template and verify RetTime formula chain."
            if same_range
            else "For a custom range, update start/target/return setpoints, trigger windows, stable hold timing, and RetTime endpoint semantics."
        )
        report_change = (
            "Preserve row-66 external endpoint formula: HeatUp=RetTime2-RetTime1-2.0; CoolDown=RetTime5-RetTime4-2.0."
            if same_range
            else "Rebuild report labels and formulas so the retained RetTime endpoints match the custom temperature range."
        )
        return (
            (
                "Config",
                "config_contract.md / CM Instrument Setup",
                "ColumnComp.CC + Thermometer1.ExtTemp_UpperCC + RetTimes1..6",
                "Confirm CC temperature control, external upper thermometer, GenericLong/model context, and all RetTimes exist.",
                "config_method_report_review.md",
            ),
            (
                "Instrument Method",
                record.instrument_method or "TEMP_HEAT_UP_DOWN_20_50_20",
                "[Run] 17 C precondition -> 20 C -> 50 C -> 20 C command flow",
                change,
                "Generic packet only now; dedicated CM script renderer not implemented for custom HeatUp/CoolDown.",
            ),
            (
                "Report Template",
                record.report_template or "Report_VTCC/VATCC",
                "`HeatUp&CoolDown` row 66 and summary cells D26/D27",
                report_change,
                "report_db_contract.tsv; config_method_report_review.md",
            ),
            (
                "Sequence",
                record.injection or "HeatUp and CoolDownTime",
                "Sequence row method/report binding",
                "Use source injection/method binding for unchanged range; for custom range, create a reviewed clone name.",
                "sequence_template.tsv",
            ),
        )

    source_report = _device_report_template(record, device) or record.report_template or "(open verification)"
    source_method = _device_instrument_method(record, device) or record.instrument_method or "(open verification)"
    return (
        (
            "Config",
            "KB required_config",
            "Required symbols/channels",
            "Close missing config evidence before editing.",
            "config_method_report_review.md",
        ),
        (
            "Instrument Method",
            source_method,
            "Decoded method flow",
            "No concrete renderer exists yet; inspect method evidence and create a reviewed clone plan.",
            "Generic review packet",
        ),
        (
            "Report Template",
            source_report,
            "Report sheets/formula objects",
            "Extract exact formulas before changing outputs.",
            "Generic review packet",
        ),
    )


def _looks_like_heatup_default_range(text: str) -> bool:
    compact = re.sub(r"\s+", "", text.lower())
    return not compact or compact in {"20->50->20", "20→50→20", "20to50to20", "20到50到20"}


def _temperature_accuracy_ret_time_for_device(device_model: str, setpoint_c: float) -> str:
    if abs(setpoint_c - 10) < 1e-9:
        return "RetTime1"
    if abs(setpoint_c - 20) < 1e-9:
        return "RetTime2"
    if abs(setpoint_c - 40) < 1e-9:
        return "RetTime3"
    if device_model == "VH-C10-A":
        if abs(setpoint_c - 80) < 1e-9:
            return "RetTime4"
        if abs(setpoint_c - 120) < 1e-9:
            return "RetTime5"
    else:
        if abs(setpoint_c - 60) < 1e-9:
            return "RetTime4"
        if abs(setpoint_c - 85) < 1e-9:
            return "RetTime5"
    return "RetTime?"


def _temperature_accuracy_report_row_for_device(device_model: str, setpoint_c: float) -> int:
    ret_time = _temperature_accuracy_ret_time_for_device(device_model, setpoint_c)
    mapping = {"RetTime1": 66, "RetTime2": 67, "RetTime3": 68, "RetTime4": 69, "RetTime5": 70}
    return mapping.get(ret_time, 0)


def _test_plan_candidate_template_lines(
    record: FoqAlignmentRecord,
    selected_records: tuple[FoqAlignmentRecord, ...],
    device_model: str,
    gate: IntentGate,
) -> tuple[str, ...]:
    rows: list[str] = []
    for item in selected_records:
        injection = _device_injection(item, device_model) or item.injection or "(open verification)"
        method = _device_instrument_method(item, device_model) or item.instrument_method or "(open verification)"
        processing = _device_processing_method(item, device_model) or item.processing_method or "(open verification)"
        report_template = _device_report_template(item, device_model) or item.report_template or "(open verification)"
        report_sheets = ", ".join(_device_report_sheets(item, device_model) or item.report_sheets) or "(not mapped)"
        db_fields = ", ".join(_device_db_fields(item, device_model) or item.db_fields) or "(not mapped)"
        rows.extend(
            (
                _markdown_table_row(
                    "Sequence row template",
                    f"`{injection}` from alignment sequence binding",
                    "Injection name, sample defaults, selected method/report binding",
                    "Keep device branch and required sequence order unless dependency model says otherwise",
                ),
                _markdown_table_row(
                    "Instrument method template",
                    f"`{method}` from CMBX method evidence",
                    _generic_method_edit_implication(item, ""),
                    "Preserve required symbols, RetTimes, waits/triggers, acquisition channels",
                ),
                _markdown_table_row(
                    "Report template",
                    f"`{report_template}` / `{report_sheets}`",
                    _generic_report_edit_implication(item, ""),
                    "Preserve FormulaOne/workbook rule semantics and display precision",
                ),
                _markdown_table_row(
                    "Config template",
                    "Required config/channel manifest from TKN + method evidence",
                    "Confirm external devices, imported variables, channels, writable symbols",
                    "; ".join(item.required_config) or "No config evidence recorded yet",
                ),
                _markdown_table_row(
                    "Downstream binding",
                    f"Processing `{processing}`; DB `{db_fields}`",
                    "Review only after Config/Method/Report plan is coherent",
                    "Do not use Processing/DB as proof that command script is correct",
                ),
            )
        )
        if item is not record:
            rows.append(
                _markdown_table_row(
                    "Merged-row note",
                    item.td_test,
                    "Selected together with anchor row",
                    "Check conflict matrix before merging",
                )
            )
    if gate.can_export_specialized_packet:
        rows.append(
            _markdown_table_row(
                "Specialized draft packet",
                "Existing Temperature Accuracy single-setpoint generator",
                "Exports reviewable method/report/config draft assets",
                "Still not a runnable CMBX claim",
            )
        )
    elif gate.can_export_generic_packet:
        rows.append(
            _markdown_table_row(
                "Generic draft packet",
                "Current alignment binding",
                "Exports review-only sequence/config/method/report plan",
                "No method/report payload rewrite",
            )
        )
    return tuple(rows)


def _test_plan_change_plan_lines(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    selected_records: tuple[FoqAlignmentRecord, ...],
    gate: IntentGate,
) -> tuple[str, ...]:
    lines: list[str] = []
    lines.extend(("### A. Decide the change", ""))
    if parameter:
        lines.append(f"- Requested parameter: `{parameter}`")
    else:
        lines.append("- No parameter entered yet; choose a concrete setpoint, range, merge target, or comparison target.")
    lines.extend(("", "Editable / locked points:"))
    lines.extend(f"- {line}" for line in record_cut_point_lines(record))
    lines.extend(("", "### B. Apply the Config -> Method -> Report check", ""))
    for item in selected_records:
        lines.append(f"- `{item.td_test}`")
        lines.append(f"  - Config: verify {', '.join(item.required_config) or 'required symbols/channels are recorded'}.")
        lines.append(f"  - Method: {_generic_method_edit_implication(item, parameter)}")
        lines.append(f"  - Report: {_generic_report_edit_implication(item, parameter)}")
    lines.extend(("", "### C. Produce review output", ""))
    if gate.can_export_specialized_packet:
        lines.append("- Export `Draft Packet` to get the specialized review assets for this intent.")
    elif gate.can_export_generic_packet:
        lines.append("- Export `Draft Packet` to get a generic review packet with `config_method_report_review.md`.")
    else:
        lines.append("- Export only the markdown review/action plan until the blocking contracts are closed.")
    lines.append("- Use `Export MD` for human review notes and `Action Plan` for closure tasks.")
    lines.extend(("", "### D. Manual test expectation", ""))
    lines.append("- In CM, recreate or clone the method/report only after P1 evidence is accepted.")
    lines.append("- Run one safe sequence/injection first and compare report cells against the review formula chain.")
    lines.append("- Only then evaluate Processing/DB behavior.")
    return tuple(lines)


def _markdown_table_row(*values: object) -> str:
    return "| " + " | ".join(_md_cell(value) for value in values) + " |"


def _intent_gate_preview_lines(gate: IntentGate) -> tuple[str, ...]:
    lines = [
        "Intent Gate:",
        f"- Status: {gate.status}",
        f"- Generic draft packet: {'available' if gate.can_export_generic_packet else 'not available'}",
        f"- Specialized draft packet: {'available' if gate.can_export_specialized_packet else 'not available'}",
        f"- Runnable generation: {'allowed' if gate.runnable_generation_allowed else 'closed until CM validation'}",
    ]
    if gate.blockers:
        lines.extend(("- Blockers:", *(f"  - {item}" for item in gate.blockers)))
    if gate.next_actions:
        lines.extend(("- Next actions:", *(f"  - {item}" for item in gate.next_actions)))
    return tuple(lines)


def render_intent_review_markdown(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> str:
    """Render a review packet for a selected intent preview."""
    selected = _unique_records((*tuple(selected_records), record))
    node = test_knowledge_node_from_record(record)
    gate = record_intent_gate(record, intent, parameter, selected_records=selected)
    preview = record_intent_preview(
        record,
        intent,
        parameter,
        selected_records=selected,
        candidate_records=tuple(candidate_records),
    )
    lines = [
        f"# Intent Review - {record.family} - {record.td_test}",
        "",
        "## Review Metadata",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Intent | {_md_cell(intent or INTENT_TOOL_OPTIONS[0])} |",
        f"| Parameter | {_md_cell(parameter.strip() or '(none)')} |",
        f"| Anchor TestIntent | `{_md_cell(record.test_intent)}` |",
        f"| Anchor Injection | `{_md_cell(record.injection or '(not bound)')}` |",
        f"| Anchor Method | `{_md_cell(record.instrument_method or '(not bound)')}` |",
        f"| Coverage | {_md_cell(record.coverage_status)} |",
        f"| Modifiability | {_md_cell(record_modifiability_summary(record))} |",
        f"| Intent Gate | {_md_cell(gate.status)} |",
        f"| Generic Draft Packet | {'available' if gate.can_export_generic_packet else 'not available'} |",
        f"| Specialized Draft Packet | {'available' if gate.can_export_specialized_packet else 'not available'} |",
        "",
        "## Boundary",
        "",
        "This review packet is non-mutating. It does not rewrite CMBX, instrument methods, processing methods, report templates, or DB mappings.",
        "Use it as a design-review artifact before deciding whether a method/report package can be generated or modified.",
        "",
        "## Selected Alignment Rows",
        "",
        "| Family | Test | Device | Injection | Method | Processing | Report | Coverage |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for item in selected:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    item.family,
                    item.td_test,
                    item.device_label,
                    item.injection or "(not bound)",
                    item.instrument_method or "(not bound)",
                    item.processing_method or "(not bound)",
                    item.report_template or "(not bound)",
                    item.coverage_status,
                )
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## Anchor Test Knowledge Node",
            "",
            "| Field | Value |",
            "|---|---|",
            f"| test_id | `{_md_cell(node.test_id)}` |",
            f"| test_name | {_md_cell(node.test_name)} |",
            f"| foq_section | {_md_cell(node.foq_section)} |",
            f"| formula_id | `{_md_cell(node.formula_id)}` |",
            f"| db_fields | {_md_cell(node.db_field_label or '(not mapped)')} |",
            f"| model_applicability | {_md_cell(node.model_label)} |",
            f"| irc_injected | {'Yes' if node.irc_injected else 'No'} |",
            "",
            "## Required Contracts",
            "",
            "### RetTimes",
            *_markdown_bullets(node.expected_ret_times, empty="- none recorded"),
            "",
            "### Channels",
            *_markdown_bullets(node.expected_channels, empty="- none recorded"),
            "",
            "### Audit / Properties",
            *_markdown_bullets(node.expected_audit_properties, empty="- none recorded"),
            "",
            "### Configuration",
            *_markdown_bullets(record.required_config, empty="- none recorded"),
            "",
            "## Intent Preview",
            "",
            "```text",
            preview,
            "```",
            "",
            "## Open Verification",
            "",
            *_markdown_bullets(record.open_gaps, empty="- none recorded"),
            "",
            "## Next Review Actions",
            "",
            "- Close open verification items in the source KB or black-box decomposition document.",
            "- Re-run FOQ Knowledge Alignment and export a new review packet.",
            "- Only after Method, Processing, Report, DB, and Config contracts are closed should generation or CMBX modification be attempted.",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def write_intent_review_markdown(
    output_path: str | Path,
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        render_intent_review_markdown(
            record,
            intent,
            parameter,
            selected_records=selected_records,
            candidate_records=candidate_records,
        ),
        encoding="utf-8",
    )
    return path


def render_intent_action_plan_markdown(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> str:
    """Render a structured modification plan derived from an intent preview."""
    selected = _unique_records((*tuple(selected_records), record))
    gate = record_intent_gate(record, intent, parameter, selected_records=selected)
    preview = record_intent_preview(
        record,
        intent,
        parameter,
        selected_records=selected,
        candidate_records=tuple(candidate_records),
    )
    plan_rows = _intent_action_plan_rows(record, intent, parameter, selected)
    conflict_rows = build_intent_conflict_rows(selected, intent=intent)
    parameter_impact_lines = record_intent_parameter_impact_lines(record, intent, parameter)
    closure_topic_rows = _intent_gate_open_verification_topic_rows(selected)
    resolution_choices = relationship_resolution_choices(record, intent, selected_records=selected)
    lines = [
        f"# Intent Action Plan - {record.family} - {record.td_test}",
        "",
        "## Scope",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Intent | {_md_cell(intent or INTENT_TOOL_OPTIONS[0])} |",
        f"| Parameter | {_md_cell(parameter.strip() or '(none)')} |",
        f"| Anchor TestIntent | `{_md_cell(record.test_intent)}` |",
        f"| Anchor Device(s) | {_md_cell(record.device_label)} |",
        f"| Anchor Injection | `{_md_cell(record.injection or '(not bound)')}` |",
        f"| Edit Gate | {_md_cell(gate.status)} |",
        f"| Generic Draft Packet | {'available' if gate.can_export_generic_packet else 'not available'} |",
        f"| Specialized Draft Packet | {'available' if gate.can_export_specialized_packet else 'not available'} |",
        f"| Runnable Generation | {'allowed' if gate.runnable_generation_allowed else 'closed until CM validation'} |",
        "",
        "## Boundary",
        "",
        "This action plan is a task list for method/report/CMBX modification. It is not an executable modification and does not claim runnable output.",
        "",
        "## Intent Gate Blockers",
        "",
        "### Blockers",
        *_markdown_bullets(gate.blockers, empty="- none recorded"),
        "",
        "### Next Actions",
        *_markdown_bullets(gate.next_actions, empty="- none recorded"),
        "",
        "## Relationship Resolution Choices",
        "",
        *_resolution_choices_markdown_lines(resolution_choices),
    ]
    lines.extend(
        [
            "",
            "## Selected Rows",
            "",
            "| Test | Device | Injection | Method | Processing | Report | Coverage |",
            "|---|---|---|---|---|---|---|",
        ]
    )
    for item in selected:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    item.td_test,
                    item.device_label,
                    item.injection or "(not bound)",
                    item.instrument_method or "(not bound)",
                    item.processing_method or "(not bound)",
                    item.report_template or "(not bound)",
                    item.coverage_status,
                )
            )
            + " |"
        )
    if parameter_impact_lines:
        lines.extend(
            [
                "",
                "## Structured Parameter Impact",
                "",
                *parameter_impact_lines,
            ]
        )
    lines.extend(
        [
            "",
            "## Layered Modification Tasks",
            "",
            "| Layer | Task | Evidence to update/verify | Status |",
            "|---|---|---|---|",
        ]
    )
    for layer, task, evidence, status in plan_rows:
        lines.append(f"| {_md_cell(layer)} | {_md_cell(task)} | {_md_cell(evidence)} | {_md_cell(status)} |")
    lines.extend(
        [
            "",
            "## Intent Conflict Matrix",
            "",
            "| Category | Aspect | Values | Status | Impact | Required Action |",
            "|---|---|---|---|---|---|",
        ]
    )
    for row in conflict_rows:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    row.category,
                    row.aspect,
                    row.values,
                    row.status,
                    row.impact,
                    row.required_action,
                )
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## Conflict-Driven Required Actions",
            "",
        ]
    )
    for row in conflict_rows:
        if row.status in {"single value", "single-row review", "none recorded"}:
            continue
        lines.append(f"- [{row.status}] {row.category} / {row.aspect}: {row.required_action}")
    lines.extend(
        [
            "",
            "## Contract Checklist",
            "",
            "### RetTimes",
            *_markdown_bullets(record.expected_ret_times, empty="- none recorded"),
            "",
            "### Channels",
            *_markdown_bullets(record.expected_channels, empty="- none recorded"),
            "",
            "### Audit / Properties",
            *_markdown_bullets(record.expected_audit_properties, empty="- none recorded"),
            "",
            "### Required Configuration",
            *_markdown_bullets(record.required_config, empty="- none recorded"),
            "",
            "## Preview Basis",
            "",
            "```text",
            preview,
            "```",
            "",
            "## Open Verification",
            "",
            *_markdown_bullets(_unique(gap for item in selected for gap in item.open_gaps), empty="- none recorded"),
            "",
            "## Open Verification Closure Queue",
            "",
            "| Test ID | Category | Topic | Likely Evidence Source | Closure Action |",
            "|---|---|---|---|---|",
        ]
    )
    if closure_topic_rows:
        for topic_row in closure_topic_rows:
            lines.append(
                "| "
                + " | ".join(
                    _md_cell(value)
                    for value in (
                        topic_row.test_id,
                        topic_row.category,
                        topic_row.topic,
                        topic_row.likely_evidence_source,
                        topic_row.closure_action,
                    )
                )
                + " |"
            )
    else:
        lines.append("| (none) | (none) | (none) | (none) | (none) |")
    lines.extend(
        [
            "",
            "## Exit Criteria Before Generation",
            "",
            "- Every task above is either completed or explicitly declared out of scope.",
            "- Method command flow has line-level evidence for each changed RetTime/channel/config symbol.",
            "- Processing method IRC/pass-action behavior is preserved or manually redesigned.",
            "- Report formula/workbook-derived cells and display precision are updated and verified.",
            "- DB field subset, SQL type, and displayed precision are reviewed against mapping.",
            "- Modified package is validated in Chromeleon before being treated as runnable.",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def write_intent_action_plan_markdown(
    output_path: str | Path,
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        render_intent_action_plan_markdown(
            record,
            intent,
            parameter,
            selected_records=selected_records,
            candidate_records=candidate_records,
        ),
        encoding="utf-8",
    )
    return path


def write_intent_draft_asset_packet(
    output_root: str | Path,
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str = "",
    device_model: str = "",
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> dict[str, Path]:
    """Write a reviewable draft asset packet for the selected alignment row.

    Dedicated generators can emit more concrete draft assets. All other rows
    still get a conservative reuse/review packet so the workbench can move from
    knowledge alignment into auditable method/report planning without claiming a
    runnable CMBX.
    """
    device = _draft_asset_device(record, device_model)
    selected = _unique_records((*tuple(selected_records), record))
    _validate_draft_asset_selection_for_device(selected, device)
    normalized_intent = (intent or "").strip().lower()
    if (
        len(selected) != 1
        or record.family != "TCC"
        or record.test_intent != "temperature_accuracy"
        or not normalized_intent.startswith("crop")
    ):
        return _write_generic_intent_draft_asset_packet(
            output_root,
            record,
            intent,
            parameter,
            device,
            selected_records=selected,
            candidate_records=candidate_records,
        )
    if not re.search(r"(-?\d+(?:\.\d+)?)", parameter or ""):
        return _write_generic_intent_draft_asset_packet(
            output_root,
            record,
            intent,
            parameter,
            device,
            selected_records=selected,
            candidate_records=candidate_records,
        )

    setpoint = _draft_asset_setpoint(record, intent, parameter)

    project = build_single_point_temperature_accuracy_project(device, setpoint)
    base = Path(output_root)
    project_dir = write_single_point_temperature_accuracy_project(project, base)
    excel_paths = write_single_point_temperature_accuracy_excel_workbooks(project, base)
    candidates = tuple(candidate_records)
    review_path = write_intent_review_markdown(
        project_dir / "intent_review.md",
        record,
        intent,
        parameter,
        selected_records=selected,
        candidate_records=candidates,
    )
    action_path = write_intent_action_plan_markdown(
        project_dir / "intent_action_plan.md",
        record,
        intent,
        parameter,
        selected_records=selected,
        candidate_records=candidates,
    )
    decision_register_path = project_dir / "relationship_decision_register.tsv"
    decision_register_path.write_text(
        relationship_decision_register_tsv(record, intent, parameter, selected_records=selected),
        encoding="utf-8",
    )
    manifest_path = project_dir / "draft_asset_packet_manifest.md"
    manifest_path.write_text(
        _draft_asset_packet_manifest_text(record, intent, parameter, device, project_dir, excel_paths),
        encoding="utf-8",
    )
    return {
        "project_dir": project_dir,
        "manifest": manifest_path,
        "intent_review": review_path,
        "intent_action_plan": action_path,
        "relationship_decision_register": decision_register_path,
        "sequence_template": project_dir / "sequence_template.tsv",
        "processing_method_binding": project_dir / "processing_method_binding.md",
        "config_method_report_review": project_dir / "config_method_report_review.md",
        **excel_paths,
    }


def _crop_intent_preview(record: FoqAlignmentRecord, parameter: str) -> str:
    requested_change = parameter.strip() or "(no explicit parameter entered)"
    lines = [
        "Intent: Crop / Modify",
        f"Selected test: {_record_label(record)}",
        f"Requested change: {requested_change}",
        f"Current modifiability: {record_modifiability_summary(record)}",
        "",
        "Required contracts to preserve:",
        *_prefixed_or_default("RetTimes", record.expected_ret_times),
        *_prefixed_or_default("Channels", record.expected_channels),
        *_prefixed_or_default("Audit/properties", record.expected_audit_properties),
        *_prefixed_or_default("Configuration", record.required_config),
        "",
        "Known editable / locked points:",
        *(f"- {line}" for line in record_cut_point_lines(record)),
    ]
    parameter_notes = _crop_parameter_notes(record, parameter)
    if parameter_notes:
        lines.extend(("", "Parameter-specific checks:", *(f"- {line}" for line in parameter_notes)))
    parameter_impact_lines = record_intent_parameter_impact_lines(record, "Crop / Modify", parameter)
    if parameter_impact_lines:
        lines.extend(("", "Structured parameter impact:", *parameter_impact_lines))
    lines.extend(
        (
            "",
            "Relationship model rules:",
            *(record_relationship_rule_lines(record, "Crop / Modify") or ("- no structured relationship rule matched",)),
            "",
            "Dependency impact:",
            *record_dependency_graph_lines(record),
            "",
            "Open verification before modifying:",
            *_gap_lines(record),
            "",
            "Preview result:",
            _intent_preview_result(record),
        )
    )
    return "\n".join(lines)


def _merge_intent_preview(
    record: FoqAlignmentRecord,
    selected_records: tuple[FoqAlignmentRecord, ...],
    candidate_records: tuple[FoqAlignmentRecord, ...],
) -> str:
    if len(selected_records) < 2:
        suggestions = _merge_suggestions(record, candidate_records)
        return "\n".join(
            [
                "Intent: Merge",
                f"Selected test: {_record_label(record)}",
                "",
                "Select at least two alignment rows to preview a real merge.",
                "",
                "Suggested nearby rows:",
                *(f"- {_record_label(item)}" for item in suggestions),
                "",
                "Baseline dependency impact:",
                *record_dependency_graph_lines(record),
            ]
        )
    shared_channels = _shared_values(item.expected_channels for item in selected_records)
    shared_ret_times = _shared_values(item.expected_ret_times for item in selected_records)
    all_gaps = _unique(gap for item in selected_records for gap in item.open_gaps)
    templates = sorted({item.report_template for item in selected_records if item.report_template})
    processing_methods = sorted({item.processing_method for item in selected_records if item.processing_method})
    inherited_gap_lines = tuple(f"- {gap}" for gap in all_gaps) if all_gaps else ("- none recorded",)
    lines = [
        "Intent: Merge",
        "Selected rows:",
        *(f"- {_record_label(item)}" for item in selected_records),
        "",
        "Shared resources:",
        *_prefixed_or_default("Shared RetTimes", shared_ret_times),
        *_prefixed_or_default("Shared Channels", shared_channels),
        "",
        "Merge conflict checks:",
        f"- Report templates involved: {', '.join(templates) or '(none)'}",
        f"- Processing methods involved: {', '.join(processing_methods) or '(none)'}",
        "- RetTime names can be shared only if the method phases remain semantically identical.",
        "- Report sheets must be redesigned if merged rows currently write to different sheets/cells.",
        "- Processing-method IRC/pass-action behavior must remain explicit; do not infer it from DB fields.",
        "",
        "Relationship model rules:",
        *_unique(rule for item in selected_records for rule in record_relationship_rule_lines(item, "Merge")),
        "",
        "Open verification inherited by merge:",
        *inherited_gap_lines,
        "",
        "Preview result:",
        "Review-only merge candidate. Generate no CMBX until Method, Processing, Report, DB, and Config contracts are all closed.",
    ]
    return "\n".join(lines)


def _compare_intent_preview(record: FoqAlignmentRecord, selected_records: tuple[FoqAlignmentRecord, ...]) -> str:
    lines = [
        "Intent: Compare",
        f"Anchor: {_record_label(record)}",
        "",
    ]
    if len(selected_records) > 1:
        lines.extend(("Selected rows:",))
        for item in selected_records:
            lines.extend(
                [
                    f"- {_record_label(item)}",
                    f"  Method: {item.instrument_method or '(not bound)'}",
                    f"  Processing: {item.processing_method or '(not bound)'}",
                    f"  Report: {item.report_template or '(not bound)'} / {item.report_sheet_label or '(not mapped)'}",
                    f"  Coverage: {item.coverage_status}",
                    f"  Modifiability: {record_modifiability_summary(item)}",
                ]
            )
    else:
        lines.extend(
            [
                "Device branch comparison:",
                "| Device | Injection | Instrument Method | Processing Method | Report Template | DB Fields |",
                "|---|---|---|---|---|---|",
            ]
        )
        for device in record.device_models:
            lines.append(
                "| "
                + " | ".join(
                    [
                        device,
                        _device_injection(record, device) or record.injection or "(not bound)",
                        _device_instrument_method(record, device) or record.instrument_method or "(not bound)",
                        _device_processing_method(record, device) or record.processing_method or "(not bound)",
                        _device_report_template(record, device) or record.report_template or "(not bound)",
                        ", ".join(_device_db_fields(record, device) or record.db_fields) or "(not mapped)",
                    ]
                )
                + " |"
            )
    lines.extend(("", "Open verification:", *_gap_lines(record)))
    return "\n".join(lines)


def _search_intent_preview(
    record: FoqAlignmentRecord,
    parameter: str,
    selected_records: tuple[FoqAlignmentRecord, ...],
    candidate_records: tuple[FoqAlignmentRecord, ...],
) -> str:
    query = parameter.strip().lower()
    source = candidate_records or selected_records or (record,)
    if query:
        matches = tuple(
            item
            for item in source
            if query
            in " ".join(
                (
                    item.family,
                    item.test_intent,
                    item.td_test,
                    item.injection,
                    item.instrument_method,
                    item.processing_method,
                    item.report_template,
                    item.db_field_label,
                    item.device_label,
                )
            ).lower()
        )
    else:
        matches = tuple(item for item in source if item.family == record.family)
    matches = matches[:12]
    return "\n".join(
        [
            "Intent: Search / Recommend",
            f"Anchor: {_record_label(record)}",
            f"Query: {parameter.strip() or '(family/current filter)'}",
            "",
            "Recommended alignment rows:",
            *(
                f"- {_record_label(item)} | {item.instrument_method or '(method open)'} | {item.report_template or '(report open)'} | {record_modifiability_summary(item)}"
                for item in matches
            ),
            "" if matches else "- no matching row in current filter",
            "",
            "Use this intent for retrieval/config recommendation only. It does not modify methods, reports, or CMBX files.",
        ]
    )


def _crop_parameter_notes(record: FoqAlignmentRecord, parameter: str) -> tuple[str, ...]:
    text = parameter.lower()
    if record.test_intent == "temperature_accuracy" and ("40" in text or "single" in text):
        return (
            "Single-point 40 C accuracy must still define the approach/baseline temperature rule.",
            "Report rows/cells for unused setpoints must be removed or marked not applicable, not silently left blank.",
            "DB output should target only the selected TempAcc field plus result/pass-fail fields that remain meaningful.",
        )
    if record.test_intent == "heatup_cooldown_20_50_20" and ("->" in text or "to" in text):
        return (
            "Method setpoints, trigger thresholds, RetTime labels, and report captions must change together.",
            "The verified report subtracts a 2.0 min hold; keep or revise this as an explicit report contract.",
        )
    if record.test_intent == "temperature_calibration":
        return ("Calibration edits are generation-blocking until the calibration report/workbook formulas are fully closed.",)
    return ()


def _intent_action_plan_rows(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    selected_records: tuple[FoqAlignmentRecord, ...],
) -> tuple[tuple[str, str, str, str], ...]:
    normalized_intent = (intent or "").strip().lower()
    if normalized_intent.startswith("crop"):
        return _crop_action_plan_rows(record, parameter)
    if normalized_intent.startswith("merge"):
        return _merge_action_plan_rows(record, selected_records)
    if normalized_intent.startswith("compare"):
        return _compare_action_plan_rows(record, selected_records)
    return _search_action_plan_rows(record, parameter)


def _crop_action_plan_rows(record: FoqAlignmentRecord, parameter: str) -> tuple[tuple[str, str, str, str], ...]:
    requested_change = parameter.strip() or "requested crop/modify parameter"
    rows = [
        (
            "Method",
            f"Draft changed method flow for {requested_change}; preserve semantic RetTime anchors.",
            ", ".join(record.expected_ret_times) or "method-flow RetTime evidence",
            "blocked until black-box contract review",
        ),
        (
            "Method",
            "Preserve or explicitly revise acquired raw channels and audit logging.",
            ", ".join((*record.expected_channels, *record.expected_audit_properties)) or "channel/audit evidence",
            "required",
        ),
        (
            "Processing",
            "Review IRC/pass-action behavior before removing or shortening related injections.",
            record.processing_method or "processing method binding",
            "open verification" if record.open_gaps else "review required",
        ),
        (
            "Report",
            "Update report sheet rows/cells/formulas and display precision for the changed scope.",
            record.report_sheet_label or "report formula trace",
            "required",
        ),
        (
            "DB",
            "Select the DB field subset that remains meaningful after the crop.",
            record.db_field_label or "DB mapping",
            "required",
        ),
        (
            "Config",
            "Verify required device symbols/channels exist in the target CM configuration.",
            "; ".join(record.required_config) or "required-symbol manifest",
            "required",
        ),
        (
            "Validation",
            "Run the modified method/report in CM and compare exported report cells against expected contract values.",
            "CM import/run, report export, FOQ DB preview",
            "not started",
        ),
    ]
    if record.test_intent == "temperature_accuracy" and ("40" in parameter.lower() or "single" in parameter.lower()):
        rows.insert(
            1,
            (
                "Method",
                "Define the 40 C approach/baseline rule before deleting neighboring accuracy setpoints.",
                "Temperature Accuracy black-box decomposition + TCC relationship model",
                "blocking design decision",
            ),
        )
        rows.insert(
            5,
            (
                "Report",
                "Remove, hide, or mark unused TempAcc setpoint rows as not applicable; do not leave stale pass/fail formulas.",
                "Report_VTCC/VATCC Temp Accuracy sheet + FormulaOne workbook rules",
                "required",
            ),
        )
    return tuple(rows)


def _merge_action_plan_rows(
    record: FoqAlignmentRecord,
    selected_records: tuple[FoqAlignmentRecord, ...],
) -> tuple[tuple[str, str, str, str], ...]:
    selected = selected_records or (record,)
    report_templates = ", ".join(sorted({item.report_template for item in selected if item.report_template})) or "report templates"
    processing_methods = ", ".join(sorted({item.processing_method for item in selected if item.processing_method})) or "processing methods"
    return (
        ("Method", "Map method phases from selected tests and identify RetTime/channel collisions.", "selected method flows", "review required"),
        ("Processing", "Confirm pass-action/IRC behavior remains valid after merge.", processing_methods, "open verification"),
        ("Report", "Design a combined report layout or preserve separate report sheets with explicit source rows.", report_templates, "required"),
        ("DB", "Decide whether merged output writes one DB row, multiple DB rows, or a new review-only output.", "FOQ mapping + DB upload contract", "design required"),
        ("Config", "Check union of required symbols and device options.", "required-symbol manifest + loaded CMBX evidence", "required"),
        ("Validation", "Run merged prototype in CM and compare each original report/DB contract.", "golden CMBX comparison", "not started"),
    )


def _compare_action_plan_rows(
    record: FoqAlignmentRecord,
    selected_records: tuple[FoqAlignmentRecord, ...],
) -> tuple[tuple[str, str, str, str], ...]:
    return (
        ("Selection", "Confirm compared rows or device branches are intentionally selected.", "Alignment table selection", "ready"),
        ("Method", "Compare instrument method names and command contracts.", "method evidence panel", "review"),
        ("Processing", "Compare processing methods and IRC involvement.", "processing binding", "review"),
        ("Report", "Compare report templates/sheets/formula IDs.", "report evidence panel", "review"),
        ("DB", "Compare DB field sets and precision/type evidence.", "DB evidence panel", "review"),
    )


def _search_action_plan_rows(record: FoqAlignmentRecord, parameter: str) -> tuple[tuple[str, str, str, str], ...]:
    return (
        ("Search", f"Use query `{parameter.strip() or record.family}` to retrieve candidate alignment rows.", "current alignment catalog", "ready"),
        ("Recommendation", "Inspect candidate modifiability and open verification before choosing an edit target.", "Intent Preview + Design Actions", "review"),
        ("Next step", "Switch to Crop / Modify, Merge, or Compare after selecting target rows.", "Alignment table selection", "manual selection"),
    )


def _intent_edit_gate(record: FoqAlignmentRecord, selected_records: tuple[FoqAlignmentRecord, ...]) -> str:
    selected = selected_records or (record,)
    if any(item.coverage_status == "not applicable" for item in selected):
        return "Closed: selected row is not applicable"
    if any(item.coverage_status in {"missing", "open verification"} for item in selected):
        return "Blocked: black-box closure required"
    if any(item.open_gaps for item in selected):
        return "Blocked: open verification remains"
    if any(record_modifiability_summary(item).startswith("🔴") for item in selected):
        return "Review required: locked/foundation step"
    return "Reviewable: contracts appear closed, CM validation still required"


def _draft_asset_device(record: FoqAlignmentRecord, device_model: str) -> str:
    device = device_model.strip()
    if device:
        if device not in record.device_models:
            raise ValueError(f"Selected device {device} is not applicable to {record.td_test}.")
        return device
    if len(record.device_models) == 1:
        return record.device_models[0]
    raise ValueError("Select exactly one device model before exporting a draft asset packet.")


def _validate_draft_asset_selection_for_device(records: Iterable[FoqAlignmentRecord], device_model: str) -> None:
    incompatible = [record.td_test for record in records if device_model not in record.device_models]
    if incompatible:
        raise ValueError(
            f"Selected device {device_model} is not applicable to: {', '.join(incompatible)}. "
            "Adjust the device filter or selected alignment rows before exporting."
        )


def _draft_asset_setpoint(record: FoqAlignmentRecord, intent: str, parameter: str) -> float:
    if record.family != "TCC" or record.test_intent != "temperature_accuracy":
        raise ValueError("Draft asset packet currently supports only TCC Temperature Accuracy.")
    if not (intent or "").strip().lower().startswith("crop"):
        raise ValueError("Draft asset packet currently supports only Crop / Modify.")
    setpoint = _parse_numeric_setpoint(parameter)
    if setpoint is None:
        raise ValueError("Enter a numeric setpoint in the parameter field, for example 40 C.")
    return setpoint


def _parse_numeric_setpoint(parameter: str) -> float | None:
    match = re.search(r"(-?\d+(?:\.\d+)?)", parameter or "")
    if not match:
        return None
    return float(match.group(1))


def _temperature_accuracy_setpoint_db_field(setpoint_c: float) -> str:
    return f"TempAcc{_format_setpoint(setpoint_c)}"


def _format_setpoint(setpoint_c: float) -> str:
    value = float(setpoint_c)
    if abs(value - int(value)) < 1e-9:
        return str(int(value))
    return f"{value:g}"


def _draft_asset_packet_manifest_text(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    device_model: str,
    project_dir: Path,
    excel_paths: dict[str, Path],
) -> str:
    files = [
        project_dir / "project_spec.json",
        project_dir / "instrument_method_draft.txt",
        project_dir / "sequence_template.tsv",
        project_dir / "processing_method_binding.md",
        project_dir / "required_configuration.md",
        project_dir / "config_method_report_review.md",
        project_dir / "report_calculation_spec.md",
        project_dir / "generation_notes.md",
        project_dir / "intent_review.md",
        project_dir / "intent_action_plan.md",
        project_dir / "relationship_decision_register.tsv",
        *excel_paths.values(),
    ]
    lines = [
        f"# Draft Asset Packet - {record.family} - {record.td_test}",
        "",
        "## Scope",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Device | `{_md_cell(device_model)}` |",
        f"| Intent | {_md_cell(intent)} |",
        f"| Parameter | {_md_cell(parameter or '(none)')} |",
        f"| Source TestIntent | `{_md_cell(record.test_intent)}` |",
        f"| Source Instrument Method | `{_md_cell(record.instrument_method)}` |",
        f"| Source Processing Method | `{_md_cell(record.processing_method)}` |",
        f"| Source Report Template | {_md_cell(record.report_template)} |",
        "",
        "## Boundary",
        "",
        "This packet contains reviewable draft assets. It is not a Chromeleon-signed method, processing method, report template, or runnable CMBX package.",
        "",
        "## Core Black-Box Review",
        "",
        "`config_method_report_review.md` is the main review entry point for this packet. It connects the three P1 contracts in the intended order:",
        "",
        "1. Instrument Config: device identity, external thermometer channels, imported variables/symbols.",
        "2. Instrument Method Script: setpoint approach, stability wait, RetTime anchor, acquisition commands.",
        "3. Report Formula: raw/audit source formulas, retained report row, display precision, pass/fail rule.",
        "",
        "Processing Method and DB mapping are recorded as downstream evidence only; they do not prove the method script is correct.",
        "",
        "## Files",
        "",
    ]
    for path in files:
        if path.exists():
            lines.append(f"- `{path.name}`")
    lines.extend(
        [
            "",
            "## Relationship Resolution Choices",
            "",
            *_resolution_choices_markdown_lines(relationship_resolution_choices(record, intent, selected_records=(record,))),
            "",
            "## Required Review Before CM Use",
            "",
            "- Confirm the setpoint approach/baseline design in the TD/method evidence.",
            "- Copy/recreate the method script in Chromeleon only after symbol/config requirements are available.",
            "- Update or clone the report template and verify FormulaOne/workbook-derived rules.",
            "- Validate exported report cells and FOQ DB preview before upload or release use.",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def _write_generic_intent_draft_asset_packet(
    output_root: str | Path,
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    device_model: str,
    selected_records: Iterable[FoqAlignmentRecord] = (),
    candidate_records: Iterable[FoqAlignmentRecord] = (),
) -> dict[str, Path]:
    """Write a conservative draft packet for reuse/review-only intents."""
    selected = _unique_records((*tuple(selected_records), record))
    candidates = tuple(candidate_records)
    project_dir = Path(output_root) / _generic_packet_dir_name(record, intent, parameter, device_model)
    project_dir.mkdir(parents=True, exist_ok=True)

    sequence_path = project_dir / "sequence_template.tsv"
    binding_path = project_dir / "method_report_binding.md"
    config_path = project_dir / "config_contract.md"
    core_review_path = project_dir / "config_method_report_review.md"
    report_db_path = project_dir / "report_db_contract.tsv"
    conflict_path = project_dir / "intent_conflict_matrix.tsv"
    decision_register_path = project_dir / "relationship_decision_register.tsv"
    boundary_path = project_dir / "generation_boundary.md"
    review_path = project_dir / "intent_review.md"
    action_path = project_dir / "intent_action_plan.md"
    manifest_path = project_dir / "draft_asset_packet_manifest.md"

    sequence_path.write_text(_generic_sequence_template_tsv(selected, device_model), encoding="utf-8")
    binding_path.write_text(_generic_method_report_binding_markdown(record, intent, parameter, device_model, selected), encoding="utf-8")
    config_path.write_text(_generic_config_contract_markdown(record, device_model, selected), encoding="utf-8")
    core_review_path.write_text(
        _generic_config_method_report_review_markdown(record, intent, parameter, device_model, selected),
        encoding="utf-8",
    )
    report_db_path.write_text(_generic_report_db_contract_tsv(selected, device_model), encoding="utf-8")
    conflict_path.write_text(_generic_intent_conflict_matrix_tsv(selected, device_model, intent), encoding="utf-8")
    decision_register_path.write_text(
        relationship_decision_register_tsv(record, intent, parameter, selected_records=selected),
        encoding="utf-8",
    )
    boundary_path.write_text(_generic_generation_boundary_markdown(record, intent, parameter, device_model, selected), encoding="utf-8")
    write_intent_review_markdown(
        review_path,
        record,
        intent,
        parameter,
        selected_records=selected,
        candidate_records=candidates,
    )
    write_intent_action_plan_markdown(
        action_path,
        record,
        intent,
        parameter,
        selected_records=selected,
        candidate_records=candidates,
    )
    manifest_path.write_text(
        _generic_draft_asset_packet_manifest_markdown(
            record,
            intent,
            parameter,
            device_model,
            (
                sequence_path,
                binding_path,
                config_path,
                core_review_path,
                report_db_path,
                conflict_path,
                decision_register_path,
                boundary_path,
                review_path,
                action_path,
            ),
        ),
        encoding="utf-8",
    )
    return {
        "project_dir": project_dir,
        "manifest": manifest_path,
        "intent_review": review_path,
        "intent_action_plan": action_path,
        "sequence_template": sequence_path,
        "method_report_binding": binding_path,
        "config_contract": config_path,
        "config_method_report_review": core_review_path,
        "report_db_contract": report_db_path,
        "intent_conflict_matrix": conflict_path,
        "relationship_decision_register": decision_register_path,
        "generation_boundary": boundary_path,
    }


def _generic_packet_dir_name(record: FoqAlignmentRecord, intent: str, parameter: str, device_model: str) -> str:
    bits = (
        device_model,
        record.family,
        record.test_intent,
        intent or "intent",
        parameter or "review",
    )
    text = "_".join(_normalize_name(bit) for bit in bits if bit)
    return text[:120] or "draft_asset_packet"


def _generic_sequence_template_tsv(records: Iterable[FoqAlignmentRecord], device_model: str) -> str:
    headers = (
        "Row",
        "Injection Name",
        "Type",
        "Status",
        "Instrument Method",
        "Processing Method",
        "Report Template",
        "Report Sheets",
        "DB Fields",
        "Review Note",
    )
    rows: list[tuple[object, ...]] = []
    for index, record in enumerate(records, start=1):
        injection = _device_injection(record, device_model) or record.injection
        method = _device_instrument_method(record, device_model) or record.instrument_method
        processing = _device_processing_method(record, device_model) or record.processing_method
        report_template = _device_report_template(record, device_model) or record.report_template
        report_sheets = _device_report_sheets(record, device_model) or record.report_sheets
        db_fields = _device_db_fields(record, device_model) or record.db_fields
        rows.append(
            (
                index,
                injection or "(open verification)",
                "Unknown",
                "Idle",
                method or "(open verification)",
                processing or "(open verification)",
                report_template or "(open verification)",
                ", ".join(report_sheets) or "(not mapped)",
                ", ".join(db_fields) or "(not mapped)",
                "Review-only sequence row; clone/recreate in CM after method/report/config contracts are closed.",
            )
        )
    return "\t".join(headers) + "\n" + "\n".join("\t".join(_tsv_escape(value) for value in row) for row in rows) + "\n"


def _generic_method_report_binding_markdown(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    device_model: str,
    selected_records: Iterable[FoqAlignmentRecord],
) -> str:
    node = test_knowledge_node_from_record(record)
    injection = _device_injection(record, device_model) or record.injection
    method = _device_instrument_method(record, device_model) or record.instrument_method
    processing = _device_processing_method(record, device_model) or record.processing_method
    report_template = _device_report_template(record, device_model) or record.report_template
    report_sheets = _device_report_sheets(record, device_model) or record.report_sheets
    db_fields = _device_db_fields(record, device_model) or record.db_fields
    lines = [
        f"# Generic Draft Asset Packet - {record.family} - {record.td_test}",
        "",
        "This packet is a review-only clone/select plan. It does not rewrite a Chromeleon method, processing method, report payload, or CMBX package.",
        "",
        "## Binding Summary",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Device | `{_md_cell(device_model)}` |",
        f"| Intent | {_md_cell(intent or '(none)')} |",
        f"| Parameter | {_md_cell(parameter or '(none)')} |",
        f"| TestIntent | `{_md_cell(record.test_intent)}` |",
        f"| FOQ section | {_md_cell(node.foq_section)} |",
        f"| Injection | `{_md_cell(injection or '(open verification)')}` |",
        f"| Instrument method | `{_md_cell(method or '(open verification)')}` |",
        f"| Processing method | `{_md_cell(processing or '(open verification)')}` |",
        f"| Report template | `{_md_cell(report_template or '(open verification)')}` |",
        f"| Report sheets | {_md_cell(', '.join(report_sheets) or '(not mapped)')} |",
        f"| Formula ID | `{_md_cell(node.formula_id)}` |",
        f"| DB fields | {_md_cell(', '.join(db_fields) or '(not mapped)')} |",
        f"| Coverage | {_md_cell(record.coverage_status)} |",
        f"| Modifiability | {_md_cell(record_modifiability_summary(record))} |",
        "",
        "## Reuse / Modify Boundary",
        "",
        "- Reuse the listed injection/method/report names only as evidence-backed anchors.",
        "- Any command change requires a method black-box contract update.",
        "- Any report output change requires report formula/workbook rule verification.",
        "- Any DB subset change requires mapping, precision, and SQL type review.",
        "- Treat all unresolved gaps as generation blockers for runnable CMBX output.",
        "",
        "## Selected Row Bindings",
        "",
        "| Test | Injection | Method | Processing | Report Template | Report Sheets | DB Fields | Coverage |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for item in selected_records:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    item.td_test,
                    _device_injection(item, device_model) or item.injection or "(open verification)",
                    _device_instrument_method(item, device_model) or item.instrument_method or "(open verification)",
                    _device_processing_method(item, device_model) or item.processing_method or "(open verification)",
                    _device_report_template(item, device_model) or item.report_template or "(open verification)",
                    ", ".join(_device_report_sheets(item, device_model) or item.report_sheets) or "(not mapped)",
                    ", ".join(_device_db_fields(item, device_model) or item.db_fields) or "(not mapped)",
                    item.coverage_status,
                )
            )
            + " |"
        )
    return "\n".join(lines).rstrip() + "\n"


def _generic_config_contract_markdown(
    record: FoqAlignmentRecord,
    device_model: str,
    selected_records: Iterable[FoqAlignmentRecord],
) -> str:
    selected = tuple(selected_records)
    required_config = _unique(value for item in selected for value in item.required_config)
    expected_channels = _unique(value for item in selected for value in item.expected_channels)
    expected_ret_times = _unique(value for item in selected for value in item.expected_ret_times)
    expected_audit = _unique(value for item in selected for value in item.expected_audit_properties)
    method_evidence = _unique(value for item in selected for value in item.method_evidence)
    lines = [
        f"# Config Contract - {record.family} - {record.td_test} - {device_model}",
        "",
        "## Required Configuration Evidence",
        "",
        *_markdown_bullets(required_config, empty="none recorded"),
        "",
        "## Expected Channels",
        "",
        *_markdown_bullets(expected_channels, empty="none recorded"),
        "",
        "## Expected RetTimes",
        "",
        *_markdown_bullets(expected_ret_times, empty="none recorded"),
        "",
        "## Expected Audit / Metadata Properties",
        "",
        *_markdown_bullets(expected_audit, empty="none recorded"),
        "",
        "## Method Evidence",
        "",
        *_markdown_bullets(method_evidence, empty="no decoded method evidence recorded"),
    ]
    return "\n".join(lines).rstrip() + "\n"


def _generic_config_method_report_review_markdown(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    device_model: str,
    selected_records: Iterable[FoqAlignmentRecord],
) -> str:
    selected = _unique_records((*tuple(selected_records), record))
    relationship_rules = _unique(rule for item in selected for rule in record_relationship_rule_lines(item, intent))
    open_gaps = _unique(gap for item in selected for gap in item.open_gaps)
    lines = [
        f"# Config -> Method -> Report Review - {record.family} - {record.td_test}",
        "",
        "## Review Boundary",
        "",
        "This is a review-only intent packet. It does not rewrite CMBX payloads, instrument methods, processing methods, or report templates.",
        "The review is ordered by the current generation priority: Instrument Config -> Instrument Method Script -> Report Formula.",
        "Processing Method and DB mapping are downstream checks; they are recorded here only when they affect the first three contracts.",
        "",
        "## Intent",
        "",
        "| Item | Value |",
        "|---|---|",
        f"| Device | `{_md_cell(device_model)}` |",
        f"| Intent | {_md_cell(intent or '(none)')} |",
        f"| Parameter | {_md_cell(parameter or '(none)')} |",
        f"| Selected rows | {len(selected)} |",
        "",
        "## 1. Instrument Config Manifest",
        "",
        "| Test | Required config | Required channels | Required variables / RetTimes | Status |",
        "|---|---|---|---|---|",
    ]
    for item in selected:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    item.td_test,
                    "; ".join(item.required_config) or "(none recorded)",
                    "; ".join(item.expected_channels) or "(none recorded)",
                    "; ".join(item.expected_ret_times) or "(none recorded)",
                    "review required" if item.coverage_status != "complete" or item.open_gaps else "evidence-backed",
                )
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 2. Instrument Method Script Contract",
            "",
            "| Test | Injection | Method | Method evidence | Edit implication |",
            "|---|---|---|---|---|",
        ]
    )
    for item in selected:
        method = _device_instrument_method(item, device_model) or item.instrument_method
        injection = _device_injection(item, device_model) or item.injection
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    item.td_test,
                    injection or "(open verification)",
                    method or "(open verification)",
                    "; ".join(item.method_evidence) or "(no decoded method evidence recorded)",
                    _generic_method_edit_implication(item, parameter),
                )
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 3. Report Formula Chain",
            "",
            "| Test | Report template | Sheets | Formula / rule evidence | Output fields | Edit implication |",
            "|---|---|---|---|---|---|",
        ]
    )
    for item in selected:
        report_template = _device_report_template(item, device_model) or item.report_template
        sheets = _device_report_sheets(item, device_model) or item.report_sheets
        db_fields = _device_db_fields(item, device_model) or item.db_fields
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    item.td_test,
                    report_template or "(open verification)",
                    ", ".join(sheets) or "(not mapped)",
                    "; ".join(item.report_evidence) or "(not recorded)",
                    ", ".join(db_fields) or "(not mapped)",
                    _generic_report_edit_implication(item, parameter),
                )
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## Relationship / Dependency Rules",
            "",
            *_markdown_bullets(relationship_rules, empty="no structured relationship rule matched"),
            "",
            "## Open Verification",
            "",
            *_markdown_bullets(open_gaps, empty="none recorded"),
            "",
            "## Minimal Verdict",
            "",
            "| Question | Answer |",
            "|---|---|",
            "| Is this reviewable as an intent? | Yes, if the Config, Method, and Report rows above are meaningful for the selected test. |",
            "| Does this prove runnable CMBX generation? | No. Binary payload generation and CM import behavior are outside this review. |",
            "| What must be closed first? | Any missing config symbol, changed method command/RetTime semantic, or changed report formula/workbook rule. |",
            "| When do Processing/DB matter? | Processing matters before CM automation/reinsertion; DB matters before database export/upload. Neither replaces method/report proof. |",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def _generic_method_edit_implication(record: FoqAlignmentRecord, parameter: str) -> str:
    if record.test_intent == "heatup_cooldown_20_50_20":
        if parameter.strip():
            return "Changing the temperature range requires new setpoints, trigger windows, RetTime semantics, and hold-time review."
        return "Reuse the 20 -> 50 -> 20 command flow only if the range is unchanged."
    if record.test_intent == "temperature_accuracy":
        return "Changing setpoints requires ladder/RetTime/report-row narrowing."
    if record.expected_ret_times:
        return "Preserve or explicitly remap all listed RetTime anchors."
    return "Review decoded method command flow before changing this test."


def _generic_report_edit_implication(record: FoqAlignmentRecord, parameter: str) -> str:
    if record.test_intent == "heatup_cooldown_20_50_20":
        return "Preserve row-66 external endpoint timing: HeatUp = RetTime2 - RetTime1 - 2.0; CoolDown = RetTime5 - RetTime4 - 2.0."
    if record.test_intent == "temperature_accuracy":
        return "Preserve external thermometer average window and narrow summary/pass-fail to retained setpoints."
    if record.report_evidence:
        return "Trace every changed output back to report cell/formula/workbook rule."
    return "Extract report formula evidence before changing report outputs."


def _generic_report_db_contract_tsv(records: Iterable[FoqAlignmentRecord], device_model: str) -> str:
    headers = (
        "Device",
        "TestIntent",
        "Test",
        "Report Template",
        "Report Sheets",
        "Formula ID",
        "DB Field",
        "Report Evidence",
        "DB Evidence",
        "Coverage",
    )
    rows = []
    for record in records:
        report_template = _device_report_template(record, device_model) or record.report_template
        report_sheets = _device_report_sheets(record, device_model) or record.report_sheets
        db_fields = _device_db_fields(record, device_model) or record.db_fields
        node = test_knowledge_node_from_record(record)
        evidence = " | ".join(record.report_evidence)
        db_evidence = " | ".join(record.db_evidence)
        for field in db_fields or ("(not mapped)",):
            rows.append(
                (
                    device_model,
                    record.test_intent,
                    record.td_test,
                    report_template or "(open verification)",
                    ", ".join(report_sheets) or "(not mapped)",
                    node.formula_id,
                    field,
                    evidence or "(not recorded)",
                    db_evidence or "(not recorded)",
                    record.coverage_status,
                )
            )
    return "\t".join(headers) + "\n" + "\n".join("\t".join(_tsv_escape(value) for value in row) for row in rows) + "\n"


def _generic_intent_conflict_matrix_tsv(records: Iterable[FoqAlignmentRecord], device_model: str, intent: str) -> str:
    rows = build_intent_conflict_rows(records, device_model=device_model, intent=intent)
    headers = (
        "Category",
        "Aspect",
        "Values",
        "Status",
        "Impact",
        "Required Action",
    )
    return "\t".join(headers) + "\n" + "\n".join(
        "\t".join(
            _tsv_escape(value)
            for value in (
                row.category,
                row.aspect,
                row.values,
                row.status,
                row.impact,
                row.required_action,
            )
        )
        for row in rows
    ) + "\n"


def build_intent_conflict_rows(
    records: Iterable[FoqAlignmentRecord],
    device_model: str = "",
    intent: str = "",
) -> tuple[IntentConflictRow, ...]:
    selected = tuple(records)
    rows: list[IntentConflictRow] = []

    def add(
        category: str,
        aspect: str,
        values: Iterable[str],
        *,
        union_required: bool = False,
        open_verification: bool = False,
        impact: str = "",
        action: str = "",
    ) -> None:
        value_tuple = _unique(values)
        status = _generic_conflict_status(
            value_tuple,
            selected_count=len(selected),
            union_required=union_required,
            open_verification=open_verification,
        )
        rows.append(
            IntentConflictRow(
                category,
                aspect,
                "; ".join(value_tuple) or "(none recorded)",
                status,
                impact or _generic_conflict_impact(status),
                action or _generic_conflict_action(status),
            )
        )

    add("Selection", "Selected tests", (record.td_test for record in selected), impact="Defines the review scope.", action="Confirm every selected test belongs in this intent.")
    add("Coverage", "Coverage states", (record.coverage_status for record in selected), impact="Partial/open rows block runnable generation.", action="Close black-box contracts before treating the packet as runnable.")
    add("Method", "Instrument Method", (_device_instrument_method(record, device_model) or record.instrument_method for record in selected), impact="Multiple methods require cloned or sequenced method assets.", action="Review method command contracts and symbol compatibility.")
    add("Processing", "Processing Method", (_device_processing_method(record, device_model) or record.processing_method for record in selected), impact="IRC/no-integration differences can change sequence behavior.", action="Review pass actions, stop actions, and inserted injection behavior.")
    add("Report", "Report Template", (_device_report_template(record, device_model) or record.report_template for record in selected), impact="Multiple templates require report selection or merged report strategy.", action="Verify template payloads, sheets, and workbook rules.")
    add("Report", "Report Sheets", (sheet for record in selected for sheet in (_device_report_sheets(record, device_model) or record.report_sheets)), union_required=True, impact="Merged reports must preserve every required sheet.", action="Check sheet/cell/formula coverage.")
    add("DB", "DB Fields", (field for record in selected for field in (_device_db_fields(record, device_model) or record.db_fields)), union_required=True, impact="DB output must include the union of selected fields.", action="Verify mapping, display precision, and SQL types.")
    add("Config", "Required Configuration", (item for record in selected for item in record.required_config), union_required=True, impact="Target CM configuration must satisfy the union of required symbols/options.", action="Check instrument setup before method recreation.")
    add("Data", "Expected RetTimes", (item for record in selected for item in record.expected_ret_times), union_required=True, impact="Report formulas depend on all listed RetTime anchors.", action="Confirm method commands log every required RetTime.")
    add("Data", "Expected Channels", (item for record in selected for item in record.expected_channels), union_required=True, impact="Raw/report evaluation depends on these channels.", action="Confirm acquisition and CM configuration provide each channel.")
    add("Audit", "Audit Properties", (item for record in selected for item in record.expected_audit_properties), union_required=True, impact="Metadata formulas depend on these audit/precondition paths.", action="Confirm audit source availability and suffix-path matching.")
    add("Relationships", "Matched relationship rules", (rule for record in selected for rule in record_relationship_rule_lines(record, intent)), union_required=True, impact="Rules explain why rows can or cannot be cropped/merged.", action="Review dependency/resource/sequence-order impact.")
    add("Generation", "Generation readiness", (record.generation_readiness for record in selected), impact="Readiness text is review guidance, not execution proof.", action="Keep runnable generation blocked until command/report/config contracts are closed.")
    open_gaps = _unique(gap for record in selected for gap in record.open_gaps)
    add("Open Verification", "Open gaps", open_gaps, open_verification=bool(open_gaps), impact="Unclosed gaps are blockers for generated assets.", action="Resolve each gap with CMBX, CM, report, or manual review evidence.")

    return tuple(rows)


def render_intent_conflict_matrix_markdown(
    records: Iterable[FoqAlignmentRecord],
    device_model: str = "",
    intent: str = "",
) -> str:
    selected = tuple(records)
    if not selected:
        return "No alignment rows are selected."
    rows = build_intent_conflict_rows(selected, device_model=device_model, intent=intent)
    lines = [
        "# Intent Conflict Matrix",
        "",
        f"Selected rows: {len(selected)}",
        f"Device branch: {device_model or '(record default / mixed)'}",
        f"Intent: {intent or 'Search / Recommend'}",
        "",
        "| Category | Aspect | Values | Status | Impact | Required Action |",
        "|---|---|---|---|---|---|",
    ]
    for row in rows:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    row.category,
                    row.aspect,
                    row.values,
                    row.status,
                    row.impact,
                    row.required_action,
                )
            )
            + " |"
        )
    return "\n".join(lines).rstrip() + "\n"


def _generic_conflict_status(
    values: tuple[str, ...],
    *,
    selected_count: int,
    union_required: bool = False,
    open_verification: bool = False,
) -> str:
    if open_verification:
        return "open verification"
    if not values:
        return "none recorded"
    if selected_count <= 1:
        return "single-row review"
    if union_required:
        return "union required"
    if len(values) == 1:
        return "single value"
    return "multiple values - review"


def _generic_conflict_impact(status: str) -> str:
    if status == "open verification":
        return "Unclosed evidence gap blocks runnable generation."
    if status == "union required":
        return "All listed values must be preserved in any cropped or merged packet."
    if status == "multiple values - review":
        return "Selection contains divergent evidence that requires manual design review."
    if status == "single value":
        return "Selection shares one evidence value."
    if status == "single-row review":
        return "Single-row packet still requires source evidence review."
    return "No evidence was recorded in the current KB/CMBX alignment."


def _generic_conflict_action(status: str) -> str:
    if status == "open verification":
        return "Close the verification item before generation."
    if status == "union required":
        return "Carry the union into method/report/DB/config review."
    if status == "multiple values - review":
        return "Choose, split, or explicitly merge the divergent values."
    if status == "single value":
        return "Reuse only after the linked contract remains valid."
    if status == "single-row review":
        return "Review the row contract before modifying it."
    return "Add evidence or mark as not applicable."


def _generic_generation_boundary_markdown(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    device_model: str,
    selected_records: Iterable[FoqAlignmentRecord],
) -> str:
    selected = _unique_records((*tuple(selected_records), record))
    gate = record_intent_gate(record, intent, parameter, selected_records=selected)
    preview = record_intent_preview(record, intent, parameter, selected_records=selected, candidate_records=selected)
    open_gaps = _unique(gap for item in selected for gap in item.open_gaps)
    lines = [
        f"# Generation Boundary - {record.family} - {record.td_test} - {device_model}",
        "",
        "## Intent Gate",
        "",
        *_intent_gate_preview_lines(gate),
        "",
        "## Current Status",
        "",
        f"- Coverage: {record.coverage_status}",
        f"- Modifiability: {record_modifiability_summary(record)}",
        f"- Generation readiness: {record.generation_readiness or '(not recorded)'}",
        "",
        "## Intent Preview",
        "",
        "```text",
        preview,
        "```",
        "",
        "## Relationship Model Rules",
        "",
        *(_unique(rule for item in selected for rule in record_relationship_rule_lines(item, intent)) or ("- no structured relationship rule matched",)),
        "",
        "## Relationship Resolution Choices",
        "",
        *_resolution_choices_markdown_lines(relationship_resolution_choices(record, intent, selected_records=selected)),
        "",
        "## Open Verification",
        "",
        *_markdown_bullets(open_gaps, empty="none recorded"),
        "",
        "## Exit Criteria",
        "",
        "- P1 / Instrument Config: target CM configuration is checked for device identity, external devices, channels, variables, and writable symbols.",
        "- P1 / Instrument Method Script: changed commands, setpoints, waits, triggers, RetTimes, and acquisition channels are closed against method evidence.",
        "- P1 / Report Formula: changed report cells, SheetObject formulas, workbook-derived rules, display precision, and pass/fail semantics are traced to source data.",
        "- P2 / Processing Method: decode or manually review IRC/pass-action behavior before CM automation or sequence insertion use.",
        "- P2 / DB: verify fields, display precision, SQL types, and upload behavior before database export/upload.",
    ]
    return "\n".join(lines).rstrip() + "\n"


def _generic_draft_asset_packet_manifest_markdown(
    record: FoqAlignmentRecord,
    intent: str,
    parameter: str,
    device_model: str,
    files: Iterable[Path],
) -> str:
    lines = [
        f"# Generic Draft Asset Packet Manifest - {record.family} - {record.td_test}",
        "",
        "## Scope",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Device | `{_md_cell(device_model)}` |",
        f"| Intent | {_md_cell(intent or '(none)')} |",
        f"| Parameter | {_md_cell(parameter or '(none)')} |",
        f"| Source TestIntent | `{_md_cell(record.test_intent)}` |",
        "",
        "## Boundary",
        "",
        "This generic packet is review-only. It preserves the current KB/CMBX binding and points to what must be checked before any method, report, or CMBX generation work.",
        "",
        "## Core Black-Box Review",
        "",
        "`config_method_report_review.md` is the main review entry point. It follows the current P1 order: Instrument Config -> Instrument Method Script -> Report Formula.",
        "Processing Method and DB mapping are downstream checks, not proof that the method script and report formula are correct.",
        "",
        "## Relationship Resolution Choices",
        "",
        *_resolution_choices_markdown_lines(relationship_resolution_choices(record, intent, selected_records=(record,))),
        "",
        "## Files",
        "",
    ]
    for path in files:
        if path.exists():
            lines.append(f"- `{path.name}`")
    return "\n".join(lines).rstrip() + "\n"


def _tsv_escape(value: object) -> str:
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def _merge_suggestions(record: FoqAlignmentRecord, candidate_records: tuple[FoqAlignmentRecord, ...]) -> tuple[FoqAlignmentRecord, ...]:
    if not candidate_records:
        return ()
    dependency_text = "\n".join(record_dependency_graph_lines(record)).lower()
    suggestions = [
        item
        for item in candidate_records
        if item is not record
        and item.family == record.family
        and (item.test_intent.lower() in dependency_text or record.test_intent.lower() in "\n".join(record_dependency_graph_lines(item)).lower())
    ]
    return tuple(suggestions[:5])


def _prefixed_or_default(label: str, values: Iterable[str], default: str = "(none recorded)") -> tuple[str, ...]:
    value_tuple = tuple(value for value in values if value)
    if not value_tuple:
        return (f"- {label}: {default}",)
    return tuple(f"- {label}: {value}" for value in value_tuple)


def _gap_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    return tuple(f"- {gap}" for gap in record.open_gaps) if record.open_gaps else ("- none recorded",)


def _intent_preview_result(record: FoqAlignmentRecord) -> str:
    if record.coverage_status == "complete" and not record.open_gaps:
        return "Candidate is reviewable for a controlled modification preview, but generated assets still require CM validation."
    if record.coverage_status == "not applicable":
        return "Not applicable; do not generate this intent for the selected branch."
    return "Blocked for runnable generation; use this preview to identify contracts that must be closed first."


def _record_label(record: FoqAlignmentRecord) -> str:
    return f"{record.family} / {record.td_test} / {record.device_label}"


def _md_cell(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def _unique_records(records: Iterable[FoqAlignmentRecord]) -> tuple[FoqAlignmentRecord, ...]:
    result: list[FoqAlignmentRecord] = []
    seen: set[tuple[str, str, str, str]] = set()
    for record in records:
        key = (record.family, record.test_intent, record.device_label, record.injection)
        if key in seen:
            continue
        seen.add(key)
        result.append(record)
    return tuple(result)


def _shared_values(groups: Iterable[Iterable[str]]) -> tuple[str, ...]:
    sets = [set(group) for group in groups if tuple(group)]
    if not sets:
        return ()
    return tuple(sorted(set.intersection(*sets)))


@dataclass(frozen=True)
class TestKnowledgeNode:
    test_id: str
    test_name: str
    foq_section: str
    purpose: str
    acceptance_criteria: tuple[str, ...]
    injection: str
    instrument_method: str
    processing_method: str
    report_template: str
    report_sheets: tuple[str, ...]
    formula_id: str
    db_fields: tuple[str, ...]
    model_applicability: tuple[str, ...]
    dependencies: tuple[str, ...]
    expected_ret_times: tuple[str, ...]
    expected_channels: tuple[str, ...]
    expected_audit_properties: tuple[str, ...]
    required_config: tuple[str, ...]
    method_evidence: tuple[str, ...]
    report_evidence: tuple[str, ...]
    db_evidence: tuple[str, ...]
    coverage_status: str
    open_gaps: tuple[str, ...]
    irc_injected: bool
    device_bindings: tuple["TestKnowledgeNodeDeviceBinding", ...]

    @property
    def model_label(self) -> str:
        return ", ".join(self.model_applicability)

    @property
    def report_sheet_label(self) -> str:
        return ", ".join(self.report_sheets)

    @property
    def db_field_label(self) -> str:
        return ", ".join(self.db_fields)

    @property
    def device_binding_label(self) -> str:
        return "; ".join(binding.summary for binding in self.device_bindings)


@dataclass(frozen=True)
class TestKnowledgeNodeDeviceBinding:
    device_model: str
    injection: str
    instrument_method: str
    processing_method: str
    report_template: str
    report_sheets: tuple[str, ...]
    report_files: tuple[str, ...]
    db_fields: tuple[str, ...]

    @property
    def report_sheet_label(self) -> str:
        return ", ".join(self.report_sheets)

    @property
    def db_field_label(self) -> str:
        return ", ".join(self.db_fields)

    @property
    def summary(self) -> str:
        return (
            f"{self.device_model}: {self.injection or '(not bound)'} / "
            f"{self.report_template or '(not bound)'} / "
            f"{', '.join(self.report_files) or '(report file follows injection)'} / "
            f"{self.db_field_label or '(no DB fields)'}"
        )


@dataclass(frozen=True)
class TestKnowledgeNodeCoverageAudit:
    test_id: str
    test_name: str
    injection: str
    method_command_status: str
    processing_method_status: str
    report_formula_status: str
    db_field_status: str
    ret_time_status: str
    channel_status: str
    audit_property_status: str
    config_status: str
    overall_status: str
    gaps: tuple[str, ...]

    @property
    def gap_label(self) -> str:
        return "; ".join(self.gaps)


@dataclass(frozen=True)
class TestKnowledgeNodeDbMappingAudit:
    test_id: str
    test_name: str
    device_model: str
    mapping_sheet: str
    injection: str
    report_sheets: tuple[str, ...]
    mapped_report_files: tuple[str, ...]
    expected_db_fields: tuple[str, ...]
    mapped_db_fields: tuple[str, ...]
    missing_expected_fields: tuple[str, ...]
    unmapped_node_fields: tuple[str, ...]
    extra_mapped_fields: tuple[str, ...]
    report_cells: tuple[str, ...]
    value_types: tuple[str, ...]
    status: str

    @property
    def mapped_report_file_label(self) -> str:
        return ", ".join(self.mapped_report_files)

    @property
    def mapped_db_field_label(self) -> str:
        return ", ".join(self.mapped_db_fields)

    @property
    def missing_label(self) -> str:
        return ", ".join(self.missing_expected_fields)

    @property
    def extra_label(self) -> str:
        return ", ".join(self.extra_mapped_fields)


@dataclass(frozen=True)
class CrossKbMappingRow:
    test_id: str
    family: str
    foq_test_name: str
    foq_section: str
    method_name: str
    processing_method: str
    report_template: str
    report_sheets: tuple[str, ...]
    formula_id: str
    db_fields: tuple[str, ...]
    model_applicability: tuple[str, ...]
    mapping_status: str

    @property
    def report_sheet_label(self) -> str:
        return ", ".join(self.report_sheets)

    @property
    def db_field_label(self) -> str:
        return ", ".join(self.db_fields)

    @property
    def model_label(self) -> str:
        return ", ".join(self.model_applicability)


@dataclass(frozen=True)
class GenerationMethodRule:
    rule_id: str
    family: str
    rule: str
    basis: str
    evidence_status: str


@dataclass(frozen=True)
class GenerationFormulaRule:
    test_name: str
    formula_id: str
    formula: str
    parameter_sources: str
    evidence_status: str


@dataclass(frozen=True)
class GenerationTemplateRule:
    test_name: str
    model_branch: str
    report_template: str
    report_sheets: tuple[str, ...]
    evidence_status: str

    @property
    def report_sheet_label(self) -> str:
        return ", ".join(self.report_sheets)


@dataclass(frozen=True)
class CrossModuleDependencyRule:
    dependency: str
    impact: str
    evidence_status: str


@dataclass(frozen=True)
class ConfigValidationRule:
    validation_item: str
    check_method: str
    failure_handling: str
    evidence_status: str


@dataclass(frozen=True)
class CmbxGenerationStrategyKb:
    method_rules: tuple[GenerationMethodRule, ...]
    formula_rules: tuple[GenerationFormulaRule, ...]
    template_rules: tuple[GenerationTemplateRule, ...]
    cross_module_dependencies: tuple[CrossModuleDependencyRule, ...]
    config_validation_rules: tuple[ConfigValidationRule, ...]


def build_test_knowledge_nodes(records: Iterable[FoqAlignmentRecord] | None = None) -> tuple[TestKnowledgeNode, ...]:
    source = tuple(records) if records is not None else base_alignment_records()
    return tuple(test_knowledge_node_from_record(record) for record in source)


def build_tkn_coverage_audits(records: Iterable[FoqAlignmentRecord] | None = None) -> tuple[TestKnowledgeNodeCoverageAudit, ...]:
    return tuple(tkn_coverage_audit_from_node(node) for node in build_test_knowledge_nodes(records))


def build_tkn_db_mapping_audits(
    records: Iterable[FoqAlignmentRecord] | None,
    mapping_path: str | Path,
    device_models: Iterable[str] = (),
) -> tuple[TestKnowledgeNodeDbMappingAudit, ...]:
    nodes = build_test_knowledge_nodes(records)
    selected_devices = {device for device in device_models if device}
    audits: list[TestKnowledgeNodeDbMappingAudit] = []
    for node in nodes:
        for device_model in node.model_applicability:
            if selected_devices and device_model not in selected_devices:
                continue
            audits.append(tkn_db_mapping_audit_from_node(node, mapping_path, device_model))
    return tuple(audits)


def tkn_db_mapping_audit_from_node(
    node: TestKnowledgeNode,
    mapping_path: str | Path,
    device_model: str,
) -> TestKnowledgeNodeDbMappingAudit:
    binding = _device_binding_for_model(node, device_model)
    injection = binding.injection if binding else node.injection
    report_sheets = binding.report_sheets if binding else node.report_sheets
    report_files = binding.report_files if binding else ()
    expected_fields = binding.db_fields if binding else node.db_fields
    try:
        from foq_result_locations import filter_locations_for_report, locations_for_device_type

        mapping_sheet, locations = locations_for_device_type(mapping_path, device_model)
    except Exception as exc:
        return TestKnowledgeNodeDbMappingAudit(
            test_id=node.test_id,
            test_name=node.test_name,
            device_model=device_model,
            mapping_sheet="",
            injection=injection,
            report_sheets=report_sheets,
            mapped_report_files=(),
            expected_db_fields=expected_fields,
            mapped_db_fields=(),
            missing_expected_fields=expected_fields,
            unmapped_node_fields=expected_fields,
            extra_mapped_fields=(),
            report_cells=(),
            value_types=(),
            status=f"mapping unavailable: {exc}",
        )

    exact_rows = _mapping_rows_for_binding(locations, injection, report_sheets, report_files, exact_injection=True)
    rows = exact_rows if injection else _mapping_rows_for_binding(locations, injection, report_sheets, report_files, exact_injection=False)
    mapped_fields = _unique(row.db_field for row in rows)
    missing_expected = tuple(field for field in expected_fields if _normalize_name(field) not in {_normalize_name(item) for item in mapped_fields})
    extra_mapped = tuple(field for field in mapped_fields if _normalize_name(field) not in {_normalize_name(item) for item in expected_fields})
    status = _db_mapping_audit_status(expected_fields, rows, missing_expected, extra_mapped)
    return TestKnowledgeNodeDbMappingAudit(
        test_id=node.test_id,
        test_name=node.test_name,
        device_model=device_model,
        mapping_sheet=mapping_sheet,
        injection=injection,
        report_sheets=report_sheets,
        mapped_report_files=_unique(row.report_file for row in rows),
        expected_db_fields=expected_fields,
        mapped_db_fields=mapped_fields,
        missing_expected_fields=missing_expected,
        unmapped_node_fields=missing_expected,
        extra_mapped_fields=extra_mapped,
        report_cells=_unique(f"{row.report_sheet}!{row.report_cell}" for row in rows),
        value_types=_unique(f"{row.db_field}: {row.value_type or 'value'} / {row.unit or 'unitless'}" for row in rows),
        status=status,
    )


def tkn_coverage_audit_from_node(node: TestKnowledgeNode) -> TestKnowledgeNodeCoverageAudit:
    method_status = _component_status(
        bool(node.instrument_method),
        bool(node.method_evidence),
        "bound to instrument method evidence",
        "method name bound; decoded command evidence open",
    )
    processing_status = _processing_component_status(node)
    report_status = _report_formula_component_status(node)
    db_status = _db_component_status(node)
    ret_time_status = _dependency_component_status(node.expected_ret_times, node.method_evidence, "RetTime")
    channel_status = _dependency_component_status(node.expected_channels, node.method_evidence + node.report_evidence, "channel")
    audit_status = _dependency_component_status(node.expected_audit_properties, node.report_evidence + node.db_evidence, "audit/property")
    config_status = _dependency_component_status(node.required_config, node.method_evidence + node.report_evidence, "configuration")
    generated_gaps = _generated_coverage_gaps(
        node,
        method_status,
        processing_status,
        report_status,
        db_status,
        ret_time_status,
        channel_status,
        audit_status,
        config_status,
    )
    gaps = _unique((*node.open_gaps, *generated_gaps))
    overall = _overall_audit_status(node.coverage_status, gaps)
    return TestKnowledgeNodeCoverageAudit(
        test_id=node.test_id,
        test_name=node.test_name,
        injection=node.injection,
        method_command_status=method_status,
        processing_method_status=processing_status,
        report_formula_status=report_status,
        db_field_status=db_status,
        ret_time_status=ret_time_status,
        channel_status=channel_status,
        audit_property_status=audit_status,
        config_status=config_status,
        overall_status=overall,
        gaps=gaps,
    )


def test_knowledge_node_from_record(record: FoqAlignmentRecord) -> TestKnowledgeNode:
    return TestKnowledgeNode(
        test_id=_test_node_id(record),
        test_name=record.td_test,
        foq_section=_foq_section(record),
        purpose=record.td_meaning,
        acceptance_criteria=_acceptance_criteria(record),
        injection=record.injection,
        instrument_method=record.instrument_method,
        processing_method=record.processing_method,
        report_template=record.report_template,
        report_sheets=record.report_sheets,
        formula_id=_formula_id(record),
        db_fields=_node_db_fields(record),
        model_applicability=record.device_models,
        dependencies=_dependencies(record),
        expected_ret_times=record.expected_ret_times,
        expected_channels=record.expected_channels,
        expected_audit_properties=record.expected_audit_properties,
        required_config=record.required_config,
        method_evidence=record.method_evidence,
        report_evidence=record.report_evidence,
        db_evidence=record.db_evidence,
        coverage_status=record.coverage_status,
        open_gaps=record.open_gaps,
        irc_injected=_is_irc_processing(record.processing_method),
        device_bindings=_device_bindings_from_record(record),
    )


def build_cross_kb_mapping_rows(records: Iterable[FoqAlignmentRecord] | None = None) -> tuple[CrossKbMappingRow, ...]:
    source = tuple(records) if records is not None else base_alignment_records()
    return tuple(cross_kb_mapping_from_record(record) for record in source)


def cross_kb_mapping_from_record(record: FoqAlignmentRecord) -> CrossKbMappingRow:
    node = test_knowledge_node_from_record(record)
    return CrossKbMappingRow(
        test_id=node.test_id,
        family=record.family,
        foq_test_name=record.td_test,
        foq_section=node.foq_section,
        method_name=record.instrument_method,
        processing_method=record.processing_method,
        report_template=record.report_template,
        report_sheets=record.report_sheets,
        formula_id=node.formula_id,
        db_fields=node.db_fields,
        model_applicability=record.device_models,
        mapping_status=_cross_kb_mapping_status(record),
    )


def build_cmbx_generation_strategy_kb() -> CmbxGenerationStrategyKb:
    return CmbxGenerationStrategyKb(
        method_rules=(
            GenerationMethodRule(
                "TCC_METH_01",
                "TCC",
                "Temperature Accuracy and Stability branches that use corrective processing must preserve IRC/corrective processing bindings.",
                "FOQ KB + CMBX sequence links: ACCURACY_IRC_STOP_H and CORRECT_* processing methods are bound to relevant injections.",
                "verified for TCC reference CMBX; exact pass action still needs deeper processing-method decode",
            ),
            GenerationMethodRule(
                "TCC_METH_02",
                "TCC",
                "If target model is VH-C10-A, include the PCC stability/performance branch.",
                "FOQ KB comparison + TCC alignment: VH maps Temperature Stability and PCC to TEMPERATURE_STABILITY_AND_PCC_70_H.",
                "verified for TCC alignment; generation command details still require method-flow validation",
            ),
            GenerationMethodRule(
                "TCC_METH_03",
                "TCC",
                "External thermometer channels must exist in the CM instrument configuration before temperature tests are runnable.",
                "FOQ KB key conditions + report formulas depend on ExtTemp_UpperCC and ExtTemp_LowerCC.",
                "verified as required symbol evidence; live CM config check still manual",
            ),
            GenerationMethodRule(
                "VDAD_METH_01",
                "VDAD",
                "Slit-width-specific tests apply only to VDAD-F where adjustable slit behavior exists.",
                "VDAD FOQ KB comparison analysis.",
                "knowledge-only; reference CMBX binding not decoded yet",
            ),
            GenerationMethodRule(
                "VDAD_METH_02",
                "VDAD",
                "3D Field tests are not applicable to VMWD-C.",
                "VDAD FOQ KB process overview and model applicability.",
                "knowledge-only; represented as not applicable in alignment",
            ),
            GenerationMethodRule(
                "VDAD_METH_03",
                "VDAD",
                "Noise testing must distinguish Diagnostic Cell and Fluidic Flow Cell evidence.",
                "VDAD FOQ KB test flow overview and troubleshooting logic.",
                "knowledge-only; reference report/method evidence not decoded yet",
            ),
        ),
        formula_rules=(
            GenerationFormulaRule(
                "TCC Temperature Accuracy",
                "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION",
                "ObservedDeviation = Observed - Nominal; result uses the larger absolute deviation from upper/lower external thermometers.",
                "Observed: ExtTemp_LowerCC and ExtTemp_UpperCC average over RetTimeN-1.0..RetTimeN-0.2; Nominal: ColumnComp.CC.Temperature.Nominal.",
                "verified in TCC evaluator",
            ),
            GenerationFormulaRule(
                "TCC Temperature Stability",
                "FORMULA_TCC_TEMP_STABILITY_SEPARATE_SENSOR_RANGE",
                "RawStability = max(max(Lower)-min(Lower), max(Upper)-min(Upper)).",
                "Lower/upper ranges from external thermometer raw channels in the report window.",
                "verified in TCC evaluator",
            ),
            GenerationFormulaRule(
                "TCC HeatUp/CoolDown",
                "FORMULA_TCC_HEATUP_COOLDOWN_RETTIME_DELTA_MINUS_HOLD",
                "Heat-up and cool-down durations are RetTime deltas minus the 2.0 min stable hold.",
                "HeatUp: RetTime2 - RetTime1 - 2.0; CoolDown: RetTime5 - RetTime4 - 2.0. RetTime3/6 remain visible internal endpoint evidence in the report layout.",
                "verified in TCC evaluator",
            ),
            GenerationFormulaRule(
                "VDAD Noise",
                "FORMULA_VDAD_NOISE_REGRESSION_DEVIATION_OPEN",
                "Average of interval maximum deviations from regression line.",
                "Expected signal source is UV/VIS detector channel such as UV_VIS_1; exact channel and report formula need VDAD CMBX evidence.",
                "open verification",
            ),
            GenerationFormulaRule(
                "VDAD Linearity",
                "FORMULA_VDAD_LINEARITY_REGRESSION_OPEN",
                "Regression coefficient r must satisfy the FOQ linearity limit.",
                "Peak area or height versus concentration regression; exact integration/report source requires processing/report decode.",
                "open verification",
            ),
        ),
        template_rules=(
            GenerationTemplateRule(
                "TCC Temperature Stability",
                "VH-C10-A",
                "Report_VTCC_V2_12",
                ("Temp Stability_Noise", "PCC"),
                "verified alignment; conceptual alias TCC_Stability_PCC.rpt remains unverified",
            ),
            GenerationTemplateRule(
                "TCC Temperature Stability",
                "VC-C10-A / VA-C10-A",
                "Report_VTCC_V2_12 for VC; Report_VATCC_V1_01 for VA",
                ("Temp Stability_Noise",),
                "partial; VC/VA no-PCC branch needs loaded CMBX confirmation",
            ),
            GenerationTemplateRule(
                "TCC Temperature Accuracy",
                "VH/VC/VA C10-A",
                "Report_VTCC_V2_12 for VC/VH; Report_VATCC_V1_01 for VA",
                ("Temp Accuracy",),
                "verified for VH/VC family, VA template named from evidence but needs full formula trace",
            ),
            GenerationTemplateRule(
                "VDAD Noise",
                "VDAD-F narrow slit",
                "VDAD_Noise_Narrow.rpt",
                ("Noise",),
                "strategy placeholder; not verified from CMBX",
            ),
            GenerationTemplateRule(
                "VDAD Noise",
                "VDAD-C / VMWD-C wide slit",
                "VDAD_Noise_Wide.rpt",
                ("Noise",),
                "strategy placeholder; not verified from CMBX",
            ),
        ),
        cross_module_dependencies=(
            CrossModuleDependencyRule(
                "TCC temperature stability -> VDAD Noise",
                "TCC temperature fluctuation can change mobile-phase refractive index and contribute to detector baseline noise.",
                "knowledge hypothesis from module physics; needs system-level validation",
            ),
            CrossModuleDependencyRule(
                "Pump flow pulsation -> VDAD Noise",
                "Pump pulsation can superimpose periodic noise on detector baseline.",
                "knowledge hypothesis from VDAD troubleshooting; needs pump evidence in system CMBX",
            ),
            CrossModuleDependencyRule(
                "TCC temperature accuracy -> VDAD wavelength accuracy",
                "Temperature can affect optical/mechanical stability and therefore wavelength accuracy interpretation.",
                "open verification; likely secondary effect and not direct method dependency",
            ),
        ),
        config_validation_rules=(
            ConfigValidationRule(
                "IRC assignment",
                "Check sequence injection processing method link and processing-method pass/action semantics.",
                "Reassign the correct processing method or stop generation until confirmed.",
                "sequence link verified; pass/action decode still partial",
            ),
            ConfigValidationRule(
                "Debug/raw diagnostic channels enabled",
                "Check required channels in CMBX evidence and later in live CM configuration.",
                "Enable required acquisition/debug configuration before running generated method.",
                "partial; CMBX evidence available, live CM check manual",
            ),
            ConfigValidationRule(
                "External thermometers exist",
                "Check required symbols/channels: ExtTemp_UpperCC and ExtTemp_LowerCC.",
                "Add/configure Generic Device thermometer channels before running TCC temperature tests.",
                "verified as TCC method/report dependency",
            ),
        ),
    )


def cmbx_generation_strategy_markdown(kb: CmbxGenerationStrategyKb | None = None) -> str:
    kb = kb or build_cmbx_generation_strategy_kb()
    lines = [
        "# CMBX Generation Strategy KB",
        "",
        "This KB connects FOQ intent, CMBX execution evidence, report formulas, and generation guardrails.",
        "It is a strategy layer, not proof that a generated method/report/CMBX is runnable.",
        "",
        "## 1. Method Generation Rules",
        "",
        "| Rule ID | Family | Rule | Basis | Evidence Status |",
        "|---|---|---|---|---|",
    ]
    for rule in kb.method_rules:
        lines.append(f"| {rule.rule_id} | {rule.family} | {rule.rule} | {rule.basis} | {rule.evidence_status} |")
    lines.extend(
        [
            "",
            "## 2. Report Formula Mapping",
            "",
            "| Test | Formula ID | Formula | Parameter Sources | Evidence Status |",
            "|---|---|---|---|---|",
        ]
    )
    for rule in kb.formula_rules:
        lines.append(f"| {rule.test_name} | `{rule.formula_id}` | `{rule.formula}` | {rule.parameter_sources} | {rule.evidence_status} |")
    lines.extend(
        [
            "",
            "## 3. Report Template Selection Rules",
            "",
            "| Test | Model Branch | Report Template | Report Sheets | Evidence Status |",
            "|---|---|---|---|---|",
        ]
    )
    for rule in kb.template_rules:
        lines.append(
            f"| {rule.test_name} | {rule.model_branch} | `{rule.report_template}` | {rule.report_sheet_label} | {rule.evidence_status} |"
        )
    lines.extend(
        [
            "",
            "## 4. Cross-Module Test Dependencies",
            "",
            "| Dependency | Impact | Evidence Status |",
            "|---|---|---|",
        ]
    )
    for rule in kb.cross_module_dependencies:
        lines.append(f"| {rule.dependency} | {rule.impact} | {rule.evidence_status} |")
    lines.extend(
        [
            "",
            "## 5. Configuration Validation Rules",
            "",
            "| Validation Item | Check Method | Failure Handling | Evidence Status |",
            "|---|---|---|---|",
        ]
    )
    for rule in kb.config_validation_rules:
        lines.append(f"| {rule.validation_item} | {rule.check_method} | {rule.failure_handling} | {rule.evidence_status} |")
    lines.extend(
        [
            "",
            "## Generation Guardrail",
            "",
            "Rows marked `open verification`, `partial`, or `strategy placeholder` must not be used as runnable generation rules until a reference CMBX method/report/formula trace confirms them.",
            "",
        ]
    )
    return "\n".join(lines)


def write_cmbx_generation_strategy_markdown(output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(cmbx_generation_strategy_markdown(), encoding="utf-8")
    return output_path


def test_knowledge_nodes_markdown(
    records: Iterable[FoqAlignmentRecord] | None = None,
    title: str = "TCC Test Knowledge Nodes",
) -> str:
    record_tuple = tuple(records) if records is not None else base_alignment_records()
    nodes = build_test_knowledge_nodes(record_tuple)
    lines = [
        f"# {title}",
        "",
        "This document renders the Test Knowledge Node contract used by the FOQ Knowledge Alignment workbench.",
        "Each node connects FOQ intent to CMBX execution evidence, report/formula evidence, and DB output fields.",
        "",
        "## Node Index",
        "",
        "| Test ID | Test | Injection | Instrument Method | Processing Method | Report Sheets | Formula ID | DB Fields | Coverage |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for node in nodes:
        lines.append(
            f"| {node.test_id} | {node.test_name} | {node.injection or '(not bound)'} | "
            f"{node.instrument_method or '(not bound)'} | {node.processing_method or '(not bound)'} | "
            f"{node.report_sheet_label or '(not mapped)'} | `{node.formula_id}` | "
            f"{node.db_field_label or '(not mapped)'} | {node.coverage_status} |"
        )
    lines.extend(
        [
            "",
            "## Coverage Audit",
            "",
            "| Test ID | Method | Processing | Report/Formula | DB | RetTimes | Channels | Audit/Metadata | Config | Overall | Gaps |",
            "|---|---|---|---|---|---|---|---|---|---|---|",
        ]
    )
    for audit in build_tkn_coverage_audits(record_tuple):
        lines.append(
            f"| {audit.test_id} | {audit.method_command_status} | {audit.processing_method_status} | "
            f"{audit.report_formula_status} | {audit.db_field_status} | {audit.ret_time_status} | "
            f"{audit.channel_status} | {audit.audit_property_status} | {audit.config_status} | "
            f"{audit.overall_status} | {audit.gap_label or 'none'} |"
        )
    default_mapping_path = Path(__file__).resolve().parents[1] / "foq" / "FOQResultLocations_V2.83.xls"
    lines.extend(
        [
            "",
            "## DB Mapping Audit",
            "",
            f"Mapping source: `{default_mapping_path}`",
            "",
            "| Test ID | Device | Injection | Mapped Report File(s) | Expected DB Fields | Mapped DB Fields | Missing Expected | Extra Mapped | Status |",
            "|---|---|---|---|---|---|---|---|---|",
        ]
    )
    if default_mapping_path.exists():
        for audit in build_tkn_db_mapping_audits(record_tuple, default_mapping_path):
            lines.append(
                f"| {audit.test_id} | {audit.device_model} | {audit.injection or '(not bound)'} | "
                f"{audit.mapped_report_file_label or '(none)'} | {', '.join(audit.expected_db_fields) or '(none)'} | "
                f"{audit.mapped_db_field_label or '(none)'} | {audit.missing_label or '(none)'} | "
                f"{audit.extra_label or '(none)'} | {audit.status} |"
            )
    else:
        lines.append(f"| n/a | n/a | n/a | n/a | n/a | n/a | n/a | n/a | mapping file not found |")
    lines.extend(["", "## Node Contracts", ""])
    for node in nodes:
        lines.extend(
            [
                f"### {node.test_id} - {node.test_name}",
                "",
                f"- FOQ Section: {node.foq_section}",
                f"- Injection: {node.injection or '(not bound)'}",
                f"- Instrument Method: {node.instrument_method or '(not bound)'}",
                f"- Processing Method: {node.processing_method or '(not bound)'}",
                f"- Report Template: {node.report_template or '(not bound)'}",
                f"- Report Sheets: {node.report_sheet_label or '(not mapped)'}",
                f"- Formula ID: `{node.formula_id}`",
                f"- DB Fields: {node.db_field_label or '(not mapped)'}",
                f"- Model Applicability: {node.model_label}",
                f"- IRC Injected: {'Yes' if node.irc_injected else 'No'}",
                f"- Coverage: {node.coverage_status}",
                "",
                "**Device Bindings**",
                "",
                "| Device | Injection | Instrument Method | Processing Method | Report Template | Report Sheets | Report Files | DB Fields |",
                "|---|---|---|---|---|---|---|---|",
                *[
                    (
                        f"| {binding.device_model} | {binding.injection or '(not bound)'} | "
                        f"{binding.instrument_method or '(not bound)'} | {binding.processing_method or '(not bound)'} | "
                        f"{binding.report_template or '(not bound)'} | {binding.report_sheet_label or '(not mapped)'} | "
                        f"{', '.join(binding.report_files) or '(follows injection)'} | "
                        f"{binding.db_field_label or '(not mapped)'} |"
                    )
                    for binding in node.device_bindings
                ],
                "",
                "**Purpose**",
                "",
                node.purpose,
                "",
                "**Acceptance Criteria**",
                "",
                *_markdown_bullets(node.acceptance_criteria),
                "",
                "**Method Command Contract**",
                "",
                "Expected RetTimes:",
                *_markdown_bullets(node.expected_ret_times),
                "",
                "Expected Channels:",
                *_markdown_bullets(node.expected_channels),
                "",
                "Expected Audit / Metadata Properties:",
                *_markdown_bullets(node.expected_audit_properties),
                "",
                "Required Configuration:",
                *_markdown_bullets(node.required_config),
                "",
                "Method Evidence:",
                *_markdown_bullets(node.method_evidence),
                "",
                "**Report / DB Contract**",
                "",
                "Report Evidence:",
                *_markdown_bullets(node.report_evidence),
                "",
                "DB Evidence:",
                *_markdown_bullets(node.db_evidence),
                "",
                "**Open Gaps**",
                "",
                *_markdown_bullets(node.open_gaps),
                "",
            ]
        )
    return "\n".join(lines)


def write_test_knowledge_nodes_markdown(
    output_path: str | Path,
    records: Iterable[FoqAlignmentRecord] | None = None,
    title: str = "TCC Test Knowledge Nodes",
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(test_knowledge_nodes_markdown(records, title=title), encoding="utf-8")
    return output_path


def family_options(records: Iterable[FoqAlignmentRecord] | None = None) -> tuple[str, ...]:
    source = tuple(records) if records is not None else base_alignment_records()
    return tuple(sorted({record.family for record in source}))


def device_options(records: Iterable[FoqAlignmentRecord]) -> tuple[str, ...]:
    devices: set[str] = set()
    for record in records:
        devices.update(record.device_models)
    return tuple(sorted(devices))


def test_intent_options(records: Iterable[FoqAlignmentRecord]) -> tuple[str, ...]:
    return tuple(sorted({record.test_intent for record in records}))


def build_foq_alignment_records(
    packages: Iterable[Any] = (),
    kb_root: str | Path = DEFAULT_KB_ROOT,
) -> tuple[FoqAlignmentRecord, ...]:
    records = _with_kb_file_status(base_alignment_records(), Path(kb_root))
    return tuple(_enrich_with_loaded_packages(record, packages) for record in records)


def filter_alignment_records(
    records: Iterable[FoqAlignmentRecord],
    family: str = "All",
    devices: Iterable[str] = (),
    test_text: str = "",
) -> tuple[FoqAlignmentRecord, ...]:
    selected_devices = {device for device in devices if device}
    needle = test_text.strip().lower()
    result: list[FoqAlignmentRecord] = []
    for record in records:
        if family and family != "All" and record.family != family:
            continue
        if selected_devices and not selected_devices.intersection(record.device_models):
            continue
        if needle and needle not in record.test_intent.lower() and needle not in record.td_test.lower():
            continue
        result.append(record)
    return tuple(result)


def build_tcc_black_box_coverage_rows(docs_root: str | Path | None = None) -> tuple[TccBlackBoxCoverageRow, ...]:
    root = Path(docs_root) if docs_root is not None else Path(__file__).resolve().parent / "docs"
    rows: list[TccBlackBoxCoverageRow] = []
    expected = (
        ("M1", "TCC_ACC_01", "Temperature Accuracy", "TCC_ACCURACY_BLACK_BOX_DECOMPOSITION.md"),
        ("M2", "TCC_CAL_01", "Temperature Calibration", "TCC_CALIBRATION_BLACK_BOX_DECOMPOSITION.md"),
        ("M2", "TCC_PRECISION_01", "Temperature Precision / Fan", "TCC_PRECISION_FAN_BLACK_BOX_DECOMPOSITION.md"),
        ("M2", "TCC_STABILITY_01", "Temperature Stability / PCC", "TCC_STABILITY_BLACK_BOX_DECOMPOSITION.md"),
        ("M2", "TCC_HEATCOOL_01", "HeatUp / CoolDown", "TCC_HEATUP_COOLDOWN_BLACK_BOX_DECOMPOSITION.md"),
        ("M2", "TCC_BURNIN_01", "BurnIn", "TCC_BURNIN_BLACK_BOX_DECOMPOSITION.md"),
        ("M3", "TCC_PREHEATER_01", "Preheater Connection", "TCC_PREHEATER_BLACK_BOX_DECOMPOSITION.md"),
        ("M3", "TCC_COL_01", "Column ID", "TCC_COLUMN_ID_BLACK_BOX_DECOMPOSITION.md"),
        ("M3", "TCC_VALVE_01", "Valve / Keypad", "TCC_VALVE_KEYPAD_BLACK_BOX_DECOMPOSITION.md"),
        ("M3", "TCC_LEAK_01", "Liquid Leak / Keypad", "TCC_LIQUID_LEAK_BLACK_BOX_DECOMPOSITION.md"),
        ("M3", "TCC_SERVICE_01", "Qualification Service", "TCC_QUALIFICATION_SERVICE_BLACK_BOX_DECOMPOSITION.md"),
        ("M3", "TCC_FACTORY_01", "Factory Default", "TCC_FACTORY_DEFAULT_BLACK_BOX_DECOMPOSITION.md"),
        ("M3", "TCC_ERRORLOG_01", "Error Log Check", "TCC_ERROR_LOG_BLACK_BOX_DECOMPOSITION.md"),
    )
    for milestone, test_id, test_name, filename in expected:
        path = root / filename
        text = path.read_text(encoding="utf-8") if path.exists() else ""
        contracts = _black_box_contract_presence(text)
        open_verification = "Open Verification" in text
        open_verification_count = _black_box_open_verification_count(text)
        open_verification_topics = _black_box_open_verification_topics(text)
        evidence_sources_present = "Evidence Sources" in text
        model_branches = _black_box_model_branch_label(text)
        mermaid_present = "```mermaid" in text
        word_count = _black_box_word_count(text)
        if not path.exists():
            status = "missing document"
        elif all(contracts) and open_verification:
            status = "documented with open verification"
        elif all(contracts):
            status = "documented"
        else:
            status = "incomplete contract headings"
        rows.append(
            TccBlackBoxCoverageRow(
                milestone,
                test_id,
                test_name,
                filename,
                path.exists(),
                contracts[0],
                contracts[1],
                contracts[2],
                contracts[3],
                contracts[4],
                contracts[5],
                open_verification,
                open_verification_count,
                open_verification_topics,
                evidence_sources_present,
                model_branches,
                mermaid_present,
                word_count,
                status,
            )
        )
    return tuple(rows)


def tcc_black_box_coverage_for_record(
    record: FoqAlignmentRecord,
    docs_root: str | Path | None = None,
) -> TccBlackBoxCoverageRow | None:
    if record.family != "TCC":
        return None
    test_id = test_knowledge_node_from_record(record).test_id
    return next((row for row in build_tcc_black_box_coverage_rows(docs_root) if row.test_id == test_id), None)


def build_tcc_open_verification_topic_rows(
    docs_root: str | Path | None = None,
    milestone: str = "",
) -> tuple[TccOpenVerificationTopicRow, ...]:
    rows: list[TccOpenVerificationTopicRow] = []
    milestone_filter = milestone.strip()
    for coverage in build_tcc_black_box_coverage_rows(docs_root):
        if milestone_filter and coverage.milestone != milestone_filter:
            continue
        for topic in coverage.open_verification_topics:
            category = _open_verification_topic_category(topic)
            rows.append(
                TccOpenVerificationTopicRow(
                    coverage.milestone,
                    coverage.test_id,
                    coverage.test_name,
                    coverage.document,
                    category,
                    topic,
                    _open_verification_likely_evidence_source(category, topic),
                    _open_verification_closure_action(category, topic),
                )
            )
    return tuple(rows)


def open_verification_topics_for_record(
    record: FoqAlignmentRecord,
    docs_root: str | Path | None = None,
) -> tuple[TccOpenVerificationTopicRow, ...]:
    coverage = tcc_black_box_coverage_for_record(record, docs_root)
    if coverage is None:
        return ()
    return tuple(row for row in build_tcc_open_verification_topic_rows(docs_root) if row.test_id == coverage.test_id)


def record_black_box_audit_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    coverage = tcc_black_box_coverage_for_record(record)
    if coverage is None:
        return (
            "No TCC black-box decomposition audit is mapped for this row.",
            "For non-TCC rows, complete module-specific method/report black-box decomposition before enabling generation tooling.",
        )
    contract_rows = (
        ("Contract 1 Method Command", coverage.contract_1_method),
        ("Contract 2 Processing Method", coverage.contract_2_processing),
        ("Contract 3 Report Formula", coverage.contract_3_report),
        ("Contract 4 DB Contract", coverage.contract_4_db),
        ("Contract 5 Config Requirement", coverage.contract_5_config),
        ("Contract 6 Open Verification", coverage.contract_6_open_verification),
    )
    topic_rows = open_verification_topics_for_record(record)
    topic_lines = tuple(
        f"- [{row.category}] {row.topic} | Evidence: {row.likely_evidence_source} | Close by: {row.closure_action}"
        for row in topic_rows
    ) or ("- no parsed open-verification topics",)
    return (
        f"Milestone: {coverage.milestone}",
        f"Test ID: {coverage.test_id}",
        f"Test: {coverage.test_name}",
        f"Document: {coverage.document}",
        f"Status: {coverage.status}",
        "",
        "Six-contract audit:",
        *(f"- {label}: {'Yes' if present else 'No'}" for label, present in contract_rows),
        "",
        "Evidence audit:",
        f"- Document exists: {'Yes' if coverage.exists else 'No'}",
        f"- Evidence Sources section: {'Yes' if coverage.evidence_sources_present else 'No'}",
        f"- Model branches mentioned: {coverage.model_branches}",
        f"- Mermaid / flow evidence: {'Yes' if coverage.mermaid_present else 'No'}",
        f"- Open Verification count: {coverage.open_verification_count}",
        f"- Open Verification topics: {'; '.join(coverage.open_verification_topics) if coverage.open_verification_topics else '(none parsed)'}",
        f"- Approximate word count: {coverage.word_count}",
        "",
        "Open Verification closure queue:",
        *topic_lines,
        "",
        "Generation implication:",
        "- This audit does not prove runnable generation.",
        "- Missing contracts, missing evidence sources, or open verification items should keep generated CMBX assets in review-only mode.",
        "- Closing an item means updating the black-box decomposition document and re-running the Alignment export.",
    )


def record_open_verification_topic_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    rows = open_verification_topics_for_record(record)
    if not rows:
        return (
            "No parsed black-box open-verification topics are mapped for this row.",
            "If generation is still blocked, check the TKN Coverage Audit and source black-box decomposition.",
        )
    lines = [
        "Open Verification Topic Queue",
        "",
        "Each item below is a closure task before runnable generation can be claimed.",
        "",
    ]
    for index, row in enumerate(rows, start=1):
        lines.extend(
            [
                f"{index}. [{row.category}] {row.topic}",
                f"   Test: {row.test_id} / {row.test_name}",
                f"   Evidence source: {row.likely_evidence_source}",
                f"   Closure action: {row.closure_action}",
                f"   Source document: {row.document}",
                "",
            ]
        )
    return tuple(lines)


def record_milestone_status_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "Milestone status is currently implemented for TCC alignment rows.",
            "VDAD and other module milestones should be added after their black-box decomposition model is established.",
        )
    coverage = tcc_black_box_coverage_for_record(record)
    rows = build_tcc_milestone_status_rows(base_alignment_records())
    lines = [
        "TCC Knowledge Engineering Milestone Status",
        "",
        f"Current Test: {record.td_test}",
        f"Current Test ID: {test_knowledge_node_from_record(record).test_id}",
        f"Current Milestone: {coverage.milestone if coverage else '(not mapped)'}",
        f"Current BlackBox Document: {coverage.document if coverage else '(not mapped)'}",
        "",
        "| Milestone | Objective | Status | Evidence | Open Work |",
        "|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if coverage and row.milestone == coverage.milestone else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    f"{row.milestone}{marker}",
                    row.objective,
                    row.status,
                    row.evidence,
                    row.open_work,
                )
            )
            + " |"
        )
    lines.extend(
        (
            "",
            "Use this view as the goal-level audit:",
            "- M1/M2/M3 prove that black-box documents exist and expose six contracts.",
            "- M4 proves relationship rules are structured and exportable.",
            "- M5 proves intent review tooling exists, while runnable CMBX generation remains closed until CM validation.",
        )
    )
    return tuple(lines)


def build_tcc_temperature_contract_matrix_rows(
    docs_root: str | Path | None = None,
) -> tuple[TccTemperatureContractMatrixRow, ...]:
    """Return M2 temperature-family contract status rows for planning tools."""
    coverage_rows = tuple(row for row in build_tcc_black_box_coverage_rows(docs_root) if row.milestone == "M2")
    topic_rows = build_tcc_open_verification_topic_rows(docs_root, milestone="M2")
    topics_by_test: dict[str, list[TccOpenVerificationTopicRow]] = {}
    for topic in topic_rows:
        topics_by_test.setdefault(topic.test_id, []).append(topic)
    rows: list[TccTemperatureContractMatrixRow] = []
    for coverage in coverage_rows:
        topics = tuple(topics_by_test.get(coverage.test_id, ()))
        categories = _unique(topic.category for topic in topics)
        actions = _unique(topic.closure_action for topic in topics)
        rows.append(
            TccTemperatureContractMatrixRow(
                coverage.test_id,
                coverage.test_name,
                coverage.document,
                _contract_status_label(coverage.contract_1_method),
                _contract_status_label(coverage.contract_2_processing),
                _contract_status_label(coverage.contract_3_report),
                _contract_status_label(coverage.contract_4_db),
                _contract_status_label(coverage.contract_5_config),
                _contract_status_label(coverage.contract_6_open_verification),
                ", ".join(categories) if categories else "(none)",
                "; ".join(actions) if actions else "No closure action recorded.",
                _temperature_template_readiness(coverage, topics),
            )
        )
    return tuple(rows)


def build_tcc_contract_closure_task_rows(
    docs_root: str | Path | None = None,
    milestone: str = "M2",
) -> tuple[TccContractClosureTaskRow, ...]:
    """Return actionable black-box closure tasks grouped by contract."""
    milestone_filter = (milestone or "").strip()
    rows: list[TccContractClosureTaskRow] = []
    for coverage in build_tcc_black_box_coverage_rows(docs_root):
        if milestone_filter and coverage.milestone != milestone_filter:
            continue
        for contract_name, present in _coverage_contract_pairs(coverage):
            if present:
                continue
            rows.append(
                TccContractClosureTaskRow(
                    coverage.milestone,
                    coverage.test_id,
                    coverage.test_name,
                    coverage.document,
                    contract_name,
                    _closure_task_priority(contract_name),
                    _closure_task_evidence_group(contract_name),
                    "Missing contract heading/evidence",
                    f"{contract_name} is not detected in {coverage.document}.",
                    _contract_likely_evidence_source(contract_name),
                    _contract_closure_action(contract_name),
                    "Yes",
                )
            )
        for topic in build_tcc_open_verification_topic_rows(docs_root, milestone=coverage.milestone):
            if topic.test_id != coverage.test_id:
                continue
            contract_name = _open_verification_category_contract(topic.category)
            rows.append(
                TccContractClosureTaskRow(
                    topic.milestone,
                    topic.test_id,
                    topic.test_name,
                    topic.document,
                    contract_name,
                    _closure_task_priority(contract_name),
                    _closure_task_evidence_group(contract_name, topic.likely_evidence_source),
                    "Open verification",
                    topic.topic,
                    topic.likely_evidence_source,
                    topic.closure_action,
                    "Yes",
                )
            )
    return tuple(sorted(rows, key=_closure_task_sort_key))


def record_contract_closure_task_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "Contract closure tasks are currently implemented for TCC black-box documents.",
            "Add module-specific coverage rules before enabling this view for other families.",
        )
    coverage = tcc_black_box_coverage_for_record(record)
    milestone = coverage.milestone if coverage else "M2"
    rows = build_tcc_contract_closure_task_rows(milestone=milestone)
    current_test_id = test_knowledge_node_from_record(record).test_id
    if not rows:
        return (
            f"No open contract closure task is recorded for milestone {milestone}.",
            "This still does not prove runnable generation; CM validation is a separate gate.",
        )
    lines = [
        f"{milestone} Contract Closure Tasks",
        "",
        "Each row is a concrete evidence task. Close it by updating the source black-box decomposition, then rerun Alignment export.",
        "",
        "| Test ID | Test | Priority | Evidence Group | Contract | Type | Topic | Evidence Source | Closure Action |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if row.test_id == current_test_id else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    f"{row.test_id}{marker}",
                    row.test_name,
                    row.priority,
                    row.evidence_group,
                    row.contract,
                    row.task_type,
                    row.topic,
                    row.likely_evidence_source,
                    row.closure_action,
                )
            )
            + " |"
        )
    return tuple(lines)


def build_tcc_next_action_queue_rows(
    docs_root: str | Path | None = None,
    milestones: Iterable[str] = ("M2", "M3"),
) -> tuple[TccNextActionQueueRow, ...]:
    """Return the global evidence queue that should drive the next work items."""
    milestone_set = {milestone for milestone in milestones if milestone}
    tasks = tuple(
        task
        for task in build_tcc_contract_closure_task_rows(docs_root, milestone="")
        if (not milestone_set or task.milestone in milestone_set) and task.generation_blocker == "Yes"
    )
    grouped: dict[tuple[str, str, str], list[TccContractClosureTaskRow]] = {}
    for task in tasks:
        grouped.setdefault((task.milestone, task.priority, task.evidence_group), []).append(task)

    rows: list[TccNextActionQueueRow] = []
    for (milestone, priority, evidence_group), group_tasks in grouped.items():
        contracts = _unique(task.contract for task in group_tasks)
        topics = _unique(task.topic for task in group_tasks)
        rows.append(
            TccNextActionQueueRow(
                0,
                milestone,
                priority,
                evidence_group,
                len(group_tasks),
                ", ".join(_unique(f"{task.test_id} / {task.test_name}" for task in group_tasks)),
                ", ".join(contracts),
                topics[0] if topics else "Open evidence task",
                _workstream_next_action(evidence_group, group_tasks),
                _workstream_unlocks(evidence_group, group_tasks),
                _next_action_generation_gate(priority, evidence_group),
            )
        )
    sorted_rows = sorted(rows, key=_next_action_sort_key)
    return tuple(replace(row, rank=index) for index, row in enumerate(sorted_rows, start=1))


def record_next_action_queue_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "Next action queue is currently implemented for TCC black-box closure tasks.",
            "Add module-specific closure-task rules before enabling this view for other families.",
        )
    rows = build_tcc_next_action_queue_rows()
    current_test_id = test_knowledge_node_from_record(record).test_id
    if not rows:
        return (
            "No global TCC next-action queue items are open.",
            "This still does not prove runnable generation; CM validation remains required.",
        )
    lines = [
        "TCC Next Action Queue",
        "",
        "This is the global evidence queue. It answers: what should be decoded next to move from review-only alignment toward reusable generation templates?",
        "",
        "| Rank | Milestone | Priority | Evidence Group | Tasks | Tests | Contracts | Primary Blocker | Next Action | Unlocks | Gate |",
        "|---:|---|---|---|---:|---|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if current_test_id in row.tests else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    row.rank,
                    row.milestone,
                    row.priority,
                    row.evidence_group,
                    row.task_count,
                    f"{row.tests}{marker}",
                    row.contracts,
                    row.primary_blocker,
                    row.next_action,
                    row.unlocks,
                    row.generation_gate,
                )
            )
            + " |"
        )
    lines.extend(
        (
            "",
            "Working rule:",
            "- Start with rank 1 unless a user intent makes another blocker more urgent.",
            "- Keep generated assets in review-only mode until all P1 queue items that affect the selected intent are closed.",
            "- Closing a queue item means updating the black-box document and rerunning this audit.",
        )
    )
    return tuple(lines)


def build_tcc_evidence_workstream_rows(
    docs_root: str | Path | None = None,
    milestone: str = "M2",
    priority: str = "",
) -> tuple[TccEvidenceWorkstreamRow, ...]:
    """Group closure tasks by evidence workstream for execution planning."""
    priority_filter = priority.strip()
    tasks = tuple(
        task
        for task in build_tcc_contract_closure_task_rows(docs_root, milestone=milestone)
        if not priority_filter or task.priority == priority_filter
    )
    grouped: dict[tuple[str, str, str], list[TccContractClosureTaskRow]] = {}
    for task in tasks:
        key = (task.milestone, task.priority, task.evidence_group)
        grouped.setdefault(key, []).append(task)
    rows: list[TccEvidenceWorkstreamRow] = []
    for (task_milestone, task_priority, evidence_group), group_tasks in grouped.items():
        rows.append(
            TccEvidenceWorkstreamRow(
                task_milestone,
                task_priority,
                evidence_group,
                len(group_tasks),
                ", ".join(_unique(f"{task.test_id} / {task.test_name}" for task in group_tasks)),
                ", ".join(_unique(task.contract for task in group_tasks)),
                "; ".join(_unique(task.likely_evidence_source for task in group_tasks)),
                "; ".join(_unique(task.closure_action for task in group_tasks)),
                _workstream_unlocks(evidence_group, group_tasks),
                _workstream_next_action(evidence_group, group_tasks),
            )
        )
    return tuple(sorted(rows, key=_workstream_sort_key))


def record_evidence_workstream_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "Evidence workstreams are currently implemented for TCC black-box closure tasks.",
            "Add module-specific task grouping before enabling this view for other families.",
        )
    coverage = tcc_black_box_coverage_for_record(record)
    milestone = coverage.milestone if coverage else "M2"
    rows = build_tcc_evidence_workstream_rows(milestone=milestone)
    if not rows:
        return (
            f"No evidence workstream is open for milestone {milestone}.",
            "This does not by itself prove runnable generation; CM validation remains required.",
        )
    lines = [
        f"{milestone} Evidence Workstreams",
        "",
        "Use this view to choose the next evidence extraction effort. P1 workstreams are the main blockers for moving from review-only to reusable generation templates.",
        "",
        "| Priority | Evidence Group | Tasks | Tests | Contracts | Unlocks | Next Action |",
        "|---|---|---:|---|---|---|---|",
    ]
    for row in rows:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    row.priority,
                    row.evidence_group,
                    row.task_count,
                    row.tests,
                    row.contracts,
                    row.unlocks,
                    row.next_action,
                )
            )
            + " |"
        )
    lines.extend(("", "Evidence sources by workstream:"))
    for row in rows:
        lines.append(f"- {row.priority} / {row.evidence_group}: {row.likely_evidence_sources}")
    return tuple(lines)


def build_tcc_p1_evidence_extraction_plan_rows(
    docs_root: str | Path | None = None,
    milestone: str = "M2",
) -> tuple[TccP1EvidenceExtractionPlanRow, ...]:
    """Expand P1 closure tasks into method/processing/report evidence extraction plans."""
    rows: list[TccP1EvidenceExtractionPlanRow] = []
    for task in build_tcc_contract_closure_task_rows(docs_root, milestone=milestone):
        if task.priority != "P1":
            continue
        rows.append(
            TccP1EvidenceExtractionPlanRow(
                task.milestone,
                task.test_id,
                task.test_name,
                task.document,
                task.contract,
                task.evidence_group,
                task.topic,
                task.likely_evidence_source,
                _p1_evidence_extraction_steps(task),
                _p1_evidence_validation_outputs(task),
                _p1_evidence_closure_update(task),
                "planned - extraction not yet executed",
            )
        )
    return tuple(rows)


def record_p1_evidence_extraction_plan_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "P1 evidence extraction plans are currently implemented for TCC M2 closure tasks.",
            "Add module-specific closure-task rules before enabling this view for other families.",
        )
    coverage = tcc_black_box_coverage_for_record(record)
    milestone = coverage.milestone if coverage else "M2"
    rows = build_tcc_p1_evidence_extraction_plan_rows(milestone=milestone)
    if not rows:
        return (
            f"No P1 extraction plan is open for milestone {milestone}.",
            "This does not prove runnable generation; it only means no P1 black-box closure task was detected.",
        )
    current_test_id = test_knowledge_node_from_record(record).test_id
    lines = [
        f"{milestone} P1 Evidence Extraction Plan",
        "",
        "This is the next-action layer for Method Command, Processing Method, and Report Formula black boxes.",
        "",
        "| Test ID | Test | Evidence Group | Contract | Topic | Status |",
        "|---|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if row.test_id == current_test_id else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    f"{row.test_id}{marker}",
                    row.test_name,
                    row.evidence_group,
                    row.contract,
                    row.topic,
                    row.status,
                )
            )
            + " |"
        )
    lines.extend(("", "Extraction steps:"))
    for row in rows:
        marker = " <- current" if row.test_id == current_test_id else ""
        lines.append(f"- {row.test_id}{marker} / {row.evidence_group}")
        lines.extend(f"  {line}" for line in row.extraction_steps.splitlines())
        lines.append(f"  Validation: {row.validation_outputs}")
        lines.append(f"  Closure update: {row.closure_update}")
    return tuple(lines)


def build_tcc_processing_method_target_rows(
    docs_root: str | Path | None = None,
    milestone: str = "M2",
    records: Iterable[FoqAlignmentRecord] = (),
) -> tuple[TccProcessingMethodTargetRow, ...]:
    """Map processing-method closure tasks onto sequence bindings and device branches."""
    record_tuple = tuple(records) or base_alignment_records()
    rows: list[TccProcessingMethodTargetRow] = []
    for task in build_tcc_contract_closure_task_rows(docs_root, milestone=milestone):
        if task.evidence_group != "Processing method decode / CM UI":
            continue
        bound_records = _records_for_closure_test_id(task.test_id, record_tuple)
        if not bound_records:
            rows.append(
                TccProcessingMethodTargetRow(
                    task.milestone,
                    task.test_id,
                    task.test_name,
                    "(not bound)",
                    "(not bound)",
                    "(not bound)",
                    "(not bound)",
                    task.topic,
                    "open - no matching alignment row",
                    "Processing method action table / IRC pass-action rows",
                    task.document,
                    task.closure_action,
                    "open - no matching alignment row",
                )
            )
            continue
        for bound in bound_records:
            for device_model in bound.device_models:
                processing_method = _device_processing_method(bound, device_model) or "(not bound)"
                rows.append(
                    TccProcessingMethodTargetRow(
                        task.milestone,
                        task.test_id,
                        task.test_name,
                        device_model,
                        _device_injection(bound, device_model) or "(not bound)",
                        _device_instrument_method(bound, device_model) or "(not bound)",
                        processing_method,
                        task.topic,
                        _processing_method_expected_behavior(processing_method),
                        _processing_method_extraction_target(processing_method, task.topic),
                        task.document,
                        task.closure_action,
                        _processing_method_target_readiness(processing_method),
                    )
                )
    return tuple(rows)


def record_processing_method_target_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "Processing method targets are currently implemented for TCC M2 closure tasks.",
            "Add module-specific processing target mapping before enabling this view for other families.",
        )
    coverage = tcc_black_box_coverage_for_record(record)
    milestone = coverage.milestone if coverage else "M2"
    rows = build_tcc_processing_method_target_rows(milestone=milestone)
    if not rows:
        return (
            f"No processing-method target is currently open for milestone {milestone}.",
            "This does not prove processing closure; it only means no processing P1 task was found.",
        )
    current_ids = _closure_test_id_aliases(test_knowledge_node_from_record(record).test_id)
    lines = [
        f"{milestone} Processing Method Targets",
        "",
        "These rows bind processing-method open verification topics to device-specific sequence rows.",
        "",
        "| Test ID | Device | Injection | Instrument Method | Processing Method | Expected Behavior | Readiness |",
        "|---|---|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if row.test_id in current_ids else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    f"{row.test_id}{marker}",
                    row.device_model,
                    row.injection,
                    row.instrument_method,
                    row.processing_method,
                    row.expected_behavior,
                    row.readiness,
                )
            )
            + " |"
        )
    lines.extend(("", "Closure actions:"))
    for row in rows:
        marker = " <- current" if row.test_id in current_ids else ""
        lines.append(f"- {row.test_id}{marker} / {row.device_model} / {row.processing_method}: {row.topic} | {row.closure_action}")
    return tuple(lines)


def build_tcc_report_formula_target_rows(
    docs_root: str | Path | None = None,
    milestone: str = "M2",
    records: Iterable[FoqAlignmentRecord] = (),
) -> tuple[TccReportFormulaTargetRow, ...]:
    """Map report-formula closure tasks onto report templates, sheets, DB fields, and formula IDs."""
    record_tuple = tuple(records) or base_alignment_records()
    rows: list[TccReportFormulaTargetRow] = []
    for task in build_tcc_contract_closure_task_rows(docs_root, milestone=milestone):
        if task.evidence_group != "Report workbook/formula extraction":
            continue
        bound_records = _records_for_closure_test_id(task.test_id, record_tuple)
        if not bound_records:
            rows.append(
                TccReportFormulaTargetRow(
                    task.milestone,
                    task.test_id,
                    task.test_name,
                    "(not bound)",
                    task.topic,
                    "(not bound)",
                    "(not mapped)",
                    "(not mapped)",
                    "FORMULA_OPEN_VERIFICATION_REQUIRED",
                    _report_formula_extraction_target(task.topic),
                    task.document,
                    task.closure_action,
                    "open - no matching alignment row",
                )
            )
            continue
        for bound in bound_records:
            node = test_knowledge_node_from_record(bound)
            for device_model in bound.device_models:
                if (
                    bound.family == "TCC"
                    and bound.test_intent == "temperature_precision_and_fan"
                    and device_model == "VA-C10-A"
                    and "fan pass/fail" in task.topic.lower()
                ):
                    continue
                report_template = _device_report_template(bound, device_model) or bound.report_template or "(not bound)"
                report_sheets = _device_report_sheets(bound, device_model) or bound.report_sheets
                db_fields = _device_db_fields(bound, device_model) or node.db_fields
                rows.append(
                    TccReportFormulaTargetRow(
                        task.milestone,
                        task.test_id,
                        task.test_name,
                        device_model,
                        task.topic,
                        report_template,
                        ", ".join(report_sheets) or "(not mapped)",
                        ", ".join(db_fields) or "(not mapped)",
                        node.formula_id,
                        _report_formula_extraction_target(task.topic),
                        task.document,
                        task.closure_action,
                        _report_formula_target_readiness(report_template, report_sheets, db_fields),
                    )
                )
    return tuple(rows)


def record_report_formula_target_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "Report formula targets are currently implemented for TCC M2 closure tasks.",
            "Add module-specific report target mapping before enabling this view for other families.",
        )
    coverage = tcc_black_box_coverage_for_record(record)
    milestone = coverage.milestone if coverage else "M2"
    rows = build_tcc_report_formula_target_rows(milestone=milestone)
    if not rows:
        return (
            f"No report-formula extraction target is currently open for milestone {milestone}.",
            "This does not prove report formula closure; it only means no parsed report-formula closure task was found.",
        )
    current_ids = _closure_test_id_aliases(test_knowledge_node_from_record(record).test_id)
    lines = [
        f"{milestone} Report Formula Extraction Targets",
        "",
        "These rows translate P1 report workstream tasks into concrete report/template targets.",
        "",
        "| Test ID | Test | Device | Report Template | Sheets | DB Fields | Formula ID | Extraction Target | Readiness |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if row.test_id in current_ids else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    f"{row.test_id}{marker}",
                    row.test_name,
                    row.device_model,
                    row.report_template,
                    row.report_sheets,
                    row.db_fields,
                    row.formula_id,
                    row.extraction_target,
                    row.readiness,
                )
            )
            + " |"
        )
    lines.extend(("", "Closure actions:"))
    for row in rows:
        marker = " <- current" if row.test_id in current_ids else ""
        lines.append(f"- {row.test_id}{marker} / {row.test_name}: {row.topic} | {row.closure_action}")
    return tuple(lines)


def build_tcc_report_formula_extraction_plan_rows(
    docs_root: str | Path | None = None,
    milestone: str = "M2",
    records: Iterable[FoqAlignmentRecord] = (),
) -> tuple[TccReportFormulaExtractionPlanRow, ...]:
    """Turn report targets into an executable extraction and validation checklist."""
    rows: list[TccReportFormulaExtractionPlanRow] = []
    for target in build_tcc_report_formula_target_rows(docs_root, milestone=milestone, records=records):
        rows.append(
            TccReportFormulaExtractionPlanRow(
                target.milestone,
                target.test_id,
                target.test_name,
                target.device_model,
                target.report_template,
                target.report_sheets,
                target.formula_id,
                target.db_fields,
                target.extraction_target,
                _report_formula_extraction_steps(target),
                _report_formula_validation_outputs(target),
                _report_formula_closure_update(target),
                "planned - extraction not yet executed",
            )
        )
    return tuple(rows)


def record_report_formula_extraction_plan_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "Report formula extraction plans are currently implemented for TCC M2 closure tasks.",
            "Add module-specific report extraction planning before enabling this view for other families.",
        )
    coverage = tcc_black_box_coverage_for_record(record)
    milestone = coverage.milestone if coverage else "M2"
    rows = build_tcc_report_formula_extraction_plan_rows(milestone=milestone)
    if not rows:
        return (
            f"No report-formula extraction plan is currently open for milestone {milestone}.",
            "This does not prove FormulaOne closure; it only means no report target was generated.",
        )
    current_ids = _closure_test_id_aliases(test_knowledge_node_from_record(record).test_id)
    lines = [
        f"{milestone} Report Formula Extraction Plan",
        "",
        "This plan converts report targets into extraction steps, validation outputs, and black-box document updates.",
        "",
        "| Test ID | Device | Template | Sheets | Extraction Target | Status |",
        "|---|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if row.test_id in current_ids else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    f"{row.test_id}{marker}",
                    row.device_model,
                    row.report_template,
                    row.report_sheets,
                    row.extraction_target,
                    row.status,
                )
            )
            + " |"
        )
    lines.extend(("", "Extraction steps:"))
    for row in rows:
        marker = " <- current" if row.test_id in current_ids else ""
        lines.append(f"- {row.test_id}{marker} / {row.device_model} / {row.report_template}")
        lines.extend(f"  {line}" for line in row.extraction_steps.splitlines())
        lines.append(f"  Validation: {row.validation_outputs}")
        lines.append(f"  Closure update: {row.closure_update}")
    return tuple(lines)


def record_temperature_contract_matrix_lines(record: FoqAlignmentRecord) -> tuple[str, ...]:
    if record.family != "TCC":
        return (
            "M2 temperature contract matrix is currently implemented for TCC only.",
            "Add module-specific black-box coverage rows before enabling this view for other families.",
        )
    rows = build_tcc_temperature_contract_matrix_rows()
    if not rows:
        return ("No M2 temperature-family contract rows are available.",)
    current_test_id = test_knowledge_node_from_record(record).test_id
    lines = [
        "M2 Temperature Contract Matrix",
        "",
        "This matrix is the planning view for temperature-family black-box closure. It does not prove runnable generation.",
        "",
        "| Test ID | Test | Method | Processing | Report | DB | Config | Open Verification | Open Topic Categories | Template Readiness |",
        "|---|---|---|---|---|---|---|---|---|---|",
    ]
    for row in rows:
        marker = " <- current" if row.test_id == current_test_id else ""
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    f"{row.test_id}{marker}",
                    row.test_name,
                    row.method_contract,
                    row.processing_contract,
                    row.report_contract,
                    row.db_contract,
                    row.config_contract,
                    row.open_verification_contract,
                    row.open_topic_categories,
                    row.template_readiness,
                )
            )
            + " |"
        )
    lines.extend(("", "Next closure actions:"))
    for row in rows:
        lines.append(f"- {row.test_id} / {row.test_name}: {row.next_closure_actions}")
    return tuple(lines)


def _contract_status_label(present: bool) -> str:
    return "present" if present else "missing"


def _coverage_contract_pairs(coverage: TccBlackBoxCoverageRow) -> tuple[tuple[str, bool], ...]:
    return (
        ("Contract 1 Method Command", coverage.contract_1_method),
        ("Contract 2 Processing Method", coverage.contract_2_processing),
        ("Contract 3 Report Formula", coverage.contract_3_report),
        ("Contract 4 DB Contract", coverage.contract_4_db),
        ("Contract 5 Config Requirement", coverage.contract_5_config),
        ("Contract 6 Open Verification", coverage.contract_6_open_verification),
    )


def _closure_task_sort_key(row: TccContractClosureTaskRow) -> tuple[int, str, str, str]:
    priority_order = {"P1": 1, "P2": 2, "P3": 3}
    priority = row.priority.split()[0]
    return (priority_order.get(priority, 9), row.evidence_group, row.test_id, row.contract)


def _closure_task_priority(contract_name: str) -> str:
    if any(token in contract_name for token in ("Method Command", "Report Formula", "Config Requirement")):
        return "P1"
    if any(token in contract_name for token in ("Processing Method", "DB Contract")):
        return "P2"
    return "P3"


def _closure_task_evidence_group(contract_name: str, evidence_source: str = "") -> str:
    text = f"{contract_name} {evidence_source}".lower()
    if "processing" in text or "procmeth" in text or "sst" in text or "irc" in text:
        return "Processing method decode / CM UI"
    if "report" in text or "formulaone" in text or "spreadsheetdata" in text or "workbook" in text or "xls" in text:
        return "Report workbook/formula extraction"
    if "db" in text or "foqresultlocations" in text or "sql" in text:
        return "DB mapping/type audit"
    if "method command" in text or "decoded instrument" in text or "method flow" in text:
        return "Method command decode"
    if "config" in text or "symbol" in text or "channel" in text or "device" in text:
        return "Configuration evidence"
    return "Black-box KB review"


def _workstream_sort_key(row: TccEvidenceWorkstreamRow) -> tuple[int, str]:
    priority_order = {"P1": 1, "P2": 2, "P3": 3}
    return (priority_order.get(row.priority, 9), row.evidence_group)


def _next_action_sort_key(row: TccNextActionQueueRow) -> tuple[int, int, int, str]:
    milestone_order = {"M2": 1, "M3": 2, "M1": 3, "M4": 4, "M5": 5}
    priority_order = {"P1": 1, "P2": 2, "P3": 3}
    evidence_order = {
        "Configuration evidence": 1,
        "Method command decode": 2,
        "Report workbook/formula extraction": 3,
        "Processing method decode / CM UI": 4,
        "DB mapping/type audit": 5,
    }
    return (
        milestone_order.get(row.milestone, 9),
        priority_order.get(row.priority, 9),
        evidence_order.get(row.evidence_group, 99),
        row.evidence_group,
    )


def _next_action_generation_gate(priority: str, evidence_group: str) -> str:
    if priority == "P1":
        return "blocks method/report generation confidence for affected intents"
    if evidence_group == "Processing method decode / CM UI":
        return "blocks CM automation fidelity but not first-pass method/report reasoning"
    if evidence_group == "DB mapping/type audit":
        return "blocks deployment/upload confidence"
    return "review before claiming reusable template"


def _workstream_unlocks(
    evidence_group: str,
    tasks: Iterable[TccContractClosureTaskRow],
) -> str:
    contracts = {task.contract for task in tasks}
    if evidence_group == "Processing method decode / CM UI":
        return "CM internal pass-action/insertion fidelity; IRC and integration side effects"
    if evidence_group == "Report workbook/formula extraction":
        return "Report formula contract; display precision; workbook-derived pass/fail"
    if evidence_group == "Method command decode":
        return "Instrument method command contract; RetTime/channel/time-condition semantics"
    if evidence_group == "DB mapping/type audit":
        return "DB upload contract; field subset/type/precision confidence"
    if evidence_group == "Configuration evidence":
        return "Instrument setup/config manifest; required external devices, variables, and channels"
    return ", ".join(sorted(contracts)) or "Black-box review readiness"


def _workstream_next_action(
    evidence_group: str,
    tasks: Iterable[TccContractClosureTaskRow],
) -> str:
    task_tuple = tuple(tasks)
    tests = ", ".join(_unique(task.test_id for task in task_tuple[:4]))
    if evidence_group == "Processing method decode / CM UI":
        return f"Defer until core config/method/report evidence is clearer; then inspect processing actions for {tests}."
    if evidence_group == "Report workbook/formula extraction":
        return f"Extract FormulaOne/workbook rules for {tests}; update Contract 3 and DB precision notes."
    if evidence_group == "Method command decode":
        return f"Add line-level command evidence for {tests}; update Contract 1 RetTime/channel semantics."
    if evidence_group == "DB mapping/type audit":
        return f"Verify mapping/type/precision for {tests}; update Contract 4."
    if evidence_group == "Configuration evidence":
        return f"Record required modules, external devices, variables, and channel imports for {tests}; update Contract 5."
    return f"Review and close {len(task_tuple)} task(s) in the source black-box documents."


def _p1_evidence_extraction_steps(task: TccContractClosureTaskRow) -> str:
    if task.evidence_group == "Processing method decode / CM UI":
        return "\n".join(
            (
                f"1. Locate the processing method bound to `{task.test_name}` in the reference CMBX or CM Associated Items.",
                "2. Export/decode the processing method action table, including SST/IRC pass-action entries.",
                "3. Capture inserted injection name, trigger condition, stop/continue behavior, and model branch if present.",
                "4. Compare the action table against TKN `irc_injected`, sequence injection order, and relationship rules.",
                "5. Record exact evidence in Contract 2 Processing Method.",
            )
        )
    if task.evidence_group == "Method command decode":
        return "\n".join(
            (
                f"1. Locate the instrument method bound to `{task.test_name}` in the reference CMBX.",
                "2. Decode/export the method command flow with method_contract.py or the Instrument Method tab.",
                "3. Capture setpoints, wait/ready gates, delays, RetTime writes, channel acquisition, and audit log commands.",
                "4. Compare decoded commands with expected RetTimes/channels/audit properties in the TKN row.",
                "5. Record line-level evidence in Contract 1 Method Command.",
            )
        )
    if task.evidence_group == "Report workbook/formula extraction":
        return "\n".join(
            (
                f"1. Locate the report template and sheet(s) bound to `{task.test_name}`.",
                "2. Extract direct ReportFormulaObject formulas and FixedChannel bindings.",
                "3. Extract SpreadSheetData / FormulaOne workbook formulas and dependency cells for DB-facing outputs.",
                "4. Compare exported report workbook cells, DB field mapping, display precision, and pass/fail cells.",
                "5. Record exact evidence in Contract 3 Report Formula and update Contract 4 if DB type/precision changes.",
            )
        )
    return "\n".join(
        (
            f"1. Inspect evidence source: {task.likely_evidence_source}.",
            f"2. Capture evidence for {task.contract}.",
            "3. Update the source black-box document and rerun the alignment audit.",
        )
    )


def _p1_evidence_validation_outputs(task: TccContractClosureTaskRow) -> str:
    if task.evidence_group == "Processing method decode / CM UI":
        return "processing method action table TSV; IRC insertion audit; updated Contract 2; relationship-rule note"
    if task.evidence_group == "Method command decode":
        return "decoded method flow TSV; RetTime/channel manifest; updated Contract 1"
    if task.evidence_group == "Report workbook/formula extraction":
        return "report formula map TSV; exported report workbook preview; DB preview trace; updated Contract 3"
    return "captured evidence note; updated black-box contract"


def _p1_evidence_closure_update(task: TccContractClosureTaskRow) -> str:
    return (
        f"Update {task.document} -> {task.contract}; close or narrow Open Verification topic `{task.topic}`; "
        "rerun FOQ Knowledge Alignment export."
    )


def _processing_method_expected_behavior(processing_method: str) -> str:
    normalized = processing_method.upper()
    if "ACCURACY_IRC_STOP_H" in normalized:
        return "VH accuracy IRC STOP branch; confirm pass-action row, inserted branch, and stop/continue behavior"
    if "ACCURACY_IRC_STOP_C" in normalized:
        return "VC/VA accuracy IRC STOP branch; confirm pass-action row, inserted branch, and stop/continue behavior"
    if "CORRECT_ACCURACY_INJ_INSERTION" in normalized:
        return "TD-backed corrective accuracy insertion: GenericBool0 pass inserts Temperature Accuracy_H, fail inserts Temperature Accuracy_C from FOQ_VX-C10_V2_00_AdditionalInjections; confirm serialized CM action row"
    if "CORRECT_STABILITY_INJ_INSERTION" in normalized:
        return "TD-backed corrective stability insertion: GenericBool0 pass inserts Temperature Stability_and_PCC_H, fail inserts Temperature Stability_C from FOQ_VX-C10_V2_00_AdditionalInjections; confirm serialized CM action row"
    if "NO_INTEGRATION" in normalized:
        return "No integration / no IRC correction expected; confirm method has no hidden pass-action dependency"
    if not processing_method or processing_method == "(not bound)":
        return "open - processing method binding missing"
    return "Decode processing action table and classify pass/fail/IRC behavior"


def _processing_method_extraction_target(processing_method: str, topic: str) -> str:
    normalized = processing_method.upper()
    if "IRC" in normalized or "STOP" in normalized:
        return "IRC pass-action rows, inserted injection target, stop/continue action"
    if normalized.startswith("CORRECT_") or "INJ_INSERTION" in normalized:
        return "Corrective injection insertion rows and trigger criteria"
    if "NO_INTEGRATION" in normalized:
        return "No-action processing-method confirmation"
    if "irc" in topic.lower() or "insertion" in topic.lower():
        return "Processing method action table / IRC pass-action rows"
    return "Processing method action table"


def _processing_method_target_readiness(processing_method: str) -> str:
    if not processing_method or processing_method == "(not bound)":
        return "open - missing processing method binding"
    return "open - decode and verify processing action evidence"


def _records_for_closure_test_id(
    test_id: str,
    records: Iterable[FoqAlignmentRecord],
) -> tuple[FoqAlignmentRecord, ...]:
    aliases = _closure_test_id_aliases(test_id)
    return tuple(record for record in records if record.family == "TCC" and test_knowledge_node_from_record(record).test_id in aliases)


def _closure_test_id_aliases(test_id: str) -> tuple[str, ...]:
    if test_id in {"TCC_STABILITY_01", "TCC_STABILITY_PCC_01"}:
        return ("TCC_STABILITY_01", "TCC_STABILITY_PCC_01")
    return (test_id,)


def _report_formula_extraction_target(topic: str) -> str:
    lowered = topic.lower()
    if "display precision" in lowered or "number format" in lowered:
        return "Workbook number formats / exported XLS display precision"
    if "formulaone" in lowered or "workbook" in lowered:
        return "FormulaOne workbook formulas and dependency cells"
    if "pass/fail" in lowered or "pass" in lowered:
        return "Workbook-derived pass/fail cells"
    if "db" in lowered:
        return "Report cell to DB field trace"
    return "ReportFormulaObject and workbook-derived rule trace"


def _report_formula_target_readiness(
    report_template: str,
    report_sheets: Iterable[str],
    db_fields: Iterable[str],
) -> str:
    if not report_template or report_template == "(not bound)" or not tuple(report_sheets):
        return "open - missing report binding"
    if not tuple(db_fields):
        return "open - missing DB field binding"
    return "open - extract and verify workbook evidence"


def _report_formula_extraction_steps(target: TccReportFormulaTargetRow) -> str:
    lowered_target = target.extraction_target.lower()
    steps = [
        f"1. Locate `{target.report_template}` in the loaded CMBX report template payloads.",
        f"2. Open report sheet(s): {target.report_sheets}.",
        "3. Extract direct `ReportFormulaObject` formulas and `FixedChannel` bindings.",
    ]
    if "formulaone" in lowered_target or "workbook" in lowered_target or "pass/fail" in lowered_target:
        steps.append("4. Extract `SpreadSheetData` / FormulaOne workbook formulas and dependency cells.")
    else:
        steps.append("4. Check whether the DB-facing cell is direct SheetObject output or workbook-derived.")
    if "number format" in lowered_target or "precision" in lowered_target:
        steps.append("5. Capture workbook number format / exported XLS display precision for mapped cells.")
    else:
        steps.append("5. Capture display precision only where mapped report cells are DB-facing.")
    steps.extend(
        (
            f"6. Trace mapped DB fields: {target.db_fields}.",
            f"7. Confirm formula ID `{target.formula_id}` against report evidence and dependency trace.",
        )
    )
    return "\n".join(steps)


def _report_formula_validation_outputs(target: TccReportFormulaTargetRow) -> str:
    outputs = [
        "exported report workbook preview",
        "report formula map TSV",
        "DB preview trace",
        "updated black-box Contract 3",
    ]
    if "precision" in target.extraction_target.lower() or "number format" in target.extraction_target.lower():
        outputs.append("display precision/type audit note")
    return "; ".join(outputs)


def _report_formula_closure_update(target: TccReportFormulaTargetRow) -> str:
    updates = [
        f"Update Contract 3 Report Formula in {target.source_document}.",
        "Move matching Contract 6 report-formula open topic toward closed evidence.",
    ]
    if "precision" in target.extraction_target.lower() or "db" in target.extraction_target.lower():
        updates.append("Update Contract 4 DB Contract if mapped field precision/type changes.")
    return " ".join(updates)


def _open_verification_category_contract(category: str) -> str:
    mapping = {
        "Method Command": "Contract 1 Method Command",
        "Processing Method": "Contract 2 Processing Method",
        "Report Formula": "Contract 3 Report Formula",
        "DB Contract": "Contract 4 DB Contract",
        "Config Requirement": "Contract 5 Config Requirement",
    }
    return mapping.get(category, "Contract 6 Open Verification")


def _contract_likely_evidence_source(contract_name: str) -> str:
    category = contract_name.replace("Contract 1 ", "").replace("Contract 2 ", "").replace("Contract 3 ", "").replace("Contract 4 ", "").replace("Contract 5 ", "")
    if "Method Command" in contract_name:
        return _open_verification_likely_evidence_source("Method Command", "")
    if "Processing" in contract_name:
        return _open_verification_likely_evidence_source("Processing Method", "")
    if "Report" in contract_name:
        return _open_verification_likely_evidence_source("Report Formula", "")
    if "DB" in contract_name:
        return _open_verification_likely_evidence_source("DB Contract", "")
    if "Config" in contract_name:
        return _open_verification_likely_evidence_source("Config Requirement", "")
    return f"source black-box document section for {category}"


def _contract_closure_action(contract_name: str) -> str:
    if "Method Command" in contract_name:
        return _open_verification_closure_action("Method Command", "")
    if "Processing" in contract_name:
        return _open_verification_closure_action("Processing Method", "")
    if "Report" in contract_name:
        return _open_verification_closure_action("Report Formula", "")
    if "DB" in contract_name:
        return _open_verification_closure_action("DB Contract", "")
    if "Config" in contract_name:
        return _open_verification_closure_action("Config Requirement", "")
    return "add an explicit Open Verification section or close all uncertainty with source-backed evidence"


def _temperature_template_readiness(
    coverage: TccBlackBoxCoverageRow,
    topics: Iterable[TccOpenVerificationTopicRow],
) -> str:
    topic_tuple = tuple(topics)
    if not coverage.exists:
        return "missing document"
    if not coverage.status.startswith("documented"):
        return "incomplete contract headings"
    if topic_tuple:
        categories = ", ".join(_unique(topic.category for topic in topic_tuple))
        return f"review-only; close {len(topic_tuple)} topic(s): {categories}"
    return "candidate template after CM validation"


def build_tcc_milestone_status_rows(
    records: Iterable[FoqAlignmentRecord] = (),
    docs_root: str | Path | None = None,
) -> tuple[TccMilestoneStatusRow, ...]:
    record_tuple = tuple(records)
    coverage_rows = build_tcc_black_box_coverage_rows(docs_root)
    relationship_rows = build_tcc_relationship_rows(record_tuple)
    tcc_records = tuple(record for record in (record_tuple or base_alignment_records()) if record.family == "TCC")
    m1_rows = tuple(row for row in coverage_rows if row.milestone == "M1")
    m2_rows = tuple(row for row in coverage_rows if row.milestone == "M2")
    m3_rows = tuple(row for row in coverage_rows if row.milestone == "M3")
    m1_complete = bool(m1_rows) and all(row.exists and row.contract_1_method and row.contract_3_report and row.contract_4_db for row in m1_rows)
    m2_documented = bool(m2_rows) and all(row.exists for row in m2_rows)
    m3_documented = bool(m3_rows) and all(row.exists for row in m3_rows)
    relationship_ready = bool(relationship_rows) and any(row.category == "Dependency" for row in relationship_rows)
    intent_tools_ready = all(tool in INTENT_TOOL_OPTIONS for tool in ("Search / Recommend", "Crop / Modify", "Merge", "Compare"))
    has_review_packets = bool(tcc_records) and any(record_intent_gate(row, "Crop / Modify", "40 C").can_export_generic_packet for row in tcc_records)
    has_specialized = any(record_intent_gate(row, "Crop / Modify", "40 C").can_export_specialized_packet for row in tcc_records)
    return (
        TccMilestoneStatusRow(
            "M1",
            "Temperature Accuracy black-box decomposition",
            _milestone_evidence_label(m1_rows),
            "complete enough for review tooling" if m1_complete else "incomplete",
            _milestone_open_work_label(m1_rows),
        ),
        TccMilestoneStatusRow(
            "M2",
            "All temperature-family TCC black-box decompositions",
            _milestone_evidence_label(m2_rows),
            "documented with open verification" if m2_documented else "incomplete",
            _milestone_open_work_label(m2_rows),
        ),
        TccMilestoneStatusRow(
            "M3",
            "All non-temperature TCC black-box decompositions",
            _milestone_evidence_label(m3_rows),
            "documented with open verification" if m3_documented else "incomplete",
            _milestone_open_work_label(m3_rows),
        ),
        TccMilestoneStatusRow(
            "M4",
            "Inter-test relationship model",
            f"{len(relationship_rows)} structured relationship rows; TCC_TEST_RELATIONSHIP_MODEL.md",
            "structured and exportable" if relationship_ready else "incomplete",
            "Continue closing relationship rows as black-box open verification closes.",
        ),
        TccMilestoneStatusRow(
            "M5",
            "Intent tools for search/crop/merge/compare and draft packet review",
            f"Tools: {', '.join(INTENT_TOOL_OPTIONS)}; generic packets: {'Yes' if has_review_packets else 'No'}; specialized Accuracy packet: {'Yes' if has_specialized else 'No'}",
            "review tooling available; runnable generation closed" if intent_tools_ready and has_review_packets else "incomplete",
            "Runnable CMBX generation still requires closed method/report/config contracts and CM validation.",
        ),
    )


def _milestone_evidence_label(rows: Iterable[TccBlackBoxCoverageRow]) -> str:
    row_tuple = tuple(rows)
    if not row_tuple:
        return "No expected rows"
    documented = sum(1 for row in row_tuple if row.exists)
    complete_contracts = sum(
        1
        for row in row_tuple
        if row.contract_1_method
        and row.contract_2_processing
        and row.contract_3_report
        and row.contract_4_db
        and row.contract_5_config
        and row.contract_6_open_verification
    )
    evidence_sources = sum(1 for row in row_tuple if row.evidence_sources_present)
    all_model_branches = sum(1 for row in row_tuple if row.model_branches == "VH-C10-A, VC-C10-A, VA-C10-A")
    return (
        f"{documented}/{len(row_tuple)} documents present; "
        f"{complete_contracts}/{len(row_tuple)} six-contract rows detected; "
        f"{evidence_sources}/{len(row_tuple)} evidence-source sections; "
        f"{all_model_branches}/{len(row_tuple)} documents mention all VH/VC/VA branches"
    )


def _milestone_open_work_label(rows: Iterable[TccBlackBoxCoverageRow]) -> str:
    row_tuple = tuple(rows)
    missing = [row.test_name for row in row_tuple if not row.exists]
    incomplete = [row.test_name for row in row_tuple if row.exists and not row.status.startswith("documented")]
    open_verification = [row.test_name for row in row_tuple if row.open_verification_present]
    open_topic_count = sum(row.open_verification_count for row in row_tuple if row.open_verification_present)
    parts: list[str] = []
    if missing:
        parts.append("Missing docs: " + ", ".join(missing))
    if incomplete:
        parts.append("Incomplete headings/evidence: " + ", ".join(incomplete))
    if open_verification:
        parts.append(f"Open verification remains ({len(open_verification)} docs / {open_topic_count} topics): " + ", ".join(open_verification))
    return "; ".join(parts) if parts else "No open work recorded in this audit."


def _black_box_contract_presence(text: str) -> tuple[bool, bool, bool, bool, bool, bool]:
    lowered = text.lower()
    method = "method command" in lowered or "instrument method" in lowered
    processing = "processing method" in lowered
    report = "report formula" in lowered or "formulaone" in lowered or "report template" in lowered
    db = "db contract" in lowered or "db field" in lowered
    config = "config requirement" in lowered or "configuration requirement" in lowered or "required configuration" in lowered
    open_verification = "open verification" in lowered
    return method, processing, report, db, config, open_verification


def _black_box_open_verification_count(text: str) -> int:
    if not text:
        return 0
    count = len(re.findall(r"Open Verification Required", text, flags=re.IGNORECASE))
    if count:
        return count
    return 1 if "open verification" in text.lower() else 0


def _black_box_open_verification_topics(text: str) -> tuple[str, ...]:
    if not text or "open verification" not in text.lower():
        return ()
    section_match = re.search(
        r"## Contract 6: Open Verification(?P<body>.*?)(?:\n## |\Z)",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    body = section_match.group("body") if section_match else text
    topics: list[str] = []
    for line in body.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        cells = [cell.strip() for cell in stripped.strip("|").split("|")]
        if len(cells) < 2:
            continue
        if not cells[0] or cells[0] in {"#", "---"} or set(cells[0]) == {"-"}:
            continue
        if cells[0].lower() in {"#", "no.", "number"} or "uncertain" in cells[1].lower():
            continue
        if not re.match(r"^\d+$", cells[0]):
            continue
        topic = re.sub(r"`", "", cells[1]).strip()
        if topic:
            topics.append(topic)
    if topics:
        return _unique(topics)
    fallback = []
    for match in re.finditer(r"([^.。\n]*Open Verification Required[^.。\n]*)", body, flags=re.IGNORECASE):
        fallback.append(match.group(1).strip(" -|"))
    return _unique(fallback)


def _open_verification_topic_category(topic: str) -> str:
    lowered = topic.lower()
    if any(token in lowered for token in ("processing", "pass-action", "pass action", "insertion", "insert", "irc", "sst", "procmeth")):
        return "Processing Method"
    if any(token in lowered for token in ("formulaone", "workbook", "formula", "cell", "cells", "display precision", "number format", "report", "sheet", "pass/fail", "definitions")):
        return "Report Formula"
    if any(token in lowered for token in ("db", "sql", "upload", "field")):
        return "DB Contract"
    if any(token in lowered for token in ("command", "script", "rettime", "ret time", "method flow", "method branch", "trigger")):
        return "Method Command"
    if any(token in lowered for token in ("config", "symbol", "channel", "modelno", "device", "genericbool", "hardware")):
        return "Config Requirement"
    return "Open Verification"


def _open_verification_likely_evidence_source(category: str, topic: str) -> str:
    if category == "Processing Method":
        return "embedded processing method payload, CM processing-method editor, pass-action/SST row decode"
    if category == "Report Formula":
        return "report template SpreadSheetData / FormulaOne workbook, report formula objects, exported XLS comparison"
    if category == "DB Contract":
        return "FOQResultLocations mapping, exported DB workbook, SQL upload type/precision audit"
    if category == "Method Command":
        return "decoded instrument method flow, CM method script/export, method contract summary"
    if category == "Config Requirement":
        return "CM instrument configuration, required-symbol manifest, CMBX precondition/audit evidence"
    return "source black-box document plus targeted CMBX/CM evidence"


def _open_verification_closure_action(category: str, topic: str) -> str:
    if category == "Processing Method":
        return "decode or manually confirm business rows, then update Contract 2 and relationship rules"
    if category == "Report Formula":
        return "extract/verify workbook formula or display format, then update Contract 3 and DB precision notes"
    if category == "DB Contract":
        return "verify field mapping/type/precision against export/upload contract, then update Contract 4"
    if category == "Method Command":
        return "add line-level command evidence and RetTime/channel semantics to Contract 1"
    if category == "Config Requirement":
        return "record required symbol/channel/model evidence in Contract 5 and required-symbol manifest"
    return "replace the open item with source-backed evidence and rerun alignment export"


def _black_box_model_branch_label(text: str) -> str:
    models = tuple(model for model in ("VH-C10-A", "VC-C10-A", "VA-C10-A") if model in text)
    return ", ".join(models) if models else "(not explicit)"


def _black_box_word_count(text: str) -> int:
    return len(re.findall(r"\b[\w./:-]+\b", text))


def build_tcc_relationship_rows(records: Iterable[FoqAlignmentRecord] = ()) -> tuple[TccRelationshipRow, ...]:
    """Return structured TCC relationship rules for workbook/UI handoff."""
    record_tuple = tuple(record for record in records if record.family == "TCC")
    by_intent = {record.test_intent: record for record in record_tuple}
    rows = (
        TccRelationshipRow(
            "Execution Order",
            "ORDER_01",
            "BurnIn",
            "Temperature Calibration",
            "hard for full FOQ; review-required for custom subset",
            "BurnIn creates thermal history and thermometer sanity context before the temperature family.",
            "Do not omit BurnIn from full FOQ clones; for custom subsets, record an explicit preconditioning decision.",
            "TCC_TEST_RELATIONSHIP_MODEL.md / BurnIn black-box decomposition",
        ),
        TccRelationshipRow(
            "Execution Order",
            "ORDER_02",
            "Temperature Calibration",
            "Accuracy, Precision, Stability, HeatUp/CoolDown",
            "hard for full FOQ",
            "Later temperature tests assume calibrated or conditioned temperature behavior.",
            "Any cropped temperature method must state whether Calibration is reused, rerun, or declared out of scope.",
            "TCC_TEST_RELATIONSHIP_MODEL.md / Calibration black-box decomposition",
        ),
        TccRelationshipRow(
            "Execution Order",
            "ORDER_03",
            "Factory Default",
            "Error Log Check",
            "hard for full FOQ",
            "Factory Default mutates final state and clears/logs state before the final audit/error-log endpoint.",
            "Keep the finalization chain together unless the intent is explicitly not a full FOQ package.",
            "TCC_TEST_RELATIONSHIP_MODEL.md / Factory Default and Error Log decompositions",
        ),
        TccRelationshipRow(
            "Dependency",
            "DEP_01",
            "Temperature Calibration",
            "Temperature Accuracy",
            "hard for full FOQ",
            "Accuracy setpoint values are interpreted in a calibrated temperature context.",
            "Single-point Accuracy crop still needs a Calibration/BurnIn decision and report/DB remapping.",
            _record_evidence_label(by_intent.get("temperature_accuracy")),
        ),
        TccRelationshipRow(
            "Dependency",
            "DEP_02",
            "Temperature Calibration",
            "Temperature Precision / Stability",
            "hard for full FOQ",
            "Precision and Stability use external thermometer windows whose interpretation depends on calibrated temperature behavior.",
            "Do not merge or crop windows without preserving separate lower/upper sensor range semantics.",
            "TCC_TEST_RELATIONSHIP_MODEL.md / Precision and Stability decompositions",
        ),
        TccRelationshipRow(
            "Dependency",
            "DEP_03",
            "Temperature Stability",
            "PCC",
            "VH-only hard branch",
            "VH Stability includes PCC cooldown/performance fields; VC/VA branch does not.",
            "Never substitute the no-PCC branch for VH when generating or filtering a full VH FOQ package.",
            _record_evidence_label(by_intent.get("temperature_stability_and_pcc")),
        ),
        TccRelationshipRow(
            "Shared Resource",
            "RES_01",
            "External thermometers",
            "BurnIn, Calibration, Accuracy, Precision, Stability, HeatUp/CoolDown",
            "hard config dependency",
            "Report formulas depend on `ExtTemp_UpperCC` and/or `ExtTemp_LowerCC` raw channels.",
            "Target CM configuration must provide Generic Device thermometer channels before method reuse.",
            "TCC_REQUIRED_SYMBOL_MANIFEST.md / report formula evidence",
        ),
        TccRelationshipRow(
            "Shared Resource",
            "RES_02",
            "AUDIT.ColumnComp.ModelNo",
            "all branch-sensitive TCC tests and DB upload",
            "hard identity dependency",
            "Device model is the source of truth for report/template/DB branch selection.",
            "Do not infer VA/VC/VH from filenames when selecting report templates or DB tables.",
            "Factory Default and DB alignment evidence",
        ),
        TccRelationshipRow(
            "Intent Rule",
            "INTENT_01",
            "Temperature Accuracy",
            "single setpoint crop",
            "editable after review",
            "RetTime ladder, report rows, and DB field subset must shrink together.",
            "Use the specialized draft packet only when a numeric setpoint is supplied; otherwise export generic review packet.",
            _record_evidence_label(by_intent.get("temperature_accuracy")),
        ),
        TccRelationshipRow(
            "Intent Rule",
            "INTENT_02",
            "HeatUp/CoolDown",
            "custom temperature range",
            "editable after review",
            "Setpoints, trigger thresholds, report labels, DB meaning, and `- 2.0 min` hold contract move together.",
            "Export generic packet until the changed RetTime/report formula contract is manually closed.",
            _record_evidence_label(by_intent.get("heatup_cooldown_20_50_20")),
        ),
        TccRelationshipRow(
            "Intent Rule",
            "INTENT_03",
            "Valve / Keypad",
            "periodic valve cycling",
            "custom method possible; FOQ test locked",
            "Periodic valve cycling can reuse command knowledge, but FOQ keypad/disconnect workflow is a different contract.",
            "Generate a custom method only after separating diagnostic valve cycling from FOQ Valve/Keypad pass/fail semantics.",
            _record_evidence_label(by_intent.get("valve_keypad")),
        ),
    )
    return rows


def record_relationship_rule_lines(record: FoqAlignmentRecord, intent: str = "") -> tuple[str, ...]:
    """Return relationship-model rules relevant to one alignment row."""
    matches: list[str] = []
    for row in record_relationship_rows(record, intent):
        matches.append(f"- [{row.item_id}] {row.category}: {row.source} -> {row.target} ({row.strength}); {row.generation_rule}")
    return tuple(matches[:8])


def record_relationship_rows(record: FoqAlignmentRecord, intent: str = "") -> tuple[TccRelationshipRow, ...]:
    """Return structured relationship-model rows relevant to one alignment row."""
    if record.family != "TCC":
        return ()
    terms = _relationship_terms(record)
    if (intent or "").strip().lower().startswith("crop"):
        terms = (*terms, "single setpoint crop", "custom temperature range")
    matches: list[TccRelationshipRow] = []
    for row in build_tcc_relationship_rows((record,)):
        text = " ".join((row.category, row.item_id, row.source, row.target, row.impact, row.generation_rule)).lower()
        if any(term in text for term in terms):
            matches.append(row)
    return tuple(matches[:12])


def record_relationship_audit_lines(record: FoqAlignmentRecord, intent: str = "") -> tuple[str, ...]:
    """Render relationship rows as a reviewable per-test audit."""
    rows = record_relationship_rows(record, intent)
    if not rows:
        return (
            "No structured relationship rows are mapped to this alignment row.",
            "Complete or extend TCC_TEST_RELATIONSHIP_MODEL.md before using this row for crop/merge decisions.",
        )
    categories = _unique(row.category for row in rows)
    lines = [
        f"Matched relationship rows: {len(rows)}",
        f"Categories: {', '.join(categories)}",
        "",
        "| Rule ID | Category | Source -> Target | Strength | Impact | Generation Rule |",
        "|---|---|---|---|---|---|",
    ]
    for row in rows:
        lines.append(
            "| "
            + " | ".join(
                _md_cell(value)
                for value in (
                    row.item_id,
                    row.category,
                    f"{row.source} -> {row.target}",
                    row.strength,
                    row.impact,
                    row.generation_rule,
                )
            )
            + " |"
        )
    lines.extend(
        (
            "",
            "How to use this audit:",
            "- Execution Order rows explain what must run before/after in a full FOQ clone.",
            "- Dependency rows explain why a test cannot be cropped independently.",
            "- Shared Resource rows explain which CM configuration, channels, or audit properties must survive.",
            "- Intent Rule rows explain crop/merge/compare-specific review gates.",
        )
    )
    return tuple(lines)


def _relationship_terms(record: FoqAlignmentRecord) -> tuple[str, ...]:
    mapping = {
        "temperature_calibration": ("temperature calibration", "calibration"),
        "temperature_accuracy": ("temperature accuracy", "accuracy", "single setpoint crop"),
        "temperature_precision_and_fan": ("temperature precision", "precision"),
        "temperature_stability_and_pcc": ("temperature stability", "stability", "pcc"),
        "temperature_stability_no_pcc": ("temperature stability", "stability"),
        "heatup_cooldown_20_50_20": ("heatup", "heatup/cooldown", "custom temperature range"),
        "valve_keypad": ("valve", "periodic valve"),
        "burn_in": ("burnin", "burnin"),
        "factory_default_metadata": ("factory default",),
        "error_log_check": ("error log",),
    }
    return mapping.get(record.test_intent, (_normalize_name(record.td_test), record.td_test.lower()))


def _record_evidence_label(record: FoqAlignmentRecord | None) -> str:
    if record is None:
        return "TCC_TEST_RELATIONSHIP_MODEL.md"
    bits = [record.td_source]
    bits.extend(record.cmbx_sources[:2])
    if record.report_evidence:
        bits.append(record.report_evidence[0])
    return "; ".join(_unique(bits)) or "TCC_TEST_RELATIONSHIP_MODEL.md"


def write_foq_alignment_workbook(
    records: Iterable[FoqAlignmentRecord],
    output_path: str | Path,
    intent: str = "",
    parameter: str = "",
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to export FOQ alignment workbooks.") from exc

    record_tuple = tuple(records)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FOQ Knowledge Alignment"
    headers = (
        "Family",
        "TestIntent",
        "TD Test",
        "Device",
        "Injection",
        "Instrument Method",
        "Processing Method",
        "Report Template",
        "Report Sheets",
        "DB Fields",
        "Coverage",
        "Modifiability",
        "Intent Gate",
        "Generic Draft Packet",
        "Specialized Draft Packet",
        "Runnable Generation",
        "Open Gaps",
        "TD Meaning",
        "Method Evidence",
        "Report Evidence",
        "DB Evidence",
        "Generation Readiness",
        "CMBX Sources",
        "TD Source",
    )
    sheet.append(headers)
    for record in record_tuple:
        gate = record_intent_gate(record, intent or "Search / Recommend", parameter, selected_records=(record,))
        sheet.append(
            (
                record.family,
                record.test_intent,
                record.td_test,
                record.device_label,
                record.injection,
                record.instrument_method,
                record.processing_method,
                record.report_template,
                record.report_sheet_label,
                record.db_field_label,
                record.coverage_status,
                record_modifiability_summary(record),
                gate.status,
                "Yes" if gate.can_export_generic_packet else "No",
                "Yes" if gate.can_export_specialized_packet else "No",
                "Yes" if gate.runnable_generation_allowed else "No",
                "\n".join(record.open_gaps),
                record.td_meaning,
                "\n".join(record.method_evidence),
                "\n".join(record.report_evidence),
                "\n".join(record.db_evidence),
                record.generation_readiness,
                "\n".join(record.cmbx_sources),
                record.td_source,
            )
        )

    gate_sheet = workbook.create_sheet("Intent Gate Matrix")
    gate_sheet.append(
        (
            "Family",
            "TestIntent",
            "TD Test",
            "Device",
            "Intent",
            "Parameter",
            "Gate Status",
            "Generic Draft Packet",
            "Specialized Draft Packet",
            "Runnable Generation",
            "Blockers",
            "Next Actions",
            "Relationship Rules",
            "Coverage",
            "Modifiability",
        )
    )
    for record in record_tuple:
        current_intent = intent or "Search / Recommend"
        gate = record_intent_gate(record, current_intent, parameter, selected_records=(record,))
        gate_sheet.append(
            (
                record.family,
                record.test_intent,
                record.td_test,
                record.device_label,
                current_intent,
                parameter,
                gate.status,
                "Yes" if gate.can_export_generic_packet else "No",
                "Yes" if gate.can_export_specialized_packet else "No",
                "Yes" if gate.runnable_generation_allowed else "No",
                "\n".join(gate.blockers),
                "\n".join(gate.next_actions),
                "\n".join(record_relationship_rule_lines(record, current_intent)),
                record.coverage_status,
                record_modifiability_summary(record),
            )
        )

    milestone_sheet = workbook.create_sheet("TCC Milestone Status")
    milestone_sheet.append(("Milestone", "Objective", "Evidence", "Status", "Open Work"))
    for milestone in build_tcc_milestone_status_rows(record_tuple):
        milestone_sheet.append(
            (
                milestone.milestone,
                milestone.objective,
                milestone.evidence,
                milestone.status,
                milestone.open_work,
            )
        )

    next_action_sheet = workbook.create_sheet("TCC Next Action Queue")
    next_action_sheet.append(
        (
            "Rank",
            "Milestone",
            "Priority",
            "Evidence Group",
            "Task Count",
            "Tests",
            "Contracts",
            "Primary Blocker",
            "Next Action",
            "Unlocks",
            "Generation Gate",
        )
    )
    for row in build_tcc_next_action_queue_rows():
        next_action_sheet.append(
            (
                row.rank,
                row.milestone,
                row.priority,
                row.evidence_group,
                row.task_count,
                row.tests,
                row.contracts,
                row.primary_blocker,
                row.next_action,
                row.unlocks,
                row.generation_gate,
            )
        )

    black_box_sheet = workbook.create_sheet("TCC BlackBox Coverage")
    black_box_sheet.append(
        (
            "Milestone",
            "Test ID",
            "Test",
            "Document",
            "Exists",
            "Contract 1 Method",
            "Contract 2 Processing",
            "Contract 3 Report",
            "Contract 4 DB",
            "Contract 5 Config",
            "Contract 6 Open Verification",
            "Open Verification Present",
            "Open Verification Count",
            "Open Verification Topics",
            "Evidence Sources Present",
            "Model Branches Mentioned",
            "Mermaid Present",
            "Word Count",
            "Status",
        )
    )
    for coverage in build_tcc_black_box_coverage_rows():
        black_box_sheet.append(
            (
                coverage.milestone,
                coverage.test_id,
                coverage.test_name,
                coverage.document,
                "Yes" if coverage.exists else "No",
                "Yes" if coverage.contract_1_method else "No",
                "Yes" if coverage.contract_2_processing else "No",
                "Yes" if coverage.contract_3_report else "No",
                "Yes" if coverage.contract_4_db else "No",
                "Yes" if coverage.contract_5_config else "No",
                "Yes" if coverage.contract_6_open_verification else "No",
                "Yes" if coverage.open_verification_present else "No",
                coverage.open_verification_count,
                "\n".join(coverage.open_verification_topics),
                "Yes" if coverage.evidence_sources_present else "No",
                coverage.model_branches,
                "Yes" if coverage.mermaid_present else "No",
                coverage.word_count,
                coverage.status,
            )
        )

    temp_matrix_sheet = workbook.create_sheet("M2 Temperature Contract Matrix")
    temp_matrix_sheet.append(
        (
            "Test ID",
            "Test",
            "Document",
            "Method Contract",
            "Processing Contract",
            "Report Contract",
            "DB Contract",
            "Config Contract",
            "Open Verification Contract",
            "Open Topic Categories",
            "Next Closure Actions",
            "Template Readiness",
        )
    )
    for row in build_tcc_temperature_contract_matrix_rows():
        temp_matrix_sheet.append(
            (
                row.test_id,
                row.test_name,
                row.document,
                row.method_contract,
                row.processing_contract,
                row.report_contract,
                row.db_contract,
                row.config_contract,
                row.open_verification_contract,
                row.open_topic_categories,
                row.next_closure_actions,
                row.template_readiness,
            )
        )

    closure_task_sheet = workbook.create_sheet("M2 Contract Closure Tasks")
    closure_task_sheet.append(
        (
            "Milestone",
            "Test ID",
            "Test",
            "Document",
            "Contract",
            "Priority",
            "Evidence Group",
            "Task Type",
            "Topic",
            "Likely Evidence Source",
            "Closure Action",
            "Generation Blocker",
        )
    )
    for row in build_tcc_contract_closure_task_rows(milestone="M2"):
        closure_task_sheet.append(
            (
                row.milestone,
                row.test_id,
                row.test_name,
                row.document,
                row.contract,
                row.priority,
                row.evidence_group,
                row.task_type,
                row.topic,
                row.likely_evidence_source,
                row.closure_action,
                row.generation_blocker,
            )
        )

    workstream_sheet = workbook.create_sheet("M2 Evidence Workstreams")
    workstream_sheet.append(
        (
            "Milestone",
            "Priority",
            "Evidence Group",
            "Task Count",
            "Tests",
            "Contracts",
            "Likely Evidence Sources",
            "Closure Actions",
            "Unlocks",
            "Next Action",
        )
    )
    for row in build_tcc_evidence_workstream_rows(milestone="M2"):
        workstream_sheet.append(
            (
                row.milestone,
                row.priority,
                row.evidence_group,
                row.task_count,
                row.tests,
                row.contracts,
                row.likely_evidence_sources,
                row.closure_actions,
                row.unlocks,
                row.next_action,
            )
        )

    p1_plan_sheet = workbook.create_sheet("M2 P1 Extraction Plan")
    p1_plan_sheet.append(
        (
            "Milestone",
            "Test ID",
            "Test",
            "Document",
            "Contract",
            "Evidence Group",
            "Topic",
            "Evidence Source",
            "Extraction Steps",
            "Validation Outputs",
            "Closure Update",
            "Status",
        )
    )
    for row in build_tcc_p1_evidence_extraction_plan_rows(milestone="M2"):
        p1_plan_sheet.append(
            (
                row.milestone,
                row.test_id,
                row.test_name,
                row.document,
                row.contract,
                row.evidence_group,
                row.topic,
                row.evidence_source,
                row.extraction_steps,
                row.validation_outputs,
                row.closure_update,
                row.status,
            )
        )

    processing_target_sheet = workbook.create_sheet("M2 Processing Targets")
    processing_target_sheet.append(
        (
            "Milestone",
            "Test ID",
            "Test",
            "Device",
            "Injection",
            "Instrument Method",
            "Processing Method",
            "Topic",
            "Expected Behavior",
            "Extraction Target",
            "Source Document",
            "Closure Action",
            "Readiness",
        )
    )
    for row in build_tcc_processing_method_target_rows(milestone="M2"):
        processing_target_sheet.append(
            (
                row.milestone,
                row.test_id,
                row.test_name,
                row.device_model,
                row.injection,
                row.instrument_method,
                row.processing_method,
                row.topic,
                row.expected_behavior,
                row.extraction_target,
                row.source_document,
                row.closure_action,
                row.readiness,
            )
        )

    report_target_sheet = workbook.create_sheet("M2 Report Formula Targets")
    report_target_sheet.append(
        (
            "Milestone",
            "Test ID",
            "Test",
            "Device",
            "Topic",
            "Report Template",
            "Report Sheets",
            "DB Fields",
            "Formula ID",
            "Extraction Target",
            "Source Document",
            "Closure Action",
            "Readiness",
        )
    )
    for row in build_tcc_report_formula_target_rows(milestone="M2", records=record_tuple):
        report_target_sheet.append(
            (
                row.milestone,
                row.test_id,
                row.test_name,
                row.device_model,
                row.topic,
                row.report_template,
                row.report_sheets,
                row.db_fields,
                row.formula_id,
                row.extraction_target,
                row.source_document,
                row.closure_action,
                row.readiness,
            )
        )

    report_plan_sheet = workbook.create_sheet("M2 Report Extraction Plan")
    report_plan_sheet.append(
        (
            "Milestone",
            "Test ID",
            "Test",
            "Device",
            "Report Template",
            "Report Sheets",
            "Formula ID",
            "DB Fields",
            "Extraction Target",
            "Extraction Steps",
            "Validation Outputs",
            "Closure Update",
            "Status",
        )
    )
    for row in build_tcc_report_formula_extraction_plan_rows(milestone="M2", records=record_tuple):
        report_plan_sheet.append(
            (
                row.milestone,
                row.test_id,
                row.test_name,
                row.device_model,
                row.report_template,
                row.report_sheets,
                row.formula_id,
                row.db_fields,
                row.extraction_target,
                row.extraction_steps,
                row.validation_outputs,
                row.closure_update,
                row.status,
            )
        )

    open_topic_sheet = workbook.create_sheet("TCC Open Verification Topics")
    open_topic_sheet.append(
        (
            "Milestone",
            "Test ID",
            "Test",
            "Document",
            "Category",
            "Topic",
            "Likely Evidence Source",
            "Closure Action",
        )
    )
    for topic_row in build_tcc_open_verification_topic_rows():
        open_topic_sheet.append(
            (
                topic_row.milestone,
                topic_row.test_id,
                topic_row.test_name,
                topic_row.document,
                topic_row.category,
                topic_row.topic,
                topic_row.likely_evidence_source,
                topic_row.closure_action,
            )
        )

    tkn_sheet = workbook.create_sheet("Test Knowledge Nodes")
    tkn_headers = (
        "test_id",
        "test_name",
        "foq_section",
        "purpose",
        "acceptance_criteria",
        "injection",
        "instrument_method",
        "processing_method",
        "report_template",
        "report_sheets",
        "formula_id",
        "db_fields",
        "model_applicability",
        "dependencies",
        "expected_ret_times",
        "expected_channels",
        "expected_audit_properties",
        "required_config",
        "method_evidence",
        "report_evidence",
        "db_evidence",
        "coverage_status",
        "open_gaps",
        "irc_injected",
        "device_bindings",
    )
    tkn_sheet.append(tkn_headers)
    for node in build_test_knowledge_nodes(record_tuple):
        tkn_sheet.append(
            (
                node.test_id,
                node.test_name,
                node.foq_section,
                node.purpose,
                "\n".join(node.acceptance_criteria),
                node.injection,
                node.instrument_method,
                node.processing_method,
                node.report_template,
                node.report_sheet_label,
                node.formula_id,
                node.db_field_label,
                node.model_label,
                "\n".join(node.dependencies),
                "\n".join(node.expected_ret_times),
                "\n".join(node.expected_channels),
                "\n".join(node.expected_audit_properties),
                "\n".join(node.required_config),
                "\n".join(node.method_evidence),
                "\n".join(node.report_evidence),
                "\n".join(node.db_evidence),
                node.coverage_status,
                "\n".join(node.open_gaps),
                "Yes" if node.irc_injected else "No",
                node.device_binding_label,
            )
        )

    mapping_sheet = workbook.create_sheet("Cross-KB Mapping")
    mapping_headers = (
        "test_id",
        "family",
        "foq_test_name",
        "foq_section",
        "method_name",
        "processing_method",
        "report_template",
        "report_sheets",
        "formula_id",
        "db_fields",
        "model_applicability",
        "mapping_status",
    )
    mapping_sheet.append(mapping_headers)
    for mapping in build_cross_kb_mapping_rows(record_tuple):
        mapping_sheet.append(
            (
                mapping.test_id,
                mapping.family,
                mapping.foq_test_name,
                mapping.foq_section,
                mapping.method_name,
                mapping.processing_method,
                mapping.report_template,
                mapping.report_sheet_label,
                mapping.formula_id,
                mapping.db_field_label,
                mapping.model_label,
                mapping.mapping_status,
            )
        )

    audit_sheet = workbook.create_sheet("TKN Coverage Audit")
    audit_headers = (
        "test_id",
        "test_name",
        "injection",
        "method_command_status",
        "processing_method_status",
        "report_formula_status",
        "db_field_status",
        "ret_time_status",
        "channel_status",
        "audit_property_status",
        "config_status",
        "overall_status",
        "gaps",
    )
    audit_sheet.append(audit_headers)
    for audit in build_tkn_coverage_audits(record_tuple):
        audit_sheet.append(
            (
                audit.test_id,
                audit.test_name,
                audit.injection,
                audit.method_command_status,
                audit.processing_method_status,
                audit.report_formula_status,
                audit.db_field_status,
                audit.ret_time_status,
                audit.channel_status,
                audit.audit_property_status,
                audit.config_status,
                audit.overall_status,
                "\n".join(audit.gaps),
            )
        )

    db_mapping_sheet = workbook.create_sheet("TKN DB Mapping Audit")
    db_mapping_headers = (
        "test_id",
        "test_name",
        "device_model",
        "mapping_sheet",
        "injection",
        "report_sheets",
        "mapped_report_files",
        "expected_db_fields",
        "mapped_db_fields",
        "missing_expected_fields",
        "extra_mapped_fields",
        "report_cells",
        "value_types",
        "status",
    )
    db_mapping_sheet.append(db_mapping_headers)
    default_mapping_path = Path(__file__).resolve().parents[1] / "foq" / "FOQResultLocations_V2.83.xls"
    if default_mapping_path.exists():
        for audit in build_tkn_db_mapping_audits(record_tuple, default_mapping_path):
            db_mapping_sheet.append(
                (
                    audit.test_id,
                    audit.test_name,
                    audit.device_model,
                    audit.mapping_sheet,
                    audit.injection,
                    ", ".join(audit.report_sheets),
                    audit.mapped_report_file_label,
                    ", ".join(audit.expected_db_fields),
                    audit.mapped_db_field_label,
                    audit.missing_label,
                    audit.extra_label,
                    ", ".join(audit.report_cells),
                    "\n".join(audit.value_types),
                    audit.status,
                )
            )
    else:
        db_mapping_sheet.append(("", "", "", "", "", "", "", "", "", "", "", "", "", f"mapping file not found: {default_mapping_path}"))

    strategy_sheet = workbook.create_sheet("Generation Strategy")
    strategy_sheet.append(("Category", "ID/Test", "Family/Branch", "Rule/Formula/Check", "Basis/Parameter/Impact", "Evidence Status"))
    strategy_kb = build_cmbx_generation_strategy_kb()
    for rule in strategy_kb.method_rules:
        strategy_sheet.append(("Method Rule", rule.rule_id, rule.family, rule.rule, rule.basis, rule.evidence_status))
    for rule in strategy_kb.formula_rules:
        strategy_sheet.append(("Formula Rule", rule.test_name, rule.formula_id, rule.formula, rule.parameter_sources, rule.evidence_status))
    for rule in strategy_kb.template_rules:
        strategy_sheet.append(
            ("Template Rule", rule.test_name, rule.model_branch, rule.report_template, rule.report_sheet_label, rule.evidence_status)
        )
    for rule in strategy_kb.cross_module_dependencies:
        strategy_sheet.append(("Cross-Module Dependency", rule.dependency, "", rule.impact, "", rule.evidence_status))
    for rule in strategy_kb.config_validation_rules:
        strategy_sheet.append(
            ("Config Validation", rule.validation_item, "", rule.check_method, rule.failure_handling, rule.evidence_status)
        )

    relationship_sheet = workbook.create_sheet("TCC Relationship Model")
    relationship_sheet.append(
        (
            "Category",
            "Rule ID",
            "Source",
            "Target",
            "Strength",
            "Impact",
            "Generation Rule",
            "Evidence",
        )
    )
    for relationship in build_tcc_relationship_rows(record_tuple):
        relationship_sheet.append(
            (
                relationship.category,
                relationship.item_id,
                relationship.source,
                relationship.target,
                relationship.strength,
                relationship.impact,
                relationship.generation_rule,
                relationship.evidence,
            )
        )

    selected_relationship_sheet = workbook.create_sheet("Selected Relationship Audit")
    selected_relationship_sheet.append(
        (
            "Family",
            "Test ID",
            "TD Test",
            "Device",
            "Intent",
            "Rule ID",
            "Category",
            "Source",
            "Target",
            "Strength",
            "Impact",
            "Generation Rule",
            "Evidence",
        )
    )
    current_intent = intent or "Search / Recommend"
    for record in record_tuple:
        node = test_knowledge_node_from_record(record)
        for relationship in record_relationship_rows(record, current_intent):
            selected_relationship_sheet.append(
                (
                    record.family,
                    node.test_id,
                    record.td_test,
                    record.device_label,
                    current_intent,
                    relationship.item_id,
                    relationship.category,
                    relationship.source,
                    relationship.target,
                    relationship.strength,
                    relationship.impact,
                    relationship.generation_rule,
                    relationship.evidence,
                )
            )

    resolution_sheet = workbook.create_sheet("Relationship Resolution Choices")
    resolution_sheet.append(
        (
            "Rule ID",
            "Relationship",
            "Decision Required",
            "Options",
            "Evidence To Capture",
            "Default Recommendation",
        )
    )
    seen_resolution_ids: set[str] = set()
    for relationship in build_tcc_relationship_rows(record_tuple):
        if "hard" not in relationship.strength.lower() or relationship.item_id in seen_resolution_ids:
            continue
        seen_resolution_ids.add(relationship.item_id)
        choice = _relationship_resolution_choice(relationship)
        resolution_sheet.append(
            (
                choice.rule_id,
                choice.relationship,
                choice.decision_required,
                "\n".join(choice.options),
                choice.evidence_to_capture,
                choice.default_recommendation,
            )
        )

    selected_resolution_sheet = workbook.create_sheet("Selected Resolution Choices")
    selected_resolution_sheet.append(
        (
            "Family",
            "Test ID",
            "TD Test",
            "Device",
            "Intent",
            "Rule ID",
            "Relationship",
            "Decision Required",
            "Options",
            "Evidence To Capture",
            "Default Recommendation",
        )
    )
    seen_selected_resolution: set[tuple[str, str, str]] = set()
    for record in record_tuple:
        node = test_knowledge_node_from_record(record)
        for choice in relationship_resolution_choices(record, current_intent, selected_records=(record,)):
            key = (node.test_id, current_intent, choice.rule_id)
            if key in seen_selected_resolution:
                continue
            seen_selected_resolution.add(key)
            selected_resolution_sheet.append(
                (
                    record.family,
                    node.test_id,
                    record.td_test,
                    record.device_label,
                    current_intent,
                    choice.rule_id,
                    choice.relationship,
                    choice.decision_required,
                    "\n".join(choice.options),
                    choice.evidence_to_capture,
                    choice.default_recommendation,
                )
            )

    decision_register_sheet = workbook.create_sheet("Resolution Decision Register")
    decision_register_sheet.append(
        (
            "Family",
            "Test ID",
            "TD Test",
            "Device",
            "Intent",
            "Rule ID",
            "Relationship",
            "Decision Required",
            "Available Options",
            "Selected Option",
            "Decision Status",
            "Evidence Path",
            "Owner",
            "Notes",
        )
    )
    seen_decision_register: set[tuple[str, str, str]] = set()
    for record in record_tuple:
        node = test_knowledge_node_from_record(record)
        for choice in relationship_resolution_choices(record, current_intent, selected_records=(record,)):
            key = (node.test_id, current_intent, choice.rule_id)
            if key in seen_decision_register:
                continue
            seen_decision_register.add(key)
            decision_register_sheet.append(
                (
                    record.family,
                    node.test_id,
                    record.td_test,
                    record.device_label,
                    current_intent,
                    choice.rule_id,
                    choice.relationship,
                    choice.decision_required,
                    "\n".join(choice.options),
                    "",
                    "Open",
                    "",
                    "",
                    choice.default_recommendation,
                )
            )

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for current_sheet in (sheet, gate_sheet, milestone_sheet, next_action_sheet, black_box_sheet, temp_matrix_sheet, closure_task_sheet, workstream_sheet, report_target_sheet, open_topic_sheet, tkn_sheet, mapping_sheet, audit_sheet, db_mapping_sheet, strategy_sheet, relationship_sheet, selected_relationship_sheet, resolution_sheet, selected_resolution_sheet, decision_register_sheet):
        for row in current_sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
    sheet.freeze_panes = "A2"
    gate_sheet.freeze_panes = "A2"
    milestone_sheet.freeze_panes = "A2"
    next_action_sheet.freeze_panes = "A2"
    black_box_sheet.freeze_panes = "A2"
    temp_matrix_sheet.freeze_panes = "A2"
    closure_task_sheet.freeze_panes = "A2"
    workstream_sheet.freeze_panes = "A2"
    report_target_sheet.freeze_panes = "A2"
    open_topic_sheet.freeze_panes = "A2"
    tkn_sheet.freeze_panes = "A2"
    mapping_sheet.freeze_panes = "A2"
    audit_sheet.freeze_panes = "A2"
    db_mapping_sheet.freeze_panes = "A2"
    strategy_sheet.freeze_panes = "A2"
    relationship_sheet.freeze_panes = "A2"
    selected_relationship_sheet.freeze_panes = "A2"
    resolution_sheet.freeze_panes = "A2"
    selected_resolution_sheet.freeze_panes = "A2"
    decision_register_sheet.freeze_panes = "A2"
    widths = {
        "A": 12,
        "B": 28,
        "C": 34,
        "D": 24,
        "E": 32,
        "F": 34,
        "G": 30,
        "H": 28,
        "I": 34,
        "J": 36,
        "K": 18,
        "L": 48,
        "M": 64,
        "N": 72,
        "O": 72,
        "P": 56,
        "Q": 64,
        "R": 64,
        "S": 64,
        "T": 64,
        "U": 64,
        "V": 56,
        "W": 64,
        "X": 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    gate_widths = {
        "A": 12,
        "B": 30,
        "C": 34,
        "D": 34,
        "E": 22,
        "F": 24,
        "G": 30,
        "H": 22,
        "I": 24,
        "J": 22,
        "K": 72,
        "L": 72,
        "M": 72,
        "N": 18,
        "O": 34,
    }
    for column, width in gate_widths.items():
        gate_sheet.column_dimensions[column].width = width
    milestone_widths = {
        "A": 12,
        "B": 52,
        "C": 52,
        "D": 34,
        "E": 86,
    }
    for column, width in milestone_widths.items():
        milestone_sheet.column_dimensions[column].width = width
    next_action_widths = {
        "A": 10,
        "B": 12,
        "C": 12,
        "D": 38,
        "E": 12,
        "F": 78,
        "G": 58,
        "H": 86,
        "I": 86,
        "J": 78,
        "K": 42,
    }
    for column, width in next_action_widths.items():
        next_action_sheet.column_dimensions[column].width = width
    black_box_widths = {
        "A": 12,
        "B": 18,
        "C": 34,
        "D": 48,
        "E": 10,
        "F": 20,
        "G": 24,
        "H": 20,
        "I": 16,
        "J": 18,
        "K": 28,
        "L": 26,
        "M": 24,
        "N": 72,
        "O": 26,
        "P": 38,
        "Q": 18,
        "R": 14,
        "S": 34,
    }
    for column, width in black_box_widths.items():
        black_box_sheet.column_dimensions[column].width = width
    temp_matrix_widths = {
        "A": 18,
        "B": 34,
        "C": 48,
        "D": 20,
        "E": 22,
        "F": 18,
        "G": 16,
        "H": 18,
        "I": 26,
        "J": 42,
        "K": 86,
        "L": 52,
    }
    for column, width in temp_matrix_widths.items():
        temp_matrix_sheet.column_dimensions[column].width = width
    closure_task_widths = {
        "A": 12,
        "B": 18,
        "C": 34,
        "D": 48,
        "E": 30,
        "F": 12,
        "G": 38,
        "H": 28,
        "I": 86,
        "J": 72,
        "K": 72,
        "L": 18,
    }
    for column, width in closure_task_widths.items():
        closure_task_sheet.column_dimensions[column].width = width
    workstream_widths = {
        "A": 12,
        "B": 12,
        "C": 38,
        "D": 12,
        "E": 72,
        "F": 44,
        "G": 72,
        "H": 72,
        "I": 72,
        "J": 72,
    }
    for column, width in workstream_widths.items():
        workstream_sheet.column_dimensions[column].width = width
    report_target_widths = {
        "A": 12,
        "B": 18,
        "C": 34,
        "D": 18,
        "E": 86,
        "F": 42,
        "G": 42,
        "H": 72,
        "I": 48,
        "J": 56,
        "K": 48,
        "L": 72,
        "M": 38,
    }
    for column, width in report_target_widths.items():
        report_target_sheet.column_dimensions[column].width = width
    open_topic_widths = {
        "A": 12,
        "B": 18,
        "C": 34,
        "D": 48,
        "E": 24,
        "F": 72,
        "G": 72,
        "H": 72,
    }
    for column, width in open_topic_widths.items():
        open_topic_sheet.column_dimensions[column].width = width
    tkn_widths = {
        "A": 18,
        "B": 34,
        "C": 42,
        "D": 72,
        "E": 72,
        "F": 34,
        "G": 34,
        "H": 34,
        "I": 42,
        "J": 42,
        "K": 44,
        "L": 54,
        "M": 34,
        "N": 72,
        "O": 32,
        "P": 40,
        "Q": 54,
        "R": 54,
        "S": 64,
        "T": 64,
        "U": 64,
        "V": 18,
        "W": 54,
        "X": 14,
    }
    for column, width in tkn_widths.items():
        tkn_sheet.column_dimensions[column].width = width
    mapping_widths = {
        "A": 18,
        "B": 12,
        "C": 34,
        "D": 42,
        "E": 34,
        "F": 34,
        "G": 42,
        "H": 42,
        "I": 44,
        "J": 54,
        "K": 34,
        "L": 18,
    }
    for column, width in mapping_widths.items():
        mapping_sheet.column_dimensions[column].width = width
    audit_widths = {
        "A": 18,
        "B": 34,
        "C": 34,
        "D": 34,
        "E": 38,
        "F": 34,
        "G": 34,
        "H": 30,
        "I": 30,
        "J": 34,
        "K": 30,
        "L": 20,
        "M": 72,
    }
    for column, width in audit_widths.items():
        audit_sheet.column_dimensions[column].width = width
    db_mapping_widths = {
        "A": 18,
        "B": 34,
        "C": 18,
        "D": 22,
        "E": 34,
        "F": 34,
        "G": 42,
        "H": 72,
        "I": 72,
        "J": 54,
        "K": 54,
        "L": 72,
        "M": 72,
        "N": 26,
    }
    for column, width in db_mapping_widths.items():
        db_mapping_sheet.column_dimensions[column].width = width
    strategy_widths = {
        "A": 24,
        "B": 34,
        "C": 30,
        "D": 72,
        "E": 72,
        "F": 32,
    }
    for column, width in strategy_widths.items():
        strategy_sheet.column_dimensions[column].width = width
    relationship_widths = {
        "A": 22,
        "B": 18,
        "C": 34,
        "D": 48,
        "E": 36,
        "F": 72,
        "G": 72,
        "H": 72,
    }
    for column, width in relationship_widths.items():
        relationship_sheet.column_dimensions[column].width = width
    selected_relationship_widths = {
        "A": 12,
        "B": 18,
        "C": 34,
        "D": 42,
        "E": 24,
        "F": 18,
        "G": 22,
        "H": 34,
        "I": 48,
        "J": 36,
        "K": 72,
        "L": 72,
        "M": 72,
    }
    for column, width in selected_relationship_widths.items():
        selected_relationship_sheet.column_dimensions[column].width = width
    resolution_widths = {
        "A": 18,
        "B": 42,
        "C": 42,
        "D": 72,
        "E": 72,
        "F": 72,
    }
    for column, width in resolution_widths.items():
        resolution_sheet.column_dimensions[column].width = width
    selected_resolution_widths = {
        "A": 12,
        "B": 18,
        "C": 34,
        "D": 42,
        "E": 24,
        "F": 18,
        "G": 42,
        "H": 42,
        "I": 72,
        "J": 72,
        "K": 72,
    }
    for column, width in selected_resolution_widths.items():
        selected_resolution_sheet.column_dimensions[column].width = width
    workbook.save(output_path)
    return output_path


def record_detail_sections(record: FoqAlignmentRecord) -> dict[str, str]:
    node = test_knowledge_node_from_record(record)
    mapping = cross_kb_mapping_from_record(record)
    audit = tkn_coverage_audit_from_node(node)
    return {
        "TKN Node": "\n".join(
            [
                f"test_id: {node.test_id}",
                f"test_name: {node.test_name}",
                f"foq_section: {node.foq_section}",
                f"injection: {node.injection or '(not bound)'}",
                f"instrument_method: {node.instrument_method or '(not bound)'}",
                f"processing_method: {node.processing_method or '(not bound)'}",
                f"report_template: {node.report_template or '(not bound)'}",
                f"report_sheets: {node.report_sheet_label or '(not mapped)'}",
                f"formula_id: {node.formula_id}",
                f"db_fields: {node.db_field_label or '(not mapped)'}",
                f"model_applicability: {node.model_label}",
                f"coverage_status: {node.coverage_status}",
                f"irc_injected: {'Yes' if node.irc_injected else 'No'}",
                "",
                "acceptance_criteria",
                "\n".join(f"- {item}" for item in node.acceptance_criteria) or "- not normalized yet",
                "",
                "dependencies",
                "\n".join(f"- {item}" for item in node.dependencies) or "- none recorded",
                "",
                "expected_ret_times",
                "\n".join(f"- {item}" for item in node.expected_ret_times) or "- none",
                "",
                "expected_channels",
                "\n".join(f"- {item}" for item in node.expected_channels) or "- none",
                "",
                "expected_audit_properties",
                "\n".join(f"- {item}" for item in node.expected_audit_properties) or "- none",
                "",
                "open_gaps",
                "\n".join(f"- {item}" for item in node.open_gaps) or "- none recorded",
            ]
        ),
        "Cross-KB Mapping": "\n".join(
            [
                f"test_id: {mapping.test_id}",
                f"FOQ: {mapping.foq_test_name} [{mapping.foq_section}]",
                f"Method: {mapping.method_name or '(not bound)'}",
                f"Processing: {mapping.processing_method or '(not bound)'}",
                f"Report: {mapping.report_template or '(not bound)'}",
                f"Report Sheets: {mapping.report_sheet_label or '(not mapped)'}",
                f"Formula: {mapping.formula_id}",
                f"DB Fields: {mapping.db_field_label or '(not mapped)'}",
                f"Models: {mapping.model_label}",
                f"Status: {mapping.mapping_status}",
            ]
        ),
        "TD Meaning": "\n".join(
            [
                f"Family: {record.family}",
                f"TestIntent: {record.test_intent}",
                f"TD Test: {record.td_test}",
                f"Device: {record.device_label}",
                f"TD Source: {record.td_source}",
                "",
                record.td_meaning,
                "",
                "Key Conditions",
                "\n".join(f"- {item}" for item in record.key_conditions) or "- none recorded",
            ]
        ),
        "Method Evidence": "\n".join(
            [
                f"Injection: {record.injection or '(not bound)'}",
                f"Instrument Method: {record.instrument_method or '(not bound)'}",
                f"Processing Method: {record.processing_method or '(not bound)'}",
                "",
                "Expected RetTimes",
                "\n".join(f"- {item}" for item in record.expected_ret_times) or "- none",
                "",
                "Expected Channels",
                "\n".join(f"- {item}" for item in record.expected_channels) or "- none",
                "",
                "Expected Audit Properties",
                "\n".join(f"- {item}" for item in record.expected_audit_properties) or "- none",
                "",
                "Method Evidence",
                "\n".join(f"- {item}" for item in record.method_evidence) or "- open verification required",
                "",
                "CMBX Sources",
                "\n".join(f"- {item}" for item in record.cmbx_sources) or "- no loaded matching CMBX evidence",
            ]
        ),
        "Report Evidence": "\n".join(
            [
                f"Report Template: {record.report_template or '(open verification required)'}",
                f"Report Sheets: {record.report_sheet_label or '(not mapped)'}",
                "",
                "Formula / Workbook Evidence",
                "\n".join(f"- {item}" for item in record.report_evidence) or "- open verification required",
            ]
        ),
        "DB Evidence": "\n".join(
            [
                f"DB Fields: {record.db_field_label or '(not mapped)'}",
                "",
                "DB Trace / Type / Precision Evidence",
                "\n".join(f"- {item}" for item in record.db_evidence) or "- open verification required",
            ]
        ),
        "Design Actions": "\n".join(
            [
                f"Modifiability: {record_modifiability_summary(record)}",
                "",
                "Known Cut / Merge / Lock Points",
                "\n".join(f"- {item}" for item in record_cut_point_lines(record)),
                "",
                "Intent Tool Readiness",
                "- Search / recommendation: available from this alignment row.",
                "- Compare: available when at least two device models are selected.",
                "- Crop / merge preview: requires the locked points above to be satisfied.",
                "- CMBX generation: blocked unless Method, Processing, Report, DB, and Config contracts are complete.",
            ]
        ),
        "Milestone Status": "\n".join(record_milestone_status_lines(record)),
        "Next Action Queue": "\n".join(record_next_action_queue_lines(record)),
        "M2 Temperature Matrix": "\n".join(record_temperature_contract_matrix_lines(record)),
        "M2 Closure Tasks": "\n".join(record_contract_closure_task_lines(record)),
        "M2 Evidence Workstreams": "\n".join(record_evidence_workstream_lines(record)),
        "M2 P1 Extraction Plan": "\n".join(record_p1_evidence_extraction_plan_lines(record)),
        "M2 Processing Targets": "\n".join(record_processing_method_target_lines(record)),
        "M2 Report Targets": "\n".join(record_report_formula_target_lines(record)),
        "M2 Report Extraction Plan": "\n".join(record_report_formula_extraction_plan_lines(record)),
        "BlackBox Audit": "\n".join(record_black_box_audit_lines(record)),
        "Open Verification Topics": "\n".join(record_open_verification_topic_lines(record)),
        "Dependency Impact": "\n".join(record_dependency_graph_lines(record)),
        "Relationship Audit": "\n".join(record_relationship_audit_lines(record)),
        "Resolution Choices": "\n".join(record_resolution_choice_lines(record)),
        "Open Verification": "\n".join(record_verification_action_lines(record)),
        "Generation Readiness": "\n".join(
            [
                f"Coverage: {record.coverage_status}",
                f"Modifiability: {record_modifiability_summary(record)}",
                f"Audit Overall: {audit.overall_status}",
                "",
                "Coverage Audit",
                f"- Method command: {audit.method_command_status}",
                f"- Processing method: {audit.processing_method_status}",
                f"- Report/formula: {audit.report_formula_status}",
                f"- DB field: {audit.db_field_status}",
                f"- RetTimes: {audit.ret_time_status}",
                f"- Channels: {audit.channel_status}",
                f"- Audit/properties: {audit.audit_property_status}",
                f"- Configuration: {audit.config_status}",
                "",
                "Required Configuration",
                "\n".join(f"- {item}" for item in record.required_config) or "- not recorded",
                "",
                "Open Gaps",
                "\n".join(f"- {item}" for item in audit.gaps) or "- none recorded",
                "",
                "Generation Readiness",
                record.generation_readiness,
            ]
        ),
    }


def base_alignment_records() -> tuple[FoqAlignmentRecord, ...]:
    records = [
        *(_tcc_records()),
        *(_vdad_records()),
    ]
    return tuple(records)


def _with_kb_file_status(records: Iterable[FoqAlignmentRecord], kb_root: Path) -> tuple[FoqAlignmentRecord, ...]:
    result: list[FoqAlignmentRecord] = []
    for record in records:
        source_name = Path(record.td_source).name
        matches = tuple(kb_root.rglob(source_name)) if kb_root.exists() else ()
        if matches:
            result.append(replace(record, td_source=str(matches[0])))
        else:
            gaps = (*record.open_gaps, f"KB source file not found below {kb_root}: {source_name}")
            result.append(_recompute_status(replace(record, open_gaps=gaps)))
    return tuple(result)


def _enrich_with_loaded_packages(record: FoqAlignmentRecord, packages: Iterable[Any]) -> FoqAlignmentRecord:
    sources: list[str] = list(record.cmbx_sources)
    gaps = list(record.open_gaps)
    method_evidence = list(record.method_evidence)
    report_evidence = list(record.report_evidence)
    found_injection = False
    found_method = False
    found_processing = False
    found_report = False

    for package in packages:
        package_path = str(getattr(package, "path", ""))
        injections = getattr(package, "injections", ()) or ()
        methods_and_reports = getattr(package, "methods_and_reports", ()) or ()
        try:
            injection_links = build_injection_method_links(package)
        except Exception as exc:
            injection_links = {}
            gaps.append(f"{Path(package_path).name}: could not decode injection-method links ({exc}).")
        for injection in injections:
            injection_name = getattr(injection, "name", "")
            if _same_name(injection_name, record.injection):
                found_injection = True
                link = injection_links.get(injection_name)
                if link:
                    sources.append(
                        f"{Path(package_path).name}: link {link.injection_name} -> "
                        f"{link.instrument_method} / {link.processing_method}"
                    )
                    if record.instrument_method and not _same_name(link.instrument_method, record.instrument_method):
                        gaps.append(
                            f"{Path(package_path).name}: injection {injection_name} links to instrument method "
                            f"{link.instrument_method}, expected {record.instrument_method}."
                        )
                    if record.processing_method and not _same_name(link.processing_method, record.processing_method):
                        gaps.append(
                            f"{Path(package_path).name}: injection {injection_name} links to processing method "
                            f"{link.processing_method}, expected {record.processing_method}."
                        )
                    method_evidence.append(
                        f"CMBX sequence link confirms injection {link.injection_name} uses "
                        f"{link.instrument_method} with processing {link.processing_method}."
                    )
                else:
                    sources.append(f"{Path(package_path).name}: injection {injection_name}")
                    gaps.append(f"{Path(package_path).name}: injection {injection_name} has no decoded method link.")
        for element in methods_and_reports:
            name = getattr(element, "name", "")
            kind = getattr(element, "kind", "")
            if kind == "instrument_method" and _same_name(name, record.instrument_method):
                found_method = True
                method_evidence.append(f"Loaded CMBX contains instrument method {name}.")
            if kind == "processing_method" and record.processing_method and _same_name(name, record.processing_method):
                found_processing = True
                evidence = inspect_processing_method(package, name)
                method_evidence.append(f"Loaded CMBX contains processing method {name}.")
                method_evidence.append(f"Processing inspector: {Path(package_path).name}: {evidence.summary}")
                method_evidence.extend(f"Processing inspector detail: {line}" for line in evidence.to_lines()[:14])
            if kind == "report_template" and record.report_template and _same_name(name, record.report_template):
                found_report = True
                report_evidence.append(f"Loaded CMBX contains report template {name}.")

    if packages:
        if record.injection and not found_injection:
            gaps.append(f"Loaded CMBX packages do not show injection {record.injection}.")
        if record.instrument_method and not found_method:
            gaps.append(f"Loaded CMBX packages do not show instrument method {record.instrument_method}.")
        if record.processing_method and not found_processing:
            gaps.append(f"Loaded CMBX packages do not show processing method {record.processing_method}.")
        if record.report_template and not found_report:
            gaps.append(f"Loaded CMBX packages do not show report template {record.report_template}.")

    return _recompute_status(
        replace(
            record,
            method_evidence=_unique(method_evidence),
            report_evidence=_unique(report_evidence),
            cmbx_sources=_unique(sources),
            open_gaps=_unique(gaps),
        )
    )


def _recompute_status(record: FoqAlignmentRecord) -> FoqAlignmentRecord:
    if record.coverage_status == "not applicable":
        return record
    if not record.open_gaps:
        return replace(record, coverage_status="complete")
    if record.injection and record.instrument_method and record.report_sheets and record.db_fields:
        return replace(record, coverage_status="partial")
    if record.injection or record.instrument_method or record.report_sheets or record.db_fields:
        return replace(record, coverage_status="open verification")
    return replace(record, coverage_status="missing")


def _test_node_id(record: FoqAlignmentRecord) -> str:
    known_ids = {
        ("TCC", "column_id"): "TCC_COL_01",
        ("TCC", "preheater_connection"): "TCC_PREHEATER_01",
        ("TCC", "valve_keypad"): "TCC_VALVE_01",
        ("TCC", "burn_in"): "TCC_BURNIN_01",
        ("TCC", "temperature_calibration"): "TCC_CAL_01",
        ("TCC", "temperature_accuracy"): "TCC_ACC_01",
        ("TCC", "temperature_precision_and_fan"): "TCC_PRECISION_01",
        ("TCC", "temperature_stability_and_pcc"): "TCC_STABILITY_PCC_01",
        ("TCC", "temperature_stability_no_pcc"): "TCC_STABILITY_01",
        ("TCC", "heatup_cooldown_20_50_20"): "TCC_HEATCOOL_01",
        ("TCC", "liquid_leak_keypad"): "TCC_LEAK_01",
        ("TCC", "qualification_service_done"): "TCC_SERVICE_01",
        ("TCC", "factory_default_metadata"): "TCC_FACTORY_01",
        ("TCC", "error_log_check"): "TCC_ERRORLOG_01",
        ("VDAD", "warm_up"): "VDAD_WARMUP_01",
        ("VDAD", "noise"): "VDAD_NOISE_01",
        ("VDAD", "linearity"): "VDAD_LINEARITY_01",
        ("VDAD", "wavelength_accuracy"): "VDAD_WAVELENGTH_01",
        ("VDAD", "3d_field"): "VDAD_3DFIELD_01",
    }
    known = known_ids.get((record.family, record.test_intent))
    if known:
        return known
    slug = _normalize_name(record.test_intent).upper() or "TEST"
    return f"{record.family}_{slug}_{record.order:03d}"


def _foq_section(record: FoqAlignmentRecord) -> str:
    known_sections = {
        ("TCC", "column_id"): "TCC FOQ sequence row 01 / TD Column ID",
        ("TCC", "preheater_connection"): "TCC FOQ sequence row 02 / TD Preheater Connection Test",
        ("TCC", "valve_keypad"): "TCC FOQ sequence row 03 / TD Valve and Keypad",
        ("TCC", "burn_in"): "TCC FOQ sequence row 04 / TD BurnIn",
        ("TCC", "temperature_calibration"): "TCC FOQ sequence row 05 / TD Temperature Calibration",
        ("TCC", "temperature_accuracy"): "TCC FOQ sequence row 06 / TD Temperature Accuracy",
        ("TCC", "temperature_precision_and_fan"): "TCC FOQ sequence row 07 / TD Temperature Precision and Fan",
        ("TCC", "temperature_stability_and_pcc"): "TCC FOQ sequence row 08 / TD Temperature Stability and PCC",
        ("TCC", "temperature_stability_no_pcc"): "TCC FOQ non-PCC branch / TD Temperature Stability",
        ("TCC", "heatup_cooldown_20_50_20"): "TCC FOQ sequence row 09 / TD HeatUp and CoolDown",
        ("TCC", "liquid_leak_keypad"): "TCC FOQ sequence row 10 / TD Liquid Leak and Keypad",
        ("TCC", "qualification_service_done"): "TCC FOQ sequence row 11 / TD Qualification Service Done",
        ("TCC", "factory_default_metadata"): "TCC FOQ sequence row 12 / TD Factory Default",
        ("TCC", "error_log_check"): "TCC FOQ sequence row 13 / TD Error Log Check",
    }
    return known_sections.get((record.family, record.test_intent), f"{record.family} TD item: {record.td_test}")


def _acceptance_criteria(record: FoqAlignmentRecord) -> tuple[str, ...]:
    known_criteria = {
        ("TCC", "column_id"): (
            "External: Column_A/B/C/D descriptions must match A/B/C/D.",
        ),
        ("TCC", "preheater_connection"): (
            "External: Preheater port RetTimes are present.",
            "External: ModulePresent = Yes and MemoryState = OK for configured preheater ports.",
        ),
        ("TCC", "temperature_accuracy"): (
            "External: max absolute temperature deviation <= Definitions!Temperature Accuracy.",
        ),
        ("TCC", "temperature_precision_and_fan"): (
            "External: max(lower sensor range, upper sensor range) <= Definitions!Temperature Precision.",
        ),
        ("TCC", "temperature_stability_and_pcc"): (
            "External: max(lower sensor range, upper sensor range) <= Definitions!Temperature Stability.",
            "External: PCC cool-down delta <= Definitions!PCC CoolDownTime.",
        ),
        ("TCC", "temperature_stability_no_pcc"): (
            "External: max(lower sensor range, upper sensor range) <= Definitions!Temperature Stability.",
        ),
        ("TCC", "heatup_cooldown_20_50_20"): (
            "External: HeatUp_Time_20to50 <= Definitions!HeatUp & Cool Down.",
            "External: CoolDown_Time_50to20 <= Definitions!HeatUp & Cool Down.",
        ),
        ("TCC", "factory_default_metadata"): (
            "External/Internal split not fully normalized: ModelNo, firmware, hardware, model variant, and serial-number checks come from audit/precondition metadata.",
        ),
    }
    criteria = known_criteria.get((record.family, record.test_intent))
    if criteria:
        return criteria
    if record.db_evidence:
        return tuple(f"Not normalized from FOQ KB yet: {item}" for item in record.db_evidence)
    return ("Not normalized from FOQ KB yet.",)


def _formula_id(record: FoqAlignmentRecord) -> str:
    known_formulas = {
        ("TCC", "column_id"): "FORMULA_TCC_COLUMN_ID_AUDIT_DESCRIPTION",
        ("TCC", "preheater_connection"): "FORMULA_TCC_PREHEATER_PORT_STATE_AND_DIFF",
        ("TCC", "valve_keypad"): "FORMULA_TCC_VALVE_KEYPAD_AUDIT_STATE_CHECKS",
        ("TCC", "burn_in"): "FORMULA_NOT_REQUIRED_TCC_BURNIN_CONDITIONING",
        ("TCC", "temperature_accuracy"): "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION",
        ("TCC", "temperature_precision_and_fan"): "FORMULA_TCC_TEMP_PRECISION_SEPARATE_SENSOR_RANGE",
        ("TCC", "temperature_stability_and_pcc"): "FORMULA_TCC_TEMP_STABILITY_AND_PCC_COOLDOWN",
        ("TCC", "temperature_stability_no_pcc"): "FORMULA_TCC_TEMP_STABILITY_SEPARATE_SENSOR_RANGE",
        ("TCC", "heatup_cooldown_20_50_20"): "FORMULA_TCC_HEATUP_COOLDOWN_RETTIME_DELTA_MINUS_HOLD",
        ("TCC", "liquid_leak_keypad"): "FORMULA_TCC_LIQUID_LEAK_AUDIT_PRECOND_CHECK",
        ("TCC", "qualification_service_done"): "FORMULA_NOT_REQUIRED_TCC_SERVICE_AUDIT_LOG",
        ("TCC", "factory_default_metadata"): "FORMULA_TCC_FACTORY_DEFAULT_METADATA",
        ("TCC", "error_log_check"): "FORMULA_NOT_REQUIRED_TCC_ERROR_LOG_AUDIT_TABLE",
    }
    known = known_formulas.get((record.family, record.test_intent))
    if known:
        return known
    if record.report_sheets or record.db_fields:
        return f"FORMULA_{record.family}_{_normalize_name(record.test_intent).upper()}_OPEN"
    return "FORMULA_OPEN_VERIFICATION_REQUIRED"


def _dependencies(record: FoqAlignmentRecord) -> tuple[str, ...]:
    values: list[str] = []
    values.extend(record.key_conditions)
    values.extend(record.required_config)
    values.extend(f"RetTime dependency: {item}" for item in record.expected_ret_times)
    values.extend(f"Channel dependency: {item}" for item in record.expected_channels)
    values.extend(f"Audit/config dependency: {item}" for item in record.expected_audit_properties)
    return _unique(values)


def _device_bindings_from_record(record: FoqAlignmentRecord) -> tuple[TestKnowledgeNodeDeviceBinding, ...]:
    return tuple(
        TestKnowledgeNodeDeviceBinding(
            device_model=device_model,
            injection=_device_injection(record, device_model),
            instrument_method=_device_instrument_method(record, device_model),
            processing_method=_device_processing_method(record, device_model),
            report_template=_device_report_template(record, device_model),
            report_sheets=_device_report_sheets(record, device_model),
            report_files=_device_report_files(record, device_model),
            db_fields=_device_db_fields(record, device_model),
        )
        for device_model in record.device_models
    )


def _node_db_fields(record: FoqAlignmentRecord) -> tuple[str, ...]:
    fields: list[str] = []
    for device_model in record.device_models:
        fields.extend(_device_db_fields(record, device_model))
    return _unique(fields)


def _device_binding_for_model(node: TestKnowledgeNode, device_model: str) -> TestKnowledgeNodeDeviceBinding | None:
    for binding in node.device_bindings:
        if binding.device_model == device_model:
            return binding
    return None


def _device_injection(record: FoqAlignmentRecord, device_model: str) -> str:
    if record.family == "TCC":
        blueprint = _tcc_sequence_blueprint(record.test_intent, device_model)
        if blueprint:
            return blueprint[0]
    if record.family == "TCC" and record.test_intent == "temperature_accuracy":
        if device_model == "VH-C10-A":
            return "Temperature Accuracy_H"
        if device_model in {"VC-C10-A", "VA-C10-A"}:
            return "Temperature Accuracy_C"
    if record.family == "TCC" and record.test_intent == "temperature_stability_no_pcc":
        if device_model in {"VC-C10-A", "VA-C10-A"}:
            return "Temperature Stability_C"
    return record.injection


def _device_instrument_method(record: FoqAlignmentRecord, device_model: str) -> str:
    if record.family == "TCC":
        blueprint = _tcc_sequence_blueprint(record.test_intent, device_model)
        if blueprint:
            return blueprint[1]
    return record.instrument_method


def _device_processing_method(record: FoqAlignmentRecord, device_model: str) -> str:
    if record.family == "TCC":
        blueprint = _tcc_sequence_blueprint(record.test_intent, device_model)
        if blueprint:
            return blueprint[2]
    if record.family == "TCC" and record.test_intent == "temperature_accuracy":
        if device_model == "VH-C10-A":
            return "ACCURACY_IRC_STOP_H"
        if device_model in {"VC-C10-A", "VA-C10-A"}:
            return "ACCURACY_IRC_STOP_C"
    return record.processing_method


def _tcc_sequence_blueprint(test_intent: str, device_model: str) -> tuple[str, str, str] | None:
    # Evidence source: TCC_FOQ_METHOD_REPORT_KNOWLEDGE_BASE.md observed
    # VA/VC/VH sequence tables. Keep this as binding evidence, not as command
    # semantics; method command interpretation is tracked separately.
    va = {
        "valve_keypad": ("Valve", "VALVES", "No_Integration"),
        "burn_in": ("VTCC_BurnIn", "BURNIN", "NO_INTEGRATION"),
        "temperature_calibration": ("Temperature Calibration", "TEMPERATURE_CALIBRATION", "NO_INTEGRATION"),
        "temperature_accuracy": ("Temperature Accuracy_C", "TEMPERATURE_ACCURACY", "ACCURACY_IRC_STOP_C"),
        "temperature_precision_and_fan": ("Temperature Precision", "TEMPERATURE_PRECISION", "NO_INTEGRATION"),
        "temperature_stability_no_pcc": ("Temperature Stability_C", "TEMPERATURE_STABILITY_70_C", "NO_INTEGRATION"),
        "heatup_cooldown_20_50_20": ("HeatUp and CoolDownTime", "TEMP_HEAT_UP_DOWN_20_50_20", "No_Integration"),
        "liquid_leak_keypad": ("LiquidLeaktest", "LIQUID LEAK", "No_Integration"),
        "qualification_service_done": ("Qualification_Service_Done", "Qualification_Service_Done", "No_Integration"),
        "factory_default_metadata": ("Factory Default", "FACTORYDEFAULT", "No_Integration"),
        "error_log_check": ("Error Log Check", "CHECKERRORLOG", "No_Integration"),
    }
    vc = {
        "column_id": ("ColumnIDs", "ColumnID", "CORRECT_ACCURACY_INJ_INSERTION"),
        "preheater_connection": ("Preheater Connection Test", "PREHEATER", "CORRECT_ACCURACY_INJ_INSERTION"),
        "valve_keypad": ("Valve", "VALVES", "CORRECT_ACCURACY_INJ_INSERTION"),
        "burn_in": ("VTCC_BurnIn", "BURNIN", "NO_INTEGRATION"),
        "temperature_calibration": ("Temperature Calibration", "TEMPERATURE_CALIBRATION", "CORRECT_ACCURACY_INJ_INSERTION"),
        "temperature_accuracy": ("Temperature Accuracy_C", "TEMPERATURE_ACCURACY", "ACCURACY_IRC_STOP_C"),
        "temperature_precision_and_fan": ("Temperature Precision_and_Fan", "TEMPERATURE_PRECISION_AND_FAN", "CORRECT_STABILITY_INJ_INSERTION"),
        "temperature_stability_no_pcc": ("Temperature Stability_C", "TEMPERATURE_STABILITY_70_C", "NO_INTEGRATION"),
        "heatup_cooldown_20_50_20": ("HeatUp and CoolDownTime", "TEMP_HEAT_UP_DOWN_20_50_20", "CORRECT_ACCURACY_INJ_INSERTION"),
        "liquid_leak_keypad": ("LiquidLeaktest", "LIQUID LEAK", "CORRECT_ACCURACY_INJ_INSERTION"),
        "qualification_service_done": ("Qualification_Service_Done", "Qualification_Service_Done", "CORRECT_ACCURACY_INJ_INSERTION"),
        "factory_default_metadata": ("Factory Default", "FACTORYDEFAULT", "CORRECT_ACCURACY_INJ_INSERTION"),
        "error_log_check": ("Error Log Check", "CHECKERRORLOG", "CORRECT_ACCURACY_INJ_INSERTION"),
    }
    vh = {
        "column_id": ("ColumnIDs", "ColumnID", "CORRECT_STABILITY_INJ_INSERTION"),
        "preheater_connection": ("Preheater Connection Test", "PREHEATER", "No_Integration"),
        "valve_keypad": ("Valve", "VALVES", "No_Integration"),
        "burn_in": ("VTCC_BurnIn", "BURNIN", "NO_INTEGRATION"),
        "temperature_calibration": ("Temperature Calibration", "TEMPERATURE_CALIBRATION", "CORRECT_ACCURACY_INJ_INSERTION"),
        "temperature_accuracy": ("Temperature Accuracy_H", "TEMPERATURE_ACCURACY", "ACCURACY_IRC_STOP_H"),
        "temperature_precision_and_fan": ("Temperature Precision_and_Fan", "TEMPERATURE_PRECISION_AND_FAN", "CORRECT_STABILITY_INJ_INSERTION"),
        "temperature_stability_and_pcc": ("Temperature Stability_and_PCC_H", "TEMPERATURE_STABILITY_AND_PCC_70_H", "NO_INTEGRATION"),
        "heatup_cooldown_20_50_20": ("HeatUp and CoolDownTime", "TEMP_HEAT_UP_DOWN_20_50_20", "No_Integration"),
        "liquid_leak_keypad": ("LiquidLeaktest", "LIQUID LEAK", "No_Integration"),
        "qualification_service_done": ("Qualification_Service_Done", "Qualification_Service_Done", "No_Integration"),
        "factory_default_metadata": ("Factory Default", "FACTORYDEFAULT", "No_Integration"),
        "error_log_check": ("Error Log Check", "CHECKERRORLOG", "No_Integration"),
    }
    tables = {
        "VA-C10-A": va,
        "VC-C10-A": vc,
        "VH-C10-A": vh,
    }
    return tables.get(device_model, {}).get(test_intent)


def _device_report_template(record: FoqAlignmentRecord, device_model: str) -> str:
    if record.family == "TCC" and record.report_template.startswith("Device-specific:"):
        if device_model == "VA-C10-A":
            return "Report_VATCC_V1_01"
        if device_model in {"VC-C10-A", "VH-C10-A"}:
            return "Report_VTCC_V2_12"
    return record.report_template


def _device_report_sheets(record: FoqAlignmentRecord, device_model: str) -> tuple[str, ...]:
    if record.family == "TCC" and record.test_intent == "temperature_calibration":
        return ("Temp_Calib_Internal",)
    if record.family == "TCC" and record.test_intent == "temperature_precision_and_fan" and device_model == "VA-C10-A":
        return ("Temp Precision",)
    return record.report_sheets


def _device_report_files(record: FoqAlignmentRecord, device_model: str) -> tuple[str, ...]:
    if record.family == "TCC" and record.test_intent == "temperature_precision_and_fan" and device_model == "VA-C10-A":
        return ("Temperature Precision_and_Fan.XLS",)
    return ()


def _device_db_fields(record: FoqAlignmentRecord, device_model: str) -> tuple[str, ...]:
    if record.family == "TCC" and record.test_intent == "temperature_accuracy":
        if device_model == "VH-C10-A":
            return ("TempAcc10", "TempAcc20", "TempAcc40", "TempAcc80", "TempAcc120", "RES_TempAccuracy")
        if device_model in {"VC-C10-A", "VA-C10-A"}:
            return ("TempAcc10", "TempAcc20", "TempAcc40", "TempAcc60", "TempAcc85", "RES_TempAccuracy")
    if record.family == "TCC" and record.test_intent == "preheater_connection":
        return (
            "Noise_PrehtLeft_Temp",
            "Noise_PrehtRight_Temp",
            "Diff_PhLeft_HtTmp",
            "Diff_PhRight_HtTmp",
            "RES_Preheater_Left_Port",
            "RES_Preheater_Right_Port",
        )
    if record.family == "TCC" and record.test_intent == "temperature_calibration":
        if device_model == "VH-C10-A":
            return (
                "TempCal120_U",
                "TempCal120_L",
                "TempCal100_U",
                "TempCal100_L",
                "TempCal80_U",
                "TempCal80_L",
                "TempCal60_U",
                "TempCal60_L",
                "TempCal40_U",
                "TempCal40_L",
                "TempCal20_U",
                "TempCal20_L",
                "TempCal10_U",
                "TempCal10_L",
                "TempCal5_U",
                "TempCal5_L",
                "TimeCal120",
                "TimeCal100",
                "TimeCal80",
                "TimeCal60",
                "TimeCal40",
                "TimeCal20",
                "TimeCal10",
                "TimeCal05",
                "Slope_Cal120_U",
                "Slope_Cal100_U",
                "Slope_Cal80_U",
                "Slope_Cal60_U",
                "Slope_Cal40_U",
                "Slope_Cal20_U",
                "Slope_Cal10_U",
                "Slope_Cal05_U",
                "Slope_Cal120_L",
                "Slope_Cal100_L",
                "Slope_Cal80_L",
                "Slope_Cal60_L",
                "Slope_Cal40_L",
                "Slope_Cal20_L",
                "Slope_Cal10_L",
                "Slope_Cal05_L",
            )
        if device_model in {"VC-C10-A", "VA-C10-A"}:
            return (
                "TempCal85_U",
                "TempCal85_L",
                "TempCal70_U",
                "TempCal70_L",
                "TempCal55_U",
                "TempCal55_L",
                "TempCal40_U",
                "TempCal40_L",
                "TempCal30_U",
                "TempCal30_L",
                "TempCal20_U",
                "TempCal20_L",
                "TempCal10_U",
                "TempCal10_L",
                "TempCal5_U",
                "TempCal5_L",
                "TimeCal85",
                "TimeCal70",
                "TimeCal55",
                "TimeCal40",
                "TimeCal30",
                "TimeCal20",
                "TimeCal10",
                "TimeCal05",
                "Slope_Cal85_U",
                "Slope_Cal70_U",
                "Slope_Cal55_U",
                "Slope_Cal40_U",
                "Slope_Cal30_U",
                "Slope_Cal20_U",
                "Slope_Cal10_U",
                "Slope_Cal05_U",
                "Slope_Cal85_L",
                "Slope_Cal70_L",
                "Slope_Cal55_L",
                "Slope_Cal40_L",
                "Slope_Cal30_L",
                "Slope_Cal20_L",
                "Slope_Cal10_L",
                "Slope_Cal05_L",
            )
    if record.family == "TCC" and record.test_intent == "temperature_precision_and_fan":
        return ("TempPrecision", "RES_TempPrecision")
    if record.family == "TCC" and record.test_intent == "temperature_stability_and_pcc":
        return (
            "TempStability",
            "Noise_CC_Temp",
            "Noise_PCC_Temp",
            "Performance_PCC",
            "RES_TempStability",
            "RES_PCC",
            "PCC_Acc_40_Step1",
            "PCC_Acc_80",
            "PCC_Acc_40_Step2",
            "PCC_Drift",
        )
    if record.family == "TCC" and record.test_intent == "temperature_stability_no_pcc":
        return ("TempStability", "Noise_CC_Temp", "RES_TempStability")
    if record.family == "TCC" and record.test_intent == "factory_default_metadata":
        return (
            "TestDate",
            "Serial",
            "TimeBase",
            "ModelNo",
            "ModelVariant",
            "HardwareVersion",
            "Firmware",
            "SubmitDate",
            "RES_SN_Check",
        )
    if record.family == "TCC" and record.test_intent in {
        "valve_keypad",
        "liquid_leak_keypad",
        "qualification_service_done",
        "error_log_check",
    }:
        return ()
    return record.db_fields


def _is_irc_processing(processing_method: str) -> bool:
    normalized = processing_method.upper()
    return "IRC" in normalized or normalized.startswith("CORRECT_")


def _cross_kb_mapping_status(record: FoqAlignmentRecord) -> str:
    if record.coverage_status == "not applicable":
        return "not applicable"
    has_foq = bool(record.td_test and record.td_source)
    has_method = bool(record.instrument_method)
    has_report = bool(record.report_template or record.report_sheets)
    has_formula = _formula_id(record) != "FORMULA_OPEN_VERIFICATION_REQUIRED"
    if has_foq and has_method and has_report and has_formula and record.coverage_status == "complete":
        return "complete"
    if has_foq and has_method and has_report and has_formula:
        return "partial"
    if has_foq and (has_method or has_report or has_formula):
        return "open verification"
    return "missing"


def _tcc_records() -> tuple[FoqAlignmentRecord, ...]:
    source = "FOQ_TCC_VX_C10_A_TD_KNOWLEDGE_MANAGEMENT.md"
    template = "Report_VTCC_V2_12"
    common_config = (
        "ColumnComp",
        "ColumnComp.CC",
        "Thermometer1.ExtTemp_UpperCC",
        "Thermometer1.ExtTemp_LowerCC",
        "AUDIT.ColumnComp.ModelNo",
    )
    return (
        FoqAlignmentRecord(
            10,
            "TCC",
            "temperature_accuracy",
            "Temperature Accuracy",
            ("VH-C10-A", "VC-C10-A", "VA-C10-A"),
            source,
            "在每个 nominal setpoint 稳定后，用外部上下温度计测得温度与 nominal 比较，取更大偏差作为准确度结果。",
            ("External thermometer stability is required.", "Device identity comes from AUDIT.ColumnComp.ModelNo."),
            "Temperature Accuracy_H",
            "TEMPERATURE_ACCURACY",
            "ACCURACY_IRC_STOP_H",
            template,
            ("Temp Accuracy",),
            (
                "K66:K70 use AUDIT.RetTime1..5.",
                "L/M rows average ExtTemp_LowerCC and ExtTemp_UpperCC over RetTimeN-1.0..RetTimeN-0.2.",
                "Workbook selects the larger absolute deviation from nominal.",
            ),
            ("TempAcc20", "TempAcc40", "TempAcc80", "TempAcc120", "RES_TempAccuracy"),
            ("DB cells resolve through Temp Accuracy report cells and display precision rules.",),
            ("RetTime1", "RetTime2", "RetTime3", "RetTime4", "RetTime5"),
            ("ExtTemp_LowerCC", "ExtTemp_UpperCC"),
            ("ColumnComp.CC.Temperature.Nominal", "ColumnComp.ModelNo"),
            common_config,
            (
                "TCC method alignment document records external stability counters and RetTimeN emissions.",
                "RetTime3 is the known 40 C anchor for VH.",
            ),
            (),
            "partial",
            ("Confirm VA/VC/VH template-specific row coverage from loaded CMBX before generation.",),
            "Reusable as a generation template only after the approach/baseline rule for custom single-point tests is confirmed.",
        ),
        FoqAlignmentRecord(
            20,
            "TCC",
            "heatup_cooldown_20_50_20",
            "HeatUp and CoolDown",
            ("VH-C10-A", "VC-C10-A", "VA-C10-A"),
            source,
            "测量 20->50 C 升温和 50->20 C 降温时间，并扣除方法脚本中的 2 min 稳定保持时间。",
            ("Upper external thermometer is the main timing reference.", "Report subtracts 2.0 min from each final duration."),
            "HeatUp and CoolDownTime",
            "TEMP_HEAT_UP_DOWN_20_50_20",
            "No_Integration",
            template,
            ("HeatUp&CoolDown",),
            (
                "Report uses RetTime anchors for heat-up and cool-down boundaries.",
                "Verified evaluator rule: heat-up = RetTime2 - RetTime1 - 2.0; cool-down = RetTime5 - RetTime4 - 2.0.",
                "RetTime3 and RetTime6 remain required for the visible row-65 internal endpoint layout.",
            ),
            ("HeatUp_Time_20to50", "CoolDown_Time_50to20", "RES_HeatUp", "RES_CoolDown"),
            ("DB output uses HeatUp&CoolDown D/E summary cells with one-decimal display for time fields.",),
            ("RetTime1", "RetTime2", "RetTime3", "RetTime4", "RetTime5", "RetTime6"),
            ("ExtTemp_UpperCC",),
            (),
            common_config,
            ("Method evidence must preserve RetTime1..6 meanings; DB timing uses external RetTime2/5 while row 65 keeps internal RetTime3/6 evidence.",),
            (),
            "partial",
            ("Keep row-65/row-66 workbook rule distinction visible until full FormulaOne parsing is complete.",),
            "Good candidate for clone/select generation because the evidence contract is compact and well understood.",
        ),
        FoqAlignmentRecord(
            30,
            "TCC",
            "upper_lower_valve_cycle",
            "Valve / Keypad",
            ("VH-C10-A", "VC-C10-A", "VA-C10-A"),
            source,
            "确认上下阀可以在目标位置间切换，并记录阀位、precision 和 keypad/FastCool 相关行为。",
            ("Upper/lower valve hardware must exist.", "Some keypad behavior remains manual/operator-driven."),
            "Valve",
            "VALVES",
            "No_Integration",
            template,
            ("Valve_Keypad",),
            (
                "Report reads AUDIT.UpperValve/LowerValve.CurrentPosition and Precision at fixed audit times.",
                "FastCool/keypad evidence is audit-driven rather than RetTime-driven.",
            ),
            ("RES_UpperValve", "RES_LowerValve"),
            ("DB field names vary by mapping file; verify exact field list from FOQResultLocations.",),
            (),
            (),
            ("UpperValve.CurrentPosition", "LowerValve.CurrentPosition", "UpperValve.Precision", "LowerValve.Precision"),
            ("ColumnComp.UpperValve", "ColumnComp.LowerValve"),
            ("Method must switch 6_1 -> 1_2 -> 6_1 and log precision.",),
            (),
            "open verification",
            ("Exact DB field names and keypad report formulas need mapping trace confirmation.",),
            "Reusable only for configurations that expose both upper and lower valve symbols.",
        ),
    )


def _make_tcc_record(
    order: int,
    test_intent: str,
    td_test: str,
    devices: tuple[str, ...],
    meaning: str,
    key_conditions: tuple[str, ...],
    injection: str,
    instrument_method: str,
    processing_method: str,
    report_sheets: tuple[str, ...],
    report_evidence: tuple[str, ...] = (),
    db_fields: tuple[str, ...] = (),
    db_evidence: tuple[str, ...] = (),
    ret_times: tuple[str, ...] = (),
    channels: tuple[str, ...] = (),
    audit_properties: tuple[str, ...] = (),
    required_config: tuple[str, ...] = (),
    method_evidence: tuple[str, ...] = (),
    coverage_status: str = "partial",
    open_gaps: tuple[str, ...] = (),
    generation_readiness: str = "Use as alignment evidence only until command flow and report formulas are verified.",
) -> FoqAlignmentRecord:
    common_config = (
        "ColumnComp",
        "AUDIT.ColumnComp.ModelNo is the device source of truth",
    )
    template = "Device-specific: Report_VTCC_V2_12 for VC/VH, Report_VATCC_V1_01 for VA"
    return FoqAlignmentRecord(
        order,
        "TCC",
        test_intent,
        td_test,
        devices,
        "FOQ_TCC_VX_C10_A_TD_KNOWLEDGE_MANAGEMENT.md",
        meaning,
        key_conditions,
        injection,
        instrument_method,
        processing_method,
        template,
        report_sheets,
        report_evidence,
        db_fields,
        db_evidence,
        ret_times,
        channels,
        audit_properties,
        (*common_config, *required_config),
        method_evidence,
        (),
        coverage_status,
        open_gaps,
        generation_readiness,
    )


def _tcc_records() -> tuple[FoqAlignmentRecord, ...]:
    all_tcc = ("VH-C10-A", "VC-C10-A", "VA-C10-A")
    vc_vh = ("VH-C10-A", "VC-C10-A")
    return (
        _make_tcc_record(
            10,
            "column_id",
            "Column ID",
            vc_vh,
            "Verify that the four column ID channels/positions are recognized and reported as A/B/C/D.",
            ("Usually the first injection in the TCC FOQ sequence.", "Uses audit metadata rather than raw signal windows."),
            "ColumnIDs",
            "ColumnID",
            "CORRECT_STABILITY_INJ_INSERTION",
            ("Column ID",),
            (
                "Report cells L46:L49 read AUDIT.Column_A/B/C/D.Description with suffix path matching.",
            ),
            ("RES_ColumnID_A", "RES_ColumnID_B", "RES_ColumnID_C", "RES_ColumnID_D"),
            ("DB fields resolve from Column ID report cells and pass when A/B/C/D descriptions match.",),
            (),
            (),
            (
                "ColumnComp.Column_A.Description",
                "ColumnComp.Column_B.Description",
                "ColumnComp.Column_C.Description",
                "ColumnComp.Column_D.Description",
            ),
            ("Column ID option/configuration enabled",),
            ("CMBX sequence row should bind ColumnIDs -> ColumnID / CORRECT_STABILITY_INJ_INSERTION.",),
            generation_readiness="Good alignment template for audit-only checks; generation needs confirmed column ID configuration.",
        ),
        _make_tcc_record(
            20,
            "preheater_connection",
            "Preheater Connection Test",
            vc_vh,
            "Check that left/right preheater connection ports are present, have valid memory state, and can generate expected heater/external temperature evidence.",
            ("Requires preheater hardware.", "VA devices without preheater hardware should be marked not applicable or open verification."),
            "Preheater Connection Test",
            "PREHEATER",
            "No_Integration",
            ("Preheater Ports_Noise",),
            (
                "Report uses RetTime windows and precond.ColumnComp.PrehtLeft/PrehtRight metadata.",
                "Temperature-difference cells compare heater temperature average against external preheater temperature average.",
            ),
            ("RES_PhLeftPort", "RES_PhRightPort", "Diff_PhLeft_HtTmp", "Diff_PhRight_HtTmp"),
            ("Left/right port pass requires RetTimes plus ModulePresent=Yes and MemoryState=OK.",),
            ("RetTime1", "RetTime2", "RetTime3", "RetTime4"),
            ("ExtTemp_PrehtLeft", "ExtTemp_PrehtRight", "PREH_L_HeaterTemp_Actual", "PREH_R_HeaterTemp_Actual"),
            (
                "precond.ColumnComp.PrehtLeft.ModulePresent",
                "precond.ColumnComp.PrehtLeft.MemoryState",
                "precond.ColumnComp.PrehtRight.ModulePresent",
                "precond.ColumnComp.PrehtRight.MemoryState",
            ),
            ("Preheater left/right modules present in instrument configuration",),
            ("CMBX sequence row should bind Preheater Connection Test -> PREHEATER / No_Integration.",),
            open_gaps=("Confirm exact channel names for VA/VC/VH variants from loaded CMBX evidence.",),
        ),
        _make_tcc_record(
            30,
            "valve_keypad",
            "Valve / Keypad",
            all_tcc,
            "Verify upper/lower valve switching behavior and keypad/FastCool related audit behavior where the device configuration supports it.",
            ("Upper/lower valves are configuration-dependent.", "Some keypad checks are operator or audit driven."),
            "Valve",
            "VALVES",
            "No_Integration",
            ("Valve_Keypad",),
            (
                "Valve_Keypad audit position checks: K49/L49 at -0.05 min, K50/L50 at 0.095 min, K51/L51 at 0.19 min, K60/L60 at 0.9 min.",
                "Valve_Keypad precision checks: U49:U51 read AUDIT.UpperValve.Precision and V49:V51 read AUDIT.LowerValve.Precision.",
                "N60 reads AUDIT.ColumnComp.FastCoolState(0.9,\"backward\") for keypad/FastCool evidence.",
            ),
            ("RES_UpperValve", "RES_LowerValve"),
            ("FOQResultLocations has no DB output rows for Valve; report/audit evidence remains procedure-facing.",),
            (),
            (),
            (
                "UpperValve.CurrentPosition",
                "LowerValve.CurrentPosition",
                "UpperValve.Precision",
                "LowerValve.Precision",
                "ColumnComp.FastCoolState",
            ),
            ("ColumnComp.UpperValve and/or ColumnComp.LowerValve configured", "Front-panel/keypad actions available to operator"),
            (
                "VALVES setup: set CC TempCtrl On, set StillAir, set UpperValve and LowerValve to 6_1, log upper/lower valve precision.",
                "VALVES run: acquire CC_Temp as audit/time carrier, switch both valves 6_1 -> 1_2 -> 6_1, log precision after each switch.",
                "VALVES operator branch: prompt disconnect/reconnect and FAST COOL / upper valve / lower valve keypad actions.",
            ),
            open_gaps=("Workbook pass/fail cells for Valve_Keypad are mapped as audit-state checks but not yet reconstructed as FormulaOne expressions.",),
            generation_readiness="Reusable only for device configurations exposing the required valve symbols.",
        ),
        _make_tcc_record(
            40,
            "burn_in",
            "VTCC BurnIn",
            all_tcc,
            "Condition the module and exercise thermal/control behavior before the final measurement-oriented tests.",
            ("Runs as an instrument method step in the FOQ sequence.", "Usually no DB field is exported directly from this injection."),
            "VTCC_BurnIn",
            "BURNIN",
            "NO_INTEGRATION",
            ("Conditioning / no DB output",),
            (
                "BurnIn is a conditioning/evidence method rather than a formula-heavy DB output injection.",
                "No direct high-value DB mapping formula was found for this row in FOQResultLocations.",
            ),
            (),
            (),
            (),
            ("CC_Temp", "ExtTemp_LowerCC", "ExtTemp_UpperCC"),
            ("ColumnComp.CC.Temperature.Nominal", "ColumnComp.ModelNo"),
            ("Column oven/temperature control configured", "External thermometers available for heating evidence"),
            (
                "BURNIN setup: ReadyTempDelta broad, EquilibrationTime 0.5, CC TempCtrl On, StillAir, LiquidLeakSensor Off.",
                "BURNIN model branch: VH uses approximately 5..120 C thermal range; VA/VC use lower high-limit range, typically up to 85 C.",
                "BURNIN run: stabilize, acquire temperature channels, heat to max, cool to min, heat again, and end after required cycles.",
            ),
            coverage_status="open verification",
            open_gaps=("Exact BURNIN trigger names and abort guard thresholds need method-script line-level extraction before generation.",),
        ),
        _make_tcc_record(
            50,
            "temperature_calibration",
            "Temperature Calibration",
            all_tcc,
            "Collect calibration evidence that supports subsequent temperature accuracy/stability interpretation.",
            ("Precedes Temperature Accuracy in the sequence.", "Uses corrective processing insertion in the existing VH evidence."),
            "Temperature Calibration",
            "TEMPERATURE_CALIBRATION",
            "CORRECT_ACCURACY_INJ_INSERTION",
            ("Temp Calibration", "Internal Use"),
            (
                "Report and DB usage depends on calibration workbook-derived cells.",
            ),
            ("TempCal120_U", "TempCal120_L", "TempCal100_U", "TempCal100_L"),
            ("DB mapping contains multiple TempCal setpoint fields for upper/lower sensors.",),
            ("RetTime1", "RetTime2", "RetTime3", "RetTime4", "RetTime5", "RetTime6", "RetTime7", "RetTime8"),
            ("CC_Temp", "ExtTemp_LowerCC", "ExtTemp_UpperCC"),
            ("ColumnComp.CC.Temperature.Nominal",),
            ("External upper/lower thermometers configured",),
            ("CMBX sequence row should bind Temperature Calibration -> TEMPERATURE_CALIBRATION / CORRECT_ACCURACY_INJ_INSERTION.",),
            open_gaps=("Need workbook-derived low-temperature reach pass/fail trace; VTCC display precision and VA Report_VATCC_V1_01 direct formula-object parity are verified.",),
        ),
        _make_tcc_record(
            60,
            "temperature_accuracy",
            "Temperature Accuracy",
            all_tcc,
            "At each nominal setpoint, compare upper/lower external thermometer readings against nominal and use the larger absolute deviation.",
            ("External thermometer stability is required.", "Device identity comes from AUDIT.ColumnComp.ModelNo."),
            "Temperature Accuracy_H",
            "TEMPERATURE_ACCURACY",
            "ACCURACY_IRC_STOP_H",
            ("Temp Accuracy",),
            (
                "K66:K70 use AUDIT.RetTime1..5.",
                "L/M rows average ExtTemp_LowerCC and ExtTemp_UpperCC over RetTimeN-1.0..RetTimeN-0.2.",
                "Workbook selects the larger absolute deviation from nominal.",
            ),
            ("TempAcc20", "TempAcc40", "TempAcc80", "TempAcc120", "RES_TempAccuracy"),
            ("DB cells resolve through Temp Accuracy report cells and display precision rules.",),
            ("RetTime1", "RetTime2", "RetTime3", "RetTime4", "RetTime5"),
            ("ExtTemp_LowerCC", "ExtTemp_UpperCC"),
            ("ColumnComp.CC.Temperature.Nominal", "ColumnComp.ModelNo"),
            ("External upper/lower thermometers configured",),
            ("RetTime3 is the known 40 C anchor for VH evidence.",),
            open_gaps=("Confirm VA/VC/VH template-specific row coverage from loaded CMBX before generation.",),
            generation_readiness="Reusable as a generation template only after approach/baseline rules for custom single-point tests are confirmed.",
        ),
        _make_tcc_record(
            70,
            "temperature_precision_and_fan",
            "Temperature Precision and Fan",
            all_tcc,
            "Measure repeatability/precision of upper and lower external thermometer readings and verify fan-related behavior where applicable.",
            ("Precision is evaluated per sensor, then the worse sensor range is used.", "Fan evidence is configuration/report dependent."),
            "Temperature Precision_and_Fan",
            "TEMPERATURE_PRECISION_AND_FAN",
            "CORRECT_STABILITY_INJ_INSERTION",
            ("Temp Precision", "Fan"),
            (
                "Correct rule: max range of lower sensor and upper sensor separately, not combined K:L range.",
            ),
            ("TempPrecision", "RES_TempPrecision", "RES_Fan"),
            ("Pass/fail uses raw precision range; displayed summary uses two decimals.",),
            ("RetTime1", "RetTime2", "RetTime3"),
            ("ExtTemp_LowerCC", "ExtTemp_UpperCC", "FanSpeed"),
            ("ColumnComp.CC.Temperature.Nominal",),
            ("External upper/lower thermometers configured", "Fan symbol available if checked"),
            ("CMBX sequence row should bind Temperature Precision_and_Fan -> TEMPERATURE_PRECISION_AND_FAN / CORRECT_STABILITY_INJ_INSERTION.",),
            open_gaps=("Fan pass/fail workbook formula and whether it has any DB fields outside current mapping need trace confirmation; VA Precision/Fan branch behavior is verified from real VATCC report XML.",),
        ),
        _make_tcc_record(
            80,
            "temperature_stability_and_pcc",
            "Temperature Stability and PCC",
            ("VH-C10-A",),
            "Measure long-window temperature stability and PCC cool-down performance on VH configurations with PCC support.",
            ("VH-only PCC performance path.", "Do not mix upper/lower thermometer offset into stability result."),
            "Temperature Stability_and_PCC_H",
            "TEMPERATURE_STABILITY_AND_PCC_70_H",
            "NO_INTEGRATION",
            ("Temp Stability_Noise", "PCC"),
            (
                "Correct stability rule: max range of lower sensor and upper sensor separately.",
                "PCC CoolDown uses K105/L105 RetTime3/RetTime4 delta.",
            ),
            ("TempStability", "Performance_PCC", "RES_TempStability", "RES_PCC"),
            ("PCC performance summary displays two decimals.",),
            ("RetTime1", "RetTime2", "RetTime3", "RetTime4"),
            ("ExtTemp_LowerCC", "ExtTemp_UpperCC", "PCC_Temp"),
            ("ColumnComp.CC.Temperature.Nominal",),
            ("PCC hardware/configuration available",),
            ("CMBX sequence row should bind Temperature Stability_and_PCC_H -> TEMPERATURE_STABILITY_AND_PCC_70_H / NO_INTEGRATION.",),
            generation_readiness="Good VH-only alignment template after PCC command stages are fully documented.",
        ),
        _make_tcc_record(
            90,
            "temperature_stability_no_pcc",
            "Temperature Stability",
            ("VC-C10-A", "VA-C10-A"),
            "Measure long-window temperature stability on configurations without the VH PCC performance path.",
            ("VC/VA path should not require PCC evidence.", "Stability evaluation still uses separate upper/lower sensor ranges."),
            "Temperature Stability",
            "TEMPERATURE_STABILITY",
            "NO_INTEGRATION",
            ("Temp Stability_Noise",),
            (
                "Expected to share the same separate-sensor stability workbook rule without PCC cells.",
            ),
            ("TempStability", "RES_TempStability"),
            ("DB mapping should omit PCC fields for non-PCC module variants.",),
            ("RetTime1", "RetTime2", "RetTime3"),
            ("ExtTemp_LowerCC", "ExtTemp_UpperCC"),
            ("ColumnComp.CC.Temperature.Nominal",),
            ("Non-PCC TCC configuration",),
            ("Binding must be confirmed from VC/VA CMBX evidence.",),
            coverage_status="open verification",
            open_gaps=("Whether Stability has any hidden IRC insertion path in production workflows still needs processing-method action-row evidence; VA Report_VATCC_V1_01 Temp Stability_Noise layout is verified.",),
        ),
        _make_tcc_record(
            100,
            "heatup_cooldown_20_50_20",
            "HeatUp and CoolDown",
            all_tcc,
            "Measure 20->50 C heat-up and 50->20 C cool-down time and subtract the method-script stable hold time.",
            ("Upper external thermometer is the main timing reference.", "Report subtracts 2.0 min from each final duration."),
            "HeatUp and CoolDownTime",
            "TEMP_HEAT_UP_DOWN_20_50_20",
            "No_Integration",
            ("HeatUp&CoolDown",),
            (
                "Verified evaluator rule: heat-up = RetTime2 - RetTime1 - 2.0; cool-down = RetTime5 - RetTime4 - 2.0.",
                "RetTime3 and RetTime6 remain required for the visible row-65 internal endpoint layout.",
            ),
            ("HeatUp_Time_20to50", "CoolDown_Time_50to20", "RES_HeatUp", "RES_CoolDown"),
            ("DB output uses one-decimal display for time fields.",),
            ("RetTime1", "RetTime2", "RetTime3", "RetTime4", "RetTime5", "RetTime6"),
            ("ExtTemp_UpperCC",),
            (),
            ("External upper thermometer configured",),
            (
                "CMBX sequence row should bind HeatUp and CoolDownTime -> TEMP_HEAT_UP_DOWN_20_50_20; VH/VA use No_Integration.",
                "Real VC 3000004.cmbx binds CORRECT_ACCURACY_INJ_INSERTION to many VC sequence rows, so HeatUp inherits VC sequence processing context rather than needing HeatUp-specific IRC logic.",
                "Keep the row-65 internal endpoint / row-66 external endpoint distinction visible when regenerating a binary-equivalent report template.",
            ),
            coverage_status="complete",
            generation_readiness="Good candidate for clone/select generation because the evidence contract is compact and well understood.",
        ),
        _make_tcc_record(
            110,
            "liquid_leak_keypad",
            "Liquid Leak / Keypad",
            all_tcc,
            "Verify the liquid leak sensor and related keypad behavior used by the module service check.",
            ("Sensor/audit evidence driven.", "May include manual/operator interaction depending on TD procedure."),
            "LiquidLeaktest",
            "LIQUID LEAK",
            "No_Integration",
            ("Liquid Leak", "Valve_Keypad"),
            (
                "Liquid Leak report evidence: M47 = AUDIT.LiquidLeak(100.000,\"backward\").",
                "Liquid Leak report evidence: K47 = precond.LiquidLeakCalibrationValue.",
                "Valve_Keypad may contribute keypad/alarm context but FOQResultLocations has no DB output row for LiquidLeaktest.",
            ),
            ("RES_LiquidLeak",),
            ("FOQResultLocations has no DB output rows for LiquidLeaktest; this is a manual/audit report check.",),
            (),
            (),
            ("LiquidLeak", "LeakSensor", "ColumnComp.Alarm", "precond.LiquidLeakCalibrationValue"),
            ("Liquid leak sensor configured", "Operator can inject/remove test liquid and confirm/mute alarm"),
            (
                "LIQUID LEAK setup: set safe CC nominal around 20 C, enable LiquidLeakSensor, initialize END_RUN trigger.",
                "LIQUID LEAK run: prompt operator to inject water, wait for LiquidLeak=Leak, log LiquidLeak, prompt alarm mute/confirmation.",
                "LIQUID LEAK cleanup: wait until ColumnComp.Alarm=NoAlarm, switch LiquidLeakSensor Off, prompt liquid removal, end run.",
            ),
            coverage_status="open verification",
            open_gaps=("Manual operator message timing and alarm-confirmation wait semantics need line-level method extraction before generation.",),
        ),
        _make_tcc_record(
            120,
            "qualification_service_done",
            "Qualification Service Done",
            all_tcc,
            "Mark or verify that the qualification service completion action/state has been performed.",
            ("Usually near the end of the sequence.", "Primarily audit/service-state evidence."),
            "Qualification_Service_Done",
            "Qualification_Service_Done",
            "No_Integration",
            ("Internal Use",),
            (
                "Qualification_Service_Done is metadata/audit driven and has no heavy raw channel calculation.",
                "Report role is qualification/service completion evidence rather than DB field output.",
            ),
            ("RES_Qualification_Service",),
            ("FOQResultLocations has no DB output rows for Qualification_Service_Done.",),
            (),
            (),
            ("ColumnComp_Wellness.Service.LastDate", "ColumnComp_Wellness.Qualification.LastDate"),
            ("Service-level command access if required",),
            (
                "Qualification_Service_Done run: log ColumnComp_Wellness.Service.LastDate.",
                "Qualification_Service_Done run: log ColumnComp_Wellness.Qualification.LastDate.",
                "Preserve this as a separate procedure-state/audit evidence injection.",
            ),
            coverage_status="open verification",
            open_gaps=("Need line-level confirmation whether method only logs service state or also writes completion state in some variants.",),
        ),
        _make_tcc_record(
            130,
            "factory_default_metadata",
            "Factory Default",
            all_tcc,
            "Verify factory default and identity metadata, including model number, model variant, serial number, firmware, and hardware revision.",
            ("Can use nearest injection audit fallback for missing audit trail.", "ModelNo must come from AUDIT.ColumnComp.ModelNo, not filename guessing."),
            "Factory Default",
            "FACTORYDEFAULT",
            "No_Integration",
            ("Definitions", "Internal Use", "Factory Default"),
            (
                "ModelVariant comes from Internal Use / ModuleHardwareRevision style metadata.",
                "Serial number check compares ColumnComp.SerialNo with sequence name.",
            ),
            ("ModelNo", "ModelVariant", "Firmware", "HardwareVersion", "RES_SN_Check"),
            ("Device detection and upload table selection use AUDIT.ColumnComp.ModelNo as source of truth.",),
            (),
            (),
            (
                "AUDIT.ColumnComp.ModelNo",
                "precond.ColumnComp.SerialNo",
                "precond.ColumnComp.FirmwareVersion",
                "precond.ColumnComp.HardwareVersion",
                "precond.ColumnComp.ModuleHardwareRevision",
            ),
            ("Module identity metadata available in precondition/audit trail",),
            ("CMBX sequence row should bind Factory Default -> FACTORYDEFAULT / No_Integration.",),
            generation_readiness="Metadata alignment is reusable, but CMBX generation must preserve source-of-truth audit/precondition paths.",
        ),
        _make_tcc_record(
            140,
            "error_log_check",
            "Error Log Check",
            all_tcc,
            "Verify that the module error log state is acceptable at the end of the FOQ run.",
            ("End-of-sequence audit/service check.", "Usually no raw signal calculation."),
            "Error Log Check",
            "CHECKERRORLOG",
            "No_Integration",
            ("Error Log", "Internal Use"),
            (
                "Error Log sheet is an audit table endpoint rather than a formula-object-heavy calculation sheet.",
                "Report formula extraction does not expose significant cell-level formulas for this sheet.",
            ),
            ("RES_ErrorLog",),
            ("FOQResultLocations has no DB output rows for Error Log Check.",),
            (),
            (),
            ("ErrorLog", "CheckErrorLog", "ColumnComp.CC.TempCtrl", "PrehtLeft.TempCtrl", "PrehtRight.TempCtrl"),
            ("Error log/service command access", "Safe end-state control over CC and preheater temp controls"),
            (
                "CHECKERRORLOG run: turn PrehtRight.TempCtrl Off, PrehtLeft.TempCtrl Off, and CC.TempCtrl Off.",
                "CHECKERRORLOG role: inspect final error-log/audit state and leave device in a safe non-running condition.",
            ),
            coverage_status="open verification",
            open_gaps=("Need line-level confirmation of error-log inspect/clear behavior and any variant-specific connection reset commands.",),
        ),
    )


def _vdad_records() -> tuple[FoqAlignmentRecord, ...]:
    source = "FOQ_VDAD_VMWD_TD_KNOWLEDGE_MANAGEMENT.md"
    common_config = ("Detector", "UV lamp", "VIS lamp where applicable", "flow cell or diagnostic cell")
    return (
        FoqAlignmentRecord(
            110,
            "VDAD",
            "warm_up",
            "Warm Up",
            ("VF-D11-A", "VC-D11-A", "VC-D12-A"),
            source,
            "确认检测器 lamp 和 optical system 在正式测试前达到可重复、稳定的工作状态。",
            ("Lamp configuration depends on detector model.", "Diagnostic cell tests precede fluidic flow-cell tests."),
            "Warm Up",
            "",
            "No_Integration",
            "",
            ("Diagnostic Cell",),
            ("TD defines warm-up as a prerequisite evidence step; CMBX method/report binding must be confirmed from a loaded VDAD package.",),
            (),
            (),
            (),
            (),
            (),
            common_config,
            (),
            (),
            "open verification",
            ("Load a VDAD CMBX and bind the actual injection, method, and report template.",),
            "Knowledge-only until VDAD method flow and report formulas are decoded from a reference CMBX.",
        ),
        FoqAlignmentRecord(
            120,
            "VDAD",
            "noise",
            "Noise",
            ("VF-D11-A", "VC-D11-A", "VC-D12-A", "VMWD-C"),
            source,
            "评估 detector baseline 噪声和漂移；故障排查时可区分 cell、lamp、pump/flow 等来源。",
            ("Diagnostic cell and flow-cell noise have different diagnostic meanings.", "Pump pulsation can propagate into detector noise."),
            "Noise",
            "",
            "No_Integration",
            "",
            ("Diagnostic Cell", "Fluidic Flow Cell"),
            ("Report formula and workbook calculation must be confirmed from VDAD/VMWD report template.",),
            ("Noise", "Drift"),
            ("DB mapping expected to contain detector noise/drift fields; exact field names require mapping confirmation.",),
            (),
            ("UV signal",),
            (),
            common_config,
            (),
            (),
            "open verification",
            ("VDAD/VMWD report template formula extraction is still required.",),
            "Use as a cross-module diagnostic alignment item before generation.",
        ),
        FoqAlignmentRecord(
            130,
            "VDAD",
            "linearity",
            "Linearity",
            ("VF-D11-A", "VC-D11-A", "VC-D12-A", "VMWD-C"),
            source,
            "通过不同浓度标准品的峰面积/峰高响应评估检测器在信号范围内的线性。",
            ("Dark Current Drift supports Linearity interpretation.", "Autosampler and pump precision can affect observed linearity."),
            "Linearity",
            "",
            "No_Integration",
            "",
            ("Fluidic Flow Cell",),
            ("Linearity depends on report integration/calibration logic that must be decoded from the report template.",),
            ("Linearity", "System Check"),
            ("DB mapping expected to include linearity-related result fields; exact report cells require mapping trace.",),
            (),
            ("UV signal",),
            (),
            ("Pump flow", "Autosampler injection volume", *common_config),
            (),
            (),
            "open verification",
            ("Processing/report integration logic must be decoded before this can become a generation template.",),
            "Not ready for generated method/report output; useful as a dependency map across pump, autosampler, and detector.",
        ),
        FoqAlignmentRecord(
            140,
            "VDAD",
            "wavelength_accuracy",
            "Wavelength Accuracy",
            ("VF-D11-A", "VC-D11-A", "VC-D12-A", "VMWD-C"),
            source,
            "用标准品或 filter/diagnostic evidence 验证检测器波长轴准确度。",
            ("Er(ClO4)3 or HoFi standards may be shared with spectral resolution checks.", "VMWD-C applicability must follow source TD distinctions."),
            "Wavelength Accuracy",
            "",
            "No_Integration",
            "",
            ("Diagnostic Cell", "Fluidic Flow Cell"),
            ("Shared standard and spectrum evaluation logic must be bound to actual report formulas.",),
            ("WavelengthAccuracy",),
            ("Exact DB fields and report cells require VDAD mapping trace.",),
            (),
            ("UV signal",),
            (),
            common_config,
            (),
            (),
            "open verification",
            ("Bind actual VDAD report template and formula objects from a loaded CMBX.",),
            "Useful first VDAD alignment target after Warm Up/Noise because TD dependencies are clear.",
        ),
        FoqAlignmentRecord(
            150,
            "VDAD",
            "3d_field",
            "3D Field",
            ("VF-D11-A", "VC-D11-A", "VC-D12-A"),
            source,
            "DAD-specific 3D field behavior; the TD explicitly distinguishes this from VMWD-C capability.",
            ("Not applicable to VMWD-C.",),
            "3D Field",
            "",
            "",
            "",
            ("3D Field",),
            ("DAD-only report/method evidence still needs CMBX binding.",),
            (),
            (),
            (),
            ("3D Field signal",),
            (),
            ("DAD detector model",),
            (),
            (),
            "not applicable",
            ("VMWD-C does not support 3D Field; do not mark this as missing for VMWD-C.",),
            "DAD-only knowledge item; no VMWD generation branch.",
        ),
    )


def _same_name(left: str, right: str) -> bool:
    return _normalize_name(left) == _normalize_name(right)


def _component_status(bound: bool, evidence_present: bool, evidence_status: str, open_status: str) -> str:
    if not bound:
        return "missing"
    if evidence_present:
        return evidence_status
    return open_status


def _processing_component_status(node: TestKnowledgeNode) -> str:
    if not node.processing_method:
        return "missing"
    if node.irc_injected:
        if any("pass/action" in gap.lower() or "processing" in gap.lower() for gap in node.open_gaps):
            return "IRC/corrective binding; pass-action decode open"
        return "IRC/corrective binding documented"
    return "bound"


def _report_formula_component_status(node: TestKnowledgeNode) -> str:
    if node.formula_id.startswith("FORMULA_NOT_REQUIRED"):
        return "no report formula contract expected"
    has_report = bool(node.report_template or node.report_sheets)
    has_formula = node.formula_id != "FORMULA_OPEN_VERIFICATION_REQUIRED" and not node.formula_id.endswith("_OPEN")
    if has_report and has_formula and node.report_evidence:
        return "report sheet and formula family mapped"
    if has_report and has_formula:
        return "mapped; formula evidence open"
    if has_report:
        return "report bound; formula open"
    if has_formula:
        return "formula family listed; report binding open"
    return "missing"


def _db_component_status(node: TestKnowledgeNode) -> str:
    if not node.db_fields:
        return "no DB contract expected"
    if node.db_evidence:
        return "DB fields mapped with evidence"
    return "DB fields mapped; trace/type/precision verification open"


def _dependency_component_status(expected: Iterable[str], evidence: Iterable[str], label: str) -> str:
    expected_values = tuple(value for value in expected if value)
    if not expected_values:
        return "not required"
    if tuple(value for value in evidence if value):
        return f"{label} contract documented"
    return f"{label} expected; evidence open"


def _generated_coverage_gaps(
    node: TestKnowledgeNode,
    method_status: str,
    processing_status: str,
    report_status: str,
    db_status: str,
    ret_time_status: str,
    channel_status: str,
    audit_status: str,
    config_status: str,
) -> tuple[str, ...]:
    status_map = {
        "method command": method_status,
        "processing method": processing_status,
        "report/formula": report_status,
        "DB field": db_status,
        "RetTime": ret_time_status,
        "channel": channel_status,
        "audit/property": audit_status,
        "configuration": config_status,
    }
    gaps: list[str] = []
    for component, status in status_map.items():
        lowered = status.lower()
        if any(marker in lowered for marker in ("missing", "open")):
            gaps.append(f"{component}: {status}")
    if node.coverage_status != "complete" and not node.open_gaps:
        gaps.append(f"coverage status is {node.coverage_status}; detailed proof gap not recorded")
    return tuple(gaps)


def _overall_audit_status(coverage_status: str, gaps: Iterable[str]) -> str:
    gap_tuple = tuple(gaps)
    if coverage_status == "not applicable":
        return "not applicable"
    if not gap_tuple and coverage_status == "complete":
        return "closed"
    if any("missing" in gap.lower() for gap in gap_tuple):
        return "missing evidence"
    if gap_tuple:
        return "open verification"
    return coverage_status


def _mapping_rows_for_binding(
    locations: list[Any],
    injection: str,
    report_sheets: tuple[str, ...],
    report_files: tuple[str, ...],
    exact_injection: bool,
) -> list[Any]:
    if not report_sheets:
        return []
    sheet_keys = {_normalize_name(sheet) for sheet in report_sheets}
    report_file_keys = {_normalize_report_file_for_alignment(report_file) for report_file in report_files if report_file}
    if exact_injection and not report_file_keys and injection:
        report_file_keys = {_normalize_report_file_for_alignment(f"{injection}.XLS")}
    rows: list[Any] = []
    for row in locations:
        if _normalize_name(getattr(row, "report_sheet", "")) not in sheet_keys:
            continue
        if exact_injection and report_file_keys and _normalize_report_file_for_alignment(getattr(row, "report_file", "")) not in report_file_keys:
            continue
        rows.append(row)
    return rows


def _db_mapping_audit_status(
    expected_fields: tuple[str, ...],
    rows: list[Any],
    missing_expected: tuple[str, ...],
    extra_mapped: tuple[str, ...],
) -> str:
    if not expected_fields and not rows:
        return "no DB contract expected"
    if expected_fields and not rows:
        return "missing mapping rows"
    if missing_expected and extra_mapped:
        return "node/mapping mismatch"
    if missing_expected:
        return "node fields not found in mapping"
    if extra_mapped:
        return "mapping has additional DB fields"
    return "closed"


def _normalize_report_file_for_alignment(value: str) -> str:
    text = str(value or "").strip()
    for suffix in (".xlsx", ".xls", ".XLSX", ".XLS"):
        if text.endswith(suffix):
            text = text[: -len(suffix)]
            break
    return _normalize_name(text)


def _normalize_name(value: str) -> str:
    return "".join(char.lower() for char in value if char.isalnum())


def _unique(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return tuple(result)


def _markdown_bullets(values: Iterable[str], empty: str = "not recorded") -> list[str]:
    items = [value for value in values if value]
    if not items:
        return [f"- {empty}"]
    return [f"- {item}" for item in items]
