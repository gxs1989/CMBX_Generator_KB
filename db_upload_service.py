from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re


@dataclass(frozen=True)
class DatabaseUploadConfig:
    server: str
    database: str
    username: str
    password: str
    schema: str = "dbo"
    table: str = "AUTO"
    driver: str = "ODBC Driver 17 for SQL Server"
    trust_server_certificate: bool = True
    dsn: str = ""


@dataclass(frozen=True)
class DbWorkbookRow:
    path: Path
    sequence_name: str
    values: dict[str, object]


@dataclass(frozen=True)
class UploadResult:
    path: Path
    sequence_name: str
    table_name: str
    field_count: int
    row_count: int


FOQ_TABLE_BY_DEVICE_TYPE = {
    "VA-C10-A": "VTCC",
    "VC-C10-A": "VTCC",
    "VH-C10-A": "VTCC",
    "VH-C10-A_GERMERING": "VTCC_Germering",
}


def test_database_connection(config: DatabaseUploadConfig) -> str:
    pyodbc = _import_pyodbc()
    with pyodbc.connect(_connection_string(config), timeout=10) as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT DB_NAME(), SUSER_SNAME()")
        database_name, user_name = cursor.fetchone()
    return f"Connected to {database_name} as {user_name}"


def list_database_tables(config: DatabaseUploadConfig) -> list[tuple[str, str]]:
    """Return user tables available to the configured read connection."""
    pyodbc = _import_pyodbc()
    with pyodbc.connect(_connection_string(config), timeout=15) as connection:
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT s.name, t.name
            FROM sys.tables t
            JOIN sys.schemas s ON s.schema_id = t.schema_id
            ORDER BY s.name, t.name
            """
        )
        return [(str(row[0]), str(row[1])) for row in cursor.fetchall()]


def fetch_table_rows(
    config: DatabaseUploadConfig,
    table: str | None = None,
    schema: str | None = None,
    limit: int = 5000,
) -> list[dict[str, object]]:
    """Read a bounded historical dataset using validated SQL identifiers."""
    pyodbc = _import_pyodbc()
    table_name = _validate_sql_identifier(table or config.table, "table")
    schema_name = _validate_sql_identifier(schema or config.schema, "schema")
    row_limit = max(1, min(int(limit), 100000))
    with pyodbc.connect(_connection_string(config), timeout=30) as connection:
        cursor = connection.cursor()
        cursor.execute(
            f"SELECT TOP ({row_limit}) * FROM {_quote_identifier(schema_name)}.{_quote_identifier(table_name)} ORDER BY 1 DESC"
        )
        columns = [str(item[0]) for item in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def upload_foq_db_workbooks(
    workbook_paths: list[str | Path],
    config: DatabaseUploadConfig,
    log=None,
) -> list[UploadResult]:
    pyodbc = _import_pyodbc()
    paths = [Path(path) for path in workbook_paths]
    if not paths:
        return []
    schema = _validate_sql_identifier(config.schema, "schema")
    rows = [read_foq_db_workbook(path) for path in paths]
    with pyodbc.connect(_connection_string(config), timeout=30) as connection:
        connection.autocommit = False
        cursor = connection.cursor()
        results: list[UploadResult] = []
        try:
            for row in rows:
                table = _resolve_target_table(row, config)
                if log:
                    log(f"Uploading {row.path.name} -> {schema}.{table}")
                if not _table_exists(cursor, schema, table):
                    if log:
                        log(f"Creating table {schema}.{table} from FOQ DB fields.")
                    _ensure_schema(cursor, schema)
                    _create_foq_table(cursor, schema, table, row.values)
                columns = _existing_columns(cursor, schema, table)
                inserted_fields, skipped_fields, truncated_fields = _insert_existing_table_row(cursor, schema, table, columns, row)
                if log and skipped_fields:
                    log(f"{row.path.name}: skipped {len(skipped_fields)} field(s) not present in {schema}.{table}: {', '.join(skipped_fields[:8])}")
                if log and truncated_fields:
                    log(f"{row.path.name}: truncated {len(truncated_fields)} string field(s) to fit table schema: {', '.join(truncated_fields[:8])}")
                results.append(UploadResult(row.path, row.sequence_name, table, inserted_fields, 1))
            connection.commit()
        except Exception:
            connection.rollback()
            raise
    return results


def read_foq_db_workbook(path: str | Path) -> DbWorkbookRow:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read FOQ DB workbooks.") from exc

    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "DB Data" not in workbook.sheetnames:
            raise ValueError(f"{workbook_path.name} has no 'DB Data' sheet")
        sheet = workbook["DB Data"]
        rows = list(sheet.iter_rows(max_row=2, values_only=True))
        if len(rows) < 2:
            raise ValueError(f"{workbook_path.name} has no DB data row")
        headers = [str(header).strip() if header is not None else "" for header in rows[0]]
        values = rows[1]
        data = {header: value for header, value in zip(headers, values) if header}
    finally:
        workbook.close()
    return DbWorkbookRow(workbook_path, _sequence_name_from_workbook(workbook_path), data)


def discover_foq_db_workbooks(folder: str | Path) -> list[Path]:
    root = Path(folder)
    if not root.exists():
        return []
    if root.is_file():
        return [root] if _is_foq_db_workbook(root) else []
    return sorted(path for path in root.rglob("*.xlsx") if _is_foq_db_workbook(path))


def _is_foq_db_workbook(path: Path) -> bool:
    name = path.name.lower()
    return name.endswith("_foq_contract_db.xlsx") or name.endswith("_foq_db.xlsx")


def _sequence_name_from_workbook(path: Path) -> str:
    name = path.stem
    for suffix in ("_FOQ_contract_DB", "_foq_contract_db", "_FOQ_DB", "_foq_db"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return name


def _connection_string(config: DatabaseUploadConfig) -> str:
    trust = "yes" if config.trust_server_certificate else "no"
    if config.dsn.strip():
        dsn = config.dsn.strip()
        dsn_key = "FILEDSN" if dsn.lower().endswith(".dsn") or "\\" in dsn or "/" in dsn else "DSN"
        return (
            f"{dsn_key}={dsn};"
            f"UID={config.username};"
            f"PWD={config.password};"
            f"DATABASE={config.database};"
            f"TrustServerCertificate={trust};"
        )
    return (
        f"DRIVER={{{config.driver}}};"
        f"SERVER={config.server};"
        f"DATABASE={config.database};"
        f"UID={config.username};"
        f"PWD={config.password};"
        f"TrustServerCertificate={trust};"
    )


def _table_exists(cursor, schema: str, table: str) -> bool:
    cursor.execute(
        """
        SELECT 1
        FROM sys.tables t
        JOIN sys.schemas s ON s.schema_id = t.schema_id
        WHERE s.name = ? AND t.name = ?
        """,
        schema,
        table,
    )
    return bool(cursor.fetchone())


def _ensure_schema(cursor, schema: str) -> None:
    cursor.execute("SELECT 1 FROM sys.schemas WHERE name = ?", schema)
    if cursor.fetchone():
        return
    cursor.execute(f"EXEC('CREATE SCHEMA {_quote_identifier(schema)}')")


def _create_foq_table(cursor, schema: str, table: str, values: dict[str, object]) -> None:
    columns = ["[ID] INT IDENTITY(1,1) NOT NULL PRIMARY KEY"]
    for header, value in values.items():
        column = _db_column_name(header)
        if column.lower() == "id":
            continue
        columns.append(f"{_quote_identifier(column)} {_sql_type_for_field(header, value)} NULL")
    column_sql = ",\n            ".join(columns)
    cursor.execute(
        f"""
        CREATE TABLE {_quote_identifier(schema)}.{_quote_identifier(table)} (
            {column_sql}
        )
        """
    )


def _existing_columns(cursor, schema: str, table: str) -> dict[str, tuple[str, str, int | None]]:
    cursor.execute(
        """
        SELECT c.name, ty.name, c.max_length
        FROM sys.columns c
        JOIN sys.tables t ON t.object_id = c.object_id
        JOIN sys.schemas s ON s.schema_id = t.schema_id
        JOIN sys.types ty ON ty.user_type_id = c.user_type_id
        WHERE s.name = ? AND t.name = ?
        """,
        schema,
        table,
    )
    rows = cursor.fetchall()
    if not rows:
        raise ValueError(f"Target table was not found: {schema}.{table}")
    result: dict[str, tuple[str, str, int | None]] = {}
    for row in rows:
        column_name = str(row[0])
        sql_type = str(row[1]).lower()
        max_length = int(row[2])
        max_chars = None
        if max_length > 0:
            max_chars = max_length // 2 if sql_type in {"nvarchar", "nchar"} else max_length
        result[column_name.lower()] = (column_name, sql_type, max_chars)
    return result


def _insert_existing_table_row(cursor, schema: str, table: str, existing_columns: dict[str, tuple[str, str, int | None]], row: DbWorkbookRow) -> tuple[int, list[str], list[str]]:
    reserved = {"id"}
    columns: list[str] = []
    values: list[object] = []
    skipped: list[str] = []
    truncated: list[str] = []
    for header, value in row.values.items():
        key = _db_column_name(header).lower()
        column_info = existing_columns.get(key)
        if not column_info or key in reserved:
            skipped.append(str(header))
            continue
        column_name, sql_type, max_chars = column_info
        columns.append(column_name)
        db_value, was_truncated = _to_database_value(value, sql_type, max_chars)
        values.append(db_value)
        if was_truncated:
            truncated.append(str(header))
    if not columns:
        raise ValueError(f"No DB Data fields matched existing columns in {schema}.{table}")
    placeholders = ", ".join("?" for _value in values)
    column_sql = ", ".join(_quote_identifier(column) for column in columns)
    cursor.execute(
        f"INSERT INTO {_quote_identifier(schema)}.{_quote_identifier(table)} ({column_sql}) VALUES ({placeholders})",
        values,
    )
    return len(columns), skipped, truncated


def _db_column_name(header: str) -> str:
    name = str(header).strip() or "Unnamed"
    if name.lower() in {"uploadid", "uploadedat", "cmbxsequencename", "sourceworkbook"}:
        name = f"DB_{name}"
    if len(name) <= 128:
        return name
    return name[:120] + "_" + str(abs(hash(name)) % 10_000_000)


def _sql_type_for_field(header: str, value: object) -> str:
    name = str(header or "").strip()
    lower = name.lower()
    if lower in {"testdate", "date", "datetime"}:
        return "DATETIME2(0)"
    if lower in {"serial", "timebase", "modelno", "modelvariant", "firmware"}:
        return "NVARCHAR(30)"
    if lower.startswith("res_") or lower.startswith("result"):
        return "NVARCHAR(30)"
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return "FLOAT"
    if isinstance(value, (datetime, date)):
        return "DATETIME2(0)"
    text = str(value or "")
    if len(text) <= 30:
        return "NVARCHAR(30)"
    if len(text) <= 100:
        return "NVARCHAR(100)"
    return "NVARCHAR(255)"


def _resolve_target_table(row: DbWorkbookRow, config: DatabaseUploadConfig) -> str:
    requested = str(config.table or "").strip()
    if requested and requested.upper() not in {"AUTO", "<AUTO>"}:
        return _validate_sql_identifier(requested, "table")
    device_type = str(row.values.get("DeviceType") or row.values.get("ModelNo") or "").strip().upper()
    mapping_sheet = str(row.values.get("MappingSheet") or "").strip()
    table = FOQ_TABLE_BY_DEVICE_TYPE.get(device_type)
    if table:
        return table
    normalized_mapping = re.sub(r"[^A-Za-z0-9_]", "", mapping_sheet)
    if normalized_mapping:
        return _validate_sql_identifier(normalized_mapping, "table")
    raise ValueError(f"Cannot resolve target table for {row.path.name}: DeviceType={device_type!r}, MappingSheet={mapping_sheet!r}")


def _to_database_value(value: object, sql_type: str, max_chars: int | None = None) -> tuple[object, bool]:
    if value is None:
        return None, False
    if isinstance(value, str) and not value.strip():
        return None, False
    if sql_type in {"float", "real", "decimal", "numeric", "money", "smallmoney"}:
        try:
            return float(value), False
        except (TypeError, ValueError):
            return None, False
    if sql_type in {"int", "bigint", "smallint", "tinyint"}:
        try:
            return int(float(value)), False
        except (TypeError, ValueError):
            return None, False
    if sql_type in {"datetime", "datetime2", "smalldatetime", "date"}:
        if isinstance(value, (datetime, date)):
            return value, False
        return str(value), False
    if sql_type == "bit":
        if isinstance(value, bool):
            return value, False
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "passed", "ok"}:
            return True, False
        if text in {"0", "false", "no", "failed"}:
            return False, False
        return None, False
    text_value = _to_database_text(value)
    if text_value is not None and max_chars and len(text_value) > max_chars:
        return text_value[:max_chars], True
    return text_value, False


def _to_database_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="milliseconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _validate_sql_identifier(value: str, label: str) -> str:
    text = str(value or "").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", text):
        raise ValueError(f"Invalid SQL {label}: {value!r}. Use letters, numbers, and underscore only.")
    return text


def _quote_identifier(value: str) -> str:
    return "[" + value.replace("]", "]]") + "]"


def _import_pyodbc():
    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError("pyodbc is required for SQL Server upload.") from exc
    return pyodbc
