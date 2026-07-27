from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from report_calculation_map import ReportCellCalculation
from report_formula_evaluator import FormulaEvaluation


@dataclass(frozen=True)
class AccuracyReportRow:
    adjusted_temperature_c: float
    time_min: float | None
    observed_lower_c: float | None
    observed_upper_c: float | None
    observed_max_deviation_temp_c: float | None
    deviation_c: float | None
    result: str


@dataclass(frozen=True)
class HeatupCooldownReport:
    criterion_min: float
    heatup_start_min: float | None
    heatup_end_min: float | None
    heatup_observed_min: float | None
    cooldown_start_min: float | None
    cooldown_end_min: float | None
    cooldown_observed_min: float | None
    t1_c: float | None
    t2_c: float | None


def write_report_workbook(
    path: str | Path,
    injection_name: str,
    sheets: dict[str, list[FormulaEvaluation]],
    calculation_map: list[ReportCellCalculation] | None = None,
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build report workbooks.") from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    if "Definitions" not in sheets:
        _write_definitions_sheet(workbook.create_sheet("Definitions"), injection_name, Font, PatternFill, Alignment, Border, Side)
    for sheet_name, rows in sheets.items():
        if sheet_name == "Temp Accuracy":
            _write_temp_accuracy_sheet(workbook.create_sheet(sheet_name), injection_name, rows, Font, PatternFill, Alignment, Border, Side)
        elif "heatup" in sheet_name.lower() and "cooldown" in sheet_name.lower():
            _write_heatup_cooldown_sheet(workbook.create_sheet(_safe_sheet_name(sheet_name)), injection_name, rows, calculation_map or [], Font, PatternFill, Alignment, Border, Side)
        elif sheet_name == "Definitions":
            _write_definitions_sheet(workbook.create_sheet(sheet_name), injection_name, Font, PatternFill, Alignment, Border, Side)
        else:
            _write_generic_formula_sheet(workbook.create_sheet(_safe_sheet_name(sheet_name)), injection_name, rows, Font, PatternFill, Alignment, Border, Side)
    if calculation_map:
        _write_calculation_map_sheet(workbook.create_sheet("Calculation Map"), injection_name, calculation_map, Font, PatternFill, Alignment, Border, Side)
    workbook.save(_windows_long_path(output_path))
    return output_path


def build_accuracy_rows(evaluations: list[FormulaEvaluation], acceptance_c: float = 0.5) -> list[AccuracyReportRow]:
    values = {row.excel_range: _to_float(row.value) for row in evaluations if row.status == "ok"}
    rows: list[AccuracyReportRow] = []
    for offset, excel_row in enumerate(range(66, 71)):
        adjusted = values.get(f"I{excel_row}")
        if adjusted is None:
            adjusted = values.get(f"J{excel_row}")
        if adjusted is None:
            adjusted = [10.0, 20.0, 40.0, 80.0, 120.0][offset]
        lower = values.get(f"L{excel_row}")
        upper = values.get(f"M{excel_row}")
        observed = _max_deviation_temperature(adjusted, lower, upper)
        deviation = (observed - adjusted) if observed is not None else None
        result = "ok" if deviation is not None and abs(deviation) <= acceptance_c else "failed"
        rows.append(
            AccuracyReportRow(
                adjusted_temperature_c=adjusted,
                time_min=values.get(f"K{excel_row}"),
                observed_lower_c=lower,
                observed_upper_c=upper,
                observed_max_deviation_temp_c=observed,
                deviation_c=deviation,
                result=result,
            )
        )
    return rows


def build_heatup_cooldown_report(
    evaluations: list[FormulaEvaluation],
    calculation_map: list[ReportCellCalculation] | None = None,
) -> HeatupCooldownReport:
    values = {row.excel_range: _to_float(row.value) for row in evaluations if row.status == "ok"}
    criterion = _criterion_from_calculation_map(calculation_map or [], "HeatUp&CoolDown", 15.0)
    heatup_start = values.get("J65")
    heatup_end = values.get("K65")
    cooldown_start = values.get("L65")
    cooldown_end = values.get("M65")
    return HeatupCooldownReport(
        criterion_min=criterion,
        heatup_start_min=heatup_start,
        heatup_end_min=heatup_end,
        heatup_observed_min=_duration_minus_hold(heatup_start, heatup_end),
        cooldown_start_min=cooldown_start,
        cooldown_end_min=cooldown_end,
        cooldown_observed_min=_duration_minus_hold(cooldown_start, cooldown_end),
        t1_c=values.get("L57"),
        t2_c=values.get("L58"),
    )


def _write_temp_accuracy_sheet(ws, injection_name: str, evaluations: list[FormulaEvaluation], Font, PatternFill, Alignment, Border, Side) -> None:
    thin = Side(style="thin", color="B7B7B7")
    medium = Side(style="medium", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    section_fill = PatternFill("solid", fgColor="EAF2F8")
    pass_fill = PatternFill("solid", fgColor="D9EAD3")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")
    rows = build_accuracy_rows(evaluations)
    max_dev = max((_summary_abs_deviation(row) for row in rows if row.observed_max_deviation_temp_c is not None), default=None)
    overall = "Test passed" if max_dev is not None and max_dev <= 0.5 else "Test failed"

    ws.sheet_view.showGridLines = False
    widths = {
        "A": 4,
        "B": 24,
        "C": 18,
        "D": 18,
        "E": 14,
        "F": 3,
        "G": 3,
        "H": 4,
        "I": 8,
        "J": 13,
        "K": 14,
        "L": 18,
        "M": 18,
        "N": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in range(1, 137):
        ws.row_dimensions[row_idx].height = 18

    ws["A5"] = "Factory Operational Qualification"
    ws["A5"].font = Font(bold=True, size=14)
    ws["B9"] = "Temperature Accuracy"
    ws["B9"].font = Font(bold=True, size=13)
    ws["B11"] = "Instruments"
    ws["B11"].font = Font(bold=True)
    instrument_rows = [
        ("Instrument Name", "Model", "Serial Number"),
        ("Column Compartment", "", ""),
        ("Chromeleon Datasystem:", "", ""),
        ("Test Date:", "", ""),
        ("Operator:", "", ""),
    ]
    start = 13
    for idx, row in enumerate(instrument_rows, start):
        for col, value in enumerate(row, 2):
            ws.cell(idx, col, value)
    ws["B23"] = "Limits, Values and Test Results"
    ws["B23"].font = Font(bold=True)
    ws["B26"] = "Temperature Accuracy"
    ws["C25"] = "Acceptance Criterion"
    ws["C26"] = 0.5
    ws["D25"] = "Observed Max. Deviation"
    ws["D26"] = max_dev
    ws["E25"] = "Result"
    ws["E26"] = overall
    ws["E26"].fill = pass_fill if overall == "Test passed" else fail_fill

    ws["B45"] = "Temperature Accuracy"
    ws["B45"].font = Font(bold=True)
    _write_accuracy_headers(ws, 63, 2, ["Adjusted ", "Observed ", "Deviation", "Result"], header_fill, border, Font, Alignment)
    _write_accuracy_headers(ws, 63, 9, ["", "Adjusted ", "Time [min]", "Observed Temperature", "Observed Temperature ", "Observed Temperature"], header_fill, border, Font, Alignment)
    ws["B64"] = "Temperature"
    ws["C64"] = "Temperature"
    ws["J64"] = "Temperature"
    ws["L64"] = " (Lower CC)"
    ws["M64"] = "(Upper CC)"
    ws["N64"] = " with Maximum Deviation"
    for cell in ("B65", "C65", "D65", "J65", "L65", "M65", "N65"):
        ws[cell] = "[deg C]"
    ws["K65"] = ""

    for idx, row in enumerate(rows, 66):
        observed = row.observed_max_deviation_temp_c
        summary_observed = round(observed, 2) if observed is not None else None
        summary_deviation = (summary_observed - row.adjusted_temperature_c) if summary_observed is not None else None
        ws.cell(idx, 2, row.adjusted_temperature_c)
        ws.cell(idx, 3, summary_observed)
        ws.cell(idx, 4, summary_deviation)
        ws.cell(idx, 5, row.result)
        ws.cell(idx, 9, row.adjusted_temperature_c)
        ws.cell(idx, 10, row.adjusted_temperature_c)
        ws.cell(idx, 11, row.time_min)
        ws.cell(idx, 12, row.observed_lower_c)
        ws.cell(idx, 13, row.observed_upper_c)
        ws.cell(idx, 14, observed)
        ws.cell(idx, 5).fill = pass_fill if row.result == "ok" else fail_fill
    ws["C71"] = "Observed Max. Deviation"
    ws["D71"] = max_dev
    ws["E71"] = "ok" if max_dev is not None and max_dev <= 0.5 else "failed"
    ws["E71"].fill = pass_fill if ws["E71"].value == "ok" else fail_fill

    for block in ("B63:E71", "I63:N70", "B25:E26"):
        for row in ws[block]:
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if cell.row in (25, 63, 64, 65):
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
    for row in ws["B13:D20"]:
        for cell in row:
            cell.border = border
    for row in ws["B63:N70"]:
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"
    for cell in ("K66", "K67", "K68", "K69", "K70"):
        ws[cell].number_format = "0.000"
    ws.freeze_panes = "A63"


def _write_heatup_cooldown_sheet(ws, injection_name: str, evaluations: list[FormulaEvaluation], calculation_map: list[ReportCellCalculation], Font, PatternFill, Alignment, Border, Side) -> None:
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    pass_fill = PatternFill("solid", fgColor="D9EAD3")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")
    report = build_heatup_cooldown_report(evaluations, calculation_map)

    ws.sheet_view.showGridLines = False
    widths = {"A": 4, "B": 26, "C": 18, "D": 18, "E": 14, "F": 3, "G": 3, "H": 4, "I": 23, "J": 14, "K": 14, "L": 14, "M": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in range(1, 90):
        ws.row_dimensions[row_idx].height = 18

    t1 = _format_temperature(report.t1_c)
    t2 = _format_temperature(report.t2_c)
    ws["A5"] = "Factory Operational Qualification"
    ws["A5"].font = Font(bold=True, size=14)
    ws["B9"] = f"Heat Up / Cool Down Time ({t1} to {t2} to {t1})"
    ws["B9"].font = Font(bold=True, size=13)
    ws["B11"] = "Instruments"
    ws["B11"].font = Font(bold=True)
    instrument_rows = [
        ("Instrument Name", "Model", "Serial Number"),
        ("Column Compartment", "", ""),
        ("Chromeleon Datasystem:", "", ""),
        ("Test Date:", "", ""),
        ("Operator:", "", ""),
    ]
    for idx, row in enumerate(instrument_rows, 13):
        for col, value in enumerate(row, 2):
            ws.cell(idx, col, value)

    ws["B23"] = "Limits, Values and Test Results"
    ws["B23"].font = Font(bold=True)
    ws["C25"] = "Acceptance Criterion"
    ws["D25"] = "Observed Time"
    ws["E25"] = "Result"
    summary_rows = [
        (26, "Heat Up Time", report.heatup_observed_min),
        (27, "Cool Down Time", report.cooldown_observed_min),
    ]
    for row_idx, label, observed in summary_rows:
        ws.cell(row_idx, 2, label)
        ws.cell(row_idx, 3, report.criterion_min)
        ws.cell(row_idx, 4, observed)
        result = _pass_fail(observed, report.criterion_min)
        ws.cell(row_idx, 5, result)
        ws.cell(row_idx, 5).fill = pass_fill if result == "Test passed" else fail_fill

    ws["B45"] = "Heat Up and Cool Down Time"
    ws["B45"].font = Font(bold=True)
    ws["I57"] = "Heat Up and Cool Down Range:"
    ws["I58"] = "Temperature T1 in [deg C]:"
    ws["J58"] = report.t1_c
    ws["K58"] = t1
    ws["I59"] = "Temperature T2 in [deg C]:"
    ws["J59"] = report.t2_c
    ws["K59"] = t2

    ws["B63"] = "Temperature"
    ws["C63"] = "Start -> End Time"
    ws["D63"] = "Time Difference"
    ws["E63"] = "Result"
    ws["B64"] = "Step"
    for cell in ("C64", "D64"):
        ws[cell] = "[min]"
    ws["E64"] = ""
    ws["B65"] = f"{t1} to {t2}"
    ws["C65"] = _format_time_range(report.heatup_start_min, report.heatup_end_min)
    ws["D65"] = report.heatup_observed_min
    ws["E65"] = _pass_fail(report.heatup_observed_min, report.criterion_min)
    ws["B66"] = f"{t2} to {t1}"
    ws["C66"] = _format_time_range(report.cooldown_start_min, report.cooldown_end_min)
    ws["D66"] = report.cooldown_observed_min
    ws["E66"] = _pass_fail(report.cooldown_observed_min, report.criterion_min)

    ws["I62"] = "Heat Up"
    ws["L62"] = "Cool Down"
    ws["I63"] = "Start Time"
    ws["J63"] = "End Time"
    ws["L63"] = "Start Time"
    ws["M63"] = "End Time"
    ws["I64"] = "[min]"
    ws["J64"] = "[min]"
    ws["L64"] = "[min]"
    ws["M64"] = "[min]"
    ws["I65"] = report.heatup_start_min
    ws["J65"] = report.heatup_end_min
    ws["L65"] = report.cooldown_start_min
    ws["M65"] = report.cooldown_end_min
    ws["I66"] = "Hold-time subtraction [min]"
    ws["J66"] = 2
    ws["L66"] = "Hold-time subtraction [min]"
    ws["M66"] = 2

    for block in ("B25:E27", "B63:E66", "I57:K59", "I62:M66"):
        for row in ws[block]:
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if cell.row in (25, 63, 64, 62, 57):
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
    for cell_ref in ("E65", "E66"):
        cell = ws[cell_ref]
        cell.fill = pass_fill if cell.value == "Test passed" else fail_fill
    for row in ws["B13:D17"]:
        for cell in row:
            cell.border = border
    for row in ws["B25:M66"]:
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    for cell in ("C26", "C27", "J58", "J59", "J66", "M66"):
        ws[cell].number_format = "0.0"
    ws.freeze_panes = "A63"


def _write_accuracy_headers(ws, row: int, col: int, headers: list[str], fill, border, Font, Alignment) -> None:
    for index, header in enumerate(headers):
        cell = ws.cell(row, col + index, header)
        cell.fill = fill
        cell.border = border
        cell.font = Font(bold=True, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_definitions_sheet(ws, injection_name: str, Font, PatternFill, Alignment, Border, Side) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Factory Operational Qualification"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Definitions"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A5"] = "Injection"
    ws["B5"] = injection_name
    ws["A7"] = "Note"
    ws["B7"] = "Generated from CMBX report template formulas, injection audit trail, and raw signal data."
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    ws["B7"].alignment = Alignment(wrap_text=True)


def _write_generic_formula_sheet(ws, injection_name: str, evaluations: list[FormulaEvaluation], Font, PatternFill, Alignment, Border, Side) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = ws.title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Injection"
    ws["B2"] = injection_name
    headers = ["ExcelRange", "ObjectType", "FixedChannel", "Formula", "Value", "Status", "Detail"]
    ws.append([])
    ws.append(headers)
    for row in evaluations:
        ws.append([row.excel_range, row.object_type, row.fixed_channel, row.formula, row.value, row.status, row.detail])
    fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[4]:
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.border = Border(bottom=thin)
    widths = [12, 22, 20, 65, 18, 14, 60]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


def _write_calculation_map_sheet(ws, injection_name: str, rows: list[ReportCellCalculation], Font, PatternFill, Alignment, Border, Side) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Report Calculation Map"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Injection"
    ws["B2"] = injection_name
    headers = ["Sheet", "Cell", "Label", "Expression", "Dependencies", "Criterion", "Rule", "Source", "Note"]
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append([row.sheet_name, row.cell, row.label, row.expression, row.dependencies, row.criterion, row.rule, row.source, row.note])
    fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[4]:
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [20, 12, 36, 46, 58, 14, 45, 46, 70]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row_cells in ws.iter_rows(min_row=5):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A5"


def _max_deviation_temperature(adjusted: float, lower: float | None, upper: float | None) -> float | None:
    candidates = [value for value in (lower, upper) if value is not None]
    if not candidates:
        return None
    return max(candidates, key=lambda value: abs(value - adjusted))


def _summary_abs_deviation(row: AccuracyReportRow) -> float:
    observed = round(row.observed_max_deviation_temp_c, 2) if row.observed_max_deviation_temp_c is not None else row.adjusted_temperature_c
    return abs(observed - row.adjusted_temperature_c)


def _duration_minus_hold(start: float | None, end: float | None, hold_min: float = 2.0) -> float | None:
    if start is None or end is None:
        return None
    return max(0.0, end - start - hold_min)


def _criterion_from_calculation_map(rows: list[ReportCellCalculation], sheet_name: str, fallback: float) -> float:
    for row in rows:
        if row.sheet_name == sheet_name and row.criterion:
            value = _to_float(row.criterion)
            if value is not None:
                return value
    for row in rows:
        if row.sheet_name == "Definitions" and row.label == "HeatUp & Cool Down":
            value = _to_float(row.expression)
            if value is not None:
                return value
    return fallback


def _pass_fail(observed: float | None, criterion: float) -> str:
    if observed is not None and observed <= criterion:
        return "Test passed"
    return "Test failed"


def _format_temperature(value: float | None) -> str:
    if value is None:
        return "n.a."
    if float(value).is_integer():
        return f"{int(value)} deg C"
    return f"{value:g} deg C"


def _format_time_range(start: float | None, end: float | None) -> str:
    if start is None or end is None:
        return "n.a."
    return f"{start:.3f} -> {end:.3f}"


def _to_float(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_sheet_name(value: str) -> str:
    cleaned = "".join("_" if ch in "[]:*?/\\'" else ch for ch in value).strip()
    return (cleaned or "Sheet")[:31]


def _windows_long_path(path: Path) -> str:
    resolved = path.resolve()
    text = str(resolved)
    if not text.startswith("\\\\?\\") and len(text) >= 240:
        return "\\\\?\\" + text
    return text
