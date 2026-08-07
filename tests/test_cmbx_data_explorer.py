from __future__ import annotations

import sys
import zipfile
import gzip
import xml.etree.ElementTree as ET
from pathlib import Path
from types import SimpleNamespace

import pytest


PROJECT_ROOT = Path(__file__).resolve().parents[2]
MODULE_ROOT = PROJECT_ROOT / "cmbx_data_explorer"
if str(MODULE_ROOT) not in sys.path:
    sys.path.insert(0, str(MODULE_ROOT))

from cmbx_container import extract_cmbx_entry, injection_for_element, iter_cmbx_entries, load_cmbx_package, report_templates_for_sequence, safe_filename, split_cmbx_sequences, summarize_package
import export_service
from export_service import export_element, export_foq_contract_report
from instrument_method_parser import discover_external_instrument_methods, parse_instrument_method_txt
from kb_index import (
    discover_kb_index_entries,
    filter_kb_index_entries,
    kb_index_category_full_markdown,
    kb_index_category_options,
    kb_index_entry_category,
    kb_index_entry_detail,
    kb_index_entry_full_markdown,
    kb_index_entry_group,
    kb_index_entries_for_scope,
    kb_index_group_options,
    kb_index_overview_markdown,
    kb_index_scope_full_markdown,
    parse_kb_index_entries,
)
from method_xml_flow import build_method_flow_from_xml, build_method_flow_rows, build_method_flow_tsv
from method_contract import build_method_contract_from_flow_rows, method_contract_summary_text, method_contracts_tsv
from method_semantic_analyzer import analyze_cm_method_rows, cm_numeric_value
from method_script_kb import load_method_script_rows_from_kb
from method_md_linter import lint_method_rows
from tools.render_cm_method_md import parse_md_to_rows
from report_template_md_compiler import compile_report_template_md_to_cmbx, parse_report_template_md
from test_intent_contract import build_test_intent_contract_coverages, contract_coverages_tsv
from chromeleon_runtime import chromeleon_bin
from embedded_report_extractor import decode_report_template_xml, extract_embedded_report_template, parse_report_sheet_objects, parse_report_sheets, report_sheet_objects_tsv, report_sheets_tsv
from foq_result_locations import (
    filter_locations_for_report,
    locations_for_device_type,
    read_device_type_mappings,
    summarize_locations,
)
from foq_contract_report import build_report_cell_value_map, dependency_trace_for_contract_value, write_foq_contract_workbook
import foq_alignment_catalog
from foq_alignment_catalog import (
    build_test_plan_modification_steps,
    build_foq_alignment_records,
    build_cmbx_generation_strategy_kb,
    build_cross_kb_mapping_rows,
    build_intent_conflict_rows,
    build_tcc_contract_closure_task_rows,
    build_tcc_evidence_workstream_rows,
    build_tcc_black_box_coverage_rows,
    build_tcc_milestone_status_rows,
    build_tcc_next_action_queue_rows,
    build_tcc_open_verification_topic_rows,
    build_tcc_p1_evidence_extraction_plan_rows,
    build_tcc_processing_method_target_rows,
    build_tcc_report_formula_extraction_plan_rows,
    build_tcc_report_formula_target_rows,
    build_tcc_relationship_rows,
    build_tcc_temperature_contract_matrix_rows,
    build_test_knowledge_nodes,
    build_tkn_db_mapping_audits,
    build_tkn_coverage_audits,
    cmbx_generation_strategy_markdown,
    filter_alignment_records,
    intent_tool_options,
    record_detail_sections,
    record_intent_gate,
    record_intent_parameter_impact,
    record_intent_preview,
    record_modifiability_summary,
    render_intent_conflict_matrix_markdown,
    render_test_plan_assistant_markdown,
    render_intent_action_plan_markdown,
    render_intent_review_markdown,
    test_knowledge_nodes_markdown as render_test_knowledge_nodes_markdown,
    write_cmbx_generation_strategy_markdown,
    write_foq_alignment_workbook,
    write_intent_draft_asset_packet,
    write_intent_action_plan_markdown,
    write_intent_review_markdown,
    write_test_knowledge_nodes_markdown,
)
from report_calculation_map import build_report_calculation_map, extract_definition_criteria
from report_formula_evaluator import (
    AuditRecord,
    FormulaEvaluation,
    SignalPoint,
    audit_ret_times,
    build_report_formula_context,
    evaluate_audit_property_formula,
    evaluate_audit_metadata_formula,
    evaluate_chm_drift,
    evaluate_chm_noise,
    evaluate_chm_signal_formula,
    evaluate_chm_signal_value,
    evaluate_chm_sig_value_average,
    formula_evaluations_tsv,
    read_audit_ret_times_tsv,
    read_audit_records_tsv,
    read_signal_tsv,
)
from external_report_spec import ExternalReportOperation, parse_external_report_md
from external_report_engine import (
    InjectionReportResult,
    ReportValue,
    audit_event_rows,
    raw_event_rows,
    safe_expression,
    write_external_report_workbook,
)
from formula_catalog import build_formula_catalog, external_scalar_block, filter_formula_catalog, unified_md_block, useful_direct_formula_catalog
from report_workbook_builder import build_accuracy_rows, build_heatup_cooldown_report, write_report_workbook
from sequence_cmd_parser import build_embedded_object_summary, build_injection_method_links, get_injection_method_link
from processing_method_inspector import inspect_processing_method
from sequence_cmd_probe import (
    sequence_cmd_hit_clusters,
    sequence_cmd_hit_clusters_tsv,
    sequence_cmd_injection_links_tsv,
    sequence_cmd_injection_record_probes,
    sequence_cmd_injection_record_probes_tsv,
    sequence_cmd_name_hits,
    sequence_cmd_name_hits_tsv,
    sequence_order_comparison,
    sequence_order_comparison_tsv,
)


def test_external_report_md_parser_reads_operations():
    text = '''
---
report_name: Valve Review
spec_version: 1.0
---
## Data Requirements
```yaml
channels: [CC_Temp, ValvePosition]
audit_paths: [ColumnComp.UpperValve.CurrentPosition]
ret_times: [1, 2]
```
### Scalar: MeanTemp
```yaml
label: Mean temperature
channel: CC_Temp
formula: chm.sig_value("average", 0, 5)
number_format: 0.00
```
### Formula: Deviation
```yaml
expression: MeanTemp - 20
```
### Audit Event Table: ValveEvents
```yaml
property_paths: [ColumnComp.UpperValve.CurrentPosition]
value_changes_only: true
```
'''
    spec = parse_external_report_md(text, from_text=True)
    assert spec.name == "Valve Review"
    assert spec.channels == ("CC_Temp", "ValvePosition")
    assert spec.ret_times == (1, 2)
    assert [operation.kind for operation in spec.operations] == ["scalar", "formula", "audit_event_table"]


def test_external_report_audit_events_keep_only_changes_and_group():
    records = [
        AuditRecord(1.0, "ColumnComp", "UpperValve.CurrentPosition", "1_2"),
        AuditRecord(1.01, "ColumnComp", "UpperValve.CurrentPosition", "1_2"),
        AuditRecord(1.02, "ColumnComp", "UpperValve.CurrentPosition", "6_1"),
        AuditRecord(1.025, "ColumnComp", "LowerValve.CurrentPosition", "6_1"),
    ]
    operation = ExternalReportOperation(
        "ValveEvents",
        "audit_event_table",
        property_paths=("UpperValve.CurrentPosition", "LowerValve.CurrentPosition"),
        value_changes_only=True,
        group_within_seconds=1.0,
    )
    rows = audit_event_rows(records, operation)
    assert len(rows) == 2
    assert rows[1]["events"] == 2
    assert "6_1" in rows[1]["value"]


def test_external_report_raw_event_detection_and_safe_formula(tmp_path):
    signal = [SignalPoint(0.0, 0.0), SignalPoint(0.1, 1.2), SignalPoint(0.2, 0.8), SignalPoint(0.3, -0.2)]
    operation = ExternalReportOperation("Edges", "raw_event_table", channel="X", threshold=1.0, edge="both")
    rows = raw_event_rows(signal, operation)
    assert [row["edge"] for row in rows] == ["rising", "falling"]
    assert safe_expression("max(a, b) - min(a, b)", {"a": 2.5, "b": 1.0}) == 1.5

    spec = parse_external_report_md(
        """---\nreport_name: Test\n---\n## Data Requirements\n```yaml\nchannels: []\n```\n### Formula: Total\n```yaml\nexpression: 1 + 2\n```""",
        from_text=True,
    )
    result = InjectionReportResult(Path("sample.cmbx"), "Seq", "Inj", True, values=[ReportValue("Total", "Total", 3, "ok")])
    output = write_external_report_workbook(spec, [result], tmp_path / "external.xlsx")
    assert output.exists()


def test_formula_catalog_search_and_insert_blocks():
    entries = build_formula_catalog(MODULE_ROOT / "docs")
    formulaone = filter_formula_catalog(entries, "ABS", engine="FormulaOne")
    external = filter_formula_catalog(entries, "signalValue", external_only=True)
    assert formulaone
    assert external
    useful = useful_direct_formula_catalog(entries)
    assert useful
    assert all(row.engine == "CM Report" and row.support == "External V1" and row.formula for row in useful)
    assert not any("RetTimeN" in row.formula or "..." in row.formula for row in useful)
    leak_calibration = next(row for row in useful if row.formula.casefold() == "precond.liquidleakcalibrationvalue")
    assert "FORMULA_INVENTORY" in leak_calibration.source.upper()
    assert "### Workbook Formula:" in unified_md_block(formulaone[0])
    assert "### Scalar:" in external_scalar_block(external[0])


def test_kb_index_parser_reads_versions_table():
    markdown = """
# CMBX Knowledge Index

## Knowledge Base Versions

| KB Name | Version | Update Date | Coverage | Status | Local File(s) |
|---|---:|---|---|---|---|
| TCC_KB | 1.0 | 2026-07-09 | VH/VC/VA-C10-A | Published | `FOQ/TCC/TCC.md`, `FOQ/TCC/Logic.md` |
| Pump_KB | 0.5 | 2026-07-10 target | Vanquish Pump | In development | `FOQ/Pump/Pump.md` |

## Other Section
"""
    entries = parse_kb_index_entries(markdown)

    assert len(entries) == 2
    assert entries[0].kb_name == "TCC_KB"
    assert entries[0].version == "1.0"
    assert entries[0].local_files == ("FOQ/TCC/TCC.md", "FOQ/TCC/Logic.md")
    assert entries[1].update_date == "2026-07-10 target"


def _cm_row(row: int, command: str, value: str = "", comment: str = "", kind: str = "Command"):
    return (str(row), kind, "", command, value, comment, "")


def test_method_script_kb_loads_full_cm_preview_rows(tmp_path: Path):
    kb_root = tmp_path / "KB"
    method_dir = kb_root / "CMBX Method Scripts" / "TCC" / "VH"
    method_dir.mkdir(parents=True)
    flow_path = method_dir / "TEMPERATURE_ACCURACY_embedded_method_flow.tsv"
    flow_path.write_text(
        "\t".join(["Action", "Level", "Time", "Stage", "Target", "Value", "Comment", "Condition"]) + "\n"
        + "\t".join(["STAGE", "0", "{Initial Time}", "InstrumentSetup", "", "", "", ""]) + "\n"
        + "\t".join(["COMMENT", "0", "", "", "", "", "Set temperature ladder", ""]) + "\n"
        + "\t".join(["IF", "0", "", "", "", "", "", 'ColumnComp.ModelNo="VH-C10-A"']) + "\n"
        + "\t".join(["COMMAND", "1", "", "", "Variables.GenericDouble3", "40.0", "Third test temperature", ""]) + "\n"
        + "\t".join(["END IF", "0", "", "", "", "", "", ""]) + "\n"
        + "\t".join(["STAGE", "0", "", "Equilibration", "", "", "Duration = 30.000 [min]", ""]) + "\n"
        + "\t".join(["END", "0", "", "", "", "", "", ""]) + "\n",
        encoding="utf-8",
    )

    rows = load_method_script_rows_from_kb("TEMPERATURE_ACCURACY", family="TCC", device_model="VH-C10-A", kb_root=kb_root)

    assert rows[0][:5] == ("0", "Stage", "{Initial Time}", "Instrument Setup", "")
    assert rows[1][1:5] == ("Comment", "", "Set temperature ladder", "")
    assert rows[2][1:5] == ("Branch", "If", "", 'ColumnComp.ModelNo="VH-C10-A"')
    assert rows[3][1:5] == ("Command", "", "    Variables.GenericDouble3", "40.0")
    assert rows[4][1:5] == ("Branch", "End If", "", "")
    assert rows[5][1:5] == ("Stage", "", "Equilibration", "Duration = 30.000 [min]")
    assert rows[6][1:5] == ("End", "", "End", "")


def test_method_semantic_analyzer_recognizes_accuracy_variable_ladder():
    rows = [
        _cm_row(1, "Variables.GenericDouble1", "10.0"),
        _cm_row(2, "Variables.GenericDouble2", "20.0"),
        _cm_row(3, "Variables.GenericDouble3", "40.0"),
        _cm_row(10, "ColumnComp.CC.Temperature.Nominal", "Variables.GenericDouble1"),
        _cm_row(11, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(12, "RetTimes.RetTime1", "System.Retention"),
        _cm_row(20, "ColumnComp.CC.Temperature.Nominal", "Variables.GenericDouble2"),
        _cm_row(21, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(22, "RetTimes.RetTime2", "System.Retention"),
        _cm_row(30, "ColumnComp.CC.Temperature.Nominal", "Variables.GenericDouble3"),
        _cm_row(31, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(32, "RetTimes.RetTime3", "System.Retention"),
        _cm_row(40, "ColumnComp.CC.Temperature.Nominal", "20.0 [degC]", "return to safe state"),
        _cm_row(41, "StabVars.TriggerStab1", "0"),
    ]

    semantic = analyze_cm_method_rows(rows)

    assert semantic.temperature_variables == (
        "Variables.GenericDouble1",
        "Variables.GenericDouble2",
        "Variables.GenericDouble3",
    )
    assert [block.setpoint.numeric_value for block in semantic.measurement_blocks[:3]] == [10.0, 20.0, 40.0]
    assert [event.ret_time for event in semantic.ret_times if event.emission] == [
        "RetTimes.RetTime1",
        "RetTimes.RetTime2",
        "RetTimes.RetTime3",
    ]
    assert semantic.blocks_for_setpoint(40.0)[0].setpoint.variable == "Variables.GenericDouble3"
    assert semantic.safety_reset_rows == (12,)


def test_method_semantic_analyzer_recognizes_heatup_cooldown_transition_anchors():
    rows = [
        _cm_row(1, "ColumnComp.CC.Temperature.Nominal", "20.0 [degC]"),
        _cm_row(2, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(3, "RetTimes.RetTime1", "System.Retention"),
        _cm_row(4, "ColumnComp.CC.Temperature.Nominal", "50.0 [degC]"),
        _cm_row(5, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(6, "RetTimes.RetTime3", "System.Retention"),
        _cm_row(7, "RetTimes.RetTime4", "System.Retention"),
        _cm_row(8, "ColumnComp.CC.Temperature.Nominal", "20.0 [degC]"),
        _cm_row(9, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(10, "RetTimes.RetTime6", "System.Retention"),
    ]

    semantic = analyze_cm_method_rows(rows)

    assert semantic.temperature_variables == ()
    assert [event.numeric_value for event in semantic.temperature_setpoints] == [20.0, 50.0, 20.0]
    assert [event.ret_time for event in semantic.ret_times if event.emission] == [
        "RetTimes.RetTime1",
        "RetTimes.RetTime3",
        "RetTimes.RetTime4",
        "RetTimes.RetTime6",
    ]
    assert semantic.blocks_for_setpoint(50.0)[0].ret_times[0].ret_time == "RetTimes.RetTime3"
    assert semantic.safety_reset_rows == ()


def test_method_semantic_analyzer_recognizes_stability_fixed_setpoint_without_ladder():
    rows = [
        _cm_row(1, "ColumnComp.CC.Temperature.Nominal", "70.0 [degC]"),
        _cm_row(2, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(3, "ColumnComp.CC_U_Temp_Actual.Data_Collection_Rate", "20"),
        _cm_row(4, "ColumnComp.CC_L_Temp_Actual.Data_Collection_Rate", "20"),
        _cm_row(5, "Delay", "70"),
    ]

    semantic = analyze_cm_method_rows(rows)

    assert semantic.temperature_variables == ()
    assert len(semantic.temperature_setpoints) == 1
    assert semantic.temperature_setpoints[0].numeric_value == 70.0
    assert len(semantic.measurement_blocks) == 1
    assert semantic.measurement_blocks[0].waits[0].condition.startswith("ColumnComp.CC.TempReady")
    assert semantic.ret_times == ()


def test_method_semantic_analyzer_recognizes_precision_aux_variables_are_not_temperature_ladder():
    rows = [
        _cm_row(1, "Variables.GenericLong9", "12", "page count for VH"),
        _cm_row(2, "Variables.GenericBool0", "1", "corrective stability insertion flag"),
        _cm_row(3, "ColumnComp.CC.Temperature.Nominal", "40.0 [degC]"),
        _cm_row(4, "Wait", "ColumnComp.CC.TempReady AND ExternalTempStable"),
        _cm_row(5, "Delay", "60"),
        _cm_row(6, "ColumnComp.FanSpeed", "High"),
    ]

    semantic = analyze_cm_method_rows(rows)

    assert [item.variable for item in semantic.variables] == [
        "Variables.GenericLong9",
        "Variables.GenericBool0",
    ]
    assert semantic.temperature_variables == ()
    assert semantic.temperature_setpoints[0].source == "literal"
    assert semantic.temperature_setpoints[0].numeric_value == 40.0
    assert semantic.ret_times == ()


def test_method_semantic_analyzer_recognizes_periodic_valve_trigger_role():
    rows = [
        _cm_row(1, "ColumnComp.CC.Temperature.Nominal", "70.0 [degC]"),
        _cm_row(2, "Variables.GenericBool1", "1"),
        _cm_row(3, "System.Trigger", "Ping, Variables.GenericBool1=1 AND System.Retention>Variables.GenericFloat1 and System.Retention>10 and System.Retention<20, TrueTime=0.1, Delay=0, AllowImmediateExecution=Yes"),
        _cm_row(4, "Variables.GenericFloat1", "System.Retention+0.1"),
        _cm_row(5, "ColumnComp.UpperValve.CurrentPosition", "6_1"),
        _cm_row(6, "ColumnComp.LowerValve.CurrentPosition", "1_2"),
        _cm_row(7, "Log", "ColumnComp.UpperValve.CurrentPosition"),
        _cm_row(8, "Log", "ColumnComp.CC_Temp"),
        _cm_row(9, "Variables.GenericBool1", "0"),
        _cm_row(10, "Variables.GenericBool2", "1"),
        _cm_row(11, "System.Trigger", "Pong, Variables.GenericBool2=1 AND System.Retention>Variables.GenericFloat1 and System.Retention>10 and System.Retention<20, TrueTime=0.1, Delay=0, AllowImmediateExecution=Yes"),
        _cm_row(12, "Variables.GenericFloat1", "System.Retention+0.1"),
        _cm_row(13, "ColumnComp.UpperValve.CurrentPosition", "1_2"),
        _cm_row(14, "ColumnComp.LowerValve.CurrentPosition", "6_1"),
    ]

    semantic = analyze_cm_method_rows(rows)

    assert len(semantic.triggers) == 2
    assert semantic.triggers[0].name == "Ping"
    assert semantic.triggers[0].bool_gate == "Variables.GenericBool1"
    assert semantic.triggers[0].scheduler_variable == "Variables.GenericFloat1"
    assert semantic.triggers[0].rearm_minutes == 0.1
    assert semantic.triggers[0].time_window_start == 10.0
    assert semantic.triggers[0].time_window_end == 20.0
    assert semantic.triggers[0].valve_positions == (
        "ColumnComp.UpperValve.CurrentPosition=6_1",
        "ColumnComp.LowerValve.CurrentPosition=1_2",
    )
    assert semantic.triggers[0].logged_properties == (
        "ColumnComp.UpperValve.CurrentPosition",
        "ColumnComp.CC_Temp",
    )


def test_kb_index_entry_detail_is_review_friendly():
    entry = parse_kb_index_entries(
        """
| KB Name | Version | Update Date | Coverage | Status | Local File(s) |
|---|---:|---|---|---|---|
| CMBX_Generation_Strategy | 1.0 | 2026-07-09 | Cross-module generation strategy | Published | `cmbx_data_explorer/docs/CMBX_GENERATION_STRATEGY_KB.md` |
"""
    )[0]
    detail = kb_index_entry_detail(entry)

    assert "KB Name: CMBX_Generation_Strategy" in detail
    assert "Cross-module generation strategy" in detail
    assert "- cmbx_data_explorer/docs/CMBX_GENERATION_STRATEGY_KB.md" in detail


def test_kb_index_classifies_filters_and_renders_local_markdown(tmp_path):
    kb_root = tmp_path / "KB"
    tcc_file = kb_root / "FOQ" / "TCC" / "TCC.md"
    detector_file = kb_root / "FOQ" / "Detector" / "Detector.md"
    tcc_file.parent.mkdir(parents=True)
    detector_file.parent.mkdir(parents=True)
    tcc_file.write_text("# TCC KB\n\nTemperature Calibration details.", encoding="utf-8")
    detector_file.write_text("# Detector KB\n\nNoise details.", encoding="utf-8")
    index_path = kb_root / "KB_INDEX.md"
    index_path.write_text(
        """
| KB Name | Version | Update Date | Coverage | Status | Local File(s) |
|---|---:|---|---|---|---|
| TCC_KB | 1.0 | 2026-07-09 | VH/VC/VA-C10-A | Published | `FOQ/TCC/TCC.md` |
| VDAD_KB | 1.0 | 2026-07-08 | VDAD-F/C, VMWD-C | Published | `FOQ/Detector/Detector.md` |
| CMBX_Formulas | 1.2 | 2026-06-20 | Report formula rules | Published | `missing.md` |
""",
        encoding="utf-8",
    )

    entries = parse_kb_index_entries(index_path.read_text(encoding="utf-8"))

    assert kb_index_entry_category(entries[0]) == "FOQ 测试知识 / TCC"
    assert kb_index_entry_category(entries[1]) == "FOQ 测试知识 / Detector"
    assert kb_index_entry_category(entries[2]) == "报告与公式 / Formula与DB"
    assert "FOQ 测试知识 / TCC" in kb_index_category_options(entries)
    assert kb_index_group_options(entries) == ("FOQ测试知识", "报告与公式")
    assert filter_kb_index_entries(entries, category="FOQ 测试知识 / TCC") == (entries[0],)
    assert filter_kb_index_entries(entries, search_text="detector") == (entries[1],)
    assert kb_index_entries_for_scope(entries, "group:FOQ测试知识") == (entries[0], entries[1])
    assert kb_index_entries_for_scope(entries, "category:FOQ 测试知识 / TCC") == (entries[0],)

    rendered = kb_index_entry_full_markdown(entries[0], index_path=index_path)

    assert "# TCC_KB" in rendered
    assert "Source:" in rendered
    assert "Temperature Calibration details." in rendered

    category_rendered = kb_index_category_full_markdown(entries, "FOQ 测试知识 / TCC", index_path=index_path)

    assert "# FOQ 测试知识 / TCC" in category_rendered
    assert "KB count: 1" in category_rendered
    assert "TCC_KB" in category_rendered
    assert "Temperature Calibration details." in category_rendered

    all_rendered = kb_index_category_full_markdown(entries, "All", index_path=index_path)

    assert "# All Knowledge Bases" in all_rendered
    assert "KB count: 3" in all_rendered
    assert "Temperature Calibration details." in all_rendered
    assert "Noise details." in all_rendered

    group_rendered = kb_index_scope_full_markdown(entries, "group:FOQ测试知识", index_path=index_path)

    assert "# FOQ测试知识 Knowledge Bases" in group_rendered
    assert "KB count: 2" in group_rendered
    assert "| FOQ测试知识 | FOQ 测试知识 / TCC | TCC_KB" in group_rendered
    assert "Temperature Calibration details." in group_rendered
    assert "Noise details." in group_rendered


def test_kb_index_discovers_unlisted_markdown_and_builds_overview(tmp_path):
    kb_root = tmp_path / "KB"
    command_file = kb_root / "CM" / "Instrument Commands" / "CM_INSTRUMENT_COMMAND_KNOWLEDGE_BASE_V2.md"
    command_file.parent.mkdir(parents=True)
    command_file.write_text("# CM Commands\n\nColumnOven command notes.", encoding="utf-8")
    index_path = kb_root / "KB_INDEX.md"
    index_path.write_text(
        """
| KB Name | Version | Update Date | Coverage | Status | Local File(s) |
|---|---:|---|---|---|---|
| TCC_KB | 1.0 | 2026-07-09 | VH/VC/VA-C10-A | Published | `missing_tcc.md` |
""",
        encoding="utf-8",
    )

    entries = parse_kb_index_entries(index_path.read_text(encoding="utf-8"))
    discovered = discover_kb_index_entries(index_path, existing_entries=entries)

    assert len(discovered) == 1
    assert discovered[0].kb_name == "Instrument Commands / CM_INSTRUMENT_COMMAND_KNOWLEDGE_BASE_V2"
    assert kb_index_entry_category(discovered[0]) == "方法脚本知识 / CM命令"
    assert kb_index_entry_group(discovered[0]) == "方法脚本知识"

    all_entries = (*entries, *discovered)
    overview = kb_index_overview_markdown(all_entries)
    assert "| 方法脚本知识 | 方法脚本知识 / CM命令 | 1 | 0 | 1 |" in overview

    rendered = kb_index_category_full_markdown(all_entries, "方法脚本知识 / CM命令", index_path=index_path)
    assert "ColumnOven command notes." in rendered


def test_foq_alignment_catalog_includes_tcc_temperature_accuracy():
    records = build_foq_alignment_records(kb_root=Path("__missing_kb_root__"))
    tcc = filter_alignment_records(records, family="TCC", devices=("VH-C10-A",), test_text="temperature_accuracy")

    assert len(tcc) == 1
    record = tcc[0]
    assert record.injection == "Temperature Accuracy_H"
    assert record.instrument_method == "TEMPERATURE_ACCURACY"
    assert record.processing_method == "ACCURACY_IRC_STOP_H"
    assert "Report_VTCC_V2_12" in record.report_template
    assert "Temp Accuracy" in record.report_sheets
    assert "RetTime3" in record.expected_ret_times


def test_foq_alignment_catalog_includes_full_tcc_sequence_map():
    records = build_foq_alignment_records(kb_root=Path("__missing_kb_root__"))
    tcc = filter_alignment_records(records, family="TCC", devices=("VH-C10-A",))
    injections = {record.injection for record in tcc}

    assert {
        "ColumnIDs",
        "Preheater Connection Test",
        "Valve",
        "VTCC_BurnIn",
        "Temperature Calibration",
        "Temperature Accuracy_H",
        "Temperature Precision_and_Fan",
        "Temperature Stability_and_PCC_H",
        "HeatUp and CoolDownTime",
        "LiquidLeaktest",
        "Qualification_Service_Done",
        "Factory Default",
        "Error Log Check",
    }.issubset(injections)
    column_id = next(record for record in tcc if record.injection == "ColumnIDs")
    assert column_id.instrument_method == "ColumnID"
    assert column_id.processing_method == "CORRECT_STABILITY_INJ_INSERTION"


def test_foq_alignment_catalog_enriches_cmbx_injection_method_links(monkeypatch):
    package = SimpleNamespace(
        path=Path("sample.cmbx"),
        injections=(SimpleNamespace(name="Temperature Accuracy_H"),),
        methods_and_reports=(
            SimpleNamespace(name="TEMPERATURE_ACCURACY", kind="instrument_method"),
            SimpleNamespace(name="Report_VTCC_V2_12", kind="report_template"),
        ),
    )
    link = SimpleNamespace(
        injection_name="Temperature Accuracy_H",
        processing_method="ACCURACY_IRC_STOP_H",
        instrument_method="TEMPERATURE_ACCURACY",
    )
    monkeypatch.setattr(foq_alignment_catalog, "build_injection_method_links", lambda _package: {"Temperature Accuracy_H": link})

    records = build_foq_alignment_records(packages=(package,), kb_root=Path("__missing_kb_root__"))
    record = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]

    assert any("link Temperature Accuracy_H -> TEMPERATURE_ACCURACY / ACCURACY_IRC_STOP_H" in item for item in record.cmbx_sources)
    assert any("sequence link confirms injection Temperature Accuracy_H" in item for item in record.method_evidence)


def test_foq_alignment_catalog_marks_vdad_3d_field_not_applicable_to_vmwd():
    records = build_foq_alignment_records(kb_root=Path("__missing_kb_root__"))
    vdad_records = filter_alignment_records(records, family="VDAD", test_text="3d_field")

    assert len(vdad_records) == 1
    record = vdad_records[0]
    assert record.coverage_status == "not applicable"
    assert "VMWD-C" not in record.device_models
    assert any("VMWD-C" in gap for gap in record.open_gaps)


def test_foq_alignment_detail_sections_and_workbook_export(tmp_path):
    record = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC", test_text="heatup")[0]
    sections = record_detail_sections(record)

    assert set(sections) == {
        "TKN Node",
        "Cross-KB Mapping",
        "TD Meaning",
        "Method Evidence",
        "Report Evidence",
        "DB Evidence",
        "Design Actions",
        "Milestone Status",
        "Next Action Queue",
        "M2 Temperature Matrix",
        "M2 Closure Tasks",
        "M2 Evidence Workstreams",
        "M2 P1 Extraction Plan",
        "M2 Processing Targets",
        "M2 Report Targets",
        "M2 Report Extraction Plan",
        "BlackBox Audit",
        "Open Verification Topics",
        "Dependency Impact",
        "Relationship Audit",
        "Resolution Choices",
        "Open Verification",
        "Generation Readiness",
    }
    assert "test_id:" in sections["TKN Node"]
    assert "FOQ:" in sections["Cross-KB Mapping"]
    assert "HeatUp" in sections["TD Meaning"]
    assert "Modifiability:" in sections["Design Actions"]
    assert "Known Cut / Merge / Lock Points" in sections["Design Actions"]
    assert "TCC Knowledge Engineering Milestone Status" in sections["Milestone Status"]
    assert "Current Milestone: M2" in sections["Milestone Status"]
    assert "M5" in sections["Milestone Status"]
    assert "runnable CMBX generation remains closed" in sections["Milestone Status"]
    assert "TCC Next Action Queue" in sections["Next Action Queue"]
    assert "Processing method decode / CM UI" in sections["Next Action Queue"]
    assert "Report workbook/formula extraction" in sections["Next Action Queue"]
    assert "blocks method/report generation confidence" in sections["Next Action Queue"]
    assert "M2 Temperature Contract Matrix" in sections["M2 Temperature Matrix"]
    assert "TCC_CAL_01" in sections["M2 Temperature Matrix"]
    assert "TCC_HEATCOOL_01 <- current" in sections["M2 Temperature Matrix"]
    assert "Next closure actions" in sections["M2 Temperature Matrix"]
    assert "M2 Contract Closure Tasks" in sections["M2 Closure Tasks"]
    assert "TCC_HEATCOOL_01 <- current" not in sections["M2 Closure Tasks"]
    assert "Evidence Source" in sections["M2 Closure Tasks"]
    assert "P1" in sections["M2 Closure Tasks"]
    assert "Report workbook/formula extraction" in sections["M2 Closure Tasks"]
    assert "M2 Evidence Workstreams" in sections["M2 Evidence Workstreams"]
    assert "Processing method decode / CM UI" in sections["M2 Evidence Workstreams"]
    assert "Report workbook/formula extraction" in sections["M2 Evidence Workstreams"]
    assert "Instrument setup/config manifest" in sections["M2 Evidence Workstreams"]
    assert "M2 P1 Evidence Extraction Plan" in sections["M2 P1 Extraction Plan"]
    assert "Configuration evidence" in sections["M2 P1 Extraction Plan"]
    assert "Method command decode" in sections["M2 P1 Extraction Plan"]
    assert "Report workbook/formula extraction" in sections["M2 P1 Extraction Plan"]
    assert "planned - extraction not yet executed" in sections["M2 P1 Extraction Plan"]
    assert "M2 Processing Method Targets" in sections["M2 Processing Targets"]
    assert "CORRECT_ACCURACY_INJ_INSERTION" in sections["M2 Processing Targets"]
    assert "TD-backed corrective accuracy insertion" in sections["M2 Processing Targets"]
    assert "GenericBool0 pass inserts Temperature Accuracy_H" in sections["M2 Processing Targets"]
    assert "M2 Report Formula Extraction Targets" in sections["M2 Report Targets"]
    assert "Temp_Calib_Internal" in sections["M2 Report Targets"]
    assert "TempCal85_U" in sections["M2 Report Targets"]
    assert "Temp Precision, Fan" in sections["M2 Report Targets"]
    assert "HeatUp&CoolDown" not in sections["M2 Report Targets"]
    assert "FormulaOne workbook formulas" in sections["M2 Report Targets"]
    assert "open - extract and verify workbook evidence" in sections["M2 Report Targets"]
    assert "M2 Report Formula Extraction Plan" in sections["M2 Report Extraction Plan"]
    assert "Locate `Report_VTCC_V2_12`" in sections["M2 Report Extraction Plan"]
    assert "SpreadSheetData` / FormulaOne" in sections["M2 Report Extraction Plan"]
    assert "updated black-box Contract 3" in sections["M2 Report Extraction Plan"]
    assert "TCC_HEATUP_COOLDOWN_BLACK_BOX_DECOMPOSITION.md" in sections["BlackBox Audit"]
    assert "Six-contract audit:" in sections["BlackBox Audit"]
    assert "Open Verification count:" in sections["BlackBox Audit"]
    assert "Evidence Sources section: Yes" in sections["BlackBox Audit"]
    assert "No parsed black-box open-verification topics are mapped for this row" in sections["Open Verification Topics"]
    assert "Exact FormulaOne workbook route" not in sections["Open Verification Topics"]
    assert "CORRECT_ACCURACY_INJ_INSERTION" not in sections["Open Verification Topics"]
    assert "Closure action:" not in sections["Open Verification Topics"]
    assert "Mermaid sketch:" in sections["Dependency Impact"]
    assert "Structured relationship rules" in sections["Dependency Impact"]
    assert "INTENT_02" in sections["Dependency Impact"]
    assert "Matched relationship rows:" in sections["Relationship Audit"]
    assert "INTENT_02" in sections["Relationship Audit"]
    assert "Shared Resource" in sections["Relationship Audit"]
    assert "Relationship Resolution Choices" in sections["Resolution Choices"]
    assert "Calibration scope decision" in sections["Resolution Choices"]
    assert "reuse existing Temperature Calibration evidence" in sections["Resolution Choices"]
    assert "Closure action:" in sections["Open Verification"]
    output = write_foq_alignment_workbook((record,), tmp_path / "alignment.xlsx", intent="Crop / Modify", parameter="20->50->20")
    assert output.exists()
    import openpyxl

    workbook = openpyxl.load_workbook(output, read_only=True)
    alignment_headers = [cell.value for cell in next(workbook["FOQ Knowledge Alignment"].iter_rows(max_row=1))]
    assert "Modifiability" in alignment_headers
    assert "Intent Gate" in alignment_headers
    assert "Generic Draft Packet" in alignment_headers
    rows = list(workbook["FOQ Knowledge Alignment"].iter_rows(values_only=True))
    assert any(row[2] == "HeatUp and CoolDown" and row[alignment_headers.index("Generic Draft Packet")] == "Yes" for row in rows)
    assert "Intent Gate Matrix" in workbook.sheetnames
    gate_rows = list(workbook["Intent Gate Matrix"].iter_rows(values_only=True))
    gate_headers = gate_rows[0]
    assert "Blockers" in gate_headers
    assert "Next Actions" in gate_headers
    assert "Relationship Rules" in gate_headers
    assert any(
        row[2] == "HeatUp and CoolDown"
        and row[gate_headers.index("Generic Draft Packet")] == "Yes"
        and "INTENT_02" in (row[gate_headers.index("Relationship Rules")] or "")
        for row in gate_rows[1:]
    )
    assert "TCC Milestone Status" in workbook.sheetnames
    milestone_rows = list(workbook["TCC Milestone Status"].iter_rows(values_only=True))
    assert {row[0] for row in milestone_rows[1:]} == {"M1", "M2", "M3", "M4", "M5"}
    assert any(row[0] == "M5" and "runnable generation closed" in row[3] for row in milestone_rows[1:])
    assert "TCC Next Action Queue" in workbook.sheetnames
    next_action_rows = list(workbook["TCC Next Action Queue"].iter_rows(values_only=True))
    next_action_headers = next_action_rows[0]
    assert "Evidence Group" in next_action_headers
    assert "Generation Gate" in next_action_headers
    assert next_action_rows[1][next_action_headers.index("Milestone")] == "M2"
    assert next_action_rows[1][next_action_headers.index("Evidence Group")] == "Configuration evidence"
    assert "blocks method/report generation confidence" in next_action_rows[1][next_action_headers.index("Generation Gate")]
    assert "TCC BlackBox Coverage" in workbook.sheetnames
    coverage_rows = list(workbook["TCC BlackBox Coverage"].iter_rows(values_only=True))
    coverage_headers = coverage_rows[0]
    assert "Contract 1 Method" in coverage_headers
    assert "Open Verification Count" in coverage_headers
    assert "Open Verification Topics" in coverage_headers
    assert "Evidence Sources Present" in coverage_headers
    assert "Model Branches Mentioned" in coverage_headers
    assert "Mermaid Present" in coverage_headers
    assert "Word Count" in coverage_headers
    assert any(row[1] == "TCC_ACC_01" and row[coverage_headers.index("Exists")] == "Yes" for row in coverage_rows[1:])
    assert any(row[1] == "TCC_ERRORLOG_01" and row[coverage_headers.index("Contract 6 Open Verification")] == "Yes" for row in coverage_rows[1:])
    assert "M2 Temperature Contract Matrix" in workbook.sheetnames
    temp_matrix_rows = list(workbook["M2 Temperature Contract Matrix"].iter_rows(values_only=True))
    temp_matrix_headers = temp_matrix_rows[0]
    assert "Template Readiness" in temp_matrix_headers
    assert "Next Closure Actions" in temp_matrix_headers
    assert any(
        row[temp_matrix_headers.index("Test ID")] == "TCC_HEATCOOL_01"
        and "candidate template after CM validation" in row[temp_matrix_headers.index("Template Readiness")]
        for row in temp_matrix_rows[1:]
    )
    assert "M2 Contract Closure Tasks" in workbook.sheetnames
    closure_task_rows = list(workbook["M2 Contract Closure Tasks"].iter_rows(values_only=True))
    closure_task_headers = closure_task_rows[0]
    assert "Contract" in closure_task_headers
    assert "Priority" in closure_task_headers
    assert "Evidence Group" in closure_task_headers
    assert "Likely Evidence Source" in closure_task_headers
    assert "Closure Action" in closure_task_headers
    assert not any(row[closure_task_headers.index("Test ID")] == "TCC_HEATCOOL_01" for row in closure_task_rows[1:])
    assert any(
        row[closure_task_headers.index("Test ID")] == "TCC_CAL_01"
        and row[closure_task_headers.index("Contract")] == "Contract 2 Processing Method"
        and row[closure_task_headers.index("Priority")] == "P2"
        for row in closure_task_rows[1:]
    )
    assert "M2 Evidence Workstreams" in workbook.sheetnames
    workstream_rows = list(workbook["M2 Evidence Workstreams"].iter_rows(values_only=True))
    workstream_headers = workstream_rows[0]
    assert "Evidence Group" in workstream_headers
    assert "Unlocks" in workstream_headers
    assert "Next Action" in workstream_headers
    assert any(
        row[workstream_headers.index("Priority")] == "P1"
        and row[workstream_headers.index("Evidence Group")] == "Configuration evidence"
        and "Instrument setup/config manifest" in row[workstream_headers.index("Unlocks")]
        for row in workstream_rows[1:]
    )
    assert "M2 P1 Extraction Plan" in workbook.sheetnames
    p1_plan_rows = list(workbook["M2 P1 Extraction Plan"].iter_rows(values_only=True))
    p1_plan_headers = p1_plan_rows[0]
    assert "Evidence Group" in p1_plan_headers
    assert "Extraction Steps" in p1_plan_headers
    assert "Validation Outputs" in p1_plan_headers
    assert any(
        row[p1_plan_headers.index("Test ID")] == "TCC_CAL_01"
        and row[p1_plan_headers.index("Evidence Group")] == "Configuration evidence"
        and "updated black-box contract" in row[p1_plan_headers.index("Validation Outputs")]
        for row in p1_plan_rows[1:]
    )
    assert "M2 Processing Targets" in workbook.sheetnames
    processing_target_rows = list(workbook["M2 Processing Targets"].iter_rows(values_only=True))
    processing_target_headers = processing_target_rows[0]
    assert "Processing Method" in processing_target_headers
    assert "Expected Behavior" in processing_target_headers
    assert "Extraction Target" in processing_target_headers
    assert any(
        row[processing_target_headers.index("Test ID")] == "TCC_CAL_01"
        and row[processing_target_headers.index("Device")] == "VH-C10-A"
        and row[processing_target_headers.index("Processing Method")] == "CORRECT_ACCURACY_INJ_INSERTION"
        and "GenericBool0 pass inserts Temperature Accuracy_H" in row[processing_target_headers.index("Expected Behavior")]
        for row in processing_target_rows[1:]
    )
    assert "M2 Report Formula Targets" in workbook.sheetnames
    report_target_rows = list(workbook["M2 Report Formula Targets"].iter_rows(values_only=True))
    report_target_headers = report_target_rows[0]
    assert "Device" in report_target_headers
    assert "Report Template" in report_target_headers
    assert "Extraction Target" in report_target_headers
    assert "Readiness" in report_target_headers
    assert any(
        row[report_target_headers.index("Test ID")] == "TCC_CAL_01"
        and row[report_target_headers.index("Device")] == "(not bound)"
        and row[report_target_headers.index("Readiness")] == "open - no matching alignment row"
        for row in report_target_rows[1:]
    )
    assert any(
        row[report_target_headers.index("Test ID")] == "TCC_PRECISION_01"
        and row[report_target_headers.index("Device")] == "(not bound)"
        and row[report_target_headers.index("Readiness")] == "open - no matching alignment row"
        for row in report_target_rows[1:]
    )
    assert not any(row[report_target_headers.index("Test ID")] == "TCC_HEATCOOL_01" for row in report_target_rows[1:])
    assert "M2 Report Extraction Plan" in workbook.sheetnames
    report_plan_rows = list(workbook["M2 Report Extraction Plan"].iter_rows(values_only=True))
    report_plan_headers = report_plan_rows[0]
    assert "Extraction Steps" in report_plan_headers
    assert "Validation Outputs" in report_plan_headers
    assert any(
        row[report_plan_headers.index("Test ID")] == "TCC_CAL_01"
        and row[report_plan_headers.index("Device")] == "(not bound)"
        and row[report_plan_headers.index("Formula ID")] == "FORMULA_OPEN_VERIFICATION_REQUIRED"
        and row[report_plan_headers.index("Status")] == "planned - extraction not yet executed"
        for row in report_plan_rows[1:]
    )
    assert not any(row[report_plan_headers.index("Test ID")] == "TCC_HEATCOOL_01" for row in report_plan_rows[1:])
    assert "TCC Open Verification Topics" in workbook.sheetnames
    topic_rows = list(workbook["TCC Open Verification Topics"].iter_rows(values_only=True))
    topic_headers = topic_rows[0]
    assert "Category" in topic_headers
    assert "Likely Evidence Source" in topic_headers
    assert "Closure Action" in topic_headers
    assert any(row[1] == "TCC_CAL_01" and row[topic_headers.index("Category")] == "Processing Method" for row in topic_rows[1:])
    assert "Test Knowledge Nodes" in workbook.sheetnames
    assert "Cross-KB Mapping" in workbook.sheetnames
    assert "TKN Coverage Audit" in workbook.sheetnames
    assert "TKN DB Mapping Audit" in workbook.sheetnames
    assert "Generation Strategy" in workbook.sheetnames
    assert "TCC Relationship Model" in workbook.sheetnames
    assert "Selected Relationship Audit" in workbook.sheetnames
    relationship_rows = list(workbook["TCC Relationship Model"].iter_rows(values_only=True))
    assert any(row[1] == "DEP_01" and row[2] == "Temperature Calibration" and row[3] == "Temperature Accuracy" for row in relationship_rows)
    assert any(row[1] == "INTENT_01" and "single setpoint crop" in row[3] for row in relationship_rows)
    selected_relationship_rows = list(workbook["Selected Relationship Audit"].iter_rows(values_only=True))
    selected_relationship_headers = selected_relationship_rows[0]
    assert "Rule ID" in selected_relationship_headers
    assert any(row[selected_relationship_headers.index("Rule ID")] == "INTENT_02" for row in selected_relationship_rows[1:])
    assert "Relationship Resolution Choices" in workbook.sheetnames
    resolution_rows = list(workbook["Relationship Resolution Choices"].iter_rows(values_only=True))
    resolution_headers = resolution_rows[0]
    assert "Decision Required" in resolution_headers
    assert any(row[0] == "ORDER_02" and "Calibration scope decision" in row[resolution_headers.index("Decision Required")] for row in resolution_rows[1:])
    assert "Selected Resolution Choices" in workbook.sheetnames
    selected_resolution_rows = list(workbook["Selected Resolution Choices"].iter_rows(values_only=True))
    selected_resolution_headers = selected_resolution_rows[0]
    assert "Evidence To Capture" in selected_resolution_headers
    assert any(row[selected_resolution_headers.index("Rule ID")] == "ORDER_02" for row in selected_resolution_rows[1:])
    assert "Resolution Decision Register" in workbook.sheetnames
    decision_rows = list(workbook["Resolution Decision Register"].iter_rows(values_only=True))
    decision_headers = decision_rows[0]
    assert "Selected Option" in decision_headers
    assert "Decision Status" in decision_headers
    assert "Evidence Path" in decision_headers
    assert any(
        row[decision_headers.index("Rule ID")] == "ORDER_02"
        and row[decision_headers.index("Decision Status")] == "Open"
        for row in decision_rows[1:]
    )


def test_foq_alignment_modifiability_marks_design_workbench_actions():
    records = build_foq_alignment_records(kb_root=Path("__missing_kb_root__"))
    heatup = filter_alignment_records(records, family="TCC", test_text="heatup")[0]
    calibration = filter_alignment_records(records, family="TCC", test_text="temperature_calibration")[0]

    assert record_modifiability_summary(heatup) == "🟡 editable after review"
    assert "temperature range" in record_detail_sections(heatup)["Design Actions"]
    assert "Temperature Calibration --> HeatUp/CoolDown" in record_detail_sections(heatup)["Dependency Impact"]
    assert "custom temperature range" in record_detail_sections(heatup)["Relationship Audit"]
    assert record_modifiability_summary(calibration) == "🟡 verify before editing"
    assert "calibration ladder writes device calibration variables" in record_detail_sections(calibration)["Design Actions"]
    assert "TCC_CALIBRATION_BLACK_BOX_DECOMPOSITION.md" in record_detail_sections(calibration)["BlackBox Audit"]
    assert "Model branches mentioned: VH-C10-A, VC-C10-A, VA-C10-A" in record_detail_sections(calibration)["BlackBox Audit"]


def test_tcc_next_action_queue_prioritizes_m2_processing_and_report_work():
    rows = build_tcc_next_action_queue_rows()

    assert rows
    assert rows[0].rank == 1
    assert rows[0].milestone == "M2"
    assert rows[0].priority == "P1"
    assert rows[0].evidence_group == "Configuration evidence"
    assert rows[0].task_count == 1
    assert "GenericBool0" in rows[0].primary_blocker
    assert "blocks method/report generation confidence" in rows[0].generation_gate
    assert any(row.milestone == "M2" and row.priority == "P1" and row.evidence_group == "Method command decode" for row in rows)
    assert any(row.milestone == "M2" and row.evidence_group == "Report workbook/formula extraction" for row in rows)
    assert any(row.milestone == "M2" and row.priority == "P2" and row.evidence_group == "Processing method decode / CM UI" for row in rows)
    assert any(row.milestone == "M3" for row in rows)


def test_foq_alignment_intent_preview_covers_crop_merge_compare_and_search():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    stability = filter_alignment_records(records, family="TCC", test_text="stability")[0]

    assert "Crop / Modify" in intent_tool_options()

    crop = record_intent_preview(
        accuracy,
        "Crop / Modify",
        "40 C",
        selected_records=(accuracy,),
        candidate_records=records,
    )
    assert "Intent: Crop / Modify" in crop
    assert "Intent Gate:" in crop
    assert "Specialized draft packet: available" in crop
    assert "Single-point 40 C accuracy" in crop
    assert "RetTimes" in crop
    assert "ExtTemp_UpperCC" in crop
    assert "DB output should target only the selected TempAcc field" in crop
    assert "Structured parameter impact" in crop
    assert "Selected DB fields: TempAcc40, RES_TempAccuracy" in crop
    assert "Removed / unused DB fields" in crop
    assert "Relationship model rules" in crop
    assert "DEP_01" in crop
    assert "INTENT_01" in crop
    assert "Blocked for runnable generation" in crop

    merge = record_intent_preview(
        accuracy,
        "Merge",
        "",
        selected_records=(accuracy, stability),
        candidate_records=records,
    )
    assert "Intent: Merge" in merge
    assert "Intent Gate:" in merge
    assert "Selected rows:" in merge
    assert "Report templates involved" in merge
    assert "Processing methods involved" in merge
    assert "Relationship model rules" in merge
    assert "Open verification inherited by merge" in merge

    compare = record_intent_preview(accuracy, "Compare", "", selected_records=(accuracy,), candidate_records=records)
    assert "Intent: Compare" in compare
    assert "| Device | Injection | Instrument Method | Processing Method | Report Template | DB Fields |" in compare
    assert "VH-C10-A" in compare
    assert "Temperature Accuracy_H" in compare
    assert "Report_VATCC_V1_01" in compare

    search = record_intent_preview(accuracy, "Search / Recommend", "heat", selected_records=(accuracy,), candidate_records=records)
    assert "Intent: Search / Recommend" in search
    assert "Intent Gate:" in search
    assert "Recommended alignment rows" in search
    assert "HeatUp" in search


def test_foq_alignment_accuracy_crop_parameter_impact_maps_setpoint_to_db_fields():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]

    vh_impact = record_intent_parameter_impact(accuracy, "Crop / Modify", "40 C", device_model="VH-C10-A")

    assert vh_impact.parameter_kind == "temperature_accuracy_setpoint"
    assert vh_impact.supported
    assert vh_impact.setpoint_c == 40.0
    assert vh_impact.affected_models == ("VH-C10-A",)
    assert "TempAcc40" in vh_impact.selected_db_fields
    assert "RES_TempAccuracy" in vh_impact.selected_db_fields
    assert "TempAcc10" in vh_impact.removed_db_fields
    assert "TempAcc120" in vh_impact.removed_db_fields

    all_branch_impact = record_intent_parameter_impact(accuracy, "Crop / Modify", "40 C")

    assert all_branch_impact.supported
    assert all_branch_impact.affected_models == ("VH-C10-A", "VC-C10-A", "VA-C10-A")
    assert "TempAcc40" in all_branch_impact.selected_db_fields
    assert "RES_TempAccuracy" in all_branch_impact.selected_db_fields
    assert "TempAcc85" in all_branch_impact.removed_db_fields
    assert "TempAcc120" in all_branch_impact.removed_db_fields

    unsupported = record_intent_parameter_impact(accuracy, "Crop / Modify", "85 C", device_model="VH-C10-A")

    assert not unsupported.supported
    assert any("does not expose TempAcc85" in note for note in unsupported.notes)


def test_foq_alignment_intent_conflict_matrix_renders_selected_merge_scope():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    stability = filter_alignment_records(records, family="TCC", test_text="stability")[0]

    conflict_rows = build_intent_conflict_rows((accuracy, stability), device_model="VH-C10-A", intent="Merge")

    assert any(row.aspect == "Instrument Method" and row.status == "multiple values - review" for row in conflict_rows)
    assert any(row.aspect == "DB Fields" and row.status == "union required" and "TempAcc40" in row.values and "TempStability" in row.values for row in conflict_rows)
    assert any(row.category == "Relationships" and row.status == "union required" and "DEP_01" in row.values for row in conflict_rows)
    assert any(row.category == "Open Verification" and row.status == "open verification" for row in conflict_rows)

    markdown = render_intent_conflict_matrix_markdown((accuracy, stability), device_model="VH-C10-A", intent="Merge")

    assert "# Intent Conflict Matrix" in markdown
    assert "Selected rows: 2" in markdown
    assert "multiple values - review" in markdown
    assert "union required" in markdown
    assert "TEMPERATURE_STABILITY_AND_PCC_70_H" in markdown


def test_foq_alignment_intent_gate_is_structured_for_generation_tools():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    heatup = filter_alignment_records(records, family="TCC", test_text="heatup")[0]

    accuracy_gate = record_intent_gate(accuracy, "Crop / Modify", "40 C", selected_records=(accuracy,))
    assert accuracy_gate.can_export_generic_packet
    assert accuracy_gate.can_export_specialized_packet
    assert not accuracy_gate.runnable_generation_allowed
    assert "Specialized Temperature Accuracy single-setpoint draft packet can be exported for review." in accuracy_gate.next_actions
    assert any("Black-box open verification topics remain by category" in blocker for blocker in accuracy_gate.blockers)
    assert any("Hard relationship rule DEP_01" in blocker for blocker in accuracy_gate.blockers)
    assert any("Hard relationship rule RES_01" in blocker for blocker in accuracy_gate.blockers)
    assert any("Close Processing Method topic for TCC_ACC_01" in action for action in accuracy_gate.next_actions)
    assert any("Resolve relationship DEP_01" in action for action in accuracy_gate.next_actions)

    heatup_gate = record_intent_gate(heatup, "Crop / Modify", "20->50->20", selected_records=(heatup,))
    assert heatup_gate.can_export_generic_packet
    assert not heatup_gate.can_export_specialized_packet
    assert not heatup_gate.runnable_generation_allowed
    assert not any("Processing Method" in blocker for blocker in heatup_gate.blockers)
    assert not any("Report Formula" in blocker for blocker in heatup_gate.blockers)
    assert any("Hard relationship rule ORDER_02" in blocker for blocker in heatup_gate.blockers)
    assert not any("Close Processing Method topic for TCC_HEATCOOL_01" in action for action in heatup_gate.next_actions)

    merge_gate = record_intent_gate(accuracy, "Merge", "", selected_records=(accuracy,))
    assert "Merge preview needs at least two selected alignment rows." in merge_gate.blockers


def test_foq_alignment_intent_review_markdown_export(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    markdown = render_intent_review_markdown(
        accuracy,
        "Crop / Modify",
        "40 C",
        selected_records=(accuracy,),
        candidate_records=records,
    )

    assert "# Intent Review - TCC - Temperature Accuracy" in markdown
    assert "This review packet is non-mutating" in markdown
    assert "| Intent | Crop / Modify |" in markdown
    assert "| Parameter | 40 C |" in markdown
    assert "| Intent Gate | blocked for runnable generation |" in markdown
    assert "| Specialized Draft Packet | available |" in markdown
    assert "## Anchor Test Knowledge Node" in markdown
    assert "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION" in markdown
    assert "## Required Contracts" in markdown
    assert "ExtTemp_UpperCC" in markdown
    assert "Single-point 40 C accuracy" in markdown
    assert "Only after Method, Processing, Report, DB, and Config contracts are closed" in markdown

    output = write_intent_review_markdown(
        tmp_path / "intent_review.md",
        accuracy,
        "Crop / Modify",
        "40 C",
        selected_records=(accuracy,),
        candidate_records=records,
    )
    assert output.exists()
    assert output.read_text(encoding="utf-8") == markdown


def test_foq_alignment_intent_action_plan_markdown_export(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    markdown = render_intent_action_plan_markdown(
        accuracy,
        "Crop / Modify",
        "40 C",
        selected_records=(accuracy,),
        candidate_records=records,
    )

    assert "# Intent Action Plan - TCC - Temperature Accuracy" in markdown
    assert "| Intent | Crop / Modify |" in markdown
    assert "| Parameter | 40 C |" in markdown
    assert "Edit Gate" in markdown
    assert "| Generic Draft Packet | available |" in markdown
    assert "| Runnable Generation | closed until CM validation |" in markdown
    assert "## Intent Gate Blockers" in markdown
    assert "Hard relationship rule DEP_01" in markdown
    assert "Resolve relationship DEP_01" in markdown
    assert "## Relationship Resolution Choices" in markdown
    assert "Accuracy calibration dependency decision" in markdown
    assert "include Temperature Calibration in the generated/cropped sequence" in markdown
    assert "## Structured Parameter Impact" in markdown
    assert "Selected DB fields: TempAcc40, RES_TempAccuracy" in markdown
    assert "Removed / unused DB fields" in markdown
    assert "Layered Modification Tasks" in markdown
    assert "| Method | Define the 40 C approach/baseline rule before deleting neighboring accuracy setpoints." in markdown
    assert "| Processing | Review IRC/pass-action behavior before removing or shortening related injections." in markdown
    assert "| Report | Remove, hide, or mark unused TempAcc setpoint rows as not applicable" in markdown
    assert "| DB | Select the DB field subset that remains meaningful after the crop." in markdown
    assert "| Config | Verify required device symbols/channels exist in the target CM configuration." in markdown
    assert "## Intent Conflict Matrix" in markdown
    assert "## Conflict-Driven Required Actions" in markdown
    assert "Open Verification / Open gaps" in markdown
    assert "## Open Verification Closure Queue" in markdown
    assert "| Test ID | Category | Topic | Likely Evidence Source | Closure Action |" in markdown
    assert "| TCC_ACC_01 |" in markdown
    assert "Processing Method" in markdown
    assert "Report Formula" in markdown
    assert "Modified package is validated in Chromeleon" in markdown

    output = write_intent_action_plan_markdown(
        tmp_path / "intent_action_plan.md",
        accuracy,
        "Crop / Modify",
        "40 C",
        selected_records=(accuracy,),
        candidate_records=records,
    )
    assert output.exists()
    assert output.read_text(encoding="utf-8") == markdown

    stability = filter_alignment_records(records, family="TCC", test_text="stability")[0]
    merge_markdown = render_intent_action_plan_markdown(
        accuracy,
        "Merge",
        "",
        selected_records=(accuracy, stability),
        candidate_records=records,
    )
    assert "## Intent Conflict Matrix" in merge_markdown
    assert "multiple values - review" in merge_markdown
    assert "union required" in merge_markdown
    assert "Method / Instrument Method" in merge_markdown
    assert "DB / DB Fields" in merge_markdown
    assert "Review method command contracts and symbol compatibility." in merge_markdown


def test_foq_alignment_draft_asset_packet_for_tcc_accuracy_40c(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]

    paths = write_intent_draft_asset_packet(
        tmp_path,
        accuracy,
        "Crop / Modify",
        "40 C",
        device_model="VH-C10-A",
        selected_records=(accuracy,),
        candidate_records=records,
    )

    project_dir = paths["project_dir"]
    assert project_dir.exists()
    assert (project_dir / "project_spec.json").exists()
    assert (project_dir / "method_script_40C_only.txt").exists()
    assert (project_dir / "config_method_report_review.md").exists()
    assert (project_dir / "report_calculation_spec.md").exists()
    assert (project_dir / "report_formula_map_40C.tsv").exists()
    assert paths["manifest"].exists()
    assert paths["intent_review"].exists()
    assert paths["intent_action_plan"].exists()
    assert paths["relationship_decision_register"].exists()
    assert paths["sequence_template"].exists()
    assert paths["processing_method_binding"].exists()
    assert paths["config_method_report_review"].exists()
    assert paths["method_excel"].exists()
    assert paths["report_excel"].exists()

    manifest = paths["manifest"].read_text(encoding="utf-8")
    assert "# Draft Asset Packet - TCC - Temperature Accuracy" in manifest
    assert "VH-C10-A" in manifest
    assert "not a Chromeleon-signed method" in manifest
    assert "sequence_template.tsv" in manifest
    assert "processing_method_binding.md" in manifest
    assert "required_configuration.md" in manifest
    assert "config_method_report_review.md" in manifest
    assert "Core Black-Box Review" in manifest
    assert "Instrument Config" in manifest
    assert "Instrument Method Script" in manifest
    assert "Report Formula" in manifest
    assert "Processing Method and DB mapping are recorded as downstream evidence only" in manifest
    assert "intent_action_plan.md" in manifest
    assert "relationship_decision_register.tsv" in manifest
    assert "Relationship Resolution Choices" in manifest
    assert "Accuracy calibration dependency decision" in manifest
    decision_register = paths["relationship_decision_register"].read_text(encoding="utf-8")
    assert "Selected Option" in decision_register
    assert "Decision Status" in decision_register
    assert "Accuracy calibration dependency decision" in decision_register
    assert "Open" in decision_register
    action_plan = paths["intent_action_plan"].read_text(encoding="utf-8")
    assert "Define the 40 C approach/baseline rule" in action_plan
    core_review = paths["config_method_report_review"].read_text(encoding="utf-8")
    assert "Config -> Method -> Report Review" in core_review
    assert "This is an intent-level review for a single Temperature Accuracy setpoint crop" in core_review
    assert "Thermometer1.ExtTemp_UpperCC" in core_review
    assert "SET ColumnComp.CC.Temperature.Nominal = 40" in core_review
    assert "AUDIT.RetTime3" in core_review
    assert "RES_TempAccuracy" in core_review
    assert "Binary method/report payload generation is not proven" in core_review
    spec = (project_dir / "project_spec.json").read_text(encoding="utf-8")
    assert '"db_field": "TempAcc40"' in spec


def test_foq_alignment_draft_asset_packet_requires_supported_intent_and_single_device(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]

    with pytest.raises(ValueError, match="Select exactly one device model"):
        write_intent_draft_asset_packet(tmp_path, accuracy, "Crop / Modify", "40 C")
    compare_paths = write_intent_draft_asset_packet(tmp_path, accuracy, "Compare", "40 C", device_model="VH-C10-A")
    assert compare_paths["method_report_binding"].exists()
    assert compare_paths["generation_boundary"].exists()
    compare_manifest = compare_paths["manifest"].read_text(encoding="utf-8")
    assert "Generic Draft Asset Packet Manifest" in compare_manifest
    assert "review-only" in compare_manifest
    assert "Relationship Resolution Choices" in compare_manifest

    missing_parameter_paths = write_intent_draft_asset_packet(tmp_path, accuracy, "Crop / Modify", "", device_model="VH-C10-A")
    assert missing_parameter_paths["sequence_template"].exists()
    missing_boundary = missing_parameter_paths["generation_boundary"].read_text(encoding="utf-8")
    assert "Generation Boundary" in missing_boundary
    assert "Blocked for runnable generation" in missing_boundary


def test_foq_alignment_generic_draft_asset_packet_for_heatup(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    heatup = filter_alignment_records(records, family="TCC", test_text="heatup")[0]

    paths = write_intent_draft_asset_packet(
        tmp_path,
        heatup,
        "Crop / Modify",
        "20->50->20",
        device_model="VH-C10-A",
        selected_records=(heatup,),
        candidate_records=records,
    )

    assert paths["manifest"].exists()
    assert paths["sequence_template"].exists()
    assert paths["method_report_binding"].exists()
    assert paths["config_contract"].exists()
    assert paths["config_method_report_review"].exists()
    assert paths["report_db_contract"].exists()
    assert paths["intent_conflict_matrix"].exists()
    assert paths["relationship_decision_register"].exists()
    sequence = paths["sequence_template"].read_text(encoding="utf-8")
    assert "HeatUp and CoolDownTime" in sequence
    assert "TEMP_HEAT_UP_DOWN_20_50_20" in sequence
    binding = paths["method_report_binding"].read_text(encoding="utf-8")
    assert "Generic Draft Asset Packet" in binding
    assert "HeatUp_Time_20to50" in paths["report_db_contract"].read_text(encoding="utf-8")
    boundary = paths["generation_boundary"].read_text(encoding="utf-8")
    assert "Intent Gate" in boundary
    assert "Generic draft packet: available" in boundary
    assert "P1 / Instrument Config" in boundary
    assert "P1 / Instrument Method Script" in boundary
    assert "P1 / Report Formula" in boundary
    assert "P2 / Processing Method" in boundary
    assert "P2 / DB" in boundary
    assert "Relationship Model Rules" in boundary
    assert "Relationship Resolution Choices" in boundary
    assert "Calibration scope decision" in boundary
    assert "INTENT_02" in boundary
    register = paths["relationship_decision_register"].read_text(encoding="utf-8")
    assert "Selected Option" in register
    assert "Decision Status" in register
    assert "Calibration scope decision" in register
    manifest = paths["manifest"].read_text(encoding="utf-8")
    assert "config_method_report_review.md" in manifest
    assert "Core Black-Box Review" in manifest
    assert "Instrument Config -> Instrument Method Script -> Report Formula" in manifest
    core_review = paths["config_method_report_review"].read_text(encoding="utf-8")
    assert "Config -> Method -> Report Review" in core_review
    assert "HeatUp and CoolDown" in core_review
    assert "TEMP_HEAT_UP_DOWN_20_50_20" in core_review
    assert "RetTime2 - RetTime1 - 2.0" in core_review
    assert "RetTime5 - RetTime4 - 2.0" in core_review
    assert "Processing Method and DB mapping are downstream checks" in core_review
    conflict = paths["intent_conflict_matrix"].read_text(encoding="utf-8")
    assert "single-row review" in conflict
    assert "Expected RetTimes" in conflict


def test_foq_alignment_test_plan_assistant_shows_intent_templates_and_change_plan():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    heatup = filter_alignment_records(records, family="TCC", test_text="heatup")[0]

    accuracy_plan = render_test_plan_assistant_markdown(
        accuracy,
        "Crop / Modify",
        "40 C",
        device_model="VH-C10-A",
        selected_records=(accuracy,),
        candidate_records=records,
    )
    assert "Test Plan Assistant" in accuracy_plan
    assert "## 1. Intent Source" in accuracy_plan
    assert "## 2. Editable Templates From Current Framework" in accuracy_plan
    assert "## 3. Modification / Change Plan" in accuracy_plan
    assert "Specialized draft packet" in accuracy_plan
    assert "Instrument method template" in accuracy_plan
    assert "Report template" in accuracy_plan
    assert "Primary proof order: Instrument Config -> Instrument Method Script -> Report Formula" in accuracy_plan
    assert "Processing Method is a downstream CM automation check" in accuracy_plan

    heatup_plan = render_test_plan_assistant_markdown(
        heatup,
        "Crop / Modify",
        "20->50->20",
        device_model="VH-C10-A",
        selected_records=(heatup,),
        candidate_records=records,
    )
    assert "Generic draft packet" in heatup_plan
    assert "TEMP_HEAT_UP_DOWN_20_50_20" in heatup_plan
    assert "RetTime2 - RetTime1 - 2.0" in heatup_plan
    assert "RetTime5 - RetTime4 - 2.0" in heatup_plan
    assert "DB mapping is a downstream export/upload check" in heatup_plan


def test_test_plan_modification_steps_are_concrete_for_accuracy_and_heatup():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    heatup = filter_alignment_records(records, family="TCC", test_text="heatup")[0]

    accuracy_steps = build_test_plan_modification_steps(
        accuracy,
        "Crop / Modify",
        "40 C",
        device_model="VH-C10-A",
    )
    joined_accuracy = "\n".join("\t".join(row) for row in accuracy_steps)
    assert "TEMPERATURE_ACCURACY" in joined_accuracy
    assert "[Equilibration] and [Run] setpoint ladder" in joined_accuracy
    assert "[Run] RetTime anchor `RetTime3`" in joined_accuracy
    assert "`Temp Accuracy` row 68" in joined_accuracy
    assert "method_script_40C.txt" in joined_accuracy
    assert "report_calculation.xlsx" in joined_accuracy
    assert "required_configuration.md" in joined_accuracy

    heatup_steps = build_test_plan_modification_steps(
        heatup,
        "Crop / Modify",
        "20->50->20",
        device_model="VH-C10-A",
    )
    joined_heatup = "\n".join("\t".join(row) for row in heatup_steps)
    assert "TEMP_HEAT_UP_DOWN_20_50_20" in joined_heatup
    assert "[Run] 17 C precondition -> 20 C -> 50 C -> 20 C command flow" in joined_heatup
    assert "RetTime2-RetTime1-2.0" in joined_heatup
    assert "RetTime5-RetTime4-2.0" in joined_heatup
    assert "dedicated CM script renderer not implemented" in joined_heatup


def test_foq_alignment_generic_draft_asset_packet_preserves_merge_selection(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    stability = filter_alignment_records(records, family="TCC", test_text="stability")[0]

    paths = write_intent_draft_asset_packet(
        tmp_path,
        accuracy,
        "Merge",
        "",
        device_model="VH-C10-A",
        selected_records=(accuracy, stability),
        candidate_records=records,
    )

    sequence = paths["sequence_template"].read_text(encoding="utf-8")
    assert "Temperature Accuracy_H" in sequence
    assert "Temperature Stability_and_PCC_H" in sequence
    binding = paths["method_report_binding"].read_text(encoding="utf-8")
    assert "Selected Row Bindings" in binding
    assert "Temperature Accuracy" in binding
    assert "Temperature Stability" in binding
    report_db = paths["report_db_contract"].read_text(encoding="utf-8")
    assert "TempAcc40" in report_db
    assert "TempStability" in report_db
    conflict = paths["intent_conflict_matrix"].read_text(encoding="utf-8")
    assert "Instrument Method" in conflict
    assert "TEMPERATURE_ACCURACY" in conflict
    assert "TEMPERATURE_STABILITY_AND_PCC_70_H" in conflict
    assert "multiple values - review" in conflict
    assert "union required" in conflict
    boundary = paths["generation_boundary"].read_text(encoding="utf-8")
    assert "Relationship Model Rules" in boundary
    assert "DEP_01" in boundary
    assert "DEP_03" in boundary


def test_foq_alignment_crop_draft_packet_preserves_multirow_selection(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    stability = filter_alignment_records(records, family="TCC", test_text="stability")[0]

    paths = write_intent_draft_asset_packet(
        tmp_path,
        accuracy,
        "Crop / Modify",
        "40 C",
        device_model="VH-C10-A",
        selected_records=(accuracy, stability),
        candidate_records=records,
    )

    assert "method_excel" not in paths
    assert paths["method_report_binding"].exists()
    assert paths["intent_conflict_matrix"].exists()
    sequence = paths["sequence_template"].read_text(encoding="utf-8")
    assert "Temperature Accuracy_H" in sequence
    assert "Temperature Stability_and_PCC_H" in sequence
    boundary = paths["generation_boundary"].read_text(encoding="utf-8")
    assert "Generic draft packet: available" in boundary
    assert "Specialized draft packet: not available" in boundary


def test_foq_alignment_draft_asset_packet_rejects_device_incompatible_selection(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    accuracy = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    vh_stability = filter_alignment_records(records, family="TCC", test_text="temperature_stability_and_pcc")[0]

    with pytest.raises(ValueError, match="not applicable"):
        write_intent_draft_asset_packet(
            tmp_path,
            accuracy,
            "Merge",
            "",
            device_model="VA-C10-A",
            selected_records=(accuracy, vh_stability),
            candidate_records=records,
        )


def test_test_knowledge_node_model_from_tcc_alignment_record():
    records = build_foq_alignment_records(kb_root=Path("__missing_kb_root__"))
    record = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    node = build_test_knowledge_nodes((record,))[0]

    assert node.test_id == "TCC_ACC_01"
    assert node.test_name == "Temperature Accuracy"
    assert "Temperature Accuracy" in node.foq_section
    assert node.injection == "Temperature Accuracy_H"
    assert node.instrument_method == "TEMPERATURE_ACCURACY"
    assert node.processing_method == "ACCURACY_IRC_STOP_H"
    assert "Report_VTCC_V2_12" in node.report_template
    assert node.report_sheets == ("Temp Accuracy",)
    assert node.formula_id == "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION"
    assert "TempAcc40" in node.db_fields
    assert "RetTime3" in node.expected_ret_times
    assert "ExtTemp_UpperCC" in node.expected_channels
    assert "ColumnComp.ModelNo" in node.expected_audit_properties
    assert node.coverage_status == "partial"
    assert "VH-C10-A" in node.model_applicability
    assert node.irc_injected
    assert any("Definitions!Temperature Accuracy" in item for item in node.acceptance_criteria)
    bindings = {binding.device_model: binding for binding in node.device_bindings}
    assert bindings["VH-C10-A"].injection == "Temperature Accuracy_H"
    assert bindings["VH-C10-A"].processing_method == "ACCURACY_IRC_STOP_H"
    assert "TempAcc120" in bindings["VH-C10-A"].db_fields
    assert bindings["VC-C10-A"].injection == "Temperature Accuracy_C"
    assert bindings["VC-C10-A"].processing_method == "ACCURACY_IRC_STOP_C"
    assert "TempAcc85" in bindings["VC-C10-A"].db_fields
    assert bindings["VA-C10-A"].report_template == "Report_VATCC_V1_01"


def test_tcc_test_knowledge_nodes_preserve_full_injection_chain():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    nodes = build_test_knowledge_nodes(records)

    assert len(nodes) >= 13
    assert all(node.injection for node in nodes)
    assert all(node.instrument_method for node in nodes)
    assert all(node.processing_method for node in nodes)
    assert {node.test_id for node in nodes}.issuperset(
        {
            "TCC_COL_01",
            "TCC_PREHEATER_01",
            "TCC_ACC_01",
            "TCC_STABILITY_PCC_01",
            "TCC_HEATCOOL_01",
            "TCC_FACTORY_01",
        }
    )
    open_nodes = [node for node in nodes if node.coverage_status != "complete"]
    assert open_nodes
    assert any(node.open_gaps for node in open_nodes)


def test_tcc_test_knowledge_nodes_markdown_export(tmp_path):
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    markdown = render_test_knowledge_nodes_markdown(records)
    output = write_test_knowledge_nodes_markdown(tmp_path / "TCC_TEST_KNOWLEDGE_NODES.md", records)

    assert "| TCC_ACC_01 | Temperature Accuracy | Temperature Accuracy_H | TEMPERATURE_ACCURACY | ACCURACY_IRC_STOP_H |" in markdown
    assert "## Coverage Audit" in markdown
    assert "## DB Mapping Audit" in markdown
    assert "report sheet and formula family mapped" in markdown
    assert "Temperature Accuracy_C.XLS" in markdown
    assert "## Node Contracts" in markdown
    assert "**Device Bindings**" in markdown
    assert "| VC-C10-A | Temperature Accuracy_C | TEMPERATURE_ACCURACY | ACCURACY_IRC_STOP_C | Report_VTCC_V2_12 | Temp Accuracy | (follows injection) | TempAcc10, TempAcc20, TempAcc40, TempAcc60, TempAcc85, RES_TempAccuracy |" in markdown
    assert "Temp_Calib_Internal" in markdown
    assert "Temperature Stability_C" in markdown
    assert "PCC_Drift" in markdown
    assert "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION" in markdown
    assert "TempAcc40" in markdown
    assert "Method Command Contract" in markdown
    assert output.exists()
    assert "TCC_HEATCOOL_01" in output.read_text(encoding="utf-8")


def test_tcc_tkn_coverage_audit_classifies_component_gaps():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=PROJECT_ROOT / "cmbx_data_explorer" / "docs"), family="TCC")
    audits = {audit.test_id: audit for audit in build_tkn_coverage_audits(records)}

    accuracy = audits["TCC_ACC_01"]
    assert accuracy.method_command_status == "bound to instrument method evidence"
    assert accuracy.processing_method_status == "IRC/corrective binding documented"
    assert accuracy.report_formula_status == "report sheet and formula family mapped"
    assert accuracy.db_field_status == "DB fields mapped with evidence"
    assert accuracy.overall_status == "open verification"
    assert any("template-specific row coverage" in gap for gap in accuracy.gaps)

    burnin = audits["TCC_BURNIN_01"]
    assert burnin.db_field_status == "no DB contract expected"
    assert burnin.report_formula_status == "no report formula contract expected"
    assert burnin.overall_status == "open verification"


def test_tcc_tkn_db_mapping_audit_uses_real_mapping_per_device():
    mapping = PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls"
    records = filter_alignment_records(build_foq_alignment_records(kb_root=PROJECT_ROOT / "cmbx_data_explorer" / "docs"), family="TCC", test_text="temperature_accuracy")
    audits = {audit.device_model: audit for audit in build_tkn_db_mapping_audits(records, mapping)}

    vh = audits["VH-C10-A"]
    assert vh.mapped_report_files == ("Temperature Accuracy_H.XLS",)
    assert "TempAcc10" in vh.mapped_db_fields
    assert "TempAcc120" in vh.expected_db_fields
    assert vh.extra_mapped_fields == ()
    assert vh.status == "closed"

    vc = audits["VC-C10-A"]
    assert vc.mapped_report_files == ("Temperature Accuracy_C.XLS",)
    assert "TempAcc85" in vc.mapped_db_fields
    assert "TempAcc85" in vc.expected_db_fields
    assert "TempAcc120" not in vc.expected_db_fields
    assert vc.missing_expected_fields == ()
    assert vc.status == "closed"

    va = audits["VA-C10-A"]
    assert va.injection == "Temperature Accuracy_C"
    assert "TempAcc60" in va.expected_db_fields
    assert va.status == "closed"


def test_tcc_tkn_db_mapping_audit_closes_temperature_family_per_device():
    mapping = PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls"
    records = filter_alignment_records(build_foq_alignment_records(kb_root=PROJECT_ROOT / "cmbx_data_explorer" / "docs"), family="TCC")
    audits = {(audit.test_id, audit.device_model): audit for audit in build_tkn_db_mapping_audits(records, mapping)}

    calibration_vh = audits[("TCC_CAL_01", "VH-C10-A")]
    assert calibration_vh.report_sheets == ("Temp_Calib_Internal",)
    assert "TempCal120_U" in calibration_vh.expected_db_fields
    assert "Slope_Cal05_L" in calibration_vh.expected_db_fields
    assert calibration_vh.status == "closed"

    calibration_vc = audits[("TCC_CAL_01", "VC-C10-A")]
    assert "TempCal85_U" in calibration_vc.expected_db_fields
    assert "Slope_Cal05_L" in calibration_vc.expected_db_fields
    assert "TempCal120_U" not in calibration_vc.expected_db_fields
    assert calibration_vc.status == "closed"

    precision = audits[("TCC_PRECISION_01", "VH-C10-A")]
    assert precision.expected_db_fields == ("TempPrecision", "RES_TempPrecision")
    assert precision.status == "closed"

    precision_va = audits[("TCC_PRECISION_01", "VA-C10-A")]
    assert precision_va.injection == "Temperature Precision"
    assert precision_va.mapped_report_files == ("Temperature Precision_and_Fan.XLS",)
    assert precision_va.status == "closed"

    pcc = audits[("TCC_STABILITY_PCC_01", "VH-C10-A")]
    assert pcc.injection == "Temperature Stability_and_PCC_H"
    assert "PCC_Drift" in pcc.expected_db_fields
    assert pcc.status == "closed"

    stability_vc = audits[("TCC_STABILITY_01", "VC-C10-A")]
    assert stability_vc.injection == "Temperature Stability_C"
    assert stability_vc.expected_db_fields == ("TempStability", "Noise_CC_Temp", "RES_TempStability")
    assert stability_vc.status == "closed"

    factory_va = audits[("TCC_FACTORY_01", "VA-C10-A")]
    assert "SubmitDate" in factory_va.expected_db_fields
    assert factory_va.status == "closed"


def test_tcc_tkn_db_mapping_audit_distinguishes_no_db_contract_nodes():
    mapping = PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls"
    records = filter_alignment_records(build_foq_alignment_records(kb_root=PROJECT_ROOT / "cmbx_data_explorer" / "docs"), family="TCC")
    nodes = {node.test_id: node for node in build_test_knowledge_nodes(records)}
    audits = {(audit.test_id, audit.device_model): audit for audit in build_tkn_db_mapping_audits(records, mapping)}

    assert nodes["TCC_VALVE_01"].db_fields == ()
    assert audits[("TCC_VALVE_01", "VH-C10-A")].status == "no DB contract expected"
    assert audits[("TCC_LEAK_01", "VC-C10-A")].status == "no DB contract expected"
    assert audits[("TCC_SERVICE_01", "VA-C10-A")].status == "no DB contract expected"
    assert audits[("TCC_ERRORLOG_01", "VH-C10-A")].status == "no DB contract expected"

    column_devices = {binding.device_model for binding in nodes["TCC_COL_01"].device_bindings}
    assert column_devices == {"VH-C10-A", "VC-C10-A"}
    assert ("TCC_COL_01", "VA-C10-A") not in audits


def test_cross_kb_mapping_row_from_tcc_alignment_record():
    records = build_foq_alignment_records(kb_root=Path("__missing_kb_root__"))
    record = filter_alignment_records(records, family="TCC", test_text="temperature_accuracy")[0]
    mapping = build_cross_kb_mapping_rows((record,))[0]

    assert mapping.test_id == "TCC_ACC_01"
    assert mapping.foq_test_name == "Temperature Accuracy"
    assert mapping.method_name == "TEMPERATURE_ACCURACY"
    assert mapping.processing_method == "ACCURACY_IRC_STOP_H"
    assert "Report_VTCC_V2_12" in mapping.report_template
    assert mapping.formula_id == "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION"
    assert "TempAcc40" in mapping.db_fields
    assert mapping.mapping_status == "partial"


def test_cmbx_generation_strategy_kb_contains_tcc_and_vdad_rules(tmp_path):
    kb = build_cmbx_generation_strategy_kb()
    method_rule_ids = {rule.rule_id for rule in kb.method_rules}
    formula_ids = {rule.formula_id for rule in kb.formula_rules}
    markdown = cmbx_generation_strategy_markdown(kb)
    output = write_cmbx_generation_strategy_markdown(tmp_path / "CMBX_GENERATION_STRATEGY_KB.md")

    assert {"TCC_METH_01", "TCC_METH_02", "VDAD_METH_02"}.issubset(method_rule_ids)
    assert "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION" in formula_ids
    assert "FORMULA_VDAD_NOISE_REGRESSION_DEVIATION_OPEN" in formula_ids
    assert "Generation Guardrail" in markdown
    assert output.exists()
    assert "CMBX Generation Strategy KB" in output.read_text(encoding="utf-8")
from semantic_generation import (
    available_capability_groups,
    blueprint_report_text,
    blueprint_to_dict,
    build_clone_select_plan,
    build_execution_sequence_plan,
    build_sequence_blueprint,
    clone_select_plan_to_dict,
    execution_sequence_plan_to_dict,
    load_tcc_semantic_catalog,
    load_tcc_symbol_manifest,
    validate_blueprint_against_package,
    validate_blueprint_capabilities,
    validate_execution_plan_against_package,
)
from clone_select_generator import (
    clone_select_candidate_validation_text,
    default_clone_select_output_path,
    validate_clone_select_candidate,
    write_clone_select_cmbx,
)
from cmbx_import_asset_generator import (
    validate_import_asset_candidate,
    write_import_asset_candidate,
)
from tcc_project_generator import (
    build_single_point_temperature_accuracy_project,
    single_point_temperature_accuracy_project_to_dict,
    write_single_point_temperature_accuracy_project,
)
from tcc_cmbx_candidate_generator import write_vh_temperature_accuracy_40c_cmbx_candidates
from vx_c10a_tcc_method_package import (
    build_vx_c10a_tcc_method_package_markdown,
    write_vx_c10a_tcc_method_package_markdown,
)


def test_vx_c10a_tcc_method_package_markdown_contains_model_branches(tmp_path):
    markdown = build_vx_c10a_tcc_method_package_markdown()
    output = write_vx_c10a_tcc_method_package_markdown(tmp_path / "VX_C10_A_TCC_FOQ_CMBX_METHOD_PACKAGE.md")

    assert "VX-C10-A TCC FOQ CMBX Test Method Package" in markdown
    assert "### VA-C10-A" in markdown
    assert "### VC-C10-A" in markdown
    assert "### VH-C10-A" in markdown
    assert "Temperature Accuracy_H" in markdown
    assert "ACCURACY_IRC_STOP_H" in markdown
    assert "TEMPERATURE_STABILITY_AND_PCC_70_H" in markdown
    assert "Report_VATCC_V1_01" in markdown
    assert "Report_VTCC_V2_12" in markdown
    assert "FORMULA_TCC_TEMP_ACCURACY_MAX_DEVIATION" in markdown
    assert "```mermaid" in markdown
    assert "Open Verification Required" in markdown
    assert output.exists()


def test_load_synthetic_cmbx_header(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd" Size="12">
    <ChromeleonElement Id="2" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection">
      <ChromeleonElement Id="3" Name="Channel_A" ItemType="Dionex.Chromeleon.Data.Signal" RawDataFilename="3.raw" Size="100" />
      <ChromeleonElement Id="4" Name="Audit" ItemType="Dionex.Chromeleon.Data.AuditTrail" RawDataFilename="4.raw" Size="200" />
    </ChromeleonElement>
    <ChromeleonElement Id="5" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="6" Name="REPORT_A" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("3.raw", b"Signal")
        archive.writestr("4.raw", b"Audit")
        archive.writestr("Seq1.cmd", b"Sequence")

    package = load_cmbx_package(cmbx)
    counts = summarize_package(package)

    assert [entry.name for entry in iter_cmbx_entries(cmbx)] == ["header.xml", "3.raw", "4.raw", "Seq1.cmd"]
    assert extract_cmbx_entry(cmbx, "3.raw") == b"Signal"
    assert counts["sequences"] == 1
    assert counts["injections"] == 1
    assert counts["channels"] == 1
    assert counts["audits"] == 1
    assert counts["instrument_methods"] == 1
    assert counts["report_templates"] == 1
    assert package.channels[0].kind == "signal"
    assert injection_for_element(package, package.channels[0]).name == "Injection A"


def test_load_cmbx_with_subfolder_root_sequences(tmp_path):
    cmbx = tmp_path / "folder_package.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="VH pilot run FOQ" ItemType="Dionex.Chromeleon.Data.SubFolder">
    <ChromeleonElement Id="2" Name="6000001" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6000001.cmd">
      <ChromeleonElement Id="3" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection">
        <ChromeleonElement Id="4" Name="CC_Temp" ItemType="Dionex.Chromeleon.Data.Signal" RawDataFilename="4.raw" />
      </ChromeleonElement>
    </ChromeleonElement>
    <ChromeleonElement Id="5" Name="6000002" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6000002.cmd">
      <ChromeleonElement Id="6" Name="Injection B" ItemType="Dionex.Chromeleon.Data.Injection">
        <ChromeleonElement Id="7" Name="Audit" ItemType="Dionex.Chromeleon.Data.AuditTrail" RawDataFilename="7.raw" />
      </ChromeleonElement>
    </ChromeleonElement>
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("4.raw", b"Signal")
        archive.writestr("7.raw", b"Audit")

    package = load_cmbx_package(cmbx)
    counts = summarize_package(package)

    assert package.root_elements[0].kind == "folder"
    assert [sequence.name for sequence in package.sequences] == ["6000001", "6000002"]
    assert counts["sequences"] == 2
    assert counts["injections"] == 2
    assert counts["channels"] == 1
    assert counts["audits"] == 1


def test_load_cmbx_with_sequence_folder_root_sequences(tmp_path):
    cmbx = tmp_path / "sequence_folder_package.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Packed Runs" ItemType="Dionex.Chromeleon.Data.SequenceFolder">
    <ChromeleonElement Id="2" Name="6000001" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6000001.cmd" />
    <ChromeleonElement Id="3" Name="6000002" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6000002.cmd" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("6000001.cmd", b"Seq 1")
        archive.writestr("6000002.cmd", b"Seq 2")

    package = load_cmbx_package(cmbx)

    assert package.root_elements[0].kind == "folder"
    assert [sequence.name for sequence in package.sequences] == ["6000001", "6000002"]


def test_split_cmbx_subfolder_sequences_to_standalone_packages(tmp_path):
    cmbx = tmp_path / "packed.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader DateCreated="Tuesday, July 7, 2026">
  <ChromeleonElement Id="1" Name="VH pilot run FOQ" ItemType="Dionex.Chromeleon.Data.SubFolder">
    <ChromeleonElement Id="2" Name="6000001" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6000001.cmd">
      <ChromeleonElement Id="3" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection">
        <ChromeleonElement Id="4" Name="CC_Temp" ItemType="Dionex.Chromeleon.Data.Signal" RawDataFilename="4.raw" />
      </ChromeleonElement>
    </ChromeleonElement>
    <ChromeleonElement Id="5" Name="6000002" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6000002.cmd">
      <ChromeleonElement Id="6" Name="Injection B" ItemType="Dionex.Chromeleon.Data.Injection">
        <ChromeleonElement Id="7" Name="Audit" ItemType="Dionex.Chromeleon.Data.AuditTrail" RawDataFilename="7.raw" />
      </ChromeleonElement>
    </ChromeleonElement>
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("6000001.cmd", b"Seq 1")
        archive.writestr("6000002.cmd", b"Seq 2")
        archive.writestr("4.raw", b"Signal")
        archive.writestr("7.raw", b"Audit")

    package = load_cmbx_package(cmbx)
    outputs = split_cmbx_sequences(package, package.sequences, tmp_path / "out")

    assert [path.name for path in outputs] == ["6000001.cmbx", "6000002.cmbx"]
    first = load_cmbx_package(outputs[0])
    second = load_cmbx_package(outputs[1])
    assert [sequence.name for sequence in first.sequences] == ["6000001"]
    assert [sequence.name for sequence in second.sequences] == ["6000002"]
    assert extract_cmbx_entry(outputs[0], "6000001.cmd") == b"Seq 1"
    assert extract_cmbx_entry(outputs[0], "4.raw") == b"Signal"
    assert extract_cmbx_entry(outputs[1], "7.raw") == b"Audit"


def test_write_import_asset_candidate_preserves_sequence_cmd_payload(tmp_path):
    cmbx = tmp_path / "asset_source.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="METHOD_A" Url="chrom://local/Seq1.seq/METHOD_A.instmeth" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="3" Name="REPORT_A" Url="chrom://local/Seq1.seq/REPORT_A.report" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"Embedded command payload with METHOD_A and REPORT_A")

    package = load_cmbx_package(cmbx)
    output = write_import_asset_candidate(package, "METHOD_A", tmp_path / "method_import.cmbx")
    validation = validate_import_asset_candidate(output)
    output_package = load_cmbx_package(output)

    assert validation.sequence_count == 1
    assert validation.instrument_methods == ("METHOD_A",)
    assert validation.report_templates == ()
    assert validation.has_sequence_payload
    assert extract_cmbx_entry(output, "Seq1.cmd") == b"Embedded command payload with METHOD_A and REPORT_A"
    assert [element.name for element in output_package.methods_and_reports] == ["METHOD_A"]


def test_write_vh_temperature_accuracy_40c_cmbx_candidates(tmp_path):
    cmbx = tmp_path / "vh_source.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="6000001" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6000001.seq_2.cmd">
    <ChromeleonElement Id="2" Name="Temperature Accuracy_H" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="TEMPERATURE_ACCURACY" Url="chrom://local/6000001.seq/TEMPERATURE_ACCURACY.instmeth" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="4" Name="ACCURACY_IRC_STOP_H" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
    <ChromeleonElement Id="5" Name="Report_VTCC_V2_12" Url="chrom://local/6000001.seq/Report_VTCC_V2_12.report" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    sequence_cmd = b"Temperature Accuracy_H TEMPERATURE_ACCURACY ACCURACY_IRC_STOP_H Report_VTCC_V2_12"
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("6000001.seq_2.cmd", sequence_cmd)

    outputs = write_vh_temperature_accuracy_40c_cmbx_candidates(cmbx, tmp_path / "out")
    reference = load_cmbx_package(outputs.reference_cmbx)
    experimental = load_cmbx_package(outputs.experimental_cmbx)
    request = load_cmbx_package(outputs.request_cmbx)

    assert outputs.reference_cmbx.exists()
    assert outputs.experimental_cmbx.exists()
    assert outputs.request_cmbx.exists()
    assert outputs.manifest.exists()
    assert [item.name for item in reference.methods_and_reports] == [
        "TEMPERATURE_ACCURACY",
        "ACCURACY_IRC_STOP_H",
        "Report_VTCC_V2_12",
    ]
    assert [item.name for item in experimental.methods_and_reports] == [
        "TEMPERATURE_ACCURACY__single_40C",
        "ACCURACY_IRC_STOP_H",
        "Report_VTCC_V2_12__single_40C",
    ]
    assert extract_cmbx_entry(outputs.experimental_cmbx, "6000001.seq_2.cmd") == sequence_cmd
    assert [injection.name for injection in request.injections] == ["Temperature Accuracy_H"]
    assert [item.name for item in request.methods_and_reports] == [
        "TEMPERATURE_ACCURACY__single_40C",
        "ACCURACY_IRC_STOP_H",
        "Report_VTCC_V2_12__single_40C",
    ]
    request_entries = {entry.name for entry in request.entries}
    assert "CMBX_DATA_EXPLORER_GENERATION/method_script_40C_only.txt" in request_entries
    assert "CMBX_DATA_EXPLORER_GENERATION/report_formula_map_40C.tsv" in request_entries
    assert "SET RetTimes.RetTime3 = System.Retention" in extract_cmbx_entry(
        outputs.request_cmbx,
        "CMBX_DATA_EXPLORER_GENERATION/method_script_40C_only.txt",
    ).decode("utf-8")


def test_safe_filename_removes_path_characters():
    assert safe_filename("Temperature Accuracy_H/ExtTemp:Upper*CC") == "Temperature Accuracy_H_ExtTemp_Upper_CC"


def test_tcc_semantic_catalog_builds_device_specific_blueprints():
    catalog = load_tcc_semantic_catalog()

    va = build_sequence_blueprint(catalog, "VA-C10-A")
    vc = build_sequence_blueprint(catalog, "VC-C10-A")
    vh = build_sequence_blueprint(catalog, "VH-C10-A")

    assert va.report_template == "Report_VATCC_V1_01"
    assert vc.report_template == "Report_VTCC_V2_12"
    assert vh.report_template == "Report_VTCC_V2_12"
    assert va.db_device_source_formula == "AUDIT.ColumnComp.ModelNo"
    assert [item.injection_name for item in va.injections][:3] == ["Valve", "VTCC_BurnIn", "Temperature Calibration"]
    assert [item.injection_name for item in vc.injections][:3] == ["ColumnIDs", "Preheater Connection Test", "Valve"]
    assert [item.injection_name for item in vh.injections][:3] == ["ColumnIDs", "Preheater Connection Test", "Valve"]
    assert any(item.instrument_method == "TEMPERATURE_STABILITY_AND_PCC_70_H" for item in vh.injections)
    assert not any(item.instrument_method == "TEMPERATURE_STABILITY_AND_PCC_70_H" for item in vc.injections)


def test_tcc_semantic_catalog_can_build_partial_requirement_blueprint():
    catalog = load_tcc_semantic_catalog()
    blueprint = build_sequence_blueprint(
        catalog,
        "VH-C10-A",
        ["upper_lower_valve_cycle", "temperature_stability_with_pcc"],
    )

    assert [item.test_intent for item in blueprint.injections] == [
        "upper_lower_valve_cycle",
        "temperature_stability_with_pcc",
    ]
    assert blueprint.required_capability_groups == (
        "core_tcc",
        "external_thermometers",
        "pcc",
        "upper_lower_valves",
    )


def test_single_point_temperature_accuracy_project_preserves_vh_40c_report_anchor(tmp_path):
    project = build_single_point_temperature_accuracy_project("VH-C10-A", 40.0)
    data = single_point_temperature_accuracy_project_to_dict(project)
    output_dir = write_single_point_temperature_accuracy_project(project, tmp_path)

    assert project.db_field == "TempAcc40"
    assert project.ret_time == "RetTime3"
    assert project.source_method == "TEMPERATURE_ACCURACY"
    assert project.source_report_template == "Report_VTCC_V2_12"
    assert project.processing_method == "ACCURACY_IRC_STOP_H"
    assert data["method_contract"]["evaluation_ret_time"] == "RetTime3"
    assert data["report_contract"]["db_field"] == "TempAcc40"
    assert (output_dir / "project_spec.json").exists()
    assert (output_dir / "sequence_template.tsv").exists()
    assert (output_dir / "processing_method_binding.md").exists()
    method_draft = (output_dir / "instrument_method_draft.txt").read_text(encoding="utf-8")
    method_script = (output_dir / "method_script_40C_only.txt").read_text(encoding="utf-8")
    report_spec = (output_dir / "report_calculation_spec.md").read_text(encoding="utf-8")
    formula_map = (output_dir / "report_formula_map_40C.tsv").read_text(encoding="utf-8")
    sequence_template = (output_dir / "sequence_template.tsv").read_text(encoding="utf-8")
    processing_binding = (output_dir / "processing_method_binding.md").read_text(encoding="utf-8")
    assert "Set RetTimes.RetTime3 = System.Retention" in method_draft
    assert "ColumnComp.CC.Temperature.Nominal = 40" in method_draft
    assert "SET RetTimes.RetTime3 = System.Retention" in method_script
    assert "Removed from the source multi-point method" in method_script
    assert "10/20/80/120 degC transitions" in method_script
    assert "AUDIT.RetTime3" in report_spec
    assert "TempAcc40" in report_spec
    assert "TempAcc40\tTemperature Accuracy_H.XLS\tTemp Accuracy\tD68" in formula_map
    assert "ExtTemp_LowerCC" in formula_map
    assert "ExtTemp_UpperCC" in formula_map
    assert "Injection Name\tType\tStatus\tInstrument Method\tProcessing Method" in sequence_template
    assert "Temperature Accuracy_H\tUnknown\tIdle\tTEMPERATURE_ACCURACY__single_40C\tACCURACY_IRC_STOP_H" in sequence_template
    assert "Processing Method Binding - ACCURACY_IRC_STOP_H" in processing_binding
    assert "does not decode or rewrite the processing method payload" in processing_binding


def test_single_point_temperature_accuracy_project_uses_vc_va_accuracy_ladder(tmp_path):
    project = build_single_point_temperature_accuracy_project("VC-C10-A", 85.0)
    data = single_point_temperature_accuracy_project_to_dict(project)
    output_dir = write_single_point_temperature_accuracy_project(project, tmp_path)

    assert project.db_field == "TempAcc85"
    assert project.ret_time == "RetTime5"
    assert project.source_report_template == "Report_VTCC_V2_12"
    assert project.processing_method == "ACCURACY_IRC_STOP_C"
    assert data["method_contract"]["required_report_output"]["report_cell"] == "D70"
    assert data["report_contract"]["db_field"] == "TempAcc85"
    method_script = (output_dir / "method_script_85C.txt").read_text(encoding="utf-8")
    formula_map = (output_dir / "report_formula_map_85C.tsv").read_text(encoding="utf-8")
    assert "10/20/40/60 degC transitions" in method_script
    assert "SET RetTimes.RetTime5 = System.Retention" in method_script
    assert "TempAcc85\tTemperature Accuracy_C.XLS\tTemp Accuracy\tD70" in formula_map

    with pytest.raises(ValueError, match="VC-C10-A temperature accuracy setpoint is not in the known FOQ sequence: 80"):
        build_single_point_temperature_accuracy_project("VC-C10-A", 80.0)


def test_build_execution_sequence_plan_excludes_report_and_db_layers():
    catalog = load_tcc_semantic_catalog()
    blueprint = build_sequence_blueprint(catalog, "VH-C10-A", ["upper_lower_valve_cycle"])
    plan = build_execution_sequence_plan(blueprint)
    data = execution_sequence_plan_to_dict(plan)

    assert data["device_model"] == "VH-C10-A"
    assert data["instrument_methods"] == ["VALVES"]
    assert data["processing_methods"] == ["No_Integration"]
    assert data["rows"] == [
        {
            "row_order": 1,
            "injection_name": "Valve",
            "instrument_method": "VALVES",
            "processing_method": "No_Integration",
            "test_intent": "upper_lower_valve_cycle",
        }
    ]
    assert "report_template" not in data


def test_tcc_semantic_blueprint_capability_validation_uses_manifest_groups():
    catalog = load_tcc_semantic_catalog()
    manifest = load_tcc_symbol_manifest()
    blueprint = build_sequence_blueprint(catalog, "VH-C10-A")

    validation = validate_blueprint_capabilities(blueprint, available_capability_groups(manifest))

    assert validation.passed
    assert validation.missing_groups == ()


def test_semantic_blueprint_serialization_and_report_text():
    catalog = load_tcc_semantic_catalog()
    blueprint = build_sequence_blueprint(catalog, "VA-C10-A", ["upper_lower_valve_cycle"])

    data = blueprint_to_dict(blueprint)
    report = blueprint_report_text(blueprint)

    assert data["device_model"] == "VA-C10-A"
    assert data["injections"][0]["instrument_method"] == "VALVES"
    assert "Device Model: VA-C10-A" in report
    assert "upper_lower_valve_cycle\tValve\tVALVES\tNo_Integration" in report


def test_validate_semantic_blueprint_against_matching_synthetic_package(tmp_path):
    cmbx = tmp_path / "golden_va.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Valve" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="VALVES" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="4" Name="No_Integration" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
    <ChromeleonElement Id="5" Name="Report_VATCC_V1_01" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    sequence_cmd = b"Valve\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x0eNo_Integration\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x06VALVES"
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd)

    catalog = load_tcc_semantic_catalog()
    blueprint = build_sequence_blueprint(catalog, "VA-C10-A", ["upper_lower_valve_cycle"])
    package = load_cmbx_package(cmbx)
    compatibility = validate_blueprint_against_package(blueprint, package)

    assert compatibility.passed
    assert compatibility.missing_injections == ()
    assert "Passed: True" in blueprint_report_text(blueprint, package_compatibility=compatibility)
    plan = build_clone_select_plan(blueprint, package, compatibility)
    plan_data = clone_select_plan_to_dict(plan)

    assert plan_data["source_package"].endswith("golden_va.cmbx")
    assert plan_data["injections"] == ["Valve"]
    assert plan_data["instrument_methods"] == ["VALVES"]
    assert plan_data["processing_methods"] == ["No_Integration"]
    assert plan_data["report_template"] == "Report_VATCC_V1_01"


def test_validate_semantic_blueprint_against_package_reports_missing_parts(tmp_path):
    cmbx = tmp_path / "bad_golden_va.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Valve" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="OTHER_METHOD" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="4" Name="No_Integration" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"Valve")

    catalog = load_tcc_semantic_catalog()
    blueprint = build_sequence_blueprint(catalog, "VA-C10-A", ["upper_lower_valve_cycle"])
    compatibility = validate_blueprint_against_package(blueprint, load_cmbx_package(cmbx))

    assert not compatibility.passed
    assert compatibility.missing_instrument_methods == ("VALVES",)
    assert compatibility.missing_report_templates == ("Report_VATCC_V1_01",)


def test_execution_compatibility_does_not_require_report_template(tmp_path):
    cmbx = tmp_path / "execution_only_va.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Valve" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="VALVES" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="4" Name="No_Integration" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    sequence_cmd = b"Valve\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x0eNo_Integration\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x06VALVES"
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd)

    catalog = load_tcc_semantic_catalog()
    blueprint = build_sequence_blueprint(catalog, "VA-C10-A", ["upper_lower_valve_cycle"])
    package = load_cmbx_package(cmbx)
    execution_plan = build_execution_sequence_plan(blueprint)
    execution_compatibility = validate_execution_plan_against_package(execution_plan, package)
    analysis_compatibility = validate_blueprint_against_package(blueprint, package)

    assert execution_compatibility.passed
    assert not analysis_compatibility.passed
    assert analysis_compatibility.missing_report_templates == ("Report_VATCC_V1_01",)


def test_write_clone_select_cmbx_filters_header_to_plan(tmp_path):
    cmbx = tmp_path / "golden_va_full.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Valve" ItemType="Dionex.Chromeleon.Data.Injection">
      <ChromeleonElement Id="20" Name="ValveSignal" ItemType="Dionex.Chromeleon.Data.Signal" RawDataFilename="20.raw" />
    </ChromeleonElement>
    <ChromeleonElement Id="3" Name="Temperature Accuracy_C" ItemType="Dionex.Chromeleon.Data.Injection">
      <ChromeleonElement Id="30" Name="AccuracySignal" ItemType="Dionex.Chromeleon.Data.Signal" RawDataFilename="30.raw" />
    </ChromeleonElement>
    <ChromeleonElement Id="4" Name="VALVES" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="5" Name="TEMPERATURE_ACCURACY" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="6" Name="No_Integration" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
    <ChromeleonElement Id="7" Name="ACCURACY_IRC_STOP_C" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
    <ChromeleonElement Id="8" Name="Report_VATCC_V1_01" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    sequence_cmd = (
        b"Valve\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x0eNo_Integration\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x06VALVES"
        b" stale Temperature Accuracy_C TEMPERATURE_ACCURACY ACCURACY_IRC_STOP_C"
    )
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd)
        archive.writestr("20.raw", b"Valve raw")
        archive.writestr("30.raw", b"Accuracy raw")

    catalog = load_tcc_semantic_catalog()
    blueprint = build_sequence_blueprint(catalog, "VA-C10-A", ["upper_lower_valve_cycle"])
    package = load_cmbx_package(cmbx)
    compatibility = validate_blueprint_against_package(blueprint, package)
    plan = build_clone_select_plan(blueprint, package, compatibility)
    output_path = write_clone_select_cmbx(package, plan, default_clone_select_output_path(plan, tmp_path / "out"))
    generated = load_cmbx_package(output_path)

    assert output_path.exists()
    assert [injection.name for injection in generated.injections] == ["Valve"]
    assert [item.name for item in generated.methods_and_reports] == ["VALVES", "No_Integration", "Report_VATCC_V1_01"]
    assert extract_cmbx_entry(output_path, "Seq1.cmd") == sequence_cmd
    assert extract_cmbx_entry(output_path, "20.raw") == b"Valve raw"
    with pytest.raises(KeyError):
        extract_cmbx_entry(output_path, "30.raw")
    validation = validate_clone_select_candidate(generated, plan, package)
    validation_text = clone_select_candidate_validation_text(validation)

    assert validation.header_passed
    assert not validation.command_payload_passed
    assert validation.stale_sequence_cmd_references == (
        "Temperature Accuracy_C",
        "TEMPERATURE_ACCURACY",
        "ACCURACY_IRC_STOP_C",
    )
    assert "Header Passed: True" in validation_text


def test_sequence_cmd_probe_finds_readable_name_offsets(tmp_path):
    cmbx = tmp_path / "probe.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    payload = b"prefix Injection A middle " + "METHOD_A".encode("utf-16le") + b" suffix"
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", payload)

    hits = sequence_cmd_name_hits(load_cmbx_package(cmbx))
    table = sequence_cmd_name_hits_tsv(hits)
    clusters = sequence_cmd_hit_clusters(hits)
    cluster_table = sequence_cmd_hit_clusters_tsv(clusters)
    link_table = sequence_cmd_injection_links_tsv(load_cmbx_package(cmbx))
    order_rows = sequence_order_comparison(load_cmbx_package(cmbx))
    order_table = sequence_order_comparison_tsv(order_rows)

    assert [hit.name for hit in hits] == ["Injection A", "METHOD_A"]
    assert hits[0].encoding == "utf-8"
    assert hits[1].encoding == "utf-16le"
    assert "Offset\tKind\tName\tEncoding" in table
    assert len(clusters) == 1
    assert "StartOffset\tEndOffset\tSpan\tHitCount\tNames" in cluster_table
    assert "Occurrence\tInjection\tProcessingMethod\tInstrumentMethod" in link_table
    assert order_rows[0].row_order == 1
    assert order_rows[0].injection_name == "Injection A"
    assert "RowOrder\tCmdOccurrence\tInjection\tProcessingMethod\tInstrumentMethod" in order_table


def test_sequence_cmd_injection_record_probe_extracts_row_fields(tmp_path):
    cmbx = tmp_path / "row_probe.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="4" Name="PROC_A" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    sequence_cmd = (
        b"\x72\x04None\x7a\x0a\x1a\x08Unlocked"
        b"\xe2\x01\x0bInjection A"
        b"\xc2\x02\x1b\x08\x00\x12\x17\x12\x0d\x1a\x0bRelativeUrl*\x06PROC_A"
        b"\xc2\x02\x1d\x08\x01\x12\x19\x12\x0d\x1a\x0bRelativeUrl*\x08METHOD_A"
        b"\xd2\x02\x4f\x0a\x1dSamples.CustomFieldCollection"
        b"\x12\x22\x0a\x0dcm6_sample_id\x12\x0c\x1a\x0aTypeString\x22\x03ABC"
        b"\x1a\x09\x1a\x07Unknown\x22\x0a\x1a\x08Finished"
    )
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd)

    rows = sequence_cmd_injection_record_probes(load_cmbx_package(cmbx))
    table = sequence_cmd_injection_record_probes_tsv(rows)

    assert rows[0].row_order == 1
    assert rows[0].cmd_occurrence_rank == 1
    assert rows[0].name_field_offset == sequence_cmd.find(b"\xe2\x01\x0bInjection A")
    assert rows[0].processing_method == "PROC_A"
    assert rows[0].instrument_method == "METHOD_A"
    assert rows[0].sample_id == "ABC"
    assert rows[0].injection_type == "Unknown"
    assert rows[0].sequence_status == "Finished"
    assert rows[0].lock_state == "Unlocked"
    assert rows[0].ext_temp_preview == "None"
    assert "SampleID" in table


def test_chromeleon_runtime_can_resolve_env_dependency_folder(tmp_path, monkeypatch):
    runtime = tmp_path / "runtime"
    runtime.mkdir()
    (runtime / "Dionex.DataCommon.dll").write_bytes(b"placeholder")
    monkeypatch.setenv("CMBX_CHROMELEON_BIN", str(runtime))

    assert chromeleon_bin() == runtime


def test_export_header_only_method_as_metadata(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence">
    <ChromeleonElement Id="5" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" Url="chrom://example/METHOD_A.instmeth" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)

    package = load_cmbx_package(cmbx)
    method = package.methods_and_reports[0]
    exported = export_element(package, method, tmp_path / "out")

    assert exported.name == "METHOD_A_summary.txt"
    assert "InstrumentMethod" in exported.read_text(encoding="utf-8")


def test_export_instrument_method_prefers_cmbx_embedded_block(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="5" Name="TEMPERATURE_ACCURACY" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" Url="chrom://example/TEMPERATURE_ACCURACY.instmeth" />
    <ChromeleonElement Id="6" Name="NEXT_METHOD" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" Url="chrom://example/NEXT_METHOD.instmeth" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    def varint(value: int) -> bytes:
        out = bytearray()
        while True:
            byte = value & 0x7F
            value >>= 7
            if value:
                out.append(byte | 0x80)
            else:
                out.append(byte)
                return bytes(out)

    def field(number: int, value: bytes) -> bytes:
        return varint((number << 3) | 2) + varint(len(value)) + value

    def method_record(name: bytes, payload: bytes) -> bytes:
        method_payload = field(1, b"TCC-VH") + field(2, b"tf-012345678901") + field(3, payload)
        return (
            field(25, b"GUID-ONE")
            + field(27, b"\x0a\x0a7.1.1.1034")
            + field(28, name)
            + field(35, b"Test")
            + field(19, field(11, method_payload))
        )

    method_block = method_record(b"TEMPERATURE_ACCURACY", b"CpXmPAYLOAD")
    next_block = method_record(b"NEXT_METHOD", b"CpXmNEXT")
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"prefix" + method_block + next_block + b"Copied TEMPERATURE_ACCURACY.instmeth audit log")
    (tmp_path / "Accuracy Method.txt").write_text(
        "\n".join(
            [
                "{Initial Time}\tInstrument Setup\t\t",
                "\tVariables.GenericDouble1\t10.0\tFirst test temperature",
                "\tRetTimes.RetTime1\tSystem.Retention\t",
            ]
        ),
        encoding="utf-8",
    )

    package = load_cmbx_package(cmbx)
    method = package.methods_and_reports[0]
    exported = export_element(package, method, tmp_path / "out")
    payload = exported.with_name("TEMPERATURE_ACCURACY_embedded_payload.bin")
    cpxm = exported.with_name("TEMPERATURE_ACCURACY_embedded_payload.cpxm.bin")
    metadata = exported.with_name("TEMPERATURE_ACCURACY_embedded_metadata.txt")
    summary = exported.with_name("TEMPERATURE_ACCURACY_summary.txt")

    assert exported.name == "TEMPERATURE_ACCURACY_embedded.instmeth.bin"
    assert exported.read_bytes().startswith(b"\xca\x01\x08GUID-ONE")
    assert b"NEXT_METHOD" not in exported.read_bytes()
    assert payload.exists()
    assert cpxm.exists()
    assert cpxm.read_bytes() == b"CpXmPAYLOAD"
    assert metadata.exists()
    assert "CpXm Payload Size" in metadata.read_text(encoding="utf-8")
    assert "embedded Chromeleon instrument method block" in metadata.read_text(encoding="utf-8")
    assert not summary.exists()


def test_export_instrument_method_uses_own_sequence_payload(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="5" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
  </ChromeleonElement>
  <ChromeleonElement Id="10" Name="Seq2" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq2.cmd">
    <ChromeleonElement Id="15" Name="METHOD_B" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""

    def varint(value: int) -> bytes:
        out = bytearray()
        while True:
            byte = value & 0x7F
            value >>= 7
            if value:
                out.append(byte | 0x80)
            else:
                out.append(byte)
                return bytes(out)

    def field(number: int, value: bytes) -> bytes:
        return varint((number << 3) | 2) + varint(len(value)) + value

    def method_record(name: bytes, payload: bytes) -> bytes:
        method_payload = field(1, b"TCC") + field(3, payload)
        return field(25, b"GUID") + field(28, name) + field(19, field(11, method_payload))

    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"prefix" + method_record(b"METHOD_A", b"CpXmA"))
        archive.writestr("Seq2.cmd", b"prefix" + method_record(b"METHOD_B", b"CpXmB"))

    package = load_cmbx_package(cmbx)
    method_b = next(method for method in package.methods_and_reports if method.name == "METHOD_B")
    exported = export_element(package, method_b, tmp_path / "out")
    cpxm = exported.with_name("METHOD_B_embedded_payload.cpxm.bin")
    summary = exported.with_name("METHOD_B_summary.txt")

    assert exported.name == "METHOD_B_embedded.instmeth.bin"
    assert cpxm.read_bytes() == b"CpXmB"
    assert not summary.exists()


def test_export_instrument_method_writes_decoded_xml_when_available(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="5" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""

    def varint(value: int) -> bytes:
        out = bytearray()
        while True:
            byte = value & 0x7F
            value >>= 7
            if value:
                out.append(byte | 0x80)
            else:
                out.append(byte)
                return bytes(out)

    def field(number: int, value: bytes) -> bytes:
        return varint((number << 3) | 2) + varint(len(value)) + value

    method_payload = field(1, b"TCC-VH") + field(3, b"CpXmPAYLOAD")
    method_record = field(25, b"GUID-ONE") + field(28, b"METHOD_A") + field(19, field(11, method_payload))
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"prefix" + method_record)

    def fake_decode(cpxm_path: Path, xml_path: Path):
        xml_path.write_text(
            '<CmData><Item type="PropertyStepNode"><SymbolPath value="ColumnComp.CC.TempCtrl" /><Value value="On" /></Item></CmData>',
            encoding="utf-8",
        )
        return export_service.MethodDecodeResult(True, "fake decode ok")

    monkeypatch.setattr(export_service, "decode_cpxm_method_xml", fake_decode)

    package = load_cmbx_package(cmbx)
    exported = export_service.export_element_paths(package, package.methods_and_reports[0], tmp_path / "out")
    names = [path.name for path in exported]
    metadata = next(path for path in exported if path.name.endswith("_embedded_metadata.txt"))
    flow_tsv = next(path for path in exported if path.name.endswith("_embedded_method_flow.tsv"))

    assert "METHOD_A_embedded_method.xml" in names
    assert "METHOD_A_embedded_method_flow.txt" in names
    assert "METHOD_A_embedded_method_flow.tsv" in names
    assert "Method\tOrder\tLevel\tStage\tTime\tNodeType\tAction\tTarget\tValue\tComment\tCondition" in flow_tsv.read_text(encoding="utf-8")
    assert "fake decode ok" in metadata.read_text(encoding="utf-8")


def test_build_method_flow_from_decoded_xml():
    xml = """<CmData>
  <Item type="CommmentNode"><Comment value="Accuracy setup" /></Item>
  <Item type="PropertyStepNode"><SymbolPath value="Variables.GenericDouble1" /><Value value="10.0" /></Item>
  <Item type="CommandStepNode"><SymbolPath value="Wait" /><Value value="ColumnComp.CC.TempReady" /></Item>
</CmData>"""

    flow = build_method_flow_from_xml(xml, "METHOD_A")
    table = build_method_flow_tsv(xml, "METHOD_A")

    assert "# Accuracy setup" in flow
    assert "SET Variables.GenericDouble1 = 10.0" in flow
    assert "RUN Wait ColumnComp.CC.TempReady" in flow
    assert "METHOD_A\t2\t0\t\t\tPropertyStepNode\tSET\tVariables.GenericDouble1\t10.0" in table
    assert "METHOD_A\t3\t0\t\t\tCommandStepNode\tRUN\tWait\tColumnComp.CC.TempReady" in table


def test_build_method_flow_preserves_cm_if_table_semantics():
    xml = """<CmData>
  <Item type="StageNode">
    <Children type="SyntaxNodeCollection">
      <Item type="TimeStepNode">
        <Time type="MethodTime"><InternalValue value="-3.00000000000000000E+001" /></Time>
        <Children type="SyntaxNodeCollection">
          <Item type="CommmentNode"><Comment value="Header comment" /></Item>
          <Item type="IfBlockNode">
            <Children type="SyntaxNodeCollection">
              <Item type="IfNode">
                <Children type="SyntaxNodeCollection">
                  <Item type="PropertyStepNode"><SymbolPath value="Variables.GenericLong9" /><Value value="12" /></Item>
                  <Item type="CommandStepNode"><Comment value="OK" /><SymbolPath value="Delay" /><Value value="1" /></Item>
                </Children>
                <Condition value="ColumnComp.ModelNo=&quot;VH-C10-A&quot;" />
              </Item>
              <Item type="ElseIfNode">
                <Children type="SyntaxNodeCollection">
                  <Item type="PropertyStepNode"><SymbolPath value="Variables.GenericLong9" /><Value value="10" /></Item>
                </Children>
                <Condition value="ColumnComp.ModelNo=&quot;VC-C10-A&quot;" />
              </Item>
              <Item type="ElseNode">
                <Children type="SyntaxNodeCollection">
                  <Item type="CommandStepNode"><SymbolPath value="System.AbortQueue" /><Value value="" /></Item>
                </Children>
              </Item>
            </Children>
          </Item>
        </Children>
      </Item>
    </Children>
    <StageName value="Equilibration" />
  </Item>
</CmData>"""

    table = build_method_flow_tsv(xml, "METHOD_A")

    assert "METHOD_A\t1\t0\tEquilibration\t-30.000\tStageNode\tSTAGE" in table
    assert "IfNode\tIF\t\t\t\tColumnComp.ModelNo=\"VH-C10-A\"" in table
    assert "Variables.GenericLong9\t12" in table
    assert "Delay\t1\tOK" in table
    assert "ElseIfNode\tELSE IF\t\t\t\tColumnComp.ModelNo=\"VC-C10-A\"" in table
    assert "IfBlockNode\tEND IF" in table
    assert "EndNode\tEND" in table
    assert "ColumnComp.ModelNo=\"VH-C10-A\" Variables.GenericLong9" not in table

    rows, error = build_method_flow_rows(xml)
    assert error is None
    from app import CmbxExplorerApp

    app = object.__new__(CmbxExplorerApp)
    cm_rows = app._cm_method_rows_from_flow_rows(rows)
    assert cm_rows[0][1] == "Stage"
    assert any(row[1] == "Comment" and "Equilibration" not in row[3] for row in cm_rows)
    assert any(row[1] == "Branch" and row[2].strip() == "If" for row in cm_rows)
    assert any(row[1] == "Command" and "Variables.GenericLong9" in row[3] for row in cm_rows)
    assert cm_rows[-1][1] == "End"


def test_build_method_flow_stage_time_accepts_nested_time_steps():
    xml = """<CmData>
  <Item type="StageNode">
    <Children type="SyntaxNodeCollection">
      <Wrapper>
        <Item type="TimeStepNode">
          <Time type="MethodTime"><InternalValue value="-1.00000000000000000E+001" /></Time>
        </Item>
        <Item type="TimeStepNode">
          <Time type="MethodTime"><InternalValue value="5.00000000000000000E+000" /></Time>
        </Item>
      </Wrapper>
    </Children>
    <StageName value="Equilibration" />
  </Item>
</CmData>"""

    table = build_method_flow_tsv(xml, "METHOD_A")

    assert "METHOD_A\t1\t0\tEquilibration\t-10.000\tStageNode\tSTAGE\t\tEquilibration\tDuration = 15.000 [min]" in table


def test_export_instrument_method_uses_payload_from_previous_record_when_linked(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="5" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""

    def varint(value: int) -> bytes:
        out = bytearray()
        while True:
            byte = value & 0x7F
            value >>= 7
            if value:
                out.append(byte | 0x80)
            else:
                out.append(byte)
                return bytes(out)

    def field(number: int, value: bytes) -> bytes:
        return varint((number << 3) | 2) + varint(len(value)) + value

    previous_payload = field(1, b"TCC") + field(3, b"CpXmMETHOD_A")
    previous_record = field(25, b"GUID-PREV") + field(28, b"PREVIOUS") + field(19, field(11, previous_payload))
    method_header = field(25, b"GUID-A") + field(28, b"METHOD_A") + field(35, b"Target")
    wrong_payload = field(1, b"TCC") + field(3, b"CpXmNEXT")
    trailing_record = field(25, b"GUID-AFTER") + field(28, b"AFTER") + field(19, field(11, wrong_payload))
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"prefix" + previous_record + method_header + trailing_record)

    package = load_cmbx_package(cmbx)
    exported = export_element(package, package.methods_and_reports[0], tmp_path / "out")
    cpxm = exported.with_name("METHOD_A_embedded_payload.cpxm.bin")

    assert cpxm.read_bytes() == b"CpXmMETHOD_A"


def test_parse_report_sheets_matches_selected_injection():
    xml = """<CmData><ReportDefinition><SpreadsheetDefinition><CmData><WorkbookData>
<SheetList>
  <Sheet><Name value="Definitions" /><Data><Id value="def-id" /></Data></Sheet>
  <Sheet><Name value="Temp Accuracy" /><Data><Id value="acc-id" /></Data></Sheet>
</SheetList>
</WorkbookData></CmData></SpreadsheetDefinition>
<PrintSettings><PrintSheetSetups><Setups>
  <Item><Name value="Definitions" /><Data type="SheetSetup"><IsActive value="N" /><SheetSetupCondition><EachInjection value="Y" /><InjectionQueryCondition><Enabled value="N" /></InjectionQueryCondition></SheetSetupCondition></Data></Item>
  <Item><Name value="Temp Accuracy" /><Data type="SheetSetup"><IsActive value="Y" /><SheetSetupCondition><EachInjection value="N" /><InjectionQueryCondition><Enabled value="Y" /><InjectionQuery><QueryRules><Item><Variable value="injname" /><Comparison value="Contains" /><Values><ValueItem value="Temperature Accuracy" /></Values></Item></QueryRules></InjectionQuery></InjectionQueryCondition></SheetSetupCondition></Data></Item>
</Setups></PrintSheetSetups></PrintSettings></ReportDefinition></CmData>"""

    sheets = parse_report_sheets(xml, "REPORT_A", "Temperature Accuracy_H")
    by_name = {sheet.sheet_name: sheet for sheet in sheets}
    table = report_sheets_tsv(sheets)

    assert by_name["Definitions"].applies_to_injection == "No"
    assert by_name["Temp Accuracy"].applies_to_injection == "Yes"
    assert "REPORT_A\tTemp Accuracy\tacc-id\tY\tN\tY\tinjname\tContains\tTemperature Accuracy\tYes" in table


def test_parse_report_sheet_objects_extracts_formula_positions():
    xml = """<CmData><ReportDefinition>
<SheetDescription type="SheetDescription">
  <Id value="{SHEET-1}" />
  <SheetName value="Temp Accuracy" />
  <SheetObject type="ReportFormulaObject">
    <Id value="CellObject1" />
    <Range type="Range`1&lt;Double&gt;"><Left value="11" /><Top value="65" /><Right value="11" /><Bottom value="65" /></Range>
    <Formula value="chm.sig_value(&quot;average&quot;)" />
    <FixedChannel value="ExtTemp_LowerCC" />
  </SheetObject>
  <SheetObject type="PlotObject">
    <Id value="Chromeleon 1" />
    <Range type="Range`1&lt;Double&gt;"><Left value="0" /><Top value="0" /><Right value="5" /><Bottom value="10" /></Range>
    <PlotType value="Chromatogram" />
  </SheetObject>
</SheetDescription>
</ReportDefinition></CmData>"""

    objects = parse_report_sheet_objects(xml, "REPORT_A")
    table = report_sheet_objects_tsv(objects)

    assert len(objects) == 2
    assert objects[0].sheet_name == "Temp Accuracy"
    assert objects[0].excel_range == "L66"
    assert objects[0].fixed_channel == "ExtTemp_LowerCC"
    assert objects[1].plot_type == "Chromatogram"
    assert "REPORT_A\tTemp Accuracy\tCellObject1\tReportFormulaObject\t11\t65\t11\t65\tL66" in table


def test_report_formula_evaluator_reads_audit_and_signal_tsv(tmp_path):
    audit = tmp_path / "audit.tsv"
    audit.write_text(
        "\n".join(
            [
                "Index\tRetentionTime(min)\tDevice\tMessage\tTriggerName\tPropertyName\tPropertyValue",
                "0\t-0.58\tColumnComp.CC\t\t\tTemperature.Nominal\t10.00",
                "1\t8.97\tRetTimes\t\t\tRetTime1\t8.970",
                "2\t8.97\tColumnComp.CC\t\t\tTemperature.Nominal\t20.00",
                "3\t22.39\tRetTimes\t\t\tRetTime2\t22.387",
            ]
        ),
        encoding="utf-8",
    )
    signal = tmp_path / "upper.tsv"
    signal.write_text(
        "\n".join(
            [
                "Channel\tExtTemp_UpperCC",
                "Time (min)\tStep (s)\tValue",
                "7.90\t0\t9.9",
                "8.00\t6\t10.0",
                "8.50\t30\t10.2",
                "8.77\t16.2\t10.4",
                "8.80\t1.8\t11.0",
            ]
        ),
        encoding="utf-8",
    )

    ret_times = read_audit_ret_times_tsv(audit)
    records = read_audit_records_tsv(audit)
    points = read_signal_tsv(signal)
    value, status, detail = evaluate_chm_sig_value_average(
        'chm.sig_value("average",AUDIT.RetTime1(1,"forward")-1,AUDIT.RetTime1(1,"forward")-0.2)',
        "ExtTemp_UpperCC",
        ret_times,
        points,
    )

    assert ret_times == {1: 8.97, 2: 22.387}
    assert records[0].device == "ColumnComp.CC"
    assert records[0].property_value == "10.00"
    nominal = evaluate_audit_property_formula(
        'AUDIT.ColumnComp.CC.Temperature.Nominal(AUDIT.RetTime1(1,"forward")-0.1)',
        ret_times,
        records,
    )
    assert nominal is not None
    assert nominal[0] == "10.00"
    assert points[0] == SignalPoint(7.90, 9.9)
    assert status == "ok"
    assert value == "10.2"
    assert "n=3" in detail
    table = formula_evaluations_tsv(
        [
            FormulaEvaluation(
                report_name="REPORT_A",
                injection_name="Injection A",
                sheet_name="Temp Accuracy",
                excel_range="M66",
                object_type="ReportFormulaObject",
                fixed_channel="ExtTemp_UpperCC",
                formula='chm.sig_value("average")',
                value=value,
                status=status,
                detail=detail,
            )
        ]
    )
    assert "Report\tInjection\tSheet\tExcelRange\tObjectType\tFixedChannel\tFormula\tValue\tStatus\tDetail" in table
    assert "REPORT_A\tInjection A\tTemp Accuracy\tM66" in table


def test_report_formula_evaluator_reads_no_argument_audit_metadata():
    records = [
        AuditRecord(None, "ColumnComp", "ModelNo", "VH-C10-A"),
        AuditRecord(None, "ColumnComp", "SerialNo", "6545251"),
    ]

    model = evaluate_audit_metadata_formula("AUDIT.ColumnComp.ModelNo", records)
    serial = evaluate_audit_metadata_formula("precond.ColumnComp.SerialNo", records)

    assert model == ("VH-C10-A", "ColumnComp.ModelNo from audit precondition.")
    assert serial == ("6545251", "ColumnComp.SerialNo from audit precondition.")


def test_report_formula_evaluator_matches_audit_property_suffix_paths():
    records = [
        AuditRecord(0.05, "ColumnComp.Column_A", "Description", "A"),
    ]

    value = evaluate_audit_property_formula('AUDIT.Column_A.Description(0,"forward")', {}, records)

    assert value == ("A", "Column_A.Description at 0 min from audit record 0.05 min.")


def test_report_formula_context_falls_back_to_previous_injection_audit(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader DateCreated="Friday, July 3, 2026">
  <ChromeleonElement Id="1" Name="Seq1" Url="chrom://server/ChromeleonLocal/Project/Seq1.seq" ItemType="Dionex.Chromeleon.Data.Sequence">
    <ChromeleonElement Id="2" Name="With Audit" ItemType="Dionex.Chromeleon.Data.Injection">
      <ChromeleonElement Id="3" Name="Audit" ItemType="Dionex.Chromeleon.Data.AuditTrail" RawDataFilename="3.raw" />
    </ChromeleonElement>
    <ChromeleonElement Id="4" Name="Factory Default" ItemType="Dionex.Chromeleon.Data.Injection" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("3.raw", b"Audit")

    def fake_export_audit_raw(_raw_path, output_path):
        output_path.write_text(
            "\n".join(
                [
                    "Index\tRetentionTime(min)\tDevice\tMessage\tTriggerName\tPropertyName\tPropertyValue",
                    "0\t\tColumnComp\t\t\tSerialNo\t6545251",
                ]
            ),
            encoding="utf-8",
        )
        return output_path

    monkeypatch.setattr("report_formula_evaluator.export_audit_raw", fake_export_audit_raw)

    package = load_cmbx_package(cmbx)
    injection = next(row for row in package.injections if row.name == "Factory Default")
    context = build_report_formula_context(package, injection)

    assert context.audit_source_injection == "With Audit"
    assert evaluate_audit_metadata_formula("precond.ColumnComp.SerialNo", context.audit_records)[0] == "6545251"


def test_report_formula_evaluator_supports_more_raw_signal_formulas():
    signal = [
        SignalPoint(0.0, 10.0),
        SignalPoint(0.5, 12.0),
        SignalPoint(1.0, 14.0),
        SignalPoint(1.5, 11.0),
        SignalPoint(2.0, 13.0),
    ]
    ret_times = {6: 1.5, 7: 2.0}

    average = evaluate_chm_signal_formula('chm.sig_value("average")', "CC_Temp", ret_times, signal)
    minimum = evaluate_chm_signal_formula('chm.signalStatistic("min",0.5,1.5)', "CC_Temp", ret_times, signal)
    maximum = evaluate_chm_signal_formula('chm.sig_value("max",0,2)', "CC_Temp", ret_times, signal)
    value = evaluate_chm_signal_value('chm.signalValue(AUDIT.RetTime6(1,"forward")-0.25)', "CC_Temp", ret_times, signal)
    noise = evaluate_chm_noise("chm.noise(0.5,1.5)", "CC_Temp", ret_times, signal)
    drift = evaluate_chm_drift('chm.drift(AUDIT.RetTime6(1,"forward")-0.5,AUDIT.RetTime6(1,"forward"))', "CC_Temp", ret_times, signal)

    assert average[:2] == ("12", "ok")
    assert minimum[:2] == ("11", "ok")
    assert maximum[:2] == ("14", "ok")
    assert value[:2] == ("12.5", "ok")
    assert noise[:2] == ("2.5", "ok")
    assert drift[:2] == ("-6", "ok")


def test_report_formula_evaluator_supports_audit_forward_backward_and_ret_time6():
    records = [
        # audit_ret_times should keep RetTime values beyond the Temperature Accuracy 1..5 case.
        # The forward/backward property formulas should choose the nearest matching audit record
        # in the requested direction.
        AuditRecord(1.0, "RetTimes", "RetTime6", "1.50"),
        AuditRecord(0.5, "UpperValve", "CurrentPosition", "A"),
        AuditRecord(1.2, "UpperValve", "CurrentPosition", "B"),
    ]
    ret_times = audit_ret_times(records)

    backward = evaluate_audit_property_formula('AUDIT.UpperValve.CurrentPosition(1.0,"backward")', ret_times, records)
    forward = evaluate_audit_property_formula('AUDIT.UpperValve.CurrentPosition(1.0,"forward")', ret_times, records)
    spaced = evaluate_audit_property_formula("AUDIT.UpperValve.CurrentPosition (1.0)", ret_times, records)

    assert ret_times == {6: 1.5}
    assert backward is not None and backward[0] == "A"
    assert forward is not None and forward[0] == "B"
    assert spaced is not None and spaced[0] == "A"


def test_export_report_template_writes_blank_xls_only(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="9" Name="REPORT_A" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""

    def varint(value: int) -> bytes:
        out = bytearray()
        while True:
            byte = value & 0x7F
            value >>= 7
            if value:
                out.append(byte | 0x80)
            else:
                out.append(byte)
                return bytes(out)

    def field(number: int, value: bytes) -> bytes:
        return varint((number << 3) | 2) + varint(len(value)) + value

    report_payload = field(15, field(1, b"CpXmREPORT"))
    report_record = field(25, b"GUID-REPORT") + field(28, b"REPORT_A") + field(19, report_payload)
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"prefix" + report_record)

    def fake_decode_report(_package, _report):
        return None, '<CmData><ReportDefinition><SpreadsheetDefinition><CmData><WorkbookData /></CmData></SpreadsheetDefinition></ReportDefinition></CmData>'

    def fake_formulaone_export(output_path, _report_xml, sheet_names=None, cell_values=None):
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"fake xls")
        return path

    monkeypatch.setattr(export_service, "decode_report_template_xml", fake_decode_report)
    monkeypatch.setattr(export_service, "export_formulaone_report_template", fake_formulaone_export)

    package = load_cmbx_package(cmbx)
    exported = export_service.export_element_paths(package, package.methods_and_reports[0], tmp_path / "out")
    names = [path.name for path in exported]

    assert names == ["REPORT_A_report_template.xls"]


def test_export_report_formula_values_writes_injection_scoped_tsv(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="9" Name="REPORT_A" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"Sequence")

    def fake_decode(_package, _report):
        return None, "<CmData />"

    def fake_evaluate(_package, injection, report_name, _xml_text, sheet_name="", context=None):
        return [
            FormulaEvaluation(
                report_name=report_name,
                injection_name=injection.name,
                sheet_name=sheet_name or "Temp Accuracy",
                excel_range="M66",
                object_type="ReportFormulaObject",
                fixed_channel="ExtTemp_UpperCC",
                formula='chm.sig_value("average")',
                value="10.02",
                status="ok",
                detail="fake",
            )
        ]

    captured = {}

    def fake_formulaone_export(output_path, _report_xml, sheet_names=None, cell_values=None):
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"fake xls")
        captured["sheet_names"] = sheet_names
        captured["cell_values"] = cell_values
        return path

    monkeypatch.setattr(export_service, "decode_report_template_xml", fake_decode)
    monkeypatch.setattr(export_service, "evaluate_report_formulas", fake_evaluate)
    monkeypatch.setattr(export_service, "export_formulaone_report_template", fake_formulaone_export)

    package = load_cmbx_package(cmbx)
    injection = next(element for element in package.injections if element.name == "Injection A")
    report = next(element for element in package.methods_and_reports if element.name == "REPORT_A")
    exported = export_service.export_report_formula_values(package, injection, report, tmp_path / "out", "Temp Accuracy")
    text = exported.read_text(encoding="utf-8")

    assert exported.name == "REPORT_A_Temp Accuracy_formula_values.tsv"
    assert exported.parent.name == "Injection A"
    assert "REPORT_A\tInjection A\tTemp Accuracy\tM66" in text
    assert "\t10.02\tok\tfake" in text


def test_export_report_workbook_writes_xlsx(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Temperature Accuracy_H" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="9" Name="REPORT_A" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"Sequence")

    report_xml = """<CmData><ReportDefinition><SpreadsheetDefinition><CmData><WorkbookData>
<SheetList>
  <Sheet><Name value="Definitions" /><Data><Id value="def-id" /></Data></Sheet>
  <Sheet><Name value="Temp Accuracy" /><Data><Id value="acc-id" /></Data></Sheet>
</SheetList></WorkbookData></CmData></SpreadsheetDefinition>
<PrintSettings><PrintSheetSetups><Setups>
  <Item><Name value="Temp Accuracy" /><Data><IsActive value="Y" /><SheetSetupCondition><EachInjection value="N" /><InjectionQueryCondition><Enabled value="Y" /><InjectionQuery><QueryRules><Item><Variable value="injname" /><Comparison value="Contains" /><Values><ValueItem value="Temperature Accuracy" /></Values></Item></QueryRules></InjectionQuery></InjectionQueryCondition></SheetSetupCondition></Data></Item>
</Setups></PrintSheetSetups></PrintSettings></ReportDefinition></CmData>"""

    def fake_decode(_package, _report):
        return None, report_xml

    def fake_evaluate(_package, injection, report_name, _xml_text, sheet_name="", context=None):
        if sheet_name != "Temp Accuracy":
            return []
        return [
            FormulaEvaluation(report_name, injection.name, sheet_name, "I66", "ReportFormulaObject", "", "", "10.00", "ok", ""),
            FormulaEvaluation(report_name, injection.name, sheet_name, "K66", "ReportFormulaObject", "", "", "8.970", "ok", ""),
            FormulaEvaluation(report_name, injection.name, sheet_name, "L66", "ReportFormulaObject", "ExtTemp_LowerCC", "", "10.02", "ok", ""),
            FormulaEvaluation(report_name, injection.name, sheet_name, "M66", "ReportFormulaObject", "ExtTemp_UpperCC", "", "10.02", "ok", ""),
        ]

    monkeypatch.setattr(export_service, "decode_report_template_xml", fake_decode)
    monkeypatch.setattr(export_service, "evaluate_report_formulas", fake_evaluate)

    package = load_cmbx_package(cmbx)
    injection = package.injections[0]
    report = package.methods_and_reports[0]
    exported = export_service.export_report_workbook(package, injection, report, tmp_path / "out")

    from openpyxl import load_workbook

    workbook = load_workbook(exported, data_only=True)
    assert exported.suffix == ".xlsx"
    assert "Definitions" in workbook.sheetnames
    assert "Temp Accuracy" in workbook.sheetnames
    assert workbook["Temp Accuracy"]["C66"].value == 10.02


def test_export_filled_report_template_writes_values_to_template_cells(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Temperature Accuracy_H" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="9" Name="REPORT_A" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"Sequence")

    report_xml = """<CmData><ReportDefinition><SpreadsheetDefinition><CmData><WorkbookData>
<SheetList>
  <Sheet><Name value="Temp Accuracy" /><Data><Id value="acc-id" /></Data></Sheet>
</SheetList></WorkbookData></CmData></SpreadsheetDefinition>
<PrintSettings><PrintSheetSetups><Setups>
  <Item><Name value="Temp Accuracy" /><Data><IsActive value="Y" /><SheetSetupCondition><EachInjection value="N" /><InjectionQueryCondition><Enabled value="Y" /><InjectionQuery><QueryRules><Item><Variable value="injname" /><Comparison value="Contains" /><Values><ValueItem value="Temperature Accuracy" /></Values></Item></QueryRules></InjectionQuery></InjectionQueryCondition></SheetSetupCondition></Data></Item>
</Setups></PrintSheetSetups></PrintSettings>
<SheetDescription type="SheetDescription">
  <SheetName value="Temp Accuracy" />
  <SheetObject type="ReportFormulaObject">
    <Id value="CellObject1" />
    <Range><Left value="12" /><Top value="65" /><Right value="12" /><Bottom value="65" /></Range>
    <Formula value="chm.sig_value(&quot;average&quot;)" />
    <FixedChannel value="ExtTemp_UpperCC" />
  </SheetObject>
</SheetDescription>
</ReportDefinition></CmData>"""

    def fake_decode(_package, _report):
        return None, report_xml

    def fake_evaluate(_package, injection, report_name, _xml_text, sheet_name="", context=None):
        return [
            FormulaEvaluation(report_name, injection.name, sheet_name, "M66", "ReportFormulaObject", "ExtTemp_UpperCC", "upper", "10.02", "ok", "fake")
        ]

    captured = {}

    def fake_formulaone_export(output_path, _report_xml, sheet_names=None, cell_values=None):
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"fake xls")
        captured["sheet_names"] = sheet_names
        captured["cell_values"] = cell_values
        return path

    monkeypatch.setattr(export_service, "decode_report_template_xml", fake_decode)
    monkeypatch.setattr(export_service, "evaluate_report_formulas", fake_evaluate)
    monkeypatch.setattr(export_service, "export_formulaone_report_template", fake_formulaone_export)

    package = load_cmbx_package(cmbx)
    injection = package.injections[0]
    report = package.methods_and_reports[0]
    exported = export_service.export_filled_report_template_workbook(package, injection, report, tmp_path / "out")

    assert exported.name == "Temperature Accuracy_H.xls"
    assert "Temp Accuracy" in captured["sheet_names"]
    written = [
        value
        for value in captured["cell_values"].values()
        if value.sheet_name == "Temp Accuracy" and value.cell == "M66"
    ]
    assert written and written[0].value == 10.02


def test_export_sequence_report_sheets_writes_one_xls_for_sequence(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Temperature Accuracy_H" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="HeatUp and CoolDownTime" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="9" Name="REPORT_A" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"Sequence")

    report_xml = """<CmData><ReportDefinition><SpreadsheetDefinition><CmData><WorkbookData>
<SheetList>
  <Sheet><Name value="Definitions" /><Data><Id value="def-id" /></Data></Sheet>
  <Sheet><Name value="Temp Accuracy" /><Data><Id value="acc-id" /></Data></Sheet>
  <Sheet><Name value="HeatUp&amp;CoolDown" /><Data><Id value="heat-id" /></Data></Sheet>
  <Sheet><Name value="COC" /><Data><Id value="coc-id" /></Data></Sheet>
</SheetList></WorkbookData></CmData></SpreadsheetDefinition></ReportDefinition></CmData>"""

    def fake_decode(_package, _report):
        return None, report_xml

    def fake_evaluate(_package, injection, report_name, _xml_text, sheet_name="", context=None):
        if sheet_name == "Temp Accuracy":
            return [FormulaEvaluation(report_name, injection.name, sheet_name, "M66", "ReportFormulaObject", "ExtTemp_UpperCC", "upper", "10.02", "ok", "fake")]
        if sheet_name == "HeatUp&CoolDown":
            return [FormulaEvaluation(report_name, injection.name, sheet_name, "J66", "ReportFormulaObject", "", "ret", "1.0", "ok", "fake")]
        return []

    captured = {}

    def fake_formulaone_export(output_path, _report_xml, sheet_names=None, cell_values=None):
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"fake xls")
        captured["sheet_names"] = sheet_names
        captured["cell_values"] = cell_values
        return path

    monkeypatch.setattr(export_service, "decode_report_template_xml", fake_decode)
    monkeypatch.setattr(export_service, "evaluate_report_formulas", fake_evaluate)
    monkeypatch.setattr(export_service, "export_formulaone_report_template", fake_formulaone_export)

    package = load_cmbx_package(cmbx)
    sequence = package.sequences[0]
    exported = export_service.export_sequence_report_sheets_workbook(package, sequence, tmp_path / "out")

    assert exported.name == "Seq1_report_sheets.xls"
    assert "Temp Accuracy" in captured["sheet_names"]
    assert "HeatUp&CoolDown" in captured["sheet_names"]
    assert "COC" in captured["sheet_names"]
    assert any(value.cell == "M66" for value in captured["cell_values"].values())


def test_export_all_report_workbooks_writes_one_workbook_per_injection(tmp_path, monkeypatch):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Temperature Accuracy_H" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="3" Name="Temperature Stability_and_PCC_H" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="9" Name="REPORT_A" ItemType="Dionex.Chromeleon.Data.ReportDefinition" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", b"Sequence")

    report_xml = """<CmData><ReportDefinition><SpreadsheetDefinition><CmData><WorkbookData>
<SheetList><Sheet><Name value="Temp Accuracy" /><Data><Id value="acc-id" /></Data></Sheet></SheetList>
</WorkbookData></CmData></SpreadsheetDefinition></ReportDefinition></CmData>"""

    monkeypatch.setattr(export_service, "decode_report_template_xml", lambda _package, _report: (None, report_xml))
    monkeypatch.setattr(export_service, "evaluate_report_formulas", lambda *_args, **_kwargs: [])

    package = load_cmbx_package(cmbx)
    exported = export_service.export_all_report_workbooks(package, tmp_path / "out")

    assert len(exported) == 2
    assert all(path.suffix == ".xlsx" for path in exported)
    assert {path.parent.name for path in exported} == {"Temperature Accuracy_H", "Temperature Stability_and_PCC_H"}


def test_report_workbook_builder_creates_temp_accuracy_sheet(tmp_path):
    evaluations = [
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "I66", "ReportFormulaObject", "", "nominal", "10.00", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "K66", "ReportFormulaObject", "", "ret", "8.970", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "L66", "ReportFormulaObject", "ExtTemp_LowerCC", "lower", "10.02", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "M66", "ReportFormulaObject", "ExtTemp_UpperCC", "upper", "10.01", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "I67", "ReportFormulaObject", "", "nominal", "20.00", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "K67", "ReportFormulaObject", "", "ret", "22.387", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "L67", "ReportFormulaObject", "ExtTemp_LowerCC", "lower", "20.03", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "M67", "ReportFormulaObject", "ExtTemp_UpperCC", "upper", "20.04", "ok", ""),
    ]

    rows = build_accuracy_rows(evaluations)
    output = write_report_workbook(tmp_path / "report.xlsx", "Temperature Accuracy_H", {"Definitions": [], "Temp Accuracy": evaluations})

    from openpyxl import load_workbook

    wb = load_workbook(output, data_only=True)
    ws = wb["Temp Accuracy"]

    assert rows[0].observed_max_deviation_temp_c == 10.02
    assert rows[1].observed_max_deviation_temp_c == 20.04
    assert ws["B9"].value == "Temperature Accuracy"
    assert ws["B66"].value == 10
    assert ws["C66"].value == 10.02
    assert ws["N67"].value == 20.04


def test_report_calculation_map_extracts_definition_criteria():
    import base64

    raw = (
        b"prefix Temperature Accuracy"
        + b"\x7e\x17\x00\x02\x00\xd5\x00\x00\x00\xe0\x3f"
        + b" spacer HeatUp & Cool Down"
        + b"\x7e\x1a\x00\x02\x00\xd6\x00\x00\x00\x2e\x40"
    )
    xml = f"""<CmData><SpreadSheetData value="{base64.b64encode(raw).decode("ascii")}" /></CmData>"""

    criteria = extract_definition_criteria(xml)
    calculation_map = build_report_calculation_map(xml, {"HeatUp&CoolDown": []})

    assert criteria["Temperature Accuracy"].value == 0.5
    assert criteria["HeatUp & Cool Down"].value == 15
    assert any(row.cell == "D65" and row.criterion == "15" for row in calculation_map)
    assert any(row.sheet_name == "Definitions" and row.label == "Temperature Accuracy" for row in calculation_map)


def test_report_workbook_builder_creates_calculation_map_sheet(tmp_path):
    rows = build_report_calculation_map("<CmData />", {"HeatUp&CoolDown": []})
    output = write_report_workbook(tmp_path / "report.xlsx", "HeatUp and CoolDownTime", {"HeatUp&CoolDown": []}, calculation_map=rows)

    from openpyxl import load_workbook

    wb = load_workbook(output, data_only=True)
    ws = wb["Calculation Map"]

    assert ws["A1"].value == "Report Calculation Map"
    assert "D65" in [cell.value for cell in ws["B"]]


def test_report_workbook_builder_creates_heatup_cooldown_sheet(tmp_path):
    evaluations = [
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "J65", "ReportFormulaObject", "", "ret", "6.252", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "K65", "ReportFormulaObject", "", "ret", "12.099", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "L65", "ReportFormulaObject", "", "ret", "12.680", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "M65", "ReportFormulaObject", "", "ret", "25.330", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "L57", "ReportFormulaObject", "", "nominal", "20.00", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "L58", "ReportFormulaObject", "", "nominal", "50.00", "ok", ""),
    ]
    calculation_map = build_report_calculation_map("<CmData />", {"HeatUp&CoolDown": evaluations})
    output = write_report_workbook(tmp_path / "heatup.xlsx", "HeatUp and CoolDownTime", {"HeatUp&CoolDown": evaluations}, calculation_map=calculation_map)
    report = build_heatup_cooldown_report(evaluations, calculation_map)

    from openpyxl import load_workbook

    wb = load_workbook(output, data_only=True)
    ws = wb["HeatUp&CoolDown"]

    assert round(report.heatup_observed_min, 3) == 3.847
    assert round(report.cooldown_observed_min, 3) == 10.65
    assert round(ws["D26"].value, 3) == 3.847
    assert round(ws["D65"].value, 3) == 3.847
    assert ws["C65"].value == "6.252 -> 12.099"
    assert ws["E26"].value == "Test passed"


def test_foq_contract_cell_value_map_includes_derived_heatup_values():
    evaluations = [
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "J65", "ReportFormulaObject", "", "ret", "6.252", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "K65", "ReportFormulaObject", "", "ret", "12.099", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "L65", "ReportFormulaObject", "", "ret", "12.680", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "M65", "ReportFormulaObject", "", "ret", "25.330", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "J66", "ReportFormulaObject", "", "ret", "6.252", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "K66", "ReportFormulaObject", "", "ret", "12.099", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "L66", "ReportFormulaObject", "", "ret", "12.680", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "M66", "ReportFormulaObject", "", "ret", "25.330", "ok", ""),
    ]
    calculation_map = build_report_calculation_map("<CmData />", {"HeatUp&CoolDown": evaluations})
    values = build_report_cell_value_map({"HeatUp&CoolDown": evaluations}, calculation_map)

    assert values[("heatup&cooldown", "D26")].value == 3.8
    assert values[("heatup&cooldown", "D27")].value == 10.7
    assert values[("heatup&cooldown", "E26")].value == "Test passed"


def test_foq_contract_temp_accuracy_result_uses_raw_report_cell_deviation():
    rows = [
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "I66", "ReportFormulaObject", "", "nominal", "10.00", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "L66", "ReportFormulaObject", "ExtTemp_LowerCC", "lower", "10.504", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Accuracy_H", "Temp Accuracy", "M66", "ReportFormulaObject", "ExtTemp_UpperCC", "upper", "10.10", "ok", ""),
    ]
    values = build_report_cell_value_map({"Temp Accuracy": rows}, [])

    assert values[("temp accuracy", "D66")].value == 0.504
    assert values[("temp accuracy", "E66")].value == "failed"


def test_foq_contract_cell_value_map_normalizes_float_artifacts():
    rows = [
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "J65", "ReportFormulaObject", "", "ret", "3.918", "ok", ""),
        FormulaEvaluation("REPORT_A", "HeatUp and CoolDownTime", "HeatUp&CoolDown", "K65", "ReportFormulaObject", "", "ret", "9.635", "ok", ""),
    ]
    values = build_report_cell_value_map({"HeatUp&CoolDown": rows}, build_report_calculation_map("<CmData />", {"HeatUp&CoolDown": rows}))

    assert values[("heatup&cooldown", "D26")].value == 3.7


def test_foq_contract_cell_value_map_includes_common_derived_sheet_values():
    precision_rows = [
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "K65", "ReportFormulaObject", "ExtTemp_LowerCC", "avg", "50.02", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "L65", "ReportFormulaObject", "ExtTemp_UpperCC", "avg", "50.03", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "K66", "ReportFormulaObject", "ExtTemp_LowerCC", "avg", "50.01", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "L66", "ReportFormulaObject", "ExtTemp_UpperCC", "avg", "50.04", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "K67", "ReportFormulaObject", "ExtTemp_LowerCC", "avg", "50.02", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "L67", "ReportFormulaObject", "ExtTemp_UpperCC", "avg", "50.03", "ok", ""),
    ]
    pcc_rows = [
        FormulaEvaluation("REPORT_A", "Temperature Stability_and_PCC_H", "PCC", "K105", "ReportFormulaObject", "", "rt", "16.619", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Stability_and_PCC_H", "PCC", "L105", "ReportFormulaObject", "", "rt", "17.491", "ok", ""),
    ]
    preheater_rows = [
        FormulaEvaluation("REPORT_A", "Preheater Connection Test", "Preheater Ports_Noise", "J72", "ReportFormulaObject", "", "rt", "1.109", "ok", ""),
        FormulaEvaluation("REPORT_A", "Preheater Connection Test", "Preheater Ports_Noise", "K72", "ReportFormulaObject", "", "rt", "1.628", "ok", ""),
        FormulaEvaluation("REPORT_A", "Preheater Connection Test", "Preheater Ports_Noise", "J117", "ReportFormulaObject", "", "present", "Yes", "ok", ""),
        FormulaEvaluation("REPORT_A", "Preheater Connection Test", "Preheater Ports_Noise", "K117", "ReportFormulaObject", "", "memory", "OK", "ok", ""),
        FormulaEvaluation("REPORT_A", "Preheater Connection Test", "Preheater Ports_Noise", "J82", "ReportFormulaObject", "PrehtLeft_Temp", "avg", "39.998", "ok", ""),
        FormulaEvaluation("REPORT_A", "Preheater Connection Test", "Preheater Ports_Noise", "K82", "ReportFormulaObject", "PREH_L_HeaterTemp_Actual", "avg", "39.682", "ok", ""),
    ]

    values = build_report_cell_value_map(
        {
            "Temp Precision": precision_rows,
            "PCC": pcc_rows,
            "Preheater Ports_Noise": preheater_rows,
        },
        build_report_calculation_map("<CmData />", {"Temp Precision": precision_rows, "PCC": pcc_rows, "Preheater Ports_Noise": preheater_rows}),
    )

    assert values[("temp precision", "D26")].value == 0.01
    assert values[("temp precision", "E26")].value == "Test passed"
    assert values[("pcc", "D26")].value == 0.87
    assert values[("preheater ports_noise", "C26")].value == "Test passed"
    assert values[("preheater ports_noise", "L82")].value == -0.3


def test_foq_contract_temp_precision_uses_sensor_repeatability_not_sensor_offset():
    rows = [
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "K65", "ReportFormulaObject", "ExtTemp_LowerCC", "avg", "49.9995435685", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "K66", "ReportFormulaObject", "ExtTemp_LowerCC", "avg", "49.9964315353", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "K67", "ReportFormulaObject", "ExtTemp_LowerCC", "avg", "49.9974688797", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "L65", "ReportFormulaObject", "ExtTemp_UpperCC", "avg", "50.009626556", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "L66", "ReportFormulaObject", "ExtTemp_UpperCC", "avg", "50.0141493776", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Precision_and_Fan", "Temp Precision", "L67", "ReportFormulaObject", "ExtTemp_UpperCC", "avg", "50.013153527", "ok", ""),
    ]
    values = build_report_cell_value_map({"Temp Precision": rows}, build_report_calculation_map("<CmData />", {"Temp Precision": rows}))

    assert values[("temp precision", "D26")].value == 0
    assert values[("temp precision", "E26")].value == "Test passed"


def test_foq_contract_cell_value_map_aliases_5c_calibration_cells():
    rows = [
        FormulaEvaluation("REPORT_A", "Temperature Calibration", "Temp_Calib_Internal", "J15", "ReportFormulaObject", "", "duration", "4.2", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Calibration", "Temp_Calib_Internal", "J16", "ReportFormulaObject", "ExtTemp_UpperCC", "upper drift", "0.001", "ok", ""),
        FormulaEvaluation("REPORT_A", "Temperature Calibration", "Temp_Calib_Internal", "J17", "ReportFormulaObject", "ExtTemp_LowerCC", "lower drift", "0.002", "ok", ""),
    ]
    values = build_report_cell_value_map({"Temp_Calib_Internal": rows}, [])

    assert values[("temp_calib_internal", "C22")].value == 4.2
    assert values[("temp_calib_internal", "D22")].value == 0.001
    assert values[("temp_calib_internal", "E22")].value == 0.002


def test_foq_contract_cell_value_map_includes_column_id_results():
    rows = [
        FormulaEvaluation("REPORT_A", "ColumnIDs", "Column ID", "L46", "ReportFormulaObject", "", "desc", "A", "ok", ""),
        FormulaEvaluation("REPORT_A", "ColumnIDs", "Column ID", "L47", "ReportFormulaObject", "", "desc", "B", "ok", ""),
        FormulaEvaluation("REPORT_A", "ColumnIDs", "Column ID", "L48", "ReportFormulaObject", "", "desc", "wrong", "ok", ""),
    ]
    values = build_report_cell_value_map({"Column ID": rows}, [])

    assert values[("column id", "C26")].value == "Test passed"
    assert values[("column id", "C27")].value == "Test passed"
    assert values[("column id", "C28")].value == "Test failed"
    assert values[("column id", "C29")].status == "missing_cell"


def test_foq_contract_cell_value_map_includes_cross_sheet_metadata_results():
    values = build_report_cell_value_map(
        {
            "Definitions": [
                FormulaEvaluation("REPORT_A", "Factory Default", "Definitions", "C15", "ReportFormulaObject", "", "model", "VH-C10-A", "ok", ""),
                FormulaEvaluation("REPORT_A", "Factory Default", "Definitions", "D15", "ReportFormulaObject", "", "serial", "6545327", "ok", ""),
            ],
            "Internal Use": [
                FormulaEvaluation("REPORT_A", "Factory Default", "Internal Use", "B20", "ReportFormulaObject", "", "seq", "6545327", "ok", ""),
                FormulaEvaluation("REPORT_A", "Factory Default", "Internal Use", "E12", "ReportFormulaObject", "", "module revision", "03", "ok", ""),
            ],
        },
        [],
    )

    assert values[("definitions", "J8")].value == "03"
    assert values[("internal use", "F10")].value == "ok"


def test_foq_contract_workbook_writes_db_data(tmp_path):
    _sheet_name, locations = locations_for_device_type(PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls", "VH-C10-A")
    selected = filter_locations_for_report(locations, "HeatUp and CoolDownTime.XLS", "HeatUp&CoolDown")
    values = [
        type("ContractValue", (), {"location": row, "injection_name": "HeatUp and CoolDownTime", "value": index, "status": "ok", "detail": "test"})()
        for index, row in enumerate(selected, 1)
    ]

    output = write_foq_contract_workbook(tmp_path / "contract.xlsx", "VH-C10-A", "VH-C10-A", values)

    from openpyxl import load_workbook

    wb = load_workbook(output, data_only=True)
    ws = wb["DB Data"]
    trace = wb["Dependency Trace"]
    headers = [cell.value for cell in ws[1]]

    assert "HeatUp_Time_20to50" in headers
    assert "RES_CoolDown" in headers
    assert trace["A1"].value == "dbField"
    assert "Calculation Tier" in [cell.value for cell in trace[1]]


def test_foq_contract_workbook_formats_test_date(tmp_path):
    _sheet_name, locations = locations_for_device_type(PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls", "VH-C10-A")
    selected = [row for row in locations if row.db_field in {"TestDate", "SubmitDate"}]
    values = [
        type("ContractValue", (), {"location": row, "injection_name": "Factory Default", "value": 46206.5, "status": "ok", "detail": "date"})()
        for row in selected
    ]

    output = write_foq_contract_workbook(tmp_path / "contract.xlsx", "VH-C10-A", "VH-C10-A", values)

    from openpyxl import load_workbook

    wb = load_workbook(output, data_only=True)
    ws = wb["DB Data"]
    headers = [cell.value for cell in ws[1]]
    test_date_column = headers.index("TestDate") + 1
    submit_date_column = headers.index("SubmitDate") + 1

    assert ws.cell(row=2, column=test_date_column).number_format == "yyyy-mm-dd"
    assert ws.cell(row=2, column=submit_date_column).number_format == "yyyy-mm-dd"


def test_foq_contract_workbook_applies_report_display_precision(tmp_path):
    _sheet_name, locations = locations_for_device_type(PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls", "VH-C10-A")
    wanted = {
        "TempAcc20": -0.0145416667000013,
        "Noise_CC_Temp": 0.000807913953068,
        "Slope_Cal120_U": -0.0101551180052,
        "HeatUp_Time_20to50": 3.847,
    }
    selected = [row for row in locations if row.db_field in wanted]
    values = [
        type("ContractValue", (), {"location": row, "injection_name": "test", "value": wanted[row.db_field], "status": "ok", "detail": "precision"})()
        for row in selected
    ]

    output = write_foq_contract_workbook(tmp_path / "contract.xlsx", "VH-C10-A", "VH-C10-A", values)

    from openpyxl import load_workbook

    wb = load_workbook(output, data_only=True)
    ws = wb["DB Data"]
    row = {header.value: ws.cell(row=2, column=header.column).value for header in ws[1]}

    assert row["TempAcc20"] == -0.01
    assert row["Noise_CC_Temp"] == 0.001
    assert row["Slope_Cal120_U"] == -0.01
    assert row["HeatUp_Time_20to50"] == 3.8


def test_foq_sequence_output_uses_sequence_name(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="6545327" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="6545327.cmd" />
</ChromeleonHeader>"""
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("6545327.cmd", b"")

    package = load_cmbx_package(cmbx)

    assert export_service._package_sequence_output_path(package, tmp_path / "out") == tmp_path / "out" / "6545327"
    assert export_service._package_sequence_output_name(package) == "6545327"


def test_dependency_trace_describes_heatup_reverse_chain():
    _sheet_name, locations = locations_for_device_type(PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls", "VH-C10-A")
    location = next(row for row in locations if row.db_field == "HeatUp_Time_20to50")
    value = type("ContractValue", (), {"location": location, "injection_name": "HeatUp and CoolDownTime", "value": 3.847, "status": "ok", "detail": "Heat-up observed time"})()

    trace = dependency_trace_for_contract_value(value)

    assert trace.calculation_tier == "2 - derived from audit RetTime cells"
    assert "J65" in trace.cm_formula_cells
    assert "Definitions!HeatUp & Cool Down" in trace.definition_source
    assert "Audit RetTimes" in trace.raw_data_sources


def test_dependency_trace_keeps_missing_column_id_source_chain():
    _sheet_name, locations = locations_for_device_type(PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls", "VH-C10-A")
    location = next(row for row in locations if row.db_field == "RES_ColumnID_A")
    value = type("ContractValue", (), {"location": location, "injection_name": "ColumnIDs", "value": "", "status": "missing_cell", "detail": "Cell is not evaluated yet"})()

    trace = dependency_trace_for_contract_value(value)

    assert trace.calculation_tier == "3 - workbook-derived result cell not mapped yet"
    assert "Column ID!L46:L49" in trace.cm_formula_cells
    assert "Column description" in trace.raw_data_sources


def test_dependency_trace_keeps_missing_preheater_port_source_chain():
    _sheet_name, locations = locations_for_device_type(PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls", "VH-C10-A")
    location = next(row for row in locations if row.db_field == "RES_Preheater_Left_Port")
    value = type("ContractValue", (), {"location": location, "injection_name": "Preheater Connection Test", "value": "", "status": "missing_cell", "detail": "Cell is not evaluated yet"})()

    trace = dependency_trace_for_contract_value(value)

    assert trace.calculation_tier == "3 - workbook-derived preheater result cell not mapped yet"
    assert "J72:J73" in trace.cm_formula_cells
    assert "module state metadata" in trace.raw_data_sources


def test_embedded_summary_reads_sequence_cmd_context(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="5" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" Url="chrom://example/METHOD_A.instmeth" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    sequence_cmd = b"prefix METHOD_A Test Instrument Method trigger RetTimes.RetTime1 condition formula suffix"
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd)

    package = load_cmbx_package(cmbx)
    method = package.methods_and_reports[0]
    summary = build_embedded_object_summary(package, method)
    exported = export_element(package, method, tmp_path / "out")

    assert summary.occurrences == [7]
    assert any("RetTimes.RetTime1" in item for item in summary.strings)
    assert exported.name == "METHOD_A_summary.txt"
    assert "Readable Context" in exported.read_text(encoding="utf-8")


def test_injection_method_links_read_relative_urls(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="2" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection" />
    <ChromeleonElement Id="5" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
    <ChromeleonElement Id="6" Name="PROC_A" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    sequence_cmd = b"Injection A\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x06PROC_A\x12\x20\x12\x0d\x1a\x0bRelativeUrl*\x08METHOD_A"
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd)

    package = load_cmbx_package(cmbx)
    links = build_injection_method_links(package)

    assert links["Injection A"].processing_method == "PROC_A"
    assert links["Injection A"].instrument_method == "METHOD_A"


def test_multi_sequence_links_and_folder_report_scope(tmp_path):
    cmbx = tmp_path / "multi.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="f" Name="Folder" ItemType="Dionex.Chromeleon.Data.SubFolder">
    <ChromeleonElement Id="s1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
      <ChromeleonElement Id="i1" Name="Injection A" ItemType="Dionex.Chromeleon.Data.Injection" />
      <ChromeleonElement Id="m1" Name="METHOD_A" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
      <ChromeleonElement Id="p1" Name="PROC_A" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
    </ChromeleonElement>
    <ChromeleonElement Id="s2" Name="Seq2" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq2.cmd">
      <ChromeleonElement Id="i2" Name="Injection B" ItemType="Dionex.Chromeleon.Data.Injection" />
      <ChromeleonElement Id="m2" Name="METHOD_B" ItemType="Dionex.Chromeleon.Data.InstrumentMethod" />
      <ChromeleonElement Id="p2" Name="PROC_B" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
    </ChromeleonElement>
    <ChromeleonElement Id="r1" Name="Shared Report" ItemType="Dionex.Chromeleon.Data.ReportDefinition" Filename="Report1.cmd" />
  </ChromeleonElement>
  <ChromeleonElement Id="r2" Name="Shared Report" ItemType="Dionex.Chromeleon.Data.ReportDefinition" Filename="Report2.cmd" />
</ChromeleonHeader>"""

    def sequence_cmd(injection: str, processing: str, method: str) -> bytes:
        return (
            injection.encode() + b"\x12\x20\x12\x0d\x1a\x0bRelativeUrl*" + bytes([len(processing)]) + processing.encode()
            + b"\x12\x20\x12\x0d\x1a\x0bRelativeUrl*" + bytes([len(method)]) + method.encode()
        )

    def varint(value: int) -> bytes:
        result = bytearray()
        while True:
            byte = value & 0x7F
            value >>= 7
            result.append(byte | (0x80 if value else 0))
            if not value:
                return bytes(result)

    def field(number: int, value: bytes) -> bytes:
        return varint((number << 3) | 2) + varint(len(value)) + value

    report_cpxm = b"CpXm-report-payload"
    report_payload = field(19, field(15, field(1, report_cpxm)) + field(28, b"Shared Report")) + field(30, b"trailing metadata")
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd("Injection A", "PROC_A", "METHOD_A"))
        archive.writestr("Seq2.cmd", sequence_cmd("Injection B", "PROC_B", "METHOD_B"))
        archive.writestr("Report1.cmd", report_payload)
        archive.writestr("Report2.cmd", report_payload)

    package = load_cmbx_package(cmbx)
    links = build_injection_method_links(package)
    injection_b = next(item for item in package.injections if item.name == "Injection B")
    link_b = get_injection_method_link(links, injection_b)
    reports = report_templates_for_sequence(package, package.sequences[1])
    embedded = extract_embedded_report_template(package, reports[0])

    assert link_b is not None
    assert link_b.processing_method == "PROC_B"
    assert link_b.instrument_method == "METHOD_B"
    assert link_b.sequence_name == "Seq2"
    assert [report.name for report in reports] == ["Shared Report"]
    assert summarize_package(package)["report_templates"] == 1
    assert embedded is not None
    assert embedded.sequence_name == "(standalone report template)"
    assert embedded.sequence_entry == "Report1.cmd"
    assert embedded.cpxm_payload == report_cpxm
    assert embedded.cpxm_end < len(report_payload)


def test_processing_method_inspector_extracts_xml_evidence(tmp_path):
    cmbx = tmp_path / "sample.cmbx"
    header = """<?xml version="1.0" encoding="UTF-8"?>
<ChromeleonHeader>
  <ChromeleonElement Id="1" Name="Seq1" ItemType="Dionex.Chromeleon.Data.Sequence" Filename="Seq1.cmd">
    <ChromeleonElement Id="6" Name="PROC_A" ItemType="Dionex.Chromeleon.Data.ProcessingMethod" />
  </ChromeleonElement>
</ChromeleonHeader>"""
    xml = """<CmData><DesignerRoot type="ProcessingMethodRootControl">
<Controls>
<Item type="SSTGrid"><Text value="SST/IRC" /><HeaderText value="Pass Actions" /><OriginalName value="PassActions" />
  <Controls />
  <IColumnVisibilityManagement type="ColumnInfo"><HeaderText value="Name" /><OriginalName value="Name" /></IColumnVisibilityManagement>
  <IColumnVisibilityManagement type="ColumnInfo"><HeaderText value="Pass Actions" /><OriginalName value="PassActions" /></IColumnVisibilityManagement>
  <IColumnVisibilityManagement type="ColumnInfo"><HeaderText value="Fail Actions" /><OriginalName value="FailActions" /></IColumnVisibilityManagement>
</Item>
<Item><HeaderText value="Fail Actions" /><OriginalName value="FailActions" /></Item>
<Item><Description value="Injection Type" /><Text value="Stop sequence on fail" /></Item>
</Controls>
</DesignerRoot></CmData>"""
    sequence_cmd = b"prefix PROC_A " + gzip.compress(xml.encode("utf-8"))
    with zipfile.ZipFile(cmbx, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("header.xml", header)
        archive.writestr("Seq1.cmd", sequence_cmd)

    package = load_cmbx_package(cmbx)
    evidence = inspect_processing_method(package, "PROC_A")

    assert evidence.found
    assert evidence.root_types == ("ProcessingMethodRootControl",)
    assert "Pass Actions" in evidence.action_columns
    assert "Fail Actions" in evidence.action_columns
    assert "SST/IRC" in evidence.tab_labels
    assert any(summary.startswith("SSTGrid: 3 column") for summary in evidence.grid_summaries)
    assert any("empty Controls" in summary for summary in evidence.grid_summaries)
    assert evidence.sst_grid_count == 1
    assert evidence.empty_sst_grid_count == 1
    assert evidence.row_candidate_count == 0
    assert dict(evidence.token_counts)["IRC"] >= 1
    assert "business rows absent" in evidence.status
    assert "empty_sst_grids=1" in evidence.summary
    assert "row_candidates=0" in evidence.summary
    assert "Processing Method: PROC_A" in "\n".join(evidence.to_lines())


def test_parse_external_instrument_method_txt(tmp_path):
    method = tmp_path / "Accuracy Method.txt"
    method.write_text(
        "\n".join(
            [
                "{Initial Time}\tInstrument Setup\t\t",
                "Trigger\t\t\"Gradient_1\",",
                "\tVariables.GenericDouble1\t10.0\tFirst test temperature",
                "\tColumnComp.CC.Temperature.Nominal\tVariables.GenericDouble1\t",
                "\tWait\tColumnComp.CC.TempReady AND StabVars.LowerReady AND StabVars.UpperReady,\t",
                "\tRetTimes.RetTime1\tSystem.Retention\t",
            ]
        ),
        encoding="utf-8",
    )

    parsed = parse_instrument_method_txt(method)
    discovered = discover_external_instrument_methods(method.parent / "sample.cmbx", ["TEMPERATURE_ACCURACY"])

    assert len(parsed.triggers) == 1
    assert len(parsed.ret_times) == 1
    assert len(parsed.setpoints) == 2
    assert "RetTime assignments: 1" in parsed.summary_text()
    assert "TEMPERATURE_ACCURACY" in discovered


def test_method_contract_summarizes_execution_evidence():
    rows = [
        {
            "Method": "METHOD_A",
            "Stage": "Equilibration",
            "Action": "SET",
            "Target": "RetTimes.RetTime1",
            "Value": "0",
        },
        {
            "Method": "METHOD_A",
            "Stage": "Run",
            "Action": "RUN",
            "Target": "Thermometer1.ExtTemp_UpperCC.AcqOn",
            "Value": "",
        },
        {
            "Method": "METHOD_A",
            "Stage": "Run",
            "Action": "RUN",
            "Target": "System.Trigger",
            "Value": '"T1", ExtTemp_UpperCC>20, TrueTime=120',
        },
        {
            "Method": "METHOD_A",
            "Stage": "Run",
            "Action": "SET",
            "Target": "RetTimes.RetTime1",
            "Value": "System.Retention",
        },
        {
            "Method": "METHOD_A",
            "Stage": "Run",
            "Action": "RUN",
            "Target": "Log",
            "Value": "ColumnComp.ModelNo",
        },
        {
            "Method": "METHOD_A",
            "Stage": "Run",
            "Action": "RUN",
            "Target": "Wait",
            "Value": "CC.TempReady",
        },
        {
            "Method": "METHOD_A",
            "Stage": "Run",
            "Action": "SET",
            "Target": "ColumnComp.CC.Temperature.Nominal",
            "Value": "50.0",
        },
    ]

    contract = build_method_contract_from_flow_rows(rows)
    summary = method_contract_summary_text(contract)
    table = method_contracts_tsv([contract])

    assert contract.method_name == "METHOD_A"
    assert contract.acquisition_on == ("Thermometer1.ExtTemp_UpperCC",)
    assert contract.ret_time_initializations == ("RetTimes.RetTime1",)
    assert contract.ret_time_emissions == ("RetTimes.RetTime1",)
    assert contract.logged_properties == ("ColumnComp.ModelNo",)
    assert contract.wait_conditions == ("CC.TempReady",)
    assert contract.trigger_definitions == ('"T1", ExtTemp_UpperCC>20, TrueTime=120',)
    assert contract.temperature_setpoints == ("ColumnComp.CC.Temperature.Nominal=50.0",)
    assert "RetTime Emissions" in summary
    assert "METHOD_A" in table


def test_test_intent_contract_coverage_links_catalog_to_method_contract():
    catalog = {
        "test_intents": {
            "sample_test": {
                "instrument_method": "METHOD_A",
                "ret_times": ["RetTime1", "RetTime2"],
                "channels": ["Thermometer1.ExtTemp_UpperCC", "ColumnComp.PCC_Temp"],
                "report_dependencies": ["AUDIT.Column_A.Description(0,\"forward\")"],
                "device_bindings": {
                    "VH-C10-A": {
                        "injection_name": "Sample Injection",
                        "processing_method": "No_Integration",
                    }
                },
            }
        }
    }
    contract = build_method_contract_from_flow_rows(
        [
            {
                "Method": "METHOD_A",
                "Stage": "Run",
                "Action": "SET",
                "Target": "RetTimes.RetTime1",
                "Value": "System.Retention",
            },
            {
                "Method": "METHOD_A",
                "Stage": "Run",
                "Action": "RUN",
                "Target": "Thermometer1.ExtTemp_UpperCC.AcqOn",
                "Value": "",
            },
            {
                "Method": "METHOD_A",
                "Stage": "Run",
                "Action": "RUN",
                "Target": "Log",
                "Value": "Column_A.Description",
            },
        ]
    )

    coverage = build_test_intent_contract_coverages(catalog, "VH-C10-A", [contract])[0]
    table = contract_coverages_tsv([coverage])

    assert not coverage.passed
    assert coverage.missing_ret_times == ("RetTime2",)
    assert coverage.missing_channels == ("ColumnComp.PCC_Temp",)
    assert coverage.missing_audit_properties == ()
    assert "sample_test" in table


def test_real_tcc_validation_cmbx_structure_if_available():
    cmbx = PROJECT_ROOT / "tcc_temperature_control_analyzer_staging" / "DATA" / "20260701_New" / "20260701_New.cmbx"
    if not cmbx.exists():
        return

    package = load_cmbx_package(cmbx)
    counts = summarize_package(package)
    injection_names = [injection.name for injection in package.injections]
    accuracy = next(injection for injection in package.injections if injection.name == "Temperature Accuracy_H")
    accuracy_channels = [child.name for child in accuracy.children if child.kind == "signal"]

    assert counts["sequences"] == 1
    assert counts["injections"] == 13
    assert counts["channels"] >= 40
    assert counts["audits"] == 10
    assert "Temperature Accuracy_H" in injection_names
    assert "ExtTemp_UpperCC" in accuracy_channels
    assert "ExtTemp_LowerCC" in accuracy_channels
    assert any(method.name == "TEMPERATURE_ACCURACY" for method in package.methods_and_reports)
    assert any(report.name.startswith("Report_") for report in package.methods_and_reports)


def test_real_tcc_injection_method_links_if_available():
    cmbx = PROJECT_ROOT / "tcc_temperature_control_analyzer_staging" / "DATA" / "20260701_New" / "20260701_New.cmbx"
    if not cmbx.exists():
        return

    package = load_cmbx_package(cmbx)
    links = build_injection_method_links(package)

    assert links["Temperature Accuracy_H"].processing_method == "ACCURACY_IRC_STOP_H"
    assert links["Temperature Accuracy_H"].instrument_method == "TEMPERATURE_ACCURACY"
    assert links["Temperature Calibration"].instrument_method == "TEMPERATURE_CALIBRATION"


def test_real_tcc_processing_method_embedded_xml_if_available():
    cmbx = PROJECT_ROOT / "tcc_temperature_control_analyzer_staging" / "DATA" / "20260701_New" / "20260701_New.cmbx"
    if not cmbx.exists():
        return

    package = load_cmbx_package(cmbx)
    method = next(method for method in package.methods_and_reports if method.name == "ACCURACY_IRC_STOP_H")
    summary = build_embedded_object_summary(package, method)

    assert summary.occurrences
    assert any(section.text.startswith("<CmData>") for section in summary.sections)
    assert any("ProcessingMethodRootControl" in section.text for section in summary.sections)


def test_real_tcc_processing_method_inspector_enriches_alignment_if_available():
    cmbx = PROJECT_ROOT / "tcc_temperature_control_analyzer_staging" / "DATA" / "20260701_New" / "20260701_New.cmbx"
    if not cmbx.exists():
        return

    package = load_cmbx_package(cmbx)
    records = build_foq_alignment_records(packages=(package,))
    accuracy = next(record for record in records if record.family == "TCC" and record.test_intent == "temperature_accuracy")

    joined = "\n".join(accuracy.method_evidence)
    assert "Processing inspector:" in joined
    assert "ACCURACY_IRC_STOP_H" in joined
    assert "ProcessingMethodRootControl" in joined
    assert "business rows absent" in joined


def test_foq_result_locations_v283_vtcc_mapping_if_available():
    mapping = PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls"
    if not mapping.exists():
        return

    sheet_name, locations = locations_for_device_type(mapping, "VH-C10-A")
    heatup = filter_locations_for_report(locations, "HeatUp and CoolDownTime.XLS", "HeatUp&CoolDown")
    accuracy = filter_locations_for_report(locations, "Temperature Accuracy_H.XLS", "Temp Accuracy")
    summary = summarize_locations(locations)

    assert sheet_name == "VH-C10-A"
    assert summary["fields"] >= 80
    assert any(row.db_field == "HeatUp_Time_20to50" and row.report_cell == "D26" for row in heatup)
    assert any(row.db_field == "CoolDown_Time_50to20" and row.report_cell == "D27" for row in heatup)
    assert any(row.db_field == "RES_HeatUp" and row.report_cell == "E26" for row in heatup)
    assert any(row.db_field == "TempAcc20" and row.report_cell == "D67" for row in accuracy)


def test_foq_result_locations_device_type_lookup_if_available():
    mapping = PROJECT_ROOT / "foq" / "FOQResultLocations_V2.83.xls"
    if not mapping.exists():
        return

    import xlrd

    workbook = xlrd.open_workbook(mapping)
    device_types = read_device_type_mappings(workbook)

    assert device_types["VH-C10-A"].sheet_name == "VH-C10-A"
    assert device_types["VN-C10-A"].sheet_name == "VN-C10-A"


def test_tcc_accuracy_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_ACCURACY_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 3: Instrument Method Command" in doc
    assert "Contract 4: Processing Method IRC" in doc
    assert "Contract 5: Report Formula" in doc
    assert "VH-C10-A" in doc
    assert "VC-C10-A" in doc
    assert "VA-C10-A" in doc
    assert "ACCURACY_IRC_STOP_H" in doc
    assert "ACCURACY_IRC_STOP_C" in doc
    assert "RetTimes.RetTime1..RetTime5" in doc
    assert "RetTimeN - 1.0" in doc
    assert "ExtTemp_LowerCC" in doc
    assert "ExtTemp_UpperCC" in doc
    assert "Open Verification Required" in doc


def test_tcc_calibration_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_CALIBRATION_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "TEMPERATURE_CALIBRATION" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "GenericBool0 = 1" in doc
    assert "Temperature Accuracy_H" in doc
    assert "Temperature Accuracy_C" in doc
    assert "FOQ_VX-C10_V2_00_AdditionalInjections" in doc
    assert "NO_INTEGRATION" in doc
    assert "RetTimes.RetTime1..RetTime8" in doc
    assert "CCCalib.CalDevU01" in doc
    assert "ExtTemp_UpperCC.Signal - CC.TempActual_Upper" in doc
    assert "TempCalibrationDeviationUpper1" in doc
    assert "TempCal120_U" in doc
    assert "TempCal85_U" in doc
    assert "Open Verification Required" in doc


def test_tcc_stability_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_STABILITY_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "TEMPERATURE_STABILITY_AND_PCC_70_H" in doc
    assert "TEMPERATURE_STABILITY_70_C" in doc
    assert "Temp Stability_Noise" in doc
    assert "K61:K75" in doc
    assert "L61:L75" in doc
    assert "Line-level decoded flow evidence for VC/VA" in doc
    assert "TEMPERATURE_STABILITY_70_C_embedded_method_flow.txt" in doc
    assert "RetTimes.RetTime2" in doc
    assert "RetTimes.RetTime3" in doc
    assert "RetTimes.RetTime4" in doc
    assert "Performance_PCC" in doc
    assert "Definitions!Temperature Stability = 0.05" in doc
    assert "Definitions!PCC CoolDownTime = 2 min" in doc
    assert "Resolved by TD / Report Contract Evidence" in doc
    assert "Workbook summary/pass behavior for `Temp Stability_Noise!D26/E26`" in doc
    assert "Temp Stability_Noise!D26 = 0.010000000000005116" in doc
    assert "PCC!D26 = 0.8299999999999983" in doc
    assert "NO_INTEGRATION" in doc
    assert "Open Verification Required" in doc


def test_tcc_precision_fan_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_PRECISION_FAN_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "TEMPERATURE_PRECISION_AND_FAN" in doc
    assert "TEMPERATURE_PRECISION" in doc
    assert "CORRECT_STABILITY_INJ_INSERTION" in doc
    assert "Variables.GenericBool0" in doc
    assert "Temperature Stability_and_PCC_H" in doc
    assert "Temperature Stability_C" in doc
    assert "FOQ_VX-C10_V2_00_AdditionalInjections" in doc
    assert "Temp Precision" in doc
    assert "Fan" in doc
    assert "K65" in doc
    assert "K66" in doc
    assert "K67" in doc
    assert "L65" in doc
    assert "14,14.8" in doc
    assert "36,36.8" in doc
    assert "58,58.8" in doc
    assert "ForcedAir" in doc
    assert "StillAir" in doc
    assert "TempPrecision" in doc
    assert "RES_TempPrecision" in doc
    assert "Definitions!Temperature Precision = 0.1" in doc
    assert "<= 0.1 C" in doc
    assert "Open Verification Required" in doc


def test_tcc_heatup_cooldown_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_HEATUP_COOLDOWN_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "TEMP_HEAT_UP_DOWN_20_50_20" in doc
    assert "HeatUp&CoolDown" in doc
    assert "RetTimes.RetTime1" in doc
    assert "RetTimes.RetTime3" in doc
    assert "RetTimes.RetTime4" in doc
    assert "RetTimes.RetTime6" in doc
    assert "T_Start_Ext" in doc
    assert "T_50_Int" in doc
    assert "T_20_Int" in doc
    assert "HeatUp_Time_20to50" in doc
    assert "CoolDown_Time_50to20" in doc
    assert "RetTime2 - RetTime1 - 2.0" in doc
    assert "RetTime5 - RetTime4 - 2.0" in doc
    assert "row 65/66" in doc.lower()
    assert "Resolved by Report Formula Evidence" in doc
    assert "row 66 uses external endpoints" in doc
    assert "C65/D65/E65" in doc
    assert "C66/D66/E66" in doc
    assert "D26 = E65" in doc
    assert "D27 = E66" in doc
    assert "C26 = 15" in doc
    assert "Calculation Map reads `Definitions / HeatUp & Cool Down = 15`" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "Resolved processing-context boundary" in doc
    assert "No HeatUp-specific numbered open verification item remains" in doc


def test_tcc_preheater_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_PREHEATER_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "PREHEATER" in doc
    assert "Preheater Ports_Noise" in doc
    assert "RetTimes.RetTime1" in doc
    assert "RetTimes.RetTime2" in doc
    assert "RetTimes.RetTime3" in doc
    assert "RetTimes.RetTime4" in doc
    assert "PrehtLeft_Temp" in doc
    assert "PrehtRight_Temp" in doc
    assert "PREH_L_HeaterTemp_Actual" in doc
    assert "PREH_R_HeaterTemp_Actual" in doc
    assert "ModulePresent" in doc
    assert "MemoryState" in doc
    assert "Diff_PhLeft_HtTmp" in doc
    assert "Diff_PhRight_HtTmp" in doc
    assert "Noise_PrehtLeft_Temp" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "Open Verification Required" in doc


def test_tcc_column_id_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_COLUMN_ID_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "ColumnID" in doc
    assert "ColumnIDs" in doc
    assert "Column ID" in doc
    assert "CORRECT_STABILITY_INJ_INSERTION" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "Column_A.CardState" in doc
    assert "Column_A.Description" in doc
    assert "Column_D.Description" in doc
    assert "AUDIT.Column_A.Description" in doc
    assert "AUDIT.Column_D.Description" in doc
    assert "RES_ColumnID_A" in doc
    assert "RES_ColumnID_D" in doc
    assert "Variables.GenericLong9" in doc
    assert "System.AbortQueue" in doc
    assert "RetTimes | not used" in doc
    assert "Open Verification Required" in doc


def test_tcc_valve_keypad_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_VALVE_KEYPAD_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "VALVES" in doc
    assert "Valve_Keypad" in doc
    assert "ColumnComp.UpperValve.CurrentPosition" in doc
    assert "ColumnComp.LowerValve.CurrentPosition" in doc
    assert "UpperValve.Precision" in doc
    assert "LowerValve.Precision" in doc
    assert "6_1" in doc
    assert "1_2" in doc
    assert "AUDIT.UpperValve.CurrentPosition" in doc
    assert "AUDIT.LowerValve.CurrentPosition" in doc
    assert "AUDIT.ColumnComp.FastCoolState" in doc
    assert "ColumnComp.Disconnect" in doc
    assert "ColumnComp.Connect" in doc
    assert "ColumnComp.Connected = Connected" in doc
    assert "VA-C10-A" in doc
    assert "lower only" in doc
    assert "no DB contract expected" in doc
    assert "fixed-time" in doc
    assert "Open Verification Required" in doc


def test_tcc_liquid_leak_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_LIQUID_LEAK_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "LIQUID LEAK" in doc
    assert "LiquidLeaktest" in doc
    assert "Liquid Leak Test" in doc
    assert "ColumnComp.LiquidLeakSensor" in doc
    assert "LiquidLeak=Leak" in doc
    assert "Log LiquidLeak" in doc
    assert "ColumnComp.Alarm = NoAlarm" in doc
    assert "MUTE ALARM" in doc
    assert "precond.LiquidLeakCalibrationValue" in doc
    assert "AUDIT.LiquidLeak(100.000,\"backward\")" in doc
    assert "Variables.GenericBool1" in doc
    assert "LEDBoard_LeakDiff" in doc
    assert "No_Integration" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "no DB contract expected" in doc
    assert "Open Verification Required" in doc


def test_tcc_qualification_service_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_QUALIFICATION_SERVICE_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "Qualification_Service_Done" in doc
    assert "QUALIFICATION_SERVICE_DONE" in doc
    assert "ColumnComp.ColumnComp_Wellness.QualificationDone" in doc
    assert "ColumnComp.ColumnComp_Wellness.ServiceDone" in doc
    assert "ColumnComp_Wellness.Service.LastDate" in doc
    assert "ColumnComp_Wellness.Qualification.LastDate" in doc
    assert "No_Integration" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "Internal Use" in doc
    assert "no RetTimes" in doc
    assert "no raw channels" in doc
    assert "no DB output contract" in doc
    assert "Open Verification Required" in doc


def test_tcc_factory_default_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_FACTORY_DEFAULT_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "FACTORYDEFAULT" in doc
    assert "Factory Default" in doc
    assert "ColumnComp.GetServiceCode" in doc
    assert "ColumnComp.ServiceCode = 87794" in doc
    assert "ColumnComp.ExceptionLogClear" in doc
    assert "ErrorLog.Clear" in doc
    assert "CC.ThermoUnitRevision" in doc
    assert "ModuleRevision" in doc
    assert "ColumnComp.CC.Temperature.Nominal = 20.00" in doc
    assert "ColumnComp.CC.TempCtrl = Off" in doc
    assert "ColumnComp.LiquidLeakSensor = On" in doc
    assert "AUDIT.ColumnComp.ModelNo" in doc
    assert "precond.ColumnComp.SerialNo" in doc
    assert "precond.ColumnComp.FirmwareVersion" in doc
    assert "precond.ColumnComp.HardwareVersion" in doc
    assert "precond.ColumnComp.ModuleHardwareRevision" in doc
    assert "TestDate" in doc
    assert "Serial" in doc
    assert "TimeBase" in doc
    assert "ModelNo" in doc
    assert "ModelVariant" in doc
    assert "HardwareVersion" in doc
    assert "Firmware" in doc
    assert "SubmitDate" in doc
    assert "RES_SN_Check" in doc
    assert "Definitions!C15" in doc
    assert "Internal Use!F10" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "Open Verification Required" in doc


def test_tcc_error_log_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_ERROR_LOG_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "CHECKERRORLOG" in doc
    assert "Error Log Check" in doc
    assert "ColumnComp.PrehtRight.TempCtrl = Off" in doc
    assert "ColumnComp.PrehtLeft.TempCtrl = Off" in doc
    assert "ColumnComp.CC.TempCtrl = Off" in doc
    assert "ColumnComp.Column_A.ActiveColumn = No" in doc
    assert "ColumnComp.Column_D.ActiveColumn = No" in doc
    assert "ColumnComp.Disconnect" in doc
    assert "ColumnComp.Connect" in doc
    assert "VA decoded flow omits" in doc
    assert "ReportTableObject" in doc
    assert "B11:D14" in doc
    assert "audittrail" in doc
    assert "No_Integration" in doc
    assert "CORRECT_ACCURACY_INJ_INSERTION" in doc
    assert "no closed DB output contract" in doc or "No mapped DB output contract" in doc
    assert "Open Verification Required" in doc


def test_tcc_burnin_black_box_decomposition_doc_covers_core_contracts():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_BURNIN_BLACK_BOX_DECOMPOSITION.md").read_text(encoding="utf-8")

    assert "Contract 1: Method Command" in doc
    assert "Contract 2: Processing Method" in doc
    assert "Contract 3: Report Formula" in doc
    assert "Contract 4: DB Contract" in doc
    assert "Contract 5: Config Requirement" in doc
    assert "Contract 6: Open Verification" in doc
    assert "VTCC_BurnIn" in doc
    assert "BURNIN" in doc
    assert "NO_INTEGRATION" in doc
    assert "ColumnComp.CC.ReadyTempDelta" in doc
    assert "ColumnComp.CC.EquilibrationTime" in doc
    assert "ColumnComp.CC.TempCtrl = On" in doc
    assert "ColumnComp.CmdString Cmd=\"PCC.TempCtrl=0\"" in doc
    assert "ColumnComp.ModelNo" in doc
    assert "Variables.GenericDouble1 = 5.0" in doc
    assert "Variables.GenericDouble2 = 120.0" in doc
    assert "Variables.GenericDouble2 = 85.0" in doc
    assert "ColumnComp.LiquidLeakSensor = Off" in doc
    assert "T_Maximum" in doc
    assert "T_Minimum" in doc
    assert "HoldTemp" in doc
    assert "Delay 7200.0" in doc
    assert "Thermometer1.ExtTemp_UpperCC" in doc
    assert "Thermometer1.ExtTemp_LowerCC" in doc
    assert "Thermometer.Environment_Temperature" in doc
    assert "LEDBoard_LeakDiff" in doc
    assert "System.AbortQueue" in doc
    assert "no DB output contract" in doc
    assert "Open Verification Required" in doc


def test_tcc_test_relationship_model_covers_dependency_and_intent_rules():
    doc = (PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_TEST_RELATIONSHIP_MODEL.md").read_text(encoding="utf-8")

    assert "Test Order Model" in doc
    assert "Execution Order Constraints" in doc
    assert "Dependency Matrix" in doc
    assert "Shared Resource Model" in doc
    assert "Modifiability Rules" in doc
    assert "Crop / Merge Impact Rules" in doc
    assert "Failure Propagation" in doc
    assert "Device Branch Summary" in doc
    assert "TCC_CAL_01" in doc
    assert "TCC_ACC_01" in doc
    assert "TCC_BURNIN_01" in doc
    assert "TCC_ERRORLOG_01" in doc
    assert "BURN --> CAL" in doc
    assert "CAL --> ACC" in doc
    assert "FACTORY --> ERR" in doc
    assert "AUDIT.ColumnComp.ModelNo" in doc
    assert "ExtTemp_UpperCC" in doc
    assert "ExtTemp_LowerCC" in doc
    assert "Single Accuracy 40 C" in doc
    assert "Periodic valve cycling" in doc
    assert "Liquid Leak" in doc
    assert "manual water injection" in doc
    assert "CORRECT_*" in doc
    assert "FormulaOne" in doc
    assert "Open Verification" in doc


def test_tcc_relationship_rows_bridge_model_to_alignment_export():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    rows = build_tcc_relationship_rows(records)

    assert any(row.item_id == "ORDER_02" and row.source == "Temperature Calibration" for row in rows)
    assert any(row.item_id == "DEP_03" and row.target == "PCC" and "VH-only" in row.strength for row in rows)
    assert any(row.item_id == "RES_01" and "ExtTemp_UpperCC" in row.impact for row in rows)
    assert any(row.item_id == "INTENT_02" and "2.0 min" in row.impact for row in rows)


def test_tcc_black_box_coverage_rows_audit_milestones_and_contracts():
    rows = build_tcc_black_box_coverage_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs")

    assert len(rows) >= 13
    accuracy = next(row for row in rows if row.test_id == "TCC_ACC_01")
    assert accuracy.milestone == "M1"
    assert accuracy.exists
    assert accuracy.contract_1_method
    assert accuracy.contract_6_open_verification
    assert accuracy.open_verification_count >= 1
    assert accuracy.evidence_sources_present
    assert accuracy.model_branches == "VH-C10-A, VC-C10-A, VA-C10-A"
    assert accuracy.mermaid_present
    assert accuracy.word_count > 1000
    assert "documented" in accuracy.status

    calibration = next(row for row in rows if row.test_id == "TCC_CAL_01")
    assert calibration.milestone == "M2"
    assert calibration.contract_3_report
    assert calibration.evidence_sources_present
    assert "VH-C10-A" in calibration.model_branches
    assert calibration.open_verification_count >= 1
    assert any("CORRECT_ACCURACY_INJ_INSERTION" in topic for topic in calibration.open_verification_topics)
    assert any("low-temperature reach" in topic for topic in calibration.open_verification_topics)
    assert not any("Display precision" in topic for topic in calibration.open_verification_topics)

    error_log = next(row for row in rows if row.test_id == "TCC_ERRORLOG_01")
    assert error_log.milestone == "M3"
    assert error_log.open_verification_present


def test_tcc_open_verification_topic_rows_classify_black_box_closure_queue():
    rows = build_tcc_open_verification_topic_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs", milestone="M2")

    assert rows
    calibration_rows = [row for row in rows if row.test_id == "TCC_CAL_01"]
    assert any(row.category == "Processing Method" and "CORRECT_ACCURACY_INJ_INSERTION" in row.topic for row in calibration_rows)
    assert any(row.category == "Report Formula" and "low-temperature reach" in row.topic for row in calibration_rows)
    assert not any("VA-specific report differences" in row.topic for row in calibration_rows)
    assert not any("Display precision" in row.topic for row in calibration_rows)
    assert all(row.likely_evidence_source for row in rows)
    assert all(row.closure_action for row in rows)


def test_tcc_temperature_contract_matrix_rows_summarize_m2_template_readiness():
    rows = build_tcc_temperature_contract_matrix_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs")

    assert {row.test_id for row in rows} == {
        "TCC_CAL_01",
        "TCC_PRECISION_01",
        "TCC_STABILITY_01",
        "TCC_HEATCOOL_01",
        "TCC_BURNIN_01",
    }
    calibration = next(row for row in rows if row.test_id == "TCC_CAL_01")
    assert calibration.method_contract == "present"
    assert calibration.processing_contract == "present"
    assert "Processing Method" in calibration.open_topic_categories
    assert "Report Formula" in calibration.open_topic_categories
    assert "review-only" in calibration.template_readiness
    assert "decode or manually confirm" in calibration.next_closure_actions
    heatup = next(row for row in rows if row.test_id == "TCC_HEATCOOL_01")
    assert heatup.report_contract == "present"
    assert "candidate template after CM validation" in heatup.template_readiness


def test_tcc_contract_closure_task_rows_bind_open_topics_to_contracts():
    rows = build_tcc_contract_closure_task_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs", milestone="M2")

    assert rows
    assert all(row.milestone == "M2" for row in rows)
    assert all(row.generation_blocker == "Yes" for row in rows)
    assert rows[0].priority == "P1"
    calibration_tasks = [row for row in rows if row.test_id == "TCC_CAL_01"]
    assert any(row.contract == "Contract 2 Processing Method" and "CORRECT_ACCURACY_INJ_INSERTION" in row.topic for row in calibration_tasks)
    assert any(row.contract == "Contract 3 Report Formula" and "low-temperature reach" in row.topic for row in calibration_tasks)
    assert not any("VA-specific report differences" in row.topic for row in calibration_tasks)
    assert not any("Display precision" in row.topic for row in calibration_tasks)
    assert any(row.evidence_group == "Processing method decode / CM UI" for row in calibration_tasks)
    assert any(row.evidence_group == "Report workbook/formula extraction" for row in calibration_tasks)
    assert not any("only binds correction context" in row.topic for row in rows)
    assert not any("inserts VH Temperature Stability_and_PCC_H" in row.topic for row in rows)
    assert any(
        row.evidence_group == "Configuration evidence"
        and "GenericBool0 downstream role" in row.topic
        for row in calibration_tasks
    )
    heatup_tasks = [row for row in rows if row.test_id == "TCC_HEATCOOL_01"]
    assert not any(row.contract == "Contract 3 Report Formula" for row in heatup_tasks)
    assert not any("external endpoint timing is used" in row.topic for row in heatup_tasks)
    assert not any("Exact Definitions!HeatUp & Cool Down" in row.topic for row in heatup_tasks)
    assert not any("Exact FormulaOne workbook route" in row.topic for row in heatup_tasks)
    precision_tasks = [row for row in rows if row.test_id == "TCC_PRECISION_01"]
    assert not any("Exact `Definitions!Temperature Precision`" in row.topic for row in precision_tasks)
    stability_tasks = [row for row in rows if row.test_id == "TCC_STABILITY_01"]
    assert not any("Exact acceptance limits in `Definitions`" in row.topic for row in stability_tasks)
    assert not any("Full line-by-line command script" in row.topic for row in stability_tasks)
    assert not any("Exact FormulaOne workbook formulas" in row.topic for row in stability_tasks)
    assert all(row.likely_evidence_source for row in rows)
    assert all(row.closure_action for row in rows)


def test_tcc_evidence_workstream_rows_group_p1_closure_tasks():
    rows = build_tcc_evidence_workstream_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs", milestone="M2", priority="P1")

    assert rows
    assert all(row.priority == "P1" for row in rows)
    config = next(row for row in rows if row.evidence_group == "Configuration evidence")
    assert config.task_count >= 1
    assert "TCC_CAL_01" in config.tests
    assert "Contract 5 Config Requirement" in config.contracts
    assert "Instrument setup/config manifest" in config.unlocks
    assert "Record required modules" in config.next_action
    method = next(row for row in rows if row.evidence_group == "Method command decode")
    assert "TCC_BURNIN_01" in method.tests
    assert "Contract 1 Method Command" in method.contracts
    report = next(row for row in rows if row.evidence_group == "Report workbook/formula extraction")
    assert report.task_count >= 1
    assert "TCC_CAL_01" in report.tests
    assert "TCC_PRECISION_01" in report.tests
    assert "TCC_HEATCOOL_01" not in report.tests
    assert "Contract 3 Report Formula" in report.contracts
    assert "display precision" in report.unlocks.lower()
    assert "Extract FormulaOne/workbook rules" in report.next_action


def test_tcc_p1_evidence_extraction_plan_rows_cover_processing_and_report_black_boxes():
    rows = build_tcc_p1_evidence_extraction_plan_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs", milestone="M2")

    assert rows
    assert all(row.milestone == "M2" for row in rows)
    assert all(row.status == "planned - extraction not yet executed" for row in rows)
    config = next(row for row in rows if row.test_id == "TCC_CAL_01" and row.evidence_group == "Configuration evidence")
    assert "Inspect evidence source" in config.extraction_steps
    assert "captured evidence note" in config.validation_outputs
    assert "Contract 5 Config Requirement" in config.closure_update
    method = next(row for row in rows if row.test_id == "TCC_BURNIN_01" and row.evidence_group == "Method command decode")
    assert "decoded method flow" in method.validation_outputs
    assert "Contract 1 Method Command" in method.closure_update
    report = next(row for row in rows if row.test_id == "TCC_CAL_01" and row.evidence_group == "Report workbook/formula extraction")
    assert "SpreadSheetData / FormulaOne" in report.extraction_steps
    assert "report formula map TSV" in report.validation_outputs
    assert "Contract 3 Report Formula" in report.closure_update


def test_tcc_processing_method_target_rows_bind_processing_tasks_to_device_branches():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    rows = build_tcc_processing_method_target_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs", milestone="M2", records=records)

    assert rows
    calibration = next(row for row in rows if row.test_id == "TCC_CAL_01" and row.device_model == "VH-C10-A")
    assert calibration.injection == "Temperature Calibration"
    assert calibration.instrument_method == "TEMPERATURE_CALIBRATION"
    assert calibration.processing_method == "CORRECT_ACCURACY_INJ_INSERTION"
    assert "GenericBool0 pass inserts Temperature Accuracy_H" in calibration.expected_behavior
    assert "Corrective injection insertion rows" in calibration.extraction_target
    assert calibration.readiness == "open - decode and verify processing action evidence"
    precision_h = next(row for row in rows if row.test_id == "TCC_PRECISION_01" and row.device_model == "VH-C10-A")
    assert precision_h.processing_method == "CORRECT_STABILITY_INJ_INSERTION"
    assert "GenericBool0 pass inserts Temperature Stability_and_PCC_H" in precision_h.expected_behavior
    assert "Corrective injection insertion rows" in precision_h.extraction_target


def test_tcc_report_formula_target_rows_bind_open_tasks_to_report_contracts():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    rows = build_tcc_report_formula_target_rows(PROJECT_ROOT / "cmbx_data_explorer" / "docs", milestone="M2", records=records)

    assert rows
    assert not any(row.test_id == "TCC_HEATCOOL_01" for row in rows)
    calibration = next(row for row in rows if row.test_id == "TCC_CAL_01" and row.device_model == "VA-C10-A")
    assert calibration.report_template == "Report_VATCC_V1_01"
    assert calibration.extraction_target == "FormulaOne workbook formulas and dependency cells"
    assert "Temp_Calib_Internal" in calibration.report_sheets
    assert "low-temperature reach" in calibration.topic
    precision_targets = [row for row in rows if row.test_id == "TCC_PRECISION_01"]
    assert {row.device_model for row in precision_targets} == {"VH-C10-A", "VC-C10-A"}
    assert all("Fan" in row.report_sheets for row in precision_targets)
    assert not any(row.test_id == "TCC_PRECISION_01" and row.device_model == "VA-C10-A" for row in rows)


def test_tcc_report_formula_extraction_plan_rows_define_steps_and_validation_outputs():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    rows = build_tcc_report_formula_extraction_plan_rows(
        PROJECT_ROOT / "cmbx_data_explorer" / "docs",
        milestone="M2",
        records=records,
    )

    assert rows
    assert not any(row.test_id == "TCC_HEATCOOL_01" for row in rows)
    calibration = next(
        row
        for row in rows
        if row.test_id == "TCC_CAL_01" and row.device_model == "VA-C10-A" and "FormulaOne workbook" in row.extraction_target
    )
    assert "Report_VATCC_V1_01" in calibration.extraction_steps
    assert "report formula map TSV" in calibration.validation_outputs
    assert "TempCal85_U" in calibration.extraction_steps
    assert "FORMULA_TCC_TEMPERATURECALIBRATION_OPEN" in calibration.extraction_steps
    assert calibration.status == "planned - extraction not yet executed"
    precision = next(row for row in rows if row.test_id == "TCC_PRECISION_01" and row.device_model == "VH-C10-A")
    assert "Fan" in precision.extraction_steps
    assert "TempPrecision" in precision.extraction_steps
    assert "Report_VTCC_V2_12" in precision.extraction_steps


def test_tcc_milestone_status_rows_summarize_goal_progress():
    records = filter_alignment_records(build_foq_alignment_records(kb_root=Path("__missing_kb_root__")), family="TCC")
    rows = build_tcc_milestone_status_rows(records, PROJECT_ROOT / "cmbx_data_explorer" / "docs")

    assert [row.milestone for row in rows] == ["M1", "M2", "M3", "M4", "M5"]
    assert next(row for row in rows if row.milestone == "M1").status == "complete enough for review tooling"
    assert "Open verification remains" in next(row for row in rows if row.milestone == "M2").open_work
    assert next(row for row in rows if row.milestone == "M4").status == "structured and exportable"
    assert "runnable generation closed" in next(row for row in rows if row.milestone == "M5").status


def test_tcc_method_role_map_classifies_accuracy_method_from_kb_script():
    from method_role_map import classify_method_role_map
    from method_script_kb import load_method_script_rows_from_kb

    rows = load_method_script_rows_from_kb(
        "TEMPERATURE_ACCURACY",
        family="TCC",
        device_model="VH-C10-A",
        workspace_root=PROJECT_ROOT,
    )
    audit = classify_method_role_map(
        rows,
        family="TCC",
        method_name="TEMPERATURE_ACCURACY",
        test_intent="temperature_accuracy",
        device_model="VH-C10-A",
    )

    assert audit.status == "complete"
    counts = audit.role_counts()
    assert counts["device_branch"] >= 1
    assert counts["temperature_ladder_assignment"] >= 5
    assert counts["accuracy_measurement_setpoint"] >= 5
    assert counts["ret_time_anchor"] >= 5
    assert counts["final_reset"] == 1


def test_tcc_method_role_map_blocks_unknown_method_contract():
    from method_role_map import classify_method_role_map

    audit = classify_method_role_map(
        [("1", "Command", "", "ColumnComp.CC.Temperature.Nominal", "70.0", "", "")],
        family="TCC",
        method_name="UNKNOWN_METHOD",
        test_intent="unknown_intent",
        device_model="VH-C10-A",
    )

    assert audit.status == "missing"
    assert audit.generation_mode == "blocked"


def test_tcc_method_role_map_json_is_valid():
    import json

    path = PROJECT_ROOT / "cmbx_data_explorer" / "docs" / "TCC_METHOD_ROLE_MAP.json"
    data = json.loads(path.read_text(encoding="utf-8"))

    assert data["version"]
    assert "TEMPERATURE_ACCURACY" in data["families"]["TCC"]["methods"]


def test_method_generator_builds_accuracy_baseline_to_target_preview():
    from app import CmbxExplorerApp

    class DummyVar:
        def __init__(self, value=""):
            self.value = value

        def get(self):
            return self.value

        def set(self, value):
            self.value = value

    app = CmbxExplorerApp.__new__(CmbxExplorerApp)
    app.foq_alignment_records = ()
    app.ai_api_key_var = DummyVar("")
    app.ai_model_var = DummyVar("gpt-5.5")
    app.ai_base_url_var = DummyVar("https://api.openai.com/v1")
    app.status_var = DummyVar("")
    app.progress_var = DummyVar(0)
    app.test_plan_parameter_var = DummyVar("")
    app._call_ui = lambda callback: callback()
    app._refresh_method_generator_ai_config_from_disk = lambda: None

    result = app._build_method_generator_result(
        "温度准确性测试，从40度开始，稳定30分钟后上升到60度，在60度测试准确性温度准确性测试",
        "TCC",
    )
    generated_rows = [tuple(row) for row, _changed in result["generated_rows"]]

    assert result["row_transformer_status"] == "generated"
    assert result["operation_validation"]["status"] == "method_ok_report_blocked"
    assert any(row[3] == "ColumnComp.CC.Temperature.Nominal" and row[4] == "40" for row in generated_rows)
    assert any(row[3].strip() == "Variables.GenericDouble4" and row[4] == "60" for row in generated_rows)
    assert any(row[3] == "RetTimes.RetTime4" and "keep target accuracy block" in row[5] for row in generated_rows)
    assert any(row[3] == "RetTimes.RetTime3" and row[6] == "removed" for row in generated_rows)


def _dummy_method_generator_app():
    from app import CmbxExplorerApp

    class DummyVar:
        def __init__(self, value=""):
            self.value = value

        def get(self):
            return self.value

        def set(self, value):
            self.value = value

    app = CmbxExplorerApp.__new__(CmbxExplorerApp)
    app.foq_alignment_records = ()
    app.ai_api_key_var = DummyVar("")
    app.ai_model_var = DummyVar("gpt-5.5")
    app.ai_base_url_var = DummyVar("https://api.openai.com/v1")
    app.status_var = DummyVar("")
    app.progress_var = DummyVar(0)
    app.test_plan_parameter_var = DummyVar("")
    app._call_ui = lambda callback: callback()
    app._refresh_method_generator_ai_config_from_disk = lambda: None
    return app


def test_method_generator_builds_stability_baseline_to_target_preview():
    app = _dummy_method_generator_app()

    result = app._build_method_generator_result(
        "\u0034\u0030\u5ea6\u7a33\u5b9a\u0033\u0030\u5206\u949f\u540e\uff0c\u4e0a\u5347\u5230\u0037\u0030\u5ea6\u6d4b\u8bd5\u7a33\u5b9a\u6027",
        "TCC",
    )
    generated_rows = [tuple(row) for row, _changed in result["generated_rows"]]

    assert result["spec"]["primary_intent"] == "temperature_stability"
    assert result["row_transformer_status"] == "generated"
    assert result["operation_validation"]["status"] == "ok"
    assert any(row[3] == "ColumnComp.CC.Temperature.Nominal" and cm_numeric_value(str(row[4])) == 40 for row in generated_rows)
    assert any(row[3] == "ColumnComp.CC.Temperature.Nominal" and cm_numeric_value(str(row[4])) == 70 for row in generated_rows)


def test_method_generator_blocks_multi_point_accuracy_until_ret_time_report_selection_exists():
    app = _dummy_method_generator_app()

    result = app._build_method_generator_result(
        "\u6e29\u5ea6\u51c6\u786e\u6027\u6d4b\u8bd5\uff0c\u9700\u8981\u4ece40\u5ea6\u5f00\u59cb\u6d4b\u8bd5\uff0c\u9700\u898140\u5ea6\u548c60\u5ea6\u7684\u51c6\u786e\u6027\u6d4b\u8bd5",
        "TCC",
    )

    assert result["spec"]["accuracy_points_c"] == [40.0, 60.0]
    assert result["operation_validation"]["status"] == "blocked"
    assert result["row_transformer_status"] == "blocked"
    assert "multi-point Temperature Accuracy" in result["operation_validation"]["summary"]


def test_method_generator_builds_heatup_transition_preview_from_temperatures():
    app = _dummy_method_generator_app()

    result = app._build_method_generator_result(
        "\u4fee\u6539heatup\u7684\u6e29\u5ea6\u70b9\u4e3a20 60 20",
        "TCC",
    )
    generated_rows = [tuple(row) for row, _changed in result["generated_rows"]]

    assert result["spec"]["primary_intent"] == "heatup_cooldown_20_50_20"
    assert result["row_transformer_status"] == "generated"
    assert any(cm_numeric_value(str(row[4])) == 60 for row in generated_rows)


def test_method_generator_marks_composite_intent_as_merge_blocked():
    app = _dummy_method_generator_app()

    result = app._build_method_generator_result(
        "\u6211\u73b0\u5728\u9700\u8981\u6d4b\u8bd5\u6e29\u5ea6\u572830\u5ea6\u7684\u7a33\u5b9a\u6027\uff0c\u53ea\u7528\u5916\u90e8\u6e29\u5ea6\u4f20\u611f\u5668\uff0c\u6d4b\u8bd5\u65f6\u95f4\u6539\u4e3a10\u5206\u949f\uff0c\u597d\u4e86\u4e4b\u540e\u5f00\u59cb\u6e29\u5ea6\u722c\u5761\uff0c\u523040\u5ea6\u6d4b\u8bd5\u51c6\u786e\u6027",
        "TCC",
    )

    assert result["spec"]["primary_intent"] == "temperature_stability"
    assert "temperature_accuracy" in result["spec"]["related_intents"]
    assert result["row_transformer_status"] == "partial_merge_blocked"


def test_report_template_v02_parses_existing_formulaone_cell_patches(tmp_path):
    source = tmp_path / "formulaone.md"
    source.write_text(
        """---
kind: cm_report_template
spec_version: 0.2
template_name: Control
reference_template:
  cmbx: carrier.cmbx
generation_mode: clone_and_patch
workbook_policy: existing_cells_only
---

## Sheet: Sheet2

### Workbook Value: A52
```yaml
operation: replace
value_type: number
value: 1
```

### Workbook Formula: D52
```yaml
operation: replace
formula: '=IF(A52=1,B52=1)'
```
""",
        encoding="utf-8",
    )
    spec = parse_report_template_md(source)

    assert spec.errors == ()
    assert [(patch.sheet_name, patch.excel_range, patch.value_type, patch.value) for patch in spec.workbook_patches] == [
        ("Sheet2", "A52", "number", "1"),
        ("Sheet2", "D52", "formula", "=IF(A52=1,B52=1)"),
    ]


def test_report_template_v10_parses_blank_sheet_cells_and_cm_formulas(tmp_path):
    source = tmp_path / "blank_report.md"
    source.write_text(
        """---
kind: cm_report_template
spec_version: 1.0
template_name: New_Report
generation_mode: create_from_blank
workbook_policy: create_static
---

## Sheet: Summary

### Sheet Settings
```yaml
active: true
each_injection: true
column_widths: [A=24, B=18]
row_heights: [1=24]
```

### Workbook Text: A1
```yaml
operation: create
value_type: text
value: Summary
style: title
```

### Workbook Formula: B4
```yaml
operation: create
value_type: formula
formula: '=A4+1'
number_format: '0.00'
```

### CM Formula: B2
```yaml
operation: create
object_type: ReportFormulaObject
formula: seq.name
fixed_channel: ''
fixed_component: ''
```
""",
        encoding="utf-8",
    )

    spec = parse_report_template_md(source)

    assert spec.errors == ()
    assert spec.reference_cmbx == ""
    assert spec.generation_mode == "create_from_blank"
    assert spec.sheets[0].name == "Summary"
    assert spec.sheets[0].column_widths == ((1, 24.0), (2, 18.0))
    assert spec.workbook_patches[0].style == "title"
    assert spec.workbook_patches[1].number_format == "0.00"
    assert spec.patches[0].formula == "seq.name"


def test_method_md_linter_rejects_unquoted_message_text():
    issues = lint_method_rows(
        [
            {"#": "1", "Kind": "Command", "Time": "", "Command": "Message", "Value": "Invalid model"},
            {"#": "2", "Kind": "Command", "Time": "", "Command": "Message", "Value": '"Invalid model"'},
        ]
    )

    assert [(issue.code, issue.row) for issue in issues] == [("MESSAGE_TEXT_UNQUOTED", "1")]


def test_method_md_linter_rejects_unquoted_cm_string_values():
    issues = lint_method_rows(
        [
            {"#": "1", "Kind": "Command", "Time": "", "Command": "Variables.GenericString0", "Value": "10300,"},
            {"#": "2", "Kind": "Command", "Time": "", "Command": "Variables.GenericString0", "Value": '"10300,"'},
            {"#": "3", "Kind": "Command", "Time": "", "Command": "PumpModule.PumpModule_Service._SendCommand", "Value": "Flow1.Blk1.Drv1.PositionMode=200000"},
            {"#": "4", "Kind": "Command", "Time": "", "Command": "PumpModule.PumpModule_Service._SendCommand", "Value": '"Flow1.Blk1.Drv1.PositionMode=200000"'},
            {"#": "5", "Kind": "Command", "Time": "", "Command": "VirtualChannel", "Value": "Volume_Loss_per_Time, Variables.GenericFloat7"},
            {"#": "6", "Kind": "Command", "Time": "", "Command": "VirtualChannel", "Value": '"Volume_Loss_per_Time", Variables.GenericFloat7'},
        ]
    )

    assert [(issue.code, issue.row) for issue in issues] == [
        ("GENERIC_STRING_UNQUOTED", "1"),
        ("SEND_COMMAND_UNQUOTED", "3"),
        ("VIRTUAL_CHANNEL_NAME_UNQUOTED", "5"),
    ]


def test_method_md_parser_keeps_trigger_explanation_as_comment(tmp_path):
    source = tmp_path / "method.md"
    source.write_text(
        """```tsv
Time\tCommand\tValue\tComment
\tTriggers T_Start_Ext and T_Start_Int ensure the same equilibration conditions.\t\t
Trigger\t\t\"T_Start_Ext\",\t
\tCondition\tVariables.GenericLong0=0\t
\tTrueTime\t120\t
\tLimit\t1\t
End Trigger\t\t\t
```""",
        encoding="utf-8",
    )

    rows = parse_md_to_rows(source)

    assert rows[0]["Kind"] == "Comment"
    assert rows[0]["Command"].startswith("Triggers T_Start_Ext")


def test_report_template_v10_parses_native_audittrail_table(tmp_path):
    source = tmp_path / "audit_report.md"
    source.write_text(
        """---
kind: cm_report_template
spec_version: 1.0
template_name: Audit_Report
generation_mode: create_from_blank
workbook_policy: create_static
---

## Sheet: Valve Events

### Dynamic Table: A3:C10
```yaml
operation: create
table_type: audittrail
body_rows: 6
audit_level: Expert
show_run: true
show_preconditions: false
show_day_time: true
day_time_format: hh:mm:ss
show_device: false
```
""",
        encoding="utf-8",
    )

    spec = parse_report_template_md(source)

    assert spec.errors == ()
    assert len(spec.dynamic_tables) == 1
    table = spec.dynamic_tables[0]
    assert (table.sheet_name, table.excel_range, table.table_type, table.body_rows) == ("Valve Events", "A3:C10", "audittrail", 6)


def test_report_template_v10_parses_peak_summary_formula_columns(tmp_path):
    source = tmp_path / "summary_report.md"
    source.write_text(
        """---
kind: cm_report_template
spec_version: 1.0
template_name: Injection_Summary
generation_mode: create_from_blank
workbook_policy: create_static
---

## Sheet: Summary

### Dynamic Table: A3:C9
```yaml
operation: create
table_type: peak_summary
body_rows: 3
sort_formula: injection.number
show_unknown: true
requires_processing: false
```

#### Table Column: Injection
```yaml
formula: injection.name
header: Injection
unit: ''
```

#### Table Column: Model
```yaml
formula: AUDIT.ColumnComp.ModelNo
header: Model
unit: ''
```

#### Table Column: Sequence
```yaml
formula: seq.name
header: Sequence
unit: ''
```
""",
        encoding="utf-8",
    )

    spec = parse_report_template_md(source)

    assert spec.errors == ()
    table = spec.dynamic_tables[0]
    assert table.table_type == "peak_summary"
    assert [column.formula for column in table.columns] == ["injection.name", "AUDIT.ColumnComp.ModelNo", "seq.name"]


def test_report_template_v11_compiles_native_dynamic_tables(tmp_path):
    source = tmp_path / "dynamic_report.md"
    source.write_text(
        """---
kind: cm_report_template
spec_version: 1.0
template_name: Dynamic_Report
generation_mode: create_from_blank
workbook_policy: create_static
---

## Sheet: Events

### Dynamic Table: A2:C8
```yaml
operation: create
table_type: audittrail
body_rows: 5
show_day_time: true
show_device: false
```

## Sheet: Summary

### Dynamic Table: A2:B7
```yaml
operation: create
table_type: peak_summary
body_rows: 3
requires_processing: false
```

#### Table Column: Injection
```yaml
formula: injection.name
header: Injection
```

#### Table Column: Model
```yaml
formula: AUDIT.ColumnComp.ModelNo
header: Model
```
""",
        encoding="utf-8",
    )
    spec = parse_report_template_md(source)
    output = tmp_path / "Dynamic_Report.cmbx"

    result = compile_report_template_md_to_cmbx(spec, output)

    assert result.ready
    package = load_cmbx_package(output)
    report = next(item for item in package.methods_and_reports if item.kind == "report_template")
    _embedded, xml = decode_report_template_xml(package, report)
    tables = [item for item in parse_report_sheet_objects(xml, report.name) if item.object_type == "ReportTableObject"]
    assert [(item.sheet_name, item.excel_range, item.table_type) for item in tables] == [
        ("Events", "A2:C8", "audittrail"),
        ("Summary", "A2:B7", "peak_summary"),
    ]


def test_report_template_v14_requires_complete_integration_contract(tmp_path):
    source = tmp_path / "invalid_integration.md"
    source.write_text(
        """---
kind: cm_report_template
spec_version: 1.0
template_name: Valve_Integration
generation_mode: create_from_blank
workbook_policy: create_static
---

## Sheet: Valve Events

### Dynamic Table: A3:A12
```yaml
operation: create
table_type: integration
body_rows: 8
```

#### Table Column: Switch Time
```yaml
formula: peak.retention_time("detected")*60
header: Switch Time
unit: s
```
""",
        encoding="utf-8",
    )

    spec = parse_report_template_md(source)

    assert any("requires requires_processing: true" in item for item in spec.errors)
    assert any("requires an exact processing_method name" in item for item in spec.errors)
    assert any("requires fixed_channel" in item for item in spec.errors)


def test_report_template_v14_compiles_valve_integration_table(tmp_path):
    source = tmp_path / "valve_integration.md"
    source.write_text(
        """---
kind: cm_report_template
spec_version: 1.0
template_name: Valve_Integration
generation_mode: create_from_blank
workbook_policy: create_static
---

## Sheet: Valve Events

### Dynamic Table: A3:C42
```yaml
operation: create
table_type: integration
body_rows: 38
requires_processing: true
processing_method: PressureSpikeEval
fixed_channel: PumpPressureVirtual
sort_formula: peak.group
include_identified_peaks: true
include_unidentified_peaks: true
```

#### Table Column: Switch Time
```yaml
formula: peak.retention_time("detected")*60
header: Switch Time
unit: s
channel: PumpPressureVirtual
```

#### Table Column: Upper Position
```yaml
formula: audit.ColumnComp.UpperValve.CurrentPosition
header: Upper Position
channel: PumpPressureVirtual
```

#### Table Column: Lower Position
```yaml
formula: audit.ColumnComp.LowerValve.CurrentPosition
header: Lower Position
channel: PumpPressureVirtual
```
""",
        encoding="utf-8",
    )
    spec = parse_report_template_md(source)
    output = tmp_path / "Valve_Integration.cmbx"

    assert spec.errors == ()
    table = spec.dynamic_tables[0]
    assert table.processing_method == "PressureSpikeEval"
    assert table.fixed_channel == "PumpPressureVirtual"
    result = compile_report_template_md_to_cmbx(spec, output)

    assert result.ready, result.errors
    package = load_cmbx_package(output)
    report = next(item for item in package.methods_and_reports if item.kind == "report_template")
    _embedded, xml = decode_report_template_xml(package, report)
    root = ET.fromstring(xml)
    table_object = next(
        item for item in root.findall(".//SheetObject")
        if item.find("ReportTableType") is not None and item.find("ReportTableType").get("value") == "integration"
    )
    assert [
        item.find("Formula/Formula").get("value")
        for item in table_object.findall(".//Columns/Item")
    ] == [
        'peak.retention_time("detected")*60',
        "audit.ColumnComp.UpperValve.CurrentPosition",
        "audit.ColumnComp.LowerValve.CurrentPosition",
    ]
    properties = table_object.find(".//Properties")
    assert properties.find("FixedChannelName").get("value") == "PumpPressureVirtual"
