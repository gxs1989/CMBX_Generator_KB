from __future__ import annotations

import ast
import math
import operator
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from cmbx_container import CmbxElement, CmbxPackage, load_cmbx_package, safe_filename
from external_report_spec import ExternalReportOperation, ExternalReportSpec
from report_formula_evaluator import (
    AuditRecord,
    ReportFormulaEvaluationContext,
    SignalPoint,
    build_report_formula_context,
    evaluate_audit_metadata_formula,
    evaluate_audit_property_formula,
    evaluate_chm_drift,
    evaluate_chm_noise,
    evaluate_chm_signal_formula,
    evaluate_chm_signal_value,
    evaluate_external_report_formula,
    load_injection_signal,
)


@dataclass(frozen=True)
class CompatibilityResult:
    package_path: Path
    injection_id: str
    sequence_name: str
    injection_name: str
    compatible: bool
    missing_channels: tuple[str, ...] = ()
    missing_audit_paths: tuple[str, ...] = ()
    missing_ret_times: tuple[int, ...] = ()
    detail: str = ""


@dataclass(frozen=True)
class ReportValue:
    operation_id: str
    label: str
    value: Any
    status: str
    detail: str = ""
    number_format: str = "General"


@dataclass
class InjectionReportResult:
    package_path: Path
    sequence_name: str
    injection_name: str
    compatible: bool
    values: list[ReportValue] = field(default_factory=list)
    tables: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    plots: dict[str, list[SignalPoint]] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


class ExternalReportEngine:
    def __init__(self, spec: ExternalReportSpec):
        self.spec = spec

    def load_packages(self, paths: list[str | Path]) -> list[CmbxPackage]:
        return [load_cmbx_package(path) for path in paths]

    def compatibility_matrix(self, packages: list[CmbxPackage]) -> list[CompatibilityResult]:
        rows: list[CompatibilityResult] = []
        for package in packages:
            for injection in package.injections:
                rows.append(self._compatibility(package, injection))
        return rows

    def execute(
        self,
        packages: list[CmbxPackage],
        selected: set[tuple[str, str]] | None = None,
    ) -> list[InjectionReportResult]:
        results: list[InjectionReportResult] = []
        for package in packages:
            for injection in package.injections:
                key = (str(package.path.resolve()), injection.id)
                if selected is not None and key not in selected:
                    continue
                compatibility = self._compatibility(package, injection)
                result = InjectionReportResult(
                    package_path=package.path,
                    sequence_name=compatibility.sequence_name,
                    injection_name=injection.name,
                    compatible=compatibility.compatible,
                )
                if not compatibility.compatible:
                    result.errors.append(compatibility.detail)
                    results.append(result)
                    continue
                try:
                    context = build_report_formula_context(package, injection)
                    namespace: dict[str, Any] = {}
                    for operation in self.spec.operations:
                        self._execute_operation(package, injection, context, operation, result, namespace)
                except Exception as exc:
                    result.errors.append(str(exc))
                results.append(result)
        return results

    def _compatibility(self, package: CmbxPackage, injection: CmbxElement) -> CompatibilityResult:
        channels = {child.name.casefold(): child.name for child in injection.children if child.kind == "signal"}
        required_channels = set(self.spec.channels)
        required_channels.update(op.channel for op in self.spec.operations if op.channel)
        missing_channels = tuple(sorted(name for name in required_channels if name.casefold() not in channels))
        required_audit_paths = set(self.spec.audit_paths)
        required_audit_paths.update(path for op in self.spec.operations if op.kind == "audit_event_table" for path in op.property_paths)
        required_audit_paths.update(path for op in self.spec.operations if op.kind == "scalar" for path in _formula_audit_paths(op.formula))
        missing_paths: list[str] = []
        missing_ret_times: list[int] = []
        audit_required = bool(required_audit_paths or self.spec.ret_times or any(op.kind == "audit_event_table" for op in self.spec.operations))
        detail_parts: list[str] = []
        if audit_required:
            try:
                context = build_report_formula_context(package, injection)
                for path in sorted(required_audit_paths):
                    if not any(_audit_path_matches(record, path) for record in context.audit_records):
                        missing_paths.append(path)
                missing_ret_times = sorted(set(self.spec.ret_times) - set(context.ret_times))
            except Exception as exc:
                detail_parts.append(f"audit read failed: {exc}")
                missing_paths.extend(path for path in required_audit_paths if path not in missing_paths)
                missing_ret_times.extend(value for value in self.spec.ret_times if value not in missing_ret_times)
        if missing_channels:
            detail_parts.append("missing channels: " + ", ".join(missing_channels))
        if missing_paths:
            detail_parts.append("missing audit paths: " + ", ".join(missing_paths))
        if missing_ret_times:
            detail_parts.append("missing RetTimes: " + ", ".join(f"RetTime{value}" for value in missing_ret_times))
        sequence = package.elements_by_id.get(injection.parent_id or "")
        compatible = not (missing_channels or missing_paths or missing_ret_times or detail_parts and detail_parts[0].startswith("audit read failed"))
        return CompatibilityResult(
            package_path=package.path,
            injection_id=injection.id,
            sequence_name=sequence.name if sequence else "",
            injection_name=injection.name,
            compatible=compatible,
            missing_channels=missing_channels,
            missing_audit_paths=tuple(missing_paths),
            missing_ret_times=tuple(missing_ret_times),
            detail="; ".join(detail_parts) or "compatible",
        )

    def _execute_operation(
        self,
        package: CmbxPackage,
        injection: CmbxElement,
        context: ReportFormulaEvaluationContext,
        operation: ExternalReportOperation,
        result: InjectionReportResult,
        namespace: dict[str, Any],
    ) -> None:
        if operation.kind == "scalar":
            value, status, detail = evaluate_external_report_formula(
                package,
                injection,
                operation.formula,
                operation.channel,
                context,
            )
            parsed = _number_or_text(value)
            namespace[operation.operation_id] = parsed
            result.values.append(ReportValue(operation.operation_id, operation.label, parsed, status, detail, operation.number_format))
        elif operation.kind == "formula":
            try:
                value = safe_expression(operation.expression, namespace)
                namespace[operation.operation_id] = value
                result.values.append(ReportValue(operation.operation_id, operation.label, value, "ok", operation.expression, operation.number_format))
            except Exception as exc:
                result.values.append(ReportValue(operation.operation_id, operation.label, "", "error", str(exc), operation.number_format))
        elif operation.kind == "audit_event_table":
            result.tables[operation.operation_id] = audit_event_rows(context.audit_records, operation)
        elif operation.kind == "raw_event_table":
            signal = load_injection_signal(package, injection, operation.channel, context)
            result.tables[operation.operation_id] = raw_event_rows(signal, operation)
        elif operation.kind == "plot":
            signal = load_injection_signal(package, injection, operation.channel, context)
            result.plots[operation.operation_id] = _window_points(signal, operation.start_min, operation.end_min)


def _evaluate_signal_formula(
    formula: str,
    channel: str,
    context: ReportFormulaEvaluationContext,
    signal: list[SignalPoint],
) -> tuple[str, str, str]:
    lowered = formula.strip().lower()
    if lowered.startswith("chm.signalvalue"):
        return evaluate_chm_signal_value(formula, channel, context.ret_times, signal)
    if lowered.startswith("chm.noise"):
        return evaluate_chm_noise(formula, channel, context.ret_times, signal)
    if lowered.startswith("chm.drift"):
        return evaluate_chm_drift(formula, channel, context.ret_times, signal)
    return evaluate_chm_signal_formula(formula, channel, context.ret_times, signal)


def _formula_audit_paths(formula: str) -> tuple[str, ...]:
    expression = formula.strip()
    match = re.match(r"(?:AUDIT|precond)\.([A-Za-z0-9_.]+)", expression, flags=re.I)
    if not match or match.group(1).casefold().startswith("rettime"):
        return ()
    return (match.group(1),)


def audit_event_rows(records: list[AuditRecord], operation: ExternalReportOperation) -> list[dict[str, Any]]:
    selected: list[AuditRecord] = []
    for record in records:
        path_match = not operation.property_paths or any(_audit_path_matches(record, path) for path in operation.property_paths)
        message_match = not operation.message_contains or any(text.casefold() in record.message.casefold() for text in operation.message_contains)
        if path_match and message_match:
            selected.append(record)
    if operation.value_changes_only:
        last_values: dict[str, str] = {}
        changed: list[AuditRecord] = []
        for record in selected:
            key = f"{record.device}.{record.property_name}".casefold()
            if last_values.get(key) != record.property_value:
                changed.append(record)
                last_values[key] = record.property_value
        selected = changed
    rows = [
        {
            "time_min": record.retention_time_min,
            "device": record.device,
            "property": record.property_name,
            "value": record.property_value,
            "message": record.message,
        }
        for record in selected
    ]
    if operation.group_within_seconds > 0:
        rows = _group_event_rows(rows, operation.group_within_seconds)
    return _select_columns(rows, operation.columns)


def raw_event_rows(signal: list[SignalPoint], operation: ExternalReportOperation) -> list[dict[str, Any]]:
    threshold = float(operation.threshold or 0.0)
    hysteresis = max(0.0, operation.hysteresis)
    points = _window_points(signal, operation.start_min, operation.end_min)
    if not points:
        return []
    state = points[0].value >= threshold
    last_time = -math.inf
    rows: list[dict[str, Any]] = []
    for point in points[1:]:
        next_state = state
        if not state and point.value >= threshold + hysteresis:
            next_state = True
        elif state and point.value <= threshold - hysteresis:
            next_state = False
        if next_state != state:
            edge = "rising" if next_state else "falling"
            if operation.edge in {"both", edge} and (point.time_min - last_time) * 60 >= operation.debounce_seconds:
                rows.append({"time_min": point.time_min, "value": point.value, "edge": edge})
                last_time = point.time_min
            state = next_state
    return _select_columns(rows, operation.columns)


def _select_columns(rows: list[dict[str, Any]], columns: tuple[str, ...]) -> list[dict[str, Any]]:
    if not columns:
        return rows
    return [{column: row.get(column, "") for column in columns} for row in rows]


def _group_event_rows(rows: list[dict[str, Any]], within_seconds: float) -> list[dict[str, Any]]:
    groups: list[list[dict[str, Any]]] = []
    for row in rows:
        time_value = row.get("time_min")
        if not groups or time_value is None or groups[-1][0].get("time_min") is None or (time_value - groups[-1][0]["time_min"]) * 60 > within_seconds:
            groups.append([row])
        else:
            groups[-1].append(row)
    output: list[dict[str, Any]] = []
    for group in groups:
        output.append(
            {
                "time_min": group[0].get("time_min"),
                "events": len(group),
                "device": " | ".join(str(row.get("device") or "") for row in group),
                "property": " | ".join(str(row.get("property") or "") for row in group),
                "value": " | ".join(str(row.get("value") or "") for row in group),
                "message": " | ".join(str(row.get("message") or "") for row in group),
            }
        )
    return output


def _window_points(points: list[SignalPoint], start: float | None, end: float | None) -> list[SignalPoint]:
    return [point for point in points if (start is None or point.time_min >= start) and (end is None or point.time_min <= end)]


def _audit_path_matches(record: AuditRecord, audit_path: str) -> bool:
    expected = audit_path.removeprefix("AUDIT.").removeprefix("precond.").casefold()
    full = f"{record.device}.{record.property_name}".strip(".").casefold()
    prop = record.property_name.casefold()
    return full == expected or full.endswith("." + expected) or expected.endswith("." + full) or prop == expected or expected.endswith("." + prop)


_BINARY = {ast.Add: operator.add, ast.Sub: operator.sub, ast.Mult: operator.mul, ast.Div: operator.truediv, ast.Pow: operator.pow, ast.Mod: operator.mod}
_UNARY = {ast.UAdd: operator.pos, ast.USub: operator.neg, ast.Not: operator.not_}
_COMPARE = {ast.Eq: operator.eq, ast.NotEq: operator.ne, ast.Lt: operator.lt, ast.LtE: operator.le, ast.Gt: operator.gt, ast.GtE: operator.ge}
_FUNCTIONS = {"abs": abs, "min": min, "max": max, "round": round, "sum": sum}


def safe_expression(expression: str, values: dict[str, Any]) -> Any:
    return _eval_ast(ast.parse(expression, mode="eval").body, values)


def _eval_ast(node: ast.AST, values: dict[str, Any]) -> Any:
    if isinstance(node, ast.Constant):
        return node.value
    if isinstance(node, ast.Name):
        if node.id not in values:
            raise ValueError(f"Unknown report value: {node.id}")
        return values[node.id]
    if isinstance(node, ast.BinOp) and type(node.op) in _BINARY:
        return _BINARY[type(node.op)](_eval_ast(node.left, values), _eval_ast(node.right, values))
    if isinstance(node, ast.UnaryOp) and type(node.op) in _UNARY:
        return _UNARY[type(node.op)](_eval_ast(node.operand, values))
    if isinstance(node, ast.Compare):
        left = _eval_ast(node.left, values)
        for op, comparator in zip(node.ops, node.comparators):
            right = _eval_ast(comparator, values)
            if type(op) not in _COMPARE or not _COMPARE[type(op)](left, right):
                return False
            left = right
        return True
    if isinstance(node, ast.BoolOp):
        items = [_eval_ast(item, values) for item in node.values]
        return all(items) if isinstance(node.op, ast.And) else any(items)
    if isinstance(node, ast.IfExp):
        return _eval_ast(node.body if _eval_ast(node.test, values) else node.orelse, values)
    if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) and node.func.id in _FUNCTIONS:
        return _FUNCTIONS[node.func.id](*[_eval_ast(arg, values) for arg in node.args])
    raise ValueError(f"Unsupported expression element: {type(node).__name__}")


def write_external_report_workbook(spec: ExternalReportSpec, results: list[InjectionReportResult], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Package", "Sequence", "Injection", "Compatible", *[op.label for op in spec.operations if op.kind in {"scalar", "formula"}]])
    _style_header(summary[1])
    for result in results:
        value_map = {value.operation_id: value.value for value in result.values}
        row = [result.package_path.name, result.sequence_name, result.injection_name, result.compatible]
        row.extend(value_map.get(op.operation_id, "") for op in spec.operations if op.kind in {"scalar", "formula"})
        summary.append(row)
        for column_index, operation in enumerate((op for op in spec.operations if op.kind in {"scalar", "formula"}), start=5):
            summary.cell(summary.max_row, column_index).number_format = operation.number_format
    summary.freeze_panes = "A2"
    for result_index, result in enumerate(results, start=1):
        for table_id, rows in result.tables.items():
            title = safe_filename(f"{result_index}_{table_id}")[:31]
            sheet = workbook.create_sheet(title)
            if not rows:
                sheet.append(["No matching events"])
                continue
            headers = list(rows[0])
            sheet.append(headers)
            _style_header(sheet[1])
            for row in rows:
                sheet.append([row.get(header, "") for header in headers])
            sheet.freeze_panes = "A2"
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            width = min(60, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(path)
    return path


def _style_header(cells) -> None:
    for cell in cells:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")
        cell.alignment = Alignment(vertical="center")


def _number_or_text(value: str) -> Any:
    try:
        return float(value)
    except (TypeError, ValueError):
        return value
